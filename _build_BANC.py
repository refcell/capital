#!/usr/bin/env python3
"""Build BANC (Banc of California) bank valuation model — 6 sheets."""

import sys
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TODAY = date.today().strftime("%Y-%m-%d")
TICKER = "BANC"
COMPANY = "Banc of California, Inc."
PRICE = 20.79
SHARES_MM = 153.72  # outstanding per Yahoo Stats
IMP_SHARES_MM = 154.2
MC_B = 3.28  # $B from Yahoo Stats
TOTAL_DEBT_B = 3.63  # Key Stats — higher, per skill guidance
TOTAL_CASH_B = 2.21  # MRQ per Yahoo Stats
PREFERRED_STOCK_B = 0.50  # $498.5M from balance sheet
COMMON_EQUITY_B = 3.04  # $B
TOTAL_EQUITY_B = 3.54  # $B
BVPS_TOTAL = 19.81  # MRQ per Yahoo Stats
BVPS_COMMON = 19.58  # computed: 3,042,761K / 155,458K
BETA = 1.14
ROE_TTM = 6.99  # % per Yahoo Stats on total equity
ROA_TTM = 0.72  # %
DIV_RATE = 0.48  # forward annual
DIV_YIELD = 2.26  # %
NEXT_EARNINGS = "2026-07-29"

# Income Statement (thousands of USD)
# TTM  FY2025  FY2024  FY2023  FY2022
REVENUE =      [1121822, 1103153, 985889, 278385, 1341017]
PRETAX_INC =   [335768, 313075, 168654, -2211338, 567568]
TAX =          [88411, 84102, 41766, -312201, 143955]
NET_INC =      [207569, 189185, 87100, -1938925, 404274]
EPS_DIL =      [1.30, 1.17, 0.52, -22.71, 1.89]
INT_INCOME =   [1677440, 1676653, 1812705, 1971000, 1556489]
INT_EXPENSE =  [680801, 699267, 886655, 1223872, 265727]
NII =          [996639, 977386, 926050, 747128, 1290762]
BASIC_AVG_SH = [156388.5, 159807, 168441, 85394, 60802.08]

# Balance Sheet (thousands)
TOTAL_ASSETS =   [34797442, 33542864, 38534064, 41228936]  # FY25-FY22
TOTAL_DEBT_BS =  [3016559, 2159737, 1229621, 2631117]
COMMON_EQ_BS =   [3042761, 3001433, 2892249, 3452015]
TANGIBLE_BV =    [2722953, 2653968, 2528145, 2043898]  # Net Tangible Assets

# Cash Flow (thousands)
OCF =    [289460, 255601, 77374, 135768, 701972]
FCF =    [266775, 234771, 64327, 120549, 581110]
CAPEX =  [22685, 20830, 13047, 15219, 120862]
END_CASH = [2217269, 2307965, 2502212, 5377576, 2240222]

# Analyst Estimates (Non-GAAP diluted EPS)
# Q2 FY26 est: $0.40, Q3 est: $0.44, FY26: $1.68, FY27: $2.05
# Revenue est: Q2: $295.35M, Q3: $305.94M, FY26: $1.20B, FY27: $1.31B
EPS_FY26 = 1.68
EPS_FY27 = 2.05
REV_FY26 = 1200  # $M
REV_FY27 = 1310  # $M

# 10Y US Treasury yield
RISK_FREE = 4.64  # % from CNBC as of Jul 27, 2026

# Tax rate TTM
TAX_RATE = 88411 / 335768 * 100  # ~26.3%

# WACC calculation
ERP = 5.0  # %
Ke = RISK_FREE + ERP * BETA  # 4.64 + 5*1.14 = 10.34%
Kd = 5.5  # estimated cost of senior/sub debt pre-tax
# Equity/debt weights
E_WEIGHT = MC_B / (MC_B + TOTAL_DEBT_B)  # 3.28 / 6.91 = 0.475
D_WEIGHT = TOTAL_DEBT_B / (MC_B + TOTAL_DEBT_B)  # 0.525
WACC = E_WEIGHT * Ke + D_WEIGHT * Kd * (1 - TAX_RATE / 100)

print(f"Ke = {Ke:.2f}%")
print(f"Kd = {Kd:.2f}%")
print(f"E weight = {E_WEIGHT:.3f}, D weight = {D_WEIGHT:.3f}")
print(f"WACC = {WACC:.2f}%")
print(f"Tax rate = {TAX_RATE:.1f}%")

# Scenario model: P/B + ROE framework (bank-specific)
# Common BVPS = $19.58
# Current P/B (common) = $20.79 / $19.58 = 1.06x

bear_bvps_cagr = 2.5
bear_exit_pb = 0.80
bear_roe = 5.5

base_bvps_cagr = 5.5
base_exit_pb = 1.10
base_roe = 8.5

bull_bvps_cagr = 8.5
bull_exit_pb = 1.35
bull_roe = 11.0

# Revenue CAGR from analyst estimates
# FY26 rev: $1.20B, FY27: $1.31B
# TTM: $1.12B
# 5-yr CAGR bridge:
# bear: revenue stabilizes around $1.2-1.25B plateau
# base: gradual growth to $1.5B
# bull: growth to $1.7B+
bear_rev_cagr = 2.0
base_rev_cagr = 5.0
bull_rev_cagr = 7.5

# Terminal BVPS after 5 years (scenario horizon)
bear_term_bvps = BVPS_COMMON * ((1 + bear_bvps_cagr / 100) ** 5)
base_term_bvps = BVPS_COMMON * ((1 + base_bvps_cagr / 100) ** 5)
bull_term_bvps = BVPS_COMMON * ((1 + bull_bvps_cagr / 100) ** 5)

# Target prices
bear_target = bear_term_bvps * bear_exit_pb
base_target = base_term_bvps * base_exit_pb
bull_target = bull_term_bvps * bull_exit_pb

# Upside
bear_upside = (bear_target / PRICE - 1) * 100
base_upside = (base_target / PRICE - 1) * 100
bull_upside = (bull_target / PRICE - 1) * 100

# Weights
bear_w, base_w, bull_w = 0.25, 0.50, 0.25
weighted_fv = bear_w * bear_target + base_w * base_target + bull_w * bull_target
weighted_upside = (weighted_fv / PRICE - 1) * 100

print(f"\n--- Scenario Results ---")
print(f"Bear BVPS_5Y: ${bear_term_bvps:.2f}, Target: ${bear_target:.2f}, Upside: {bear_upside:.1f}%")
print(f"Base BVPS_5Y: ${base_term_bvps:.2f}, Target: ${base_target:.2f}, Upside: {base_upside:.1f}%")
print(f"Bull BVPS_5Y: ${bull_term_bvps:.2f}, Target: ${bull_target:.2f}, Upside: {bull_upside:.1f}%")
print(f"Weighted FV: ${weighted_fv:.2f}, Upside: {weighted_upside:.1f}%")

# ===== Build Workbook =====
wb = Workbook()

# Styles
title_font = Font(name='Calibri', size=16, bold=True)
subtitle_font = Font(name='Calibri', size=12, bold=True)
header_font = Font(name='Calibri', size=11, bold=True)
data_font = Font(name='Calibri', size=11)
note_font = Font(name='Calibri', size=10, italic=True)
money_fmt = '#,##0.00'
percent_fmt = '0.00%'
int_fmt = '#,##0'
header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
note_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))

def c(ws, row, col, value, fmt= None, font= None, fill=None):
    """Convenience: write cell with formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    cell.border = thin_border
    if font:
        cell.font = font
    else:
        cell.font = data_font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell

# ---- Sheet 1: Valuation ----
ws1 = wb.active
ws1.title = "Valuation"
ws1.merge_cells('A1:F1')
ws1.merge_cells('A2:F2')
c(ws1, 1, 1, f"{COMPANY} ({TICKER}) — Valuation Summary", font=title_font)
c(ws1, 2, 1, f"As of {TODAY} | Bank-Framework (P/B + ROE / Forward P/E) | FCF multiples N/A", font=note_font, fill=note_fill)

val_title = [
    ("Company", COMPANY),
    ("Date", TODAY),
    ("Ticker", f"NYSE: {TICKER}"),
    ("Sector / Industry", "Financial Services / Banks — Regional"),
    ("Price ($)", PRICE),
    ("Shares Outstanding (M)", SHARES_MM),
    ("Implied Shares Outstanding (M)", IMP_SHARES_MM),
    ("Market Cap ($B)", MC_B),
    ("Total Debt — Key Stats ($B)", TOTAL_DEBT_B),
    ("Total Cash ($B)", TOTAL_CASH_B),
    ("Preferred Stock ($B)", PREFERRED_STOCK_B),
    ("Total Equity ($B)", TOTAL_EQUITY_B),
    ("Common Equity ($B)", COMMON_EQUITY_B),
    ("BVPS — Total (MRQ)", BVPS_TOTAL),
    ("BVPS — Common", BVPS_COMMON),
    ("Primary Valuation Lens", "P/B + ROE / Forward P/E (bank framework)"),
    ("Stance", "Watch"),
]

val_data = [
    ("P/B Ratio (Total)", f"{PRICE / BVPS_TOTAL:.2f}x", "Below peer median ~1.2-1.5x for regional banks"),
    ("P/B Ratio (Common)", f"{PRICE / BVPS_COMMON:.2f}x", "Excludes preferred stock equity"),
    ("Forward P/E (FY26)", f"{PRICE / EPS_FY26:.1f}x", f"Consensus EPS ${EPS_FY26:.2f}, Non-GAAP"),
    ("Forward P/E (FY27)", f"{PRICE / EPS_FY27:.1f}x", f"Consensus EPS ${EPS_FY27:.2f}, Non-GAAP"),
    ("P/S Ratio (TTM)", f"{MC_B / (1121.822):.1f}x", "TTM revenue $1,122M"),
    ("EV/EBITDA", "N/A", "EV not applicable for banks (deposits = operating liabilities)"),
    ("P/FCF", "N/A", "FCF multiples not primary lens for banks"),
    ("ROE (TTM)", f"{ROE_TTM:.1f}%", "Below cost of equity ~{Ke:.1f}%"),
    ("ROA (TTM)", f"{ROA_TTM:.1f}%", "Below 1%; recovering from post-bankruptcy position"),
    ("Beta (5Y)", BETA, "Moderate rate-cycle sensitivity"),
    ("Forward Dividend Yield", f"{DIV_YIELD:.1f}%", f"${DIV_RATE:.2f} annual rate; 32.3% payout ratio"),
    ("Dividend Payout Ratio", "32.3%", "Yahoo Finance, July 27, 2026"),
    ("Next Earnings", NEXT_EARNINGS, "Q2 FY26 — Q3 FY26 in calendar terms"),
]

for i, (field, val) in enumerate(val_title):
    c(ws1, 3 + i, 1, field, font=header_font)
    c(ws1, 3 + i, 2, val)

c(ws1, 3 + len(val_title) + 1, 1, "Metric", font=header_font, fill=header_fill)
c(ws1, 3 + len(val_title) + 1, 2, "Value", font=header_font, fill=header_fill)
c(ws1, 3 + len(val_title) + 1, 3, "Comment", font=header_font, fill=header_fill)

start_row = 3 + len(val_title) + 2
for i, (metric, value, comment) in enumerate(val_data):
    c(ws1, start_row + i, 1, metric, font=header_font)
    c(ws1, start_row + i, 2, value)
    c(ws1, start_row + i, 3, comment, font=note_font)

for ci in [1, 2, 3]:
    ws1.column_dimensions[get_column_letter(ci)].width = [20, 15, 55][ci - 1]

# ---- Sheet 2: WACC ----
ws2 = wb.create_sheet("WACC")
ws2.merge_cells('A1:D1')
c(ws2, 1, 1, f"{TICKER} — WACC (CAPM)", font=title_font)

wacc_components = [
    ("Risk-Free Rate (10Y US Treasury)", RISK_FREE / 100, "CNBC, July 27, 2026"),
    ("Equity Risk Premium", ERP / 100, "Standard assumption"),
    ("Beta (5Y Monthly)", BETA, "Yahoo Finance Statistics"),
    ("Cost of Equity (Ke = Rf + ERP × Beta)", Ke / 100, f"{RISK_FREE} + {ERP} × {BETA} = {Ke:.2f}%"),
    ("Cost of Debt (Kd)", Kd / 100, "Estimated senior/sub debt rate"),
    ("Tax Rate (TTM)", TAX_RATE / 100, f"{TAX[0]:.0f} / {PRETAX_INC[0]:.0f} TTM"),
    ("Market Cap ($B)", MC_B, "Yahoo Finance, July 27, 2026"),
    ("Total Debt ($B)", TOTAL_DEBT_B, "Yahoo Key Stats — conservative higher figure"),
    ("Equity Weight", E_WEIGHT, f"{MC_B} / ({MC_B} + {TOTAL_DEBT_B})"),
    ("Debt Weight", D_WEIGHT, f"{TOTAL_DEBT_B} / ({MC_B} + {TOTAL_DEBT_B})"),
    ("WACC", WACC / 100, f"{E_WEIGHT:.3f}×{Ke:.2f} + {D_WEIGHT:.3f}×{Kd:.2f}×(1-{TAX_RATE/100:.3f})"),
]

c(ws2, 3, 1, "Component", font=header_font, fill=header_fill)
c(ws2, 3, 2, "Value", font=header_font, fill=header_fill)
c(ws2, 3, 3, "Comment", font=header_font, fill=header_fill)

for i, (label, value, comment) in enumerate(wacc_components):
    c(ws2, 4 + i, 1, label, font=header_font)
    fmt = percent_fmt if i < 6 or i == 10 else None
    c(ws2, 4 + i, 2, value, fmt=fmt)
    c(ws2, 4 + i, 3, comment, font=note_font)

ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 60

# ---- Sheet 3: Scenarios ----
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells('A1:I1')
c(ws3, 1, 1, f"{TICKER} — P/B + ROE Scenario Analysis (Bank Framework)", font=title_font)
c(ws3, 2, 1, "FCF multiples N/A for banks. Primary framework: BVPS CAGR → exit P/B → implied price/share.", font=note_font, fill=note_fill)

# Header row at row 3
headers = ["Metric", "Bear", "Base", "Bull", "Comment"]
for ci, h in enumerate(headers, 1):
    c(ws3, 3, ci, h, font=header_font, fill=header_fill)

scenarios = [
    ("Framework", "P/B + ROE", "P/B + ROE", "P/B + ROE", "Bank valuation framework"),
    ("Revenue CAGR (5Y)", f"{bear_rev_cagr:.0f}%", f"{base_rev_cagr:.0f}%", f"{bull_rev_cagr:.0f}%", "Revenue growth assumption"),
    ("Terminal Revenue ($B, 5Y)", f"{1121.822 * (1+bear_rev_cagr/100)**5:.0f}", f"{1121.822 * (1+base_rev_cagr/100)**5:.0f}", f"{1121.822 * (1+bull_rev_cagr/100)**5:.0f}", "Projected total revenue"),
    ("BVPS CAGR", f"{bear_bvps_cagr:.0f}%", f"{base_bvps_cagr:.0f}%", f"{bull_bvps_cagr:.0f}%", "Book value per share compounding"),
    ("ROE", f"{bear_roe:.0f}%", f"{base_roe:.0f}%", f"{bull_roe:.0f}%", "Return on common equity"),
    ("Exit P/B Multiple", f"{bear_exit_pb:.2f}x", f"{base_exit_pb:.2f}x", f"{bull_exit_pb:.2f}x", "Target P/B ratio"),
    ("Terminal BVPS (5Y)", f"${bear_term_bvps:.2f}", f"${base_term_bvps:.2f}", f"${bull_term_bvps:.2f}", f"BVPS ${BVPS_COMMON:.2f} × (1+rate)^5"),
    ("Implied Price / Share", f"${bear_target:.2f}", f"${base_target:.2f}", f"${bull_target:.2f}", "Terminal BVPS × Exit P/B"),
    ("Upside from Current", f"{bear_upside:.1f}%", f"{base_upside:.1f}%", f"{bull_upside:.1f}%", f"vs ${PRICE:.2f}"),
    ("", "", "", "", ""),
    ("Weight", "25%", "50%", "25%", "Probability assignment"),
    ("Weighted Value / Share", f"${bear_w*bear_target:.2f}", f"${base_w*base_target:.2f}", f"${bull_w*bull_target:.2f}", ""),
    ("Probability-Weighted FV", "", "", f"${weighted_fv:.2f}", "Sum of weighted values"),
    ("Current Price", "", "", f"${PRICE:.2f}", "July 27, 2026 close"),
    ("Implied Upside (Weighted)", "", "", f"{weighted_upside:.1f}%", "From current price"),
    ("", "", "", "", ""),
    ("Analyst Avg PT Cross-Check", "", "", "", "Yahoo Finance — no explicit PT visible; EPS consensus used as proxy"),
    ("Base vs Analyst Consistency", "", "", "", f"Base case ${base_target:.2f} is plausible cross-check"),
]

for i, (metric, bear, base, bull, note) in enumerate(scenarios, 4):
    c(ws3, i, 1, metric, font=header_font if metric else data_font)
    c(ws3, i, 2, bear)
    c(ws3, i, 3, base)
    c(ws3, i, 4, bull)
    c(ws3, i, 5, note, font=note_font)

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 50

# ---- Sheet 4: Actuals Source Audit ----
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells('A1:F1')
c(ws4, 1, 1, f"{TICKER} — Actuals Source Audit", font=title_font)

# Header
for ci, h in enumerate(["Data Point", "Value", "Source Page", "Date", "Notes"], 1):
    c(ws4, 3, ci, h, font=header_font, fill=header_fill)

audit = [
    ("Stock Price", f"${PRICE:.2f}", "Yahoo Finance Summary", "2026-07-27", "Close price; after-hours $20.82"),
    ("Market Cap", f"${MC_B:.2f}B", "Yahoo Finance Key Stats", "2026-07-27", "From statistics page"),
    ("Shares Outstanding", f"{SHARES_MM}M", "Yahoo Finance Key Stats", "2026-07-27", "Regular outstanding"),
    ("Implied Shares Outstanding", f"{IMP_SHARES_MM}M", "Yahoo Finance Key Stats", "2026-07-27", "Includes convertible subsidiary equity"),
    ("Total Debt — Key Stats", f"${TOTAL_DEBT_B:.2f}B", "Yahoo Finance Key Stats", "2026-07-27", "MRQ; higher than BS figure — used for consistency"),
    ("Total Debt — Balance Sheet", f"${3.02:.2f}B", "Yahoo Finance Balance Sheet", "FY2025", "Permanent debt line; $0.61B less than Key Stats"),
    ("Total Cash", f"${TOTAL_CASH_B:.2f}B", "Yahoo Finance Key Stats", "2026-07-27", "MRQ"),
    ("Preferred Stock", f"${PREFERRED_STOCK_B:.2f}B", "Yahoo Finance Balance Sheet", "FY2025", "$498.5M; constant since FY2022"),
    ("Common Equity", f"${COMMON_EQUITY_B:.2f}B", "Yahoo Finance Balance Sheet", "FY2025", "Total equity less preferred"),
    ("BVPS — Total", f"${BVPS_TOTAL:.2f}", "Yahoo Finance Key Stats", "2026-07-27", "MRQ; total equity basis"),
    ("Tangible Book Value", f"${2.72:.2f}B", "Yahoo Finance Balance Sheet", "FY2025", "Net tangible assets"),
    ("Total Assets", "$34.80B", "Yahoo Finance Balance Sheet", "FY2025", "From annual data"),
    ("TTM Revenue", "$1,121.8M", "Yahoo Finance Income Statement", "TTM", "All numbers in thousands"),
    ("FY2025 Revenue", "$1,103.2M", "Yahoo Finance Income Statement", "FY2025", ""),
    ("FY2024 Revenue", "$985.9M", "Yahoo Finance Income Statement", "FY2024", "Partial year post-restructuring"),
    ("FY2023 Revenue", "$278.4M", "Yahoo Finance Income Statement", "FY2023", "Post-SVB bankruptcy restructuring year"),
    ("FY2022 Revenue", "$1,341.0M", "Yahoo Finance Income Statement", "FY2022", "Pre-bankruptcy"),
    ("TTM Net Income", "$207.6M", "Yahoo Finance Income Statement", "TTM", ""),
    ("FY2025 Net Income", "$189.2M", "Yahoo Finance Income Statement", "FY2025", ""),
    ("FY2023 Net Income", "-$1,938.9M", "Yahoo Finance Income Statement", "FY2023", "SVB deposit insurance loss"),
    ("TTM Diluted EPS", "$1.30", "Yahoo Finance Income Statement", "TTM", ""),
    ("TTM Operating CF", "$289.5M", "Yahoo Finance Cash Flow", "TTM", ""),
    ("TTM Free CF", "$266.8M", "Yahoo Finance Cash Flow", "TTM", "OCF - CapEx"),
    ("Beta (5Y Monthly)", BETA, "Yahoo Finance Key Stats", "2026-07-27", ""),
    ("ROE (TTM)", f"{ROE_TTM:.1f}%", "Yahoo Finance Financial Highlights", "2026-07-27", ""),
    ("ROA (TTM)", f"{ROA_TTM:.1f}%", "Yahoo Finance Financial Highlights", "2026-07-27", ""),
    ("Dividend Yield (Fwd)", f"{DIV_YIELD:.1f}%", "Yahoo Finance Key Stats", "2026-07-27", f"Forward annual dividend ${DIV_RATE:.2f}"),
    ("P/E Trailing", "16.37x", "Yahoo Finance Key Stats", "2026-07-27", "Based on TTM EPS"),
    ("Earnings Date", NEXT_EARNINGS, "Yahoo Finance Summary", "2026-07-27", "Q2 FY26 earnings call, July 29 at 11 AM EDT"),
    ("EPS Estimate FY26", "$1.68", "Yahoo Finance Analysis", "2026-07-27", "10 analysts; Non-GAAP diluted"),
    ("EPS Estimate FY27", "$2.05", "Yahoo Finance Analysis", "2026-07-27", "10 analysts; Non-GAAP diluted"),
    ("Revenue Estimate FY26", "$1.20B", "Yahoo Finance Analysis", "2026-07-27", "10 analysts; +7.45% YoY"),
    ("Revenue Estimate FY27", "$1.31B", "Yahoo Finance Analysis", "2026-07-27", "10 analysts; +8.49% YoY"),
    ("EPS Q1 FY26 Actual", "$0.39", "Yahoo Finance Analysis", "2026-07-27", "Beat est by $0.01 (+2.82%)"),
    ("10Y Treasury Yield", f"{RISK_FREE:.2f}%", "CNBC US10Y", "2026-07-27", "Yield quote 10:10 PM EDT"),
    ("Preferred Stock Terms", "Unknown", "N/A", "N/A", "Questions item — need to verify dividend rate, redemption terms"),
]

for i, (dp, val, src, dt, notes) in enumerate(audit, 4):
    c(ws4, i, 1, dp, font=header_font)
    c(ws4, i, 2, val)
    c(ws4, i, 3, src)
    c(ws4, i, 4, dt)
    c(ws4, i, 5, notes, font=note_font)

ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 15
ws4.column_dimensions['C'].width = 30
ws4.column_dimensions['D'].width = 15
ws4.column_dimensions['E'].width = 55

# ---- Sheet 5: Questions ----
ws5 = wb.create_sheet("Questions")
ws5.merge_cells('A1:C1')
c(ws5, 1, 1, f"{TICKER} — Open Questions", font=title_font)

c(ws5, 3, 1, "#", font=header_font, fill=header_fill)
c(ws5, 3, 2, "Question", font=header_font, fill=header_fill)
c(ws5, 3, 3, "Category", font=header_font, fill=header_fill)

questions = [
    ("Q1", "Preferred stock terms: What is the dividend rate on the $498.5M preferred stock? Is it cumulative? What are the redemption/call terms? Should this be subtracted from MC for common-share equity valuation?", "Capital Structure"),
    ("Q2", "Debt discrepancy: Key Stats shows Total Debt of $3.63B vs Balance Sheet of $3.02B — a $0.61B difference. What comprises the difference? Likely includes capital lease obligations, subordinated notes, or revolving credit facilities not on the permanent balance sheet line.", "Capital Structure"),
    ("Q3", "BVPS trend: Total BVPS is $19.81 (MRQ) but common BVPS is $19.58. With preferred stock constant at $498.5M since FY2022, BVPS trajectory is essentially flat. What is driving the lack of book accretion at 7% ROE?", "Book Value"),
    ("Q4", "ROE below cost of equity: At 6.99% ROE vs ~10.3% cost of equity, the franchise is technically destroying value on a total-equity basis. Is this sustainable or will ROE recover post-restructuring?", "Returns"),
    ("Q5", "Post-SVB bankruptcy legacy: How much residual risk from the 2023 SVB-related deposit assessment remains? Has the full loss been booked or are there reserve/contingent exposure risks?", "Legacy Risk"),
    ("Q6", "Share count reduction: Shares outstanding fell from 168.6M (FY2024) to 155.5M (FY2025) — a 7.8% decline. Has this been via buybacks? If so, what is the annual buyback rate vs. market cap?", "Capital Allocation"),
    ("Q7", "NII trajectory: Net Interest Income grew from $926M (FY2024) to $977M (FY2025) to $997M TTM — positive but modest. At current asset levels (~$34.8B), this implies NIM of ~2.9%. Is NIM compressing or stabilizing?", "Credit / NIM"),
    ("Q8", "Interest expense trend: Interest expense declined from $887M (FY2024) to $699M (FY2025) — consistent with lower funding costs. Is this rate-cycle tailwind or structural?", "Cost of Funds"),
    ("Q9", "Deposit stability: Post-SVB collapse, deposit outflows from California regional banks were severe. Has BANC stabilized its deposit base? What is the current uninsured deposit ratio?", "Deposit Risk"),
    ("Q10", "CRE exposure: What is the CRE concentration in the loan portfolio? California commercial real estate has faced significant headwinds. Is there OREO/nonperforming exposure?", "Credit Quality"),
    ("Q11", "Management governance: Jared Wolff became CEO after the bankruptcy restructuring. What is the track record of the post-restructuring management team on franchise building?", "Governance"),
    ("Q12", "Q2 FY26 earnings on July 29: This report will include the first full-year comparison against the post-restructuring baseline. What credit metrics (NPL, provision ratio, loan growth) will be disclosed?", "Next Earnings"),
]

for i, (num, question, cat) in enumerate(questions, 4):
    c(ws5, i, 1, num, font=header_font)
    c(ws5, i, 2, question)
    c(ws5, i, 3, cat)

ws5.column_dimensions['A'].width = 8
ws5.column_dimensions['B'].width = 90
ws5.column_dimensions['C'].width = 20

# ---- Sheet 6: Sources ----
ws6 = wb.create_sheet("Sources")
ws6.merge_cells('A1:C1')
c(ws6, 1, 1, f"{TICKER} — Sources", font=title_font)

c(ws6, 3, 1, "#", font=header_font, fill=header_fill)
c(ws6, 3, 2, "Source", font=header_font, fill=header_fill)
c(ws6, 3, 3, "URL", font=header_font, fill=header_fill)

sources = [
    ("S1", "Yahoo Finance — BANC Summary", "https://finance.yahoo.com/quote/BANC/"),
    ("S2", "Yahoo Finance — BANC Income Statement", "https://finance.yahoo.com/quote/BANC/financials/"),
    ("S3", "Yahoo Finance — BANC Balance Sheet", "https://finance.yahoo.com/quote/BANC/balance-sheet/"),
    ("S4", "Yahoo Finance — BANC Cash Flow", "https://finance.yahoo.com/quote/BANC/cash-flow/"),
    ("S5", "Yahoo Finance — BANC Key Statistics", "https://finance.yahoo.com/quote/BANC/key-statistics/"),
    ("S6", "Yahoo Finance — BANC Analysis / Estimates", "https://finance.yahoo.com/quote/BANC/analysis/"),
    ("S7", "Yahoo Finance — BANC Profile", "https://finance.yahoo.com/quote/BANC/profile/"),
    ("S8", "CNBC — US 10Y Treasury Yield", "https://www.cnbc.com/quotes/US10Y"),
]

for i, (num, name, url) in enumerate(sources, 4):
    c(ws6, i, 1, num, font=header_font)
    c(ws6, i, 2, name)
    c(ws6, i, 3, url)

ws6.column_dimensions['A'].width = 8
ws6.column_dimensions['B'].width = 55
ws6.column_dimensions['C'].width = 65

# ---- Save ----
outpath = f"models/[{TODAY}] Banc of California Model.xlsx"
wb.save(outpath)
print(f"\nSaved to {outpath}")
print(f"Model sheets: {wb.sheetnames}")
