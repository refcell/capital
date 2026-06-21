"""Add a BABA (Alibaba Group) tab to projections/Projections.xlsx.

Clones the last existing tab (ORCL) to preserve styling/formulas, then
overwrites the input cells with Alibaba numbers.

Alibaba's fiscal year ends March 31, so FY2026 (ended 3/31/2026) is already
reported.  The "2026" column is FY2026 actual and the model projects
FY2027-FY2031.  BABA trades as ADSs (each ADS = 8 ordinary shares) at ~$110;
everything here is in USD on a per-ADS basis: net income in USD millions,
share count = diluted ADS count (~2,400M), EPS = diluted earnings per ADS.

Basis decision: the stock's quoted valuation (TTM P/E ~19x, forward ~14x) is
struck on GAAP diluted EPS per ADS ($6.38) / net income attributable to ordinary
shareholders ($15.4B, a 10.4% net margin), so this tab uses GAAP net income
attributable.  Note Alibaba's Non-GAAP (core operating) net income collapsed 62%
in FY2026 to $8.8B as it funded a major investment cycle (quick-commerce
subsidies via Taobao Instant Commerce, AI model/Qwen user acquisition, and cloud
capex); GAAP was buoyed by mark-to-market equity gains.  Both measures are
temporarily distorted; the forward thesis is an earnings recovery as
quick-commerce losses narrow and Cloud Intelligence (revenue +40%, AI products
+triple digits) scales with operating leverage.  All $ figures in millions.

Sources (Q4/FY2026 results 5/13/2026; analyst consensus; quotes Jun 2026):
  - FY2026 revenue $148,401M (+8% Y/Y; +~10% like-for-like ex-disposals)
    GAAP net income $14,805M; net income attributable to ordinary $15,353M
    GAAP diluted EPS/ADS $6.38; net margin ~10.4%
    Non-GAAP net income $8,794M (-62%); Non-GAAP EPS/ADS $3.89 (-59%)
    Cloud Intelligence external revenue +40%; AI products ~30% of cloud revenue
    diluted ordinary shares ~19,200M (~2,400M ADS); cash/liquid inv. ~$75.5B
  - FY2025 revenue $137,300M; net income attributable ~$17.5B; EPS/ADS RMB55.12
  - Analyst consensus: revenue +~9-10% p.a.; net income +~17% FY2027;
    EPS growth ~12-14%/yr; forward P/E ~11-14x; FY2026 net margin ~10.5%
  - Price ~$107-115 (Jun 2026); market cap ~$288B; 52-wk $103.71-$192.67;
    42 analysts consensus Strong Buy, avg 12-mo PT ~$190 (high $225 / low $148)
  - Risks (bear): China macro/regulatory + geopolitics (DoD "Chinese Military
    Company" designation), e-commerce competition (PDD, JD, Meituan/quick
    commerce price war), prolonged investment drag on margins
"""

from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "projections" / "Projections.xlsx"

# Shared inputs (USD millions; shares = diluted ADS count in millions)
SHARES = 2400         # diluted ADS (~19.2B ordinary shares / 8); ~flat on buybacks
REV_2026 = 148401     # FY2026 actual total revenue
NI_2026 = 15400       # FY2026 actual GAAP net income attributable to ordinary
NI_G_2026 = -0.18     # net income / EPS-per-ADS down ~18-20% Y/Y in FY2026

# Analyst GAAP net-income-attributable estimates FY2026-FY2029 (USD M)
# (FY27 ~+17% to ~$18B; then ~12-14% EPS growth; EPS/ADS ~$7.5 / $8.8 / $10.0.)
ANALYST = {"C": 15400, "D": 18000, "E": 21000, "F": 24000}

# Per-case assumptions.  Each template block is offset by 24 rows:
#   Bull base row 4, Base 28, Bear 52.  Within a block, relative rows:
#     +2 revenue, +3 revenue growth, +5 net income, +6 NI growth,
#     +8 analyst est, +10 NI margins, +14 PE low, +15 PE high, H(+0) shares.
# NOTE: revenue growth applies with a one-column lag (D6 = C6*(1+C7)), so the
# column-C growth rate drives FY2027 revenue, column-D drives FY2028, etc.
# Revenue grows high-single/low-double digits (cloud/AI re-accelerating, China
# e-commerce low-single).  Net margin recovers off the FY2026 trough (~10% GAAP)
# as quick-commerce losses narrow and cloud scales (bull: faster monetization +
# multiple re-rating; bear: prolonged investment drag + China/geopolitical risk).
CASES = {
    "bull": {
        "base": 4,
        # cloud/AI re-acceleration sustains low-teens growth
        "rev_g": {"C": 0.13, "D": 0.13, "E": 0.12, "F": 0.11, "G": 0.10, "H": 0.09},
        "margins": {"D": 0.13, "E": 0.15, "F": 0.16, "G": 0.17, "H": 0.18},
        "pe_low": {c: 16 for c in "CDEFGH"},
        "pe_high": {c: 22 for c in "CDEFGH"},
    },
    "base": {
        "base": 28,
        # ~10% revenue growth; steady margin recovery toward mid-teens
        "rev_g": {"C": 0.10, "D": 0.10, "E": 0.09, "F": 0.08, "G": 0.08, "H": 0.07},
        "margins": {"D": 0.115, "E": 0.125, "F": 0.13, "G": 0.135, "H": 0.14},
        "pe_low": {c: 12 for c in "CDEFGH"},
        "pe_high": {c: 18 for c in "CDEFGH"},
    },
    "bear": {
        "base": 52,
        # competition + macro cap growth; investment drag persists on margins
        "rev_g": {"C": 0.07, "D": 0.06, "E": 0.05, "F": 0.05, "G": 0.04, "H": 0.04},
        "margins": {"D": 0.09, "E": 0.10, "F": 0.10, "G": 0.105, "H": 0.11},
        "pe_low": {c: 8 for c in "CDEFGH"},
        "pe_high": {c: 12 for c in "CDEFGH"},
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    src = wb[wb.sheetnames[-1]]          # clone the last tab (ORCL)
    ws = wb.copy_worksheet(src)
    ws.title = "BABA"

    # Header
    ws["B2"] = "BABA"
    ws["C2"] = datetime(2026, 6, 21)

    for cfg in CASES.values():
        b = cfg["base"]
        ws[f"H{b}"] = SHARES                      # diluted ADS count

        # Revenue base (FY2026 actual) + growth row
        ws[f"C{b + 2}"] = REV_2026
        for col, g in cfg["rev_g"].items():
            ws[f"{col}{b + 3}"] = g

        # GAAP net income attributable base (FY2026 actual) + base-year growth
        ws[f"C{b + 5}"] = NI_2026
        ws[f"C{b + 6}"] = NI_G_2026

        # Analyst estimates (FY2026-FY2029)
        for col, v in ANALYST.items():
            ws[f"{col}{b + 8}"] = v

        # NI margins (D..H; C stays as formula =NI/Rev)
        for col, m in cfg["margins"].items():
            ws[f"{col}{b + 10}"] = m

        # PE low / high
        for col, v in cfg["pe_low"].items():
            ws[f"{col}{b + 14}"] = v
        for col, v in cfg["pe_high"].items():
            ws[f"{col}{b + 15}"] = v

    wb.save(XLSX)
    print(f"Added BABA tab. Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
