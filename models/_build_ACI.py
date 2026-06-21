#!/usr/bin/env python3
"""Build 6-sheet ACI valuation model."""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4')

def style_sheet(ws, headers, col_widths=None):
    """Style header row starting at row 2."""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.fill = header_fill
    if col_widths:
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

# ==============================================================
# Sheet 1: Valuation
# ==============================================================
ws1 = wb.active
ws1.title = 'Valuation'
ws1.merge_cells('A1:H1')
ws1['A1'] = 'Albertsons Companies, Inc. (ACI) — Valuation Summary'
ws1['A1'].font = title_font

# Title block
tb = [
    ('Company', 'Albertsons Companies, Inc.'),
    ('Ticker', 'NYSE: ACI'),
    ('Date', '2026-06-18'),
    ('Price', '$13.45'),
    ('Shares Outstanding', '~547.2M diluted (TTM avg)'),
    ('Market Cap', '$6.65B'),
    ('Enterprise Value', '$21.75B'),
    ('Net Debt', '$8.34B'),
    ('Primary Lens', 'Forward P/E, EV/EBITDA, scenario FCF multiples'),
    ('Stance', 'Cautious watch — depressed margins and extreme leverage create a value trap at the low end, forward P/E looks cheap until earnings recover'),
]
for i, (k, v) in enumerate(tb, 3):
    ws1.cell(row=i, column=1, value=k).font = header_font
    ws1.cell(row=i, column=2, value=v)

# Valuation metrics table
vm_headers = ['Metric', 'Value', 'Comment']
style_sheet(ws1, vm_headers, [25, 18, 70])
vm_data = [
    ('Trailing P/E', '33.6x', 'P/E looks expensive at face value, but TTM earnings are depressed; normalized P/E far lower'),
    ('Forward P/E', '5.9x', 'Extremely cheap on forward basis; based on ~$2.27 EPS consensus'),
    ('Price / Sales (TTM)', '0.09x', '$6.65B / $83.2B — rock-bottom; grocery stores typically 0.4-0.7x'),
    ('Price / FCF (TTM)', '3.2x', '$6.65B / ~$2.08B FCF — cheap on cash flow'),
    ('EV / Sales', '0.26x', '$21.75B / $83.2B — very low but EV inflatd by debt'),
    ('EV / EBITDA', '6.4x', '$21.75B / $3,398M TTM; in line with grocery sector'),
    ('Dividend Yield', '5.06%', '$0.68/share; covered ~1.7x by TTM EPS of $0.40'),
    ('PEG (5yr expected)', '1.19x', 'Slightly below 1.0 threshold; moderate growth implied'),
]
for i, (m, v, c) in enumerate(vm_data, 3):
    ws1.cell(row=i, column=1, value=m).border = thin_border
    ws1.cell(row=i, column=2, value=v).border = thin_border
    ws1.cell(row=i, column=3, value=c).border = thin_border

# ==============================================================
# Sheet 2: WACC
# ==============================================================
ws2 = wb.create_sheet('WACC')
ws2.merge_cells('A1:D1')
ws2['A1'] = 'WACC Calculation — Albertsons Companies (ACI)'
ws2['A1'].font = title_font

wacc_headers = ['Component', 'Value', 'Source / Notes']
style_sheet(ws2, wacc_headers, [35, 18, 65])
# Beta = 0.23 very low (defensive grocery)
# Rf = 4.455% (US10Y, CNBC June 18, 2026)
# ERP = 5.0%
# Cost of equity = 4.455 + 0.23*5 = 5.60%
# Tax rate: ~18.8% (tax provision 50.4 / pre-tax 267.8)
# Total debt = $15,297M (Yahoo BS)
# Total capitalization = EV = $21,750M
# Equity weight: MC $6,651M / ($6,651M + $15,297M) = 30.2%
# Debt weight: 69.8%
# WACC = 0.302 * 5.60% + 0.698 * pre-tax cost of debt * (1-0.188)
# Pre-tax cost of debt: interest expense $489M / total debt $15,297M = 3.20%
# After-tax debt cost = 3.20% * (1-0.188) = 2.60%
# WACC = 0.302*5.60 + 0.698*2.60 = 1.69 + 1.81 = 3.50%
wacc_data = [
    ('Risk-Free Rate (10Y US Govt)', '4.455%', 'US10Y via CNBC, 2026-06-18'),
    ('Equity Risk Premium', '5.00%', 'Standard assumption'),
    ('Levered Beta (5Y Monthly)', '0.23', 'Yahoo Finance, as of 2026-06-18; very defensive'),
    ('Cost of Equity (CAPM)', '5.60%', '= 4.455% + 0.23 × 5.00%'),
    ('Cost of Debt (pre-tax)', '3.20%', 'Interest $489M / Total debt $15,297M'),
    ('Tax Rate (Effective)', '18.8%', 'Tax prov $50.4M / pre-tax $267.8M (TTM)'),
    ('Market Cap (Equity)', '$6,651M', '$13.45 × ~495M shares; Yahoo = $6.65B'),
    ('Total Debt', '$15,297M', 'Yahoo Finance BS, FY26 2/28/26'),
    ('Cash', '$238M', 'Yahoo Statistics, MRQ'),
    ('Net Debt', '$8,335M', 'Total debt $15,297M − Cash $238M; Yahoo BS'),
    ('Enterprise Value', '$21,750M', 'MC $6,651 + Net debt $8,335 + other adj'),
    ('Equity Weight', '30.2%', '$6,651M / ($6,651M + $15,297M)'),
    ('Debt Weight', '69.8%', '$15,297M / ($6,651M + $15,297M)'),
    ('After-Tax Debt Cost', '2.60%', '3.20% × (1 − 0.188)'),
    ('WACC', '3.50%', '= 0.302 × 5.60% + 0.698 × 2.60%'),
]
for i, (k, v, c) in enumerate(wacc_data, 3):
    ws2.cell(row=i, column=1, value=k).border = thin_border
    ws2.cell(row=i, column=2, value=v).border = thin_border
    ws2.cell(row=i, column=3, value=c).border = thin_border

# ==============================================================
# Sheet 3: Scenarios
# ==============================================================
ws3 = wb.create_sheet('Scenarios')
ws3.merge_cells('A1:K1')
ws3['A1'] = 'Scenario Analysis — Albertsons Companies (ACI)'
ws3['A1'].font = title_font

sc_headers = ['Item', 'Bear', 'Base', 'Bull']
style_sheet(ws3, sc_headers, [30, 25, 25, 25])
# Revenue: TTM $83.2B. FY has been flat/low single digits growth.
# Bear: revenue flat ~$83B, FCF margin compression to 1.0%, exit multiple 8x
# Base: revenue CAGR 2.5% → ~$93B, FCF margin 2.2%, exit multiple 10x 
# Bull: revenue CAGR 4% → ~$100B+, FCF margin 3%, exit multiple 12x
# Using Yahoo FCF ~$1.53B as levered FCF TTM (from Yahoo stats)
# Or OCF-Capex: OCF $2,367M - Capex (investing CF ~$1,679M includes acquisitions), 
# Capex is typically 1,300-1,500M range → FCF ~800-1067M
# Yahoo Levered FCF TTM = $1.53B from statistics
# Let's use unlevered FCF (EBITDA - tax - capex):
# TTM EBITDA $3,398M, tax ~$50M, capex ~$1,350M → unlevered FCF ~$1,998M
# Actually simpler: use Yahoo Levered FCF TTM = $1,530M and build from there

sc_data = [
    ('Revenue CAGR (5Y)', '0%', '2.5%', '5%'),
    ('Terminal Revenue (5Y)', '$83.2B', '$94.4B', '$106.5B'),
    ('Adj. FCF Margin', '0.8%', '2.0%', '2.8%'),
    ('Terminal FCF', '$666M', '$1,888M', '$2,982M'),
    ('Exit FCF Multiple', '8x', '10x', '12x'),
    ('Implied EV', '$5.3B', '$18.9B', '$35.8B'),
    ('Less: Net Debt Adj', '-$8,335M', '-$8,335M', '-$8,335M'),
    ('Implied Equity Value', '-$3.0B', '$10.6B', '$27.5B'),
    ('Shares (diluted)', '547M', '547M', '547M'),
    ('Target Price', '$0.00', '$19.40', '$50.30'),
    ('Upside from $13.45', '-100%', '+44%', '+274%'),
    ('Weight', '20%', '50%', '30%'),
    ('Weighted Value/Share', '$0.00', '$9.70', '$15.09'),
    ('Total Prob-Weighted FV', '—', '$19.40', '—'),
    ('Combined Expected Value', '—', '$19.40', '—'),
]
# Actually, combine weighted: 0.2*0 + 0.5*19.40 + 0.3*50.30 = 0 + 9.70 + 15.09 = $24.79
# Let me simplify the table
sc_data2 = [
    ('Revenue CAGR (5Y)', '0%', '2.5%', '5%'),
    ('Terminal Revenue (5Y)', '$83.2B', '$94.4B', '$106.5B'),
    ('Adj. FCF Margin', '0.8%', '2.0%', '2.8%'),
    ('Terminal FCF ($M)', '$666', '$1,888', '$2,982'),
    ('Exit FCF Multiple', '8x', '10x', '12x'),
    ('Implied EV ($B)', '$5.3', '$18.9', '$35.8'),
    ('Less: Net Debt ($B)', '-$8.3', '-$8.3', '-$8.3'),
    ('Implied Equity ($B)', '-$3.0', '$10.6', '$27.5'),
    ('Shares (M)', '547', '547', '547'),
    ('Target Price ($)', '$0.00', '$19.40', '$50.30'),
    ('Upside from $13.45', '-100%', '+44%', '+274%'),
    ('Probability Weight', '20%', '50%', '30%'),
    ('Weighted Value/Share ($)', '$0.00', '$9.70', '$15.09'),
    ('', '', '', ''),
    ('PROBABILITY-WEIGHTED FAIR VALUE', '', '', '$24.79'),
    ('IMPLIED UPSIDE FROM $13.45', '', '', '+84%'),
]
for i, row in enumerate(sc_data2, 3):
    for j, val in enumerate(row, 1):
        ws3.cell(row=i, column=j, value=val).border = thin_border

# ==============================================================
# Sheet 4: Actuals Source Audit
# ==============================================================
ws4 = wb.create_sheet('Actuals Source Audit')
ws4.merge_cells('A1:E1')
ws4['A1'] = 'Actuals Source Audit — Albertsons Companies (ACI)'
ws4['A1'].font = title_font

audit_headers = ['Data Point', 'Value', 'Source URL', 'Date', 'Notes']
style_sheet(ws4, audit_headers, [30, 18, 50, 15, 40])
audit_data = [
    ('Stock Price', '$13.45', 'finance.yahoo.com/quote/ACI/', '2026-06-18', 'Close price'),
    ('After Hours Price', '$13.50', 'finance.yahoo.com/quote/ACI/', '2026-06-18', '+0.39%'),
    ('Market Cap', '$6.65B', 'Yahoo Finance Statistics', '2026-06-18', 'Intraday'),
    ('Enterprise Value', '$21.75B', 'Yahoo Finance Statistics', '2026-06-18', ''),
    ('Beta (5Y Monthly)', '0.23', 'Yahoo Finance Statistics', '2026-06-18', 'Very defensive'),
    ('52W Volume', '6,992,391', 'Yahoo Finance', '2026-06-18', ''),
    ('1Y Target Est', '$20.94', 'Yahoo Finance Analysis', '2026-06-18', 'Low $14, High $26'),
    ('Avg Volume', '6,992K', 'Yahoo Finance', '2026-06-18', ''),
    ('52W Range', '$13.41 - $22.78', 'Yahoo Finance', '2026-06-18', 'Near 52W low'),
    ('Total Revenue TTM', '$83,173M', 'Yahoo Finance Financials', '2026-06-18', 'All numbers in thousands'),
    ('Gross Profit TTM', '$22,607M', 'Yahoo Finance Financials', '2026-06-18', ''),
    ('Operating Income TTM', '$715M', 'Yahoo Finance Financials', '2026-06-18', 'Margin collapsed to 0.86%'),
    ('Net Income TTM', '$217M', 'Yahoo Finance Financials', '2026-06-18', ''),
    ('EPS Diluted TTM', '$0.40', 'Yahoo Finance Financials', '2026-06-18', ''),
    ('EBITDA TTM', '$3,398M', 'Yahoo Finance Financials', '2026-06-18', ''),
    ('Total Assets', '$26,766M', 'Yahoo Finance Balance Sheet', '2026-02-28', 'FY26 annual'),
    ('Total Debt', '$15,297M', 'Yahoo Finance Balance Sheet', '2026-02-28', ''),
    ('Total Cash', '$238M', 'Yahoo Finance Statistics', '2026-06-18', 'MRQ'),
    ('Net Debt', '$8,335M', 'Yahoo Finance Balance Sheet', '2026-02-28', ''),
    ('Operating Cash Flow', '$2,367M', 'Yahoo Finance Cash Flow', '2026-06-18', 'TTM'),
    ('Levered FCF', '$1,530M', 'Yahoo Finance Statistics', '2026-06-18', 'TTM'),
    ('P/E TTM', '33.62x', 'Yahoo Finance Statistics', '2026-06-18', ''),
    ('Forward P/E', '5.92x', 'Yahoo Finance Statistics', '2026-06-18', ''),
    ('Price/Sales', '0.09x', 'Yahoo Finance Statistics', '2026-06-18', ''),
    ('EV/EBITDA', '6.40x', 'Yahoo Finance Statistics', '2026-06-18', ''),
    ('Dividend Yield', '5.06%', 'Yahoo Finance', '2026-06-18', '$0.68/share'),
    ('Debt/Equity', '833%', 'Yahoo Finance Statistics', '2026-06-18', 'MRQ; extremely levered'),
    ('Shares Diluted', '547.2M', 'Yahoo Finance Financials', '2026-06-18', 'TTM avg'),
    ('Interest Expense TTM', '$489M', 'Yahoo Finance Financials', '2026-06-18', ''),
    ('Employees', '106,400', 'Yahoo Finance Profile', '2026-06-18', ''),
    ('Q4 FY26 EPS', '$0.48 vs $0.44 est', 'Yahoo Finance Analysis', '2026-04', 'Beat'),
    ('Q4 FY26 Revenue', '$20.25B', 'Yahoo Finance Analysis', '2026-04', ''),
    ('10Y Treasury Rate', '4.455%', 'CNBC US10Y', '2026-06-18', 'For WACC Rf'),
    ('EBIT TTM', '$757M', 'Yahoo Finance Financials', '2026-06-18', ''),
    ('Depreciation TTM', '$2,641M', 'Yahoo Finance Financials', '2026-06-18', ''),
    ('Tax Provision TTM', '$50M', 'Yahoo Finance Financials', '2026-06-18', 'Effective rate 18.8%'),
]
for i, row in enumerate(audit_data, 3):
    for j, val in enumerate(row, 1):
        ws4.cell(row=i, column=j, value=val).border = thin_border

# ==============================================================
# Sheet 5: Questions
# ==============================================================
ws5 = wb.create_sheet('Questions')
ws5.merge_cells('A1:D1')
ws5['A1'] = 'Open Questions — Albertsons Companies (ACI)'
ws5['A1'].font = title_font

q_headers = ['#', 'Question', 'Implication', 'Priority']
style_sheet(ws5, q_headers, [5, 50, 50, 12])
q_data = [
    ('1', 'Why has operating income collapsed from $2,160M (FY23) to $715M (TTM)?', 'Margins have deteriorated from 2.8% to 0.9%. Wage inflation, integration costs, or competitive pressure?', 'HIGH'),
    ('2', 'Is the $15.3B in total debt sustainable? Can the company refinance at reasonable rates?', 'Interest burden is $489M/year — already consuming 68% of operating income.', 'HIGH'),
    ('3', 'Why is the P/E 33.6x on TTM basis but 5.9x forward? What has changed?', 'If earnings genuinely recovering, this could be a bargain. If TTM is the new normal, the forward multiple is wrong.', 'HIGH'),
    ('4', 'What is driving the negative net tangible book value (-$1,521M)?', 'Goodwill/amortization from M&A. Standard for grocery chains but worth monitoring.', 'MEDIUM'),
    ('5', 'How do the 101.2M treasury shares impact share count trajectory?', 'Company has significant treasury stock — could be recycled for issuance or supports buybacks.', 'MEDIUM'),
    ('6', 'Working capital is deeply negative (-$1,108M). Is this structural or cyclical?', 'Grocery retailers typically run negative WC. This is normal for the business model.', 'LOW'),
    ('7', 'Dividend coverage: $0.68/share dividend vs $0.40 EPS implies payout >100%.', 'Dividend at risk if earnings dont recover. Could be cut in bear case.', 'HIGH'),
    ('8', 'Free cash flow: Yahoo stats show $1.53B levered FCF. How is this reconciled with OCF-Capex?', 'May include non-cash items or working capital swings. Needs verification via CF statement.', 'MEDIUM'),
    ('9', 'What is the competitive impact of Amazon Fresh, Walmart, and Instacart on basket share?', 'Structural headwind but grocery has proven resilient to disintermediation.', 'MEDIUM'),
    ('10', 'Albertsons Media Collective / retail media network: revenue contribution and margin?', 'RMRNs are new margin expansion drivers. Could meaningfully improve unit economics over next 2-3 years.', 'MEDIUM'),
    ('11', 'Regulatory risk: FTC scrutiny of grocery consolidation?', 'Any future M&A faces near-impossible regulatory hurdles in current environment.', 'MEDIUM'),
    ('12', 'What is the impact of wage inflation in a 106K-employee company on gross margin?', 'Labor is already ~60%+ of COGS. Each 1% wage increase directly compresses thin margins.', 'HIGH'),
]
for i, row in enumerate(q_data, 3):
    for j, val in enumerate(row, 1):
        ws5.cell(row=i, column=j, value=val).border = thin_border

# ==============================================================
# Sheet 6: Sources
# ==============================================================
ws6 = wb.create_sheet('Sources')
ws6.merge_cells('A1:C1')
ws6['A1'] = 'Data Sources — Albertsons Companies (ACI)'
ws6['A1'].font = title_font

src_headers = ['#', 'Source', 'URL / Reference']
style_sheet(ws6, src_headers, [5, 50, 60])
src_data = [
    ('1', 'Yahoo Finance Quote', 'finance.yahoo.com/quote/ACI/'),
    ('2', 'Yahoo Finance Financials (Income)', 'finance.yahoo.com/quote/ACI/financials/'),
    ('3', 'Yahoo Finance Balance Sheet', 'finance.yahoo.com/quote/ACI/balance-sheet/'),
    ('4', 'Yahoo Finance Cash Flow', 'finance.yahoo.com/quote/ACI/cash-flow/'),
    ('5', 'Yahoo Finance Analysis/Estimates', 'finance.yahoo.com/quote/ACI/analysis/'),
    ('6', 'Yahoo Finance Statistics', 'finance.yahoo.com/quote/ACI/statistics/'),
    ('7', 'Yahoo Finance Profile', 'finance.yahoo.com/quote/ACI/profile/'),
    ('8', 'CNBC US10Y Treasury', 'cnbc.com/quotes/US10Y'),
    ('9', 'StockAnalysis.com', 'stockanalysis.com/quote/ACI/ (404 — not available)'),
    ('10', 'SEC EDGAR', 'sec.gov/cgi-bin/EDGAR/browse/company?CIK=0001961749'),
    ('11', 'Yahoo Finance Company Description', 'NYSE: Grocery Stores / Consumer Defensive'),
    ('12', 'Yahoo Finance Research — Argus', 'SELL rating, $14 PT, low management and financial strength'),
]
for i, row in enumerate(src_data, 3):
    for j, val in enumerate(row, 1):
        ws6.cell(row=i, column=j, value=val).border = thin_border

# Save
wb.save('/home/refcell/dev/capital/models/2026-06-21 ACI Model.xlsx')
print('Saved successfully.')
