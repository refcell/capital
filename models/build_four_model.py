import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()

title_font = Font(name="Calibri", size=14, bold=True, color="e94560")
header_font = Font(name="Calibri", size=11, bold=True, color="e94560")
normal_font = Font(name="Calibri", size=10, color="e0e0e0")
section_font = Font(name="Calibri", size=11, bold=True, color="e94560")
accent_fill = PatternFill(start_color="0f3460", end_color="0f3460", fill_type="solid")
pct_fmt = '0.0%'
dollar_fmt = '#,##0'
dollar2_fmt = '#,##0.0'

thin = Border(left=Side(style='thin', color='333333'), right=Side(style='thin', color='333333'),
              top=Side(style='thin', color='333333'), bottom=Side(style='thin', color='333333'))

def sc(ws, r, c, fill=None):
    cell = ws.cell(row=r, column=c)
    cell.font = normal_font
    if fill:
        cell.fill = fill
    cell.border = thin
    return cell

def hdr(ws, r, n=15):
    for c in range(1, n + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = header_font
        cell.fill = accent_fill
        cell.border = thin

# ========== SHEET 1: Valuation ==========
ws = wb.active
ws.title = "Valuation"
ws.merge_cells('A1:O1')
ws['A1'].value = "Shift4 Payments (FOUR) Valuation Model"
ws['A1'].font = title_font

# Key inputs
ws['A3'].value = "Ticker"
ws['B3'].value = "FOUR"
ws['A4'].value = "Exchange"
ws['B4'].value = "NYSE"
ws['A5'].value = "CIK"
ws['B5'].value = "0001794669"
ws['A6'].value = "Model date"
ws['B6'].value = "2026-05-28"
ws['A7'].value = "Quote used"
ws['B7'].value = "2026-05-22 close"
ws['A8'].value = "Units"
ws['B8'].value = "USD millions except per-share"
for r in range(3, 9):
    for c in [1, 2]:
        sc(ws, r, c)

ws['D3'].value = "Current price"
ws['E3'].value = 43.24
ws['D4'].value = "Shares out (M)"
ws['E4'].value = 81.0
ws['D5'].value = "Market cap ($M)"
ws['E5'].value = "=E3*E4"
ws['D6'].value = "Cash & ST investments ($M)"
ws['E6'].value = 964.0
ws['D7'].value = "Debt principal ($M)"
ws['E7'].value = 4589.0
ws['D8'].value = "Mandatory conv preferred ($M)"
ws['E8'].value = 1000.0
ws['D9'].value = "Net debt ($M)"
ws['E9'].value = "=E7-E6"
ws['D10'].value = "Enterprise value ($M)"
ws['E10'].value = "=E5+E9+E8"
for r in range(3, 11):
    for c in [4, 5]:
        sc(ws, r, c)

# TTM multiples
ws['D12'].value = "TTM Gross revenue ($M)"
ws['E12'].value = 4452    # approx
ws['D13'].value = "FY2025 Adj EBITDA ($M)"
ws['E13'].value = 970
ws['D14'].value = "EV / Adj EBITDA (FY2025)"
ws['E14'].value = "=E10/E13"
for r in [12, 13, 14]:
    sc(ws, r, 4)
    sc(ws, r, 5)

# Historical actuals
r0 = 16
years = ["", "FY2022", "FY2023", "FY2024", "FY2025", "Q1'26", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
for c, y in enumerate(years, 1):
    ws.cell(row=r0, column=c, value=y)
hdr(ws, r0, len(years))

rows = [
    ("Gross revenue ($M)", [1990, 2570, 3330, 4180, 1121, 5200, 5890, 6480, 7060, 7628]),
    ("YoY revenue growth", [None, 29.1, 29.6, 34.5, 25.5, 24.4, 13.3, 10.0, 8.9, 8.0]),
    ("GRLNF ($M)", [None, None, None, 1981, 549, 2550, 2945, 3240, 3532, 3815]),
    ("GRLNF growth", [None, None, None, None, None, 29.0, 15.5, 10.0, 9.0, 8.0]),
    ("Adj EBITDA ($M)", [None, None, 678, 970, 234, 1190, 1377, 1513, 1648, 1783]),
    ("Adj EBITDA margin (of GRLNF)", [None, None, None, 0.49, 0.427, 0.467, 0.468, 0.467, 0.467, 0.467]),
    ("GAAP net income ($M)", [None, None, 295, 147, 12, 180, 220, 270, 320, 380]),
    ("Adj FCF ($M)", [None, None, None, 500, None, 500, 578, 635, 692, 749]),
    ("Payment volume ($B)", [76.2, 109.9, 165.0, 209.0, 56.0, None, None, None, None, None]),
]

for i, (label, vals) in enumerate(rows):
    r = r0 + 1 + i
    ws.cell(row=r, column=1, value=label)
    sc(ws, r, 1)
    for c, v in enumerate(vals, 2):
        ws.cell(row=r, column=c, value=v)
        cell = sc(ws, r, c)
        if v is not None and isinstance(v, float) and v < 1:
            cell.number_format = '0.0%'

# Balance sheet & capital structure
r2 = r0 + 12
ws.cell(row=r2, column=1, value="Balance Sheet & Capital Structure")
ws.cell(row=r2, column=1).font = section_font

bs_rows = [
    ("Cash ($M)", [None, None, None, 964, None, None, None, None, None]),
    ("Debt principal ($M)", [None, None, None, 4589, None, None, None, None, None]),
    ("Net debt ($M)", [None, None, None, 3625, None, None, None, None, None]),
    ("Preferred ($M)", [None, None, None, 1000, None, None, None, None, None]),
    ("Shares out (M)", [None, None, None, 81, None, 77.8, 74.7, 71.7, 68.8, 66.0]),
]
for i, (label, vals) in enumerate(bs_rows):
    r = r2 + 1 + i
    ws.cell(row=r, column=1, value=label)
    sc(ws, r, 1)
    for c, v in enumerate(vals, 2):
        ws.cell(row=r, column=c, value=v)
        sc(ws, r, c)

# Management guidance
r3 = r2 + 8
ws.cell(row=r3, column=1, value="Management Guidance (FY2026)")
ws.cell(row=r3, column=1).font = section_font
mgmt = [
    ("Gross revenue consensus", "$5.20B", "+24.4%"),
    ("GRLNF guide", "$2.50B - $2.60B", "+26% to +31%"),
    ("Adj EBITDA guide", "$1.165B - $1.215B", "+20% to +25%"),
    ("Adj FCF guide", "$490M - $510M", "~flat vs FY2025"),
    ("Payment volume guide", "$240B - $260B", "+15% to +24%"),
    ("Non-GAAP EPS guide", "$5.50 - $5.70", ""),
]
for i, (label, val, growth) in enumerate(mgmt):
    r = r3 + 1 + i
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=val)
    ws.cell(row=r, column=3, value=growth)
    for c in [1, 2, 3]:
        sc(ws, r, c)

# Analyst consensus
r4 = r3 + 9
ws.cell(row=r4, column=1, value="Analyst Consensus")
ws.cell(row=r4, column=1).font = section_font
analyst = [
    ("Num analysts", "24"),
    ("Consensus", "Buy"),
    ("Avg price target", "$60.81 (StockAnalysis) / $71 (MarketBeat)"),
    ("Low target", "$40"),
    ("High target", "$120"),
    ("Short interest", "~15M shares (~18.5% of float)"),
]
for i, (label, val) in enumerate(analyst):
    r = r4 + 1 + i
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=val)
    sc(ws, r, 1)
    sc(ws, r, 2)

# Forward assumptions
r5 = r4 + 9
ws.cell(row=r5, column=1, value="Forward Assumptions (Base Case)")
ws.cell(row=r5, column=1).font = section_font
assumptions = [
    ("GRLNF margin (adj EBITDA / GRLNF)", 0.467),
    ("Adj FCF / adj EBITDA conversion", 0.42),
    ("Annual interest expense ($M)", 250),
    ("Preferred dividends ($M)", 60),
    ("Tax rate", 0.21),
    ("Share count annual decline", -0.04),
    ("Discount rate (WACC)", 0.115),
    ("Terminal FCF yield", 0.0575),
]
for i, (label, val) in enumerate(assumptions):
    r = r5 + 1 + i
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=val)
    sc(ws, r, 1)
    cell = sc(ws, r, 2)
    if isinstance(val, float) and val < 1:
        cell.number_format = pct_fmt

# Owner earnings bridge
r6 = r5 + 11
ws.cell(row=r6, column=1, value="Owner Earnings Bridge (Base Case)")
ws.cell(row=r6, column=1).font = section_font
oe_years = ["", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
for c, y in enumerate(oe_years, 1):
    ws.cell(row=r6 + 1, column=c, value=y)
hdr(ws, r6 + 1, len(oe_years))

oe_data = [
    ("Adj FCF ($M)", [500, 578, 635, 692, 749]),
    ("Less: Interest ($M)", [250, 240, 230, 220, 210]),
    ("Less: Preferred div ($M)", [60, 55, 50, 45, 40]),
    ("Owner earnings ($M)", [190, 283, 355, 427, 499]),
    ("Shares out (M)", [77.8, 74.7, 71.7, 68.8, 66.0]),
    ("Owner earnings / share", [2.44, 3.79, 4.95, 6.21, 7.56]),
]

for i, (label, vals) in enumerate(oe_data):
    r = r6 + 2 + i
    ws.cell(row=r, column=1, value=label)
    sc(ws, r, 1)
    for c, v in enumerate(vals, 2):
        ws.cell(row=r, column=c, value=v)
        cell = sc(ws, r, c)
        if v is not None:
            cell.number_format = dollar2_fmt

# Terminal valuation
r7 = r6 + 10
ws.cell(row=r7, column=1, value="Terminal Valuation (Base Case)")
ws.cell(row=r7, column=1).font = section_font
term = [
    ("FY2030 Adj FCF ($M)", 749),
    ("Terminal FCF yield", 0.0575),
    ("Implied terminal EV ($M)", 13026),
    ("FY2030 net debt & preferred adj ($M)", -2500),
    ("Implied terminal equity ($M)", 10526),
    ("FY2030 shares (M)", 66.0),
    ("Implied price / share", 159.5),
    ("5Y CAGR from $43.24", 0.240),
]
for i, (label, val) in enumerate(term):
    r = r7 + 1 + i
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=val)
    sc(ws, r, 1)
    cell = sc(ws, r, 2)
    if isinstance(val, float) and val < 1:
        cell.number_format = pct_fmt

# ========== SHEET 2: WACC ==========
ws = wb.create_sheet("WACC")
wacc_data = [
    ("WACC / Discount Rate", None, None),
    ("Risk-free rate", 0.045, "FRED DGS10, 2026-05-26 observation."),
    ("Equity risk premium", 0.050, "Public-market ERP; elevated for levered name."),
    ("Beta", 1.15, "StockAnalysis/Yahoo 5Y monthly beta; elevated for payments leverage."),
    ("Cost of equity", None, "Risk-free + beta * ERP."),
    ("Pre-tax debt/pref cost", 0.065, "Blended cost for senior notes, convertibles, preferred."),
    ("Tax rate", 0.21, "Normalized statutory cash tax assumption."),
    ("Market value equity", None, "Valuation!E5."),
    ("Debt + preferred", None, "Debt principal + preferred."),
    ("Equity weight", None, "Market equity / enterprise capital."),
    ("Debt weight", None, "(Debt + preferred) / enterprise capital."),
    ("Mechanical WACC", None, "Cost of equity and after-tax debt weighted."),
    ("Selected discount rate", 0.115, "Rounded above mechanical WACC for leverage / M&A / GAAP complexity."),
]

for i, (label, val, note) in enumerate(wacc_data):
    r = i + 1
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=3, value=note)
    sc(ws, r, 1)
    sc(ws, r, 3)
    if val is not None:
        c = ws.cell(row=r, column=2, value=val)
        sc(ws, r, 2)
        if isinstance(val, float) and val < 1:
            c.number_format = pct_fmt

# Formulas
ws['B5'].value = "=B2+B3*B4"
ws['B8'].value = "=Valuation!E5"
ws['B9'].value = "=Valuation!E7+Valuation!E8"
ws['B10'].value = "=B8/(B8+B9)"
ws['B11'].value = "=B9/(B8+B9)"
ws['B12'].value = "=B10*B5+B11*B6*(1-B7)"

for r in [5, 8, 9, 10, 11, 12]:
    sc(ws, r, 2)

# ========== SHEET 3: Scenarios ==========
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells('A1:E1')
ws3['A1'].value = "Bear / Base / Bull Scenario Summary"
ws3['A1'].font = title_font

ws3['A3'].value = "Current price"
ws3['B3'].value = "=Valuation!E3"
ws3['A4'].value = "Current shares (M)"
ws3['B4'].value = "=Valuation!E4"

sc_row = 7
sc_headers = ["", "Bear", "Base", "Bull", "Notes"]
for c, h in enumerate(sc_headers, 1):
    ws3.cell(row=sc_row, column=c, value=h)
hdr(ws3, sc_row, 5)

scenarios = [
    ("Narrative",
     "Acquired growth normalizes, organic stalls. Adj FCF ~$450M. Leverage elevated.",
     "Guidance hits, organic 8-12%. FCF rises with deleveraging.",
     "Strong organic, Global Blue synergies, FCF compounds with buybacks.",
     ""),
    ("FY2026 rev growth", 0.08, 0.24, 0.28, "FY2026 vs FY2025"),
    ("FY2027-2030 rev CAGR", 0.05, 0.10, 0.14, "Annualized"),
    ("FY2030 revenue ($M)", 5180, 6600, 8050, "Terminal year"),
    ("FY2030 adj FCF margin (of rev)", 0.07, 0.098, 0.110, "Adj FCF / gross revenue"),
    ("FY2030 adj FCF ($M)", 363, 647, 886, ""),
    ("Exit adj FCF yield", 0.065, 0.0575, 0.050, "Terminal yield"),
    ("Exit enterprise value ($M)", 5580, 11250, 17700, "= FCF / yield"),
    ("Net debt & pref adj ($M)", -3000, -2500, -2000, "Cash - debt - pref"),
    ("FY2030 shares (M)", 78.0, 66.0, 60.0, "After buybacks"),
    ("FY2030 target price ($)", 33, 113, 215, "Equity / shares"),
    ("5Y CAGR from current", -0.058, 0.190, 0.338, "Annualized return"),
]

for i, (label, bear, base, bull, note) in enumerate(scenarios):
    r = sc_row + 1 + i
    ws3.cell(row=r, column=1, value=label)
    ws3.cell(row=r, column=2, value=bear)
    ws3.cell(row=r, column=3, value=base)
    ws3.cell(row=r, column=4, value=bull)
    ws3.cell(row=r, column=5, value=note)
    for c in range(1, 6):
        sc(ws3, r, c)
    # Pct format for rate rows
    if label in ("FY2026 rev growth", "FY2027-2030 rev CAGR", "FY2030 adj FCF margin (of rev)",
                 "Exit adj FCF yield", "5Y CAGR from current"):
        for c in [2, 3, 4]:
            ws3.cell(row=r, column=c).number_format = pct_fmt

# Scenario drivers
dd_row = sc_row + 14
ws3.cell(row=dd_row, column=1, value="Scenario Drivers")
ws3.cell(row=dd_row, column=1).font = section_font

drivers = [
    ("Metric", "Bear", "Base", "Bull"),
    ("Organic GRLNF growth", "<5%", "8-12%", ">12%"),
    ("GAAP/adjusted gap", "Persists", "Narrows slowly", "Narrows significantly"),
    ("Deleveraging", "Stagnant", "Gradual", "Aggressive + buybacks"),
    ("Global Blue", "Disappoint", "Integration works", "Synergy unlock"),
    ("Take-rate / spread", "Compression", "Stable", "Improves with mix"),
    ("Competition", "Erode share", "Durable niche", "Category leader"),
]

for i, (label, bear, base, bull) in enumerate(drivers):
    r = dd_row + 1 + i
    ws3.cell(row=r, column=1, value=label)
    ws3.cell(row=r, column=2, value=bear)
    ws3.cell(row=r, column=3, value=base)
    ws3.cell(row=r, column=4, value=bull)
    if i == 0:
        hdr(ws3, r, 4)
    for c in range(1, 5):
        sc(ws3, r, c)

# ========== SHEET 4: Actuals Source Audit ==========
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.cell(row=1, column=1, value="Line item")
ws4.cell(row=1, column=2, value="Period")
ws4.cell(row=1, column=3, value="Model value")
ws4.cell(row=1, column=4, value="Unit")
ws4.cell(row=1, column=5, value="Primary source")
ws4.cell(row=1, column=6, value="Secondary source")
ws4.cell(row=1, column=7, value="Difference / resolution")
hdr(ws4, 1, 7)

audit = [
    ("Gross revenue", "FY2023", 2570, "USD mm", "FOUR.md / 10-K", "StockAnalysis", "Ties."),
    ("Gross revenue", "FY2024", 3330, "USD mm", "FOUR.md / 10-K", "StockAnalysis", "Ties."),
    ("Gross revenue", "FY2025", 4180, "USD mm", "FOUR.md / 10-K", "StockAnalysis", "Ties."),
    ("Gross revenue", "Q1 2026", 1121, "USD mm", "Q1 2026 release", "StockAnalysis", "Ties."),
    ("GRLNF", "FY2025", 1981, "USD mm", "FOUR.md", "StockAnalysis", "Full-year FY2025."),
    ("GRLNF", "Q1 2026", 549, "USD mm", "Q1 2026 release", "StockAnalysis", "Q1 actual."),
    ("Adj EBITDA", "FY2024", 678, "USD mm", "FOUR.md", "StockAnalysis", "Ties."),
    ("Adj EBITDA", "FY2025", 970, "USD mm", "FOUR.md / 10-K", "StockAnalysis", "Ties."),
    ("Adj EBITDA", "Q1 2026", 234, "USD mm", "Q1 2026 release", "StockAnalysis", "Ties."),
    ("GAAP net income", "FY2024", 295, "USD mm", "FOUR.md / 10-K", "StockAnalysis", "Ties."),
    ("GAAP net income", "FY2025", 147, "USD mm", "FOUR.md / 10-K", "StockAnalysis", "Ties."),
    ("GAAP net income", "Q1 2026", 12, "USD mm", "Q1 2026 release", "StockAnalysis", "Ties."),
    ("Debt principal", "FY2025 YE", 4589, "USD mm", "FOUR.md / 10-K", "StockAnalysis", "All principal."),
    ("Cash", "FY2025 YE", 964, "USD mm", "FOUR.md / 10-K", "StockAnalysis", "Cash & equivalents."),
    ("Preferred", "Issued 2025", 1000, "USD mm", "FOUR.md / 10-K", "StockAnalysis", "6.00% mconv."),
    ("Payment volume", "FY2025", 209.0, "USD B", "FOUR.md", "StockAnalysis", "Full-year."),
    ("Payment volume", "Q1 2026", 56.0, "USD B", "Q1 release", "StockAnalysis", "Q1 actual."),
    ("Shares outstanding", "2026-05-22", 81.0, "M", "StockAnalysis", "Yahoo", "Post-buyback approx."),
]

for i, (item, period, val, unit, primary, secondary, diff) in enumerate(audit):
    r = i + 2
    ws4.cell(row=r, column=1, value=item)
    ws4.cell(row=r, column=2, value=period)
    ws4.cell(row=r, column=3, value=val)
    ws4.cell(row=r, column=4, value=unit)
    ws4.cell(row=r, column=5, value=primary)
    ws4.cell(row=r, column=6, value=secondary)
    ws4.cell(row=r, column=7, value=diff)
    for c in range(1, 8):
        sc(ws4, r, c)

# ========== SHEET 5: Questions ==========
ws5 = wb.create_sheet("Questions")
ws5.cell(row=1, column=1, value="Q Bank")
ws5.cell(row=1, column=1).font = title_font

questions = [
    "Is the GAAP/adjusted earnings gap structural or transitional? Can GAAP net income meaningfully close to adjusted EBITDA?",
    "How much of GRLNF growth is truly organic vs. acquired? Is SkyTab and organic hospitality growth compounding?",
    "Can Global Blue deliver the cross-sell and margin expansion management is banking on?",
    "Does the Q2 2026 guide imply conservatism, organic deceleration, or continued competitive pressure?",
    "Can adjusted FCF actually reduce net leverage instead of just maintaining it via aggressive buybacks?",
    "What happens to take-rate and GRLNF margin if large enterprises and agencies press for lower fees?",
    "Will buybacks more than offset SBC dilution without consuming strategic cash flexibility?",
    "What valuation multiple is justified if owner earnings normalize to high-single-digit growth?",
    "Which leading indicators should be monitored: organic GRLNF, spread, Global Blue contribution, net leverage, share count?",
    "Is the preferred stock likely to convert at current price levels, creating dilution?",
]

for i, q in enumerate(questions):
    r = i + 3
    ws5.cell(row=r, column=1, value=i + 1)
    ws5.cell(row=r, column=2, value=q)
    sc(ws5, r, 1)
    sc(ws5, r, 2)

# ========== SHEET 6: Sources ==========
ws6 = wb.create_sheet("Sources")
ws6.cell(row=1, column=1, value="Source")
ws6.cell(row=1, column=2, value="Type")
ws6.cell(row=1, column=3, value="Date / as-of")
ws6.cell(row=1, column=4, value="Usage")
ws6.cell(row=1, column=5, value="URL / note")
hdr(ws6, 1, 5)

sources = [
    ("Research/FOUR.md", "Primary research", "2026-05-24", "Historical actuals, capital structure, analyst consensus, scenarios.", "research/FOUR.md"),
    ("Shift4 FY2025 Form 10-K", "Primary filing", "Filed 2026-02", "FY2025 actuals, balance sheet, cash flow, leases.", "https://www.sec.gov/Archives/edgar/data/1794669/000179466926000010/four-20251231.htm"),
    ("Shift4 Q1 2026 earnings release", "Company release", "2026-05", "Q1 2026 actuals, guidance.", "https://investors.shift4.com/"),
    ("StockAnalysis FOUR", "Market data", "2026-05-24", "Price, shares, beta, analyst targets.", "https://stockanalysis.com/stocks/four/"),
    ("MarketBeat FOUR", "Market data", "2026-05-24", "Analyst targets, ownership, short interest.", "https://www.marketbeat.com/stocks/NYSE/FOUR/"),
    ("Bob Hammel Substack", "Secondary analysis", "2025", "Bull case; flagged leverage, Global Blue, transition.", "https://bobhammel.substack.com/p/shift4-payments-positioned-to-win"),
    ("Shift4 investor relations", "Company materials", "2026-05-24", "Earnings releases, presentations.", "https://investors.shift4.com/financial-information/financial-results"),
    ("FRED DGS10", "Rate source", "2026-05-26", "10-year Treasury: 4.50%.", "https://fred.stlouisfed.org/series/DGS10"),
    ("Seeking Alpha FOUR", "Analyst debate", "2026-05-24", "Sell-side articles (partially paywalled).", "https://seekingalpha.com/symbol/FOUR/analysis"),
]

for i, (name, typ, date_val, usage, url) in enumerate(sources):
    r = i + 2
    ws6.cell(row=r, column=1, value=name)
    ws6.cell(row=r, column=2, value=typ)
    ws6.cell(row=r, column=3, value=date_val)
    ws6.cell(row=r, column=4, value=usage)
    ws6.cell(row=r, column=5, value=url)
    for c in range(1, 6):
        sc(ws6, r, c)

# Column widths
for w in wb.worksheets:
    w.column_dimensions['A'].width = 35
    w.column_dimensions['B'].width = 18
    w.column_dimensions['C'].width = 15
    w.column_dimensions['D'].width = 25
    w.column_dimensions['E'].width = 18
    w.column_dimensions['F'].width = 18

output_path = "/home/refcell/dev/capital/models/[2026-05-28] Shift4 Payments Model.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print(f"Sheet names: {wb.sheetnames}")
