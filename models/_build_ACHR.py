#!/usr/bin/env python3
"""Build 6-sheet valuation model for Archer Aviation (ACHR).
Pre-commercial eVTOL developer — Cash NAV floor + Pipeline optionality framework.
Date: 2026-07-22
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──
title_font = Font(name='Calibri', size=16, bold=True)
subtitle_font = Font(name='Calibri', size=12, bold=True)
header_font = Font(name='Calibri', size=11, bold=True)
normal_font = Font(name='Calibri', size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
section_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

def s(ws, row, col, val, font=None, fill=None, border=None):
    cell = ws.cell(row=row, column=col, value=val)
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    return cell

def c(ws, row, col, val):
    return s(ws, row, col, val, normal_font, border=thin_border)

# ═══════════════════════════════════════════════
# Sheet 1: Valuation
# ═══════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Valuation'
ws1.merge_cells('A1:F1')
s(ws1, 1, 1, 'Archer Aviation Inc. (ACHR) — Valuation Model', title_font)

title_data = [
    ('Ticker', 'ACHR'),
    ('Exchange', 'NYSE'),
    ('Sector / Industry', 'Industrials / Aerospace & Defense'),
    ('Date', '2026-07-22'),
    ('Price', '$5.18'),
    ('Shares Outstanding', '762.86M'),
    ('Market Cap', '$3.953B'),
    ('Enterprise Value', '$2.37B (net cash position)'),
    ('Primary Valuation Lens', 'Cash NAV floor + Pipeline optionality (pre-commercial developer)'),
    ('Current Stance', 'Watch'),
    ('Analyst Coverage', 'Limited — 8-12 estimates visible'),
    ('Description', 'eVTOL aircraft for urban air taxi; pre-revenue, pre-certification'),
]

for i, (k, v) in enumerate(title_data, 2):
    s(ws1, i, 1, k, header_font)
    s(ws1, i, 2, v, normal_font)

metrics = [
    ('Trailing P/E', 'N/A (negative earnings — diluted EPS -$1.10 TTM)', 'Not meaningful; company is deeply unprofitable'),
    ('Forward P/E', 'N/A', 'No positive earnings forecast; analysts project continued losses through FY27'),
    ('P/S (Trailing)', '2,079x ($3.95B / $1.9M TTM)', 'Pre-revenue multiple; essentially meaningless as valuation metric'),
    ('P/FCF', 'N/A (FCF structurally negative)', 'Company burning $487M TTM OCF — FCF multiple not applicable'),
    ('EV/FCF', 'N/A', 'Same reason — structural negative cash flow by design'),
    ('EV/Sales', '1,247x ($2.37B / $1.9M TTM)', 'Enterprise value appears lower than MC due to net cash, but still meaningless vs negligible revenue'),
    ('EV/EBITDA', '-0.39x (Yahoo Finance quarterly stats)', 'Negative — EBITDA is -$813.9M; ratio is meaningless for pre-commercial developer'),
    ('Price/Book', '1.94x', 'Most applicable multiple for pre-commercial — 1.94x vs P/B of 3.38x at Q4 FY25'),
    ('Cash/Share NAV', '$2.34', 'Total cash $1.78B / 762.86M shares; this is the actual floor'),
    ('Optionality Premium', '$2.84/share ($2.17B total)', 'Market cap minus cash NAV = what market pays for FAA certification option'),
]

s(ws1, 15, 1, 'Key Valuation Metrics', subtitle_font)
for i, (k, v, note) in enumerate(metrics, 16):
    s(ws1, i, 1, k, header_font)
    s(ws1, i, 2, v, normal_font)
    s(ws1, i, 3, note, normal_font)

# ═══════════════════════════════════════════════
# Sheet 2: WACC
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet('WACC')
ws2.merge_cells('A1:D1')
s(ws2, 1, 1, 'WACC — CAPM (Archer Aviation ACHR)', title_font)

# Risk-free rate from CNBC 10Y Treasury
rf = 4.661
erp = 5.0
beta = 3.19
cost_of_equity = rf + beta * erp  # 4.661 + 15.95 = 19.661%
cost_of_debt = 4.0  # assumed — low debt level means this is immaterial
tax_rate = 21.0  # US corporate tax rate
mc_b = 3.953  # $B
total_debt = 0.122  # $121.8M
equity_weight = mc_b / (mc_b + total_debt)
debt_weight = total_debt / (mc_b + total_debt)
wacc = equity_weight * cost_of_equity + debt_weight * cost_of_debt * (1 - tax_rate / 100)

wacc_data = [
    ('Risk-Free Rate (10Y US Treasury)', f'{rf:.3f}%'),
    ('Equity Risk Premium', f'{erp:.1f}%'),
    ('Beta (5Y Monthly)', f'{beta:.2f}'),
    ('Cost of Equity (Ke = Rf + Beta * ERP)', f'{cost_of_equity:.2f}%'),
    ('Cost of Debt (assumed)', f'{cost_of_debt:.1f}%'),
    ('Tax Rate (US corporate)', f'{tax_rate:.0f}%'),
    ('Market Cap ($B)', f'{mc_b:.3f}'),
    ('Total Debt ($B)', f'{total_debt:.3f}'),
    ('Equity Weight', f'{equity_weight:.4f}'),
    ('Debt Weight', f'{debt_weight:.4f}'),
    ('WACC', f'{wacc:.2f}%'),
    ('Note', 'High beta reflects extreme volatility of pre-commercial developer. Debt is minimal so WACC ≈ Ke.'),
]

for i, (k, v) in enumerate(wacc_data, 3):
    s(ws2, i, 1, k, header_font)
    s(ws2, i, 2, v, normal_font)

print(f"WACC: {wacc:.2f}%")

# ═══════════════════════════════════════════════
# Sheet 3: Scenarios — Cash NAV + Pipeline Optionality
# ═══════════════════════════════════════════════
ws3 = wb.create_sheet('Scenarios')
ws3.merge_cells('A1:H1')
s(ws3, 1, 1, 'Scenarios — Cash NAV Floor + Pipeline Optionality Framework', title_font)

s(ws3, 3, 1, 'Framework: Pre-commercial eVTOL developer. Standard DCF/FCF multiples are inapplicable.', normal_font)
s(ws3, 4, 1, 'Scenarios model: (a) annual burn rate, (b) dilution factor, (c) NAV floor per share, (d) probability-weighted pipeline NPV.', normal_font)
s(ws3, 5, 1, 'Key input: Cash $1.78B / 762.86M shares = $2.34 cash/share NAV floor. Total returnability from FAA certification.', normal_font)

scenarios = [
    ('Metric', 'Bear', 'Base', 'Bull', 'Units'),
    ('Annual Burn Rate', '$600M', '$500M', '$450M', '$M/yr'),
    ('Cash Runway (years)', '3.0', '3.6', '4.0', 'years'),
    ('Dilution Factor', '2.00x', '1.40x', '1.25x', 'multiplier'),
    ('Post-Dilution Shares (M)', '1,525.7', '1,068.0', '954.6', 'M shares'),
    ('Revenue FY2027 (analyst avg)', '$114M', '$114M', '$114M', '$M (Yahoo analysis)'),
    ('Revenue FY2028 (analyst avg)', '$221M', '$221M', '$221M', '$M (yahoo analysis)'),
    ('Terminal Revenue (Year 5)', '$80M', '$300M', '$800M', '$M'),
    ('Terminal Operating Margin', '-100%', '-30%', '+5%', '%'),
    ('NAV Floor Per Share (post-dilution)', '$1.16', '$1.66', '$1.86', '$/share'),
    ('Pipeline NPV Per Share', '$0', '$7.00', '$25.00', '$/share'),
    ('Target Price', '$1.16', '$8.66', '$26.86', '$/share'),
    ('Upside from $5.18', '-77.6%', '+67.4%', '+417.8%', '%'),
    ('Weight', '25%', '50%', '25%', '%'),
    ('Weighted Value/Share', '$0.29', '$4.33', '$6.72', '$/share'),
    ('Probability-Weighted FV', '$11.34', '', '', '$/share'),
    ('Upside from Current $5.18', '+118.7%', '', '', '%'),
]

for i, row_data in enumerate(scenarios, 7):
    font = header_font if i == 7 else normal_font
    fill = header_fill if i == 7 else None
    for j, val in enumerate(row_data, 1):
        s(ws3, i, j, val, font, fill=fill, border=thin_border)

fv = 11.34
print(f"Probability-Weighted FV: ${fv:.2f}")
print(f"Upside from $5.18: {(fv/5.18 - 1)*100:.1f}%")

# ═══════════════════════════════════════════════
# Sheet 4: Actuals Source Audit
# ═══════════════════════════════════════════════
ws4 = wb.create_sheet('Actuals Source Audit')
ws4.merge_cells('A1:E1')
s(ws4, 1, 1, 'Actuals Source Audit', title_font)

audit = [
    ('Stock Price', '$5.18 close', 'Yahoo Finance /quote/ACHR/', '2026-07-22', 'NYSE real-time'),
    ('Market Cap', '$3.953B', 'Yahoo Finance /quote/ACHR/', '2026-07-22', 'Intraday; stats page shows $4.03B Q1 FY26'),
    ('Enterprise Value', '$2.37B', 'Yahoo Finance Statistics', '2026-07-22', 'Net cash position due to $1.78B cash'),
    ('Shares Outstanding', '762.86M', 'Yahoo Finance Statistics', '2026-07-22', 'Per MRQ'),
    ('Revenue TTM', '$1.9M', 'Yahoo Finance Financials', 'TTM as of Q1 FY26', 'In thousands on page'),
    ('Revenue FY2025', '$0.3M', 'Yahoo Finance Financials', 'Annual', 'Minimal — pre-revenue'),
    ('Net Income TTM', '-$742.5M', 'Yahoo Finance Financials', 'TTM', 'Diluted EPS -$1.10'),
    ('Total Assets', '$2,465.9M', 'Yahoo Finance Balance Sheet', '12/31/2025', 'In thousands'),
    ('Total Equity', '$2,202.8M', 'Yahoo Finance Balance Sheet', '12/31/2025', 'In thousands'),
    ('Total Cash (MRQ)', '$1.78B', 'Yahoo Finance Statistics', 'Q1 FY26', 'Cash per share $2.34'),
    ('Total Debt (MRQ)', '$121.8M', 'Yahoo Finance Statistics', 'Q1 FY26', 'Debt/Equity 5.86%'),
    ('Operating Cash Flow TTM', '-$487.4M', 'Yahoo Finance Cash Flow', 'TTM', 'In thousands'),
    ('Beta (5Y Monthly)', '3.19', 'Yahoo Finance Statistics', '2026-07-22', 'High volatility; pre-commercial risk'),
    ('P/B Ratio', '1.94x', 'Yahoo Finance Statistics', 'Current/Quarterly', 'Most relevant valuation multiple'),
    ('10Y US Treasury', '4.661%', 'CNBC /quotes/US10Y', '2026-07-22', 'For WACC calculation'),
    ('Analyst Coverage', 'Limited', 'Yahoo Finance Analysis', '2026-07-22', 'No forward P/E visible; estimates gated'),
    ('Earnings Date', 'Aug 10, 2026', 'Yahoo Finance Profile', '2026-07-22', 'Next catalyst'),
]

audit_headers = ['Data Point', 'Value', 'Source URL', 'Date', 'Notes']
for j, h in enumerate(audit_headers, 1):
    s(ws4, 2, j, h, header_font, fill=header_fill, border=thin_border)

for i, (k, v, src, dt, note) in enumerate(audit, 3):
    s(ws4, i, 1, k, normal_font, border=thin_border)
    s(ws4, i, 2, v, normal_font, border=thin_border)
    s(ws4, i, 3, src, normal_font, border=thin_border)
    s(ws4, i, 4, dt, normal_font, border=thin_border)
    s(ws4, i, 5, note, normal_font, border=thin_border)

# ═══════════════════════════════════════════════
# Sheet 5: Questions
# ═══════════════════════════════════════════════
ws5 = wb.create_sheet('Questions')
ws5.merge_cells('A1:C1')
s(ws5, 1, 1, 'Open Questions', title_font)

questions = [
    ('1', 'FAA Certification Timeline', 'What is the current date for FAA Type Certificate application and expected certification? This is THE binary event for the company. Management has guided to ~2027-2028 for certification — has this slipped?'),
    ('2', 'Cash Burn Trajectory', 'TTM OCF is -$487.4M. At this rate, the $1.78B cash pile provides ~3.6 years of runway. Is the burn rate accelerating with the Long Beach manufacturing facility coming online?'),
    ('3', 'Dilution History and Future Raises', 'Shares grew from ~240M in FY2022 to 762.86M now — a 217% expansion. Will another dilutive raise be needed before certification? At what price?'),
    ('4', 'Revenue Composition', '$1.9M TTM revenue vs $729M operating expense in FY2025. What drives the tiny revenue — government contracts, consulting, or aircraft sales? Is any of it recurring?'),
    ('5', 'Partnership Commitments', 'Uber, Toyota, and other partners have MOUs and pilot agreements. What is the binding revenue commitment from each? MOUs ≠ purchase orders.'),
    ('6', 'Manufacturing Capacity', 'What is the current and planned production capacity at the Long Beach facility? Is the facility generating material CapEx that accelerates the burn?'),
    ('7', 'Investing Cash Flow Spike', 'Investing CF of -$1.087B TTM and -$1.176B in FY2025 — what is this spending on? Manufacturing facility, R&D, or something else?'),
    ('8', 'Debt Composition', '$121.8M total debt at Q1 FY26. What is this — convertible notes, capital leases, or term debt? Convertible debt would create additional dilution risk.'),
    ('9', 'Competitive Landscape', 'How does Archer position against Joby Aviation (the other major public eVTOL), Eve Air Mobility, and Vertical Aerospace? Is there first-mover advantage?'),
    ('10', 'Insurance and Regulatory Risk', 'Beyond FAA certification, what are the insurance, liability, and operational certification challenges? Urban air taxi introduces entirely new regulatory categories.'),
    ('11', 'Revenue Recognition', 'The revenue jumped from $0 in FY2023/FY2024 to $300K in FY2025 to $1.9M TTM. Is this from actual aircraft operations or related-party contracts?'),
    ('12', 'SBC and Option Dilution', 'With 1,160 employees and tech-sector compensation norms, how much annual SBC is there? Option dilution can be silent shareholder destruction.'),
]

for j, h in enumerate(['#', 'Topic', 'Question'], 1):
    s(ws5, 2, j, h, header_font, fill=header_fill, border=thin_border)

for i, (num, topic, q) in enumerate(questions, 3):
    s(ws5, i, 1, num, normal_font, border=thin_border)
    s(ws5, i, 2, topic, header_font, border=thin_border)
    s(ws5, i, 3, q, normal_font, border=thin_border)

# ═══════════════════════════════════════════════
# Sheet 6: Sources
# ═══════════════════════════════════════════════
ws6 = wb.create_sheet('Sources')
ws6.merge_cells('A1:C1')
s(ws6, 1, 1, 'Sources', title_font)

sources = [
    ('1', 'Yahoo Finance — ACHR Quote Page', 'https://finance.yahoo.com/quote/ACHR/'),
    ('2', 'Yahoo Finance — ACHR Income Statement', 'https://finance.yahoo.com/quote/ACHR/financials/'),
    ('3', 'Yahoo Finance — ACHR Balance Sheet', 'https://finance.yahoo.com/quote/ACHR/balance-sheet/'),
    ('4', 'Yahoo Finance — ACHR Cash Flow', 'https://finance.yahoo.com/quote/ACHR/cash-flow/'),
    ('5', 'Yahoo Finance — ACHR Key Statistics', 'https://finance.yahoo.com/quote/ACHR/key-statistics/'),
    ('6', 'Yahoo Finance — ACHR Analyst Estimates', 'https://finance.yahoo.com/quote/ACHR/analysis/'),
    ('7', 'Yahoo Finance — ACHR Profile', 'https://finance.yahoo.com/quote/ACHR/profile/'),
    ('8', 'CNBC — US10Y Treasury Yield', 'https://www.cnbc.com/quotes/US10Y'),
    ('9', 'StockAnalysis — 404 (not available)', 'https://stockanalysis.com/quote/ACHR/'),
    ('10', 'Company Website', 'https://www.archer.com'),
]

for j, h in enumerate(['#', 'Description', 'URL'], 1):
    s(ws6, 2, j, h, header_font, fill=header_fill, border=thin_border)

for i, (num, desc, url) in enumerate(sources, 3):
    s(ws6, i, 1, num, normal_font, border=thin_border)
    s(ws6, i, 2, desc, normal_font, border=thin_border)
    s(ws6, i, 3, url, normal_font, border=thin_border)

# ── Column widths ──
for ws in [ws1, ws2, ws3, ws4, ws5, ws6]:
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30

# ── Save ──
fpath = '/home/refcell/dev/capital/models/2026-07-22 Archer Aviation Model.xlsx'
wb.save(fpath)
print(f'Saved to {fpath}')

# ── Verify ──
wb2 = openpyxl.load_workbook(fpath)
print(f'Sheets: {wb2.sheetnames}')
for sn in wb2.sheetnames:
    ws = wb2[sn]
    print(f'  {sn}: {ws.max_row} rows x {ws.max_column} cols')
