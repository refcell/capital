#!/usr/bin/env python3
"""Build 6-sheet valuation model for Agilent Technologies (A). Agilent Technologies (A) is a diagnostics & research / life sciences company providing LC/MS/MS systems, genomics, pathology, diagnostics, CrossLab services (consumables, software, maintenance), and Applied Markets (gas chromatography, mass spectrometry, vacuum, spectroscopy)."""
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Style helpers ──
bold = Font(bold=True)
header_font = Font(bold=True, underline='single')
title_font = Font(bold=True, size=14)
subtitle_font = Font(bold=True, size=11, italic=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

def c(ws, row, col, value, font=None, border=None, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if border:
        cell.border = border
    if fill:
        cell.fill = fill
    return cell

# ── Sheet 1: Valuation ──
ws1 = wb.active
ws1.title = "Valuation"

# Merge title
ws1.merge_cells("A1:D1")
c(ws1, 1, 1, "Agilent Technologies (A) - Valuation Summary", title_font)
ws1.merge_cells("A2:D2")
c(ws1, 2, 1, "As of August 28, 2026 close", subtitle_font)

title_data = [
    ("Ticker", "NYSE: A", ""),
    ("Company", "Agilent Technologies, Inc.", "Diagnostics & Research / Life Sciences"),
    ("Date", "2026-08-28", ""),
    ("Stock Price", "$153.84", "Yahoo Finance close"),
    ("Shares Outstanding", "282.0M", "Yahoo Key Stats"),
    ("Market Cap", "$44.46B", ""),
    ("Enterprise Value", "$46.65B", "Yahoo Key Stats"),
    ("Net Debt", "$2.19B", "EV - MC; Total debt $3.95B, cash $1.76B"),
    ("Primary Valuation Lens", "Forward P/E", "Analyst consensus supported"),
    ("Stance", "Watch", "Quality name at reasonable P/E; await Q3 catalyst"),
]

for i, (field, value, note) in enumerate(title_data, 3):
    c(ws1, i, 1, field, bold, thin_border)
    c(ws1, i, 2, value, None, thin_border)
    c(ws1, i, 3, note, None, thin_border)

# Valuation metrics table
c(ws1, 14, 1, "Valuation Metrics", bold)
ws1.merge_cells("A14:B14")

val_headers = ["Metric", "Value", "Comment"]
for j, h in enumerate(val_headers, 1):
    c(ws1, 15, j, h, header_font, thin_border, header_fill)

val_data = [
    ("P/E (Trailing)", "31.10x", "TTM basis; elevated but normal for quality"),
    ("P/E (Forward)", "23.64x", "FY2026 consensus; reasonable for medtech/life sciences"),
    ("P/S", "6.06x", "Mid-range; premium for quality and recurring revenue"),
    ("P/FCF", "35.1x", "On $1.27B TTM FCF; elevated relative to growth rate"),
    ("EV/FCF", "36.8x", "EV $46.65B / FCF $1.267B"),
    ("EV/Sales", "6.33x", "Enterprise value on revenue; premium name traded fairly"),
    ("EV/EBITDA", "22.93x", "S&P calculated EBITDA ~$2.035B; moderate premium"),
    ("PEG (5Y Expected)", "1.41x", "Below 1.5x threshold; growth-justified multiple"),
    ("P/B", "6.04x", "On BVPS $23.82; elevated but ROE 20.97% justifies"),
    ("Dividend Yield", "0.65%", "$1.02 forward annual; low yield typical for growth"),
    ("Beta (5Y)", "1.23", "Slightly above market; healthcare sensitivity"),
    ("ROE (TTM)", "20.97%", "High returns on $6.74B equity; well-managed capital base"),
    ("Operating Margin (TTM)", "22.25%", "High gross margin (53.8%) supports this well"),
    ("Net Margin (TTM)", "19.53%", "On $7.37B revenue; quality earnings conversion"),
]

for i, (metric, val, comment) in enumerate(val_data, 16):
    c(ws1, i, 1, metric, None, thin_border)
    c(ws1, i, 2, val, None, thin_border)
    c(ws1, i, 3, comment, None, thin_border)

# Set column widths
for col_letter, width in [("A", 25), ("B", 22), ("C", 50)]:
    ws1.column_dimensions[col_letter].width = width

# ── Sheet 2: WACC ──
ws2 = wb.create_sheet("WACC")
ws2.merge_cells("A1:B1")
c(ws2, 1, 1, "Weighted Average Cost of Capital (WACC)", title_font)
ws2.merge_cells("A2:B2")
c(ws2, 2, 1, "CAPM Approach - Agilent Technologies (A)", subtitle_font)

# WACC inputs
risk_free = 4.73  # 10Y US Treasury
erp = 5.0
beta = 1.23
cost_of_equity = risk_free + beta * erp  # 4.73 + 1.23*5 = 10.88
pre_tax_cost_of_debt = 3.25  # estimate for investment grade corp
tax_rate = 13.55  # based on TTM: $218M tax / $1,658M pre-tax
total_debt_B = 3.95  # Key Stats figure
market_cap_B = 44.46
total_cap_B = market_cap_B + total_debt_B
equity_weight = market_cap_B / total_cap_B
debt_weight = total_debt_B / total_cap_B
after_tax_cost_of_debt = pre_tax_cost_of_debt * (1 - tax_rate / 100)
wacc = equity_weight * cost_of_equity + debt_weight * after_tax_cost_of_debt

wacc_data = [
    ("Risk-Free Rate (10Y US Treasury)", f"{risk_free:.2f}%", "CNBC quote Aug 28, 2026"),
    ("Equity Risk Premium", f"{erp:.1f}%", "Standard ERP assumption"),
    ("Beta (5Y Monthly)", f"{beta:.2f}", "Yahoo Finance Key Stats"),
    ("Cost of Equity (Ke)", f"{cost_of_equity:.2f}%", f"Rf + Beta * ERP = {risk_free:.2f} + {beta:.2f} * {erp:.1f}"),
    ("Pre-Tax Cost of Debt (Kd)", f"{pre_tax_cost_of_debt:.2f}%", "Estimate for investment grade corp"),
    ("Tax Rate", f"{tax_rate:.2f}%", f"TTM: $218M tax / $1,658M pre-tax income"),
    ("", "", ""),
    ("Market Cap", f"${market_cap_B:.2f}B", "Yahoo Finance"),
    ("Total Debt", f"${total_debt_B:.2f}B", "Yahoo Key Stats"),
    ("Total Capitalization", f"${total_cap_B:.2f}B", ""),
    ("Equity Weight", f"{equity_weight:.2%}", ""),
    ("Debt Weight", f"{debt_weight:.2%}", ""),
    ("", "", ""),
    ("After-Tax Cost of Debt", f"{after_tax_cost_of_debt:.2f}%", f"Kd * (1 - t) = {pre_tax_cost_of_debt:.2f} * (1 - {tax_rate:.2f}%/100)"),
    ("", "", ""),
    ("WACC", f"{wacc:.2f}%", f"Equity weight * Ke + Debt weight * after-tax Kd"),
    ("", f"{equity_weight:.4f} * {cost_of_equity:.2f} + {debt_weight:.4f} * {after_tax_cost_of_debt:.2f} = {wacc:.2f}%", ""),
]

for i, (field, value, note) in enumerate(wacc_data, 3):
    c(ws2, i, 1, field, bold if field == "WACC" else None, thin_border)
    c(ws2, i, 2, value, None, thin_border)
    c(ws2, i, 3, note, None, thin_border)

for col_letter, width in [("A", 35), ("B", 22), ("C", 55)]:
    ws2.column_dimensions[col_letter].width = width

# ── Sheet 3: Scenarios ──
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells("A1:I1")
c(ws3, 1, 1, "Scenario Analysis - Forward P/E Framework", title_font)
ws3.merge_cells("A2:I2")
c(ws3, 2, 1, "Agilent Technologies (A) - Bear / Base / Bull Cases", subtitle_font)

# Scenario logic:
# Current revenue FY25: $6.948B. TTM ~$7.372B (includes Q1 FY26).
# FY26 consensus EPS ~$6.51 (from forward P/E 23.64x: $153.84 / 23.64)
# Bear: lower growth, lower multiple. Forward P/E bear 17x, base 21x, bull 25x.
# 5Y terminal: Bear EPS $7.2 (0.4% CAGR from ~$6.51 implied fwd), Base $8.2 (2.4% CAGR), Bull $10.0 (5.5% CAGR)

bear_terminal_eps = 7.2
bear_exit_pe = 17
bear_target = bear_terminal_eps * bear_exit_pe  # $122.40

base_terminal_eps = 8.2
base_exit_pe = 21
base_target = base_terminal_eps * base_exit_pe  # $172.20

bull_terminal_eps = 10.0
bull_exit_pe = 25
bull_target = bull_terminal_eps * bull_exit_pe  # $250.00

current_price = 153.84

bear_weight = 0.2
base_weight = 0.5
bull_weight = 0.3

weighted_fv = bear_weight * bear_target + base_weight * base_target + bull_weight * bull_target

# Net debt for scenario sheet - using EV - MC
net_debt_B = 2.19  # From Key Stats: debt $3.95B - cash $1.76B
shares_mm = 282.0
net_debt_per_share = net_debt_B * 1000 / shares_mm  # $7.77/share

bear_upside = (bear_target - current_price) / current_price * 100
base_upside = (base_target - current_price) / current_price * 100
bull_upside = (bull_target - current_price) / current_price * 100

sc_headers = [
    "Metric", "Bear (20%)", "Base (50%)", "Bull (30%)",
    "Notes",
]

sc_data = [
    ("Revenue CAGR (5Y)", "2%", "4%", "6.5%", ""),
    ("Terminal EPS (5Y)", f"${bear_terminal_eps:.2f}", f"${base_terminal_eps:.2f}", f"${bull_terminal_eps:.2f}", ""),
    ("Exit P/E", f"{bear_exit_pe}x", f"{base_exit_pe}x", f"{bull_exit_pe}x", ""),
    ("Target Price", f"${bear_target:.2f}", f"${base_target:.2f}", f"${bull_target:.2f}", ""),
    ("Upside from Current", f"{bear_upside:.1f}%", f"{base_upside:.1f}%", f"{bull_upside:.1f}%", ""),
    ("Weight", f"{bear_weight:.0%}", f"{base_weight:.0%}", f"{bull_weight:.0%}", ""),
    ("Weighted Value", f"${bear_weight * bear_target:.2f}", f"${base_weight * base_target:.2f}", f"${bull_weight * bull_target:.2f}", ""),
    ("", "", "", "", ""),
    ("Probability-Weighted FV", "", "", f"${weighted_fv:.2f}", "Sum of weighted values"),
    ("Current Price", "", "", f"${current_price:.2f}", ""),
    ("Probability-Weighted Upside", "", "", f"{(weighted_fv - current_price) / current_price * 100:.1f}%", ""),
]

for j, h in enumerate(sc_headers, 1):
    c(ws3, 3, j, h, header_font, thin_border, header_fill)

for i, row_data in enumerate(sc_data, 4):
    for j, val in enumerate(row_data, 1):
        f = bold if "Weighted FV" in str(row_data[0]) or "Upside" in str(row_data[0]) else None
        c(ws3, i, j, val, f, thin_border)

# Net cash note
c(ws3, 14, 1, "Framework Note", bold)
c(ws3, 15, 1, "Forward P/E on analyst consensus is the primary framework. Agilent generates $1.27B TTM FCF")
c(ws3, 16, 1, "on $7.37B revenue with 22% operating margins. The key variable is earnings growth trajectory.")
c(ws3, 17, 1, "Forward P/E 23.64x (current) is reasonable for a quality medtech/life sciences name.")
c(ws3, 18, 1, f"Net debt: ${net_debt_B:.2f}B (Key Stats debt $3.95B - cash $1.76B).")
c(ws3, 19, 1, f"WACC (from Sheet 2): {wacc:.2f}%")
c(ws3, 20, 1, f"Probability-weighted fair value: ${weighted_fv:.2f} implies {(weighted_fv - current_price) / current_price * 100:.1f}% upside from ${current_price:.2f}")

for col_letter, width in [("A", 35), ("B", 18), ("C", 18), ("D", 18), ("E", 50)]:
    ws3.column_dimensions[col_letter].width = width

# ── Sheet 4: Actuals Source Audit ──
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells("A1:D1")
c(ws4, 1, 1, "Actuals Source Audit - Agilent Technologies (A)", title_font)
ws4.merge_cells("A2:D2")
c(ws4, 2, 1, "Every data point with source URL, date, and notes", subtitle_font)

c(ws4, 3, 1, "Data Point", header_font, thin_border, header_fill)
c(ws4, 3, 2, "Value", header_font, thin_border, header_fill)
c(ws4, 3, 3, "Source", header_font, thin_border, header_fill)
c(ws4, 3, 4, "Notes", header_font, thin_border, header_fill)

audit_data = [
    # Stock data
    ("Stock Price", "$153.84", "Yahoo Finance, Aug 28 2026", "Close price NYSE"),
    ("Market Cap", "$44.46B", "Yahoo Finance Key Stats", "Quarterly stats, Aug 28 2026"),
    ("Enterprise Value", "$46.65B", "Yahoo Finance Key Stats", "Includes total debt $3.95B"),
    ("Shares Outstanding", "281.97M", "Yahoo Finance Key Stats", "Basic shares outstanding"),
    ("", "", "", ""),
    # Income statement data
    ("Revenue (TTM)", "$7.37B", "Yahoo Finance Income Statement", "FY25: $6.948B, TTM: $7.372B"),
    ("Cost of Revenue (TTM)", "$3.40B", "Yahoo Finance Income Statement", ""),
    ("Gross Profit (TTM)", "$3.97B", "Yahoo Finance Income Statement", "Gross margin 53.8%"),
    ("Gross Margin (TTM)", "53.8%", "Yahoo Finance Income Statement", "Improvement from 52.4% FY25"),
    ("Operating Expense (TTM)", "$2.33B", "Yahoo Finance Income Statement", ""),
    ("Operating Income (TTM)", "$1.64B", "Yahoo Finance Income Statement", "Op margin 22.25%"),
    ("EBITDA (TTM)", "$2.04B", "Yahoo Finance Income Statement", "Normalized EBITDA"),
    ("Pretax Income (TTM)", "$1.66B", "Yahoo Finance Income Statement", ""),
    ("Tax Provision (TTM)", "$218M", "Yahoo Finance Income Statement", "Effective rate ~13.1%"),
    ("Net Income (TTM)", "$1.44B", "Yahoo Finance Income Statement", "Diluted EPS $5.07"),
    ("EPS Diluted (TTM)", "$5.07", "Yahoo Finance Income Statement", "Basic EPS $5.09"),
    ("", "", "", ""),
    # Balance sheet data
    ("Total Assets", "$12.73B", "Yahoo Finance Balance Sheet", "FY25: $11.846B"),
    ("Total Liabilities", "$5.99B", "Yahoo Finance Balance Sheet", ""),
    ("Total Equity", "$6.74B", "Yahoo Finance Balance Sheet", "Common stock equity"),
    ("Total Debt", "$3.95B", "Yahoo Finance Key Stats", "BS shows $3.35B; Key Stats higher"),
    ("Total Cash", "$1.76B", "Yahoo Finance Key Stats", "Per share $6.24"),
    ("Total Debt (BS)", "$3.35B", "Yahoo Finance Balance Sheet", "FY25: $3.39B"),
    ("Goodwill", "$4.92B", "Yahoo Finance Balance Sheet", "Large goodwill = intangible base"),
    ("Net Tangible Assets", "$1.82B", "Yahoo Finance Balance Sheet", ""),
    ("Book Value Per Share", "$23.82", "Yahoo Finance Key Stats", ""),
    ("", "", "", ""),
    # Cash flow data
    ("Operating Cash Flow (TTM)", "$1.61B", "Yahoo Finance Cash Flow", "FY25: $1.559B"),
    ("Capex (TTM)", "($342M)", "Yahoo Finance Cash Flow", "FY25: ($407M)"),
    ("Free Cash Flow (TTM)", "$1.27B", "Yahoo Finance Cash Flow", "FCF margin 17.2%"),
    ("Repayment of Debt (TTM)", "($9M)", "Yahoo Finance Cash Flow", "Minimal repayment"),
    ("Repurchase of Stock (TTM)", "($380M)", "Yahoo Finance Cash Flow", "Moderate buyback program"),
    ("", "", "", ""),
    # Analyst estimates
    ("Forward P/E", "23.64x", "Yahoo Finance Key Stats", "Nasdaq Real-Time stats"),
    ("Beta (5Y)", "1.23", "Yahoo Finance Key Stats", "5Y monthly beta"),
    ("Dividend Yield", "0.65%", "Yahoo Finance Key Stats", "Forward annual $1.02"),
    ("Analyst PT Avg", "$165.28", "Yahoo Finance Analysis", "12 analysts"),
    ("Analyst PT High", "$188.00", "Yahoo Finance Analysis", ""),
    ("Analyst PT Low", "$143.00", "Yahoo Finance Analysis", ""),
    ("Recommendation", "Buy (1.95)", "Yahoo Finance Analysis", "12 analysts: 6 strong buy, 4 buy, 2 hold"),
    ("", "", "", ""),
    # Earnings dates
    ("Next Earnings Date", "Nov 25, 2026", "Yahoo Finance Profile", "3:00 PM EST"),
    ("Last Earnings Date", "May 2026 Q1 FY26", "Yahoo Finance Profile", "Revenue beat reported"),
    ("", "", "", ""),
    # 10Y Treasury
    ("10Y US Treasury Yield", "4.73%", "CNBC, Aug 28 2026", "Used for WACC calculation"),
]

for i, (point, value, source, note) in enumerate(audit_data, 4):
    c(ws4, i, 1, point, None, thin_border)
    c(ws4, i, 2, value, None, thin_border)
    c(ws4, i, 3, source, None, thin_border)
    c(ws4, i, 4, note, None, thin_border)

for col_letter, width in [("A", 30), ("B", 20), ("C", 35), ("D", 50)]:
    ws4.column_dimensions[col_letter].width = width

# ── Sheet 5: Questions ──
ws5 = wb.create_sheet("Questions")
ws5.merge_cells("A1:B1")
c(ws5, 1, 1, "Open Questions - Agilent Technologies (A)", title_font)

questions = [
    ("Q1", "Key Stats vs Balance Sheet total debt: Key Stats shows $3.95B but BS shows $3.35B — a $600M gap. Does Key Stats include capital leases, convertible notes, or other obligations outside statutory permanent debt?"),
    ("Q2", "Debt trajectory: Total debt increased from $2.77B (FY22) to $3.35B (FY25) to $3.95B Key Stats. Is debt growing for acquisitions, buybacks, or organic growth funding? What is the maturity profile?"),
    ("Q3", "Goodwill magnitude: $4.92B in goodwill on $12.73B total assets (38.7%). What acquisitions generated this? Any historical impairment charges?"),
    ("Q4", "Share count trajectory: Shares declining from 285.2M (FY25) to 281.97M (current). At what annual pace? Is the buyback program sized at $380M TTM sustainable long-term?"),
    ("Q5", "Customer concentration: Any single customer or end-market (>10% revenue)? Life sciences vs. Diagnostics vs. Applied Markets concentration risk?"),
    ("Q6", "Competitive differentiation vs. Waters (WAT), Thermo Fisher (TMO), Revvity (RVTY), Danaher (DHR): Where does Agilent win or lose on instrumentation, consumables, and services?"),
    ("Q7", "SBC and dilution: What is the annual stock-based compensation expense? Does it offset buyback dilution? Is the net diluted share count trend directionally declining?"),
    ("Q8", "Gross margin trajectory: Gross margin expanded from 52.4% (FY25) to 53.8% (TTM). Is this structural (mix shift to CrossLab/services) or cyclical? How durable is the higher margin?"),
    ("Q9", "Capital allocation priority: TTM buybacks of $380M but only $9M in debt repayment. Is management prioritizing share count reduction over deleveraging?"),
    ("Q10", "Next earnings catalyst: November 25, 2026 Q3 FY26 earnings. Any guidance given on segment mix, capex, or forward revenue? Did Q1 beat patterns continue?"),
]

for i, (num, question) in enumerate(questions, 2):
    c(ws5, i, 1, num, bold, thin_border)
    ws5.merge_cells(f"B{i}:D{i}")
    c(ws5, i, 2, question, None, thin_border)

for col_letter, width in [("A", 6), ("B", 100)]:
    ws5.column_dimensions[col_letter].width = width

# ── Sheet 6: Sources ──
ws6 = wb.create_sheet("Sources")
ws6.merge_cells("A1:B1")
c(ws6, 1, 1, "Sources - Agilent Technologies (A)", title_font)

sources = [
    ("1", "Yahoo Finance - Profile", "https://finance.yahoo.com/quote/A/profile/"),
    ("2", "Yahoo Finance - Income Statement", "https://finance.yahoo.com/quote/A/financials/"),
    ("3", "Yahoo Finance - Balance Sheet", "https://finance.yahoo.com/quote/A/balance-sheet/"),
    ("4", "Yahoo Finance - Cash Flow", "https://finance.yahoo.com/quote/A/cash-flow/"),
    ("5", "Yahoo Finance - Key Statistics", "https://finance.yahoo.com/quote/A/key-statistics/"),
    ("6", "Yahoo Finance - Analysis / Estimates", "https://finance.yahoo.com/quote/A/analysis/"),
    ("7", "Yahoo Finance - News", "https://finance.yahoo.com/quote/A/news/"),
    ("8", "CNBC - US 10Y Treasury", "https://www.cnbc.com/quotes/US10Y"),
    ("9", "Agilent Investor Relations", "https://www.agilent.com/"),
    ("10", "SEC EDGAR Filings (for reference)", "https://www.sec.gov/cgi-bin/browse-edgar?CIK=0001090872"),
]

for i, (num, name, url) in enumerate(sources, 2):
    c(ws6, i, 1, num, None, thin_border)
    c(ws6, i, 2, name, None, thin_border)
    c(ws6, i, 3, url, None, thin_border)

for col_letter, width in [("A", 4), ("B", 40), ("C", 60)]:
    ws6.column_dimensions[col_letter].width = width

# ── Save ──
outpath = "/home/refcell/dev/capital/models/2026-08-28 Agilent Model.xlsx"
wb.save(outpath)
print(f"Saved workbook to {outpath}")

# ── Verification ──
from openpyxl import load_workbook
wb_verify = load_workbook(outpath)
print(f"Sheets: {wb_verify.sheetnames}")
print(f"WACC = {wacc:.2f}%")
print(f"Bear target = ${bear_target:.2f} (upside: {bear_upside:.1f}%)")
print(f"Base target = ${base_target:.2f} (upside: {base_upside:.1f}%)")
print(f"Bull target = ${bull_target:.2f} (upside: {bull_upside:.1f}%)")
print(f"Weighted FV = ${weighted_fv:.2f}")
print(f"Prob. weighted upside = {(weighted_fv - current_price) / current_price * 100:.1f}%")
wb_verify.close()