#!/usr/bin/env python3
"""
Build THG (The Hanover Insurance Group) 6-sheet valuation model.
Bank/Insurance-specific valuation: P/B and ROE lens, not DCF.
FCF marked N/A per skill rules for financial institutions.
Date: 2026-06-29
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
import os, datetime

wb = Workbook()

# ── Styles ─────────────────────────────────────────────
bold = Font(bold=True)
title_font = Font(bold=True, size=14)
heading_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(), right=Side(), top=Side(), bottom=Side()
)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, color="FFFFFF")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def style_data(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

# ═══════════════════════════════════════════════════════
# DATA (all from Yahoo Finance as of 2026-06-29 close)
# ═══════════════════════════════════════════════════════
PRICE = 214.54
SHARES_MM = 35.0  # implied: MC $7.44B / $214.54 ~ 34.68, use 35.0 ~ diluted avg
MC_B = 7.44  # market cap intraday $7.44B
EV_B = 8.04  # enterprise value
TOTAL_DEBT_MM = 843.8  # Q1 2026 mrq total debt $843.8M
TOTAL_CASH_MM = 243.5  # end cash Q1 2026
NET_DEBT_MM = TOTAL_DEBT_MM - TOTAL_CASH_MM  # ~600.3M
EQUITY_MM = 3571.5  # FY2025 common stock equity $3.57B (from balance sheet)
BVPS_CURRENT = 100.86  # from Yahoo Finance statistics mrq
BETA = 0.32
RISK_FREE = 4.374  # 10Y US Treasury from CNBC
ERP = 5.0
TAX_RATE = 21.0  # ~22% from FY2025: tax prov $183.1M / pretax $843.8M = 21.7%
TAX_RATE_USED = 21.7

# Revenue (in $thousands from Yahoo, convert to $M)
REV_FY2022 = 5432.9
REV_FY2023 = 5963.5
REV_FY2024 = 6216.8
REV_FY2025 = 6567.3
REV_TTM = 6662.0

# Net Income (in $M)
NI_FY2022 = 116.0
NI_FY2023 = 35.3   # one-time items
NI_FY2024 = 426.0
NI_FY2025 = 662.5
NI_TTM = 721.1

# EPS diluted
EPS_FY2022 = 3.21
EPS_FY2023 = 0.98
EPS_FY2024 = 11.70
EPS_FY2025 = 18.16
EPS_TTM = 19.86

# Operating cash flow (in $M)
OCF_FY2022 = 722.3
OCF_FY2023 = 361.7
OCF_FY2024 = 806.4
OCF_FY2025 = 1178.1
OCF_TTM = 1258.0

# Capex (in $M, absolute)
CAPEX_FY2022 = 17.8
CAPEX_FY2023 = 11.9
CAPEX_FY2024 = 10.2
CAPEX_FY2025 = 7.7
CAPEX_TTM = 9.4

# Book value equity (in $M)
BV_FY2022 = 2333.7
BV_FY2023 = 2465.6
BV_FY2024 = 2841.8
BV_FY2025 = 3571.5

# Analyst estimates
ANALYST_EPS_FY2026 = 18.60
ANALYST_EPS_FY2027 = 18.59
ANALYST_REV_FY2026 = 6640.0  # $6.64B avg
ANALYST_REV_FY2027 = 6930.0  # $6.93B avg

# Next earnings
NEXT_EARNINGS = "2026-07-28"

DATE_STR = "2026-06-29"

# ── WACC calculation ──
cost_of_equity = RISK_FREE + BETA * ERP  # 4.374 + 0.32*5 = 5.974
# cost of debt ~ 4.5% (from interest expense $33.3M / debt ~$783M in FY2024)
cost_of_debt = 4.5
# weights: equity = MC / (MC + Debt)
eq_weight = MC_B / (MC_B + TOTAL_DEBT_MM / 1000)
debt_weight = 1 - eq_weight
WACC = eq_weight * cost_of_equity + debt_weight * cost_of_debt * (1 - TAX_RATE_USED / 100)

# ═══════════════════════════════════════════════════════
# SHEET 1: Valuation
# ═══════════════════════════════════════════════════════
ws = wb.active
ws.title = "Valuation"

ws.merge_cells("A1:G1")
ws["A1"] = "The Hanover Insurance Group, Inc. (THG) — P/B & ROE Valuation Model"
ws["A1"].font = title_font
ws.merge_cells("A2:G2")
ws["A2"] = f"Date: {DATE_STR} | Ticker: NYSE: THG | Price: ${PRICE:,.2f} | Shares: {SHARES_MM:.1f}M"
ws["A2"].font = bold
ws.merge_cells("A3:G3")
ws["A3"] = f"Market Cap: ${MC_B:.2f}B | Enterprise Value: ${EV_B:.2f}B | Primary Lens: P/B & ROE (Insurance)"
ws["A3"].font = bold
ws.merge_cells("A4:G4")
ws["A4"] = "Insurance company — FCF multiples marked N/A. Deposits are operating liabilities, not financial debt."
ws["A4"].font = Font(italic=True, color="C00000")

# Title block data
title_data = [
    ("Company", "The Hanover Insurance Group, Inc."),
    ("Date", DATE_STR),
    ("Ticker", "NYSE: THG"),
    ("Price", f"${PRICE:,.2f}"),
    ("Shares Outstanding (dil avg)", f"{SHARES_MM:.2f}M"),
    ("Market Cap", f"${MC_B:.2f}B"),
    ("Enterprise Value", f"${EV_B:.2f}B"),
    ("Primary Lens", "P/B and ROE (Insurance/Financial Services)"),
    ("Stance", "Watch — strong ROE growth, but insurance cycles matter"),
]

for i, (label, val) in enumerate(title_data, start=6):
    ws.cell(row=i, column=1, value=label).font = bold
    ws.cell(row=i, column=2, value=val)
    ws.cell(row=i, column=1).border = thin_border
    ws.cell(row=i, column=2).border = thin_border

# Valuation metrics table
ws.cell(row=16, column=1, value="Metric").font = heading_font
ws.cell(row=16, column=2, value="Value").font = heading_font
ws.cell(row=16, column=3, value="Comment").font = heading_font
style_header(ws, 16, 3)

metrics = [
    ("Trailing P/E", f"{MC_B*1000/NI_TTM:.1f}x", f"P/E ~{MC_B*1000/NI_TTM:.1f}x on TTM NI of ${NI_TTM:.0f}M. Reasonable for insurance."),
    ("Forward P/E (FY2026)", f"{MC_B*1000/(ANALYST_EPS_FY2026*SHARES_MM):.1f}x", f"Using FY26 consensus EPS ${ANALYST_EPS_FY2026} × {SHARES_MM}M"),
    ("Forward P/E (FY2027)", f"{MC_B*1000/(ANALYST_EPS_FY2027*SHARES_MM):.1f}x", f"Using FY27 consensus EPS ${ANALYST_EPS_FY2027} × {SHARES_MM}M"),
    ("P/Sales (TTM)", f"{MC_B*1000/REV_TTM:.2f}x", f"MC ${MC_B}B / Rev ${REV_TTM/1000:.2f}B"),
    ("P/Book", f"{MC_B*1000/BV_FY2025:.2f}x", f"MC ${MC_B}B / BV ${BV_FY2025/1000:.2f}B. Current P/B {MC_B*1000/BV_FY2025:.2f}x vs. historical 1.5-2.5x range."),
    ("BVPS (Current Q1 2026)", f"${BVPS_CURRENT:.2f}", "Yahoo Finance mrq book value per share"),
    ("ROE (TTM)", f"{NI_TTM/BV_FY2025*100:.1f}%", f"TTM NI ${NI_TTM:.0f}M / FY25 BV ${BV_FY2025/1000:.0f}B = ~{NI_TTM/BV_FY2025*100:.1f}%"),
    ("P/FCF", "N/A", "Insurance — FCF is not a meaningful valuation metric"),
    ("EV/FCF", "N/A", "Insurance — FCF is not a meaningful valuation metric"),
    ("P/FCF (Yahoo LFCF)", "N/A", "Yahoo shows 'Levered FCF TTM $1B' but this is OCF-Capex for an insurer; deposits offset loan origination"),
    ("EV/Sales", f"{EV_B*1000/REV_TTM:.2f}x", f"EV ${EV_B}B / Rev ${REV_TTM/1000:.2f}B"),
    ("EV/EBITDA", "N/A", "Not standard for insurance companies"),
]

for i, (metric, value, comment) in enumerate(metrics, start=17):
    ws.cell(row=i, column=1, value=metric).font = bold
    ws.cell(row=i, column=2, value=value)
    ws.cell(row=i, column=3, value=comment)
    style_data(ws, i, 3)

ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 65

# ═══════════════════════════════════════════════════════
# SHEET 2: WACC
# ═══════════════════════════════════════════════════════
ws2 = wb.create_sheet("WACC")

ws2.merge_cells("A1:D1")
ws2["A1"] = "THG — WACC (CAPM) — Insurance Company Override"
ws2["A1"].font = title_font
ws2.merge_cells("A2:D2")
ws2["A2"] = "Note: FCF-based DCF is NOT appropriate for insurance/financial. WACC shown for Residual Income framework only."
ws2["A2"].font = Font(italic=True, color="C00000")

wacc_data = [
    ["Component", "Value", "Source / Notes"],
    ["Risk-Free Rate (10Y US Treasury)", f"{RISK_FREE:.3f}%", "CNBC US10Y as of 2026-06-29"],
    ["Equity Risk Premium", f"{ERP:.1f}%", "Standard assumption"],
    ["Beta (Levered, 5Y Monthly)", f"{BETA:.2f}", "Yahoo Finance Statistics"],
    ["Cost of Equity (CAPM)", f"{cost_of_equity:.2f}%", f"={RISK_FREE:.3f} + {BETA:.2f} × {ERP:.1f}"],
    ["Cost of Debt", f"{cost_of_debt:.1f}%", f"Interest exp ~$33.3M / Debt ~$783M (FY2024)"],
    ["Tax Rate", f"{TAX_RATE_USED:.1f}%", f"FY2025: Tax prov $183.1M / Pretax $843.8M"],
    ["", "", ""],
    ["CAPITAL STRUCTURE", "", ""],
    ["Market Cap", f"${MC_B:.2f}B", "Yahoo Finance intraday"],
    ["Total Debt (Q1 2026 mrq)", f"${TOTAL_DEBT_MM/1000:.1f}B", "Yahoo Finance Balance Sheet"],
    ["Total Market Value", f"${MC_B + TOTAL_DEBT_MM/1000:.2f}B", ""],
    ["Equity Weight", f"{eq_weight:.4f}", ""],
    ["Debt Weight", f"{debt_weight:.4f}", ""],
    ["", "", ""],
    ["WACC", f"{WACC:.2f}%", f"={eq_weight:.4f}×{cost_of_equity:.2f}% + {debt_weight:.4f}×{cost_of_debt:.1f}%×(1-{TAX_RATE_USED:.1f}%)"],
    ["", "", ""],
    ["Primary Framework", "Residual Income (ROE vs. Cost of Equity)", "Insurance company — P/B is the anchor"],
    ["Fair P/B = ROE / (Cost of Equity - g)", f"ROE ~{NI_TTM/BV_FY2025*100:.1f}% vs Ke ~{cost_of_equity:.1f}%", "Residual Income formula for banks/insurers"],
]

for i, row_data in enumerate(wacc_data, 4):
    for j, val in enumerate(row_data, 1):
        cell = ws2.cell(row=i, column=j, value=val)
        if i == 4:
            cell.font = heading_font
        cell.border = thin_border
        if j == 1:
            cell.font = bold

ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 55

# ═══════════════════════════════════════════════════════
# SHEET 3: Scenarios (P/B based for insurance)
# ═══════════════════════════════════════════════════════
ws3 = wb.create_sheet("Scenarios")

ws3.merge_cells("A1:L1")
ws3["A1"] = "THG — Bear / Base / Bull Scenarios (P/B + ROE Framework)"
ws3["A1"].font = title_font
ws3.merge_cells("A2:L2")
ws3["A2"] = "Insurance company: valuations based on BVPS CAGR, exit P/B multiples, and implied price per share."
ws3["A2"].font = Font(italic=True, color="C00000")

# Scenario parameters
# Current BVPS ~$100.86, current shares ~35M
# Current equity $3.57B
# ROE TTM ~20.2% on FY25 BV

# Bear: BVPS CAGR 4%, exit P/B 1.6x (insurance cyclical downturn)
# Base: BVPS CAGR 8%, exit P/B 2.0x (current multiple maintenance)
# Bull: BVPS CAGR 12%, exit P/B 2.4x (ROE premium re-rating)

current_bvps = BVPS_CURRENT
current_price = PRICE

scenarios = {
    "Bear": {"bvps_cagr": 0.04, "exit_pb": 1.60, "weight": 0.20},
    "Base": {"bvps_cagr": 0.08, "exit_pb": 2.00, "weight": 0.50},
    "Bull": {"bvps_cagr": 0.12, "exit_pb": 2.40, "weight": 0.30},
}

# Headers
headers = [
    "Scenario", "BVPS CAGR (5Y)", "Terminal BVPS (5Y)",
    "Exit P/B", "Target Price",
    "Upside from Current", "Weight", "Weighted Value/Share",
    "Key Assumptions"
]
for j, h in enumerate(headers, 1):
    ws3.cell(row=4, column=j, value=h)
style_header(ws3, 4, len(headers))

row_num = 5
for name, params in scenarios.items():
    terminal_bvps = current_bvps * (1 + params["bvps_cagr"])**5
    target = terminal_bvps * params["exit_pb"]
    upside = (target / current_price - 1) * 100
    weighted = target * params["weight"]

    ws3.cell(row=row_num, column=1, value=name).font = bold
    ws3.cell(row=row_num, column=2, value=params["bvps_cagr"])
    ws3.cell(row=row_num, column=2).number_format = '0.0%'
    ws3.cell(row=row_num, column=3, value=terminal_bvps)
    ws3.cell(row=row_num, column=3).number_format = '$#,##0.00'
    ws3.cell(row=row_num, column=4, value=params["exit_pb"])
    ws3.cell(row=row_num, column=4).number_format = '0.0'
    ws3.cell(row=row_num, column=5, value=target)
    ws3.cell(row=row_num, column=5).number_format = '$#,##0.00'
    ws3.cell(row=row_num, column=6, value=upside / 100)
    ws3.cell(row=row_num, column=6).number_format = '0.0%'
    ws3.cell(row=row_num, column=7, value=params["weight"])
    ws3.cell(row=row_num, column=7).number_format = '0.0%'
    ws3.cell(row=row_num, column=8, value=weighted)
    ws3.cell(row=row_num, column=8).number_format = '$#,##0.00'

    assumptions = {
        "Bear": "Hard insurance cycle; combined ratio rises; ROE drops to 12-14%; P/B contracts to 1.6x. Book growth slows on reserves.",
        "Base": "Moderate cycle; ROE holds at 18-20%; P/B maintains current 2.0x. Steady buybacks support BVPS growth.",
        "Bull": "Soft cycle; pricing power; combined ratio improves; ROE expands to 22-25%; P/B re-rates to 2.4x on quality premium.",
    }
    ws3.cell(row=row_num, column=9, value=assumptions[name])
    style_data(ws3, row_num, len(headers))
    row_num += 1

# Totals row
ws3.cell(row=row_num+1, column=1, value="Total Probability-Weighted FV").font = bold
total_wv = sum(
    current_bvps * (1 + p["bvps_cagr"])**5 * p["exit_pb"] * p["weight"]
    for p in scenarios.values()
)
ws3.cell(row=row_num+1, column=8, value=total_wv)
ws3.cell(row=row_num+1, column=8).number_format = '$#,##0.00'
total_upside = (total_wv / current_price - 1) * 100
ws3.cell(row=row_num+1, column=6, value=total_upside / 100)
ws3.cell(row=row_num+1, column=6).number_format = '0.0%'
for j in range(1, len(headers) + 1):
    ws3.cell(row=row_num+1, column=j).font = bold
    ws3.cell(row=row_num+1, column=j).fill = green_fill
    ws3.cell(row=row_num+1, column=j).border = thin_border

ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 10
ws3.column_dimensions["E"].width = 14
ws3.column_dimensions["F"].width = 18
ws3.column_dimensions["G"].width = 10
ws3.column_dimensions["H"].width = 18
ws3.column_dimensions["I"].width = 70

# ═══════════════════════════════════════════════════════
# SHEET 4: Actuals Source Audit
# ═══════════════════════════════════════════════════════
ws4 = wb.create_sheet("Actuals Source Audit")

ws4.merge_cells("A1:E1")
ws4["A1"] = "THG — Actuals Source Audit"
ws4["A1"].font = title_font

audit_headers = ["Data Point", "Value", "Source", "Date Accessed", "Notes"]
style_header(ws4, 3, 5)
for j, h in enumerate(audit_headers, 1):
    ws4.cell(row=3, column=j, value=h)

audit_data = [
    ["Stock Price", f"${PRICE:.2f}", "Yahoo Finance /quote/THG/", "2026-06-29", "Close price June 29, 2026"],
    ["After Hours Price", "$216.25", "Yahoo Finance /quote/THG/", "2026-06-29", "After hours 6:02 PM EDT"],
    ["Market Cap", f"${MC_B:.2f}B", "Yahoo Finance Statistics", "2026-06-29", "Intraday market cap"],
    ["Enterprise Value", f"${EV_B:.2f}B", "Yahoo Finance Statistics", "2026-06-29", ""],
    ["Shares Outstanding", f"{SHARES_MM:.2f}M diluted", "Yahoo Finance Statistics", "2026-06-29", "Implied shares outstanding"],
    ["Beta (5Y Monthly)", f"{BETA:.2f}", "Yahoo Finance Statistics", "2026-06-29", ""],
    ["52 Week Range", "$160.70 - $214.76", "Yahoo Finance Summary", "2026-06-29", "Stock at 52-week high"],
    ["", "", "", "", ""],
    ["Revenue FY2022", f"${REV_FY2022:.1f}M", "Yahoo Finance Income Statement", "2026-06-29", "All numbers in thousands"],
    ["Revenue FY2023", f"${REV_FY2023:.1f}M", "Yahoo Finance Income Statement", "2026-06-29", ""],
    ["Revenue FY2024", f"${REV_FY2024:.1f}M", "Yahoo Finance Income Statement", "2026-06-29", ""],
    ["Revenue FY2025", f"${REV_FY2025:.1f}M", "Yahoo Finance Income Statement", "2026-06-29", ""],
    ["Revenue TTM", f"${REV_TTM:.1f}M", "Yahoo Finance Income Statement", "2026-06-29", ""],
    ["Net Income FY2022", f"${NI_FY2022:.1f}M", "Yahoo Finance Income Statement", "2026-06-29", ""],
    ["Net Income FY2023", f"${NI_FY2023:.1f}M", "Yahoo Finance Income Statement", "2026-06-29", "Severe one-time items — not representative"],
    ["Net Income FY2024", f"${NI_FY2024:.1f}M", "Yahoo Finance Income Statement", "2026-06-29", ""],
    ["Net Income FY2025", f"${NI_FY2025:.1f}M", "Yahoo Finance Income Statement", "2026-06-29", ""],
    ["Net Income TTM", f"${NI_TTM:.1f}M", "Yahoo Finance Income Statement", "2026-06-29", ""],
    ["", "", "", "", ""],
    ["Total Assets FY2025", "$16,945.9M", "Yahoo Finance Balance Sheet", "2026-06-29", "Q1 2026 mrq — jumped 11% from FY25. Possible acquisition?"],
    ["Total Debt Q1 2026", f"${TOTAL_DEBT_MM:.1f}M", "Yahoo Finance Statistics", "2026-06-29", ""],
    ["Total Equity FY2025", f"${EQUITY_MM:.1f}M", "Yahoo Finance Balance Sheet", "2026-06-29", ""],
    ["BVPS mrq", f"${BVPS_CURRENT:.2f}", "Yahoo Finance Statistics", "2026-06-29", ""],
    ["", "", "", "", ""],
    ["OCF FY2025", f"${OCF_FY2025:.1f}M", "Yahoo Finance Cash Flow", "2026-06-29", ""],
    ["OCF TTM", f"${OCF_TTM:.1f}M", "Yahoo Finance Cash Flow", "2026-06-29", ""],
    ["Capex TTM", f"${CAPEX_TTM:.1f}M", "Yahoo Finance Cash Flow", "2026-06-29", "Minimal capex — insurance company"],
    ["", "", "", "", ""],
    ["FY26 Rev Estimate", f"${ANALYST_REV_FY2026:.0f}M", "Yahoo Finance Analysis", "2026-06-29", "6 analysts, avg $6.64B"],
    ["FY27 Rev Estimate", f"${ANALYST_REV_FY2027:.0f}M", "Yahoo Finance Analysis", "2026-06-29", "6 analysts, avg $6.93B"],
    ["FY26 EPS Estimate", f"${ANALYST_EPS_FY2026:.2f}", "Yahoo Finance Analysis", "2026-06-29", "8 analysts, avg $18.60"],
    ["FY27 EPS Estimate", f"${ANALYST_EPS_FY2027:.2f}", "Yahoo Finance Analysis", "2026-06-29", "8 analysts, avg $18.59"],
    ["Q1 FY26 EPS Actual", "$5.25", "Yahoo Finance Analysis", "2026-06-29", "Beat of +24.5% est $4.22"],
    ["", "", "", "", ""],
    ["10Y Treasury Rate", f"{RISK_FREE:.3f}%", "CNBC US10Y", "2026-06-29", ""],
    ["Next Earnings Date", NEXT_EARNINGS, "Yahoo Finance Profile", "2026-06-29", "July 28, 2026"],
    ["Dividend Yield", "1.74%", "Yahoo Finance Statistics", "2026-06-29", "Trailing annual $3.70"],
    ["Payout Ratio", "18.67%", "Yahoo Finance Statistics", "2026-06-29", "Very low — room for buybacks + dividends"],
]

for i, row_data in enumerate(audit_data, 4):
    for j, val in enumerate(row_data, 1):
        cell = ws4.cell(row=i, column=j, value=val)
        cell.border = thin_border
        if j == 1:
            cell.font = bold

ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 20
ws4.column_dimensions["C"].width = 45
ws4.column_dimensions["D"].width = 15
ws4.column_dimensions["E"].width = 60

# ═══════════════════════════════════════════════════════
# SHEET 5: Questions
# ═══════════════════════════════════════════════════════
ws5 = wb.create_sheet("Questions")

ws5.merge_cells("A1:C1")
ws5["A1"] = "THG — Open Questions"
ws5["A1"].font = title_font

ws5.cell(row=3, column=1, value="#")
ws5.cell(row=3, column=2, value="Question")
ws5.cell(row=3, column=3, value="Priority")
style_header(ws5, 3, 3)

questions = [
    ["1", "Asset jump: Total assets surged $16,946M (Q1 2026) vs $15,275M (FY2025) — an 11% jump in one quarter. Was this investment portfolio appreciation, a reinsurance acquisition, or something else?", "HIGH"],
    ["2", "Debt increase: Total debt rose from $784M (FY2025) to $844M (Q1 2026). Is this deal-related or ordinary refinancing?", "MEDIUM"],
    ["3", "FY2023 earnings collapse: Net income crashed to just $35.3M (EPS $0.98) in FY2023 — down 71% from FY2022. What one-time items or catastrophic combined ratio drove this?", "HIGH"],
    ["4", "Combined ratio trend: What is the combined ratio trajectory under current hard insurance cycle? Does this mean structural improvement or cyclical temporary?", "HIGH"],
    ["5", "Share count: Basic shares 35.6M diluted vs 36.3M diluted avg. Is the company consistently executing buybacks or is there option dilution offsetting?", "MEDIUM"],
    ["6", "Insurance cycle position: Where is the US property & casualty insurance cycle? Hardening pricing power supports margin expansion — but how long can it last?", "HIGH"],
    ["7", "Investment portfolio: As an insurer, THG holds a large float/investment portfolio. What returns on float does it generate? Duration/credit risk?", "MEDIUM"],
    ["8", "Segment concentration: Core Commercial vs Specialty vs Personal Lines — any customer or segment concentration risk?", "MEDIUM"],
    ["9", "Guidance: Management guidance for FY2026? Q1 showed a 24.5% EPS beat — is this repeatable?", "HIGH"],
    ["10", "Geographic exposure: What is the geographic distribution of risk? Hurricane/ catastrophe exposure given Worcester, MA HQ?", "MEDIUM"],
    ["11", "Next earnings: July 28, 2026 — Q2 results could confirm whether Q1 beat was sustainable. Key catalyst for P/B re-rating view.", "HIGH"],
]

for i, (num, q, pri) in enumerate(questions, 4):
    ws5.cell(row=i, column=1, value=num)
    ws5.cell(row=i, column=2, value=q)
    ws5.cell(row=i, column=3, value=pri)
    for j in range(1, 4):
        ws5.cell(row=i, column=j).border = thin_border
    if pri == "HIGH":
        ws5.cell(row=i, column=3).fill = red_fill
    else:
        ws5.cell(row=i, column=3).fill = yellow_fill

ws5.column_dimensions["A"].width = 5
ws5.column_dimensions["B"].width = 90
ws5.column_dimensions["C"].width = 12

# ═══════════════════════════════════════════════════════
# SHEET 6: Sources
# ═══════════════════════════════════════════════════════
ws6 = wb.create_sheet("Sources")

ws6.merge_cells("A1:C1")
ws6["A1"] = "THG — Sources"
ws6["A1"].font = title_font

ws6.cell(row=3, column=1, value="#")
ws6.cell(row=3, column=2, value="Source")
ws6.cell(row=3, column=3, value="URL")
style_header(ws6, 3, 3)

sources = [
    ["1", "Yahoo Finance — THG Quote & Summary", "https://finance.yahoo.com/quote/THG/"],
    ["2", "Yahoo Finance — THG Income Statement", "https://finance.yahoo.com/quote/THG/financials/"],
    ["3", "Yahoo Finance — THG Balance Sheet", "https://finance.yahoo.com/quote/THG/balance-sheet/"],
    ["4", "Yahoo Finance — THG Cash Flow", "https://finance.yahoo.com/quote/THG/cash-flow/"],
    ["5", "Yahoo Finance — THG Key Statistics", "https://finance.yahoo.com/quote/THG/key-statistics/"],
    ["6", "Yahoo Finance — THG Analysis / Estimates", "https://finance.yahoo.com/quote/THG/analysis/"],
    ["7", "Yahoo Finance — THG Company Profile", "https://finance.yahoo.com/quote/THG/profile/"],
    ["8", "CNBC — US 10-Year Treasury (US10Y)", "https://cnbc.com/quotes/US10Y"],
    ["9", "StockAnalysis.com — THG (404 — not available)", "N/A"],
    ["10", "ISS Governance Quality Score (via Yahoo Finance)", "https://finance.yahoo.com/quote/THG/profile/"],
]

for i, (num, source, url) in enumerate(sources, 4):
    ws6.cell(row=i, column=1, value=num)
    ws6.cell(row=i, column=2, value=source)
    ws6.cell(row=i, column=3, value=url)
    for j in range(1, 4):
        ws6.cell(row=i, column=j).border = thin_border

ws6.column_dimensions["A"].width = 5
ws6.column_dimensions["B"].width = 60
ws6.column_dimensions["C"].width = 65

# ── Save ─────────────────────────────────────────────
OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       f"[{DATE_STR}] The Hanover Insurance Group Model.xlsx")
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"WACC: {WACC:.2f}%")
print(f"Probability-weighted FV: ${total_wv:.2f}/share ({total_upside:+.1f}%)")
print(f"Scenarios: Bear ${current_bvps*(1.04)**5*1.6:.2f}, Base ${current_bvps*(1.08)**5*2.0:.2f}, Bull ${current_bvps*(1.12)**5*2.4:.2f}")
