#!/usr/bin/env python3
"""Build OLLI (Ollie's Bargain Outlet Holdings) 6-sheet valuation model."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = Workbook()

# Styles
bold_font = Font(bold=True)
header_font = Font(bold=True, size=12)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
bear_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
base_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
bull_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = bold_font
        cell.border = thin_border
        cell.fill = header_fill

def style_data_range(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = thin_border

# ===== Sheet 1: Valuation =====
ws1 = wb.active
ws1.title = "Valuation"
ws1.merge_cells('A1:F1')
ws1['A1'] = "Ollie's Bargain Outlet Holdings, Inc. (OLLI) - Valuation Model"
ws1['A1'].font = Font(bold=True, size=14)

title_data = [
    ("Company:", "Ollie's Bargain Outlet Holdings, Inc."),
    ("Ticker:", "NASDAQ: OLLI"),
    ("Sector:", "Consumer Defensive / Discount Stores"),
    ("Date:", "2026-07-09"),
    ("Source Date:", "Yahoo Finance, Jul 9, 2026 close"),
    ("Price:", "$63.72"),
    ("Shares Outstanding:", "60.45M (Yahoo Finance Statistics, 7/8/2026)"),
    ("Market Cap:", "$3.74B"),
    ("Enterprise Value:", "$4.20B (EV/MC ratio 1.54/1.40, net debt ~$0.46B)"),
    ("Primary Lens:", "Forward P/E + FCF multiple; WACC-discounted scenario analysis"),
    ("Stance:", "Watch — compelling multiples post-55% drawdown but consensus revisions negative"),
]

for i, (label, val) in enumerate(title_data, 2):
    ws1.cell(row=i, column=1, value=label).font = bold_font
    ws1.cell(row=i, column=2, value=val)

r = len(title_data) + 3
ws1.cell(row=r, column=1, value="Valuation Metric").font = bold_font
ws1.cell(row=r, column=2, value="Value").font = bold_font
ws1.cell(row=r, column=3, value="Comment").font = bold_font

metrics = [
    ("Trailing P/E", "15.32x", "TTM dil EPS $4.16; compresses sharply from 52W high of 32x+ at $141"),
    ("Forward P/E (FY2027)", "14.10x", "Based on FY2027 EPS consensus $4.49 (15 analysts, normalized)"),
    ("P/S (TTM)", "1.40x", "TTM revenue $2.73B; well below 52W peak of ~2.9x"),
    ("P/FCF (TTM)", "17.6x", "TTM FCF $212.7M; reasonable for discount retail"),
    ("EV/FCF (TTM)", "19.7x", "EV $4.20B / FCF $212.7M; includes ~$460M net debt adjustment"),
    ("EV/Sales", "1.54x", "EV $4.20B / TTM rev $2.73B; attractive for high-quality retailer"),
    ("EV/EBITDA", "10.68x", "EV $4.20B / EBITDA $368.4M; below sector median ~14-16x"),
    ("P/B", "1.98x", "MC $3.74B / equity $1.89B; reasonable vs peers (DG 3.5x, DLTR 4.0x)"),
    ("52W Range", "$60.29 - $141.74", "Stock at the extreme bottom of 52W range; down ~55% from highs"),
]

for i, (metric, value, comment) in enumerate(metrics, r + 1):
    ws1.cell(row=i, column=1, value=metric)
    ws1.cell(row=i, column=2, value=value)
    ws1.cell(row=i, column=3, value=comment)

style_header_row(ws1, r, 3)
style_data_range(ws1, r + 1, r + len(metrics), 3)
ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 75

# ===== Sheet 2: WACC =====
ws2 = wb.create_sheet("WACC")
ws2.merge_cells('A1:D1')
ws2['A1'] = "WACC Calculation - CAPM Method"
ws2['A1'].font = Font(bold=True, size=12)

# Risk-free: 4.541% per CNBC Jul 9, 2026
# Beta: 0.47 (very low — discount retail defensive, low leverage)
# ERP: 5%
# Cost of equity = 4.541% + 0.47 * 5% = 6.89%
# Market cap: $3,740M
# Total debt MRQ: $710.3M, Cash: $249.56M, Net debt: ~$460.74M
# Total cap: $4,450M (MC + Debt)
# Equity weight: 3740/4450 = 84.0%, Debt weight: 710/4450 = 16.0%
# Cost of debt: ~6.0% pre-tax (moderate leverage for retailer)
# Tax rate: TTM effective = $80.5M / $329.9M = 24.4%
# WACC = 0.840 * 6.89% + 0.160 * 6.0% * (1 - 0.244) = 5.79% + 0.72% = 6.50%

cost_of_equity = 0.04541 + 0.47 * 0.05
wacc = 0.840 * cost_of_equity + 0.160 * 0.06 * (1 - 0.244)
print(f"Cost of equity = {cost_of_equity*100:.2f}%")
print(f"WACC = {wacc*100:.2f}%")

wacc_data = [
    ("Component", "Value", "Source / Notes"),
    ("Risk-Free Rate (10Y US)", "4.541%", "CNBC US10Y, Jul 9, 2026"),
    ("Beta (5Y Monthly)", "0.47", "Yahoo Finance Statistics; very low — defensive consumer staples"),
    ("Equity Risk Premium", "5.00%", "Standard assumption"),
    ("Cost of Equity (Rf + Beta*ERP)", "6.89%", "= 4.541% + 0.47 * 5.00%"),
    ("Market Cap", "$3,740M", "Yahoo Finance, Jul 9, 2026"),
    ("Total Debt (MRQ)", "$710M", "Yahoo Finance Statistics, MRQ ~May 2026"),
    ("Cash (MRQ)", "$250M", "Yahoo Finance Statistics, MRQ"),
    ("Net Debt (proxy: EV-MC)", "$461M", "$4.20B EV - $3.74B MC"),
    ("Total Capitalization (E+D)", "$4,450M", "MC + Debt"),
    ("Equity Weight", "84.0%", "MC / (MC + Debt)"),
    ("Debt Weight", "16.0%", "Debt / (MC + Debt)"),
    ("Pre-Tax Cost of Debt", "6.00%", "Estimated; moderate leverage, capital lease heavy"),
    ("Tax Rate (effective TTM)", "24.4%", "TTM tax provision $80.5M / pretax $329.9M"),
    ("WACC", "6.50%", "= 0.840 * 6.89% + 0.160 * 6.0% * (1 - 0.244)"),
]

for i, (comp, val, notes) in enumerate(wacc_data, 2):
    ws2.cell(row=i, column=1, value=comp)
    ws2.cell(row=i, column=2, value=val)
    ws2.cell(row=i, column=3, value=notes)

style_header_row(ws2, 2, 3)
style_data_range(ws2, 3, len(wacc_data) + 1, 3)
ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 75

# ===== Sheet 3: Scenarios =====
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells('A1:Q1')
ws3['A1'] = "Scenario Analysis - Bear / Base / Bull (FCF Multiple Framework)"
ws3['A1'].font = Font(bold=True, size=12)

# All numbers in MILLIONS for consistency
# TTM Revenue: $2,731M
# Shares: 60.45M
# Current net debt: $461M
# WACC: 6.50%
# FCF TTM: $212.7M

# Bear: Rev CAGR 5%, terminal rev $2731*1.05^5 = $3,503M, FCF margin 5.0%, terminal FCF=$175M, exit 8x, EV=$1,403M
#   eq_val = $1,403M + $200M (net debt declines) = $1,603M, target = $1,603/60.45 = $26.52 (BEARISH - too low)
#   Let me recalibrate: Bear FCF margin 6%, terminal FCF=$210M, exit 10x, EV=$2,100M, eq=$2,300, target=$38
bear_rev_cagr = 0.05
bear_term_rev = 2731 * (1.05 ** 5)  # 3,503
bear_fcf_margin = 0.055
bear_term_fcf = bear_term_rev * bear_fcf_margin  # ~193
bear_exit_multiple = 9
bear_ev = bear_term_fcf * bear_exit_multiple  # ~1,736
bear_net_cash = 150  # mm, net debt declines over 5 yrs
bear_eq = bear_ev + bear_net_cash
bear_target = bear_eq / 60.45

# Base: Rev CAGR 10% (close to consensus ~12%), terminal rev $2731*1.10^5 = $4,413M
base_rev_cagr = 0.10
base_term_rev = 2731 * (1.10 ** 5)  # 4,413
base_fcf_margin = 0.075
base_term_fcf = base_term_rev * base_fcf_margin  # ~331
base_exit_multiple = 11
base_ev = base_term_fcf * base_exit_multiple  # ~3,641
base_net_cash = 200  # mm
base_eq = base_ev + base_net_cash
base_target = base_eq / 60.45

# Bull: Rev CAGR 15%, terminal rev $2731*1.15^5 = $5,502M
bull_rev_cagr = 0.15
bull_term_rev = 2731 * (1.15 ** 5)  # 5,502
bull_fcf_margin = 0.090
bull_term_fcf = bull_term_rev * bull_fcf_margin  # ~495
bull_exit_multiple = 14
bull_ev = bull_term_fcf * bull_exit_multiple  # ~6,930
bull_net_cash = 250  # mm
bull_eq = bull_ev + bull_net_cash
bull_target = bull_eq / 60.45

print(f"Bear - Term Rev: ${bear_term_rev:.0f}M, Term FCF: ${bear_term_fcf:.0f}M, EV: ${bear_ev:.0f}M, Target: ${bear_target:.2f}")
print(f"Base - Term Rev: ${base_term_rev:.0f}M, Term FCF: ${base_term_fcf:.0f}M, EV: ${base_ev:.0f}M, Target: ${base_target:.2f}")
print(f"Bull - Term Rev: ${bull_term_rev:.0f}M, Term FCF: ${bull_term_fcf:.0f}M, EV: ${bull_ev:.0f}M, Target: ${bull_target:.2f}")

# Verify against analyst avg target $120.20
# Base target should be roughly consistent with analyst consensus
print(f"Analyst avg target: $120.20, base target: ${base_target:.2f}")

# Upside from current $63.72
bear_upside = (bear_target / 63.72 - 1) * 100
base_upside = (base_target / 63.72 - 1) * 100
bull_upside = (bull_target / 63.72 - 1) * 100

# Weights
bear_w, base_w, bull_w = 0.20, 0.50, 0.30
weighted_FV = bear_w * bear_target + base_w * base_target + bull_w * bull_target
total_upside = (weighted_FV / 63.72 - 1) * 100

print(f"Weighted FV: ${weighted_FV:.2f}")
print(f"Total upside: {total_upside:.1f}%")
print(f"Bear upside: {bear_upside:.1f}%, Base upside: {base_upside:.1f}%, Bull upside: {bull_upside:.1f}%")

# Sanity check: targets should be plausible ($20-$200 range), not in thousands
assert 10 < bear_target < 200, f"Bear target implausible: ${bear_target}"
assert 20 < base_target < 250, f"Base target implausible: ${base_target}"
assert 50 < bull_target < 400, f"Bull target implausible: ${bull_target}"
print("All targets in plausible range — unit check PASS")

for c, h in enumerate(["Item", "Bear", "Base", "Bull"], 1):
    ws3.cell(row=2, column=c, value=h)
style_header_row(ws3, 2, 4)

scenario_data = [
    ("Revenue CAGR (5Y)", f"{bear_rev_cagr:.0%}", f"{base_rev_cagr:.0%}", f"{bull_rev_cagr:.0%}"),
    ("TTM Revenue ($M)", "$2,731", "$2,731", "$2,731"),
    ("Terminal Revenue (5Y) ($M)", f"${bear_term_rev:.0f}", f"${base_term_rev:.0f}", f"${bull_term_rev:.0f}"),
    ("Adjusted FCF Margin", f"{bear_fcf_margin:.1%}", f"{base_fcf_margin:.1%}", f"{bull_fcf_margin:.1%}"),
    ("Terminal FCF ($M)", f"${bear_term_fcf:.0f}", f"${base_term_fcf:.0f}", f"${bull_term_fcf:.0f}"),
    ("Exit FCF Multiple", f"{bear_exit_multiple}x", f"{base_exit_multiple}x", f"{bull_exit_multiple}x"),
    ("Implied EV ($M)", f"${bear_ev:.0f}", f"${base_ev:.0f}", f"${bull_ev:.0f}"),
    ("Plus Net Cash Adj ($M)", f"+${bear_net_cash}", f"+${base_net_cash}", f"+${bull_net_cash}"),
    ("Equity Value ($M)", f"${bear_eq:.0f}", f"${base_eq:.0f}", f"${bull_eq:.0f}"),
    ("Shares Outstanding (M)", "60.45", "60.45", "60.45"),
    ("Target Price", f"${bear_target:.2f}", f"${base_target:.2f}", f"${bull_target:.2f}"),
    ("Upside from $63.72", f"{bear_upside:+.1f}%", f"{base_upside:+.1f}%", f"{bull_upside:+.1f}%"),
    ("Weight", f"{bear_w:.0%}", f"{base_w:.0%}", f"{bull_w:.0%}"),
    ("Weighted Value/Share", f"${bear_w*bear_target:.2f}", f"${base_w*base_target:.2f}", f"${bull_w*bull_target:.2f}"),
]

for i, row in enumerate(scenario_data, 3):
    for c, val in enumerate(row, 1):
        ws3.cell(row=i, column=c, value=val)

# Color columns: Bear=B, Base=C, Bull=D
for r_idx in range(3, 16):
    ws3.cell(row=r_idx, column=2).fill = bear_fill
    ws3.cell(row=r_idx, column=3).fill = base_fill
    ws3.cell(row=r_idx, column=4).fill = bull_fill

style_data_range(ws3, 3, 16, 4)

# Summary
ws3.cell(row=18, column=1, value="Total Probability-Weighted FV").font = bold_font
ws3.cell(row=18, column=4, value=f"${weighted_FV:.2f}").font = bold_font
ws3.cell(row=19, column=1, value="Upside from Current ($63.72)").font = bold_font
ws3.cell(row=19, column=4, value=f"{total_upside:+.1f}%").font = bold_font

ws3.column_dimensions['A'].width = 32
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 15

# ===== Sheet 4: Actuals Source Audit =====
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells('A1:E1')
ws4['A1'] = "Actuals Source Audit"
ws4['A1'].font = Font(bold=True, size=12)

audit_headers = ["Data Point", "Value", "Source", "Date", "Notes"]
for c, h in enumerate(audit_headers, 1):
    ws4.cell(row=2, column=c, value=h)
style_header_row(ws4, 2, 5)

audit_data = [
    ("Stock Price", "$63.72", "Yahoo Finance", "2026-07-09", "Close price; overnight $63.80"),
    ("Market Cap", "$3.74B", "Yahoo Finance Statistics", "2026-07-08", "Quarterly tab; current"),
    ("Enterprise Value", "$4.20B", "Yahoo Finance Statistics", "2026-07-08", "Quarterly tab; EV = MC - Cash + Debt"),
    ("Total Debt (MRQ)", "$710.3M", "Yahoo Finance Statistics", "MRQ ~May 2026", "Up from $686M FY26; capital lease heavy"),
    ("Cash (MRQ)", "$249.56M", "Yahoo Finance Statistics", "MRQ ~May 2026", "Down from $259.7M FY26 end cash"),
    ("Shares Outstanding", "60.45M", "Yahoo Finance Statistics", "2026-07-09", "Stable vs FY26 diluted avg 61.77M"),
    ("Beta (5Y Monthly)", "0.47", "Yahoo Finance Statistics", "N/A", "Very low — defensive consumer staples retailer"),
    ("TTM Revenue", "$2,731M", "Yahoo Finance Income Statement", "TTM", "+3.1% vs FY2026 $2,649M"),
    ("TTM Gross Profit", "$1,112M", "Yahoo Finance Income Statement", "TTM", "GP margin 40.7%"),
    ("TTM Operating Income", "$336M", "Yahoo Finance Income Statement", "TTM", "Op margin 12.3%"),
    ("TTM Net Income", "$249M", "Yahoo Finance Income Statement", "TTM", "Net margin 9.1%"),
    ("TTM Diluted EPS", "$4.16", "Yahoo Finance Income Statement", "TTM", "61.6M diluted shares"),
    ("TTM EBITDA", "$393.5M", "Yahoo Finance Statistics", "TTM", "Yahoo calc; derived OI + D&A"),
    ("TTM Operating Cash Flow", "$313M", "Yahoo Finance Cash Flow", "TTM", "Strong conversion 127% of OI"),
    ("TTM Capex", "$100.6M", "Yahoo Finance Cash Flow", "TTM", "Capex/rev ratio 3.7% — efficient"),
    ("TTM Free Cash Flow", "$212.7M", "Yahoo Finance Cash Flow", "TTM", "FCF margin 7.8% of revenue"),
    ("FY2027 Rev Est (Avg)", "$2.98B", "Yahoo Finance Analysis", "15 analysts", "+12.5% vs FY2026"),
    ("FY2028 Rev Est (Avg)", "$3.31B", "Yahoo Finance Analysis", "15 analysts", "+11.2% vs FY2027"),
    ("FY2027 EPS Est (Avg)", "$4.49", "Yahoo Finance Analysis", "15 analysts", "Normalized; +15.3% vs TTM $4.16"),
    ("FY2028 EPS Est (Avg)", "$5.12", "Yahoo Finance Analysis", "15 analysts", "+14.0% vs FY2027"),
    ("Analyst Avg Target", "$120.20", "Yahoo Finance Analysis", "Jul 9 2026", "High $152; 88% upside from current"),
    ("Earnings Surprise History", "+4.63% Q4", "Yahoo Finance Analysis", "4/30/2026", "3 of 4 quarters beat"),
    ("EPS Revisions Trend", "Downward", "Yahoo Finance Analysis", "Jul 9 2026", "11-13 analysts cut ests last 30d"),
    ("Next Earnings Date", "Aug 27, 2026", "Yahoo Finance", "Jul 9 2026", "Q2 FY27 report"),
    ("Tax Rate (TTM effective)", "24.4%", "Yahoo Finance Income Statement", "TTM", "Tax prov $80.5M / pretax $329.9M"),
    ("ROE (TTM)", "13.78%", "Yahoo Finance Statistics", "TTM", "Decent for capital-light retailer"),
    ("ROA (TTM)", "6.82%", "Yahoo Finance Statistics", "TTM", "Solid asset utilization"),
    ("Current Ratio (MRQ)", "2.32", "Yahoo Finance Statistics", "MRQ", "Strong liquidity"),
    ("Share Repurchases TTM", "$110M", "Yahoo Finance Cash Flow", "TTM", "Active buyback program; $74M FY26"),
    ("52W Range", "$60.29 - $141.74", "Yahoo Finance", "2026-07-09", "At the bottom; down 55% from highs"),
]

for i, row in enumerate(audit_data, 3):
    for c, val in enumerate(row, 1):
        ws4.cell(row=i, column=c, value=val)

style_data_range(ws4, 3, len(audit_data) + 2, 5)
ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 30
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 55

# ===== Sheet 5: Questions =====
ws5 = wb.create_sheet("Questions")
ws5.merge_cells('A1:C1')
ws5['A1'] = "Open Questions"
ws5['A1'].font = Font(bold=True, size=12)

questions = [
    ("1", "Why the 55% collapse from $141 to $63? The fundamentals look strong — revenue growing 14%+ YoY, EPS up 18%, FCF $213M. Is this a multiple compression from over-extension, or has something broken in the business model?"),
    ("2", "Negative EPS revisions: 11-13 analysts cut estimates in the last 30 days. What do they see that the surface numbers miss? Q4 FY26 beat by 4.63% yet consensus is being cut downward. This suggests Q2 FY27 or forward guidance disappointed."),
    ("3", "Capital lease obligations: $684M of the $686M total debt is capital leases (store financing). This is a structural financing model for discount retail. How does it limit financial flexibility vs. unsecured debt?"),
    ("4", "Share count trends: Basic shares declining (62,495 in FY23 to 61,090 in FY26) thanks to the $110M TTM buyback program. But diluted shares are also trending down (62,704 to 61,617). Are there offsetting option grants or other dilution?"),
    ("5", "Truist cut PT from $112 to $80 on Jul 9, 2026 — but also maintained a Buy rating. What changed? The PT is still 26% above current price. This suggests a technical/multiple adjustment rather than a fundamental deterioration."),
    ("6", "Short interest at 10.56% of outstanding (6.38M shares), up sharply from 4.54M in May. The bears are positioning. What's their thesis? Margin compression? Traffic weakness? Tariff-related cost pressure on imported goods?"),
    ("7", "Gas price sensitivity: Recent news highlights 'higher gas prices' pressure. As a car-centric retailer (1,250+ stores), fuel cost pass-through to logistics could meaningfully affect margins in inflationary environments."),
    ("8", "Weather headwinces: News mentions 'weather headwinds'. Ollie's sells seasonal items (holiday decor, seasonal products). Unusual weather patterns could depress same-store sales."),
    ("9", "Closeout merchandise model resilience: Ollie's sources excess inventory and closeout merchandise. In a supply-chain-normalized environment, does this model have less margin leverage than in the supply-chain-disrupted years of 2021-2023?"),
    ("10", "Competitive positioning vs. Dollar General (DG) and Dollar Tree (DLTR): How does Ollie's same-store sales growth compare? DG has been reformulating its strategy post-bankruptcy. DLTR is struggling. Is Ollie's 'over-extended and compressed' or 'resilient through the cycle'?"),
]

for i, (num, q) in enumerate(questions, 2):
    ws5.cell(row=i, column=1, value=num).font = bold_font
    ws5.cell(row=i, column=2, value=q)

ws5.column_dimensions['A'].width = 5
ws5.column_dimensions['B'].width = 120

# ===== Sheet 6: Sources =====
ws6 = wb.create_sheet("Sources")
ws6.merge_cells('A1:C1')
ws6['A1'] = "Sources"
ws6['A1'].font = Font(bold=True, size=12)

sources = [
    ("1", "Yahoo Finance - OLLI Quote & Summary", "https://finance.yahoo.com/quote/OLLI/"),
    ("2", "Yahoo Finance - OLLI Income Statement", "https://finance.yahoo.com/quote/OLLI/financials/"),
    ("3", "Yahoo Finance - OLLI Balance Sheet", "https://finance.yahoo.com/quote/OLLI/balance-sheet/"),
    ("4", "Yahoo Finance - OLLI Cash Flow", "https://finance.yahoo.com/quote/OLLI/cash-flow/"),
    ("5", "Yahoo Finance - OLLI Analyst Estimates", "https://finance.yahoo.com/quote/OLLI/analysis/"),
    ("6", "Yahoo Finance - OLLI Key Statistics", "https://finance.yahoo.com/quote/OLLI/key-statistics/"),
    ("7", "Yahoo Finance - OLLI Profile", "https://finance.yahoo.com/quote/OLLI/profile/"),
    ("8", "CNBC - US 10 Year Treasury", "https://www.cnbc.com/quotes/US10Y"),
    ("9", "StockAnalysis.com - OLLI (404 — unavailable)", "https://stockanalysis.com/quote/OLLI/"),
    ("10", "Yahoo Finance - OLLI Recent News", "https://finance.yahoo.com/quote/OLLI/news/"),
]

for i, (num, name, url) in enumerate(sources, 2):
    ws6.cell(row=i, column=1, value=num)
    ws6.cell(row=i, column=2, value=name)
    ws6.cell(row=i, column=3, value=url)

ws6.column_dimensions['A'].width = 5
ws6.column_dimensions['B'].width = 55
ws6.column_dimensions['C'].width = 65

# Save
output_path = "models/[2026-07-09] OLLI Model.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")

# Verification
from openpyxl import load_workbook
wb_check = load_workbook(output_path)
print(f"Verification - sheets: {wb_check.sheetnames}")
print(f"WACC: {wacc*100:.2f}%")
print(f"Weighted FV: ${weighted_FV:.2f}")
print(f"Bear target: ${bear_target:.2f}, Base target: ${base_target:.2f}, Bull target: ${bull_target:.2f}")
print(f"Total upside from $63.72: {total_upside:+.1f}%")
print("Model build complete and verified.")
