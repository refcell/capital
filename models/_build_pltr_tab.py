"""Add a PLTR (Palantir Technologies) tab to projections/Projections.xlsx.

Clones the last existing tab (NU) to preserve styling/formulas, then
overwrites the input cells with Palantir numbers.

Palantir (NASDAQ: PLTR) has a December fiscal year and reports in USD.  Like the
other December-FY tabs (HOOD, NFLX, NU), the "2026" column is a FY2026 ESTIMATE
(company guidance / consensus) and the model projects FY2027-FY2031.  Basis:
GAAP net income / diluted EPS -- Palantir is now solidly GAAP-profitable and its
headline valuation (TTM P/E ~144x, forward ~81x) and the published consensus
earnings line are struck on GAAP, so GAAP is the cleaner anchor here.  Note
adjusted EPS runs slightly below GAAP EPS (Q1'26: GAAP $0.34 vs adj $0.33)
because GAAP is helped by interest income on ~$8B cash and a low tax rate, while
adjusted adds back heavy stock-based comp (~$200M/qtr).  All $ in millions.

The thesis: hyper-growth operational-AI software (AIP) compounding off explosive
U.S. demand -- Q1'26 revenue +85% Y/Y (U.S. +104%, U.S. commercial +133%),
Rule of 40 = 145%, adjusted operating margin 60%, NDR 150% -- with FY2026
revenue guidance raised to ~$7.66B (+71%).  The whole debate is valuation, not
fundamentals: the multiple is extreme, so the model pairs a fast (but
decelerating) growth path and ~45% GAAP net margin with very high P/E multiples
(bull: sustained growth + premium multiple; bear: law-of-large-numbers
deceleration + multiple compression).

Sources (Q1'26 results 5/4/2026; FY2026 guidance; consensus; quotes Jun 2026):
  - FY2025 revenue $4,475M (+56%); GAAP net income $1,625M; GAAP EPS ~$0.66
  - Q1'26: revenue $1,633M (+85% Y/Y, +16% QoQ); GAAP net income $871M (53%
    margin); GAAP EPS $0.34; adj EPS $0.33; adj op margin 60%; adj FCF $925M;
    diluted shares (adj) ~2,571M; $8.0B cash; Rule of 40 = 145%; NDR 150%
  - FY2026 guidance: revenue $7.650-7.662B (+71%); adj op income $4.44-4.452B;
    adj FCF $4.2-4.4B; GAAP operating + net income positive every quarter
  - Consensus (GAAP net income): FY2026 rev $7,716M / NI $3,449M;
    FY2027 rev $11,197M / NI $4,829M; FY2028 rev $16,040M / NI $7,307M
  - Price $128.47 (6/18/2026, -27% YTD); market cap ~$308B; ~2,570M dil shares;
    TTM P/E ~144x, forward P/E ~81x; 52-wk $122.68-$207.52;
    32 analysts Buy, avg 12-mo PT ~$183 (Street targets ~$190-230)
  - Risks (bear): extreme valuation / multiple compression, growth
    deceleration (law of large numbers), single-market (U.S.) concentration,
    international/government headwinds (UK NHS, Germany, Switzerland, France),
    SBC dilution, government contract termination clauses
"""

from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "projections" / "Projections.xlsx"

# Shared inputs (USD millions; shares = diluted shares in millions)
SHARES = 2570         # diluted shares (~2,571M Q1'26); slow SBC-driven growth
REV_2026 = 7716       # FY2026E revenue (consensus; guidance ~$7.66B, +71%)
NI_2026 = 3449        # FY2026E GAAP net income (~44.7% margin, EPS ~$1.34)
NI_G_2026 = 1.12      # GAAP net income growth FY2025->FY2026E ($1,625M -> $3,449M)

# Analyst GAAP net income estimates FY2026-FY2029 (USD M)
# (FY26 $3.45B / FY27 $4.83B / FY28 $7.31B consensus; FY29 extrapolated.)
ANALYST = {"C": 3449, "D": 4829, "E": 7307, "F": 9500}

# Per-case assumptions.  Each template block is offset by 24 rows:
#   Bull base row 4, Base 28, Bear 52.  Within a block, relative rows:
#     +2 revenue, +3 revenue growth, +5 net income, +6 NI growth,
#     +8 analyst est, +10 NI margins, +14 PE low, +15 PE high, H(+0) shares.
# NOTE: revenue growth applies with a one-column lag (D6 = C6*(1+C7)), so the
# column-C growth rate drives FY2027 revenue, column-D drives FY2028, etc.
# Revenue decelerates from the ~70% guided FY26 pace but stays very high; GAAP
# net margin holds in the mid-40s% on software operating leverage.  PE multiples
# are exceptionally high, reflecting how PLTR actually trades (bull: sustained
# hyper-growth + premium multiple; bear: deceleration + multiple compression).
CASES = {
    "bull": {
        "base": 4,
        # AIP land-and-expand sustains ~50%/40%+ growth longer
        "rev_g": {"C": 0.52, "D": 0.48, "E": 0.40, "F": 0.33, "G": 0.27, "H": 0.23},
        "margins": {"D": 0.45, "E": 0.47, "F": 0.48, "G": 0.49, "H": 0.50},
        "pe_low": {c: 100 for c in "CDEFGH"},
        "pe_high": {c: 150 for c in "CDEFGH"},
    },
    "base": {
        "base": 28,
        # ties to consensus: FY27 ~$11.2B, FY28 ~$16.0B, then decelerates
        "rev_g": {"C": 0.45, "D": 0.43, "E": 0.35, "F": 0.28, "G": 0.23, "H": 0.20},
        "margins": {"D": 0.43, "E": 0.45, "F": 0.46, "G": 0.47, "H": 0.47},
        "pe_low": {c: 70 for c in "CDEFGH"},
        "pe_high": {c: 110 for c in "CDEFGH"},
    },
    "bear": {
        "base": 52,
        # law of large numbers bites; multiple compresses toward software norms
        "rev_g": {"C": 0.38, "D": 0.32, "E": 0.26, "F": 0.21, "G": 0.18, "H": 0.15},
        "margins": {"D": 0.41, "E": 0.42, "F": 0.42, "G": 0.43, "H": 0.43},
        "pe_low": {c: 40 for c in "CDEFGH"},
        "pe_high": {c: 65 for c in "CDEFGH"},
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    src = wb[wb.sheetnames[-1]]          # clone the last tab (NU)
    ws = wb.copy_worksheet(src)
    ws.title = "PLTR"

    # Header
    ws["B2"] = "PLTR"
    ws["C2"] = datetime(2026, 6, 21)

    for cfg in CASES.values():
        b = cfg["base"]
        ws[f"H{b}"] = SHARES                      # diluted share count

        # Revenue base (FY2026E) + growth row
        ws[f"C{b + 2}"] = REV_2026
        for col, g in cfg["rev_g"].items():
            ws[f"{col}{b + 3}"] = g

        # GAAP net income base (FY2026E) + base-year growth
        ws[f"C{b + 5}"] = NI_2026
        ws[f"C{b + 6}"] = NI_G_2026

        # Analyst estimates (FY2026-FY2029)
        for col, v in ANALYST.items():
            ws[f"{col}{b + 8}"] = v

        # NI margins (D..H; C stays as formula =NI/Rev)
        for col, m in cfg["margins"].items():
            ws[f"{col}{b + 10}"] = m

        # PE low / high
        for col, v in cfg["pe_low"].items():
            ws[f"{col}{b + 14}"] = v
        for col, v in cfg["pe_high"].items():
            ws[f"{col}{b + 15}"] = v

    wb.save(XLSX)
    print(f"Added PLTR tab. Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
