"""
TU (TELUS Corporation) - DCF / FCF Multiple Valuation Model
Built: 2026-07-02
Data as of July 2, 2026 (NYSE close)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
import os

wb = Workbook()

# ── Styles ──
header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)
section_font = Font(bold=True, size=11, color="FFFFFF")
section_fill = PatternFill("solid", fgColor="2F5496")
neg_fill = PatternFill("solid", fgColor="FFC7CE")
bear_fill = PatternFill("solid", fgColor="FFE6E6")
base_fill = PatternFill("solid", fgColor="E6F0FF")
bull_fill = PatternFill("solid", fgColor="E6F9E6")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

def style_data_block(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            ws.cell(r, c).border = thin_border

def write_title_block(ws, company, ticker, exchange, date_str, price, shares_mm, mc_b, ev_b, primary_lens, stance):
    ws.merge_cells('A1:G1')
    ws['A1'] = company
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")
    meta = [
        ("Ticker", f"{exchange}:{ticker}"),
        ("Date", date_str),
        ("Price", f"${price:.2f}"),
        ("Shares Outstanding", f"{shares_mm:.2f}M"),
        ("Market Cap", f"${mc_b:.2f}B"),
        ("Enterprise Value", f"${ev_b:.2f}B"),
        ("Primary Valuation Lens", primary_lens),
        ("Stance", stance),
    ]
    for i, (k, v) in enumerate(meta, 2):
        ws.cell(i, 1, k).font = header_font
        ws.cell(i, 2, v)

# ═══════════════════════════════════════════
#  Key data anchors (all in USD unless noted)
# ═══════════════════════════════════════════
PRICE = 10.19
SHARES_MM = 1_560  # ~1.56B
MC_B = 15.91  # intraday market cap from YF
EV_B = 37.45  # enterprise value from YF

# Yahoo Finance stats page provides EV and MC directly.
# EV - MC = net debt proxy = $21.54B
NET_DEBT_B = EV_B - MC_B  # 21.54B

# Financials in CAD (Yahoo) - converting at ~0.70
CAD_RATE = 0.70

# Revenue history (CAD in $)
REV_2025_CAD = 20_346_000 / 1000  # $M CAD
REV_2024_CAD = 20_139_000 / 1000
REV_2023_CAD = 20_005_000 / 1000
REV_2022_CAD = 18_292_000 / 1000

# FCF
FCF_TTM_CAD_M = 1_903  # $M CAD
FCF_2025_CAD_M = 2_351
FCF_2024_CAD_M = 1_460
FCF_2023_CAD_M = 1_288
FCF_2022_CAD_M = 1_164

FCF_TTM_USD_M = FCF_TTM_CAD_M * CAD_RATE  # ~1332
FCF_2025_USD_M = FCF_2025_CAD_M * CAD_RATE

# EBITDA
EBITDA_TTM_CAD_M = 6_618
EBITDA_TTM_USD_M = EBITDA_TTM_CAD_M * CAD_RATE

# Revenue USD (approx - Yahoo shows $20.32B but that's CAD labeled as USD)
REV_TTM_USD_M = 20_317 * CAD_RATE  # ~14,222

# Operating income
OP_INC_TTM_CAD_M = 2_504
OP_INC_TTM_USD_M = OP_INC_TTM_CAD_M * CAD_RATE

# Net income
NI_TTM_CAD_M = 928
NI_TTM_USD_M = NI_TTM_CAD_M * CAD_RATE  # ~650

# Analyst estimates (USD)
FWD_EPS_Y1 = 0.68   # FY2026
FWD_EPS_Y2 = 0.72   # FY2027
FWD_REV_Y1_B = 20.77  # Yahoo shows this as USD but it's CAD
FWD_REV_Y1_USD_B = FWD_REV_Y1_B * CAD_RATE

# ── SHEET 1: Valuation ──
ws1 = wb.active
ws1.title = "Valuation"

write_title_block(ws1, "TELUS Corporation", "TU", "NYSE", "2026-07-02",
                  PRICE, SHARES_MM / 1000, MC_B, EV_B,
                  "Forward P/E (Analyst Consensus)", "Watch")

# Valuation metrics table
ws1.cell(11, 1, "Valuation Metric").font = header_font
ws1.cell(11, 2, "Value").font = header_font
ws1.cell(11, 3, "Commentary").font = header_font

metrics = [
    ("P/E (Trailing)", f"{24.89:.2f}", "Elevated trailing P/E due to depressed TTM earnings post-capex cycle; 52-week decline from 16.74 to 10.19 compressed the multiple denominator"),
    ("Forward P/E (FY2026)", f"{16.05:.2f}", "Primary lens. Based on $0.68 consensus EPS. Reasonable for a stable Canadian telecom; trades below historical avg ~18-20x"),
    ("Forward P/E (FY2027)", f"{EV_B / (FWD_EPS_Y2 * SHARES_MM / 1000):.2f}", "Using FY27 EPS $0.72. Implied Fwd P/E on next year estimates — still reasonable"),
    ("P/S (TTM)", f"{1.14:.2f}", "Low for a telecom; reflects depressed stock price vs stable ~$20B revenue base"),
    ("P/B (mrq)", f"{1.50:.2f}", "Reasonable; below historical range of 1.8-2.5x for Canadian wireless"),
    ("EV/EBITDA (TTM)", f"{8.03:.2f}", "Compressed vs historical 9-10x; capex cycle concerns dragged EV multiple down"),
    ("EV/Sales (TTM)", f"{2.62:.2f}", "Below peer avg; BCE at ~3.0x+; reflects stock underperformance"),
    ("Dividend Yield (Fwd)", f"{11.41:.1f}%", "Extremely high — 278% payout ratio suggests dividend is at risk if earnings do not recover"),
]

for i, (m, v, c) in enumerate(metrics, 12):
    ws1.cell(i, 1, m).font = header_font
    ws1.cell(i, 2, v)
    ws1.cell(i, 2).alignment = Alignment(horizontal="center")
    ws1.cell(i, 3, c).alignment = Alignment(wrap_text=True)

style_data_block(ws1, 11, 19, 3)
ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 70

# ── SHEET 2: WACC ──
ws2 = wb.create_sheet("WACC")
ws2.merge_cells('A1:B1')
ws2['A1'] = "WACC Calculation (CAPM)"
ws2['A1'].font = title_font

rfr = 4.485  # 10Y Treasury from CNBC
erp = 5.0
beta = 0.73
cost_equity = rfr + beta * erp  # 4.485 + 0.73*5 = 8.135%

total_debt_b = 31_459 * CAD_RATE / 1000  # ~$22.02B
# Use EV - MC as the cleanest net debt proxy per skill guidance
# Total debt USD approx
total_equity_b = MC_B  # ~$15.91B

wdt = (total_equity_b) / (total_equity_b + total_debt_b)        # ~0.42
wdd = (total_debt_b) / (total_equity_b + total_debt_b)         # ~0.58

# Telus has ~$22B debt. Yahoo EV of $37.45B implies net debt of ~$21.54B.
# This is high for a telecom because of spectrum/capex borrowing.
# For scenario analysis, we need realistic FCF assumptions that can support this leverage.
# Use EBITDA multiple approach instead - telecoms are typically valued on EV/EBITDA.

# Cost of debt - telco debt yield ~4.5-5% for investment grade
cost_debt_pct = 4.80
tax_rate = 27.0  # Canadian corporate effective rate

wacc = wdt * cost_equity + wdd * cost_debt_pct * (1 - tax_rate / 100)
# 0.42 * 8.135 + 0.58 * 4.80 * 0.73 = 3.42 + 2.03 = 5.45%

wacc_data = [
    ("Component", "Value"),
    ("Risk-Free Rate (10Y US Treasury)", f"{rfr:.3f}%"),
    ("Equity Risk Premium", f"{erp:.1f}%"),
    ("Beta (5Y Monthly)", f"{beta:.2f}"),
    ("Cost of Equity (Rf + Beta*ERP)", f"{cost_equity:.2f}%"),
    ("", ""),
    ("Market Cap (Equity)", f"${total_equity_b:.2f}B"),
    ("Total Debt", f"${total_debt_b:.2f}B"),
    ("Equity Weight", f"{wdt:.4f}"),
    ("Debt Weight", f"{wdd:.4f}"),
    ("Cost of Debt (est)", f"{cost_debt_pct:.2f}%"),
    ("Tax Rate (Canadian Corp Eff.)", f"{tax_rate:.1f}%"),
    ("", ""),
    ("WACC", f"{wacc:.2f}%"),
]

for i, (k, v) in enumerate(wacc_data, 2):
    ws2.cell(i, 1, k).font = header_font if k else Font()
    ws2.cell(i, 2, v)
    ws2.cell(i, 2).alignment = Alignment(horizontal="center")
    if k == "WACC":
        ws2.cell(i, 1).font = Font(bold=True, size=12)
        ws2.cell(i, 2).font = Font(bold=True, size=12)
        ws2.cell(i, 1).fill = base_fill
        ws2.cell(i, 2).fill = base_fill

style_data_block(ws2, 2, 14, 2)
ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 20

# ── SHEET 3: Scenarios ──
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells('A1:L1')
ws3['A1'] = "Scenario Analysis - Bear / Base / Bull"
ws3['A1'].font = title_font

shares_mm = SHARES_MM / 1000  # 1560 M

# ── Scenario framework: Forward P/E on analyst consensus ──
# Telus is a mature Canadian telecom with ~$22B debt and ~$1.3B USD FCF.
# The FCF multiple approach produces negative equity values at reasonable multiples
# because net debt ($21.5B) overwhelms the FCF-generated implied EV.
# Primary lens: Forward P/E on FY2027 analyst consensus ($0.72 EPS).
# Cross-check: EV/EBITDA multiples.

# Forward P/E approach
bear_eps = 0.60   # Conservative: below FY26 consensus
base_eps = 0.72   # FY27 consensus
bull_eps = 0.82   # Above FY27, assumes cadence recovery

bear_pe = 12.0   # Compressed - earnings disappointment
base_pe = 15.0   # In-line with historical for Canadian wireless
bull_pe = 18.0   # Expansion - capex cycle ends, earnings normalize

bear_target = bear_eps * bear_pe
base_target = base_eps * base_pe
bull_target = bull_eps * bull_pe

bear_upside = (bear_target - PRICE) / PRICE * 100
base_upside = (base_target - PRICE) / PRICE * 100
bull_upside = (bull_target - PRICE) / PRICE * 100

# EBITDA cross-check (USD, $M)
EBITDA_TTM_USD_M = EBITDA_TTM_CAD_M * CAD_RATE  # ~4633
shares_mm = SHARES_MM / 1000  # 1560

bear_ev_ebitda = 6.5
base_ev_ebitda = 8.0
bull_ev_ebitda = 9.5

bear_implied_ev_mm = EBITDA_TTM_USD_M * bear_ev_ebitda
base_implied_ev_mm = EBITDA_TTM_USD_M * base_ev_ebitda
bull_implied_ev_mm = EBITDA_TTM_USD_M * bull_ev_ebitda

# Net debt in $M
net_debt_mm = (EV_B - MC_B) * 1000  # 21540 $M

bear_eq_ev = (bear_implied_ev_mm - net_debt_mm) / shares_mm
base_eq_ev = (base_implied_ev_mm - net_debt_mm) / shares_mm
bull_eq_ev = (bull_implied_ev_mm - net_debt_mm) / shares_mm

# Weights
bear_wt = 0.25
base_wt = 0.50
bull_wt = 0.25

weighted_value = bear_wt * bear_target + base_wt * base_target + bull_wt * bull_target
weighted_upside = (weighted_value - PRICE) / PRICE * 100

headers = [
    "Metric", "Bear", "Base", "Bull"
]
scenario_data = [
    headers,
    ("Framework", "Forward P/E", "Forward P/E", "Forward P/E"),
    ("Forward EPS", f"${bear_eps:.2f}", f"${base_eps:.2f}", f"${bull_eps:.2f}"),
    ("Forward P/E Multiple", f"{bear_pe:.1f}x", f"{base_pe:.1f}x", f"{bull_pe:.1f}x"),
    ("Target Price ($ USD)", f"${bear_target:.2f}", f"${base_target:.2f}", f"${bull_target:.2f}"),
    ("Upside from Current", f"{bear_upside:.1f}%", f"{base_upside:.1f}%", f"{bull_upside:.1f}%"),
    ("", "", "", ""),
    ("EV/EBITDA Cross-Check", f"{bear_ev_ebitda:.1f}x", f"{base_ev_ebitda:.1f}x", f"{bull_ev_ebitda:.1f}x"),
    ("TTM EBITDA ($M USD)", f"${EBITDA_TTM_USD_M:,.0f}", f"${EBITDA_TTM_USD_M:,.0f}", f"${EBITDA_TTM_USD_M:,.0f}"),
    ("Implied EV ($M USD)", f"${bear_implied_ev_mm:,.0f}", f"${base_implied_ev_mm:,.0f}", f"${bull_implied_ev_mm:,.0f}"),
    ("Less: Net Debt ($M USD)", f"${net_debt_mm:,.0f}", f"${net_debt_mm:,.0f}", f"${net_debt_mm:,.0f}"),
    ("Implied Price/Share ($)", f"${bear_eq_ev:.2f}", f"${base_eq_ev:.2f}", f"${bull_eq_ev:.2f}"),
    ("", "", "", ""),
    ("Scenario Weight", f"{bear_wt:.0%}", f"{base_wt:.0%}", f"{bull_wt:.0%}"),
    ("Weighted Value/Share (P/E)", f"${bear_wt*bear_target:.2f}", f"${base_wt*base_target:.2f}", f"${bull_wt*bull_target:.2f}"),
]

for i, row in enumerate(scenario_data, 2):
    for j, val in enumerate(row, 1):
        ws3.cell(i, j, val)
        ws3.cell(i, j).border = thin_border
    if i == 2:
        style_header_row(ws3, i, 4)
    elif i in (5, 14):
        for j in range(1, 5):
            ws3.cell(i, j).font = Font(bold=True)

# Color coding
for r in range(3, 15):
    ws3.cell(r, 2).fill = bear_fill
    ws3.cell(r, 3).fill = base_fill
    ws3.cell(r, 4).fill = bull_fill

# Summary
ws3.cell(16, 1, "Total Probability-Weighted FV ($)").font = Font(bold=True, size=12)
ws3.cell(16, 2, f"{weighted_value:.2f}").font = Font(bold=True, size=12)
ws3.merge_cells('B16:D16')
ws3.cell(17, 1, "Upside from Current Price").font = Font(bold=True)
ws3.cell(17, 2, f"{weighted_upside:.1f}%").font = Font(bold=True)
ws3.merge_cells('B17:D17')

# Note about FCF insufficiency
ws3.cell(19, 1, "NOTE: FCF multiple approach produces negative equity values due to").font = Font(italic=True, size=9)
ws3.cell(20, 1, "high leverage ($22B debt vs $1.3B FCF). Primary lens is Forward P/E, not FCF.").font = Font(italic=True, size=9)
ws3.cell(21, 1, "EV/EBITDA cross-check shown above for reference.").font = Font(italic=True, size=9)

# WACC reference
ws3.cell(18, 1, "WACC (for DCF cross-check)").font = header_font
ws3.cell(18, 2, f"{wacc:.2f}%").alignment = Alignment(horizontal="center")

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 15

# Print WACC and FV for verification
print(f"WACC: {wacc:.2f}%")
print(f"Bear target: ${bear_target:.2f}")
print(f"Base target: ${base_target:.2f}")
print(f"Bull target: ${bull_target:.2f}")
print(f"Weighted FV: ${weighted_value:.2f}")
print(f"Upside: {weighted_upside:.1f}%")

# ── SHEET 4: Actuals Source Audit ──
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells('A1:E1')
ws4['A1'] = "Actuals Source Audit"
ws4['A1'].font = title_font

audit_headers = ["Data Point", "Value", "Source URL", "Date Accessed", "Notes"]
audit_data = [
    audit_headers,
    ("Stock Price (close)", "$10.19 USD", "https://finance.yahoo.com/quote/TU/", "2026-07-02", "NYSE delayed quote, at close 4:00 PM EDT"),
    ("After Hours Price", "$10.20", "https://finance.yahoo.com/quote/TU/", "2026-07-02", "After hours 7:56 PM EDT"),
    ("Market Cap (intraday)", "$15.91B", "https://finance.yahoo.com/quote/TU/", "2026-07-02", "Yahoo Finance summary page"),
    ("Enterprise Value", "$37.45B", "https://finance.yahoo.com/quote/TU/key-statistics/", "2026-07-02", "Yahoo Finance key statistics"),
    ("Shares Outstanding", "1.56B", "https://finance.yahoo.com/quote/TU/key-statistics/", "2026-07-02", "Shares Outstanding from key statistics"),
    ("Beta (5Y Monthly)", "0.73", "https://finance.yahoo.com/quote/TU/key-statistics/", "2026-07-02", "Yahoo Finance key statistics"),
    ("Revenue TTM", "$20,317M CAD", "https://finance.yahoo.com/quote/TU/financials/", "2026-07-02", "Yahoo Finance income statement, all numbers in thousands CAD"),
    ("Operating Income TTM", "$2,504M CAD", "https://finance.yahoo.com/quote/TU/financials/", "2026-07-02", "Yahoo Finance income statement"),
    ("Net Income TTM", "$928M CAD", "https://finance.yahoo.com/quote/TU/financials/", "2026-07-02", "Yahoo Finance income statement"),
    ("EBITDA TTM", "$6,618M CAD", "https://finance.yahoo.com/quote/TU/financials/", "2026-07-02", "Yahoo Finance income statement"),
    ("FCF TTM", "$1,903M CAD", "https://finance.yahoo.com/quote/TU/cash-flow/", "2026-07-02", "OCF $4,839M - CapEx $2,936M CAD"),
    ("Total Debt (FY2025)", "$31,459M CAD", "https://finance.yahoo.com/quote/TU/balance-sheet/", "2026-07-02", "Yahoo Finance balance sheet"),
    ("Total Equity (FY2025)", "$16,579M CAD", "https://finance.yahoo.com/quote/TU/balance-sheet/", "2026-07-02", "Yahoo Finance balance sheet"),
    ("Rev Estimate FY2026", "$20.77B (CAD)", "https://finance.yahoo.com/quote/TU/analysis/", "2026-07-02", "15 analysts avg, CAD currency"),
    ("Rev Estimate FY2027", "$21.25B (CAD)", "https://finance.yahoo.com/quote/TU/analysis/", "2026-07-02", "14 analysts avg, CAD currency"),
    ("EPS FY2026", "$0.68", "https://finance.yahoo.com/quote/TU/analysis/", "2026-07-02", "16 analysts avg, USD"),
    ("EPS FY2027", "$0.72", "https://finance.yahoo.com/quote/TU/analysis/", "2026-07-02", "15 analysts avg, USD"),
    ("P/E Trailing", "24.89", "https://finance.yahoo.com/quote/TU/key-statistics/", "2026-07-02", "Yahoo Finance key statistics"),
    ("Forward P/E", "16.05", "https://finance.yahoo.com/quote/TU/key-statistics/", "2026-07-02", "Yahoo Finance key statistics"),
    ("EV/EBITDA", "8.03", "https://finance.yahoo.com/quote/TU/key-statistics/", "2026-07-02", "Yahoo Finance key statistics"),
    ("P/B", "1.50", "https://finance.yahoo.com/quote/TU/key-statistics/", "2026-07-02", "Yahoo Finance key statistics"),
    ("Dividend Rate (Fwd)", "$1.20", "https://finance.yahoo.com/quote/TU/key-statistics/", "2026-07-02", "Yahoo Finance key statistics"),
    ("Dividend Yield (Fwd)", "11.41%", "https://finance.yahoo.com/quote/TU/key-statistics/", "2026-07-02", "Yahoo Finance key statistics"),
    ("Payout Ratio", "278.23%", "https://finance.yahoo.com/quote/TU/key-statistics/", "2026-07-02", "Yahoo Finance key statistics - above 100%"),
    ("StockExchange", "NYSE", "https://finance.yahoo.com/quote/TU/", "2026-07-02", "NYSE listing (US ADR)"),
    ("StockAnalysis", "N/A", "https://stockanalysis.com/quote/NASDAQ/TU/", "2026-07-02", "404 - page not found; Canadian/ADR ticker not covered"),
    ("Treasury 10Y Yield", "4.485%", "https://www.cnbc.com/quotes/US10Y", "2026-07-02", "CNBC quote page, 2:30 PM EDT"),
]

for i, row in enumerate(audit_data, 2):
    for j, val in enumerate(row, 1):
        ws4.cell(i, j, val)
        ws4.cell(i, j).border = thin_border
        ws4.cell(i, j).alignment = Alignment(wrap_text=True)

style_header_row(ws4, 2, 5)
ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 55
ws4.column_dimensions['D'].width = 15
ws4.column_dimensions['E'].width = 50

# ── SHEET 5: Questions ──
ws5 = wb.create_sheet("Questions")
ws5.merge_cells('A1:B1')
ws5['A1'] = "Open Questions"
ws5['A1'].font = title_font

questions = [
    ("Q1", "Dividend Sustainability at 278% Payout: Forward dividend yield of 11.41% with 278% payout ratio on EPS. Is the dividend funded from FCF, and if so, at what coverage ratio? The dividend is ~$1.20/yr × 1.56B shares = $1.87B. TTM FCF ~$1.33B USD after conversion. Dividend appears to EXCEED FCF — how is this sustainable?"),
    ("Q2", "Cadbury / CAD Currency Risk: All financial statements are in CAD. US investors face significant currency exposure. CAD/USD has been ~0.70 — if CAD strengthens to 0.75, effectively the company's USD earnings increase proportionally. Inverse applies."),
    ("Q3", "Debt Growth Trajectory: Total debt grew from $25,141M (FY2022) to $31,459M (FY2025) CAD — a 25% increase in 3 years while net tangible assets turned more negative. The capex cycle (spectrum auctions, 5G buildout) is funded by incremental debt. When does this reverse?"),
    ("Q4", "Capex Cycle Timing: FY2025 CapEx was lower ($2,515M) vs FY2024 ($3,387M). Is the spectrum/buildout peak behind them? FY2025 FCF rebounded to $2,351M — is this a recovery signal or temporary?"),
    ("Q5", "Competition from Rogers-Superior deal: Rogers has been acquiring competitors. Any regulatory or competitive threats to Telus's market share in Quebec or British Columbia?"),
    ("Q6", "Negative Net Tangible Assets: NTAs of -$15B (FY2025) is structural for telecoms (high intangible assets from spectrum licenses, goodwill). Not immediately concerning, but limits balance sheet flexibility."),
    ("Q7", "Stock Price Context: Down 37.5% from 52-week high ($16.74) to current ($10.19). This is the 52-week LOW. What drove the decline: macro, sector rotation, company-specific factors?"),
    ("Q8", "Short Interest: 69.05M shares short (6/15/2026) with 13.1 short ratio — extremely elevated short interest. Bears are making a strong bet; is this overdone?"),
    ("Q9", "Earnings Quality: FY2022 showed $1.615B net income on $18.3B revenue but FY2025 shows $1.113B on $20.3B. What drove the margin compression? Tax provisions? Interest expense growth?"),
    ("Q10", "Next Earnings Date: When is the next earnings report? Current Q2 FY2026 (ending Jun 2026) results will be the first catalyst to check."),
]

for i, (q, t) in enumerate(questions, 2):
    ws5.cell(i, 1, q).font = header_font
    ws5.cell(i, 2, t).alignment = Alignment(wrap_text=True)

ws5.column_dimensions['A'].width = 8
ws5.column_dimensions['B'].width = 100

# ── SHEET 6: Sources ──
ws6 = wb.create_sheet("Sources")
ws6.merge_cells('A1:B1')
ws6['A1'] = "Sources"
ws6['A1'].font = title_font

sources = [
    ("1", "Yahoo Finance - Summary", "https://finance.yahoo.com/quote/TU/"),
    ("2", "Yahoo Finance - Income Statement", "https://finance.yahoo.com/quote/TU/financials/"),
    ("3", "Yahoo Finance - Balance Sheet", "https://finance.yahoo.com/quote/TU/balance-sheet/"),
    ("4", "Yahoo Finance - Cash Flow", "https://finance.yahoo.com/quote/TU/cash-flow/"),
    ("5", "Yahoo Finance - Key Statistics", "https://finance.yahoo.com/quote/TU/key-statistics/"),
    ("6", "Yahoo Finance - Analyst Estimates", "https://finance.yahoo.com/quote/TU/analysis/"),
    ("7", "StockAnalysis (404 - not available)", "https://stockanalysis.com/quote/NASDAQ/TU/"),
    ("8", "CNBC - 10-Year Treasury Yield", "https://www.cnbc.com/quotes/US10Y"),
]

for i, (n, name, url) in enumerate(sources, 2):
    ws6.cell(i, 1, n)
    ws6.cell(i, 2, name)
    ws6.cell(i, 3, url)

ws6.column_dimensions['A'].width = 5
ws6.column_dimensions['B'].width = 40
ws6.column_dimensions['C'].width = 55

# ── Save ──
out = "/home/refcell/dev/capital/models/[2026-07-02] TELUS Model.xlsx"
wb.save(out)
print(f"\nSaved to {out}")
print("Done.")
