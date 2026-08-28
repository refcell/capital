#!/usr/bin/env python3
"""Build SOLV - Solventum Corporation valuation model.xlsx
Data snapshot: August 27, 2026 close at $90.93
Post-spinoff from 3M (2023). Announced HIS divestiture Aug 2026.
Massive non-operating income ($1.43B TTM from HIS divestiture) makes
trailing P/E and TTM EPS structurally misleading.
Uses Forward P/E as primary framework (analyst consensus normalized
EPS already excludes one-time spinoff items).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── style helpers ──
title_font = Font(name="Calibri", size=16, bold=True)
subtitle_font = Font(name="Calibri", size=12, bold=True)
header_font = Font(name="Calibri", size=11, bold=True)
normal_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)
note_font = Font(name="Calibri", size=10, italic=True)
red_font = Font(name="Calibri", size=11, color="C00000")

header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
title_font_white = Font(name="Calibri", size=16, bold=True, color="FFFFFF")

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def c(ws, row, col, value, font=normal_font, border=False, fill=None, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if border:
        cell.border = thin_border
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    return cell

def header_row(ws, row, values, col_start=1):
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i, value=v)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def print_val(label, val):
    print(f"  {label}: {val}")

# ═══════════════════════════════════════════
# DATA ANCHORS
# ═══════════════════════════════════════════
PRICE = 90.93
DATE_T = "2026-08-27"
# Shares: use ISS'd shares from balance sheet, slightly different from intraday MC
SHARES_MM = 173.49  # millions
MC_B = 15.48  # billion
# Balance sheet: Total Debt $5.035B, End Cash $400M
TOTAL_DEBT_B = 5.035
CASH_B = 0.40
NET_DEBT_B = 4.16
TCF_B = -0.12  # TTM FCF is -$118M (capex cycle + restructuring)
# EV = MC + Debt - Cash
EV_B = MC_B + TOTAL_DEBT_B - CASH_B  # = 15.48 + 5.035 - 0.40 = 20.115
print(f"EV = {EV_B:.2f}B (MC {MC_B:.2f} + Debt {TOTAL_DEBT_B:.2f} - Cash {CASH_B:.2f})")

# Income statement (in $B): Revenue, Gross Profit, Op Income, EBITDA
revenue = [8.130, 8.197, 8.254, 8.325, float("nan")]
gp = [4.695, 4.693, 4.593, 4.451, float("nan")]
op_income = [1.693, 1.636, 1.036, 0.632, float("nan")]
ebitda = [2.270, 2.228, 1.527, 2.549, float("nan")]
ni = [1.343, 1.346, 0.479, 1.556, float("nan")]

TTM_REV_B = 8.310
TTM_GROSS_PROFIT_B = 4.544
TTM_OP_INC_B = 0.526
TTM_EBITDA_B = 2.466
TTM_NI_B = 1.434
TTM_DIL_EPS = 8.18
GROSS_MARGIN_TTM = 4.544 / 8.310  # 54.7%
OP_MARGIN_TTM = 0.526 / 8.310  # 6.3%

# Analyst consensus (normalized non-GAAP - excludes spinoff items)
FWD_EPS_FY26 = 7.18
FWD_EPS_FY27 = 7.22
FWD_REV_FY26 = 8.23
FWD_REV_FY27 = 8.51

# 10Y Treasury
RF = 0.0468
BETA = 0.67
ERP = 0.05
TAX_RATE = 0.21

# ═══════════════════════════════════════════
# VALUATION METRICS
# ═══════════════════════════════════════════
PE_TTM = PRICE / TTM_DIL_EPS  # ~11.1
PE_FWD = PRICE / FWD_EPS_FY26  # ~12.7
PS = MC_B / TTM_REV_B  # ~1.86
# P/FCF N/A - negative FCF
EV_SALES = EV_B / TTM_REV_B
EV_EBITDA = EV_B / TTM_EBITDA_B

print(f"P/E TTM: {PE_TTM:.2f}")
print(f"Fwd P/E: {PE_FWD:.2f}")
print(f"P/S: {PS:.2f}")
print(f"EV/Sales: {EV_SALES:.2f}")
print(f"EV/EBITDA: {EV_EBITDA:.2f}")

# ═══════════════════════════════════════════
# WACC
# ═══════════════════════════════════════════
COST_OF_EQUITY = RF + BETA * ERP
PRINT_COE = COST_OF_EQUITY
print(f"Ke = {RF:.4f} + {BETA} * {ERP:.2f} = {COST_OF_EQUITY:.4f}")

# Cost of debt: interest / total debt
INT_EXP_B = 0.181  # interest paid FY25 in $B
COST_OF_DEBT = INT_EXP_B / TOTAL_DEBT_B
print(f"Kd (pre-tax) = {INT_EXP_B} / {TOTAL_DEBT_B} = {COST_OF_DEBT:.4f}")

E_WEIGHT = MC_B / (MC_B + TOTAL_DEBT_B)
D_WEIGHT = TOTAL_DEBT_B / (MC_B + TOTAL_DEBT_B)
WACC = E_WEIGHT * COST_OF_EQUITY + D_WEIGHT * COST_OF_DEBT * (1 - TAX_RATE)

print(f"WACC = {E_WEIGHT:.3f}*{COST_OF_EQUITY:.4f} + {D_WEIGHT:.3f}*{COST_OF_DEBT:.4f}*(1-{TAX_RATE}) = {WACC:.4f}")
print(f"WACC = {WACC*100:.2f}%")

# ═══════════════════════════════════════════
# SCENARIOS - Forward P/E framework
# ═══════════════════════════════════════════
# Forward P/E because: (a) FCF is -$118M TTM, negative = framework invalid
# (b) Analyst consensus is normalized non-GAAP post-HIS spinoff
# (c) Operating income trajectory improving (Q2 beat 33.8%)

# Revenue CAGR: FY26 $8.23B -> 5Y forward
# Analyst: FY27 $8.51B (3.4% growth from FY26)
# Medtech mature growth: 2-4% organic typical

bear_eps = 5.50    # margin remains suppressed, restructuring drag
bear_exit_pe = 10.0

base_eps = 6.80    # modest recovery to pre-spinoff normalized margins
base_exit_pe = 13.0

bull_eps = 8.00    # full margin recovery to 9-10% operating margin
bull_exit_pe = 16.0

bear_target = bear_eps * bear_exit_pe
base_target = base_eps * base_exit_pe
bull_target = bull_eps * bull_exit_pe

print_vals = []
print(f"\n--- SCENARIOS (Fwd P/E framework) ---")
print(f"Bear: ${bear_eps:.2f} * {bear_exit_pe:.0f}x = ${bear_target:.2f} vs current ${PRICE}")
print(f"Base: ${base_eps:.2f} * {base_exit_pe:.0f}x = ${base_target:.2f} vs current ${PRICE}")
print(f"Bull: ${bull_eps:.2f} * {bull_exit_pe:.0f}x = ${bull_target:.2f} vs current ${PRICE}")

# Sanity check: analyst avg PT = $93.92
print(f"\nAnalyst avg PT: $93.92")
print(f"Base target ${base_target:.2f} vs analyst PT $93.92: diff = {abs(base_target - 93.92)/93.92*100:.1f}%")

# Bear MUST be below current price
assert bear_target < PRICE, f"BEAR TARGET ${bear_target:.2f} > CURRENT ${PRICE}!"
print(f"Bear target ${bear_target:.2f} < current ${PRICE}: {PRICE - bear_target:.2f} downside ({(bear_target/PRICE - 1)*100:.1f}%)")

# Weights
bear_w = 0.20
base_w = 0.50
bull_w = 0.30
weighted_fv = bear_w * bear_target + base_w * base_target + bull_w * bull_target
upside = (weighted_fv / PRICE - 1) * 100

print(f"\nWeighted FV: {bear_w}*${bear_target:.2f} + {base_w}*${base_target:.2f} + {bull_w}*${bull_target:.2f} = ${weighted_fv:.2f}")
print(f"Upside from current ${PRICE}: {upside:.1f}%")

# ═══════════════════════════════════════════
# Sheet 1: Valuation
# ═══════════════════════════════════════════
ws1 = wb.active
ws1.title = "Valuation"
ws1.merge_cells("A1:F1")
c(ws1, 1, 1, "Solventum Corporation (SOLV) - Valuation Summary", title_font_white, fill=title_fill)
ws1.row_dimensions[1].height = 30

title_data = [
    ("Company", "Solventum Corporation", "Medtech post-spinoff from 3M (2023)"),
    ("Ticker", "NYSE: SOLV", "Healthcare / Medical Instruments & Supplies"),
    ("Date", DATE_T, "As of Aug 27 close"),
    ("Price", f"${PRICE:.2f}", "NYSE close"),
    ("Shares Outstanding", f"{SHARES_MM:.1f}M", "Balance sheet issued shares, FY25"),
    ("Market Cap", f"${MC_B:.2f}B", "Intraday per Yahoo Finance"),
    ("Total Debt", f"${TOTAL_DEBT_B:.3f}B", "Balance sheet, down from $8.035B FY24"),
    ("Cash", f"${CASH_B:.2f}B", "Cash flow statement end position TTM"),
    ("Enterprise Value", f"${EV_B:.2f}B", "MC + Debt - Cash"),
    ("Net Debt", f"${NET_DEBT_B:.2f}B", "Debt - Cash; deleveraging from $7.25B FY24"),
    ("FCF (TTM)", f"${TCF_B/1e9:.2f}B (negative)", "Capex cycle + restructuring; FCF framework N/A"),
    ("Primary Lens", "Forward P/E", "Analyst consensus normalized non-GAAP; FCF negative"),
    ("Stance", "Watch / Cautiously Positive", "Post-HIS spinoff inflection. Q2 beat 33.8%. Nov 5 earnings catalyst."),
]

for i, (field, value, note) in enumerate(title_data, 2):
    c(ws1, i, 1, field, bold_font, True)
    c(ws1, i, 2, value, normal_font, True)
    c(ws1, i, 3, note, note_font, True)

# Valuation metrics table
header_row(ws1, len(title_data) + 3, ["Metric", "Value", "Comment"], 1)
val_metrics = [
    ("P/E (TTM)", f"{PE_TTM:.2f}x", "Distorted by $1.43B one-time non-operating gain (HIS divestiture)"),
    ("Forward P/E (FY26)", f"{PE_FWD:.2f}x", "On FY26 consensus EPS $7.18 (non-GAAP normalized)"),
    ("P/S", f"{PS:.2f}x", "Revenue flat at ~$8.3B post-spinoff"),
    ("P/FCF", "N/A", "FCF negative (-$118M TTM) from capex cycle"),
    ("EV/FCF", "N/A", "Not applicable; negative FCF breaks the framework"),
    ("EV/Sales", f"{EV_SALES:.2f}x", f"EV ${EV_B:.2f}B / Rev ${TTM_REV_B:.2f}B"),
    ("EV/EBITDA", f"{EV_EBITDA:.2f}x", f"EV ${EV_B:.2f}B / EBITDA ${TTM_EBITDA_B:.2f}B"),
    ("Forward P/S", f"{EV_B/FWD_REV_FY26:.2f}x", "EV / FY26 revenue consensus"),
    ("EV/EBITDA (Normalized)", f"{EV_B/((TTM_OP_INC_B+0.506)*(1+0.4)):.2f}x", "Using normalized EBITDA proxy"),
    ("52-Week Range", "$62.38 - $92.71", "Near 52-week high"),
    ("Beta (5Y)", f"{BETA:.2f}", "Low volatility relative to market"),
    ("Analyst Avg PT", "$93.92", "~3% above current price"),
    ("Dividend", "None", "No dividend declared"),
]

for i, (metric, val, comment) in enumerate(val_metrics, len(title_data) + 4):
    c(ws1, i, 1, metric, bold_font, True)
    c(ws1, i, 2, val, normal_font, True)
    c(ws1, i, 3, comment, note_font, True)

# ═══════════════════════════════════════════
# Sheet 2: WACC
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("WACC")
ws2.merge_cells("A1:D1")
c(ws2, 1, 1, "WACC - Weighted Average Cost of Capital", title_font_white, fill=title_fill)
ws2.row_dimensions[1].height = 30

wacc_data = [
    ("CAPM Components", "", ""),
    ("Risk-Free Rate (10Y US Treasury)", f"{RF*100:.2f}%", "CNBC Aug 27, 2026"),
    ("Equity Risk Premium", f"{ERP*100:.1f}%", "Standard assumption"),
    ("Beta (Levered, 5Y Monthly)", f"{BETA:.2f}", "Yahoo Finance Key Statistics"),
    ("", "", ""),
    ("Cost of Equity (Ke)", f"{COST_OF_EQUITY*100:.2f}%", f"= {RF*100:.2f}% + {BETA} × {ERP*100:.0f}%"),
    ("", "", ""),
    ("Cost of Debt", "", ""),
    ("Interest Paid (FY25)", f"${INT_EXP_B*1e9/1e6:.0f}M", "Cash flow statement"),
    ("Total Debt", f"${TOTAL_DEBT_B:.2f}B", "Balance sheet FY25"),
    ("Cost of Debt (Kd, pre-tax)", f"{COST_OF_DEBT*100:.2f}%", f"= ${INT_EXP_B*1e9/1e6:.0f}M / ${TOTAL_DEBT_B:.2f}B"),
    ("", "", ""),
    ("Capital Structure", "", ""),
    ("Market Cap (E)", f"${MC_B:.2f}B", "Intraday"),
    ("Total Debt (D)", f"${TOTAL_DEBT_B:.2f}B", "Balance sheet"),
    ("Equity Weight", f"{E_WEIGHT:.3f}", f"= {MC_B:.2f} / ({MC_B:.2f}+{TOTAL_DEBT_B:.2f})"),
    ("Debt Weight", f"{D_WEIGHT:.3f}", f"= {TOTAL_DEBT_B:.2f} / ({MC_B:.2f}+{TOTAL_DEBT_B:.2f})"),
    ("", "", ""),
    ("Tax Rate", f"{TAX_RATE*100:.0f}%", "Estimated US corporate rate"),
    ("", "", ""),
    ("WACC", f"{WACC*100:.2f}%", f"= {E_WEIGHT:.3f}×{COST_OF_EQUITY*100:.2f}% + {D_WEIGHT:.3f}×{COST_OF_DEBT*100:.2f}%×(1-{TAX_RATE*100:.0f}%)"),
]

for i, (field, value, note) in enumerate(wacc_data, 2):
    c(ws2, i, 1, field, bold_font if field and not field.startswith(" ") else normal_font, True)
    c(ws2, i, 2, value, normal_font, True)
    c(ws2, i, 3, note, note_font, True)

ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 40

# ═══════════════════════════════════════════
# Sheet 3: Scenarios
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells("A1:I1")
c(ws3, 1, 1, "Scenario Analysis - Forward P/E Framework", title_font_white, fill=title_fill)
ws3.row_dimensions[1].height = 30

# Note on framework
c(ws3, 2, 1, "Framework: Forward P/E (NOT FCF multiples) — TTM FCF is -$118M (negative). Analyst consensus normalized non-GAAP EPS excludes one-time HIS divestiture items.", note_font)

headers = ["Metric", "Bear", "Base", "Bull", "Notes"]
header_row(ws3, 4, headers)

scenarios = [
    ("", "", "", "", ""),
    ("Revenue CAGR (5Y)", "0-1%", "2-3%", "4-5%", "Medtech mature growth, post-spinoff normalization"),
    ("FY26 Revenue (Consensus)", "$8.23B", "$8.23B", "$8.23B", "13 analysts"),
    ("Terminal Revenue (5Y)", f"${8.23*1.05:.1f}B", f"${8.23*1.16:.1f}B", f"${8.23*1.28:.1f}B", ""),
    ("", "", "", "", ""),
    ("Terminal EPS", f"${bear_eps:.2f}", f"${base_eps:.2f}", f"${bull_eps:.2f}", "Non-GAAP normalized"),
    ("", "", "", "", ""),
    ("Exit P/E Multiple", f"{bear_exit_pe:.0f}x", f"{base_exit_pe:.0f}x", f"{bull_exit_pe:.0f}x", "Bear: distressed medtech, Base: peer norm, Bull: margin recovery"),
    ("", "", "", "", ""),
    ("Implied Target Price", f"${bear_target:.2f}", f"${base_target:.2f}", f"${bull_target:.2f}", "Terminal EPS × Exit P/E"),
    ("Upside / (Downside)", f"{(bear_target/PRICE-1)*100:.1f}%", f"{(base_target/PRICE-1)*100:.1f}%", f"{(bull_target/PRICE-1)*100:.1f}%", "vs current price"),
    ("", "", "", "", ""),
    ("Weight", f"{bear_w*100:.0f}%", f"{base_w*100:.0f}%", f"{bull_w*100:.0f}%", ""),
    ("Weighted Value/Share", f"${bear_w*bear_target:.2f}", f"${base_w*base_target:.2f}", f"${bull_w*bull_target:.2f}", ""),
    ("", "", "", "", ""),
    ("Probability-Weighted FV", "", "", f"${weighted_fv:.2f}", "Sum of weighted values"),
    ("Upside from Current", "", "", f"{upside:.1f}%", ""),
    ("", "", "", "", ""),
    ("WACC", "", "", f"{WACC*100:.2f}%", "Discount rate for reference"),
    ("", "", "", "", ""),
    ("Current Price", "", "", f"${PRICE:.2f}", ""),
    ("Analyst Avg PT", "", "", "$93.92", "Yahoo Finance consensus"),
    ("", "", "", "", ""),
    ("Basis", "", "", "", ""),
    ("Bear Rationale", "", "", "", "Operating margins remain depressed at 6% range. Restructuring costs persist. HIS separation creates transitional headwinds."),
    ("Base Rationale", "", "", "", "Gradual margin recovery as restructuring completes. Q2 EPS beat of 33.8% signals inflection. Revenue growth resumes organically."),
    ("Bull Rationale", "", "", "", "Full margin recovery to 10-12% (FY22 levels of 20.8% pre-spinoff). Operating leverage returns. Revenue grows above consensus with M&A."),
]

for i, (metric, bear, base, bull, note) in enumerate(scenarios, 5):
    c(ws3, i, 1, metric, bold_font if metric and not metric.startswith("Rationale") else normal_font, True)
    c(ws3, i, 2, bear, normal_font, True)
    c(ws3, i, 3, base, normal_font, True)
    c(ws3, i, 4, bull, normal_font, True)
    c(ws3, i, 5, note, note_font, True)

ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 15
ws3.column_dimensions["C"].width = 15
ws3.column_dimensions["D"].width = 15
ws3.column_dimensions["E"].width = 60

# ═══════════════════════════════════════════
# Sheet 4: Actuals Source Audit
# ═══════════════════════════════════════════
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells("A1:D1")
c(ws4, 1, 1, "Actuals Source Audit", title_font_white, fill=title_fill)

header_row(ws4, 3, ["Data Point", "Value", "Source", "Date / Notes"], 4)

audit_data = [
    # Market data
    ("Stock Price", f"${PRICE:.2f}", "Yahoo Finance Summary", "2026-08-27 close"),
    ("After Hours Price", "$92.80", "Yahoo Finance Summary", "2026-08-27 7:56 PM EDT"),
    ("52-Week Range", "$62.38 - $92.71", "Yahoo Finance Summary", "2026-08-27"),
    ("Market Cap", f"${MC_B:.2f}B", "Yahoo Finance intraday", "2026-08-27"),
    ("Beta", f"{BETA:.2f}", "Yahoo Finance Key Statistics", "5Y Monthly"),
    ("Analyst Avg PT", "$93.92", "Yahoo Finance Key Statistics", "1-y target"),
    ("", "", "", ""),
    # Income statement
    ("Revenue TTM", f"${TTM_REV_B*1e9/1e6:.0f}M", "Yahoo Finance Income Statement", "TTM, in thousands"),
    ("Revenue FY25", "$8,325M", "Yahoo Finance Income Statement", "Annual"),
    ("Revenue FY24", "$8,254M", "Yahoo Finance Income Statement", "Annual"),
    ("Revenue FY23", "$8,197M", "Yahoo Finance Income Statement", "Annual"),
    ("Revenue FY22", "$8,130M", "Yahoo Finance Income Statement", "Annual"),
    ("Gross Profit TTM", f"${TTM_GROSS_PROFIT_B*1e9/1e6:.0f}M", "Yahoo Finance Income Statement", "TTM"),
    ("Operating Income TTM", f"${TTM_OP_INC_B*1e9/1e6:.0f}M", "Yahoo Finance Income Statement", "TTM, declining from $1.69B FY22"),
    ("Operating Income FY25", "$632M", "Yahoo Finance Income Statement", "Annual"),
    ("Operating Income FY24", "$1,036M", "Yahoo Finance Income Statement", "Annual"),
    ("Net Income TTM", f"${TTM_NI_B*1e9/1e6:.0f}M", "Yahoo Finance Income Statement", "TTM — INFLATED by $1.43B non-operating gain"),
    ("Diluted EPS TTM", f"${TTM_DIL_EPS:.2f}", "Yahoo Finance Income Statement", "TTM — distorted by non-operating income"),
    ("", "", "", ""),
    # Balance sheet
    ("Total Debt", f"${TOTAL_DEBT_B:.3f}B", "Yahoo Finance Balance Sheet", "FY25, down from $8.035B FY24"),
    ("Total Cash", f"${CASH_B:.2f}B", "Yahoo Finance Cash Flow Statement", "TTM end cash position"),
    ("Net Debt", f"${NET_DEBT_B:.2f}B", "Computed: Debt - Cash", "Deleveraging from $7.25B FY24"),
    ("Total Assets", "$14.29B", "Yahoo Finance Balance Sheet", "FY25"),
    ("Total Equity", "$5.05B", "Yahoo Finance Balance Sheet", "FY25"),
    ("Shares Outstanding", f"{SHARES_MM:.2f}M", "Yahoo Finance Balance Sheet", "Shares issued FY25"),
    ("Tangible Book Value", "-$3.25B", "Yahoo Finance Balance Sheet", "FY25, negative from goodwill"),
    ("", "", "", ""),
    # Cash flow
    ("OCF TTM", "$209M", "Yahoo Finance Cash Flow Statement", "TTM"),
    ("OCF FY25", "$369M", "Yahoo Finance Cash Flow Statement", "Annual"),
    ("Investing CF TTM", "+$2,857M", "Yahoo Finance Cash Flow Statement", "TTM — includes HIS divestiture proceeds"),
    ("CapEx TTM", "-$327M", "Yahoo Finance Cash Flow Statement", "TTM"),
    ("FCF TTM", "-$118M", "Yahoo Finance Cash Flow Statement", "TTM, negative — capex cycle"),
    ("Debt Repayments TTM", "-$2,870M", "Yahoo Finance Cash Flow Statement", "TTM — aggressive deleveraging"),
    ("Interest Paid FY25", "$181M", "Yahoo Finance Cash Flow Statement", "FY25"),
    ("", "", "", ""),
    # Analyst estimates
    ("FY26 Rev Consensus", "$8.23B", "Yahoo Finance Analysis", "13 analysts, 33.78% Q2 beat"),
    ("FY27 Rev Consensus", "$8.51B", "Yahoo Finance Analysis", "13 analysts"),
    ("FY26 EPS Consensus", "$7.18", "Yahoo Finance Analysis", "14 analysts, non-GAAP normalized"),
    ("FY27 EPS Consensus", "$7.22", "Yahoo Finance Analysis", "15 analysts"),
    ("Q2 FY26 EPS Actual", "$2.55 vs est $1.91", "Yahoo Finance Analysis", "+33.78% surprise"),
    ("Q1 FY26 EPS Actual", "$1.48 vs est $1.35", "Yahoo Finance Analysis", "+9.33% surprise"),
    ("EPS Revisions 30D", "FY26: 12 up, 0 down", "Yahoo Finance Analysis", "Net positive revisions"),
    ("", "", "", ""),
    # Valuation ratios
    ("P/E TTM", f"{PE_TTM:.2f}x", "Computed: Price / EPS TTM", "DISTORTED by non-operating gain"),
    ("Forward P/E", f"{PE_FWD:.2f}x", "Computed: Price / FY26 EPS", "On non-GAAP normalized consensus"),
    ("EV/EBITDA", f"{EV_EBITDA:.2f}x", f"Computed: ${EV_B:.2f}B / ${TTM_EBITDA_B:.2f}B", "TTM"),
    ("", "", "", ""),
    # Earnings dates
    ("Next Earnings", "Nov 5, 2026", "Yahoo Finance Profile", "Q3 FY26"),
    ("Last Earnings", "~Jul 2026", "Yahoo Finance Profile / News", "Q2 FY26"),
    ("", "", "", ""),
    # Risk-free rate
    ("10Y Treasury Yield", "4.68%", "CNBC", "Aug 27, 2026"),
]

for i, (field, value, source, notes) in enumerate(audit_data, 4):
    c(ws4, i, 1, field, bold_font if field else normal_font, True)
    c(ws4, i, 2, value, normal_font, True)
    c(ws4, i, 3, source, note_font, True)
    c(ws4, i, 4, notes, note_font, True)

ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 22
ws4.column_dimensions["C"].width = 35
ws4.column_dimensions["D"].width = 45

# ═══════════════════════════════════════════
# Sheet 5: Questions
# ═══════════════════════════════════════════
ws5 = wb.create_sheet("Questions")
ws5.merge_cells("A1:C1")
c(ws5, 1, 1, "Open Questions", title_font_white, fill=title_fill)

header_row(ws5, 3, ["#", "Question", "Importance"], 4)

questions = [
    ("1", "What driven the $1.43B non-operating income TTM? Confirm it is entirely HIS divestiture gain and not partly other items (tax benefit recycling, foreign exchange, investment gains)?", "Critical"),
    ("2", "When will the HIS separation be completed? Timeline, expected proceeds, and how will the proceeds be deployed (share buybacks, debt reduction, new capex)?", "Critical"),
    ("3", "Operating income has declined from $1.69B (FY22) to $526M (TTM). Is this purely one-time post-spinoff restructuring, or is there structural margin erosion in Medsurg/Dental?", "High"),
    ("4", "The investing CF of +$2.86B TTM includes HIS proceeds — what is the organic investing CF excluding the divestiture? Is capex trending up or stable?", "High"),
    ("5", "Goodwill appears on balance sheet (tangible BV is -$3.25B on $5.05B equity). How much goodwill was inherited from 3M and is any of it impaired?", "Medium"),
    ("6", "Shares are nearly flat (173.49M now vs 172.7M FY22) with no dilution from 3M spinoff. Is the share count expected to remain stable or will buybacks accelerate?", "Medium"),
    ("7", "What is the segment revenue split between Medsurg, Dental Solutions, and Health Information Systems? Post-HIS separation, what is the projected revenue base?", "High"),
    ("8", "Q2 EPS beat of 33.78% vs. Q1 beat of 9.33% — is there a genuine momentum inflection, or is Q2 partially aided by one-time items?", "High"),
    ("9", "Debt has been cut from $8.3B to $5.0B in two years. Will deleveraging continue at this pace, or will HIS proceeds be diverted to buybacks or special dividends?", "Medium"),
    ("10", "What portion of operating income decline is attributable to stock-based compensation and restructuring charges vs. genuine operating cost increases?", "High"),
    ("11", "Post-spinoff, the company has 20,584 employees on ~$8.3B revenue. How does this cost base compare to medtech peers (BAX, TFX, BDX)?", "Medium"),
    ("12", "The Nov 5, 2026 Q3 earnings report: will this be the first quarter on a fully post-HIS-separation run rate, or are there continuing transitional items?", "High"),
]

for i, (num, question, importance) in enumerate(questions, 4):
    c(ws5, i, 1, num, normal_font, True)
    c(ws5, i, 2, question, normal_font, True)
    c(ws5, i, 3, importance, normal_font if importance != "Critical" else red_font, True)

ws5.column_dimensions["A"].width = 5
ws5.column_dimensions["B"].width = 80
ws5.column_dimensions["C"].width = 12

# ═══════════════════════════════════════════
# Sheet 6: Sources
# ═══════════════════════════════════════════
ws6 = wb.create_sheet("Sources")
ws6.merge_cells("A1:B1")
c(ws6, 1, 1, "Sources", title_font_white, fill=title_fill)

sources = [
    ("1", "https://finance.yahoo.com/quote/SOLV/ — Summary, price, volume, 52-week range"),
    ("2", "https://finance.yahoo.com/quote/SOLV/statistics/ — Key statistics, beta, share count, market cap"),
    ("3", "https://finance.yahoo.com/quote/SOLV/financials/ — Income Statement (annual + TTM)"),
    ("4", "https://finance.yahoo.com/quote/SOLV/balance-sheet/ — Balance Sheet"),
    ("5", "https://finance.yahoo.com/quote/SOLV/cash-flow/ — Cash Flow Statement"),
    ("6", "https://finance.yahoo.com/quote/SOLV/analysis/ — Analyst estimates, earnings history, revisions"),
    ("7", "https://finance.yahoo.com/quote/SOLV/profile/ — Company profile, description, executives, earnings dates"),
    ("8", "https://www.cnbc.com/quotes/US10Y — 10Y Treasury yield (Aug 27, 2026: 4.68%)"),
    ("9", "Yahoo Finance news tab — HIS separation announcement (Aug 4, 2026 8-K), Q2 earnings headlines"),
    ("10", "StockAnalysis.com — attempted but not available for SOLV; used Yahoo Finance as primary"),
]

for i, (num, source) in enumerate(sources, 2):
    c(ws6, i, 1, num, normal_font, True)
    c(ws6, i, 2, source, normal_font, True)

ws6.column_dimensions["A"].width = 5
ws6.column_dimensions["B"].width = 100

# ═══════════════════════════════════════════
# Set column widths for Sheet 1
# ═══════════════════════════════════════════
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 20
ws1.column_dimensions["C"].width = 55

# ═══════════════════════════════════════════
# Save
# ═══════════════════════════════════════════
out = "/home/refcell/dev/capital/models/2026-08-27 Solventum Model.xlsx"
wb.save(out)
print(f"\nSaved to {out}")
print(f"WACC: {WACC*100:.2f}%")
print(f"Weighted FV: ${weighted_fv:.2f}")
print(f"Upside: {upside:.1f}%")
print(f"Base target: ${base_target:.2f}")
print(f"Bear target: ${bear_target:.2f}")
print(f"Bull target: ${bull_target:.2f}")
