#!/usr/bin/env python3
"""
Build script for [2026-08-19] Repligen Corporation (RGEN) Model
6-sheet workbook: Valuation, WACC, Scenarios, Actuals Source Audit, Questions, Sources
Data source: StockAnalysis.com (Aug 19, 2026)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Styles ──
thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
hdr_font = Font(bold=True, underline="single", size=11)
tit_font = Font(bold=True, size=14)
sub_font = Font(italic=True, size=10, color="666666")
hdr_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
sec_font = Font(bold=True, size=11, color="1F4E79")


def c(ws, row, col, value, font=None, border=False, fill=None, align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if border:
        cell.border = thin
    if fill:
        cell.fill = fill
    if align in ("center", "right"):
        cell.alignment = Alignment(horizontal=align)
    return cell


# ── Constants ──
price = 176.36
shares_mm = 56.44
mc_mm = 9950  # $M
ev_mm = 9830  # $M
cash_mm = 810.45
debt_mm = 690.94
net_cash = cash_mm - debt_mm  # 119.51
beta = 1.01
rf = 4.637  # CNBC US10Y Aug 19 2026
erp = 5.0
ke = rf + beta * erp
kd = 4.5
tax_rate = 20.35
total_cap = mc_mm + debt_mm
we = mc_mm / total_cap
wd = debt_mm / total_cap
wacc = we * ke + wd * kd * (1 - tax_rate / 100)

fcf_ttm = 111.97  # $M
ocf_ttm = 134.86
ebitda_ttm = 144.10
op_income_ttm = 65.32
ni_ttm = 41.54
rev_ttm = 785.10

# ── Sheet 1: Valuation ──
ws1 = wb.active
ws1.title = "Valuation"
ws1.merge_cells("A1:D1")
c(ws1, 1, 1, "[2026-08-19] Repligen Corporation (RGEN) — Valuation Model", tit_font)
ws1.merge_cells("A2:D2")
c(ws1, 2, 1, "Primary framework: FCF multiple — trailing P/E of 241.66x is post-loss distortion", sub_font)

vdata = [
    ("Company", "Repligen Corporation", "Life Sciences / Bioprocessing"),
    ("Ticker / Exchange", "RGEN / NASDAQ", "NASDAQ-listed"),
    ("Date", "August 19, 2026", "Market close"),
    ("Stock Price", "$176.36", "+10.21 (+6.15%)"),
    ("Shares Outstanding", "56.44M", "+0.94% YoY"),
    ("Market Cap", "$9,950M", "~$9.95B"),
    ("Enterprise Value", "$9,830M", "MC + debt $690.9M - cash $810.5M"),
    ("Net Cash", "$119.5M", "Cash-heavy; debt/equity 0.33"),
    ("Forward P/E", "77.70x", "Implies ~$2.27 EPS FY27"),
    ("PEG", "3.67", "Forward P/E / EPS growth; well above 1.0"),
    ("Stance", "HOLD", "See scenarios"),
    ("", "", ""),
    ("VALUATION METRICS", "", ""),
    ("P/E (Trailing)", "241.66x", "Distorted; FY24 loss ($25.5M NI) drags TTM EPS to $0.73"),
    ("P/FCF", "88.89x", "Very expensive; reflects scarcity premium"),
    ("EV/FCF", "87.83x", "Enterprise-level view"),
    ("EV/Sales", "12.53x", "Premium for niche growth"),
    ("EV/EBITDA", "68.24x", "High EBITDA multiple; justified if 15%+ growth sustains"),
    ("P/S", "12.68x", "Forward P/S 11.20x"),
    ("P/B", "4.71x", "BVPS $37.43; premium for intangibles"),
]

for i, (m, v, n) in enumerate(vdata, 3):
    f = sec_font if m.startswith("VALUATION METRICS") else None
    c(ws1, i, 1, m, font=f, border=True, fill=hdr_fill if f else None)
    c(ws1, i, 2, v, border=True)
    c(ws1, i, 3, n, border=True)

for ci in range(1, 4):
    ws1.column_dimensions[get_column_letter(ci)].width = [32, 18, 60][ci - 1]


# ── Sheet 2: WACC ─────────
ws2 = wb.create_sheet("WACC")
ws2.merge_cells("A1:D1")
c(ws2, 1, 1, "WACC — CAPM Framework", tit_font)
ws2.merge_cells("A2:D2")
c(ws2, 2, 1,
    f"Computed WACC: {wacc:.3f}% vs StockAnalysis reported 9.36% (close match)",
    sub_font)

wdata = [
    ("Risk-Free Rate (10Y US)", f"{rf:.3f}%", "CNBC US10Y Aug 19 2026"),
    ("Equity Risk Premium", f"{erp:.1f}%", "Standard"),
    ("Beta (5Y)", f"{beta:.2f}", "StockAnalysis"),
    ("Cost of Equity (Ke)", f"{ke:.3f}%", f"{rf:.3f}+{beta:.2f}*{erp:.1f}"),
    ("Cost of Debt (Kd)", f"{kd:.1f}%", "Investment-grade est."),
    ("Tax Rate", f"{tax_rate:.2f}%", "TTM effective"),
    ("Market Cap", f"{mc_mm:.0f}", "$9.95B"),
    ("Total Debt", f"{debt_mm:.2f}", "$690.9M"),
    ("Equity Weight", f"{we*100:.2f}%", f"{mc_mm:.0f}/{total_cap:.0f}"),
    ("Debt Weight", f"{wd*100:.2f}%", f"{debt_mm:.2f}/{total_cap:.0f}"),
    ("", "", ""),
    ("WACC", f"{wacc:.3f}%",
        f"{we*100:.1f}%*{ke:.3f}%+{wd*100:.1f}%*{kd:.1f}%(1-{tax_rate:.2f}%)"),
]

for i, (m, v, n) in enumerate(wdata, 3):
    f = hdr_font if m == "WACC" else None
    c(ws2, i, 1, m, font=f, border=True, fill=hdr_fill if f else None)
    c(ws2, i, 2, v, border=True)
    c(ws2, i, 3, n, border=True)

for ci in range(1, 4):
    ws2.column_dimensions[get_column_letter(ci)].width = [35, 18, 55][ci - 1]

print(f"WACC: {wacc:.3f}%")


# ── Sheet 3: Scenarios ──
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells("A1:E1")
c(ws3, 1, 1,
    "Scenario Analysis — FCF Multiple Framework (5Y Horizon)", tit_font)
ws3.merge_cells("A2:E2")
c(ws3, 2, 1,
    "Primary framework: FCF multiple. Trailing P/E of 241.66x is uninformative.",
    sub_font)

hdrs = ["Metric", "Bear", "Base", "Bull", "Notes"]
for ci, h in enumerate(hdrs, 1):
    c(ws3, 3, ci, h, font=hdr_font, border=True, fill=hdr_fill, align="center")

s = [
    {"name": "Bear", "cagr": 10, "mult": 35, "w": 0.20},
    {"name": "Base", "cagr": 15, "mult": 48, "w": 0.50},
    {"name": "Bull", "cagr": 20, "mult": 62, "w": 0.30},
]

fcf5 = [fcf_ttm * (1 + x["cagr"] / 100) ** 5 for x in s]
ev5 = [fc * x["mult"] for fc, x in zip(fcf5, s)]
eq5 = [e + net_cash for e in ev5]
prc5 = [e / shares_mm for e in eq5]

weighted = sum(x["w"] * p for x, p in zip(s, prc5))
up_w = (weighted - price) / price * 100

rows = [
    ("Revenue CAGR (5Y)", f"{s[0]['cagr']}%", f"{s[1]['cagr']}%", f"{s[2]['cagr']}%",
     "Growth assumption"),
    (f"Terminal FCF (Y5, $M)", f"${fcf5[0]:.1f}", f"${fcf5[1]:.1f}", f"${fcf5[2]:.1f}",
     f"FCF {fcf_ttm:.2f}M * (1+cagr)^5"),
    ("Exit EV/FCF Multiple", f"{s[0]['mult']}x", f"{s[1]['mult']}x", f"{s[2]['mult']}x",
     "Below/at/above current comp."),
    (f"Implied EV ($M)", f"${ev5[0]:.0f}", f"${ev5[1]:.0f}", f"${ev5[2]:.0f}",
     "FCF * multiple"),
    ("Less Net Debt Adj. ($M)", f"${net_cash:.1f}", f"${net_cash:.1f}", f"${net_cash:.1f}",
     f"Cash {cash_mm}M - debt {debt_mm}M"),
    (f"Target Price/Share", f"${prc5[0]:.2f}", f"${prc5[1]:.2f}", f"${prc5[2]:.2f}",
     "Eq / shares"),
    (f"Upside %", f"{(prc5[0]-price)/price*100:.1f}%",
     f"{(prc5[1]-price)/price*100:.1f}%",
     f"{(prc5[2]-price)/price*100:.1f}%",
     f"From ${price:.2f}"),
    ("", "", "", "", ""),
    ("Weight", f"{s[0]['w']*100:.0f}%", f"{s[1]['w']*100:.0f}%", f"{s[2]['w']*100:.0f}%",
     ""),
    ("Weighted Val/Share", f"${s[0]['w']*prc5[0]:.2f}", f"${s[1]['w']*prc5[1]:.2f}",
     f"${s[2]['w']*prc5[2]:.2f}", ""),
    ("", "", "", "", ""),
    ("Probability-Weighted FV", "", "", f"${weighted:.2f}",
     f"Upside: {up_w:.2f}% from ${price:.2f}"),
    ("Current Price", "", "", f"${price:.2f}", "Aug 19 2026"),
    ("Analyst Avg PT", "", "", "$185.70", "22 analysts"),
    ("", "", "", "", ""),
    ("FRAMEWORK NOTE", "", "", "",
     "Post-FY24 loss drags trailing P/E to 241x. Forward P/E 77.7x (PEG 3.67) "
     "is elevated. FCF margin of 14.3% and 16%+ revenue growth are the story. "
     "Bear-case multiple of 35x still reflects premium positioning. "
     "If growth disappoints, downside risk is material at 87x EV/FCF."),
]

for i, (m, b, ba, bu, n) in enumerate(rows, 4):
    f = hdr_font if m in ("Probability-Weighted FV", "FRAMEWORK NOTE") else None
    c(ws3, i, 1, m, font=f, border=True, fill=hdr_fill if f else None)
    c(ws3, i, 2, b, border=True, align="center")
    c(ws3, i, 3, ba, border=True, align="center")
    c(ws3, i, 4, bu, border=True, align="center")
    c(ws3, i, 5, n, border=True)

for ci in range(1, 6):
    ws3.column_dimensions[get_column_letter(ci)].width = [28, 14, 14, 14, 65][ci - 1]

print(f"Scenarios: Bear=${prc5[0]:.2f}, Base=${prc5[1]:.2f}, Bull=${prc5[2]:.2f}")
print(f"Weighted FV: ${weighted:.2f}  Upside: {up_w:.2f}%")


# ── Sheet 4: Actuals Source Audit ────
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells("A1:D1")
c(ws4, 1, 1, "Actuals Source Audit", tit_font)

hdrs4 = ["Data Point", "Value", "Source", "Notes"]
for ci, h in enumerate(hdrs4, 1):
    c(ws4, 2, ci, h, font=hdr_font, border=True, fill=hdr_fill, align="center")

audit = [
    ("Stock Price", "$176.36",
     "stockanalysis.com/stocks/rgen/",
     "Aug 19 2026 close; +10.21 (+6.15%)"),
    ("Market Cap", "$9.95B",
     "stockanalysis.com/stocks/rgen/statistics/",
     "9,950M"),
    ("Enterprise Value", "$9.83B",
     "stockanalysis.com/stocks/rgen/statistics/",
     "9,830M"),
    ("Shares Outstanding", "56.44M",
     "stockanalysis.com/stocks/rgen/statistics/",
     "+0.94% YoY; float 53.47M"),
    ("Revenue TTM", "$785.1M",
     "stockanalysis.com/stocks/rgen/financials/",
     "+16.5% YoY from $673.8M"),
    ("Gross Profit TTM", "$422.7M",
     "stockanalysis.com/stocks/rgen/financials/",
     "53.84% margin"),
    ("Operating Income TTM", "$65.3M",
     "stockanalysis.com/stocks/rgen/financials/",
     "8.32% margin"),
    ("Net Income TTM", "$41.5M",
     "stockanalysis.com/stocks/rgen/financials/",
     "5.29% margin"),
    ("EPS Diluted TTM", "$0.73",
     "stockanalysis.com/stocks/rgen/financials/",
     "TTM EPS; FY25 EPS $0.86"),
    ("EBITDA TTM", "$144.1M",
     "stockanalysis.com/stocks/rgen/financials/",
     "18.35% margin"),
    ("D&A TTM", "$78.8M",
     "stockanalysis.com/stocks/rgen/financials/",
     "For EBITDA calculation"),
    ("OCF TTM", "$134.9M",
     "stockanalysis.com/stocks/rgen/statistics/",
     "Operating cash flow"),
    ("CapEx TTM", "-$22.9M",
     "stockanalysis.com/stocks/rgen/statistics/",
     "Capital expenditures"),
    ("FCF TTM", "$112.0M",
     "stockanalysis.com/stocks/rgen/statistics/",
     "OCF - CapEx = $134.9 - $22.9"),
    ("Cash & Equiv", "$810.5M",
     "stockanalysis.com/stocks/rgen/statistics/",
     "Cash 606.8M + ST Inv 203.7M"),
    ("Total Debt", "$690.9M",
     "stockanalysis.com/stocks/rgen/statistics/",
     "Debt/equity 0.33"),
    ("Beta (5Y)", "1.01",
     "stockanalysis.com/stocks/rgen/statistics/",
     "5-year monthly beta"),
    ("Forward P/E", "77.70",
     "stockanalysis.com/stocks/rgen/statistics/",
     "Implies ~$2.27 EPS"),
    ("PEG Ratio", "3.67",
     "stockanalysis.com/stocks/rgen/statistics/",
     "Well above 1.0"),
    ("Analyst Consensus", "Buy",
     "stockanalysis.com/stocks/rgen/forecast/",
     "22 analysts; 15 SB, 3 B, 4 H"),
    ("Avg Price Target", "$185.70",
     "stockanalysis.com/stocks/rgen/forecast/",
     "Low $145, High $220"),
    ("Earnings Date", "Jul 28, 2026",
     "stockanalysis.com/stocks/rgen/statistics/",
     "Already released prior to model build"),
    ("Tax Rate TTM", "20.35%",
     "stockanalysis.com/stocks/rgen/statistics/",
     "Effective tax rate"),
    ("Interest Expense TTM", "-$23.8M",
     "stockanalysis.com/stocks/rgen/financials/",
     "Interest expense; cov = 2.74x"),
    ("Rev Growth FY25->TTM", "16.49%",
     "stockanalysis.com/stocks/rgen/financials/",
     "+16.5% revenue growth"),
    ("Restructuring TTM", "-$4.2M",
     "stockanalysis.com/stocks/rgen/financials/",
     "M&A integration charges continuing"),
]

for i, (m, v, s, n) in enumerate(audit, 3):
    c(ws4, i, 1, m, border=True)
    c(ws4, i, 2, v, border=True)
    c(ws4, i, 3, s, border=True)
    c(ws4, i, 4, n, border=True)

for ci in range(1, 5):
    ws4.column_dimensions[get_column_letter(ci)].width = [30, 15, 48, 45][ci - 1]


# ── Sheet 5: Questions ───────────
ws5 = wb.create_sheet("Questions")
ws5.merge_cells("A1:C1")
c(ws5, 1, 1, "Open Questions", tit_font)

qs = [
    ("Q1", "Revenue spike in FY2022: The FY2022 revenue of $801.5M was a one-time peak driven by a major licensing deal or acquisition event. FY2023 dropped 21% to $632.4M. What was the FY2022 spike from? Was it a one-time licensing/royalty event, acquisition revenue, or a large one-off contract?"),
    ("Q2", "Restructuring/M&A charges: TTM restructuring charges of $4.2M and FY2024 had $46.9M in merger/restructuring charges. What M&A activity drove the $46.9M charge in FY2024? Is the continuing $4M+ TTM charge related to integration or amortization of acquired intangibles that will continue draining operating income?"),
    ("Q3", "Net income volatility pattern: FY2022 NI $186M, FY2023 NY 35.6M, FY2024 -$25.5M (loss), FY2025 $48.9M, TTM $41.5M. The $186M in FY2022 was a high-water mark that never returned. Are these structural (one-time tax benefits, litigation settlements, asset sales) or cyclical? Trailing P/E of 241x is largely meaningless."),
    ("Q4", "D&A composition: TTM D&A of $78.8M plus other amortization of $17.3M. How much D&A is from organic facilities vs acquired intangibles? Amortization of purchased IP or acquired licenses would suggest operating income is structurally above reported. For FCF valuation, D&A is a non-cash add-back anyway."),
    ("Q5", "Interest expense trajectory: Interest expense has risen from -$2.98M FY2023 to -$23.8M TTM. Interest coverage is only 2.74x ($65.3M op inc / $23.8M int exp). What is the debt composition and maturity? With $691M total debt, is this variable-rate exposure that could spike with rising rates?"),
    ("Q6", "Short interest at 10.87% of shares outstanding: High short interest (6.13M shares). What is the short thesis — multiple compression, growth deceleration, or competitive displacement? The 5.12-day cover ratio means shorts can exit quickly."),
    ("Q7", "Institutional ownership at 118.05%: The institutional ownership figure exceeds 100%, indicating significant short positions held by institutions (net borrowing). Is this unusual and does it signal bearish positioning or standard repo activity?"),
    ("Q8", "FCF multiple compression risk: At 88x P/FCF, RGEN is trading well above growth SaaS comparables. If FCF growth slows from 16% to single digits, the multiple collapses. What protects the multiple? The company is not yet a high-margin software business."),
    ("Q9", "Buyback yield is negative (-0.94%): Dilution of 0.94% YoY. Given the premium valuation, why is the company not returning capital to shareholders via buybacks or dividends when ROIC is only 2.6%? Management incentives for capital deployment?"),
    ("Q10", "FY26 earnings already released Jul 28: Q3 FY26 results were released prior to the Aug 19 model date. What guidance did management provide for FY27/28? The market has already incorporated Q3 outcomes. What is the next earnings catalyst? (~Oct 2026 for Q4 FY26?"),
]

for i, (q, t) in enumerate(qs, 2):
    c(ws5, i, 1, q, font=sec_font, border=True, align="center")
    c(ws5, i, 2, t, border=True)

ws5.column_dimensions["A"].width = 8
ws5.column_dimensions["B"].width = 120


# ── Sheet 6: Sources ─────
ws6 = wb.create_sheet("Sources")
ws6.merge_cells("A1:B1")
c(ws6, 1, 1, "Sources", tit_font)

srcs = [
    ("1", "StockAnalysis.com — RGEN Overview", "https://stockanalysis.com/stocks/rgen/"),
    ("2", "StockAnalysis.com — RGEN Financials", "https://stockanalysis.com/stocks/rgen/financials/"),
    ("3", "StockAnalysis.com — RGEN Income Statement", "https://stockanalysis.com/stocks/rgen/financials/income-statement/"),
    ("4", "StockAnalysis.com — RGEN Balance Sheet", "https://stockanalysis.com/stocks/rgen/financials/balance-sheet/"),
    ("5", "StockAnalysis.com — RGEN Cash Flow", "https://stockanalysis.com/stocks/rgen/financials/cash-flow-statement/"),
    ("6", "StockAnalysis.com — RGEN Statistics", "https://stockanalysis.com/stocks/rgen/statistics/"),
    ("7", "StockAnalysis.com — RGEN Forecast/Analysts", "https://stockanalysis.com/stocks/rgen/forecast/"),
    ("8", "CNBC — US10Y Treasury Yield", "https://cnbc.com/quotes/US10Y"),
]

for i, (n, name, url) in enumerate(srcs, 2):
    c(ws6, i, 1, n, border=True, align="center")
    c(ws6, i, 2, f"{name} — {url}", border=True)

ws6.column_dimensions["A"].width = 8
ws6.column_dimensions["B"].width = 120


# ── Save ──
fname = "models/[2026-08-19] Repligen Model.xlsx"
wb.save(fname)
print(f"Saved: {fname}")
print(f"WACC: {wacc:.3f}%")
print(f"FV: ${weighted:.2f}  Upside: {up_w:.2f}%")
