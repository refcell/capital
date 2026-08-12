"""Build [2026-08-11] ITa\xfa Unibanco Model.xlsx — 6-sheet bank valuation."""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# === Style helpers ===
thin = Side(style="thin")
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_f = Font(bold=True, size=11)
ttl_f = Font(bold=True, size=14, color="FFFFFF")
sub_f = Font(bold=True, size=12, italic=True)
dat_f = Font(size=10)
hdr_fill = PatternFill("solid", fgColor="4472C4")
lite_fill = PatternFill("solid", fgColor="D9E2F3")

# === Constants (Yahoo Finance, 2026-08-11) ===
PRICE      = 7.50
SHARES_M   = 11027   # millions; from BS ordinary shares
MC_B       = PRICE * SHARES_M / 1000   # ~82.7 B
TOT_DEBT_B = 60.62   # ~563.7B BRL / 9.3
TOT_CASH_B = 12.35   # ~114.9B BRL / 9.3
NET_DEBT_B = TOT_DEBT_B - TOT_CASH_B   # ~48.27
BETA       = 0.16
RISK_FREE  = 0.04682
ERP        = 0.05
TAX        = 0.124
KE         = RISK_FREE + BETA * ERP
KD_NET     = 0.10 * (1 - TAX)
WACC       = (MC_B / (MC_B + TOT_DEBT_B)) * KE + \
             (TOT_DEBT_B / (MC_B + TOT_DEBT_B)) * KD_NET

# === Workbook ===
wb = openpyxl.Workbook()

# =========================================================
# Sheet 1: Valuation
# =========================================================
ws = wb.active
ws.title = "Valuation"
ws.merge_cells("A1:E1")
c = ws["A1"]
c.value = "Itau Unibanco (ITUB) - Bank Valuation Model"
c.font = ttl_f
c.fill = hdr_fill
c.alignment = Alignment(horizontal="center")

pairs = [
    ("Company",   "Itau Unibanco Holding S.A."),
    ("Ticker",    "NYSE: ITUB"),
    ("Sector",    "Financials / Banks (Brazil)"),
    ("Date",      date.today().strftime("%Y-%m-%d")),
    ("Price",     "$7.50"),
    ("Shares (M)", str(SHARES_M)),
    ("Market Cap (B USD)", str(round(MC_B, 1))),
    ("Total Debt (B USD)",  str(TOT_DEBT_B)),
    ("Net Debt (B USD)",    str(round(NET_DEBT_B, 1))),
    ("EV (B USD)",         str(round(MC_B + NET_DEBT_B, 1))),
    ("Primary Lens",       "P/B + ROE (Bank framework)"),
    ("Stance",             "Watch"),
]
for i, (label, val) in enumerate(pairs, start=3):
    ws.cell(row=i, column=1, value=label).font = hdr_f
    ws.cell(row=i, column=2, value=val).font = dat_f

met_data = [
    ["Metric", "Value", "Comment"],
    ["Trailing P/E", "9.55x", "Yahoo Key Stats (TTM)"],
    ["Forward P/E",  "9.43x", "Yahoo Key Stats (FY27E)"],
    ["P/B Ratio",    "2.03x", "Yahoo Key Stats"],
    ["P/S",          "N/A",   "Bank - deposits offset loans"],
    ["P/FCF",        "N/A",   "FCF meaningless for banks"],
    ["EV/EBITDA",    "N/A",   "Deposits are operating liab."],
    ["ROE (TTM)",    "21.48%", "Yahoo Key Stats"],
    ["ROA (TTM)",    "1.58%",  "Yahoo Key Stats"],
    ["Net Margin",   "32.58%", "Yahoo Key Stats"],
    ["BVPS",         "$3.88",  "Yahoo Key Stats"],
    ["Fwd Div Yield","2.21%",  "Yahoo Key Stats"],
    ["Payout Ratio", "73.96%"," Yahoo Key Stats"],
    ["Beta (5Y)",    "0.16",   "Yahoo Key Stats - very low"],
]
ws["A15"] = "Key Valuation Metrics"
ws["A15"].font = sub_f
for i, row in enumerate(met_data, start=16):
    for j, txt in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=txt)
        c.font = hdr_f if i == 16 else dat_f
        c.border = bdr
        if i == 16:
            c.fill = hdr_fill

# =========================================================
# Sheet 2: WACC
# =========================================================
ws = wb.create_sheet("WACC")
ws.merge_cells("A1:C1")
c = ws["A1"]
c.value = "WACC - CAPM Components"
c.font = ttl_f
c.fill = hdr_fill

wacc_rows = [
    ("Component", "Value", "Source / Notes"),
    ("Risk-Free (10Y US)",   str(round(RISK_FREE * 100, 3)) + "%", "CNBC US10Y"),
    ("Equity Risk Premium",  str(round(ERP * 100, 1)) + "%",       "Assumed standard"),
    ("Beta (levered, 5Y)",   str(BETA),                           "Yahoo Key Stats"),
    ("Cost of Equity (Ke)",  str(round(KE * 100, 2)) + "%",        "CAPM: Rf + B*ERP"),
    ("Gross Cost of Debt",   "10.00%",                             "Market estimate"),
    ("After-Tax Kd",         str(round(KD_NET * 100, 2)) + "%",    "Kd*(1-T)"),
    ("Market Cap (B USD)",   str(round(MC_B, 1)),                  "Price*Shares"),
    ("Total Debt (B USD)",   str(TOT_DEBT_B),                      "From BS"),
    ("Equity Weight",        str(round(MC_B / (MC_B + TOT_DEBT_B), 3)), "MC/(MC+D)"),
    ("Debt Weight",          str(round(TOT_DEBT_B / (MC_B + TOT_DEBT_B), 3)), "D/(MC+D)"),
    ("WACC",                 str(round(WACC * 100, 2)) + "%",      "EqWt*Ke + DtWt*Kd(1-T)"),
]
for i, row in enumerate(wacc_rows, 2):
    for j, txt in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=txt)
        c.font = hdr_f if i == 2 else dat_f
        c.border = bdr
        if i == 2:
            c.fill = hdr_fill

# =========================================================
# Sheet 3: Scenarios (P/B + ROE)
# =========================================================
ws = wb.create_sheet("Scenarios")
ws.merge_cells("A1:L1")
c = ws["A1"]
c.value = "Scenario Analysis - P/B and ROE Framework"
c.font = ttl_f
c.fill = hdr_fill
ws["A2"] = "Banks: FCF is meaningless. Correct framework is Residual Income: BVPS CAGR + exit P/B."
ws["A2"].font = Font(italic=True, size=10)
ws["A2"].alignment = Alignment(wrap_text=True)

hdr_cols = ["Scenario", "BVPS CAGR", "Term BVPS",
            "Exit P/B", "Term FCF Margin", "Implied EV",
            "Less Net Debt", "Shares (M)",
            "Target Price", "Upside %", "Weight", "Wtd Val/Shr"]
scen_data = [
    ["Bear", "3.0%", "$4.02", "1.5x", "N/A", "N/A", "N/A", "11,027", "$6.03", "-19.6%", "25%", "$1.51"],
    ["Base", "5.5%", "$4.57", "2.0x", "N/A", "N/A", "N/A", "11,027", "$9.14", "+21.9%", "50%", "$4.57"],
    ["Bull", "7.0%", "$4.75", "2.5x", "N/A", "N/A", "N/A", "11,027", "$11.88", "+58.4%", "25%", "$2.97"],
    ["", "", "", "", "", "", "", "", "", "", "Prob-FV:", "$9.05"],
    ["", "", "", "", "", "", "", "", "", "", "Current:", "$7.50"],
    ["", "", "", "", "", "", "", "", "", "", "Upside:", "+20.7%"],
]
for i, row in enumerate([hdr_cols] + scen_data, start=4):
    for j, txt in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=txt)
        c.font = hdr_f if i == 4 else dat_f
        c.border = bdr
        if i == 4:
            c.fill = hdr_fill

# =========================================================
# Sheet 4: Actuals Source Audit
# =========================================================
ws = wb.create_sheet("Actuals Source Audit")
ws.merge_cells("A1:E1")
c = ws["A1"]
c.value = "Actuals Source Audit - ITUB"
c.font = ttl_f
c.fill = hdr_fill

audit = [
    ("Item", "Value", "Source", "Date", "Notes"),
    ("Price Close", "$7.50", "Yahoo Summary", "2026-08-11", "Down -4.82% that day"),
    ("Market Cap", "~$82.7B", "Yahoo Key Stats", "2026-08-11", "MC at close"),
    ("Shares Outstanding", "11,027M", "Yahoo BS", "FY2025", "Ordinary shares count"),
    ("Revenue TTM", "169.37B BRL", "Yahoo IS", "Q2 FY26 TTM", "~$18.2B USD"),
    ("Pretax Income TTM", "54.69B BRL", "Yahoo IS", "Q2 FY26 TTM", ""),
    ("Net Income TTM", "46.83B BRL", "Yahoo IS", "Q2 FY26 TTM", "~$5.04B USD"),
    ("Diluted EPS TTM", "4.20 BRL", "Yahoo IS", "Q2 FY26 TTM", ""),
    ("Operating CF TTM", "158.99B BRL", "Yahoo CF", "Q2 FY26 TTM", "~$17.1B USD"),
    ("Total Assets FY25", "3.066T BRL", "Yahoo BS", "2025-12-31", "Assets grew 7.4% YoY"),
    ("Total Debt FY25", "563.70B BRL", "Yahoo BS", "2025-12-31", "~$60.6B USD"),
    ("Common Equity FY25","204.50B BRL", "Yahoo BS", "2025-12-31", "~$21.99B USD"),
    ("Preferred Stock", "~$1.13B", "Total-Common Equity", "2025-12-31", "Capital structure item"),
    ("Effective Tax Rate","~12.4%", "4401/50250", "FY2025", "Preferencial regime"),
    ("ROE TTM", "21.48%", "Yahoo Key Stats", "2026-08-11", "Strong for LATAM"),
    ("BVPS", "$3.88", "Yahoo Key Stats", "2026-08-11", "From MC/Shares"),
    ("Forward P/E", "9.43x", "Yahoo Key Stats", "2026-08-11", "FY27E basis"),
    ("Analyst FY26 EPS","$0.89", "Yahoo Analysis", "10 analysts", "Non-GAAP normalized"),
    ("Analyst FY27 EPS","$0.99-1.01", "Yahoo Analysis", "Range", "Non-GAAP"),
    ("Analyst Rev FY26","193.69B BRL", "Yahoo Analysis", "10 analysts",""),
    ("Analyst Rev FY27","208.49B BRL", "Yahoo Analysis", "11 analysts",""),
    ("P/B Ratio", "2.03x", "Yahoo Key Stats", "2026-08-11", ""),
    ("Div Yield Fwd", "2.21%", "Yahoo Key Stats", "2026-08-11", ""),
    ("10Y Treasury", "4.682%", "CNBC US10Y", "2026-08-11", "WACC risk-free input"),
]
for i, row in enumerate(audit, 2):
    for j, txt in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=txt)
        c.font = hdr_f if i == 2 else dat_f
        c.border = bdr
        if i == 2:
            c.fill = hdr_fill

# =========================================================
# Sheet 5: Questions
# =========================================================
ws = wb.create_sheet("Questions")
ws.merge_cells("A1:C1")
c = ws["A1"]
c.value = "Open Questions - ITUB"
c.font = ttl_f
c.fill = hdr_fill

questions = [
    ("#", "Question"),
    ("1", "Preferred stock = ~$1.13B (Total-Common Equity). Terms, dividend obligations, fixed-charge coverage? Subtract from MC for common value?"),
    ("2", "Key Stats says 5.4B shares, BS shows 11.027B. Explain the discrepancy - ADS ratio? Timing mismatch?"),
    ("3", "OCF swung wildly: 159B TTM vs 34.5B FY25 vs 7.07B FY24. Working capital / loan growth timing?"),
    ("4", "Brazil SELIC at ~10.5%. How does rate trajectory affect NIM? Net gainer or loser from falling rates?"),
    ("5", "Buyback pace vs OCF? Payout 73.96% plus buybacks. Is buyback sustainable given OCF volatility?"),
    ("6", "Interest expense surged: 212.4B TTM vs 116.7B FY22. Deposit mix shift? Interbank funding increase?"),
    ("7", "Effective tax ~12.4% below statutory 25-30%. IRRF preferential regime? Deferred tax dynamics?"),
    ("8", "Customer concentration: Large corporate vs retail depositor base? SME exposure?"),
    ("9", "Competitive differentiation vs BB (Bradesco), Santander, Nubank? Tech, geography, fee mix?"),
    ("10", "Management guidance on FY26/FY27? Any explicit EPS/revenue/profit margin targets?"),
    ("11", "Next earnings date? Q2 FY26 reported Aug 2026 - when is Q3? Guidance update expected?"),
    ("12", "FX exposure: BRL/USD at ~9.3. Hedges? How does currency translate to ADR price?"),
    ("13", "NPL ratio trajectory, provision for credit losses, coverage ratio. Credit quality check?"),
    ("14", "Share count trend: 11.100M to 11.027M. Buybacks shrinking float? Rate of retiro?"),
]
for i, row in enumerate(questions, 2):
    for j, txt in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=txt)
        c.font = hdr_f if i == 2 else dat_f
        c.border = bdr
        if i == 2:
            c.fill = hdr_fill
ws.column_dimensions["B"].width = 100

# =========================================================
# Sheet 6: Sources
# =========================================================
ws = wb.create_sheet("Sources")
ws.merge_cells("A1:B1")
c = ws["A1"]
c.value = "Sources - ITUB Research"
c.font = ttl_f
c.fill = hdr_fill

sources = [
    ("#", "Source"),
    ("1", "Yahoo Finance - Income Statement: finance.yahoo.com/quote/ITUB/financials/"),
    ("2", "Yahoo Finance - Balance Sheet: finance.yahoo.com/quote/ITUB/balance-sheet/"),
    ("3", "Yahoo Finance - Cash Flow: finance.yahoo.com/quote/ITUB/cash-flow/"),
    ("4", "Yahoo Finance - Key Statistics: finance.yahoo.com/quote/ITUB/key-statistics/"),
    ("5", "Yahoo Finance - Analysis/Estimates: finance.yahoo.com/quote/ITUB/analysis/"),
    ("6", "Yahoo Finance - Profile: finance.yahoo.com/quote/ITUB/profile/"),
    ("7", "Yahoo Finance - Summary: finance.yahoo.com/quote/ITUB/"),
    ("8", "CNBC - 10Y Treasury: www.cnbc.com/quotes/US10Y"),
    ("9", "StockAnalysis - NOT available (404)"),
]
for i, row in enumerate(sources, 2):
    for j, txt in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=txt)
        c.font = hdr_f if i == 2 else dat_f
        c.border = bdr
        if i == 2:
            c.fill = hdr_fill
ws.column_dimensions["B"].width = 80

# =========================================================
# Save
# =========================================================
path = "models/[2026-08-11] Itau Unibanco Model.xlsx"
wb.save(path)
print("Saved:", path)
print("WACC:", round(WACC * 100, 2), "%")
print("Ke:", round(KE * 100, 2), "%")
print("Beta:", BETA)
print("Prob-Weighted FV: $9.05 vs Current $7.50 => +20.7%")
