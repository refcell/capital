#!/usr/bin/env python3
"""
Build ICE (Intercontinental Exchange) 6-sheet valuation model.
Standard exchange/data services valuation: FCF + EV/EBITDA lens.
Date: 2026-06-30
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
import os, datetime

wb = Workbook()

# ── Styles ──
bold = Font(bold=True)
title_font = Font(bold=True, size=14)
heading_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(), right=Side(), top=Side(), bottom=Side()
)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, color="FFFFFF")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def style_data(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

def write_valuation_row(ws, row, data, comment=None):
    for c, val in enumerate(data, 1):
        ws.cell(row=row, column=c, value=val)
    style_data(ws, row, len(data))
    if comment:
        ws.cell(row=row, column=len(data)+1, value=comment)

# ═════════════════════════════════════════════════════════
# DATA (all from Yahoo Finance + CNBC as of 2026-06-30 close)
# ═════════════════════════════════════════════════════════
PRICE = 123.11
SHARES_MM = 567.0  # ordinary shares outstanding 12/31/2025 from BS
MC_B = 69.51  # market cap $69.51B (Yahoo Finance statistics 6/29/2026)
EV_B = 86.55  # enterprise value $86.55B (Yahoo Finance statistics)
BETA = 0.92  # 5-year monthly beta
RISK_FREE = 4.461  # 10Y US Treasury from CNBC 6/30/2026
ERP = 5.0
TAX_RATE = 22.9  # TTM: prov $1,186M / pretax $5,176M = 22.9%
COST_OF_DEBT = 3.94  # interest expense $800M / total debt $20.28B ≈ 3.94%

# Net debt: use EV - MC as the cleanest measure (avoids restricted clearing cash ambiguity)
NET_DEBT_B = EV_B - MC_B  # ~$17.04B

# Income statement (in $millions)
REV = {"FY22": 9636, "FY23": 9903, "FY24": 11761, "FY25": 12640, "TTM": 13077}
OP_INCOME = {"FY22": 3731, "FY23": 3963, "FY24": 4413, "FY25": 4999, "TTM": 5452}
NI = {"FY22": 1446, "FY23": 2368, "FY24": 2754, "FY25": 3315, "TTM": 3931}
EPS_DIL = {"FY22": 2.58, "FY23": 4.19, "FY24": 4.78, "FY25": 5.77, "TTM": 6.87}
EBITDA = {"FY22": 3455, "FY23": 4917, "FY24": 6075, "FY25": 6709, "TTM": 7531}

# Cash flow (in $millions)
OCF = {"FY22": 3554, "FY23": 3542, "FY24": 4609, "FY25": 4662, "TTM": 5022}
CAPEX = {"FY22": 482, "FY23": 489, "FY24": 752, "FY25": 791, "TTM": 778}
FCF = {"FY22": 3072, "FY23": 3053, "FY24": 3857, "FY25": 3871, "TTM": 4244}

# Balance sheet highlights (in $millions)
TOTAL_DEBT = {
    "FY22": 18376, "FY23": 22912, "FY24": 20703, "FY25": 20279
}
EQUITY = {
    "FY22": 22706, "FY23": 25717, "FY24": 27647, "FY25": 28915
}

# Analyst estimates (normalized/consensus)
REV_FY26 = 10980  # avg estimate, 12 analysts
REV_FY27 = 11530  # avg estimate, 12 analysts
EPS_FY26 = 8.14   # avg estimate, 14 analysts
EPS_FY27 = 8.81   # avg estimate, 15 analysts

# Valuation multiples (from Yahoo Finance statistics 6/29/2026)
TRAILING_PE = 17.89
FWD_PE = 17.83
P_SALES = 5.39
P_BOOK = 2.36
EV_REVENUE = 6.62
EV_EBITDA = 11.49
# EV/FCF derived: EV $86.55B / FCF TTM $4.24B = 20.4x
EV_FCF = EV_B / (FCF["TTM"] / 1000)

# Analyst price targets
AVG_PRICE_TARGET = 153.00
HIGH_PRICE_TARGET = 251.00
ONE_YEAR_TARGET_EST = 195.00

# ═════════════════════════════════════════════════════════
# WACC COMPUTATION
# ═════════════════════════════════════════════════════════
# CAPM: Cost of Equity = Rf + Beta * ERP
COST_OF_EQUITY = RISK_FREE + BETA * ERP  # 4.461 + 0.92*5 = 9.061%

# After-tax cost of debt
AT_COST_OF_DEBT = COST_OF_DEBT * (1 - TAX_RATE / 100)

# Capital structure weights (market-based)
# Equity weight = MC / (MC + Net Debt) = 69.51 / 86.55 = 80.3%
# Debt weight = Net Debt / (MC + Net Debt) = 17.04 / 86.55 = 19.7%
# Actually: Total capitalization = MC + Total Debt (not net debt)
# Use EV-based weights: Equity/EV + Debt/EV
DEBT_B = NET_DEBT_B  # for WACC purposes, use net debt as the relevant leverage measure
WEIGHT_EQ = MC_B / (MC_B + DEBT_B) if (MC_B + DEBT_B) > 0 else 0.803
WEIGHT_DEBT = DEBT_B / (MC_B + DEBT_B) if (MC_B + DEBT_B) > 0 else 0.197

WACC = WEIGHT_EQ * COST_OF_EQUITY + WEIGHT_DEBT * AT_COST_OF_DEBT

print(f"WACC: {WACC:.2f}%")
print(f"Cost of Equity: {COST_OF_EQUITY:.2f}%")
print(f"After-tax Cost of Debt: {AT_COST_OF_DEBT:.2f}%")

# ═════════════════════════════════════════════════════════
# SCENARIO COMPUTATION
# ═════════════════════════════════════════════════════════
# Base revenue CAGR: bridge from FY25 $12.64B toward consensus trajectory
# FY26 est $10.98B seems low vs actual FY25 — likely pro-forma/restated basis
# Conservative: use ~7% CAGR reflecting blended organic + acquisition integration
# FCF margin TTM: $4.24B/$13.08B = 32.4%

SCENARIOS = {
    "Bear": {
        "rev_cagr_5y": 4.0,
        "fcs_margin": 28.0,
        "exit_fcf_mult": 15.0,
        "weight": 0.25,
    },
    "Base": {
        "rev_cagr_5y": 7.0,
        "fcs_margin": 32.4,
        "exit_fcf_mult": 18.0,
        "weight": 0.50,
    },
    "Bull": {
        "rev_cagr_5y": 10.0,
        "fcs_margin": 35.0,
        "exit_fcf_mult": 22.0,
        "weight": 0.25,
    },
}

# Compute terminal values
for name, s in SCENARIOS.items():
    terminal_rev = REV["TTM"] * ((1 + s["rev_cagr_5y"]/100) ** 5)
    terminal_fcf = terminal_rev * s["fcs_margin"] / 100  # in $M
    implied_ev_b = terminal_fcf * s["exit_fcf_mult"] / 1000  # in $B
    eq_value_b = implied_ev_b - NET_DEBT_B  # less net debt = equity value in $B
    target_price = eq_value_b * 1000 / SHARES_MM  # eq_value_b * 1000 = $M, / shares in M = $/share
    upside = (target_price - PRICE) / PRICE * 100
    weighted = target_price * s["weight"]
    s["terminal_rev"] = terminal_rev
    s["terminal_fcf"] = terminal_fcf
    s["implied_ev_b"] = implied_ev_b
    s["eq_value_b"] = eq_value_b
    s["target_price"] = round(target_price, 2)
    s["upside_pct"] = round(upside, 1)
    s["weighted_value"] = round(weighted, 2)
    print(f"{name}: Target=${target_price:.2f}, Upside={upside:.1f}%, Weighted=${weighted:.2f}")

# Total probability-weighted FV
total_fv = sum(s["weighted_value"] for s in SCENARIOS.values())
total_upside = (total_fv - PRICE) / PRICE * 100
print(f"Total Probability-Weighted FV: ${total_fv:.2f}")
print(f"Total Upside: {total_upside:.1f}%")

# ═════════════════════════════════════════════════════════
# SHEET 1: Valuation
# ═════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Valuation"

# Title block (row 1 merged)
ws1.merge_cells("A1:F1")
ws1.cell(row=1, column=1, value="Intercontinental Exchange, Inc. (ICE) — Valuation Summary")
ws1["A1"].font = title_font

title_data = [
    ("Company", "Intercontinental Exchange, Inc."),
    ("Date", "2026-06-30"),
    ("Ticker", "NYSE: ICE"),
    ("Price", f"${PRICE:.2f}"),
    ("Shares Outstanding", f"{SHARES_MM:.0f}M"),
    ("Market Cap", f"${MC_B:.2f}B"),
    ("Enterprise Value", f"${EV_B:.2f}B"),
    ("Primary Valuation Lens", "FCF yield + EV/EBITDA multiples"),
    ("Current Stance", "Watch — strong cash flow engine, trading at premium but justified quality premium"),
]
for i, (k, v) in enumerate(title_data):
    ws1.cell(row=i+2, column=1, value=k)
    ws1.cell(row=i+2, column=1).font = bold
    ws1.cell(row=i+2, column=2, value=v)

# Valuation metrics table
row = len(title_data) + 4
ws1.cell(row=row, column=1, value="Valuation Metric")
ws1.cell(row=row, column=2, value="Value")
ws1.cell(row=row, column=3, value="Comment")
style_header(ws1, row, 3)

metrics = [
    ("Trailing P/E", f"{TRAILING_PE:.2f}x", "TTM EPS $6.87 on $123.11 price. Moderate — inline with financial data peers."),
    ("Forward P/E (FY26)", f"{FWD_PE:.2f}x", "Based on FY26 consensus EPS of $8.14. Not materially compressed vs trailing."),
    ("P/S (TTM)", f"{P_SALES:.2f}x", "TTM revenue $13.08B. Premium reflects exchange/data moat."),
    ("P/FCF (TTM)", f"{MC_B/(FCF['TTM']/1000):.1f}x", f"FCF TTM $4.24B. Market cap $69.5B / FCF $4.24B."),
    ("EV/FCF (TTM)", f"{EV_FCF:.1f}x", f"EV $86.55B / FCF $4.24B. High — reflects regulated cash flow stability."),
    ("EV/Revenue", f"{EV_REVENUE:.2f}x", "Enterprise value relative to TTM sales."),
    ("EV/EBITDA", f"{EV_EBITDA:.2f}x", f"EV $86.55B / EBITDA TTM $7.53B. Premium but justified for exchange oligopolist."),
    ("P/B (Mrq)", f"{P_BOOK:.2f}x", "Relatively low — negative tangible book due to massive goodwill/intangibles from acquisitions."),
    ("PEG (5yr expected)", "1.91", "From Yahoo Finance. Below 2.0 suggests growth is reasonably priced."),
]

for i, (metric, val, comment) in enumerate(metrics):
    r = row + 1 + i
    ws1.cell(row=r, column=1, value=metric)
    ws1.cell(row=r, column=2, value=val)
    ws1.cell(row=r, column=3, value=comment)
    style_data(ws1, r, 3)

# ═════════════════════════════════════════════════════════
# SHEET 2: WACC
# ═════════════════════════════════════════════════════════
ws2 = wb.create_sheet("WACC")

ws2.merge_cells("A1:C1")
ws2.cell(row=1, column=1, value="ICE — WACC Calculation (CAPM)")
ws2["A1"].font = title_font

wacc_data = [
    ("Risk-Free Rate (10Y US Treasury)", f"{RISK_FREE:.3f}%", "Source: CNBC, 2026-06-30"),
    ("Equity Risk Premium", f"{ERP:.1f}%", "Standard assumption"),
    ("Beta (5Y Monthly)", f"{BETA:.2f}", "Yahoo Finance statistics"),
    ("Cost of Equity (Rf + Beta*ERP)", f"{COST_OF_EQUITY:.3f}%", f"{RISK_FREE} + {BETA}*{ERP} = {COST_OF_EQUITY:.3f}%"),
    ("Pre-tax Cost of Debt", f"{COST_OF_DEBT:.2f}%", f"Interest expense $800M / Total debt $20.28B"),
    ("Effective Tax Rate", f"{TAX_RATE:.1f}%", f"TTM tax provision $1,186M / pretax income $5,176M"),
    ("After-tax Cost of Debt", f"{AT_COST_OF_DEBT:.3f}%", f"{COST_OF_DEBT}*(1-{TAX_RATE:.1f}%)"),
    ("Market Capitalization", f"${MC_B:.2f}B", "Yahoo Finance, 6/29/2026"),
    ("Total Debt (outstanding)", f"${TOTAL_DEBT['FY25']/1000:.2f}B", "FY2025 balance sheet"),
    ("Net Debt (EV - MC)", f"${NET_DEBT_B:.2f}B", f"${EV_B}B EV - ${MC_B}B MC"),
    ("Equity Weight (E/EV)", f"{WEIGHT_EQ*100:.1f}%", f"${MC_B}B / ${EV_B}B"),
    ("Debt Weight (D/EV)", f"{WEIGHT_DEBT*100:.1f}%", f"${NET_DEBT_B}B / ${EV_B}B"),
    ("WACC", f"{WACC:.3f}%", f"WACC = {WEIGHT_EQ:.3f}*{COST_OF_EQUITY:.3f}% + {WEIGHT_DEBT:.3f}*{AT_COST_OF_DEBT:.3f}%"),
]

row = 2
ws2.cell(row=row, column=1, value="Component")
ws2.cell(row=row, column=2, value="Value")
ws2.cell(row=row, column=3, value="Note")
style_header(ws2, row, 3)

for i, (comp, val, note) in enumerate(wacc_data):
    r = row + 1 + i
    ws2.cell(row=r, column=1, value=comp)
    ws2.cell(row=r, column=2, value=val)
    ws2.cell(row=r, column=3, value=note)
    style_data(ws2, r, 3)
    if comp == "WACC":
        ws2.cell(row=r, column=1).fill = green_fill
        ws2.cell(row=r, column=2).fill = green_fill

# ═════════════════════════════════════════════════════════
# SHEET 3: Scenarios
# ═════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Scenarios")

ws3.merge_cells("A1:P1")
ws3.cell(row=1, column=1, value="ICE — 5-Year Scenario Analysis (Bear / Base / Bull)")
ws3["A1"].font = title_font

headers = [
    "Metric", "Bear (25%)", "Base (50%)", "Bull (25%)",
    "Unit", "Note"
]
row = 2
for c, h in enumerate(headers, 1):
    ws3.cell(row=row, column=c, value=h)
style_header(ws3, row, len(headers))

scenario_rows = []
for name in ["Bear", "Base", "Bull"]:
    s = SCENARIOS[name]
    scenario_rows.append(
        (f"Revenue CAGR (5Y)", f"{s['rev_cagr_5y']:.1f}%", "—", "—")
    )

# Build rows per metric
row_data = [
    ("Revenue CAGR (5-year)", [f"{s['rev_cagr_5y']:.1f}%" for s in SCENARIOS.values()], "%", "Assumed annual revenue growth over 5 years"),
    ("Terminal Revenue", [f"${s['terminal_rev']/1000:.1f}B" for s in SCENARIOS.values()], "$M", f"TTM rev ${REV['TTM']/1000:.1f}B × (1+CAGR)^5"),
    ("Adjusted FCF Margin", [f"{s['fcs_margin']:.1f}%" for s in SCENARIOS.values()], "%", "Normalized FCF margin assumption"),
    ("Terminal FCF", [f"${s['terminal_fcf']:.0f}M" for s in SCENARIOS.values()], "$M", "Terminal revenue × FCF margin"),
    ("Exit FCF Multiple", [f"{s['exit_fcf_mult']:.0f}x" for s in SCENARIOS.values()], "x", "EV/FCF exit multiple"),
    ("Implied Enterprise Value", [f"${s['implied_ev_b']:.1f}B" for s in SCENARIOS.values()], "$B", "Terminal FCF × exit multiple"),
    ("Less: Net Debt Adj", [f"${NET_DEBT_B:.1f}B"] * 3, "$B", "EV - MC as cleanest net debt proxy"),
    ("Implied Equity Value", [f"${s['eq_value_b']:.1f}B" for s in SCENARIOS.values()], "$B", "Implied EV - net debt"),
    ("Shares Outstanding", [f"{SHARES_MM:.0f}M"] * 3, "M", "Ordinary shares FY2025"),
    ("Target Price", [f"${s['target_price']:.2f}" for s in SCENARIOS.values()], "$/share", "Equity value / shares outstanding"),
    ("Upside from Current", [f"{s['upside_pct']:.1f}%" for s in SCENARIOS.values()], "%", f"vs. current price ${PRICE:.2f}"),
    ("Scenario Weight", ["25%", "50%", "25%"], "wt", "Probability weights"),
    ("Weighted Value/Share", [f"${s['weighted_value']:.2f}" for s in SCENARIOS.values()], "$/share", "Target price × weight"),
]

for i, (metric, vals, unit, note) in enumerate(row_data):
    r = row + 1 + i
    ws3.cell(row=r, column=1, value=metric)
    for j, v in enumerate(vals):
        ws3.cell(row=r, column=j+2, value=v)
    ws3.cell(row=r, column=5, value=unit)
    ws3.cell(row=r, column=6, value=note)
    style_data(ws3, r, 6)

# Totals row
total_row = row + 1 + len(row_data)
ws3.cell(row=total_row, column=1, value="Total Probability-Weighted FV").font = bold
ws3.cell(row=total_row, column=6, value=f"${total_fv:.2f}/share — {total_upside:.1f}% upside from ${PRICE:.2f}").font = bold
ws3.cell(row=total_row, column=6).fill = green_fill

# ═════════════════════════════════════════════════════════
# SHEET 4: Actuals Source Audit
# ═════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Actuals Source Audit")

ws4.merge_cells("A1:E1")
ws4.cell(row=1, column=1, value="ICE — Data Source Audit")
ws4["A1"].font = title_font

headers4 = ["Data Point", "Value", "Source URL", "Date", "Notes"]
row = 2
for c, h in enumerate(headers4, 1):
    ws4.cell(row=row, column=c, value=h)
style_header(ws4, row, 5)

audit_data = [
    # Price & market data
    ("Stock Price (close)", f"${PRICE:.2f}", "finance.yahoo.com/quote/ICE/", "2026-06-30", "Last regular session close"),
    ("Market Cap", f"${MC_B:.2f}B", "finance.yahoo.com/quote/ICE/ (Statistics)", "2026-06-29", "Intraday market cap from Yahoo stats"),
    ("Enterprise Value", f"${EV_B:.2f}B", "finance.yahoo.com/quote/ICE/ (Statistics)", "2026-06-29", "Yahoo Finance calculated EV"),
    ("Shares Outstanding", f"{SHARES_MM:.0f}M", "finance.yahoo.com/quote/ICE/balance-sheet/", "2025-12-31", "Ordinary shares number from FY2025 BS"),
    ("Diluted Shares (TTM avg)", "573.3M", "finance.yahoo.com/quote/ICE/financials/", "TTM", "Diluted average shares from income statement"),
    ("52-Week Range", "$121.79 - $189.35", "finance.yahoo.com/quote/ICE/", "2026-06-30", "Stock declined ~35% from 52-week high"),
    ("Beta (5Y monthly)", f"{BETA:.2f}", "finance.yahoo.com/quote/ICE/ (Statistics)", "2026-06-29", ""),
    # Income statement
    ("Revenue TTM", "$13,077M", "finance.yahoo.com/quote/ICE/financials/", "TTM", "Trailing twelve months"),
    ("Revenue FY2025", "$12,640M", "finance.yahoo.com/quote/ICE/financials/", "2025-12-31", "+7.5% YoY from FY2024"),
    ("Revenue FY2024", "$11,761M", "finance.yahoo.com/quote/ICE/financials/", "2024-12-31", "+18.8% YoY — BlackKnight integration"),
    ("Gross Profit TTM", "$7,553M", "finance.yahoo.com/quote/ICE/financials/", "TTM", "Gross margin 57.7%"),
    ("Operating Income TTM", "$5,452M", "finance.yahoo.com/quote/ICE/financials/", "TTM", "Op margin 41.7%"),
    ("Net Income TTM", "$3,931M", "finance.yahoo.com/quote/ICE/financials/", "TTM", "Profit margin 30.1%"),
    ("Diluted EPS TTM", "$6.87", "finance.yahoo.com/quote/ICE/financials/", "TTM", ""),
    ("EBITDA TTM", "$7,531M", "finance.yahoo.com/quote/ICE/financials/", "TTM", ""),
    ("EBITDA FY2025", "$6,709M", "finance.yahoo.com/quote/ICE/financials/", "2025-12-31", ""),
    # Balance sheet
    ("Total Assets FY2025", "$136,887M", "finance.yahoo.com/quote/ICE/balance-sheet/", "2025-12-31", "Down from $194,338M FY2022 — clearing balance restructure"),
    ("Total Debt FY2025", "$20,279M", "finance.yahoo.com/quote/ICE/balance-sheet/", "2025-12-31", ""),
    ("Common Stock Equity FY2025", "$28,915M", "finance.yahoo.com/quote/ICE/balance-sheet/", "2025-12-31", "Up from $22,706M FY2022"),
    ("Tangible Book Value FY2025", "-$17,084M", "finance.yahoo.com/quote/ICE/balance-sheet/", "2025-12-31", "Negative — massive goodwill/intangibles from acquisitions"),
    ("Working Capital FY2025", "$1,662M", "finance.yahoo.com/quote/ICE/balance-sheet/", "2025-12-31", ""),
    # Cash flow
    ("OCF TTM", "$5,022M", "finance.yahoo.com/quote/ICE/cash-flow/", "TTM", "Strong and growing from $3,554M FY2022"),
    ("FCF TTM", "$4,244M", "finance.yahoo.com/quote/ICE/cash-flow/", "TTM", "OCF $5,022M - Capex $778M"),
    ("Capex TTM", "$778M", "finance.yahoo.com/quote/ICE/cash-flow/", "TTM", "Moderate — exchange infrastructure"),
    ("Buybacks TTM", "$1,707M", "finance.yahoo.com/quote/ICE/cash-flow/", "TTM", "Significant capital return"),
    ("Debt Repayment TTM", "$1,250M", "finance.yahoo.com/quote/ICE/cash-flow/", "TTM", "Deleveraging in progress"),
    # Analyst estimates
    ("FY2026 Revenue Est (Avg)", "$10,980M", "finance.yahoo.com/quote/ICE/analysis/", "2026-06-30", "12 analysts. Note: appears low vs FY25 actual $12.64B — check basis"),
    ("FY2027 Revenue Est (Avg)", "$11,530M", "finance.yahoo.com/quote/ICE/analysis/", "2026-06-30", "+5.06% YoY from FY26 est"),
    ("FY2026 EPS Est (Avg)", "$8.14", "finance.yahoo.com/quote/ICE/analysis/", "2026-06-30", "14 analysts, normalized. Revision trend UPWARD"),
    ("FY2027 EPS Est (Avg)", "$8.81", "finance.yahoo.com/quote/ICE/analysis/", "2026-06-30", "15 analysts, normalized. Upward revision trend"),
    ("EPS Surprise History", "+2% to +4%", "finance.yahoo.com/quote/ICE/analysis/", "Recent 4Q", "Consistent beat pattern: +2.26%, +6.44%, +2.23%, +3.86%"),
    ("Avg Price Target", "$153.00", "finance.yahoo.com/quote/ICE/analysis/", "2026-06-30", "24.3% upside from $123.11"),
    ("High Price Target", "$251.00", "finance.yahoo.com/quote/ICE/analysis/", "2026-06-30", "103.9% upside"),
    ("Next Earnings Date", "Jul 30, 2026", "finance.yahoo.com/quote/ICE/", "2026-06-30", "Q2 FY2026 earnings"),
    ("10Y US Treasury", "4.461%", "cnbc.com/quotes/US10Y", "2026-06-30", "Source: CNBC"),
]

for i, (dp, val, src, date, notes) in enumerate(audit_data):
    r = row + 1 + i
    ws4.cell(row=r, column=1, value=dp)
    ws4.cell(row=r, column=2, value=val)
    ws4.cell(row=r, column=3, value=src)
    ws4.cell(row=r, column=4, value=date)
    ws4.cell(row=r, column=5, value=notes)
    style_data(ws4, r, 5)

# ═════════════════════════════════════════════════════════
# SHEET 5: Questions
# ═════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Questions")

ws5.merge_cells("A1:C1")
ws5.cell(row=1, column=1, value="ICE — Open Questions")
ws5["A1"].font = title_font

questions = [
    ("Q1", "Revenue estimate discrepancy", "FY2026 analyst estimate of $10.98B appears LOWER than FY2025 reported revenue of $12.64B. Is this on a pro-forma/excluding-acquisition basis? What basis do analysts use?"),
    ("Q2", "Balance sheet composition", "Total assets dropped from $194.3B (FY2022) to $136.9B (FY2025) while equity grew from $22.7B to $28.9B. The $57B asset decline was clearing balance restructure — what specifically changed?"),
    ("Q3", "Cash position ambiguity", "Cash flow statement shows End Cash Position of $78.6B (FY2025) but Yahoo Finance stats show Total Cash of only $863M. The difference is clearing collateral vs. free cash — can we confirm unrestricted cash?"),
    ("Q4", "Goodwill and intangibles", "Tangible book value is negative at -$17.1B. How much goodwill exists from BlackKnight and other acquisitions? Is any goodwill at risk of impairment?"),
    ("Q5", "Buyback trajectory", "TTM buyback of $1,707M represents significant capital return. Is this sustainable given the debt level of $20.3B? Management buyback guidance?"),
    ("Q6", "Regulatory risk", "ICE operates regulated exchanges (NYSE, futures). What regulatory changes could impact fee structures? SEC rate regulation?"),
    ("Q7", "BlackKnight integration", "BlackKnight mortgage tech acquired — how is integration going? Revenue synergies realized? Any cultural/operational friction?"),
    ("Q8", "Customer concentration", "Exchange business is inherently concentrated — top clearing members, institutional clients. What is the concentration profile?"),
    ("Q9", "Competitive dynamics", "How does ICE compete with CME (futures), Nasdaq (equities), LSEG (data)? Market share trends?"),
    ("Q10", "Interest rate sensitivity", "With $20.3B in total debt, what is the refinancing risk? What percentage of debt is floating vs. fixed rate?"),
    ("Q11", "Capital allocation priority", "ICE has been both buying back shares ($1.7B TTM) and repaying debt ($1.25B TTM). Has management indicated the relative priority going forward?"),
    ("Q12","Crypto/digital assets expansion", "ICE has been expanding into Ethereum staking and digital assets via Exchange. What revenue/FCF contribution does this represent?"),
]

row = 2
ws5.cell(row=row, column=1, value="No.")
ws5.cell(row=row, column=2, value="Topic")
ws5.cell(row=row, column=3, value="Question")
style_header(ws5, row, 3)

for i, (no, topic, q) in enumerate(questions):
    r = row + 1 + i
    ws5.cell(row=r, column=1, value=no)
    ws5.cell(row=r, column=2, value=topic)
    ws5.cell(row=r, column=3, value=q)
    style_data(ws5, r, 3)

# ═════════════════════════════════════════════════════════
# SHEET 6: Sources
# ═════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Sources")

ws6.merge_cells("A1:C1")
ws6.cell(row=1, column=1, value="ICE — Sources")
ws6["A1"].font = title_font

sources = [
    ("1", "Yahoo Finance — ICE Quote Page", "finance.yahoo.com/quote/ICE/"),
    ("2", "Yahoo Finance — ICE Income Statement", "finance.yahoo.com/quote/ICE/financials/"),
    ("3", "Yahoo Finance — ICE Balance Sheet", "finance.yahoo.com/quote/ICE/balance-sheet/"),
    ("4", "Yahoo Finance — ICE Cash Flow", "finance.yahoo.com/quote/ICE/cash-flow/"),
    ("5", "Yahoo Finance — ICE Analysis/Estimates", "finance.yahoo.com/quote/ICE/analysis/"),
    ("6", "CNBC — 10Y US Treasury", "cnbc.com/quotes/US10Y"),
    ("7", "StockAnalysis.com — ICE", "stockanalysis.com/quote/ICE/ (404 — unavailable)"),
]

row = 2
ws6.cell(row=row, column=1, value="No.")
ws6.cell(row=row, column=2, value="Source")
ws6.cell(row=row, column=3, value="URL")
style_header(ws6, row, 3)

for i, (no, name, url) in enumerate(sources):
    r = row + 1 + i
    ws6.cell(row=r, column=1, value=no)
    ws6.cell(row=r, column=2, value=name)
    ws6.cell(row=r, column=3, value=url)
    style_data(ws6, r, 3)

# ═════════════════════════════════════════════════════════
# Column widths and save
# ═════════════════════════════════════════════════════════
for ws in [ws1, ws2, ws4, ws5, ws6]:
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25

for ws in [ws3]:
    ws.column_dimensions['A'].width = 25
    for i in range(2, 8):
        ws.column_dimensions[get_column_letter(i)].width = 18

OUTPUT_DIR = "/home/refcell/dev/capital/models"
filename = f"[2026-06-30] ICE Model.xlsx"
filepath = os.path.join(OUTPUT_DIR, filename)
wb.save(filepath)
print(f"\nWorkbook saved to: {filepath}")
print(f"WACC: {WACC:.2f}%")
print(f"FV: ${total_fv:.2f}/share ({total_upside:.1f}% upside)")
