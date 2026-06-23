"""Add an INTR (Inter & Co) tab to projections/Projections.xlsx.

Clones the last existing tab (PLTR) to preserve styling/formulas, then
overwrites the input cells with Inter & Co numbers.

Inter & Co, Inc. (NASDAQ: INTR) is a Brazilian digital bank ("Inter")
headquartered in Belo Horizonte, offering banking, cards, lending,
investments, insurance and an in-app marketplace ("Inter Shop"), plus a
nascent U.S. (Miami) operation.  It reports in BRL under IFRS with a December
fiscal year.  Because the stock is quoted in USD, this model is built on a
USD-per-share basis (the basis in which Yahoo/Zacks quote INTR EPS): BRL
results translated to USD at ~5.5 BRL/USD.  Consistent with the other
December-FY tabs, the "2026" column is a FY2026 ESTIMATE built off FY2025
actuals and the Q1'26 run-rate; the model then projects FY2027-FY2031.
Basis: IFRS (GAAP) net income / diluted EPS per share (USD).  $ in millions.

The thesis: a smaller, cheaper LatAm fintech than Nu, executing a "60-30-30"
plan (60M clients, ~30% efficiency ratio, ~30% ROE by 2027).  Credit book
+36% in 2025 with +45% net income growth; ROE climbing toward the high-teens
and management-targeted 30%.  Net margin sits in the low-20s% and expands on
operating leverage; bull assumes faster client/credit growth + a re-rating
toward Nu-like multiples, bear assumes Brazil credit-cycle stress and BRL drag.

Sources (FY2025 results 2/11/2026; Q1'26 results 5/7/2026; consensus):
  - FY2025: "record results, +36% credit expansion, +45% net income growth";
    43M+ customers; net income ~R$1.4B (~$258M USD); USD EPS (ttm) ~$0.62;
    net margin ~22%; ROE ~15-16% (rising)
  - USD net revenue (ttm) ~$1.15B; market cap ~$2.38B at ~$5.4 -> ~442M shares
  - Q1'26: USD adj. EPS $0.17 (in line); revenue $463M (+48% Y/Y); ROE rising
  - FY2026E consensus (USD): EPS ~$0.76; net income ~$335M (~23% margin)
  - Analyst views: avg PT ~$9.38 (UBS $9.4, Itau BBA $10, high $12.5,
    low $4.8); Strong Buy consensus; "<15x earnings, ~66% upside" (SA)
  - Price ~$5.4 (Jun 2026, near 52-wk low of $5.36; 52-wk high $10.36)
  - Risks (bear): Brazil household-debt/credit cycle, BRL depreciation,
    competition vs Nu/incumbents, U.S. expansion execution
"""

from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "projections" / "Projections.xlsx"

# Shared inputs (USD millions; shares = diluted shares in millions)
SHARES = 442          # ~442M shares (mkt cap ~$2.38B / ~$5.39)
REV_2026 = 1450       # FY2026E USD net revenue (~+26% off ~$1.15B ttm)
NI_2026 = 335         # FY2026E IFRS net income (~23% margin, EPS ~$0.76)
NI_G_2026 = 0.30      # net income growth FY2025->FY2026E (~$258M -> ~$335M)

# Analyst IFRS net income estimates FY2026-FY2029 (USD M)
# (EPS ~$0.76 / $0.97 / $1.22 / $1.47 on ~442M shares.)
ANALYST = {"C": 335, "D": 430, "E": 540, "F": 650}

# Per-case assumptions.  Each template block is offset by 24 rows:
#   Bull base row 4, Base 28, Bear 52.  Within a block, relative rows:
#     +2 revenue, +3 revenue growth, +5 net income, +6 NI growth,
#     +8 analyst est, +10 NI margins, +14 PE low, +15 PE high, H(+0) shares.
# NOTE: revenue growth applies with a one-column lag (D6 = C6*(1+C7)), so the
# column-C growth rate drives FY2027 revenue, column-D drives FY2028, etc.
# Brazilian bank: low P/E multiples, compressing as growth decelerates.  Margin
# expands from low-20s% on operating leverage + efficiency-ratio gains.
CASES = {
    "bull": {
        "base": 4,
        # faster client/credit growth + re-rating toward Nu-like multiples
        "rev_g": {"C": 0.30, "D": 0.27, "E": 0.23, "F": 0.20, "G": 0.17, "H": 0.15},
        "margins": {"D": 0.24, "E": 0.255, "F": 0.265, "G": 0.27, "H": 0.28},
        "pe_low": {"C": 13, "D": 12, "E": 12, "F": 11, "G": 10, "H": 10},
        "pe_high": {"C": 18, "D": 17, "E": 16, "F": 15, "G": 14, "H": 13},
    },
    "base": {
        "base": 28,
        # steady 20s% growth decelerating; gradual margin expansion
        "rev_g": {"C": 0.24, "D": 0.21, "E": 0.18, "F": 0.16, "G": 0.14, "H": 0.12},
        "margins": {"D": 0.235, "E": 0.245, "F": 0.25, "G": 0.255, "H": 0.26},
        "pe_low": {"C": 11, "D": 10, "E": 10, "F": 9, "G": 8, "H": 8},
        "pe_high": {"C": 15, "D": 14, "E": 13, "F": 12, "G": 11, "H": 11},
    },
    "bear": {
        "base": 52,
        # Brazil credit-cycle stress + BRL drag cap growth and margins
        "rev_g": {"C": 0.17, "D": 0.14, "E": 0.12, "F": 0.10, "G": 0.09, "H": 0.08},
        "margins": {"D": 0.215, "E": 0.215, "F": 0.22, "G": 0.22, "H": 0.225},
        "pe_low": {"C": 8, "D": 8, "E": 7, "F": 7, "G": 6, "H": 6},
        "pe_high": {"C": 11, "D": 10, "E": 10, "F": 9, "G": 8, "H": 8},
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    src = wb[wb.sheetnames[-1]]          # clone the last tab (PLTR)
    ws = wb.copy_worksheet(src)
    ws.title = "INTR"

    # Header
    ws["B2"] = "INTR"
    ws["C2"] = datetime(2026, 6, 23)

    for cfg in CASES.values():
        b = cfg["base"]
        ws[f"H{b}"] = SHARES                      # diluted share count

        # Revenue base (FY2026E) + growth row
        ws[f"C{b + 2}"] = REV_2026
        for col, g in cfg["rev_g"].items():
            ws[f"{col}{b + 3}"] = g

        # IFRS net income base (FY2026E) + base-year growth
        ws[f"C{b + 5}"] = NI_2026
        ws[f"C{b + 6}"] = NI_G_2026

        # Analyst estimates (FY2026-FY2029)
        for col, v in ANALYST.items():
            ws[f"{col}{b + 8}"] = v

        # NI margins (D..H; C stays as formula =NI/Rev)
        for col, m in cfg["margins"].items():
            ws[f"{col}{b + 10}"] = m

        # PE low / high
        for col, v in cfg["pe_low"].items():
            ws[f"{col}{b + 14}"] = v
        for col, v in cfg["pe_high"].items():
            ws[f"{col}{b + 15}"] = v

    wb.save(XLSX)
    print(f"Added INTR tab. Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
