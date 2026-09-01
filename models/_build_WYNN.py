#!/usr/bin/env python3
"""Build the Wynn Resorts six-sheet valuation workbook."""
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT = Path(__file__).with_name("2026-09-01 Wynn Resorts Model.xlsx")
PRICE = 91.28
SHARES = 101.46  # millions
MARKET_CAP = 9.26  # billions
ENTERPRISE_VALUE = 19.50  # billions
DEBT = 12.342  # standardized total debt, billions; includes leases
CASH = 2.101  # cash and short-term investments, billions
NET_DEBT = ENTERPRISE_VALUE - MARKET_CAP
TTM_REVENUE = 7.413  # billions
TTM_EBITDA = 1.81  # billions
TTM_FCF = 0.79187  # billions
TTM_EPS = 4.17
FY26_EPS = 4.70  # adjusted diluted consensus
FY27_EPS = 5.16  # adjusted diluted consensus

# Forward P/E is the primary framework because net debt is 12.9x TTM FCF.
SCENARIOS = {
    "Bear": {"rev_cagr": 0.01, "terminal_eps": 4.60, "multiple": 14.0, "weight": 0.20},
    "Base": {"rev_cagr": 0.04, "terminal_eps": 6.50, "multiple": 18.0, "weight": 0.50},
    "Bull": {"rev_cagr": 0.065, "terminal_eps": 8.00, "multiple": 21.0, "weight": 0.30},
}
for case in SCENARIOS.values():
    case["target"] = case["terminal_eps"] * case["multiple"]
weighted_value = sum(case["target"] * case["weight"] for case in SCENARIOS.values())

assert SCENARIOS["Bear"]["target"] < PRICE
assert abs(SCENARIOS["Bull"]["target"] - 132.58) / 132.58 < 0.30

rf = 4.75
erp = 5.00
beta = 1.00
cost_equity = rf + beta * erp
pretax_debt_cost = 5.00
tax_rate = 15.81
equity_weight = MARKET_CAP / (MARKET_CAP + DEBT)
debt_weight = 1 - equity_weight
wacc = equity_weight * cost_equity + debt_weight * pretax_debt_cost * (1 - tax_rate / 100)

wb = Workbook()
navy = "17365D"
blue = "D9EAF7"
gold = "D8B34B"
gray = "E7E6E6"
white = "FFFFFF"
thin = Side(style="thin", color="A6A6A6")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def put(ws, row, col, value, *, bold=False, fill=None, color="000000", size=10, wrap=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Aptos", size=size, bold=bold, color=color)
    cell.fill = PatternFill("solid", fgColor=fill) if fill else PatternFill(fill_type=None)
    cell.border = border
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    return cell


def title(ws, text, end_col=5, subtitle=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    put(ws, 1, 1, text, bold=True, fill=navy, color=white, size=15)
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        put(ws, 2, 1, subtitle, bold=True, fill=blue, size=10)


def header(ws, row, values):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=True, fill=navy, color=white)


def widths(ws, values):
    for col, width in enumerate(values, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A3"


# 1. Valuation
ws = wb.active
ws.title = "Valuation"
title(ws, "Wynn Resorts (WYNN) — Valuation", 4, "Quote date: August 31, 2026 | Model date: September 1, 2026")
facts = [
    ("Company", "Wynn Resorts, Limited", "Integrated luxury resorts in Macau, Las Vegas and Boston; UAE development opening targeted September 2027"),
    ("Ticker", "NASDAQ: WYNN", "Consumer Discretionary / Resorts & Casinos"),
    ("Price", PRICE, "StockAnalysis August 31 close"),
    ("Shares outstanding (M)", SHARES, "StockAnalysis filing-date shares"),
    ("Market capitalization ($B)", MARKET_CAP, "StockAnalysis"),
    ("Enterprise value ($B)", ENTERPRISE_VALUE, "StockAnalysis"),
    ("Net debt proxy ($B)", NET_DEBT, "Enterprise value less market capitalization"),
    ("Primary valuation lens", "Forward P/E", "Net debt is 12.9x TTM FCF; FCF-multiple equity targets are distorted"),
    ("Stance", "Watch / Positive", "Compelling recovery and Al Marjan option value, offset by leverage and project execution risk"),
]
header(ws, 3, ["Field", "Value", "Comment"])
for row, values in enumerate(facts, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))

row = 15
header(ws, row, ["Valuation metric", "Value", "Interpretation"])
metrics = [
    ("Trailing P/E", 21.89, "GAAP TTM; usable but property win-rate and derivative movements create noise"),
    ("Forward P/E", 20.61, "FY2026 adjusted diluted consensus; primary market multiple"),
    ("P/S", 1.25, "Low relative to luxury positioning because leverage absorbs enterprise value"),
    ("P/FCF", 11.70, "Attractive equity FCF yield, but Al Marjan construction makes the run rate volatile"),
    ("EV/FCF", 24.63, "Shows debt burden; not the primary equity framework"),
    ("EV/Sales", 2.63, "Capital-intensive resort portfolio"),
    ("EV/EBITDA", 10.78, "Useful cross-check for casinos; debt/EBITDA is 6.62x"),
    ("Interest coverage", 1.92, "Thin buffer; refinancing and project spending matter"),
    ("FCF yield", 0.0855, "$791.9M TTM FCF / $9.26B market cap"),
]
for row, values in enumerate(metrics, row + 1):
    for col, value in enumerate(values, 1):
        cell = put(ws, row, col, value, bold=(col == 1))
        if values[0] == "FCF yield" and col == 2:
            cell.number_format = "0.0%"
widths(ws, [28, 24, 76, 14])

# 2. WACC
ws = wb.create_sheet("WACC")
title(ws, "Wynn Resorts — WACC", 4, "CAPM and debt-weighted cost of capital")
header(ws, 3, ["Component", "Value", "Source / formula"])
wacc_rows = [
    ("Risk-free rate", rf / 100, "U.S. Treasury 10-year CMT, August 31, 2026"),
    ("Equity risk premium", erp / 100, "Standard model assumption"),
    ("Levered beta", beta, "StockAnalysis five-year beta"),
    ("Cost of equity", cost_equity / 100, "Risk-free rate + beta × ERP"),
    ("Pre-tax cost of debt", pretax_debt_cost / 100, "Conservative normalized estimate; cash interest / debt is approximately 4.6%"),
    ("Effective tax rate", tax_rate / 100, "StockAnalysis TTM"),
    ("Market cap ($B)", MARKET_CAP, "StockAnalysis"),
    ("Total debt ($B)", DEBT, "StockAnalysis standardized total debt, including leases"),
    ("Equity weight", equity_weight, "Market cap / (market cap + debt)"),
    ("Debt weight", debt_weight, "Debt / (market cap + debt)"),
    ("After-tax debt cost", pretax_debt_cost / 100 * (1 - tax_rate / 100), "Kd × (1 − tax rate)"),
    ("WACC", wacc / 100, "E/V × Ke + D/V × Kd × (1 − t)"),
]
for row, values in enumerate(wacc_rows, 4):
    for col, value in enumerate(values, 1):
        cell = put(ws, row, col, value, bold=(values[0] == "WACC" or col == 1), fill=(gold if values[0] == "WACC" else None))
        if col == 2 and values[0] not in {"Levered beta", "Market cap ($B)", "Total debt ($B)"}:
            cell.number_format = "0.00%"
widths(ws, [30, 20, 72, 14])

# 3. Scenarios
ws = wb.create_sheet("Scenarios")
title(ws, "Wynn Resorts — Forward P/E Scenarios", 6, "Five-year terminal adjusted EPS framework; EV/EBITDA used only as a cross-check")
header(ws, 3, ["Metric", "Bear", "Base", "Bull", "Notes"])
scenario_rows = [
    ("Revenue CAGR (5Y)", *(SCENARIOS[k]["rev_cagr"] for k in ("Bear", "Base", "Bull")), "FY2026 revenue consensus is $7.49B; FY2027 visible anchor is $7.76B"),
    ("Terminal revenue ($B)", *(7.49 * (1 + SCENARIOS[k]["rev_cagr"]) ** 5 for k in ("Bear", "Base", "Bull")), "Five years from FY2026 consensus"),
    ("Terminal adjusted EPS", *(SCENARIOS[k]["terminal_eps"] for k in ("Bear", "Base", "Bull")), "Per-share earnings absorb buybacks and operating leverage"),
    ("Exit P/E", *(SCENARIOS[k]["multiple"] for k in ("Bear", "Base", "Bull")), "14x stress / 18x normalized / 21x successful Al Marjan ramp"),
    ("Target price", *(SCENARIOS[k]["target"] for k in ("Bear", "Base", "Bull")), "Terminal adjusted EPS × exit P/E"),
    ("Upside / (downside)", *((SCENARIOS[k]["target"] / PRICE - 1) for k in ("Bear", "Base", "Bull")), "Versus $91.28 close"),
    ("Probability", *(SCENARIOS[k]["weight"] for k in ("Bear", "Base", "Bull")), "20% / 50% / 30%"),
    ("Weighted value/share", *(SCENARIOS[k]["target"] * SCENARIOS[k]["weight"] for k in ("Bear", "Base", "Bull")), "Contribution to fair value"),
]
for row, values in enumerate(scenario_rows, 4):
    for col, value in enumerate(values, 1):
        cell = put(ws, row, col, value, bold=(col == 1))
        if col in (2, 3, 4) and values[0] in {"Revenue CAGR (5Y)", "Upside / (downside)", "Probability"}:
            cell.number_format = "0.0%"
        elif col in (2, 3, 4) and values[0] in {"Terminal revenue ($B)", "Terminal adjusted EPS", "Target price", "Weighted value/share"}:
            cell.number_format = "$0.00"
put(ws, 13, 1, "Probability-weighted fair value", bold=True, fill=gold)
put(ws, 13, 2, weighted_value, bold=True, fill=gold).number_format = "$0.00"
put(ws, 14, 1, "Upside from current price", bold=True, fill=gold)
put(ws, 14, 2, weighted_value / PRICE - 1, bold=True, fill=gold).number_format = "0.0%"
put(ws, 16, 1, "Framework note", bold=True, fill=gray)
put(ws, 16, 2, "Net debt of $10.24B exceeds 10× TTM FCF of $7.92B, so an FCF multiple would understate equity value after debt subtraction. Forward P/E is primary; current EV/EBITDA of 10.78x and interest coverage of 1.92x are leverage cross-checks.", fill=gray)
ws.merge_cells("B16:E16")
widths(ws, [30, 18, 18, 18, 78, 14])

# 4. Actuals Source Audit
ws = wb.create_sheet("Actuals Source Audit")
title(ws, "Wynn Resorts — Actuals Source Audit", 5, "All dollar figures are USD; financial statement figures are in millions unless noted")
header(ws, 3, ["Data point", "Value", "Source URL", "Source date", "Notes"])
sa = "https://stockanalysis.com/stocks/wynn"
audit = [
    ("Stock price", "$91.28", f"{sa}/", "2026-08-31", "Official close; after-hours excluded"),
    ("Market cap", "$9.26B", f"{sa}/statistics/", "2026-09-01", "Price-sensitive daily statistic"),
    ("Enterprise value", "$19.50B", f"{sa}/statistics/", "2026-09-01", "EV less MC implies $10.24B net debt"),
    ("Shares outstanding", "101.46M", f"{sa}/statistics/", "2026-09-01", "Down 3.24% YoY"),
    ("Revenue TTM", "$7,413M", f"{sa}/financials/", "2026-06-30", "Up 6.36%"),
    ("Operating income TTM", "$1,184M", f"{sa}/financials/", "2026-06-30", "15.98% margin"),
    ("Net income TTM", "$448.89M", f"{sa}/financials/", "2026-06-30", "GAAP"),
    ("EPS TTM", "$4.17", f"{sa}/financials/", "2026-06-30", "GAAP diluted provider figure"),
    ("Cash and investments", "$2,101M", f"{sa}/financials/balance-sheet/", "2026-06-30", "$1,573M cash plus $527M short-term investments"),
    ("Total debt", "$12,342M", f"{sa}/financials/balance-sheet/", "2026-06-30", "Standardized figure includes lease liabilities; company Q2 release reports $10.72B current and long-term debt excluding leases"),
    ("Common equity", "-$169.45M", f"{sa}/financials/balance-sheet/", "2026-06-30", "Negative from retained deficit and treasury stock; not a liquidity metric"),
    ("Operating cash flow TTM", "$1,459M", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "Cash conversion exceeds GAAP net income"),
    ("Capital expenditures TTM", "$667.32M", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "Includes development and maintenance spending"),
    ("Free cash flow TTM", "$791.87M", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "10.68% FCF margin"),
    ("FY2026 revenue consensus", "$7.49B", f"{sa}/forecast/", "2026-08-18", "18 analysts; +4.94%"),
    ("FY2026 adjusted EPS consensus", "$4.70", f"{sa}/forecast/", "2026-08-18", "Non-GAAP diluted; range $3.81-$5.68"),
    ("FY2027 adjusted EPS consensus", "$5.16", f"{sa}/forecast/", "2026-08-18", "Visible headline anchor"),
    ("Average analyst target", "$132.58", f"{sa}/forecast/", "2026-08-18", "20 analysts; range $116-$145"),
    ("Beta", "1.00", f"{sa}/statistics/", "2026-09-01", "Five-year beta"),
    ("10Y Treasury", "4.75%", "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/TextView?type=daily_treasury_yield_curve&field_tdr_date_value=2026", "2026-08-31", "U.S. Treasury CMT"),
    ("Q2 revenue", "$1,856.9M", "https://www.prnewswire.com/news-releases/wynn-resorts-limited-reports-second-quarter-2026-results-302842811.html", "2026-08-04", "+6.9% YoY"),
    ("Q2 adjusted EPS", "$1.24", "https://www.prnewswire.com/news-releases/wynn-resorts-limited-reports-second-quarter-2026-results-302842811.html", "2026-08-04", "Non-GAAP diluted; versus $1.09 prior year"),
    ("Q2 property EBITDAR", "$568.3M", "https://www.prnewswire.com/news-releases/wynn-resorts-limited-reports-second-quarter-2026-results-302842811.html", "2026-08-04", "+2.9% YoY"),
    ("Al Marjan opening target", "September 2027", "https://www.prnewswire.com/news-releases/wynn-resorts-limited-reports-second-quarter-2026-results-302842811.html", "2026-08-04", "$1.06B Wynn life-to-date cash contributions at June 30"),
]
for row, values in enumerate(audit, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
widths(ws, [30, 20, 86, 16, 66])

# 5. Questions
ws = wb.create_sheet("Questions")
title(ws, "Wynn Resorts — Open Questions", 4, "Items that can change the valuation or risk assessment")
header(ws, 3, ["#", "Question", "Why it matters", "Best evidence / next check"])
questions = [
    (1, "What drove the standardized $12.34B total-debt figure versus the company's $10.72B current-and-long-term debt disclosure?", "The $1.62B gap appears largely lease-related; valuation and fixed-charge coverage depend on consistent treatment.", "Reconcile the 10-Q debt and lease footnotes."),
    (2, "How much additional Wynn cash equity is required before Wynn Al Marjan opens in September 2027?", "The project is the largest growth option and the largest near-term capital call.", "Quarterly JV contribution and remaining construction budget."),
    (3, "What explains the reported $600M increase in Al Marjan budget referenced on the Q2 call, and who bears it?", "Cost overruns can consume buyback capacity and delay equity returns.", "Updated project budget, partner funding split and contingency."),
    (4, "What normalized mass-market win rate should be used for Wynn Palace after Q2's 29.7%?", "Q2 Palace growth benefited from unusually favorable hold; extrapolating it would overstate earnings.", "Rolling four-quarter table drop, win rate and EBITDAR."),
    (5, "Can Wynn Macau reverse declining VIP turnover and weak 1H win rates without sacrificing mix quality?", "Wynn Macau EBITDAR fell despite higher mass-market activity.", "Mass table drop, slot handle and normalized hold."),
    (6, "Why did Las Vegas EBITDAR fall 8.3% on 0.7% revenue growth in Q2?", "ADR strength did not translate to margin; labor, entertainment mix and comps need separation.", "Property expense bridge and group/convention calendar."),
    (7, "Is Encore Boston's 12.2% Q2 EBITDAR decline temporary hold noise or structural competitive pressure?", "Boston's capital productivity is below the portfolio's stronger assets.", "Massachusetts gaming data and normalized hold."),
    (8, "How much of the $667M TTM capex is maintenance versus growth?", "Normalized owner earnings require separating recurring upkeep from Al Marjan and enhancement spend.", "Capex guidance and project disclosures."),
    (9, "Will buybacks continue while interest coverage is only 1.92x and Al Marjan is under construction?", "Repurchases are accretive at low prices but compete with deleveraging.", "Board authorization, debt maturities and quarterly cash allocation."),
    (10, "What is the maturity ladder and expected refinancing rate through 2029?", "A high-rate refinancing cycle can absorb operating growth.", "10-Q debt table and bond yields."),
    (11, "How should negative common equity be interpreted after $2.78B of treasury stock?", "Buybacks create accounting deficits, but the capital structure still reduces covenant and downside flexibility.", "Covenants, restricted payments and parent liquidity."),
    (12, "What is normalized stock-based compensation after the Q2 decline?", "SBC was $86M TTM and must be covered by repurchases to avoid dilution.", "Annual proxy and cash-flow statement."),
    (13, "How concentrated is premium gaming revenue among a small number of customers and junket channels?", "High-end gaming can create volatility and credit-loss risk.", "Receivables aging and jurisdictional disclosures."),
    (14, "How durable are Macau concessions and regulatory economics through the current concession term?", "Required non-gaming investment can reduce returns despite demand recovery.", "Macau concession commitments."),
    (15, "When is the next earnings release and what guidance will establish FY2027 expectations?", "The August 4 quarter is already public; the next quarter should test normalization after favorable Palace hold.", "Company investor-relations calendar."),
]
for row, values in enumerate(questions, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
widths(ws, [7, 78, 66, 58])

# 6. Sources
ws = wb.create_sheet("Sources")
title(ws, "Wynn Resorts — Sources", 4, "Public sources accessed August 31–September 1, 2026")
header(ws, 3, ["#", "Source", "URL", "Use"])
sources = [
    (1, "StockAnalysis overview", f"{sa}/", "Price, company summary, market cap and headline results"),
    (2, "StockAnalysis financials", f"{sa}/financials/", "Income statement, segments, margins and historical overview"),
    (3, "StockAnalysis balance sheet", f"{sa}/financials/balance-sheet/", "Cash, debt, assets, liabilities and equity"),
    (4, "StockAnalysis cash flow", f"{sa}/financials/cash-flow-statement/", "Operating cash flow, capex, FCF, buybacks and interest"),
    (5, "StockAnalysis statistics", f"{sa}/statistics/", "Valuation, beta, EV, leverage and analyst summary"),
    (6, "StockAnalysis forecast", f"{sa}/forecast/", "Adjusted diluted EPS, revenue consensus and targets"),
    (7, "StockAnalysis profile", f"{sa}/company/", "Segments, management and company identity"),
    (8, "Wynn Q2 2026 release", "https://www.prnewswire.com/news-releases/wynn-resorts-limited-reports-second-quarter-2026-results-302842811.html", "Property results, liquidity, buybacks and Al Marjan timing"),
    (9, "U.S. Treasury daily yield curve", "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/TextView?type=daily_treasury_yield_curve&field_tdr_date_value=2026", "Risk-free rate"),
    (10, "Wynn investor relations", "https://investors.wynnresorts.com/", "Primary company materials and future updates"),
]
for row, values in enumerate(sources, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
widths(ws, [7, 34, 100, 65])

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = sheet.dimensions

wb.save(OUT)

# Reload verification: workbook must open and contain the exact six-sheet contract.
check = load_workbook(OUT, data_only=False)
expected = ["Valuation", "WACC", "Scenarios", "Actuals Source Audit", "Questions", "Sources"]
assert check.sheetnames == expected, check.sheetnames
assert check["Scenarios"]["B13"].value == weighted_value
print(f"WACC: {wacc:.2f}%")
for name, case in SCENARIOS.items():
    print(f"{name}: ${case['target']:.2f} ({case['target'] / PRICE - 1:.1%})")
print(f"Probability-weighted fair value: ${weighted_value:.2f} ({weighted_value / PRICE - 1:.1%})")
print(f"Wrote and verified {OUT}")
