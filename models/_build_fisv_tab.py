"""Add a FISV (Fiserv) tab to projections/Projections.xlsx.

Clones the last existing tab (CRM) to preserve styling/formulas, then
overwrites the input cells with Fiserv numbers.

Fiserv (NASDAQ: FISV) basis = adjusted net income / adjusted EPS, which is how
the Street values the stock (analyst estimates and P/E multiples are quoted on
this basis). All $ figures in millions.

Sources (FY2025 10-K earnings release, Feb 10 2026; Investor Day May 14 2026):
  - FY2025 total revenue: $21,193M; adjusted net income: $4,745M; adj EPS $8.64
  - Diluted shares ~535M (Q1'26 535.4M) and falling on buybacks -> use 530M
  - 2026 guide: organic rev growth 1-3%, adjusted EPS $8.00-8.30 (down modestly)
  - 2026-2029 targets: adj revenue CAGR 4-6%, adj op margin >37% by 2029,
    adjusted EPS >$12 in 2029 (double-digit EPS growth 2027-2029)
  - Analyst adj NI est: 2026 ~4.4B, 2027 ~4.85B, 2028 ~5.46B, 2029 ~6.1B
  - Price ~$48-50 (Jun 2026); adj P/E ~6.7x (vs 5yr avg ~32x)
"""

from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "projections" / "Projections.xlsx"

# Shared inputs
SHARES = 530          # diluted shares outstanding (M), declining on buybacks
REV_2026 = 21500      # 2026E total revenue (~1.5% on 2025's 21,193)
NI_2026 = 4400        # 2026E adjusted net income (= analyst est, EPS ~$8.30)
NI_G_2026 = -0.07     # adj NI growth 2025->2026 (EPS "down modestly")

# Analyst adjusted net income estimates 2026-2029 (consensus, same all cases)
ANALYST = {"C": 4400, "D": 4850, "E": 5460, "F": 6100}

# Per-case assumptions.  Each block of the template is offset by 24 rows:
#   Bull base row 4, Base 28, Bear 52.  Within a block, relative rows:
#     +6 revenue, +7 revenue growth, +9 net income, +10 NI growth,
#     +12 analyst est, +14 NI margins, +18 PE low, +19 PE high, H(+0) shares.
CASES = {
    "bull": {
        "base": 4,
        # revenue growth D..H (C is informational only)
        "rev_g": {"C": 0.05, "D": 0.06, "E": 0.06, "F": 0.065, "G": 0.065, "H": 0.07},
        # adj net margins D..H (C is formula =NI/Rev)
        "margins": {"D": 0.21, "E": 0.22, "F": 0.23, "G": 0.24, "H": 0.25},
        "pe_low": {c: 14 for c in "CDEFGH"},
        "pe_high": {"C": 16, "D": 17, "E": 18, "F": 18, "G": 19, "H": 19},
    },
    "base": {
        "base": 28,
        "rev_g": {"C": 0.04, "D": 0.04, "E": 0.045, "F": 0.05, "G": 0.05, "H": 0.05},
        "margins": {"D": 0.20, "E": 0.20, "F": 0.21, "G": 0.21, "H": 0.22},
        "pe_low": {c: 10 for c in "CDEFGH"},
        "pe_high": {"C": 12, "D": 12, "E": 13, "F": 13, "G": 13, "H": 13},
    },
    "bear": {
        "base": 52,
        "rev_g": {"C": 0.02, "D": 0.02, "E": 0.02, "F": 0.03, "G": 0.03, "H": 0.03},
        "margins": {"D": 0.19, "E": 0.19, "F": 0.19, "G": 0.19, "H": 0.20},
        "pe_low": {c: 7 for c in "CDEFGH"},
        "pe_high": {c: 9 for c in "CDEFGH"},
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    src = wb[wb.sheetnames[-1]]          # clone the last tab (CRM)
    ws = wb.copy_worksheet(src)
    ws.title = "FISV"

    # Header
    ws["B2"] = "FISV"
    ws["C2"] = datetime(2026, 6, 20)

    for cfg in CASES.values():
        b = cfg["base"]
        ws[f"H{b}"] = SHARES                      # share count

        # Revenue base (2026E) + growth row
        ws[f"C{b + 2}"] = REV_2026
        for col, g in cfg["rev_g"].items():
            ws[f"{col}{b + 3}"] = g

        # Net income base (2026E) + base-year growth
        ws[f"C{b + 5}"] = NI_2026
        ws[f"C{b + 6}"] = NI_G_2026

        # Analyst estimates (2026-2029)
        for col, v in ANALYST.items():
            ws[f"{col}{b + 8}"] = v

        # NI margins (D..H; C stays as formula =NI/Rev)
        for col, m in cfg["margins"].items():
            ws[f"{col}{b + 10}"] = m

        # PE low / high
        for col, v in cfg["pe_low"].items():
            ws[f"{col}{b + 14}"] = v
        for col, v in cfg["pe_high"].items():
            ws[f"{col}{b + 15}"] = v

    wb.save(XLSX)
    print(f"Added FISV tab. Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
