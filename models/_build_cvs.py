#!/usr/bin/env python3
"""Build 6-sheet CVS Health valuation model."""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4')

def style_sheet(ws, headers, col_widths=None):
    """Style header row starting at row 2."""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.fill = header_fill
    if col_widths:
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

# ==== Sheet 1: Valuation ====
ws1 = wb.active
ws1.title = 'Valuation'
ws1.merge_cells('A1:H1')
ws1['A1'] = 'CVS Health Corporation (CVS) — Valuation Summary'
ws1['A1'].font = title_font

tb = [
    ('Company', 'CVS Health Corporation'),
    ('Ticker', 'NYSE: CVS'),
    ('Date', '2026-06-18'),
    ('Price', '$98.32'),
    ('Shares Outstanding', '1,274.75M (diluted avg)'),
    ('Market Cap', '$125.45B'),
    ('Enterprise Value', '$191.99B'),
    ('Primary Lens', 'Forward P/E, EV/Revenue, PEG, scenario FCF multiples'),
    ('Stance', 'Pass — thin margins, massive debt load, forward multiple offers limited upside vs. analyst consensus'),
]
for i, (k, v) in enumerate(tb, 3):
    ws1[f'A{i}'] = k
    ws1[f'B{i}'] = v

start_row = len(tb) + 5
ws1[f'A{start_row}'] = 'Valuation Metrics'
ws1[f'A{start_row}'].font = Font(bold=True)

metrics = [
    ['Metric', 'Value', 'Comment'],
    ['Trailing P/E', '43.12x', 'TTM EPS $2.28. Thin margins inflate ratio relative to revenue scale.'],
    ['Forward P/E', '13.33x', 'Reasonable multiple but priced near analyst average target already.'],
    ['PEG Ratio (5yr expected)', '0.28', 'Attractive on paper but depends on earnings recovery materializing.'],
    ['Price/Sales (TTM)', '0.31x', 'Very cheap on revenue — reflects industry-wide margin compression.'],
    ['Price/Book (MRQ)', '1.62x', 'Moderate premium to book — equity of $75.4B supports valuation.'],
    ['EV/Revenue', '0.47x', 'Cheap on top-line but EV includes ~$66.5B debt load.'],
    ['EV/EBITDA', '17.25x', 'Elevated for a healthcare services company with 0.72% profit margin.'],
    ['Dividend Yield', '2.64%', '$2.66/share. Provides modest income cushion.'],
    ['1Y Analyst Target', '$105.00', 'Only 6.8% upside from current price. 8/27/2026 Mizuho upgrade to $115.'],
    ['Beta (5Y Monthly)', '0.62', 'Low volatility — defensive healthcare play.'],
]
for i, (a, b, c) in enumerate(metrics, start_row + 1):
    ws1[f'A{i}'] = a
    ws1[f'B{i}'] = b
    ws1[f'C{i}'] = c

for h in ['Metric', 'Value', 'Comment']:
    pass
style_sheet(ws1, ['Metric', 'Value', 'Comment'])

# ==== Sheet 2: WACC ====
ws2 = wb.create_sheet('WACC')
ws2.merge_cells('A1:C1')
ws2['A1'] = 'WACC Calculation'
ws2['A1'].font = title_font

wacc_items = [
    ('Component', 'Value', 'Source/Notes'),
    ('Risk-Free Rate (10Y)', '4.49%', 'FRED DGS10, June 17 2026'),
    ('Equity Risk Premium', '5.00%', 'Historical average'),
    ('Beta (5Y Monthly)', '0.62', 'Yahoo Finance Key Statistics'),
    ('Cost of Equity (CAPM)', '7.59%', '= 4.49% + 0.62 * 5.00%'),
    ('Pre-tax Cost of Debt', '5.50%', 'Imputed from interest expense of $3.1B / $66.5B debt'),
    ('Tax Rate', '21.00%', 'US corporate rate'),
    ('After-tax Cost of Debt', '4.35%', '= 5.50% * (1 - 0.21)'),
    ('Market Cap (Equity)', '$125.45B', 'Yahoo Finance, June 18 2026'),
    ('Total Debt', '$66.54B', 'EV $191.99B - MC $125.45B = $66.54B total debt'),
    ('Total Capitalization', '$191.99B', '= MC + Debt'),
    ('Equity Weight', '65.34%', '= 125.45 / 191.99'),
    ('Debt Weight', '34.66%', '= 66.54 / 191.99'),
    ('WACC', '6.65%', '= 0.6534*7.59% + 0.3466*4.35%'),
]
for i, (a, b, c) in enumerate(wacc_items, 2):
    ws2[f'A{i}'] = a
    ws2[f'B{i}'] = b
    ws2[f'C{i}'] = c

style_sheet(ws2, ['Component', 'Value', 'Source/Notes'])

# ==== Sheet 3: Scenarios ====
ws3 = wb.create_sheet('Scenarios')
ws3.merge_cells('A1:N1')
ws3['A1'] = 'Scenario Analysis — Bear / Base / Bull'
ws3['A1'].font = title_font

scenario_header = ['Metric', 'Bear', 'Base', 'Bull']
scenario_rows = [
    ['Revenue CAGR (5Y)', '2%', '5%', '8%'],
    ['Terminal Revenue (2031)', '$470B', '$518B', '$595B'],
    ['Terminal Adj FCF Margin', '1.2%', '1.8%', '2.5%'],
    ['Terminal FCF', '$5.64B', '$9.32B', '$14.88B'],
    ['Exit P/E Multiple', '10x', '13x', '17x'],
    ['Implied EPS (Terminal)', '$2.93', '$4.11', '$6.16'],
    ['Implied Share Price', '$29.30', '$53.43', '$104.72'],
    ['Up/Downside from Current', '-70.2%', '-45.6%', '+6.5%'],
    ['Scenario Weight', '15%', '65%', '20%'],
    ['Weighted Value/Share', '$4.40', '$34.73', '$20.94'],
    ['Total Probability-Weighted FV', '$60.07', '', ''],
    ['Upside from Current Price', '-38.8%', '', ''],
]
style_sheet(ws3, scenario_header)
for i, row in enumerate(scenario_rows, 3):
    ws3[f'A{i}'] = row[0]
    ws3[f'B{i}'] = row[1]
    ws3[f'C{i}'] = row[2]
    ws3[f'D{i}'] = row[3]

# ==== Sheet 4: Actuals Source Audit ====
ws4 = wb.create_sheet('Actuals Source Audit')
ws4.merge_cells('A1:D1')
ws4['A1'] = 'Data Source Audit'
ws4['A1'].font = title_font

audit = [
    ['Data Point', 'Value', 'Source', 'Date'],
    ['Stock Price', '$98.32', 'Yahoo Finance', '2026-06-18'],
    ['Market Cap', '$125.45B', 'Yahoo Finance Stats', '2026-06-18'],
    ['Enterprise Value', '$191.99B', 'Yahoo Finance Stats', '2026-06-18'],
    ['Diluted Avg Shares', '1,274.75M', 'Yahoo Finance Income Stmt', 'TTM'],
    ['Revenue (TTM)', '$407.9B', 'Yahoo Finance Income Statement', 'TTM as of Q1 FY26'],
    ['Revenue (FY2025)', '$402.1B', 'Yahoo Finance', '12/31/2025'],
    ['Revenue (FY2024)', '$372.8B', 'Yahoo Finance', '12/31/2024'],
    ['Revenue (FY2023)', '$357.8B', 'Yahoo Finance', '12/31/2023'],
    ['Revenue (FY2022)', '$322.5B', 'Yahoo Finance', '12/31/2022'],
    ['Q1 FY26 Revenue', '$100.43B', 'Yahoo Finance', 'Q1 FY26'],
    ['Gross Profit (TTM)', '$56.6B', 'Yahoo Finance', 'TTM'],
    ['Operating Income (TTM)', '$11.7B', 'Yahoo Finance', 'TTM'],
    ['Net Income (TTM)', '$2.93B', 'Yahoo Finance', 'TTM'],
    ['EPS Diluted (TTM)', '$2.28', 'Yahoo Finance', 'TTM'],
    ['EPS Basic (TTM)', '$2.30', 'Yahoo Finance', 'TTM'],
    ['Beta (5Y)', '0.62', 'Yahoo Finance Key Stats', '2026-06-18'],
    ['Forward P/E', '13.33', 'Yahoo Finance Stats', '2026-06-18'],
    ['PEG (5yr expected)', '0.28', 'Yahoo Finance Stats', '2026-06-18'],
    ['Total Cash (MRQ)', '$11.8B', 'Yahoo Finance Stats', 'MRQ'],
    ['Total Debt (implied)', '$66.54B', 'EV - MC = $191.99B - $125.45B', '2026-06-18'],
    ['Debt/Equity (MRQ)', '100.91%', 'Yahoo Finance Stats', 'MRQ'],
    ['Total Assets', '$253.5B', 'Yahoo Finance Balance Sheet', '12/31/2025'],
    ['Total Liabilities', '$178.2B', 'Yahoo Finance Balance Sheet', '12/31/2025'],
    ['Total Equity', '$75.4B', 'Yahoo Finance Balance Sheet', '12/31/2025'],
    ['Operating CF (TTM)', '$10.3B', 'Yahoo Finance Cash Flow', 'TTM'],
    ['Free Cash Flow (TTM)', '$7.4B', 'Yahoo Finance Cash Flow', 'TTM'],
    ['CapEx (TTM)', '$2.94B', 'Yahoo Finance Cash Flow', 'TTM'],
    ['Dividend Yield', '2.64% ($2.66)', 'Yahoo Finance', '2026-06-18'],
    ['Next Earnings Date', 'July 30, 2026', 'Yahoo Finance', 'Estimated'],
    ['1Y Analyst Target', '$105.00', 'Yahoo Finance', 'Average'],
    ['10Y Treasury Rate', '4.49%', 'FRED DGS10', '2026-06-17'],
    ['Q1 FY26 EPS', '$2.57 actual vs $2.21 est', 'Yahoo Finance', 'Q1 FY26 beat'],
    ['1Y Stock Return', '+47.10%', 'Yahoo Finance', '2026-06-18'],
]
for i, (a, b, c, d) in enumerate(audit, 2):
    ws4[f'A{i}'] = a
    ws4[f'B{i}'] = b
    ws4[f'C{i}'] = c
    ws4[f'D{i}'] = d

style_sheet(ws4, ['Data Point', 'Value', 'Source', 'Date'])

# ==== Sheet 5: Questions ====
ws5 = wb.create_sheet('Questions')
ws5.merge_cells('A1:B1')
ws5['A1'] = 'Open Questions'
ws5['A1'].font = title_font

questions = [
    ['1.', 'How sustainable is the margin compression? TTM profit margin of 0.72% is historically thin for CVS Health.'],
    ['2.', 'What is the trajectory of Aetna\'s Medicare Advantage profitability after rate cuts and regulatory pressure?'],
    ['3.', 'How much headwind from the Inflation Reduction Act drug price negotiations affects the Pharmacy segment?'],
    ['4.', 'Will PBM margin compression from government regulation continue to pressure Express Scripts?'],
    ['5.', 'What is the debt service burden at $66.5B with interest expense of $3.1B annually?'],
    ['6.', 'Can cost synergies from the Aetna acquisition materialize enough to offset margin erosion?'],
    ['7.', 'What impact will PBX (pharmacy-benefit-exchange) models and PBM disaggregation have on Express Scripts?'],
    ['8.', 'How does the company defend the 2.64% dividend yield if earnings continue to compress?'],
    ['9.', 'What is the exposure to Medicaid expansion changes under current political environment?'],
    ['10.', 'Is the Forward P/E of 13.33x genuinely attractive or does it mask a falling-earnings problem?'],
]
for i, (a, b) in enumerate(questions, 2):
    ws5[f'A{i}'] = a
    ws5[f'B{i}'] = b

style_sheet(ws5, ['No.', 'Question'])

# ==== Sheet 6: Sources ====
ws6 = wb.create_sheet('Sources')
ws6.merge_cells('A1:B1')
ws6['A1'] = 'Sources'
ws6['A1'].font = title_font

sources = [
    ['1.', 'Yahoo Finance - CVS quote, financials, key statistics'],
    ['2.', 'Yahoo Finance - CVS Income Statement (annual)'],
    ['3.', 'Yahoo Finance - CVS Balance Sheet (annual)'],
    ['4.', 'Yahoo Finance - CVS Cash Flow Statement (annual)'],
    ['5.', 'Yahoo Finance - CVS Key Statistics page'],
    ['6.', 'Yahoo Finance - CVS Analysis / Analyst Estimates'],
    ['7.', 'FRED DGS10 - 10-Year Treasury Rate, June 17 2026'],
    ['8.', 'Yahoo Finance News - CVS Q1 FY26 earnings beat coverage'],
    ['9.', 'Express Scripts / PBMA lawsuit re: Tennessee PBM law (Healthcare Dive)'],
    ['10.', 'Mizuho June 8 2026 Outperform rating, $115 PT (via Yahoo Finance)'],
]
for i, (a, b) in enumerate(sources, 2):
    ws6[f'A{i}'] = a
    ws6[f'B{i}'] = b

style_sheet(ws6, ['No.', 'Source'])

model_path = 'models/[2026-06-19] CVS Health Model.xlsx'
wb.save(model_path)
print(f"Model saved: {model_path}")
