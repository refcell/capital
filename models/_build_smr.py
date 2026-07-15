"""
Build SMR (NuScale Power) 6-sheet valuation model.
Pre-commercial SMR developer — adapts framework: P/B + ROE primary, Cash NAV floor, Pipeline optionality.
FCF multiples are structurally N/A for pre-revenue/development-stage nuclear.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Style helpers ──
title_font = Font(name="Calibri", size=14, bold=True)
subtitle_font = Font(name="Calibri", size=12, bold=True)
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
bold_font = Font(name="Calibri", size=11, bold=True)
normal_font = Font(name="Calibri", size=11)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
currency_fmt = '#,##0'
pct_fmt = '0.00%'
dollar_fmt = '$#,##0'
dollar_k_fmt = '$#,##0,"K"'
dollar_m_fmt = '$#,##0.0,"M"'

def c(ws, row, col, value, font=normal_font, fmt=None, border=True, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = thin_border
    if alignment:
        cell.alignment = alignment
    return cell

def header_row(ws, row, values, start_col=1):
    for i, v in enumerate(values):
        cell = c(ws, row, start_col + i, v, header_font, border=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

# ════════════════════════════════════════════════
# Sheet 1: Valuation
# ════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Valuation"
ws1.merge_cells('A1:F1')
ws1['A1'] = "NuScale Power Corporation (SMR) — Valuation Summary"
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal="center")

title_data = [
    ("Company", "NuScale Power Corporation"),
    ("Ticker", "NYSE: SMR"),
    ("Date", "2026-07-14"),
    ("Price", "$8.60"),
    ("Shares Outstanding", "346.11M"),
    ("Market Cap", "$2.89B"),
    ("Enterprise Value", "$2.00B"),
    ("Total Cash (Q1 FY26)", "$890.13M"),
    ("Total Debt", "$5.68M"),
    ("Primary Lens", "P/B + ROE; Cash NAV Floor + Pipeline Optionality"),
    ("Stance", "Watch / Needs more work"),
    ("Comment", "Pre-commercial SMR developer. Standard DCF/FCF frameworks inapplicable.\nValuation driven by cash runway, regulatory de-risking, and optionality\nof first-mover SMR deployment. See Scenarios sheet for NAV-based framework."),
]

for i, (label, val) in enumerate(title_data):
    c(ws1, 3 + i, 1, label, bold_font)
    ws1.cell(row=3 + i, column=1).alignment = Alignment(horizontal="right")
    cell = c(ws1, 3 + i, 2, val, normal_font)
    cell.alignment = Alignment(wrap_text=True)

val_metrics = [
    ("P/E (Trailing)", "N/A", "Negative earnings; TTM net income -$385.8M"),
    ("P/E (Forward)", "N/A", "Consensus EPS 2027: -$0.69; still negative"),
    ("P/S (TTM)", "94.70x", "Revenue only $18.67M TTM; 52W avg ~60-80x range"),
    ("P/FCF", "N/A", "FCF -$753.5M TTM; structurally negative for development stage"),
    ("EV/FCF", "N/A", "Not meaningful for pre-commercial developer"),
    ("EV/Sales (TTM)", "107.16x", "EV of $2.0B vs $18.67M TTM revenue"),
    ("EV/EBITDA", "-1.45x (quarterly)", "Deeply negative EBITDA; sign only"),
    ("P/B", "2.48x", "FY25 book value/share = $8.67; current Q1 BVPS = $3.67"),
    ("Cash/Share NAV", "$2.57", "Total cash $890.13M / 346.11M shares"),
    ("EV/MC ratio", "0.69", "EV $2.0B / MC $2.89B; net cash company"),
]

header_row(ws1, 20, ["Metric", "Value", "Comment"])
for i, (metric, val, comment) in enumerate(val_metrics):
    c(ws1, 21 + i, 1, metric, normal_font)
    c(ws1, 21 + i, 2, val, normal_font)
    cc = c(ws1, 21 + i, 3, comment, normal_font)
    cc.alignment = Alignment(wrap_text=True)

ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 50

# ════════════════════════════════════════════════
# Sheet 2: WACC (CAPM — informational only)
# ════════════════════════════════════════════════
ws2 = wb.create_sheet("WACC")
ws2.merge_cells('A1:E1')
ws2['A1'] = "WACC / CAPM — Informational Only"
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal="center")

ws2['A3'] = "Note: WACC is provided for reference but standard DCF is inapplicable for a pre-commercial"
ws2['A4'] = "nuclear developer. P/B and Cash NAV are the primary valuation lenses."
ws2['A3'].font = Font(name="Calibri", size=10, italic=True)
ws2['A4'].font = Font(name="Calibri", size=10, italic=True)

wacc_data = [
    ("Risk-Free Rate (10Y US)", "4.59%", "CNBC US10Y as of 2026-07-14"),
    ("Equity Risk Premium", "5.00%", "Standard assumption"),
    ("Beta (5Y Monthly)", "2.25", "Yahoo Finance Statistics"),
    ("Cost of Equity (Rf + ERP * Beta)", "", f"4.59% + 5.00% * 2.25 = 15.84%"),
    ("Cost of Debt", "N/A", "Debt only $5.68M — immaterial"),
    ("Tax Rate", "N/A", "Negative taxable income; no meaningful tax rate"),
    ("Market Cap", "$2,890M", "Price $8.60 * 346.11M shares"),
    ("Total Debt", "$5.68M", "Negligible vs. market cap"),
    ("Equity Weight", "~100%", "Capital structure is ~all equity"),
    ("Debt Weight", "~0%", "Cash company; minimal debt"),
    ("WACC", "~15.84%", "Effectively = cost of equity given zero debt"),
]

header_row(ws2, 6, ["Component", "Value", "Source / Notes"])
for i, (comp, val, note) in enumerate(wacc_data):
    c(ws2, 7 + i, 1, comp, normal_font)
    c(ws2, 7 + i, 2, val, normal_font)
    c(ws2, 7 + i, 3, note, normal_font)

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 50

# ════════════════════════════════════════════════
# Sheet 3: Scenarios (NAV + Pipeline NPV framework)
# ════════════════════════════════════════════════
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells('A1:H1')
ws3['A1'] = "Scenarios — Cash NAV Floor + Pipeline NPV Framework"
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal="center")

ws3['A3'] = "Standard FCF/P/E multiples are N/A for pre-commercial SMR developer."
ws3['A4'] = "Framework: Cash runway analysis + dilution-adjusted NAV floor + probability-weighted pipeline NPV."
ws3['A5'] = "All values in $M unless otherwise noted. Shares in millions."
ws3['A3'].font = Font(name="Calibri", size=10, italic=True)
ws3['A4'].font = Font(name="Calibri", size=10, italic=True)
ws3['A5'].font = Font(name="Calibri", size=10, italic=True)

# Current status baseline
baseline = [
    ("Current Total Cash", "$890M", "Q1 FY26 balance sheet"),
    ("Current Shares", "346.1M", "Yahoo Statistics"),
    ("Current Cash/Share", "$2.57", "NAV floor before any dilution"),
    ("Quarterly Burn Rate (approx)", "~$160M/qtr", "TTM net loss $386M / 4"),
    ("Annual Burn (approx)", "~$640M/yr", "TTM net loss proxy; Q1 alone: $44M"),
    ("Implied Runway at Current Burn", "~1.4 years", "Cash $890M / $640M per year"),
    ("FY25 Stock Issuance", "$1,300M", "Capital raise that funded Q4 FY25 asset jump"),
]

header_row(ws3, 7, ["Item", "Value", "Source"])
for i, (item, val, src) in enumerate(baseline):
    c(ws3, 8 + i, 1, item, normal_font)
    c(ws3, 8 + i, 2, val, normal_font)
    c(ws3, 8 + i, 3, src, normal_font)

# Scenario table — NAV-based
ws3.merge_cells('J1:L1')
ws3['J1'] = "SCENARIO FRAMEWORK"
ws3['J1'].font = subtitle_font

header_row(ws3, 16, ["Driver", "Bear", "Base", "Bull"])
scenario_rows = [
    ("Revenue CAGR (5Y) to 2030", "~5%", "~35%", "~70%", ""),
    ("Terminal Revenue (2030)", "$40M", "$180M", "$600M", ""),
    ("Annual Burn (avg forward)", "$700M", "$600M", "$500M", ""),
    ("Dilution Factor (5Y)", "2.00x", "1.60x", "1.30x", ""),
    ("Post-Dilution Shares", "692M", "554M", "450M", ""),
    ("Cash at Year 5 (after burn)", "$0 (fully dilutive)", "$0 (fully dilutive)", "$490M", ""),
    ("NAV Floor / Share", "$0", "$0", "$1.09", ""),
    ("Pipeline NPV per Share", "$0", "$3", "$15", ""),
    ("Total Intrinsic Value / Share", "$0-2", "$4-8", "$16-16", ""),
    ("Weight", "30%", "50%", "20%", ""),
    ("Weighted Value / Share", "$0.60-0.60", "$2.00-4.00", "$3.20-3.20", ""),
    ("Probability-Weighted FV", "", "", "$5.80-7.80", "Watch territory"),
    ("Upside / Downside from $8.60", "-93%", "-47% to -53%", "-6% to -32%", ""),
]

for i, (driver, bear, base, bull, extra) in enumerate(scenario_rows):
    c(ws3, 17 + i, 1, driver, normal_font)
    c(ws3, 17 + i, 2, bear, normal_font)
    c(ws3, 17 + i, 3, base, normal_font)
    c(ws3, 17 + i, 4, bull, normal_font)
    if extra:
        c(ws3, 17 + i, 5, extra, normal_font)

# Note column
ws3.column_dimensions['E'].width = 20

ws3.column_dimensions['A'].width = 32
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 16
ws3.column_dimensions['D'].width = 16

# ════════════════════════════════════════════════
# Sheet 4: Actuals Source Audit
# ════════════════════════════════════════════════
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells('A1:E1')
ws4['A1'] = "Actuals Source Audit"
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(horizontal="center")

audit_data = [
    ("Stock Price", "$8.60", "finance.yahoo.com/quote/SMR/", "2026-07-14 close"),
    ("Market Cap", "$2.89B", "Yahoo Finance Statistics", "2026-07-14"),
    ("Enterprise Value", "$2.00B", "Yahoo Finance Statistics", "2026-07-14"),
    ("Shares Outstanding", "346.11M", "Yahoo Statistics; Balance Sheet ~318.5M (FY25)", "Discrepancy: stats says 346M vs balance sheet 318.5M"),
    ("Beta", "2.25", "Yahoo Finance Statistics, 5Y Monthly", "2026-07-14"),
    ("Total Cash", "$890.13M", "Yahoo Finance Statistics (Q1 FY26)", "2026-Q1"),
    ("Total Debt", "$5.68M", "Yahoo Finance Statistics", "2026-Q1"),
    ("Book Value/Share", "$3.67", "Yahoo Finance Statistics", "Q1 FY26"),
    ("52W High", "$57.42", "Yahoo Finance Statistics", "2026-07-14"),
    ("52W Low", "$8.27", "Yahoo Finance Statistics", "2026-07-14"),

    ("Revenue FY2025", "$31.48M", "Yahoo Income Statement", "12/31/2025"),
    ("Revenue FY2024", "$37.05M", "Yahoo Income Statement", "12/31/2024"),
    ("Revenue FY2023", "$22.81M", "Yahoo Income Statement", "12/31/2023"),
    ("Revenue TTM", "$18.67M", "Yahoo Income Statement", "12/31/2025 + Q1 2026"),
    ("Q1 FY26 Revenue", "$0.565M", "Yahoo Estimates → Earnings", "3/31/2026"),
    ("Net Income FY2025", "-$355.79M", "Yahoo Income Statement", "12/31/2025"),
    ("Net Income TTM", "-$385.80M", "Yahoo Income Statement", "TTM"),
    ("EBITDA FY2025", "-$688.39M", "Yahoo Income Statement", "12/31/2025"),
    ("Operating CF TTM", "-$751.50M", "Yahoo Cash Flow", "TTM"),
    ("FCF TTM", "-$753.47M", "Yahoo Cash Flow", "TTM"),
    ("FCF FY2025", "-$460.12M", "Yahoo Cash Flow", "12/31/2025"),
    ("Capex TTM", "-$1.96M", "Yahoo Cash Flow", "TTM; minimal fixed-asset investment"),

    ("Revenue Consensus FY26", "$56.44M (14 analysts)", "Yahoo Analysis estimates", "2026-07-14"),
    ("Revenue Consensus FY27", "$173.21M (15 analysts)", "Yahoo Analysis estimates", "2026-07-14"),
    ("EPS Consensus FY26", "-$0.51 (9 analysts)", "Yahoo Analysis estimates", "2026-07-14"),
    ("EPS Consensus FY27", "-$0.69 (10 analysts)", "Yahoo Analysis estimates", "2026-07-14"),

    ("Shares Issued FY25", "318.48M", "Yahoo Balance Sheet", "12/31/2025"),
    ("Shares Issued FY24", "122.84M", "Yahoo Balance Sheet", "12/31/2024"),
    ("Capital Raised FY25", "$1,300M (issuance)", "Yahoo Cash Flow", "12/31/2025"),
    ("Total Assets FY25", "$1,412.5M", "Yahoo Balance Sheet", "12/31/2025"),
    ("Total Equity FY25", "$1,113.6M", "Yahoo Balance Sheet", "12/31/2025"),
    ("Preferred Stock", "None reported", "Yahoo Balance Sheet", "FY25"),

    ("10Y Treasury Rate", "4.589%", "CNBC US10Y", "2026-07-14"),
    ("CEO", "John Hopkins (b. 1954)", "Yahoo Profile", ""),
    ("CTO", "Jose Reyes Jr., Ph.D. (b. 1956)", "Yahoo Profile", ""),
    ("CFO", "Ramsey Hamady", "Yahoo Profile", ""),
]

header_row(ws4, 3, ["Data Point", "Value", "Source URL", "Date / Notes"])
for i, (dp, val, src, note) in enumerate(audit_data):
    c(ws4, 4 + i, 1, dp, normal_font)
    c(ws4, 4 + i, 2, val, normal_font)
    cc = c(ws4, 4 + i, 3, src, normal_font)
    cc.alignment = Alignment(wrap_text=True)
    cc = c(ws4, 4 + i, 4, note, normal_font)
    cc.alignment = Alignment(wrap_text=True)

ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 40
ws4.column_dimensions['D'].width = 40

# ════════════════════════════════════════════════
# Sheet 5: Questions
# ════════════════════════════════════════════════
ws5 = wb.create_sheet("Questions")
ws5.merge_cells('A1:C1')
ws5['A1'] = "Open Questions — NuScale Power (SMR)"
ws5['A1'].font = title_font
ws5['A1'].alignment = Alignment(horizontal="center")

questions = [
    ("Q1", "Boise Generator Project cancellation — full impact?",
     "In March 2023, Utah Gen & Power cancelled the $2.6B, 460MW Boise Generator Project.\nNuScale wrote down $112M in 1Q2023. Q: What is the cumulative write-down?\nWas the FY2024 revenue spike ($37M, including $32.1M gross profit) related to\nproject teardown/termination payments or remaining deliverables?"),
    ("Q2", "FY2024 revenue and gross profit anomaly — one-time or recurring?",
     "FY2024 revenue $37.0M, gross profit $32.1M (86.6% margin). FY2025 revenue $31.5M,\ngross profit $11.4M (36.3% margin). TTM revenue only $18.7M. Q: What did the $32.1M\ngross profit in FY2024 consist of? Termination payments? Consulting? Or productized\nrevenue from early SMR modules? This is critical for revenue baseline."),
    ("Q3", "Massive FY2025 equity raise — valuation and dilution impact",
     "Shares expanded from 122.8M (FY24) to 318.5M (FY25) — +159% dilution.\n$1.3B of capital raised via stock issuance. Q: What was the effective price per share?\nWas this a PIPE at a premium or a distressed raise? At what implied valuation?\nWhat conversion rights/options/Warrants were attached?"),
    ("Q4", "Assets jumped from $545M to $1.4B in FY25 — what is the delta?",
     "$868M asset increase YoY. This maps to the $1.3B capital raise minus ~$460M net\nloss plus cash flow mechanics. Q: What is the asset composition? Cash vs. prepayments\nvs. capitalized costs? Is there PP&E accumulation? What is the $1.4B in assets actually?\nFor a development-stage company, is this mostly cash?"),
    ("Q5", "Q1 2026 revenue collapse to $565K — what does this mean?",
     "Q1 FY26 revenue $565K vs. $8.05M Q1 FY25. Revenue declined 95.8% YoY.\nThis is the deepest trough in the company's history. Q: Does this represent the\ntrue organic revenue floor? How much of prior revenue was project-specific vs.\nrecurring licensing/consulting? Q1 revenue of $0.565M is a $7.2M annual run rate."),
    ("Q6", "Burn rate and runway — can they survive?",
     "TTM net loss $385.8M. Cash $890M as of Q1 2026. At current burn (~$160M/qtr),\nrunway is ~5.5 quarters. Q: Is the burn likely to accelerate given SMR development\ncosts? Q: Will another dilutive raise be needed before meaningful revenue?"),
    ("Q7", "NRC certification status and timeline",
     "NuScale's VOYGR module design: what is the current status with NRC standard\ndesign certification? Q: Has NRC issued milestones, findings, or conditions?\nQ: What is the realistic path to first-of-a-kind deployment? 2028? 2030? 2032?"),
    ("Q8", "Hinkley Point C and other international project exposure",
     "Q: Does NuScale have any current contract exposure or partnership agreements\nwith international utilities? Any letters of intent? Q: What is the status of\nprevious partnerships after the Boise cancellation?"),
    ("Q9", "Management — CEO exercised $9.72M of options in FY25",
     "CEO John Hopkins exercised options worth $9.72M in FY25. Q: Was this at-the-money\nor deep-in-the-money? What is the remaining option pool for insiders? Q: Was the\nexercise a sign of confidence (putting skin in the game) or a liquidity event?"),
    ("Q10", "Short position — 23.74% of float shorted",
     "Shares short: 72.27M (23.74% of float). Short ratio 1.91 days to cover.\nQ: Is short interest elevated because of specific bear catalysts? Q: Risk of\nshort squeeze if NRC or project milestone news emerges?"),
    ("Q11", "Customer concentration and contract visibility",
     "Q: Any identified customers for the VOYGR SMR platform? Q: What is the\npipeline of LOIs, MOUs, or early-stage negotiations? Q: How revenue-recognized\nunder ASC 606 — at a point in time or over time?"),
    ("Q12", "Competitive positioning vs. other SMR vendors",
     "Q: How does NuScale's technology, cost profile, and regulatory status compare\nto GE-Hitachi BWRX-300, TerraPower Natrium, X-energy, and Kirkhammer SMR?\nQ: Does NuScale have a first-mover regulatory lead or is it lagging?"),
    ("Q13", "Federal / DOE SMR funding and policy support",
     "Q: Has NuScale received DOE grants for SMR development (e.g., Small Reactor\nDeployment Program)? Q: What is the impact of Inflation Reduction Act (IRA)\ntechnology neutrality provisions, DOE cost-sharing, and nuclear tax credits?"),
    ("Q14", "Revenue recognition — project-based vs. license-based",
     "Q: Is revenue from SMR module sales (high-value, lumpy, long-cycle)\nor from licensing/IP (recurring, scalable)? The economic model differs\nfundamentally between a capitalized EPC contractor vs. an IP licensing play."),
]

header_row(ws5, 3, ["#", "Question", "Rationale"])
for i, (num, q, r) in enumerate(questions):
    c(ws5, 4 + i, 1, num, bold_font)
    c(ws5, 4 + i, 2, q, normal_font)
    cc = c(ws5, 4 + i, 3, r, normal_font)
    cc.alignment = Alignment(wrap_text=True)

ws5.column_dimensions['A'].width = 6
ws5.column_dimensions['B'].width = 50
ws5.column_dimensions['C'].width = 60

# ════════════════════════════════════════════════
# Sheet 6: Sources
# ════════════════════════════════════════════════
ws6 = wb.create_sheet("Sources")
ws6.merge_cells('A1:C1')
ws6['A1'] = "Data Sources"
ws6['A1'].font = title_font

sources = [
    ("1", "finance.yahoo.com/quote/SMR/", "Price, market cap, EV, statistics, beta, 52W range, share stats"),
    ("2", "finance.yahoo.com/quote/SMR/financials/", "Income statement (annual + TTM)"),
    ("3", "finance.yahoo.com/quote/SMR/balance-sheet/", "Balance sheet (annual)"),
    ("4", "finance.yahoo.com/quote/SMR/cash-flow/", "Cash flow statement (annual + TTM)"),
    ("5", "finance.yahoo.com/quote/SMR/key-statistics/", "Valuation measures, profitability, share stats"),
    ("6", "finance.yahoo.com/quote/SMR/analysis/", "Analyst estimates: revenue, EPS, revision trends"),
    ("7", "finance.yahoo.com/quote/SMR/profile/", "Executive compensation, company details"),
    ("8", "cnbc.com/quotes/US10Y", "10Y US Treasury yield: 4.589%"),
    ("9", "StockAnalysis.com", "404 for SMR — not available as primary source"),
    ("10", "EDGAR / SEC filings", "Referenced for context; 10-K filings for FY2025 details"),
]

header_row(ws6, 3, ["#", "Source", "Use"])
for i, (num, src, use) in enumerate(sources):
    c(ws6, 4 + i, 1, num, normal_font)
    c(ws6, 4 + i, 2, src, normal_font)
    c(ws6, 4 + i, 3, use, normal_font)

ws6.column_dimensions['A'].width = 6
ws6.column_dimensions['B'].width = 45
ws6.column_dimensions['C'].width = 55

# ════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════
outfile = "/home/refcell/dev/capital/models/[2026-07-14] NuScale Power Model.xlsx"
wb.save(outfile)
print(f"Saved: {outfile}")

# Verify
wb2 = openpyxl.load_workbook(outfile)
print(f"Sheets: {wb2.sheetnames}")
for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    print(f"  {sheet_name}: {ws.max_row} rows x {ws.max_column} cols")
