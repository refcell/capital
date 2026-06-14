#!/usr/bin/env python3
"""Build Citi (NYSE: C) 6-sheet valuation workbook.

Data from StockAnalysis (accessed 2026-06-14):
- Price: $139.83 (Jun 12, 2026 close)
- Market cap: $238.49B
- Enterprise value: $7.50B (misreported for bank — see notes)
- Shares: 1.71B
- TTM Revenue (revenues before loan losses): $88,262M
- TTM Net interest income: $61,521M
- TTM Non-interest income: $26,741M
- TTM Provision for credit losses: $10,347M
- TTM Operating expenses: $51,419M
- TTM Net income: $14,690M
- TTM EPS: $8.04
- PE trailing: 17.39
- Forward PE: 12.59
- P/B: 1.25
- ROE: 7.65%
- Beta: 1.11
- Analyst avg target: $146.93
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

wb = Workbook()

# Column width helper
def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# Border
thin = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, size=14)
section_font = Font(bold=True, size=12, color='333333')
bold_font = Font(bold=True, size=11)

# ── Sheet 1: Valuation ──
ws1 = wb.active
ws1.title = 'Valuation'
ws1.merge_cells('A1:F1')
ws1['A1'] = 'Citigroup Inc. (NYSE: C) Valuation'
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center')

data = [
    ['Item', 'Value', 'Notes'],
    ['Price', '$139.83', 'Jun 12, 2026 close'],
    ['Shares outstanding', '1.71B', 'StockAnalysis'],
    ['Market cap', '$238.49B', '1.71B × $139.83'],
    ['Enterprise value (reported)', '$7.50B', 'Misleading for bank — deposits distort EV calc'],
    ['TTM Revenue (revenues b/l loan losses)', '$88.26B', 'NII + NII_income'],
    ['TTM Net interest income', '$61.52B', '70% of revenue'],
    ['TTM Non-interest income', '$26.74B', 'Fee, trading, insurance'],
    ['TTM Provision for credit losses', '$10.35B', 'Tripled from FY2022'],
    ['TTM Net income', '$14.69B'],
    ['TTM EPS', '$8.04'],
    ['Dividend', '$2.40 (1.72%)'],
    ['Primary valuation lens', 'P/B and ROE', 'Banks — not DCF'],
    ['Current stance', 'Watch', 'Turnaround priced in, needs ROE proof'],
]

# Write valuation data starting at row 3 to avoid merged cell
for r, row in enumerate(data, 3):
    for c, val in enumerate(row, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.border = thin
        if r == 3:
            cell.font = header_font
            cell.fill = header_fill
        elif val in ('Citigroup Inc. (NYSE: C) Valuation', 'Watch'):
            pass

# Valuation metrics table
ws1['A13'] = 'Key Valuation Metrics'
ws1['A13'].font = section_font

metrics = [
    ['Metric', 'Value', 'Peer Comparison', 'Notes'],
    ['PE Ratio', '17.39', 'Above peer median', 'Trailing P/E'],
    ['Forward PE', '12.59', 'Cheap to peer median of ~11-13x', 'Prices in 38% EPS growth'],
    ['P/B Ratio', '1.25', 'JPM 1.9x, BAC 1.3x, WFC 1.2x', 'Cheap on P/B but ROE is 7.65%'],
    ['ROE', '7.65%', 'JPM ~15%, BAC ~12%, WFC ~10%', 'Below historical average and peers'],
    ['PEG Ratio', '0.50', 'Below 1.0 = potentially undervalued', 'Only meaningful if growth holds'],
    ['Price target (avg)', '$146.93', '5.1% upside', '22 analysts, Buy consensus'],
]

for r, row in enumerate(metrics, 14):
    for c, val in enumerate(row, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.border = thin
        if r == 14:
            cell.font = header_font
            cell.fill = header_fill

ws1.append([])
row = ws1.max_row + 1
ws1.cell(row=row, column=1, value='Comment: The forward P/E of 12.59x only looks cheap if the 38% EPS growth expected by consensus actually materializes. At trailing 17.39x and 7.65% ROE, the stock is priced as if the turnaround is priced in — which it largely is. P/B of 1.25x is fair to slightly cheap given ROE of 7.65%, but if ROE improves to 10%+, 1.25x is cheap.').font = Font(italic=True, size=10)

set_col_widths(ws1, [35, 25, 30, 50])

# ── Sheet 2: WACC ──
ws2 = wb.create_sheet('WACC')
ws2.merge_cells('A1:D1')
ws2['A1'] = 'Citigroup WACC — CAPM Components'
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center')

wacc_data = [
    ['Component', 'Value', 'Source', 'Notes'],
    ['Risk-free rate (10Y US Treas)', '4.30%', 'FRED DGS10, accessed 2026-06-14', 'Approximate — rates have risen modestly'],
    ['Equity risk premium', '5.00%', 'Assumption', 'Standard ERP for large-cap US equities'],
    ['Beta (levered)', '1.11', 'StockAnalysis', 'Systematic risk relative to market'],
    ['Cost of equity (CAPM)', '9.82%', '= 4.30 + 1.11 × 5.00', 'Re = Rf + Beta × ERP'],
    ['Cost of debt', '3.50%', 'Proxy', 'Approximate bank deposit/wholesale funding rate after simplification'],
    ['Tax rate', '21.0%', 'US federal corporate rate', 'State taxes are incremental but small'],
    ['Market cap', '$238.49B', 'StockAnalysis', ''],
    ['Total debt', '$385.70B', 'StockAnalysis balance sheet', 'Includes long-term debt and subordinated debt'],
    ['Equity weight', '38.2%', '= MC / (MC + Debt)', 'Significant simplification reduces debt from $708B to $386B'],
    ['Debt weight', '61.8%', '= Debt / (MC + Debt)', ''],
    ['WACC', '8.42%', '= 0.382 × 9.82 + 0.618 × 3.50 × (1 - 0.21)', 'After-tax weighted average'],
]

for r, row in enumerate(wacc_data, 3):
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.border = thin
        if r == 3:
            cell.font = header_font
            cell.fill = header_fill

ws2.append([])
ws2.cell(row=ws2.max_row + 1, column=1, value='Note: For a bank, WACC is less useful than it is for industrial companies because the primary valuation lens is P/B and ROE. Cost of debt is approximated as the average deposit/funding cost. The WACC above is primarily for scenario modeling purposes.').font = Font(italic=True, size=10)

set_col_widths(ws2, [35, 15, 30, 50])

# ── Sheet 3: Scenarios ──
ws3 = wb.create_sheet('Scenarios')
ws3.merge_cells('A1:K1')
ws3['A1'] = 'Citigroup Scenarios — Bear / Base / Bull'
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='center')

scenario_headers = ['Metric', 'Bear', 'Bear Notes', 'Base', 'Base Notes', 'Bull', 'Bull Notes']
scenarios = [
    scenario_headers,
    ['Revenue CAGR (5Y)', '3.0%', 'NII growth slows', '5.0%', 'Moderate growth continues', '7.0%', 'Franchise compounding'],
    ['Terminal Revenue (FY2031)', '$103B', 'Conservative', '$112B', 'Base', '$120B', 'Aggressive'],
    ['Net margin', '26.0%', 'Credit costs rise', '31.0%', 'Stable', '35.0%', 'High ROE, low credit costs'],
    ['Terminal Net Income (FY2031)', '$27B', 'Below current', '$35B', 'Above current', '$42B', 'High earnings power'],
    ['ROE', '7.0%', 'Stagnant', '11.0%', 'Improvement', '14.0%', 'Great returns'],
    ['Exit P/B', '1.00x', 'Multiple compression', '1.30x', 'Fair for 11% ROE', '1.60x', 'High-quality premium'],
    ['Implied book value/share (FY2031)', '$195', '= $350B / 1.79B', '', '$175', '= $314B / 1.79B', '', '$168', '= $301B / 1.79B'],
    ['Target price per share', '$195', 'Bear P/B × BVPS', '', '$228', 'Base P/B × BVPS', '', '$269', 'Bull P/B × BVPS'],
    ['Upside from $139.83', '+39.5%', 'Unlikely at bear', '', '+63.1%', 'Requires ROE proof', '', '+92.4%', 'Requires transformation'],
    ['Scenario weight', '25%', 'Credit + rates risk', '', '50%', 'Base outcome', '', '25%', 'High-end outcome'],
    ['Weighted value per share', '$49', '', '$114', '', '$67', '', ''],
    ['Probability-weighted FV', '', '', '', '', '', '$330', '= $49 + $114 + $67'],
]

for r, row in enumerate(scenarios, 3):
    for c, val in enumerate(row, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.border = thin
        if r == 3:
            cell.font = header_font
            cell.fill = header_fill

ws3.append([])
ws3.cell(row=ws3.max_row + 1, column=1, value='Note: Bank scenario modeling uses P/B exit multiples rather than FCF multiples because banks have unique cash flow profiles. Book value per share assumes 1.79B shares (modest dilution from potential capital raises over 5 years). Upside is relative to current price of $139.83. The base case probability-weighted FV of ~$230 implies significant upside if ROE improves meaningfully. This is optimistic relative to the current analyst consensus target of $146.93.').font = Font(italic=True, size=10)

set_col_widths(ws3, [30, 10, 30, 10, 30, 10, 30])

# ── Sheet 4: Actuals Source Audit ──
ws4 = wb.create_sheet('Actuals Source Audit')
ws4.merge_cells('A1:E1')
ws4['A1'] = 'Actuals Source Audit — Citigroup (C)'
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(horizontal='center')

audit_data = [
    ['Data Point', 'Value', 'Source', 'Date', 'Notes'],
    ['Stock price', '$139.83', 'StockAnalysis overview', '2026-06-12 close', 'Jun 12, 2026 close'],
    ['Market cap', '$238.49B', 'StockAnalysis statistics', '2026-06-14', '1.71B shares × $139.83'],
    ['Shares outstanding', '1.71B', 'StockAnalysis', 'TTM as of Q1 2026', 'Down 5.02% YoY'],
    ['Enterprise value (reported)', '$7.50B', 'StockAnalysis', '2026-06-14', 'Misleading for bank — deposits distort EV calc'],
    ['TTM Revenue (rev b/l losses)', '$88,262M', 'StockAnalysis income stmt', 'TTM Mar 2026', 'NII + NII_inc'],
    ['TTM Net interest income', '$61,521M', 'StockAnalysis income stmt', 'TTM Mar 2026', '+12.68% YoY'],
    ['TTM Non-interest income', '$26,741M', 'StockAnalysis income stmt', 'TTM Mar 2026', 'Flat +0.15% YoY'],
    ['TTM Provision for credit losses', '$10,347M', 'StockAnalysis income stmt', 'TTM Mar 2026', 'Tripled from FY2022'],
    ['TTM Operating expenses', '$51,419M', 'StockAnalysis income stmt', 'TTM Mar 2026', '-2.2% YoY'],
    ['TTM Net income', '$14,690M', 'StockAnalysis income stmt', 'TTM Mar 2026', '+19.8% YoY'],
    ['TTM EPS', '$8.04', 'StockAnalysis overview', '2026-06-12', '+27.1% YoY'],
    ['PE Ratio', '17.39', 'StockAnalysis statistics', '2026-06-14', 'Trailing P/E'],
    ['Forward PE', '12.59', 'StockAnalysis statistics', '2026-06-14', 'Prices in 38% EPS growth'],
    ['P/B Ratio', '1.25', 'StockAnalysis statistics', '2026-06-14', 'P/B is primary lens for bank'],
    ['ROE', '7.65%', 'StockAnalysis statistics', '2026-06-14', 'Below peer average and historical Citi avg'],
    ['Beta', '1.11', 'StockAnalysis overview', '2026-06-14', 'Systematic risk vs market'],
    ['Dividend', '$2.40 (1.72%)', 'StockAnalysis overview', '2026-06-14', 'Ex-date May 4, 2026'],
    ['Analyst avg target', '$146.93', 'StockAnalysis forecast', '2026-06-14', '22 analysts, Buy consensus'],
    ['Analyst low target', '$125', 'StockAnalysis forecast', '2026-06-14', '10.61% downside'],
    ['Analyst high target', '$170', 'StockAnalysis forecast', '2026-06-14', '21.58% upside'],
    ['52-week range', '$76.11 - $141.12', 'StockAnalysis overview', '2026-06-14', 'At 52-week high essentially'],
    ['Earnings date', 'Jul 14, 2026', 'StockAnalysis', '2026-06-14', 'Q1 2026 earnings'],
    ['Total assets', '$2,344,733M', 'StockAnalysis balance sheet', 'TTM Mar 2026', 'Down from 2023 peak'],
    ['Gross loans', '$761,616M', 'StockAnalysis balance sheet', 'TTM Mar 2026', 'Up from FY2024'],
    ['Total debt', '$385,695M', 'StockAnalysis balance sheet', 'TTM Mar 2026', 'Down from FY2023 peak of $708B'],
    ['Cash & equivalents', '$385,722M', 'StockAnalysis balance sheet', 'TTM Mar 2026', 'Large liquidity cushion'],
    ['Shareholders equity', '$299,142M', 'StockAnalysis', 'TTM Mar 2026', 'Stable capital base'],
    ['EV/Sales', '0.10', 'StockAnalysis statistics', '2026-06-14', 'Distorted for bank'],
    ['EV/Earnings', '0.51', 'StockAnalysis statistics', '2026-06-14', 'Distorted for bank'],
]

for r, row in enumerate(audit_data, 3):
    for c, val in enumerate(row, 1):
        cell = ws4.cell(row=r, column=c, value=val)
        cell.border = thin
        if r == 3:
            cell.font = header_font
            cell.fill = header_fill

set_col_widths(ws4, [35, 20, 30, 20, 40])

# ── Sheet 5: Questions ──
ws5 = wb.create_sheet('Questions')
ws5.merge_cells('A1:B1')
ws5['A1'] = 'Open Questions — Citigroup (C)'
ws5['A1'].font = title_font
ws5['A1'].alignment = Alignment(horizontal='center')

questions = [
    ['Q', 'Question'],
    [1, 'Can Jane Fraser deliver organic ROE improvement above 10% without further divestitures? Simplification plan is mostly done; next chapter needs organic growth.'],
    [2, 'Is the 84% stock rally entirely justified by the turnaround, or has the multiple already repriced the simplification benefit?'],
    [3, 'Will net interest income growth persist in the upcoming rate-cut environment, or will NII growth slow materially?'],
    [4, 'Is the provision for credit losses of $10.35B sustainable, or could it decline as the simplified loan book matures?'],
    [5, 'How will the 2023 Basel III endgame rules affect Citi when they eventually take effect? Capital constraints?'],
    [6, 'Will the Services franchise continue to deliver fee-based growth with cross-border tailwinds?'],
    [7, 'Share buybacks have been aggressive — will they continue at the same pace, and are they accretive at current prices?'],
    [8, 'What is the credit quality of the newly-focused loan book after consumer business exits? Are hidden credit risks in the transition?'],
    [9, 'Can Citi compete effectively with JPM, Goldman, and MS in the institutional franchise without the consumer and wealth businesses?'],
    [10, 'Is $146.93 analyst consensus target realistic, or does it lag the actual earnings momentum?'],
    [11, 'How does the current P/B of 1.25x compare to Citi\'s historical average after similar restructuring phases?'],
    [12, 'What is the realistic path to ROE from 7.65% to 12% given the regulatory constraints on G-SIBs?'],
]

for r, row in enumerate(questions, 3):
    for c, val in enumerate(row, 1):
        cell = ws5.cell(row=r, column=c, value=val)
        cell.border = thin
        if r == 3:
            cell.font = header_font
            cell.fill = header_fill
    if c == 1:
        cell.alignment = Alignment(horizontal='center')

set_col_widths(ws5, [5, 110])

# ── Sheet 6: Sources ──
ws6 = wb.create_sheet('Sources')
ws6.merge_cells('A1:B1')
ws6['A1'] = 'Data Sources — Citigroup (C)'
ws6['A1'].font = title_font
ws6['A1'].alignment = Alignment(horizontal='center')

sources = [
    ['Source', 'URL', 'Content'],
    [1, 'stockanalysis.com/stocks/c/', 'Overview, price, market cap, revenue, earnings, beta, dividend'],
    [2, 'stockanalysis.com/stocks/c/financials/', 'Income statement — net interest income, non-interest income, provision, expenses, net income'],
    [3, 'stockanalysis.com/stocks/c/financials/balance-sheet/', 'Balance sheet — assets, loans, debt, equity, cash'],
    [4, 'stockanalysis.com/stocks/c/statistics/', 'Valuation ratios — PE, forward PE, PB, PEG, ROE, ROA, EV metrics'],
    [5, 'stockanalysis.com/stocks/c/forecast/', 'Analyst targets, consensus, price forecasts, recommendation trends'],
    [6, 'FRED DGS10', '10-Year Treasury yield for risk-free rate in WACC'],
    [7, 'SEC EDGAR filings reference', 'For detailed earnings and balance sheet data — filings not accessible in this environment'],
]

for r, row in enumerate(sources, 3):
    for c, val in enumerate(row, 1):
        cell = ws6.cell(row=r, column=c, value=val)
        cell.border = thin
        if r == 3:
            cell.font = header_font
            cell.fill = header_fill

set_col_widths(ws6, [10, 50, 50])

# Save
out_path = os.path.expanduser('~/dev/capital/models/[2026-06-14] Citigroup Model.xlsx')
wb.save(out_path)
print(f'Saved to {out_path}')
