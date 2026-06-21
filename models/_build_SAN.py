#!/usr/bin/env python3
"""
Build SAN (Banco Santander) 6-sheet valuation model.
Data collected: Jun 18, 2026 from StockAnalysis + CNBC.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# --- Style helpers ---
hdr_font = Font(bold=True, size=12)
title_font = Font(bold=True, size=14)
small_font = Font(size=9)
note_font = Font(size=9, italic=True)
bdr = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
hdr_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
orange_fill = PatternFill(start_color='FCD5B4', end_color='FCD5B4', fill_type='solid')

EUR_MILLION = '#,##0'
PERCENT_FMT = '0.00%'
DOLLAR_FMT = '$#,##0.00'
DOLLAR_B_FMT = '$#,##0.00"B"'


def style_row(ws, row, values, fmt=None, fill=None, font=None):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = bdr
        if fmt and i > 1:
            c.number_format = fmt
        if fill:
            c.fill = fill
        if font:
            c.font = font


def style_header(ws, row, values):
    style_row(ws, row, values, fill=hdr_fill, font=hdr_font)


# ==========================================================================
# Sheet 1: Valuation
# ==========================================================================
ws1 = wb.active
ws1.title = "Valuation"

# Title block - merged row 1
ws1.merge_cells('A1:F1')
ws1['A1'] = 'Banco Santander, S.A. (SAN) — Valuation Model'
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center')

# Title block data rows starting at 2
title_data = [
    ['Company', 'Banco Santander, S.A.'],
    ['Date', '2026-06-18'],
    ['Ticker', 'NYSE: SAN'],
    ['Price (USD)', '$13.50'],
    ['Shares Outstanding', '14.54B'],
    ['Market Cap', '$194.16B'],
    ['Enterprise Value', 'N/A (bank — deposits not debt)'],
    ['Primary Lens', 'P/B and ROE, normalized P/E'],
    ['Stance', 'Watch — fairly valued at current levels'],
]
for i, row in enumerate(title_data, 2):
    style_row(ws1, i, row)
    ws1.cell(row=i, column=1).font = Font(bold=True)

# Separator
ws1['A12'] = 'Key Valuation Metrics'
ws1['A12'].font = hdr_font

# Metrics table
metrics_header = ['Metric', 'Value', 'Comment']
style_header(ws1, 13, metrics_header)

metrics = [
    ['P/E (Trailing)', 10.84, 'Below S&P 500 average (~20x); reflects bank sector discount & EU regulatory overhang'],
    ['Forward P/E', 11.08, 'Modest forward premium vs trailing — consensus sees modest FY26 growth'],
    ['P/S', 3.55, 'Standard for diversified banks; not a meaningful multiple isolation'],
    ['P/B', 1.50, 'Below historical average; implies ROE > cost of equity at current levels'],
    ['P/Tangible BV', 1.90, 'Reasonable for Spanish bank with strong NII growth'],
    ['EV/FCF', 'N/A', 'Banks do not produce FCF in the traditional sense — deposit funding offsets capex'],
    ['EV/Sales', 'N/A', 'Enterprise value not applicable to banking — deposit liabilities are operating, not financial debt'],
    ['EV/EBITDA', 'N/A', 'Same — bank economics are spread + fee, not margin-based manufacturing'],
    ['ROE', 12.89, 'Solid for EU bank; above cost of equity (~13.3%)'],
    ['ROA', 0.78, 'Healthy — large diversified banks typically run 0.5-1.0%'],
    ['Dividend Yield', '1.48%', 'Low — management returning capital via buybacks over higher dividend'],
    ['Beta', 0.95, 'Near-market sensitivity; financial sector beta'],
    ['Analyst Consensus', 'Hold', '3 analysts covering; avg target $12.16 (-9.9% vs $13.50)'],
]
for i, row in enumerate(metrics, 14):
    style_row(ws1, i, row)

# Column widths
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 70

# ==========================================================================
# Sheet 2: WACC
# ==========================================================================
ws2 = wb.create_sheet("WACC")

ws2.merge_cells('A1:D1')
ws2['A1'] = 'SAN — WACC (CAPM) Calculation'
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center')

wacc_header = ['Component', 'Value', 'Source / Notes']
style_header(ws2, 2, wacc_header)

# For banks: WACC is less relevant because equity IS the funding.
# But we still compute it for completeness.
wacc_data = [
    ['Risk-Free Rate (10Y US)', 0.04455, 'CNBC US10Y, Jun 18 2026'],
    ['Equity Risk Premium', 0.05, 'Assumed 5% (standard)'],
    ['Beta (Levered)', 0.95, 'StockAnalysis, Jun 18 2026'],
    ['Cost of Equity (Rf + Beta*ERP)', 0.04455 + 0.95 * 0.05, 'CAPM formula'],
    ['Cost of Debt', 0.025, 'European bank funding cost approximation; deposits at ~1-2% + wholesale at ~2.5%'],
    ['Tax Rate', 0.2528, 'FY2025 effective tax rate from StockAnalysis'],
    ['Market Cap (USD)', 194.16, 'Billion; $13.50 * 14.54B shares'],
    ['Total Debt (EUR)', 324.39, 'Billion; Long-term debt FY2025 BS'],
    ['Total Debt (USD est)', 324.39 * 1.1, 'Billion; ~15% FX uplift EUR->USD'],
    ['Debt + Equity', '', ''],
    ['Equity Weight', '', ''],
    ['Debt Weight', '', ''],
    ['WACC', '', ''],
]

# Compute derived values
total_cap = 194.16 + 324.39 * 1.1
eq_weight = 194.16 / total_cap
debt_weight = (324.39 * 1.1) / total_cap
cost_of_equity = 0.04455 + 0.95 * 0.05
wacc = eq_weight * cost_of_equity + debt_weight * 0.025 * (1 - 0.2528)

wacc_data[8][1] = round(324.39 * 1.1, 2)
wacc_data[9] = ['Total Capitalization (USD B)', round(total_cap, 2), 'Equity + debt in USD']
wacc_data[10] = ['Equity Weight', round(eq_weight, 4), f'{eq_weight*100:.1f}%']
wacc_data[11] = ['Debt Weight', round(debt_weight, 4), f'{debt_weight*100:.1f}%']
wacc_data[12] = ['WACC', round(wacc, 4), f'{wacc*100:.1f}%']

for i, row in enumerate(wacc_data, 3):
    style_row(ws2, i, row)
    ws2.cell(row=i, column=1).font = Font(bold=True)

ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 60

# ==========================================================================
# Sheet 3: Scenarios
# ==========================================================================
ws3 = wb.create_sheet("Scenarios")

ws3.merge_cells('A1:L1')
ws3['A1'] = 'SAN — Bear / Base / Bull Scenario Analysis (5-Year Horizon)'
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='center')

current_price = 13.50

# Scenario assumptions (5-year horizon, 2031)
# Bear: NIM compression, credit costs rise, ECB rate cuts outpace loan repricing
# Base: Gradual NII growth, stable credit costs, continued EPS growth
# Bull: NIM expansion via volume, ROE improvement to 15%+, margin re-rating

scenarios_header = [
    'Driver', 'Bear', 'Base', 'Bull'
]
style_header(ws3, 2, scenarios_header)

# FY2025 EPS was ~EUR 0.90. FY2026E consensus ~EUR 1.01. FY2027E ~EUR 1.22.
# 5Y EPS at CAGR rates.
# Bear: 3% CAGR EPS -> 1.01 * 1.03^5 ~ 1.17 EUR
# Base: 8% CAGR EPS -> 1.01 * 1.08^5 ~ 1.49 EUR
# Bull: 15% CAGR EPS -> 1.01 * 1.15^5 ~ 2.04 EUR

# P/B for scenarios
# Bear: 0.9x (stress pricing)
# Base: 1.4x (near current)
# Bull: 1.9x (premium ROE repricing)

# BVPS FY2025: EUR 6.89
# Book value grows ~ROE - dividend payout
# Approximate book value CAGR:
# Bear: 5% (high provision drag)
# Base: 9% (ROE ~13%, payout ~25%, retention drives ~9.75% BV growth)
# Bull: 14% (ROE ~15%, lower payout ratio)

bevps_2025 = 6.89
bear_bvps_y5 = bevps_2025 * (1.05 ** 5)  # 8.86
base_bvps_y5 = bevps_2025 * (1.09 ** 5)  # 10.70
bull_bvps_y5 = bevps_2025 * (1.14 ** 5)  # 13.65

bear_target_eur = bear_bvps_y5 * 0.9   # ~7.97
base_target_eur = base_bvps_y5 * 1.4   # ~14.98
bull_target_eur = bull_bvps_y5 * 1.9   # ~25.94

fx_rate = 1.1  # EUR to USD approx (June 2026)

bear_target_usd = bear_target_eur * fx_rate
base_target_usd = base_target_eur * fx_rate
bull_target_usd = bull_target_eur * fx_rate

bear_upside = (bear_target_usd / current_price) - 1
base_upside = (base_target_usd / current_price) - 1
bull_upside = (bull_target_usd / current_price) - 1

# Probability-weighted fair value
bear_wt = 0.25
base_wt = 0.50
bull_wt = 0.25
fvalue = bear_wt * bear_target_usd + base_wt * base_target_usd + bull_wt * bull_target_usd
fupside = (fvalue / current_price) - 1

scenario_rows = [
    ['Revenue CAGR (5Y)', '2%', '5%', '10%'],
    ['Terminal Revenue (2031, EUR B)', round(58.559 * 1.02**5, 1), round(58.559 * 1.05**5, 1), round(58.559 * 1.10**5, 1)],
    ['BVPS CAGR (5Y)', '5%', '9%', '14%'],
    ['Terminal BVPS (2031, EUR)', round(bear_bvps_y5, 2), round(base_bvps_y5, 2), round(bull_bvps_y5, 2)],
    ['Exit P/B Multiple', 0.9, 1.4, 1.9],
    ['Implied Price (EUR)', round(bear_target_eur, 2), round(base_target_eur, 2), round(bull_target_eur, 2)],
    ['Implied Price (USD, ~1.10 fx)', round(bear_target_usd, 2), round(base_target_usd, 2), round(bull_target_usd, 2)],
    ['Upside from $13.50', bear_upside, base_upside, bull_upside],
    ['Probability Weight', bear_wt, base_wt, bull_wt],
    ['Weighted Value/Share', round(bear_wt * bear_target_usd, 2), round(base_wt * base_target_usd, 2), round(bull_wt * bull_target_usd, 2)],
    ['---', '---', '---'],
    ['Total Probability-Weighted FV', '', '', round(fvalue, 2)],
    ['Upside from Current Price', '', '', fupside],
]

for i, row in enumerate(scenario_rows, 3):
    fill = None
    if row[0] == 'Upside from $13.50':
        for j in range(1, 4):
            pass
    style_row(ws3, i, row, fmt=PERCENT_FMT if 'upside' in row[0].lower() or 'weight' in row[0].lower() or 'cagr' in row[0].lower() else None)

# Fill the FV and upside cells specially
fs_fv = ws3.cell(row=13, column=4, value=round(fvalue, 2))
fs_fv.font = Font(bold=True, size=12)
fs_fv.fill = green_fill
fs_up = ws3.cell(row=14, column=4, value=fupside)
fs_up.number_format = PERCENT_FMT
fs_up.font = Font(bold=True, size=12)
fs_up.fill = green_fill if fupside > 0 else orange_fill

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 15

# ==========================================================================
# Sheet 4: Actuals Source Audit
# ==========================================================================
ws4 = wb.create_sheet("Actuals Source Audit")

ws4.merge_cells('A1:E1')
ws4['A1'] = 'SAN — Data Source Audit'
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(horizontal='center')

audit_header = ['Data Point', 'Value', 'Source URL', 'Date Accessed', 'Notes']
style_header(ws4, 2, audit_header)

audit_data = [
    ['Stock Price', '$13.50', 'stockanalysis.com/stocks/san/', '2026-06-18', 'Close Jun 18, 2026, 4:00 PM EDT'],
    ['After-Hours Price', '$13.48', 'stockanalysis.com/stocks/san/', '2026-06-18', 'Jun 18, 2026, 7:59 PM EDT'],
    ['Market Cap', '$194.16B', 'stockanalysis.com/stocks/san/', '2026-06-18', ''],
    ['Shares Outstanding', '14.54B', 'stockanalysis.com/stocks/san/', '2026-06-18', ''],
    ['Revenue TTM', 'EUR 54.62B', 'stockanalysis.com/stocks/san/', '2026-06-18', '+1.1% YoY'],
    ['Net Income TTM', 'EUR 17.90B', 'stockanalysis.com/stocks/san/', '2026-06-18', '+23.1% YoY'],
    ['EPS TTM', 'EUR 1.20', 'stockanalysis.com/stocks/san/', '2026-06-18', '+28.3% YoY'],
    ['P/E Ratio', '10.84', 'stockanalysis.com/stocks/san/statistics/', '2026-06-18', 'Trailing'],
    ['Forward P/E', '11.08', 'stockanalysis.com/stocks/san/statistics/', '2026-06-18', ''],
    ['P/B Ratio', '1.50', 'stockanalysis.com/stocks/san/statistics/', '2026-06-18', ''],
    ['P/TBV Ratio', '1.90', 'stockanalysis.com/stocks/san/statistics/', '2026-06-18', ''],
    ['Beta', '0.95', 'stockanalysis.com/stocks/san/', '2026-06-18', ''],
    ['ROE', '12.89%', 'stockanalysis.com/stocks/san/statistics/', '2026-06-18', ''],
    ['ROA', '0.78%', 'stockanalysis.com/stocks/san/statistics/', '2026-06-18', ''],
    ['Dividend', '$0.20 (1.48%)', 'stockanalysis.com/stocks/san/', '2026-06-18', ''],
    ['Ex-Dividend Date', 'May 4, 2026', 'stockanalysis.com/stocks/san/', '2026-06-18', ''],
    ['Earnings Date', 'Jul 29, 2026', 'stockanalysis.com/stocks/san/', '2026-06-18', ''],
    ['Analyst Consensus', 'Hold, $12.16 avg', 'stockanalysis.com/stocks/san/forecast/', '2026-06-18', '3 analysts via S&P Global'],
    ['52-Week Range', '$7.86 - $13.78', 'stockanalysis.com/stocks/san/', '2026-06-18', ''],
    ['FY2025 Revenue', 'EUR 58.559B', 'stockanalysis.com/stocks/san/financials/', '2026-06-18', ''],
    ['FY2025 Net Income', 'EUR 13.958B', 'stockanalysis.com/stocks/san/financials/', '2026-06-18', ''],
    ['FY2025 EPS Diluted', 'EUR 0.90', 'stockanalysis.com/stocks/san/financials/', '2026-06-18', ''],
    ['FY2026E Revenue', 'EUR 63.08B', 'stockanalysis.com/stocks/san/forecast/', '2026-06-18', 'Upgraded from 62.39B'],
    ['FY2026E EPS', 'EUR 1.01', 'stockanalysis.com/stocks/san/forecast/', '2026-06-18', 'Upgraded from 0.90, non-GAAP adj'],
    ['FY2027E Revenue', 'EUR 68.01B', 'stockanalysis.com/stocks/san/forecast/', '2026-06-18', 'Upgraded, +7.81%'],
    ['FY2027E EPS', 'EUR 1.22', 'stockanalysis.com/stocks/san/forecast/', '2026-06-18', 'Upgraded from 1.01, non-GAAP adj'],
    ['10Y US Treasury', '4.455%', 'cnbc.com/quotes/US10Y', '2026-06-18', ''],
    ['Total Assets', 'EUR 1,867.52B', 'stockanalysis.com/stocks/san/financials/balance-sheet/', '2026-06-18', 'FY2025'],
    ['Total Deposits', 'EUR 1,005.08B', 'stockanalysis.com/stocks/san/financials/balance-sheet/', '2026-06-18', 'FY2025'],
    ['Long-Term Debt', 'EUR 324.39B', 'stockanalysis.com/stocks/san/financials/balance-sheet/', '2026-06-18', 'FY2025'],
    ['BVPS', 'EUR 6.89', 'stockanalysis.com/stocks/san/financials/balance-sheet/', '2026-06-18', 'FY2025'],
    ['TBVPS', 'EUR 5.73', 'stockanalysis.com/stocks/san/financials/balance-sheet/', '2026-06-18', 'FY2025'],
    ['Shares Change YoY', '-3.64%', 'stockanalysis.com/stocks/san/statistics/', '2026-06-18', 'Buyback-driven reduction'],
    ['Revenue Growth FY25', '0.30%', 'stockanalysis.com/stocks/san/financials/', '2026-06-18', 'Stagnant top line'],
    ['Net Income Growth FY25', '12.14%', 'stockanalysis.com/stocks/san/financials/', '2026-06-18', 'Driven by margin/expense control'],
]

for i, row in enumerate(audit_data, 3):
    style_row(ws4, i, row, font=small_font)

ws4.column_dimensions['A'].width = 22
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 50
ws4.column_dimensions['D'].width = 14
ws4.column_dimensions['E'].width = 40

# ==========================================================================
# Sheet 5: Questions
# ==========================================================================
ws5 = wb.create_sheet("Questions")

ws5.merge_cells('A1:C1')
ws5['A1'] = 'SAN — Open Questions for Due Diligence'
ws5['A1'].font = title_font
ws5['A1'].alignment = Alignment(horizontal='center')

questions_header = ['#', 'Question', 'Notes / Context']
style_header(ws5, 2, questions_header)

questions = [
    ['1', 'Why is TTM FCF deeply negative (-EUR 20.7B) while net income is strongly positive?',
     'Bank FCF = OCF - Capex is not a meaningful metric for banks. Deposits offset loan origination. FCF margin shows -36.7%. Need to verify whether this is an accounting artifact of how StockAnalysis maps bank cash flows.'],
    ['2', 'Shares outstanding jumped from 14.89B (FY2025) to 17.57B (TTM) — what drove this?',
     'StockAnalysis shows basic shares 17.40B TTM vs 14.89B FY2025. This is +14.08% YoY — opposite to the -3.64% on the stats page. Possible discontinued operations adjustment or share restructuring.'],
    ['3', 'Revenue growth stalled at 0.30% in FY2025 after 7.56% in FY2024.',
     'Was this NIM normalization post-rate-hike cycle? NII declined -3.29% despite rate environment. What are the NIM drivers?'],
    ['4', 'How does the ECB rate path impact NII trajectory?',
     'Santander is Europe-centric; ECB rate cuts flow through to margins faster than volume growth can offset. Need Q2 2026 NIM data.'],
    ['5', 'How does the credit cost / NPL trajectory?',
     'Provision for credit losses disappeared from recent TTM/FY25 in StockAnalysis data (shows blank/dash). Is this an IFRS 9 classification issue or actually low credit costs?'],
    ['6', 'Goodwill declined from EUR 13.4B (FY2024) to EUR 11.96B (FY2025).',
     'Impairment charge or divestiture? Check for goodwill write-downs — common in EU banks.'],
    ['7', 'Minority interest is EUR 9.6B (FY2025) — what subsidiaries is this?',
     'Santander UK, Brazil, US operations likely create minority interests. Check consolidated vs non-consolidated JV treatments.'],
    ['8', 'What percentage of revenue comes from Spain vs UK vs LatAm vs US?',
     'Segment breakdown is critical for regional risk allocation (UK: Brexit, Spain: sovereign, LatAm: FX/currency, US: regulatory).'],
    ['9', 'Is the buyback program meaningful or cosmetic?',
     'Shares decreased 3.64% YoY per stats page, but TTM basic shares show increase. Data discrepancy needs resolution.'],
    ['10', 'What is the dividend policy and buyback authorization?',
     'Dividend per share rose to EUR 0.24 in FY2025 (+14.3%). Payout ratio ~27% on EPS. Room for return of capital?'],
    ['11', 'How does Santander\'s CET1 ratio compare to peers?',
     'Critical for bank valuation — capital buffers limit buybacks/dividends. Need to check Q1 2026 capital ratios.'],
    ['12', 'What is the UK divestiture status?',
     'Santander UK has been subject to UK regulatory pressure for sale/divestiture. Any material update?'],
]

for i, row in enumerate(questions, 3):
    style_row(ws5, i, row, font=small_font)
    ws5.cell(row=i, column=1).font = Font(bold=True, size=9)

ws5.column_dimensions['A'].width = 4
ws5.column_dimensions['B'].width = 45
ws5.column_dimensions['C'].width = 70

# ==========================================================================
# Sheet 6: Sources
# ==========================================================================
ws6 = wb.create_sheet("Sources")

ws6.merge_cells('A1:C1')
ws6['A1'] = 'SAN — Data Sources'
ws6['A1'].font = title_font
ws6['A1'].alignment = Alignment(horizontal='center')

sources_header = ['#', 'Source', 'URL / Reference']
style_header(ws6, 2, sources_header)

sources = [
    ['1', 'StockAnalysis — Overview', 'stockanalysis.com/stocks/san/'],
    ['2', 'StockAnalysis — Income Statement', 'stockanalysis.com/stocks/san/financials/'],
    ['3', 'StockAnalysis — Balance Sheet', 'stockanalysis.com/stocks/san/financials/balance-sheet/'],
    ['4', 'StockAnalysis — Statistics', 'stockanalysis.com/stocks/san/statistics/'],
    ['5', 'StockAnalysis — Forecast', 'stockanalysis.com/stocks/san/forecast/'],
    ['6', 'CNBC — 10Y US Treasury', 'cnbc.com/quotes/US10Y'],
    ['7', 'S&P Global Market Intelligence (via StockAnalysis)', ''],
    ['8', 'TipRanks (via StockAnalysis forecasts)', ''],
]

for i, row in enumerate(sources, 3):
    style_row(ws6, i, row, font=small_font)

ws6.column_dimensions['A'].width = 4
ws6.column_dimensions['B'].width = 45
ws6.column_dimensions['C'].width = 50

# Save
outpath = '/home/refcell/dev/capital/models/[2026-06-21] Banco Santander Model.xlsx'
wb.save(outpath)
print(f'Saved: {outpath}')
