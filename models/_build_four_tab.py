"""Add a FOUR (Shift4 Payments) tab to projections/Projections.xlsx.

Clones the last existing tab (INTR) to preserve styling/formulas, then
overwrites the input cells with Shift4 Payments numbers.

Shift4 Payments, Inc. (NYSE: FOUR) is a U.S. integrated payments / commerce
technology company ("powers the experience economy" - restaurants, hotels,
stadiums, retail), with December fiscal year, reporting in USD.  Following the
2025 acquisitions of Global Blue and (Mar 2026) Bambora, GAAP earnings are
heavily depressed by acquisition D&A and interest, so the company is valued on
a non-GAAP basis.  Consistent with the other tabs, the "2026" column is the
FY2026 guide/estimate; the model then projects FY2027-FY2031.

Basis: NON-GAAP adjusted net income / non-GAAP EPS.  The "Revenue" line is
Gross Revenue Less Network Fees (GRLNF) - Shift4's key top-line metric and the
figure management guides on (network fees are pass-through interchange), which
makes net margin meaningful (~21%) rather than the ~10% it would be on gross
revenue.  All $ figures in millions.

The thesis: a high-growth payments consolidator compounding GRLNF in the
20-30%s with mid-40s% Adjusted EBITDA margins, now de-rated to a depressed
~7x forward non-GAAP EPS after a Q4'25 miss and acquisition-integration noise.
Bull assumes successful Global Blue/Bambora integration + international ramp +
a re-rating; bear assumes integration drag, leverage/interest headwinds and
slower volume growth.

Sources (FY2025 results; Q1'26 results 5/7/2026; guidance; consensus):
  - FY2025: gross revenue $4.18B (+34% Q4 Y/Y); GRLNF $1.98B (+46%);
    Adjusted EBITDA $970M (+43%, ~49% margin); volume $209B
  - FY2026 guidance: GRLNF $2.5-2.6B (+26-31%); Adjusted EBITDA $1.165-1.215B
    (+20-25%); non-GAAP EPS $5.50-5.70
  - Q1'26: gross revenue $1,121M (+32%); GRLNF $549M; Adjusted EBITDA $234M;
    net income attrib. $15M (GAAP); non-GAAP EPS ~$0.97
  - Shares: 79.3M Class A (Apr 2026) + ~17M Class C/LLC exchangeable units;
    adj. FCF/share implies ~95-97M fully-diluted economic shares; $1B buyback
    authorized Nov 2025 (~7.7M repurchased Q4'25-Q1'26)
  - Price ~$40.65 (~May 2026, depressed post Q4 miss); avg analyst PT ~$64
    (range $50-$120); Outperform consensus
  - Risks (bear): Global Blue/Bambora integration, leverage/interest expense,
    FX, slower consumer/travel volume, payments competition
"""

from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "projections" / "Projections.xlsx"

# Shared inputs (USD millions; shares = fully-diluted economic shares in millions)
SHARES = 95           # ~95M fully-diluted economic shares (Class A + C + buybacks)
REV_2026 = 2550       # FY2026E GRLNF (midpoint of $2.5-2.6B guide)
NI_2026 = 532         # FY2026E non-GAAP net income (EPS ~$5.60 on 95M shares)
NI_G_2026 = 0.24      # non-GAAP net income growth FY2025->FY2026E

# Analyst non-GAAP net income estimates FY2026-FY2029 (USD M)
# (EPS ~$5.60 / $6.91 / $8.31 / $9.80 on ~95M shares.)
ANALYST = {"C": 532, "D": 656, "E": 789, "F": 931}

# Per-case assumptions.  Each template block is offset by 24 rows:
#   Bull base row 4, Base 28, Bear 52.  Within a block, relative rows:
#     +2 revenue, +3 revenue growth, +5 net income, +6 NI growth,
#     +8 analyst est, +10 NI margins, +14 PE low, +15 PE high, H(+0) shares.
# NOTE: revenue growth applies with a one-column lag (D6 = C6*(1+C7)), so the
# column-C growth rate drives FY2027 revenue, column-D drives FY2028, etc.
# Payments multiples are moderate; the stock is depressed (~7x fwd) so floors
# stay reasonable with only mild compression as GRLNF growth decelerates.
CASES = {
    "bull": {
        "base": 4,
        # Global Blue/Bambora integration + international ramp + re-rating
        "rev_g": {"C": 0.27, "D": 0.24, "E": 0.21, "F": 0.18, "G": 0.16, "H": 0.14},
        "margins": {"D": 0.22, "E": 0.23, "F": 0.24, "G": 0.245, "H": 0.25},
        "pe_low": {"C": 15, "D": 15, "E": 14, "F": 14, "G": 13, "H": 13},
        "pe_high": {"C": 21, "D": 20, "E": 20, "F": 19, "G": 18, "H": 18},
    },
    "base": {
        "base": 28,
        # 20s% GRLNF growth decelerating; gradual margin expansion
        "rev_g": {"C": 0.22, "D": 0.19, "E": 0.17, "F": 0.15, "G": 0.13, "H": 0.12},
        "margins": {"D": 0.215, "E": 0.22, "F": 0.225, "G": 0.23, "H": 0.235},
        "pe_low": {"C": 11, "D": 11, "E": 11, "F": 10, "G": 10, "H": 10},
        "pe_high": {"C": 15, "D": 14, "E": 14, "F": 13, "G": 13, "H": 13},
    },
    "bear": {
        "base": 52,
        # integration drag + leverage/interest + slower volume cap growth
        "rev_g": {"C": 0.16, "D": 0.14, "E": 0.12, "F": 0.10, "G": 0.09, "H": 0.08},
        "margins": {"D": 0.205, "E": 0.205, "F": 0.21, "G": 0.21, "H": 0.215},
        "pe_low": {"C": 7, "D": 7, "E": 7, "F": 7, "G": 7, "H": 7},
        "pe_high": {"C": 9, "D": 9, "E": 9, "F": 9, "G": 9, "H": 9},
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    src = wb[wb.sheetnames[-1]]          # clone the last tab (INTR)
    ws = wb.copy_worksheet(src)
    ws.title = "FOUR"

    # Header
    ws["B2"] = "FOUR"
    ws["C2"] = datetime(2026, 6, 23)

    for cfg in CASES.values():
        b = cfg["base"]
        ws[f"H{b}"] = SHARES                      # diluted share count

        # Revenue base (FY2026E GRLNF) + growth row
        ws[f"C{b + 2}"] = REV_2026
        for col, g in cfg["rev_g"].items():
            ws[f"{col}{b + 3}"] = g

        # Non-GAAP net income base (FY2026E) + base-year growth
        ws[f"C{b + 5}"] = NI_2026
        ws[f"C{b + 6}"] = NI_G_2026

        # Analyst estimates (FY2026-FY2029)
        for col, v in ANALYST.items():
            ws[f"{col}{b + 8}"] = v

        # NI margins (D..H; C stays as formula =NI/Rev)
        for col, m in cfg["margins"].items():
            ws[f"{col}{b + 10}"] = m

        # PE low / high
        for col, v in cfg["pe_low"].items():
            ws[f"{col}{b + 14}"] = v
        for col, v in cfg["pe_high"].items():
            ws[f"{col}{b + 15}"] = v

    wb.save(XLSX)
    print(f"Added FOUR tab. Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
