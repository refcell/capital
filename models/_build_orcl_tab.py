"""Add an ORCL (Oracle Corporation) tab to projections/Projections.xlsx.

Clones the last existing tab (UBER) to preserve styling/formulas, then
overwrites the input cells with Oracle numbers.

Oracle's fiscal year ends May 31, so FY2026 (ended 5/31/2026) is already
reported.  The "2026" column is therefore FY2026 actual and the model projects
FY2027-FY2031.  The Street quotes Oracle on a Non-GAAP (adjusted) basis -- the
$90B FY2027 revenue / $8.05 Non-GAAP EPS guidance and consensus forward P/E are
all Non-GAAP -- so this tab uses Non-GAAP Net Income / Non-GAAP EPS.  All $
figures in millions.

The central modeling story: Oracle is re-rating from a legacy database vendor
into an AI cloud-infrastructure (OCI) builder.  Revenue growth accelerates
sharply ($638B RPO, +363% Y/Y; FY27 guide $90B / +34%, FY28 consensus ~$130B),
but Non-GAAP net margin COMPRESSES from ~33% as the ~$55B+/yr data-center capex
drives a huge depreciation ramp (FY26 capex $55.7B, FY26 FCF -$23.7B).  The
bull case assumes margins hold up better; the bear assumes deeper compression
and slower backlog conversion.

Sources (Q4/FY2026 results 6/10/2026; FY2027 guidance; analyst consensus):
  - FY2026 revenue $67,357M (+17% Y/Y); Cloud $34.0B (+39%), OCI/IaaS $18.1B (+77%)
    GAAP net income $17,087M / GAAP EPS $5.83 (+34%)
    Non-GAAP net income avail. to common $22.2B; Non-GAAP EPS $7.63 (+27%)
    diluted shares ~2,910M; RPO $638B (+363%); op cash flow $32.0B; capex $55.7B
  - FY2025 revenue $57,399M; GAAP NI $12,443M; Non-GAAP EPS $6.01 (NI ~$17.2B)
  - FY2027 guidance: total revenue ~$90B; Non-GAAP EPS $8.05 (+18% adj.)
    Q1 FY27 guide: revenue +27%-29%; Non-GAAP EPS $1.72-$1.76
  - Analyst consensus (Non-GAAP): FY27 rev ~$89-90B, NI ~$23.3B, EPS ~$8.06;
    FY28 rev ~$130B, NI ~$28.2B, EPS ~$10.9 (Trefis); forward P/E ~22.6x FY27
  - Price ~$184 (Jun 2026); market cap ~$530B; 52-wk $134.57-$345.72;
    30 analysts consensus Strong Buy, avg 12-mo PT ~$248-269
  - Risks (bear): massive capex/leverage (raising ~$40B debt+equity in FY27,
    incl. $20B ATM), negative free cash flow, customer concentration in large
    AI prepay contracts, margin drag from depreciation
"""

from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "projections" / "Projections.xlsx"

# Shared inputs
SHARES = 2910         # diluted shares (M); ~flat (ATM equity offsets buybacks)
REV_2026 = 67357      # FY2026 actual total revenue
NI_2026 = 22200       # FY2026 actual Non-GAAP net income avail. to common
NI_G_2026 = 0.29      # Non-GAAP NI growth FY2025->FY2026 (~$17.2B -> $22.2B)

# Analyst Non-GAAP net income estimates FY2026-FY2029
# (FY27 ~$23.3B / EPS $8.05; FY28 ~$28.2B / EPS ~$9.7-10.9; FY29 extrapolated.)
ANALYST = {"C": 22200, "D": 23300, "E": 28200, "F": 33500}

# Per-case assumptions.  Each template block is offset by 24 rows:
#   Bull base row 4, Base 28, Bear 52.  Within a block, relative rows:
#     +2 revenue, +3 revenue growth, +5 net income, +6 NI growth,
#     +8 analyst est, +10 NI margins, +14 PE low, +15 PE high, H(+0) shares.
# Revenue accelerates on OCI/AI backlog conversion then decelerates.  Non-GAAP
# net margin compresses from ~33% as data-center depreciation ramps (bull holds
# margins better; bear sees deeper compression + slower backlog conversion).
CASES = {
    # NOTE: revenue growth applies with a one-column lag (D6 = C6*(1+C7)), so the
    # column-C growth rate drives FY2027 revenue, column-D drives FY2028, etc.
    "bull": {
        "base": 4,
        # backlog converts fast; OCI scales above guidance
        "rev_g": {"C": 0.36, "D": 0.45, "E": 0.35, "F": 0.27, "G": 0.21, "H": 0.18},
        "margins": {"D": 0.27, "E": 0.25, "F": 0.25, "G": 0.26, "H": 0.27},
        "pe_low": {c: 28 for c in "CDEFGH"},
        "pe_high": {c: 38 for c in "CDEFGH"},
    },
    "base": {
        "base": 28,
        # FY27 ~$90B (guidance), FY28 ~$128B (consensus), then decelerates
        "rev_g": {"C": 0.34, "D": 0.42, "E": 0.30, "F": 0.24, "G": 0.18, "H": 0.15},
        "margins": {"D": 0.26, "E": 0.22, "F": 0.22, "G": 0.225, "H": 0.23},
        "pe_low": {c: 22 for c in "CDEFGH"},
        "pe_high": {c: 30 for c in "CDEFGH"},
    },
    "bear": {
        "base": 52,
        # slower backlog conversion; depreciation drag deepens margin compression
        "rev_g": {"C": 0.30, "D": 0.30, "E": 0.20, "F": 0.15, "G": 0.12, "H": 0.10},
        "margins": {"D": 0.25, "E": 0.20, "F": 0.19, "G": 0.185, "H": 0.18},
        "pe_low": {c: 14 for c in "CDEFGH"},
        "pe_high": {c: 20 for c in "CDEFGH"},
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    src = wb[wb.sheetnames[-1]]          # clone the last tab (UBER)
    ws = wb.copy_worksheet(src)
    ws.title = "ORCL"

    # Header
    ws["B2"] = "ORCL"
    ws["C2"] = datetime(2026, 6, 21)

    for cfg in CASES.values():
        b = cfg["base"]
        ws[f"H{b}"] = SHARES                      # share count

        # Revenue base (FY2026 actual) + growth row
        ws[f"C{b + 2}"] = REV_2026
        for col, g in cfg["rev_g"].items():
            ws[f"{col}{b + 3}"] = g

        # Non-GAAP net income base (FY2026 actual) + base-year growth
        ws[f"C{b + 5}"] = NI_2026
        ws[f"C{b + 6}"] = NI_G_2026

        # Analyst estimates (FY2026-FY2029)
        for col, v in ANALYST.items():
            ws[f"{col}{b + 8}"] = v

        # NI margins (D..H; C stays as formula =NI/Rev)
        for col, m in cfg["margins"].items():
            ws[f"{col}{b + 10}"] = m

        # PE low / high
        for col, v in cfg["pe_low"].items():
            ws[f"{col}{b + 14}"] = v
        for col, v in cfg["pe_high"].items():
            ws[f"{col}{b + 15}"] = v

    wb.save(XLSX)
    print(f"Added ORCL tab. Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
