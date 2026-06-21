"""Add a QXO (QXO, Inc.) tab to projections/Projections.xlsx.

Clones the last existing tab (NOW) to preserve styling/formulas, then
overwrites the input cells with QXO numbers.

QXO is Brad Jacobs' building-products distribution roll-up (formerly a small
software company).  It acquired Beacon Roofing (Apr 2025), Kodiak Building
Partners (Apr 2026) and announced TopBuild (~$17B, expected close Q3 2026), and
targets $50B in annual revenue within a decade via M&A + organic growth.  The
company is GAAP-loss-making (acquisition amortization, transaction/transformation
costs, integration investments), so -- like the ZETA tab -- this model uses an
adjusted (non-GAAP) net income basis.  All $ figures in millions.

Sources (FY2025 results Feb 2026; Q1 2026 results May 2026; 10-Q/10-K):
  - FY2025 net sales $6,842M (Beacon consolidated from Apr 29, 2025)
    GAAP net loss $(279.4)M; GAAP diluted EPS $(0.63)
    Adjusted Net Income $362.7M (~5.3% margin); Adjusted Diluted EPS $0.34
    Adjusted diluted weighted-avg shares 727.3M
  - Q1 2026 net sales $1,730M; Adjusted Net Loss $(57.2)M; adj. EPS $(0.12)
    (industry softness + tech/sales-capacity investments); ~744M shares
  - M&A: Kodiak $2.25B (closed Apr 1 2026, ~$2.4B revenue, cash + 13.2M shares);
    TopBuild ~$17B announced (expected close Q3 2026, ~$5.4B revenue);
    Beacon standalone ~$9.8B revenue.  Combined run-rate ~$17.6B post-deals.
  - Long-term target: $50B annual revenue within ~a decade
  - Price ~$17 (Jun 2026); market cap ~$12B; ~710M common shares + preferred
  - Analyst consensus: 18 analysts, Moderate Buy, avg PT $32 (range $26-$50);
    forward P/E ~38x on depressed near-term earnings; Street 2026 adj. EPS ~$0.50
    rising sharply in 2027 as Kodiak/TopBuild contribute full years
"""

from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "projections" / "Projections.xlsx"

# Shared inputs
SHARES = 800          # diluted shares (M), blended forward incl. M&A stock issuance
REV_2026 = 13000      # FY2026E net sales (Beacon FY + Kodiak ~9mo + TopBuild ~1Q)
NI_2026 = 450         # FY2026E adjusted net income (~3.5% margin; depressed year)
NI_G_2026 = 0.24      # adjusted NI growth 2025->2026 ($363M -> $450M)

# Analyst adjusted net income estimates 2026-2029.
# (Street adj. EPS ~$0.56/$0.98/$1.44/$1.88 x 800 shares; earnings inflect up as
#  acquisitions annualize and the building-products cycle recovers.)
ANALYST = {"C": 450, "D": 780, "E": 1150, "F": 1500}

# Per-case assumptions.  Each template block is offset by 24 rows:
#   Bull base row 4, Base 28, Bear 52.  Within a block, relative rows:
#     +2 revenue, +3 revenue growth, +5 net income, +6 NI growth,
#     +8 analyst est, +10 NI margins, +14 PE low, +15 PE high, H(+0) shares.
# Revenue growth is M&A-heavy in 2027 (full year of Kodiak + TopBuild + tuck-ins),
# then decelerates as the base scales toward the $50B target.  Adjusted net margin
# expands from a thin distributor base (~3.5%) as Jacobs drives procurement, pricing
# and logistics gains (bear: integration struggles + a building-products downturn).
CASES = {
    "bull": {
        "base": 4,
        # aggressive accretive M&A + strong organic/margin execution
        "rev_g": {"C": 0.65, "D": 0.38, "E": 0.28, "F": 0.22, "G": 0.18, "H": 0.16},
        "margins": {"D": 0.055, "E": 0.065, "F": 0.075, "G": 0.08, "H": 0.085},
        "pe_low": {"C": 34, "D": 30, "E": 27, "F": 24, "G": 21, "H": 19},
        "pe_high": {"C": 44, "D": 40, "E": 36, "F": 32, "G": 28, "H": 25},
    },
    "base": {
        "base": 28,
        # steady roll-up toward $50B target with gradual margin expansion
        "rev_g": {"C": 0.54, "D": 0.30, "E": 0.22, "F": 0.18, "G": 0.15, "H": 0.13},
        "margins": {"D": 0.05, "E": 0.055, "F": 0.06, "G": 0.065, "H": 0.07},
        "pe_low": {"C": 24, "D": 22, "E": 20, "F": 18, "G": 16, "H": 15},
        "pe_high": {"C": 32, "D": 30, "E": 27, "F": 24, "G": 22, "H": 20},
    },
    "bear": {
        "base": 52,
        # slower M&A, integration drag, cyclical building-products weakness
        "rev_g": {"C": 0.40, "D": 0.20, "E": 0.13, "F": 0.10, "G": 0.08, "H": 0.07},
        "margins": {"D": 0.04, "E": 0.04, "F": 0.045, "G": 0.05, "H": 0.05},
        "pe_low": {"C": 16, "D": 14, "E": 12, "F": 11, "G": 10, "H": 9},
        "pe_high": {"C": 24, "D": 22, "E": 19, "F": 17, "G": 15, "H": 14},
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    src = wb[wb.sheetnames[-1]]          # clone the last tab (NOW)
    ws = wb.copy_worksheet(src)
    ws.title = "QXO"

    # Header
    ws["B2"] = "QXO"
    ws["C2"] = datetime(2026, 6, 21)

    for cfg in CASES.values():
        b = cfg["base"]
        ws[f"H{b}"] = SHARES                      # share count

        # Revenue base (2026E) + growth row
        ws[f"C{b + 2}"] = REV_2026
        for col, g in cfg["rev_g"].items():
            ws[f"{col}{b + 3}"] = g

        # Adjusted net income base (2026E) + base-year growth
        ws[f"C{b + 5}"] = NI_2026
        ws[f"C{b + 6}"] = NI_G_2026

        # Analyst estimates (2026-2029)
        for col, v in ANALYST.items():
            ws[f"{col}{b + 8}"] = v

        # NI margins (D..H; C stays as formula =NI/Rev)
        for col, m in cfg["margins"].items():
            ws[f"{col}{b + 10}"] = m

        # PE low / high
        for col, v in cfg["pe_low"].items():
            ws[f"{col}{b + 14}"] = v
        for col, v in cfg["pe_high"].items():
            ws[f"{col}{b + 15}"] = v

    wb.save(XLSX)
    print(f"Added QXO tab. Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
