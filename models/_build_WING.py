
#!/usr/bin/env python3
"""
Build 6-sheet Excel valuation model for Wingstop Inc. (WING)
Date: 2026-08-14, Quote: $126.12 close
Primary lens: Forward P/E — Fallen Angel framework
Negative equity (-$737M common), massive analyst estimate cuts (20-24 down in 30d),
63.4% drawdown from 52-week high, FY2025 one-time $87M income spike.
Source: finance.yahoo.com/quote/WING/ (StockAnalysis 404)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2F5496")
BOLD = Font(name="Calibri", size=10, bold=True)
DATA = Font(name="Calibri", size=10)
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def hdr_row(ws, row, cols):
    """Style a header row with dark blue fill, white bold text, borders."""
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN

def title_block(ws, text, row=1):
    """Merge A-D with title text."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center")

def write_table(ws, start_row, headers, rows_data):
    """Write headers + data rows with borders. Returns last row used."""
    nr = len(headers)
    hdr_row(ws, start_row, nr)
    for ci, h in enumerate(headers):
        ws.cell(row=start_row, column=ci+1, value=h)
    for ri, row_vals in enumerate(rows_data):
        for ci, v in enumerate(row_vals):
            cell = ws.cell(row=start_row + 1 + ri, column=ci+1, value=v)
            cell.font = BOLD if isinstance(row_vals[0], str) and row_vals[0] != "" and row_vals[0][0].isupper() and len(row_vals[0]) > 10 else DATA
            cell.border = THIN
    return start_row + 1 + len(rows_data)

# ════════════════════════════════════════════════════════════
# Sheet 1: Valuation
# ════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Valuation"

title_block(ws, "Wingstop Inc. (WING) — Valuation Summary")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
c2 = ws.cell(row=2, column=1, value="Fallen Angel Framework — Forward P/E Primary Lens")
c2.font = Font(name="Calibri", size=11, italic=True, color="666666")
c2.alignment = Alignment(horizontal="center")

title_fields = [
    ("Field", "Value", "Source / Notes"),
    ("Ticker", "NASDAQ: WING", ""),
    ("Date", "2026-08-14", ""),
    ("Price", "$126.12", "Close 14 Aug 2026"),
    ("Shares Outstanding", "27.24M", "Yahoo Key Stats"),
    ("Market Cap", "$3.44B", "27.24M x $126.12"),
    ("Enterprise Value", "$4.35B", "Yahoo Key Stats"),
    ("Total Debt", "$1.27B", "BS 12/31/2025"),
    ("Cash", "$127.5M", "Key Stats MRQ"),
    ("Net Debt (EV − MC)", "$0.91B", "$4.35B − $3.44B"),
    ("Primary Lens", "Forward P/E", "Negative equity disqualifies P/B; FCF gray zone"),
    ("Stance", "Watch", "Fallen angel — base case near current price"),
    ("52-Week Range", "$110.44 – $345.81", "−63.4% from high"),
    ("Beta (5Y Mo)", "1.81", "Yahoo Key Stats — high vol"),
    ("Dividend Yield", "1.16%", "$1.32 forward annual"),
]

write_table(ws, 4, title_fields[0], title_fields[1:])

# Valuation metrics
ws.cell(row=len(title_fields) + 5, column=1, value="Key Valuation Metrics").font = Font(name="Calibri", size=12, bold=True)

metrics = [
    ("Metric", "Value", "Comment"),
    ("P/E Trailing", "27.80x", "From $4.23 TTM EPS"),
    ("Forward P/E (FY2026)", "28.3x", "$126.12 / $4.45 consensus (28 analysts)"),
    ("Forward P/E (FY2027)", "23.5x", "$126.12 / $5.36 consensus"),
    ("P/S (TTM)", "4.51x", "Revenue $720.7M"),
    ("EV/EBITDA", "18.6x", "S&P EBITDA $234M"),
    ("EV/Sales", "6.03x", "Key Stats"),
    ("PEG (5yr expected)", "1.82x", "Key Stats"),
    ("FCF Yield (TTM)", "3.74%", "$128.5M FCF / $3.44B MC"),
    ("Dividend Yield", "1.16%", "$1.32 forward annual / $126.12"),
    ("Short % of Float", "15.64%", "7/31/2026 — highly contested"),
]

write_table(ws, len(title_fields) + 6, metrics[0], metrics[1:])

# Revenue history
wr = len(title_fields) + 6 + len(metrics)
ws.cell(row=wr + 1, column=1, value="Revenue History — Income Statement (annual)").font = Font(name="Calibri", size=12, bold=True)

rev_history = [
    ("Year / Period", "Revenue ($M)", "Gross Profit ($M)", "Op Income ($M)", "Net Income ($M)", "Diluted EPS"),
    ("FY2022", "$357.5", "$171.1", "$93.1", "$52.9", "$1.77"),
    ("FY2023", "$460.1", "$222.8", "$112.7", "$70.2", "$2.35"),
    ("FY2024", "$625.8", "$300.9", "$164.6", "$108.7", "$3.70"),
    ("FY2025*", "$696.9", "$339.3", "$185.8", "$174.3", "$6.21"),
    ("TTM", "$720.7", "$356.2", "$200.8", "$116.4", "$4.23"),
]
write_table(ws, wr + 2, rev_history[0], rev_history[1:])

ws.cell(row=wr + 2 + len(rev_history), column=1,
        value="* FY2025 includes $87.1M one-time gain (Other Income/Expense). Normalized NI ~$87.1M, EPS ~$3.93.").font = Font(name="Calibri", size=9, italic=True, color="CC0000")

for c in range(1, 7):
    ws.column_dimensions[get_column_letter(c)].width = 20

# ════════════════════════════════════════════════════════════
# Sheet 2: WACC
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("WACC")
title_block(ws2, "WACC — Wingstop Inc. (WING)")
ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
ws2.cell(row=2, column=1).font = Font(name="Calibri", size=11, italic=True, color="666666")

# CAPM calculation
Rf = 0.04692       # CNBC US10Y Aug 14, 2026
ERP = 0.05
Beta = 1.81        # Yahoo Key Stats
Tax = 0.27         # Historical blended (23.6%-28.3%)
Ke = Rf + Beta * ERP                       # 13.74%
Kd_pre = 0.030   # $38M interest / $1,270M debt = 3.00%
Kd_after = Kd_pre * (1 - Tax)             # 2.19%
MC = 3.44        # $B
TD = 1.27        # $B
V = MC + TD
We = MC / V     # 73.1%
Wd = TD / V     # 26.9%
WACC = We * Ke + Wd * Kd_after            # 10.63%

print(f"WACC: Rf={Rf:.4f}, Ke={Ke:.4f}, Kd_after={Kd_after:.4f}, We={We:.4f}, Wd={Wd:.4f}, WACC={WACC:.4f}")

wacc_rows = [
    ("Component", "Value", "Calculation / Source"),
    ("Risk-Free Rate (10Y US)", f"{Rf*100:.2f}%", "CNBC US10Y, Aug 14 2026"),
    ("Equity Risk Premium", "5.00%", "Assumed"),
    ("Levered Beta (5Y Monthly)", f"{Beta:.2f}", "Yahoo Key Stats"),
    ("Cost of Equity (CAPM)", f"{Ke*100:.2f}%", f"= {Rf*100:.2f}% + {Beta} x 5.00%"),
    ("Cost of Debt (pre-tax)", f"{Kd_pre*100:.2f}%", f"= $38M int / $1,270M debt"),
    ("Tax Rate", "27.00%", "Blended historical"),
    ("Cost of Debt (after-tax)", f"{Kd_after*100:.2f}%", f"= 3.00% x (1 − 0.27)"),
    ("Market Cap", "$3.44B", "27.24M x $126.12"),
    ("Total Debt", "$1.27B", "BS 12/31/2025"),
    ("Equity Weight", f"{We*100:.1f}%", f"= 3.44 / (3.44 + 1.27)"),
    ("Debt Weight", f"{Wd*100:.1f}%", f"= 1.27 / (3.44 + 1.27)"),
    ("", "", ""),
    ("WACC", f"{WACC*100:.2f}%", f"= {We*100:.1f}% x {Ke*100:.2f}% + {Wd*100:.1f}% x {Kd_after*100:.2f}%"),
]
write_table(ws2, 4, wacc_rows[0], wacc_rows[1:])
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 50

# ════════════════════════════════════════════════════════════
# Sheet 3: Scenarios
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Scenarios")
title_block(ws3, "Scenario Analysis — WING (Forward P/E Primary Framework)")
ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
ws3.cell(row=2, column=1, value=(
    "Primary: Forward P/E. FCF multiples used as secondary cross-check. "
    "Negative common equity (−$737M) and debt/FCF ratio of 7.5x make FCF-multiple "
    "bear targets unreliable (implied EV < net debt)."))
ws3.cell(row=2, column=1).font = Font(name="Calibri", size=10, italic=True)
ws3.cell(row=2, column=1).alignment = Alignment(wrap_text=True)

# Parameters
# FY2027 EPS consensus: $5.36 | Revenue FY2027: $887.9M
# Growth to 2031 = 4 years from FY2027
fwd_eps_2027 = 5.36
rev_2027 = 887.9  # $M

# Bear: 5% EPS CAGR, 15x exit P/E
# Base: 9% EPS CAGR, 20x exit P/E  
# Bull: 15% EPS CAGR, 27x exit P/E
bear_cagr, bear_pe = 0.05, 15
base_cagr, base_pe = 0.09, 20
bull_cagr, bull_pe = 0.15, 27

bear_eps31 = fwd_eps_2027 * (1 + bear_cagr)**4
base_eps31 = fwd_eps_2027 * (1 + base_cagr)**4
bull_eps31 = fwd_eps_2027 * (1 + bull_cagr)**4

bear_tgt = bear_eps31 * bear_pe
base_tgt = base_eps31 * base_pe
bull_tgt = bull_eps31 * bull_pe

# Revenue CAGRs mirror EPS
bear_rev31 = rev_2027 * (1 + bear_cagr)**4
base_rev31 = rev_2027 * (1 + base_cagr)**4
bull_rev31 = rev_2027 * (1 + bull_cagr)**4

# FCF marginal margins
bear_fm, base_fm, bull_fm = 0.14, 0.18, 0.22

# Weighted FV
w_b, w_s, w_u = 0.20, 0.50, 0.30
fv = w_b * bear_tgt + w_s * base_tgt + w_u * bull_tgt
current = 126.12

print(f"\nScenario Targets (P/E Primary):")
print(f"  Bear:  EPS=${bear_eps31:.2f}, {bear_pe}x, Target=${bear_tgt:.2f}")
print(f"  Base:  EPS=${base_eps31:.2f}, {base_pe}x, Target=${base_tgt:.2f}")
print(f"  Bull:  EPS=${bull_eps31:.2f}, {bull_pe}x, Target=${bull_tgt:.2f}")
print(f"  FV = ${fv:.2f}  (current=${current})")

# Revenue / FCF rows
bear_fcf31 = bear_rev31 * bear_fm
base_fcf31 = base_rev31 * base_fm
bull_fcf31 = bull_rev31 * bull_fm

net_debt_mm = 914   # EV − MC in $M

scen_headers = ["Metric", "Bear (20%)", "Base (50%)", "Bull (30%)", "Weighted", "Notes"]
scen_data = [
    ("Revenue CAGR (FY2027–2031)", f"{bear_cagr:.0%}", f"{base_cagr:.0%}", f"{bull_cagr:.0%}", "", "From $887.9M FY2027 cons."),
    ("Terminal Revenue (2031, $M)", f"${bear_rev31:.0f}", f"${base_rev31:.0f}", f"${bull_rev31:.0f}", "", ""),
    ("", "", "", "", "", ""),
    ("EPS CAGR (FY2027–2031)", f"{bear_cagr:.0%}", f"{base_cagr:.0%}", f"{bull_cagr:.0%}", "", "From $5.36 FY2027 cons."),
    ("Terminal EPS (2031)", f"${bear_eps31:.2f}", f"${base_eps31:.2f}", f"${bull_eps31:.2f}", "", "Normalized — excl. FY25 one-time"),
    ("Exit P/E (primary)", str(bear_pe), str(base_pe), str(bull_pe), "", "Primary framework"),
    ("Target Price (P/E)", f"${bear_tgt:.2f}", f"${base_tgt:.2f}", f"${bull_tgt:.2f}", "", "= Term EPS x Exit P/E"),
    ("Upside from Current", f"{(bear_tgt/current-1):.1%}", f"{(base_tgt/current-1):.1%}", f"{(bull_tgt/current-1):.1%}", "", f"From ${current:.2f}"),
    ("Weight", "20%", "50%", "30%", "", ""),
    ("Weighted Value / Share", f"${bear_tgt * w_b:.2f}", f"${base_tgt * w_s:.2f}", f"${bull_tgt * w_u:.2f}", f"${fv:.2f}", "Sum = weighted FV"),
    ("", "", "", "", "", ""),
    ("Probability-Weighted FV", "", "", "", f"${fv:.2f}", "Primary estimate"),
    ("Up/Downside from Current", "", "", "", f"{(fv/current-1):.1%}", f"From ${current:.2f}"),
    ("", "", "", "", "", ""),
    ("--- FCF Cross-Check (secondary) ---", "", "", "", "", ""),
    ("Implied FCF Margin", f"{bear_fm:.0%}", f"{base_fm:.1%}", f"{bull_fm:.0%}", "", "TTM = 17.8%"),
    ("Terminal FCF (2031, $M)", f"${bear_fcf31:.0f}", f"${base_fcf31:.0f}", f"${bull_fcf31:.0f}", "", "= Rev x FCF margin"),
    ("Exit FCF Multiple", "8", "10", "14", "", "Cross-check"),
    ("Implied EV ($M)", f"${bear_fcf31*8:.0f}", f"${base_fcf31*10:.0f}", f"${bull_fcf31*14:.0f}", "", "= FCF x Multiple"),
    ("Less Net Debt ($M)", "-$914", "-$914", "-$914", "", "EV − MC"),
    ("FCF Target Price", f"${(bear_fcf31*8 - net_debt_mm)/27.24:.2f}", f"${(base_fcf31*10 - net_debt_mm)/27.24:.2f}", f"${(bull_fcf31*14 - net_debt_mm)/27.24:.2f}", "", "Cross-check only"),
]

write_table(ws3, 4, scen_headers, scen_data)
for c in range(1, 7):
    ws3.column_dimensions[get_column_letter(c)].width = 22

# ════════════════════════════════════════════════════════════
# Sheet 4: Actuals Source Audit
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Actuals Source Audit")
title_block(ws4, "Actuals Source Audit — WING")

audit_hdr = ["Data Point", "Value", "Source", "Date", "Notes"]
audit_rows = [
    ("Stock Price", "$126.12", "yahoo.com/quote/WING/", "2026-08-14", "Close; +11.05% intraday"),
    ("After Hours", "$126.20", "yahoo.com/quote/WING/", "2026-08-14", "7:56 PM EDT"),
    ("Market Cap", "$3.44B", "27.24M x $126.12", "2026-08-14", ""),
    ("Enterprise Value", "$4.35B", "yahoo.com/quote/WING/key-statistics", "2026-08-14", ""),
    ("Shares Outstanding (common)", "27.24M", "Key Stats", "2026-08-14", ""),
    ("Implied Shares Outstanding", "27.24M", "Key Stats", "2026-08-14", "= common; no convertibles"),
    ("52-Week High", "$345.81", "Key Stats", "2026-08-14", ""),
    ("52-Week Low", "$110.44", "Key Stats", "2026-08-14", "Near current price"),
    ("Beta (5Y Monthly)", "1.81", "Key Stats", "2026-08-14", "High vol; amplified drawdown"),
    ("Forward P/E", "27.17x", "Key Stats", "2026-08-14", ""),
    ("Trailing P/E", "27.80x", "Key Stats", "2026-08-14", "On TTM EPS $4.23"),
    ("PEG (5yr expected)", "1.82x", "Key Stats", "2026-08-14", ""),
    ("Short % of Float", "15.64%", "Key Stats", "2026-08-14", "7/31/2026"),
    ("", "", "", "", ""),
    ("Revenue TTM", "$720.7M", "IS page", "2026-08-14", "720,719K"),
    ("Revenue FY2025", "$696.9M", "IS page", "2026-08-14", "12/31/2025"),
    ("Revenue FY2024", "$625.8M", "IS page", "2026-08-14", "12/31/2024"),
    ("Revenue FY2023", "$460.1M", "IS page", "2026-08-14", "12/31/2023"),
    ("Revenue FY2022", "$357.5M", "IS page", "2026-08-14", "12/31/2022"),
    ("Revenue 3Y CAGR", "+32.6%", "Calculated", "2026-08-14", "$357.5M → $696.9M"),
    ("Diluted EPS TTM", "$4.23", "IS / Key Stats", "2026-08-14", ""),
    ("Diluted EPS FY2025", "$6.21", "IS page", "2026-08-14", "Inflated by $87.1M one-time gain"),
    ("Diluted EPS FY2024", "$3.70", "IS page", "2026-08-14", ""),
    ("Diluted EPS FY2023", "$2.35", "IS page", "2026-08-14", ""),
    ("Net Income FY2025", "$174.3M", "IS page", "2026-08-14", "Before-tax $237.2M includes $87.1M other income"),
    ("Op Income FY2025", "$185.8M", "IS page", "2026-08-14", ""),
    ("Interest Expense TTM", "$38.0M", "IS page", "2026-08-14", "Up from $21.2M FY2022"),
    ("EBITDA TTM", "$227.1M", "IS page", "2026-08-14", "Normalized $227.5M"),
    ("", "", "", "", ""),
    ("Total Debt", "$1.27B", "BS page", "2026-08-14", "12/31/2025: $1,270,406K"),
    ("Capital Lease Obligations", "$61.3M", "BS page", "2026-08-14", "From $2.3M FY2022"),
    ("Total Cash", "$127.5M", "Key Stats", "2026-08-14", "MRQ"),
    ("Common Stock Equity", "−$736.8M", "BS page", "2026-08-14", "NEGATIVE — buyback driven"),
    ("Net Tangible Assets", "−$853.3M", "BS page", "2026-08-14", "Goodwill + intangibles heavy"),
    ("Total Liabilities", "$1,430M", "BS page", "2026-08-14", ""),
    ("Current Ratio", "2.97", "Key Stats", "2026-08-14", ""),
    ("", "", "", "", ""),
    ("Operating Cash Flow TTM", "$189.5M", "CF page", "2026-08-14", "189,490K"),
    ("CapEx TTM", "$61.0M", "CF page", "2026-08-14", "60,996K"),
    ("FCF (OCF-CapEx) TTM", "$128.5M", "Computed", "2026-08-14", "189.5 − 61.0"),
    ("Repurchase of Stock TTM", "−$179.6M", "CF page", "2026-08-14", "Aggressive; 179,605K"),
    ("Issuance of Debt FY2025", "+$500M", "CF page", "2026-08-14", "Funds buybacks / growth"),
    ("", "", "", "", ""),
    ("Rev Estimate FY2026", "$765.7M", "Analysis", "2026-08-14", "26 analysts"),
    ("Rev Estimate FY2027", "$887.9M", "Analysis", "2026-08-14", "28 analysts"),
    ("EPS Estimate FY2026", "$4.45", "Analysis", "2026-08-14", "28 analysts"),
    ("EPS Estimate FY2027", "$5.36", "Analysis", "2026-08-14", "28 analysts"),
    ("EPS Revisions (30d down)", "20–24 analysts", "Analysis", "2026-08-14", "Sharp downward revision trend"),
    ("Quarterly EPS Beats (last 4Q)", "4 consecutive beats", "Analysis", "2026-08-14", "+15% to +20% avg surprise"),
    ("Dividend Forward Annual", "$1.32", "Key Stats", "2026-08-14", "Yield 1.16%"),
    ("Dividend Payout Ratio", "28.37%", "Key Stats", "2026-08-14", ""),
    ("10Y Treasury Rate", "4.69%", "cnbc.com/quotes/US10Y", "2026-08-14", "5:05 PM EDT"),
    ("Last Earnings Date", "Jul 29, 2026", "Key Stats", "2026-08-14", "Q2 FY26"),
    ("Next Earnings (est)", "Oct 2026", "Inferred", "2026-08-14", "Q3 FY26"),
]
write_table(ws4, 4, audit_hdr, audit_rows)
for c in range(1, 6):
    ws4.column_dimensions[get_column_letter(c)].width = 28

# ════════════════════════════════════════════════════════════
# Sheet 5: Questions
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Questions")
title_block(ws5, "Open Questions — WING")

q_hdr = ["#", "Question", "Why It Matters"]
questions = [
    ("Q1",
     "What was the $87.1M one-time gain in FY2025 'Other Income/Expense'? Sale of investments, insurance recovery, or accounting adjustment?",
     "Defines the normalized earnings baseline. Without it, FY2025 EPS of $6.21 overstates true operating performance by 58%."),
    ("Q2",
     "Why the debt surge: $1.27B total (from $732.5M FY2023)? CF shows $500M issued FY2024, $500M FY2025. Used for buybacks or store expansion?",
     "If debt funded buybacks at $200–$345/share (peak prices), up to $400–600M of intrinsic shareholder value was destroyed. If funding growth capex, trajectory changes."),
    ("Q3",
     "Common equity is −$737M — what accounts for the negative retained earnings? WING generates $189.5M OCF, $128.5M FCF TTM.",
     "Negative equity disqualifies P/B. Cumulative ~$842M in buybacks (FY2023: $125M, FY2024: $315M, FY2025: $222M, TTM: $180M) drove it. Key question is buyback price vs. value created per share."),
    ("Q4",
     "15.64% short interest (7/31/2026) — positioning for further downside or priced in?",
     "High short interest can fuel short squeezes (WING +11% today) but also indicates significant bear conviction. The underlying thesis of deceleration + multiple compression may remain valid."),
    ("Q5",
     "What is the same-store sales trajectory? Q2 FY2026 revenue growth was +6.4% YoY — is this decelerating from the prior 15–20% growth that drove the $345 high?",
     "Wingstop's entire prior valuation depended on high-single-to-double-digit sss growth. A deceleration to 5–8% fundamentally changes the EPS trajectory and supports multiple compression."),
    ("Q6",
     "International exposure and franchise vs. company-owned mix?",
     "International growth often outpaces US but with different unit economics. Company-owned has higher margin but higher capex. Mix shift affects FCF trajectory."),
    ("Q7",
     "Poultry/chicken cost inflation impact on gross margin? TTM gross margin = 49.4% — is this pressured vs. FY2024 levels?",
     "Food cost drives variable COGS. Prior QSR weakness was partly cost-driven. If menu price increases lag commodity inflation, GM compression drives multiple contraction."),
    ("Q8",
     "20–24 analysts cut estimates in 30 days. Q3 FY26 EPS estimate dropped $1.19 → $1.07 in 90 days (10%). Macro QSR slowdown or WING-specific?",
     "Structurally significant. Combined with consecutive quarterly beats (+15–20% surprise) over the past 4 quarters, the revision pattern suggests either a macro deceleration cycle or an internal inflection."),
    ("Q9",
     "Capital lease obligations up 26x: $2.3M (FY2022) → $61.3M (FY2025). What does this signal for real estate / store expansion?",
     "Likely reflects shift from leased to company-owned real estate. Positive long-term for margins, but increases leverage. Standard QSR expansion pattern."),
    ("Q10",
     "Is buyback deployment at $126/share more accretive than prior rounds at $200+? 27.24M shares now vs. 29.97M in FY2022 (9%/share dilution offset by buybacks).",
     "Management has a proven track record of returning capital. The question is whether the $126 price represents better value creation than $250–345 buybacks. If management continues buybacks at current levels, per-share accretion could be significant."),
]
write_table(ws5, 4, q_hdr, questions)
ws5.column_dimensions["A"].width = 8
ws5.column_dimensions["B"].width = 75
ws5.column_dimensions["C"].width = 65

# ════════════════════════════════════════════════════════════
# Sheet 6: Sources
# ════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Sources")
title_block(ws6, "Sources — WING")

s_hdr = ["#", "Source", "URL", "Date"]
sources = [
    ("1", "Summary / Price", "finance.yahoo.com/quote/WING/", "2026-08-14"),
    ("2", "Income Statement", "finance.yahoo.com/quote/WING/financials/", "2026-08-14"),
    ("3", "Balance Sheet", "finance.yahoo.com/quote/WING/balance-sheet/", "2026-08-14"),
    ("4", "Cash Flow", "finance.yahoo.com/quote/WING/cash-flow/", "2026-08-14"),
    ("5", "Key Statistics", "finance.yahoo.com/quote/WING/key-statistics/", "2026-08-14"),
    ("6", "Analysis / Estimates", "finance.yahoo.com/quote/WING/analysis/", "2026-08-14"),
    ("7", "Company Profile", "finance.yahoo.com/quote/WING/profile/", "2026-08-14"),
    ("8", "10-Year Treasury Yield", "cnbc.com/quotes/US10Y", "2026-08-14"),
    ("9", "StockAnalysis.com (404)", "stockanalysis.com/quote/WING/", "2026-08-14"),
    ("10", "Peer Tickers (SHAK, DPZ, TXRH, EAT, DRI, PZZA, CAVA, YUM, QSR)", "Key Stats — Related", "2026-08-14"),
]
write_table(ws6, 4, s_hdr, sources)
for c in range(1, 5):
    ws6.column_dimensions[get_column_letter(c)].width = 25

# ── Save ──
wb.save("/home/refcell/dev/capital/models/[2026-08-14] Wingstop Model.xlsx")
print("Saved: [2026-08-14] Wingstop Model.xlsx")
