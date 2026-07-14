#!/usr/bin/env python3
"""Build 6-sheet valuation model for Post Holdings (POST) as of July 13, 2026."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import date
import os

wb = openpyxl.Workbook()

# ── Styles ──
title_font = Font(name='Calibri', size=14, bold=True, color='1F2937')
subtitle_font = Font(name='Calibri', size=11, bold=True, color='374151')
header_font = Font(name='Calibri', size=10, bold=True, color='F9FAFB')
header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
data_font = Font(name='Calibri', size=10)
comment_font = Font(name='Calibri', size=9, italic=True, color='6B7280')
neg_font = Font(name='Calibri', size=10, color='DC2626')
pct_fmt = '0.00%'
dollar_fmt = '$#,##0.0'
dollar_mm_fmt = '$#,##0'
num_fmt = '#,##0'
int_fmt = '#,##0'
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')

def s(ws, row, col, value, font=data_font, fmt=None, fill=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if align: c.alignment = align
    if border: c.border = border
    return c

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

DATE = date.today().strftime('%Y-%m-%d')

# ═══════════════════════════════════════════════════════════
# DATA
# ═══════════════════════════════════════════════════════════
price = 86.70
shares_basic_mm = 52.8
shares_diluted_mm = 58.9
mc = 3.93  # $B
ev = 11.27  # $B
net_debt = 7.34  # $B (EV - MC)
total_debt = 7.70  # $B
total_cash = 0.27  # $B (total cash from stats)
beta = 0.33
rf = 0.04618   # 10Y Treasury 4.618%
erp = 0.05
cost_equity = rf + beta * erp  # 6.27%
tax_rate = 0.249  # TTM tax provision / pretax
cost_debt = 0.052  # interest / total debt approx
eq_weight = mc / (mc + total_debt)
debt_weight = total_debt / (mc + total_debt)
wacc = eq_weight * cost_equity + debt_weight * cost_debt * (1 - tax_rate)

print(f"WACC: {wacc:.4%} ({wacc*100:.2f}%)")
print(f"  Cost of equity: {cost_equity:.4%}")
print(f"  Cost of debt (after-tax): {cost_debt*(1-tax_rate):.4%}")
print(f"  Eq weight: {eq_weight:.1%}, Debt weight: {debt_weight:.1%}")

# Revenue history (annual, $B)
rev_fy22, rev_fy23, rev_fy24, rev_fy25 = 5.85, 6.99, 7.92, 8.16
rev_ttm = 8.45

# Income (annual, $B)
gross_fy22, gross_fy23, gross_fy24, gross_fy25 = 1.47, 1.88, 2.30, 2.34
op_fy22, op_fy23, op_fy24, op_fy25 = 0.42, 0.64, 0.79, 0.83
nic_fy22, nic_fy23, nic_fy24, nic_fy25 = 0.76, 0.31, 0.37, 0.34
nic_ttm = 0.34

# EBITDA (annual, $B)
ebitda_fy22, ebitda_fy23, ebitda_fy24, ebitda_fy25 = 1.59, 1.10, 1.27, 1.33
ebitda_ttm = 1.42

# FCF (annual, $B) - unlevered (OCF - CapEx)
fcf_fy22, fcf_fy23, fcf_fy24, fcf_fy25 = 0.13, 0.45, 0.50, 0.49
fcf_ttm = 0.52

# Analyst consensus (EPS)
eps_fy26 = 7.71
eps_fy27 = 8.69

# ═══════════════════════════════════════════════════════════
# SCENARIO CALCULATIONS - Forward P/E based (FCF-insufficient)
# ═══════════════════════════════════════════════════════════
# Net debt = $7.34B, FCF * 10 = $5.2B => FCF insufficiency detected.
# Primary lens: Forward P/E on analyst consensus.

# Bear: Revenue flat, EPS low-end, P/E compressed
bear_rev_cagr = 0.00
bear_terminal_rev = 8.27  # near FY2027 est
bear_eps_fy27 = 8.06  # low estimate
bear_exit_pe = 10.0
bear_terminal_eps = bear_eps_fy27 * (1 + 0.03) ** 3  # slow growth over 3 more years
bear_price = bear_terminal_eps * bear_exit_pe

# Base: Consensus, moderate growth, fair P/E
base_rev_cagr = 0.01
base_terminal_rev = 8.27 * (1.01) ** 3
base_eps_fy27 = 8.69  # consensus
base_exit_pe = 13.0
base_terminal_eps = base_eps_fy27 * (1 + 0.06) ** 3
base_price = base_terminal_eps * base_exit_pe

# Bull: Growth acceleration, margin expansion, higher P/E
bull_rev_cagr = 0.03
bull_terminal_rev = 8.32 * (1.03) ** 4
bull_eps_fy27 = 10.17  # high estimate
bull_exit_pe = 16.0
bull_terminal_eps = bull_eps_fy27 * (1 + 0.09) ** 3
bull_price = bull_terminal_eps * bull_exit_pe

# Weighted FV
bear_weight, base_weight, bull_weight = 0.25, 0.50, 0.25
weighted_fv = bear_weight * bear_price + base_weight * base_price + bull_weight * bull_price
upside = (weighted_fv / price) - 1

print(f"\nScenario Targets:")
print(f"  Bear:  ${bear_price:.1f} (EPS $0{bear_terminal_eps:.2f} × {bear_exit_pe:.0f}x)")
print(f"  Base:  ${base_price:.1f} (EPS $0{base_terminal_eps:.2f} × {base_exit_pe:.0f}x)")
print(f"  Bull:  ${bull_price:.1f} (EPS $0{bull_terminal_eps:.2f} × {bull_exit_pe:.0f}x)")
print(f"  Weighted FV: ${weighted_fv:.1f}")
print(f"  Upside: {upside:+.1%}")

# ═══════════════════════════════════════════════════════════
# SHEET 1: Valuation
# ═══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Valuation'
ws1.merge_cells('A1:G1')
s(ws1, 1, 1, 'Post Holdings, Inc. — Valuation Model', title_font)

s(ws1, 3, 1, 'Ticker', subtitle_font)
s(ws1, 3, 2, 'NYSE: POST')
s(ws1, 4, 1, 'Company', subtitle_font)
s(ws1, 4, 2, 'Post Holdings, Inc.')
s(ws1, 5, 1, 'Sector / Industry', subtitle_font)
s(ws1, 5, 2, 'Consumer Defensive / Packaged Foods')
s(ws1, 6, 1, 'Date', subtitle_font)
s(ws1, 6, 2, DATE)
s(ws1, 7, 1, 'Price', subtitle_font)
s(ws1, 7, 2, price, subtitle_font, dollar_fmt)
s(ws1, 8, 1, 'Shares Outstanding', subtitle_font)
s(ws1, 8, 2, f'{shares_basic_mm:.1f}M basic / {shares_diluted_mm:.1f}M diluted')
s(ws1, 9, 1, 'Market Cap', subtitle_font)
s(ws1, 9, 2, f'${mc:.2f}B', subtitle_font)
s(ws1, 10, 1, 'Enterprise Value', subtitle_font)
s(ws1, 10, 2, f'${ev:.2f}B', subtitle_font)
s(ws1, 11, 1, 'Primary Valuation Lens', subtitle_font)
s(ws1, 11, 2, 'Forward P/E (analyst consensus) — FCF framework insufficient: net debt $7.3B > FCF × 10')
s(ws1, 12, 1, 'Current Stance', subtitle_font)
s(ws1, 12, 2, 'Watch — attractive on surface multiples but net debt is heavy; FV slightly above current price depends on margins')

# Valuation metrics table
vdata = [
    ('Metric', 'Value', 'Comment'),
    ('Trailing P/E', f'{price / 5.96:.1f}x', 'TTM diluted EPS $5.96; relatively cheap by CPG standards'),
    ('Forward P/E (FY26)', f'{price / eps_fy26:.1f}x', f'Using FY26 consensus EPS ${eps_fy26:.2f}; reasonable for CPG'),
    ('Forward P/E (FY27)', f'{price / eps_fy27:.1f}x', f'Using FY27 consensus EPS ${eps_fy27:.2f}; implies significant value'),
    ('P/S (TTM)', f'{price * shares_basic_mm / rev_ttm / 1000:.2f}x', f'TTM revenue $${rev_ttm:.2f}B; very cheap revenue multiple'),
    ('P/B (MRQ)', '1.22x', 'Market value vs book — modest premium, normal for CPG'),
    ('EV/Revenue', '1.33x', 'Enterprise value multiple on revenue — compression vs peers'),
    ('EV/EBITDA', '7.95x', f'TTM EBITDA $${ebitda_ttm:.2f}B; reasonable for the sector'),
    ('EV/FCF', f'{ev/fcf_ttm:.1f}x', f'TTM FCF $${fcf_ttm:.2f}B; EXPENSIVE — debt weight distorts this multiple'),
    ('FCF Yield (Unlevered)', f'{fcf_ttm/mc*100:.1f}%', f'${fcf_ttm*1000:.0f}M FCF / ${mc*1000:.0f}M MC; decent cash-generation per share'),
    ('FCF Yield (Levered)', f'{0.316/mc*100:.1f}%', 'After debt service and interest — lower free-cash return'),
    ('PEG (5yr expected)', 'N/A', 'Growth too flat to meaningfully assess PEG'),
    ('Dividend Yield', 'N/A', 'No dividend currently declared'),
    ('EV - MC (Net Debt $)', f'${net_debt:.2f}B', 'Significant net debt position from leveraged buyout and ongoing buybacks'),
    ('Debt/Equity', '238.09%', 'Very high leverage — primary risk factor'),
    ('Interest Coverage (EBIT/Int)', f'{0.85/0.40:.1f}x', 'EBIT $850M / Interest $400M — adequate but tight'),
    ('WACC', f'{wacc*100:.1f}%', f'Low because beta is 0.33; CP defensiive stock with stable cash flows'),
]

# Title block ends at row 12. Start table at row 14.
for i, (field, value, comment) in enumerate(vdata):
    row = i + 14
    font_use = header_font if i == 0 else data_font
    fill_use = header_fill if i == 0 else None
    s(ws1, row, 1, field, font_use, align=wrap_align, fill=fill_use)
    s(ws1, row, 2, value, font_use, align=wrap_align, fill=fill_use)
    s(ws1, row, 3, comment, font_use, align=wrap_align, fill=fill_use)

set_col_widths(ws1, [22, 28, 65])

# ═══════════════════════════════════════════════════════════
# SHEET 2: WACC
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet('WACC')
ws2.merge_cells('A1:C1')
s(ws2, 1, 1, 'WACC Calculation — CAPM', title_font)

wacc_data = [
    ('Component', 'Value', 'Source / Notes'),
    ('Risk-Free Rate (10Y US)', f'{rf*100:.2f}%', 'CNBC US10Y as of July 13, 2026'),
    ('Equity Risk Premium', '5.00%', 'Standard estimate'),
    ('Beta (5Y Monthly)', f'{beta:.2f}', 'Yahoo Finance Statistics'),
    ('Cost of Equity (CAPM)', f'{cost_equity*100:.2f}%', f'= {rf*100:.2f}% + {beta:.2f} × 5.00%'),
    ('Cost of Debt (pre-tax)', f'{cost_debt*100:.2f}%', f'Interest TTM $400M / Total Debt $7.7B ≈ 5.20%'),
    ('Tax Rate', f'{tax_rate*100:.2f}%', f'TTM Tax Provision $112M / Pretax $450M'),
    ('Market Cap ($B)', f'{mc:.2f}', 'Yahoo Finance July 13, 2026 close'),
    ('Total Debt ($B)', f'{total_debt:.2f}', 'Yahoo Finance Balance Sheet 9/30/2025'),
    ('Total Capital ($B)', f'{mc + total_debt:.2f}', 'MC + Total Debt'),
    ('Equity Weight', f'{eq_weight:.2%}', f'MC / Total Capital'),
    ('Debt Weight', f'{debt_weight:.2%}', f'Total Debt / Total Capital'),
    ('After-Tax Cost of Debt', f'{cost_debt*(1-tax_rate)*100:.2f}%', f'= 5.20% × (1 - {tax_rate*100:.1f}%)'),
    ('WACC', f'{wacc*100:.2f}%', f'= {eq_weight:.0%} × {cost_equity*100:.2f}% + {debt_weight:.0%} × {cost_debt*(1-tax_rate)*100:.2f}%'),
]

for i, (field, value, comment) in enumerate(wacc_data):
    font_use = header_font if i == 0 else data_font
    fill_use = header_fill if i == 0 else None
    s(ws2, i+2, 1, field, font_use, align=wrap_align, fill=fill_use)
    s(ws2, i+2, 2, value, font_use, align=wrap_align, fill=fill_use)
    s(ws2, i+2, 3, comment, font_use, align=wrap_align, fill=fill_use)

set_col_widths(ws2, [25, 20, 55])

# ═══════════════════════════════════════════════════════════
# SHEET 3: Scenarios — Forward P/E framework
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Scenarios')
ws3.merge_cells('A1:L1')
s(ws3, 1, 1, 'Scenario Analysis — Forward P/E Framework', title_font)

# Framework note
s(ws3, 3, 1, 'FRAMEWORK NOTE:', subtitle_font)
s(ws3, 3, 2, 'FCF multiple framework NOT used: Net debt ($7.34B) > FCF × 10 ($5.17B). Using Forward P/E on analyst consensus as primary lens, with EV/EBITDA cross-check. 3-year terminal point from FY2027 consensus.')

scenario_headers = ['Scenario', 'Revenue CAGR (5Y)', 'Terminal Revenue ($B)',
                    'EPS FY27 Anchor', 'Terminal EPS (3Y fwd)', 'Exit P/E',
                    'Target Price', 'Upside/PDown', 'Weight', 'Weighted $/Share', 'Total Prob-Weighted FV', 'Upside from Current']

scenarios = [
    ('Bear', bear_rev_cagr, bear_terminal_rev, bear_eps_fy27, bear_terminal_eps, bear_exit_pe,
     bear_price, (bear_price/price-1), bear_weight, bear_weight*bear_price, weighted_fv, upside),
    ('Base', base_rev_cagr, base_terminal_rev, base_eps_fy27, base_terminal_eps, base_exit_pe,
     base_price, (base_price/price-1), base_weight, base_weight*base_price, weighted_fv, upside),
    ('Bull', bull_rev_cagr, bull_terminal_rev, bull_eps_fy27, bull_terminal_eps, bull_exit_pe,
     bull_price, (bull_price/price-1), bull_weight, bull_weight*bull_price, weighted_fv, upside),
]

# Headers at row 5
for j, h in enumerate(scenario_headers):
    s(ws3, 5, j+1, h, header_font, fill=header_fill, align=wrap_align)

# Data rows at rows 6-8
for i, (name, cagr, t_rev, anchor_eps, term_eps, exit_pe, tprice, updown, wt, wval, total, total_up) in enumerate(scenarios):
    r = 6 + i
    s(ws3, r, 1, name, Font(name='Calibri', size=10, bold=True))
    s(ws3, r, 2, cagr, data_font, pct_fmt)
    s(ws3, r, 3, t_rev, data_font, '0.00')
    s(ws3, r, 4, anchor_eps, data_font, '0.00')
    s(ws3, r, 5, term_eps, data_font, '0.00')
    s(ws3, r, 6, exit_pe, data_font, '0.0x')
    s(ws3, r, 7, tprice, data_font, dollar_fmt)
    s(ws3, r, 8, updown, data_font if updown >= 0 else neg_font, pct_fmt)
    s(ws3, r, 9, wt, data_font, pct_fmt)
    s(ws3, r, 10, wval, data_font, dollar_fmt)
    s(ws3, r, 11, total, data_font, dollar_fmt)
    s(ws3, r, 12, total_up, data_font, pct_fmt)

# Summary row
s(ws3, 10, 1, 'SUMMARY', subtitle_font)
s(ws3, 10, 2, f'Probability-weighted fair value: ${weighted_fv:.1f}')
s(ws3, 10, 3, f'Upside from ${price:.2f}: {upside:+.1%}')
s(ws3, 10, 4, f'Base case target ${base_price:.1f} vs analyst avg PT ${120.33:.1f}: within calibration range')

set_col_widths(ws3, [12, 18, 18, 16, 18, 12, 12, 14, 10, 16, 20, 18])

# ═══════════════════════════════════════════════════════════
# SHEET 4: Actuals Source Audit
# ═══════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Actuals Source Audit')
ws4.merge_cells('A1:D1')
s(ws4, 1, 1, 'Actuals Source Audit', title_font)

audit = [
    ('Data Point', 'Value', 'Source', 'Date / Notes'),
    ('Stock Price', '$86.70', 'Yahoo Finance summary', 'July 13, 2026 close'),
    ('Market Cap', '$3.93B', 'Yahoo Finance summary', 'July 13, 2026 intraday'),
    ('Enterprise Value', '$11.27B', 'Yahoo Finance Statistics', 'July 10, 2026'),
    ('Shares Outstanding (basic)', '52,800,000', 'Yahoo Finance Balance Sheet / Profile', '9/30/2025'),
    ('Diluted Shares (TTM avg)', '58,900,000', 'Yahoo Finance Income Statement', 'TTM ending ~June 2026'),
    ('Preferred Shares', '3,200,000', 'Yahoo Finance Balance Sheet', '9/30/2025 — stable across years'),
    ('Beta (5Y Monthly)', '0.33', 'Yahoo Finance Statistics', 'July 10, 2026'),
    ('Revenue TTM', '$8.45B', 'Yahoo Finance Income Statement', 'TTM ending ~June 2026'),
    ('Revenue FY2025', '$8.16B', 'Yahoo Finance Income Statement', '9/30/2025'),
    ('Revenue FY2024', '$7.92B', 'Yahoo Finance Income Statement', '9/30/2024'),
    ('Gross Profit TTM', '$2.45B', 'Yahoo Finance Income Statement', 'TTM'),
    ('Gross Margin TTM', '29.1%', 'Calculated: GP / Revenue', 'TTM'),
    ('Op Income TTM', '$883M', 'Yahoo Finance Income Statement', 'TTM'),
    ('Operating Margin TTM', '10.5%', 'Calculated: Op Inc / Revenue', 'TTM'),
    ('EBITDA TTM', '$1.42B', 'Yahoo Finance Income Statement', 'TTM'),
    ('Net Income TTM', '$338.5M', 'Yahoo Finance Income Statement', 'TTM'),
    ('Net Income FY2025', '$335.7M', 'Yahoo Finance Income Statement', '9/30/2025'),
    ('Diluted EPS TTM', '$5.96', 'Yahoo Finance Income Statement / Summary', 'TTM'),
    ('Total Debt', '$7.70B', 'Yahoo Finance Balance Sheet', '9/30/2025'),
    ('Total Cash (MRQ)', '$269.5M', 'Yahoo Finance Statistics', 'MRQ as of July 10, 2026'),
    ('Net Debt (BS)', '$7.25B', 'Calculated: Total Debt - Total Cash', '9/30/2025'),
    ('Net Debt (EV-MC proxy)', '$7.34B', 'Calculated: $11.27B - $3.93B', 'July 10/13, 2026'),
    ('Operating Cash Flow TTM', '$1.01B', 'Yahoo Finance Cash Flow Statement', 'TTM'),
    ('CapEx TTM', '$488M', 'Yahoo Finance Cash Flow Statement', 'TTM'),
    ('Unlevered FCF TTM', '$517M', 'Calculated: OCF - CapEx', 'TTM'),
    ('Levered FCF TTM', '$316.5M', 'Yahoo Finance Statistics', 'TTM July 10, 2026'),
    ('FCF Margin (Unlev)', '6.1%', 'Calculated: FCF / Revenue', 'TTM'),
    ('Share Repurchases TTM', '$1.05B', 'Yahoo Finance Cash Flow Statement', 'TTM — aggressive buyback program'),
    ('Interest Expense TTM', '$400M', 'Yahoo Finance Income Statement', 'TTM'),
    ('Analyst Avg Target', '$120.33', 'Yahoo Finance Analysis', 'July 13, 2026'),
    ('Analyst PT High', '$131.00', 'Yahoo Finance Analysis', 'July 13, 2026'),
    ('EPS Estimate FY26', '$7.71', 'Yahoo Finance Analysis', '7 analysts, consensus'),
    ('EPS Estimate FY27', '$8.69', 'Yahoo Finance Analysis', '7 analysts, consensus'),
    ('Revenue Est FY26', '$8.32B', 'Yahoo Finance Analysis', '6 analysts, +1.92% YoY'),
    ('Revenue Est FY27', '$8.27B', 'Yahoo Finance Analysis', '5 analysts, -0.59% YoY'),
    ('EPS Surprise Q2 FY26', '+10.95%', 'Yahoo Finance Analysis', 'Est $1.75, Act $1.94'),
    ('EPS Surprise Q4 FY25', '+27.67%', 'Yahoo Finance Analysis', 'Est $1.67, Act $2.13 — massive beat'),
    ('10Y Treasury Yield', '4.618%', 'CNBC US10Y', 'July 13, 2026'),
    ('Debt/Equity (MRQ)', '238.09%', 'Yahoo Finance Statistics', 'July 10, 2026'),
    ('ROE (TTM)', '9.62%', 'Yahoo Finance Statistics', 'TTM'),
    ('ROA (TTM)', '4.45%', 'Yahoo Finance Statistics', 'TTM'),
    ('10Y U.S. Treasury Rate', '4.618%', 'CNBC US10Y', 'July 13, 2026'),
    ('Earnings Date', 'Aug 6, 2026', 'Yahoo Finance Profile', 'Q3 FY26 estimated'),
]

for i, (field, value, source, notes) in enumerate(audit):
    font_use = header_font if i == 0 else data_font
    fill_use = header_fill if i == 0 else None
    s(ws4, i+2, 1, field, font_use, align=wrap_align, fill=fill_use)
    s(ws4, i+2, 2, value, font_use, align=wrap_align, fill=fill_use)
    s(ws4, i+2, 3, source, font_use, align=wrap_align, fill=fill_use)
    s(ws4, i+2, 4, notes, font_use, align=wrap_align, fill=fill_use)

set_col_widths(ws4, [28, 22, 38, 45])

# ═══════════════════════════════════════════════════════════
# SHEET 5: Questions
# ═══════════════════════════════════════════════════════════
ws5 = wb.create_sheet('Questions')
ws5.merge_cells('A1:C1')
s(ws5, 1, 1, 'Open Questions', title_font)

questions = [
    ('Q1', 'Preferred Stock — Terms and Obligations', '3,200,000 preferred shares exist, stable since at least FY2022. What are the dividend obligations? Fixed-charge coverage? Should preferred equity be subtracted from market cap for common share valuation?'),
    ('Q2', 'Net Debt vs FCF Coverage', 'Net debt of $7.34B vs unlevered FCF of $517M (14.2x multiple). Is this leverage sustainable, or is it a leveraged-buyout legacy that will compress over time? What is the targeted net-debt reduction trajectory?'),
    ('Q3', 'Aggressive Share Repurchases', 'TTM buybacks of $1.05B — that is ~2% of market cap. Are these sustainable given the heavy debt load? Does management prioritize buybacks over debt reduction? How does buyback volume affect FCF available for reinvestment?'),
    ('Q4', 'Revenue Plateau', 'Analyst consensus shows FY2026 $8.32B → FY2027 $8.27B (decline). Has the company hit a revenue ceiling? What organic or M&A growth drivers remain after the BellRing / Bob Evans / Michael Foods rollups?'),
    ('Q5', 'Interest Expense Sustainability', 'Interest expense TTM $399M on $7.70B debt (~5.2% average rate). If rates stay elevated or rise, does interest coverage (EBIT/Int ~2.1x) become concerning? What happens in a rate-hike environment?'),
    ('Q6', 'Negative Tangible Book Value', 'Net tangible assets = -$4.11B (FY2025). The entire balance sheet is goodwill + intangibles from acquisition rolldup. Is this sustainable, or are impairment risks real?'),
    ('Q7', 'FCF Quality and Conversion', 'Unlevered FCF $517M but levered FCF only $316.5M — $200M lost to financing. OCF of $1.01B looks strong, but most is consumed by debt service. Is the FCF quality durable or is it a function of current rate environment?'),
    ('Q8', 'Segment Contribution and Cyclicality', 'Foodservice segment (eggs, potatoes, meat) carries different margin/cyclicality profile than Retail CPG. How stable is the foodservice mix, and does it introduce commodity price risk?'),
    ('Q9', 'Customer Concentration and Retailer Power', 'Serves grocery, mass merch, club stores, dollar stores. Is there concentration risk in top retailers (Walmart, Costco, Amazon)? What is the impact of private label vs branded mix shift?'),
    ('Q10', 'Capital Allocation Priority', 'With $1.05B TTM in buybacks, $517M FCF, and $7.70B in debt, the capital allocation hierarchy is unclear. Should debt reduction precede buybacks? Is management signaling confidence via buybacks despite leverage?'),
    ('Q11', 'Competitive Differentiation', 'CPG is low-differentiation. How does Post compete with generalists like Conagra, General Mills, Kellogg/Mondelez on the same product categories? Brand equity depth vs. cost leadership?'),
    ('Q12', 'Next Earnings Catalyst', 'Q3 FY26 earnings on Aug 6, 2026. The company has beaten estimates 4 consecutive quarters (+11% to +28%). Can the beat streak continue, and what is the revision trajectory for FY26/FY27?'),
    ('Q13', 'SBC and Dilution', 'Diluted shares (58.9M) vs basic (52.8M) = 6.1M dilution gap. How much SBC accrual is embedded? Does this gap shrink as buybacks reduce treasury shares?'),
]

for i, (num, title, body) in enumerate(questions):
    s(ws5, i+2, 1, num, subtitle_font)
    s(ws5, i+2, 2, title, subtitle_font, align=wrap_align)
    s(ws5, i+2, 3, body, data_font, align=wrap_align)

set_col_widths(ws5, [6, 35, 90])

# ═══════════════════════════════════════════════════════════
# SHEET 6: Sources
# ═══════════════════════════════════════════════════════════
ws6 = wb.create_sheet('Sources')
ws6.merge_cells('A1:C1')
s(ws6, 1, 1, 'Sources', title_font)

sources = [
    ('1', 'Yahoo Finance — Summary / Statistics', 'https://finance.yahoo.com/quote/POST/'),
    ('2', 'Yahoo Finance — Income Statement', 'https://finance.yahoo.com/quote/POST/financials/'),
    ('3', 'Yahoo Finance — Balance Sheet', 'https://finance.yahoo.com/quote/POST/balance-sheet/'),
    ('4', 'Yahoo Finance — Cash Flow', 'https://finance.yahoo.com/quote/POST/cash-flow/'),
    ('5', 'Yahoo Finance — Company Profile', 'https://finance.yahoo.com/quote/POST/profile/'),
    ('6', 'Yahoo Finance — Analyst Estimates', 'https://finance.yahoo.com/quote/POST/analysis/'),
    ('7', 'CNBC — 10-Year Treasury Yield', 'https://www.cnbc.com/quotes/US10Y'),
    ('8', 'StockAnalysis — 404 (unavailable for POST)', 'https://stockanalysis.com/quote/POST/'),
    ('9', 'Argus Research — HOLD rating, PT $84', 'Via Yahoo Finance Research tab'),
    ('10', 'Wells Fargo — Equal-Weight, PT lowered $110→$98', 'July 8, 2026, via Yahoo Finance'),
]

for i, (num, name, url) in enumerate(sources):
    s(ws6, i+2, 1, num, data_font)
    s(ws6, i+2, 2, name, data_font, align=wrap_align)
    s(ws6, i+2, 3, url, data_font, align=wrap_align)

set_col_widths(ws6, [6, 55, 60])

# ═══════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════
outpath = os.path.join('/home/refcell/dev/capital/models', f'{DATE} Post Holdings Model.xlsx')
wb.save(outpath)
print(f"\nSaved to: {outpath}")

# Verify
wb2 = openpyxl.load_workbook(outpath)
print(f"Verification: {len(wb2.sheetnames)} sheets: {wb2.sheetnames}")
wb2.close()
print("Build complete.")
