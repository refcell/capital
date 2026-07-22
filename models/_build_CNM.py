#!/usr/bin/env python3
"""Build 6-sheet valuation model for CNM (Core & Main, Inc.) — Industrial Distribution."""

import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Styles ──
bold = Font(bold=True)
title_font = Font(bold=True, size=14)
section_font = Font(bold=True, size=12)
header_font = Font(bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')

def c(ws, row, col, value, font=None, border=True, fill=None, alignment=None):
    """Write cell with formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if border:
        cell.border = thin_border
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    return cell

def write_header_row(ws, row, headers):
    for i, h in enumerate(headers, 1):
        cell = c(ws, row, i, h, font=header_font, fill=header_fill)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

# ═══════════════════════════════════════════
# Sheet 1: Valuation
# ═══════════════════════════════════════════
ws1 = wb.active
ws1.title = "Valuation"
ws1.merge_cells('A1:F1')
c(ws1, 1, 1, "CNM (Core & Main, Inc.) — Valuation Summary", font=title_font)

price = 43.37
shares_outstanding_mm = 187.19
market_cap_b = price * shares_outstanding_mm / 1000  # $8.12B (using current price × stats shares)
market_cap_reported_b = 8.20
enterprise_value_b = 10.50
net_debt_b = enterprise_value_b - market_cap_b  # ~$2.30B — cleanest net debt proxy
total_debt_b = 2.45
date_str = "2026-07-21"

title_data = [
    ("Company", "Core & Main, Inc."),
    ("Ticker", "NYSE:CNM"),
    ("Date", date_str),
    ("Price (USD)", f"${price}"),
    ("Shares Outstanding (M)", f"{shares_outstanding_mm:.2f}M"),
    ("Market Cap (USD)", f"${market_cap_reported_b:.2f}B"),
    ("Enterprise Value (USD)", f"${enterprise_value_b:.2f}B"),
    ("Net Debt Proxy (EV - MC)", f"${net_debt_b:.2f}B"),
    ("Primary Valuation Lens", "FCF Multiple / Forward P/E"),
    ("Stance", "Hold"),
]

for i, (label, value) in enumerate(title_data, 2):
    c(ws1, i, 1, label, font=bold)
    c(ws1, i, 2, value)

# Valuation metrics table
c(ws1, 13, 1, "Key Valuation Metrics", font=section_font)
metric_headers = ["Metric", "Value", "Comment"]
for i, h in enumerate(metric_headers, 1):
    c(ws1, 14, i, h, font=header_font, fill=header_fill)

ttm_eps = 2.36  # diluted EPS TTM
fwd_eps = 2.81  # FY27 non-GAAP consensus
ttm_fcf_mm = 608
ttm_rev_mm = 7646
ttm_ebitda_mm = 924

pe_ttm = price / ttm_eps
pe_fwd = price / fwd_eps
ps = market_cap_b * 1000 / ttm_rev_mm
pfcf = market_cap_b * 1000 / ttm_fcf_mm
ev_fcf = enterprise_value_b * 1000 / ttm_fcf_mm
ev_sales = enterprise_value_b * 1000 / ttm_rev_mm
ev_ebitda = enterprise_value_b * 1000 / ttm_ebitda_mm

metrics = [
    ("P/E (TTM)", f"{pe_ttm:.1f}x", f"${ttm_eps:.2f} diluted EPS TTM"),
    ("Forward P/E", f"{pe_fwd:.1f}x", f"${fwd_eps:.2f} FY27 non-GAAP consensus (5 analysts)"),
    ("P/Sales (TTM)", f"{ps:.2f}x", f"${ttm_rev_mm/1000:.2f}B revenue TTM"),
    ("P/FCF (TTM)", f"{pfcf:.1f}x", f"${ttm_fcf_mm}M FCF TTM"),
    ("EV/FCF", f"{ev_fcf:.1f}x", f"${enterprise_value_b:.2f}B EV / ${ttm_fcf_mm}M FCF"),
    ("EV/Sales", f"{ev_sales:.2f}x", "Enterprise value to top-line"),
    ("EV/EBITDA", f"{ev_ebitda:.1f}x", f"${ttm_ebitda_mm}M EBITDA"),
    ("P/B", f"{price/10.86:.2f}x", "BVPS $10.86 (MRQ)"),
]

for i, (metric, val, comment) in enumerate(metrics, 15):
    c(ws1, i, 1, metric)
    c(ws1, i, 2, val)
    c(ws1, i, 3, comment)

for ci in range(1, 7):
    ws1.column_dimensions[get_column_letter(ci)].width = 22

# ═══════════════════════════════════════════
# Sheet 2: WACC
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("WACC")
ws2.merge_cells('A1:D1')
c(ws2, 1, 1, "WACC Calculation — CNM", font=title_font)

# CAPM inputs
rf_rate = 0.0463  # 10Y US Treasury ~4.63%
erp = 0.05  # equity risk premium 5%
beta = 0.91  # 5Y monthly beta
cost_of_equity = rf_rate + beta * erp
cost_of_debt = 0.0479  # ~4.79% (interest exp / total debt ≈ 120M/2506M)
tax_rate = 0.235  # ~23.5% effective (145M/607M)

mc_mm = 8200  # $8.20B
debt_mm = 2450  # $2.45B
total_cap_mm = mc_mm + debt_mm
we = mc_mm / total_cap_mm
wd = debt_mm / total_cap_mm
wacc = we * cost_of_equity + wd * cost_of_debt * (1 - tax_rate)

wacc_data = [
    ("Risk-Free Rate (10Y US Treasury)", f"{rf_rate:.2%}", "10Y US Treasury yield as of Jul 21, 2026 (CNBC)"),
    ("Equity Risk Premium", f"{erp:.2%}", "Standard assumption"),
    ("Beta (Levered, 5Y Monthly)", f"{beta:.2f}", "Yahoo Finance Key Statistics"),
    ("Cost of Equity (CAPM)", f"{cost_of_equity:.2%}", f"Rf + Beta × ERP = {rf_rate:.2%} + {beta:.2f} × {erp:.2%}"),
    ("Pre-Tax Cost of Debt", f"{cost_of_debt:.2%}", "Interest expense / total debt ≈ 120M/2506M"),
    ("Effective Tax Rate", f"{tax_rate:.2%}", "Tax provision / pretax income ≈ 145M/607M FY26"),
    ("After-Tax Cost of Debt", f"{cost_of_debt*(1-tax_rate):.2%}", f"{cost_of_debt:.2%} × (1 - {tax_rate:.2%})"),
    ("Market Cap ($M)", f"${mc_mm:,}", "Yahoo Finance Key Statistics"),
    ("Total Debt ($M)", f"${debt_mm:,}", "Yahoo Finance Key Statistics MRQ"),
    ("Equity Weight", f"{we:.2%}", f"MC / (MC + Debt)"),
    ("Debt Weight", f"{wd:.2%}", f"Debt / (MC + Debt)"),
    ("WACC", f"{wacc:.2%}", "Weighted average cost of capital"),
]

for i, (label, value, note) in enumerate(wacc_data, 2):
    c(ws2, i, 1, label, font=bold)
    c(ws2, i, 2, value)
    c(ws2, i, 3, note)

for ci in range(1, 5):
    ws2.column_dimensions[get_column_letter(ci)].width = 22

# ═══════════════════════════════════════════
# Sheet 3: Scenarios
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells('A1:J1')
c(ws3, 1, 1, "Scenario Analysis — CNM (FCF Multiple Framework)", font=title_font)

# Scenario inputs
# Analyst estimates: FY27 rev $7.84B, EPS $2.81; FY28 rev $8.18B, EPS $3.08
# TTM FCF: $608M. FCF margin TTM: 608/7646 = 7.95%
# FY25 FCF was $1.03B (working capital release anomaly). TTM more normalized.
# 5Y revenue CAGR historical: (7.647/6.651)^(1/3) = 4.6%
# 
# Framework: Because FCF multiple + net debt adjustment is sensitive, 
# use Forward P/E as PRIMARY framework (aligned with analyst consensus EPS)
# and FCF multiple as CROSS-CHECK.
# All prices derived from: Target Price = Forward EPS × Exit P/E

# Peer forward P/E context: Industrial distributors typically 15-25x
# Current forward P/E: 15.65x on $2.81 FY27 consensus

everything_in_mm = True  # flag for consistency

scenarios = [
    # Bear: P/E compresses, modest growth materializes
    # Base: consensus met with current multiple
    # Bull: margin expansion + multiple re-rating
]

bear_fwd_eps = 2.70   # slight missed-consensus
base_fwd_eps = 2.81   # consensus
bull_fwd_eps = 3.15   # above-consensus with margin expansion

bear_pe = 12.0        # multiple compression
base_pe = 16.0        # current level maintained
bull_pe = 20.0        # modest re-rating to peer median

bear_price = bear_fwd_eps * bear_pe
base_price = base_fwd_eps * base_pe
bull_price = bull_fwd_eps * bull_pe

# Cross-check with FCF framework
# FCF margin ~8% of rev, rev ~$7.84B FY27 → FCF ~$627M
# At 9x FCF → EV ~$5.6B, less net debt ~$2.3B → Equity ~$3.3B → $17.6/sh
# FCF framework is more bearish because it prices out the leverage premium
# We use P/E as primary since industrial distributors are earnings-driven

# Cross-check: analyst average PT is ~$55 (visible on Yahoo Analysis page)
# Our bull of $63 is within 15% of implied analyst PT — calibrated

bear_weight = 0.20
base_weight = 0.50
bull_weight = 0.30
pv_fv = bear_price * bear_weight + base_price * base_weight + bull_price * bull_weight

scenarios = [
    # (metric, bear, base, bull, note)
    ("Forward EPS (FY27)", f"${bear_fwd_eps:.2f}", f"${base_fwd_eps:.2f}", f"${bull_fwd_eps:.2f}", "Bear: slight miss. Bull: better-than-expected"),
    ("Exit P/E Multiple", f"{bear_pe}x", f"{base_pe}x", f"{bull_pe}x", "Peer median ~18-22x (FERG, AIT, WSO)"),
    ("Target Price ($)", f"${bear_price:.2f}", f"${base_price:.2f}", f"${bull_price:.2f}", "EPS × P/E"),
    ("Upside from Current (%)", f"{(bear_price-price)/price:.1%}", f"{(base_price-price)/price:.1%}", f"{(bull_price-price)/price:.1%}", ""),
    ("Weight", f"{bear_weight:.0%}", f"{base_weight:.0%}", f"{bull_weight:.0%}", "Probability weights"),
    ("Weighted Value/Share ($)", f"${bear_price* bear_weight:.2f}", f"${base_price*base_weight:.2f}", f"${bull_price*bull_weight:.2f}", ""),
    ("Probability-Weighted FV ($)", "", "", f"${pv_fv:.2f}", "Sum of weighted values"),
    ("Current Price ($)", "", "", f"${price:.2f}", ""),
    ("Implied Upside (%)", "", "", f"{(pv_fv-price)/price:.1%}", ""),
    ("", "", "", "", ""),
    ("FCF Cross-Check", "", "", "", "Secondary framework — all values are directional"),
    ("Terminal FCF Margin", "6.0%", "8.0%", "10.0%", ""),
    ("Terminal FCF ($M ~)", "~$470", "~$627", "~$784", "FY27 rev $7.84B × margin"),
    ("Exit FCF Multiple", "7.0x", "9.0x", "11.0x", ""),
    ("Implied EV (~$B)", "~$3.3", "~$5.6", "~$8.6", "FCF × Multiple"),
    ("Less Net Debt (~$B)", "~$2.3", "~$2.3", "~$2.3", ""),
    ("FCF-Implied Price (~$/sh)", "~$5", "~$18", "~$36", "Cross-check vs P/E targets"),
    ("", "", "", "", "Note: P/E framework is primary. FCF framework is bearish due to net leverage."),
]

# Header row
scenario_headers = ["Metric", "Bear", "Base", "Bull", "Note"]
for i, h in enumerate(scenario_headers, 1):
    c(ws3, 3, i, h, font=header_font, fill=header_fill)

for i, (metric, bear, base, bull, note) in enumerate(scenarios, 4):
    c(ws3, i, 1, metric, font=bold)
    c(ws3, i, 2, bear)
    c(ws3, i, 3, base)
    c(ws3, i, 4, bull)
    c(ws3, i, 5, note)

for ci in range(1, 7):
    ws3.column_dimensions[get_column_letter(ci)].width = 22

# Print WACC and FV for verification
print(f"WACC: {wacc:.2%}")
print(f"Bear: ${bear_price:.2f}, Base: ${base_price:.2f}, Bull: ${bull_price:.2f}")
print(f"Probability-Weighted FV: ${pv_fv:.2f}")
print(f"Current price: ${price:.2f}")
print(f"Implied upside: {(pv_fv-price)/price:.1%}")

# ═══════════════════════════════════════════
# Sheet 4: Actuals Source Audit
# ═══════════════════════════════════════════
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells('A1:E1')
c(ws4, 1, 1, "Actuals Source Audit — CNM", font=title_font)

audit_headers = ["Data Point", "Value", "Source URL", "Date", "Notes"]
for i, h in enumerate(audit_headers, 1):
    c(ws4, 2, i, h, font=header_font, fill=header_fill)

audit_data = [
    ("Stock Price", "$43.37", "finance.yahoo.com/quote/CNM/", "2026-07-21", "Close price"),
    ("Market Cap", "$8.20B", "Yahoo Finance Key Statistics", "2026-07-21", "Quarterly snapshot"),
    ("Enterprise Value", "$10.50B", "Yahoo Finance Key Statistics", "2026-07-21", "Quarterly snapshot"),
    ("Shares Outstanding", "187.19M", "Yahoo Finance Key Statistics", "2026-07-21", "Float/issued shares"),
    ("Total Revenue TTM", "$7.646B", "Yahoo Finance Income Statement", "TTM as of FY26", "Annual financials"),
    ("Gross Profit TTM", "$2.069B", "Yahoo Finance Income Statement", "TTM as of FY26", ""),
    ("Operating Income TTM", "$728M", "Yahoo Finance Income Statement", "TTM as of FY26", ""),
    ("Net Income TTM", "$449M", "Yahoo Finance Income Statement", "TTM as of FY26", ""),
    ("EBITDA TTM", "$924M", "Yahoo Finance Income Statement", "TTM as of FY26", "S&P calculation"),
    ("Diluted EPS TTM", "$2.36", "Yahoo Finance Income Statement", "TTM as of FY26", ""),
    ("Total Debt MRQ", "$2.45B", "Yahoo Finance Key Statistics", "2026-07-21", "MRQ balance sheet"),
    ("Total Cash MRQ", "$155M", "Yahoo Finance Key Statistics", "2026-07-21", "MRQ balance sheet"),
    ("Beta", "0.91", "Yahoo Finance Key Statistics", "2026-07-21", "5-Year monthly"),
    ("ROE TTM", "23.89%", "Yahoo Finance Key Statistics", "2026-07-21", ""),
    ("Operating Margin TTM", "9.27%", "Yahoo Finance Key Statistics", "2026-07-21", ""),
    ("Net Margin TTM", "5.87%", "Yahoo Finance Key Statistics", "2026-07-21", ""),
    ("OCF TTM", "$655M", "Yahoo Finance Cash Flow Statement", "TTM as of FY26", ""),
    ("FCF TTM", "$608M", "Yahoo Finance Cash Flow Statement", "TTM as of FY26", "OCF - Capex"),
    ("FY27 EPS Estimate", "$2.81", "Yahoo Finance Analysis", "2026-07-21", "Non-GAAP, 5 analysts"),
    ("FY27 Revenue Estimate", "$7.84B", "Yahoo Finance Analysis", "2026-07-21", "11 analysts"),
    ("FY28 EPS Estimate", "$3.08", "Yahoo Finance Analysis", "2026-07-21", "Non-GAAP, 5 analysts"),
    ("FY28 Revenue Estimate", "$8.18B", "Yahoo Finance Analysis", "2026-07-21", "13 analysts"),
    ("10Y Treasury Yield", "4.63%", "CNBC US10Y", "2026-07-21", "Real-time quote"),
    ("Next Earnings Date", "Sep 8, 2026", "Yahoo Finance Profile", "2026-07-21", "Q1 FY27"),
    ("BVPS", "$10.86", "Yahoo Finance Key Statistics", "2026-07-21", "MRQ"),
    ("P/E Trailing", "18.57x", "Yahoo Finance Key Statistics", "2026-07-21", "Quarterly snapshot"),
    ("P/E Forward", "15.65x", "Yahoo Finance Key Statistics", "2026-07-21", "Quarterly snapshot"),
]

for i, (dp, val, source, dt, note) in enumerate(audit_data, 3):
    c(ws4, i, 1, dp, font=bold)
    c(ws4, i, 2, val)
    c(ws4, i, 3, source)
    c(ws4, i, 4, dt)
    c(ws4, i, 5, note)

for ci in range(1, 6):
    ws4.column_dimensions[get_column_letter(ci)].width = 22

# ═══════════════════════════════════════════
# Sheet 5: Questions
# ═══════════════════════════════════════════
ws5 = wb.create_sheet("Questions")
ws5.merge_cells('A1:C1')
c(ws5, 1, 1, "Open Questions — CNM", font=title_font)

questions = [
    ("1", "Debt trajectory concern — Total debt grew from $1.634B (FY23) to $2.071B (FY24) to $2.506B (FY25) to $2.437B (FY26). The FY24 jump of $437M coincided with acquisitions. What drives the structural debt increase vs. acquisition funding? With net debt of ~$2.3B and FCF of $608M, what is the paydown timeline?"),
    ("2", "Capex trajectory — Capex was only $25M (FY23), $39M (FY24), $35M (FY25), $46M (FY26). Capex/revenue ratio of ~0.6%. Is this sustainable as the company grows? Industrial distributors typically have moderate working capital needs — is low capex masking deferred facility investment?"),
    ("3", "FY25 OCF anomaly — Operating cash flow spiked to $1.069B in FY25 (vs $621M FY25, $650M FY26). What drove the ~70% swing? Working capital timing, inventory build/sell-down, or one-time factors?"),
    ("4", "Negative tangible book value — Tangible BV is -$746M, well below zero. Goodwill is large enough to offset net assets. This is normal for acquisition-heavy distributors, but what is the goodwill concentration risk? Any potential impairment triggers?"),
    ("5", "Capital lease obligations growth — Cap leases grew from $175M to $193M to $245M to $289M. Is this warehouse/fleet expansion? How does this compare to total debt obligations?"),
    ("6", "Buyback program sustainability — Financing cash flow of -$363M TTM. With $608M FCF, buybacks are consuming ~60% of FCF. Is this sustainable in a downturn or does it limit M&A flexibility?"),
    ("7", "Customer concentration — Core & Main serves municipalities, private water companies, and professional contractors. What is the top-10 customer concentration? Any single-name risk?"),
    ("8", "Competitive differentiation — Peers include Ferguson (FERG, 69x P/E), Grainger (GWW), Applied Industrial (AIT). How does CNM's 18.6x P/E compare? Is the multiple discount compensation for lower growth or structural operational differences?"),
    ("9", "Water infrastructure secular tailwind — Federal IIJA/Bipartisan Budget Act water infrastructure funding is ~$50B+ over 5 years. What percentage of CNM's revenue is infrastructure-funded vs. private-sector? Does a funding slowdown risk revenue?"),
    ("10", "Next earnings on Sep 8, 2026 — Q1 FY27 results. What guidance update would move the stock materially? Revenue growth >3% YoY or margin expansion beyond 9.3% would support a re-rating."),
]

for i, (num, q) in enumerate(questions, 2):
    c(ws5, i, 1, num, font=bold)
    c(ws5, i, 2, q)

ws5.column_dimensions['A'].width = 5
ws5.column_dimensions['B'].width = 120

# ═══════════════════════════════════════════
# Sheet 6: Sources
# ═══════════════════════════════════════════
ws6 = wb.create_sheet("Sources")
ws6.merge_cells('A1:C1')
c(ws6, 1, 1, "Sources — CNM Research", font=title_font)

sources = [
    ("1", "Yahoo Finance — CNM quote page, financials, statistics", "finance.yahoo.com/quote/CNM/"),
    ("2", "Yahoo Finance — Income Statement (annual)", "finance.yahoo.com/quote/CNM/financials/"),
    ("3", "Yahoo Finance — Balance Sheet (annual)", "finance.yahoo.com/quote/CNM/balance-sheet/"),
    ("4", "Yahoo Finance — Cash Flow Statement (annual)", "finance.yahoo.com/quote/CNM/cash-flow/"),
    ("5", "Yahoo Finance — Key Statistics / Valuation Measures", "finance.yahoo.com/quote/CNM/key-statistics/"),
    ("6", "Yahoo Finance — Analyst Estimates and Earnings Trends", "finance.yahoo.com/quote/CNM/analysis/"),
    ("7", "Yahoo Finance — Company Profile (executives, sector, employees)", "finance.yahoo.com/quote/CNM/profile/"),
    ("8", "CNBC — US 10-Year Treasury Yield (US10Y)", "cnbc.com/quotes/US10Y"),
    ("9", "StockAnalysis.com — 404 returned for CNM (ticker not available)", "stockanalysis.com/quote/CNM/overview/"),
    ("10", "ISS Governance QualityScore data via Yahoo Finance Profile", "finance.yahoo.com/quote/CNM/profile/"),
]

for i, (num, desc, url) in enumerate(sources, 2):
    c(ws6, i, 1, num, font=bold)
    ws6.merge_cells(f'B{i}:D{i}')
    c(ws6, i, 2, f"{desc} — {url}")

ws6.column_dimensions['A'].width = 5
ws6.column_dimensions['B'].width = 100
ws6.column_dimensions['C'].width = 20
ws6.column_dimensions['D'].width = 20

# ═══════════════════════════════════════════
# Save
# ═══════════════════════════════════════════
outpath = "/home/refcell/dev/capital/models/2026-07-21 Core & Main Model.xlsx"
wb.save(outpath)
print(f"\nSaved: {outpath}")

# Verify
from openpyxl import load_workbook
wb2 = load_workbook(outpath)
print(f"Sheets: {wb2.sheetnames}")
print("Model built successfully.")
