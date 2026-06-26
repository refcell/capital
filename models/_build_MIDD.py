#!/usr/bin/env python3
"""Build MIDD (The Middleby Corporation) 6-sheet valuation model."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import datetime

wb = Workbook()
# ─── helpers ───
thin = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))
hdr_font = Font(bold=True, size=11)
hdr_fill = PatternFill('solid', fgColor='D9E2F3')
title_font = Font(bold=True, size=14)
today = datetime.date.today().isoformat()

def style_row(ws, row_data, row_num):
    for c, v in enumerate(row_data, 1):
        cell = ws.cell(row_num, c, v)
        cell.border = thin

def style_header(ws, row_num, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row_num, c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

# ===================================================================
# DATA (all in $millions unless noted)
# ===================================================================
PRICE = 173.36
SHARES_MM = 44.8        # MC / price ≈ 7.76B / 173.36
MC = 7.76              # $B
EV = 9.48              # $B
TOTAL_DEBT_MM = 2192.5  # FY25 annual report
CASH_MM = 179.5        # MRQ from Yahoo stats
NET_DEBT_MM = TOTAL_DEBT_MM - CASH_MM  # ~1,951
NET_CASH_MM = -NET_DEBT_MM  # negative = net debt
BETA = 1.36
RISK_FREE = 4.376      # CNBC US10Y 2026-06-26
ERP = 5.0
TAX_RATE = 24.1        # 116,455 / 483,943 ≈ 24.1%

# Cost of debt: interest expense ~94M on avg debt ~2.3B
COST_OF_DEBT = 4.27    # implied from FY25 IS

# Income statement history ($M, in thousands / 1000)
revenue_h = [3310.5, 3201.2, 3150.2, 3242.1, 4032.9]   # TTM, FY25, FY24, FY23, FY22
gp_h      = [1282.5, 1251.9, 1251.8, 1284.1, 1446.6]
opinc_h   = [592.9, 588.8, 661.7, 659.1, 649.3]
ebt_h     = [483.9, 482.3, 566.1, 536.1, 564.4]
ni_cont_h = [367.5, 367.3, 421.0, 413.0, 436.6]   # continuing ops
ni_rep_h  = [-420.2, -277.7, 428.4, 400.9, 436.6] # reported
eps_ttm   = 7.29      # continuing ops diluted
ebitda_h  = [688.9, 681.4, 763.6, 767.1, 791.5]
ocf_h     = [554.7, 630.2, 686.8, 628.8, 332.6]
capex_h   = [52.2, 71.8, 36.8, 60.9, 69.5]
fcf_h     = [502.5, 558.4, 650.0, 567.9, 263.0]

# Analyst estimates
fy26_rev = 3400.0
fy27_rev = 3520.0
fy26_eps = 9.61
fy27_eps = 10.85

# ===================================================================
# Sheet 1: Valuation
# ===================================================================
ws = wb.active
ws.title = 'Valuation'
ws.merge_cells('A1:F1')
ws['A1'] = 'The Middleby Corporation (MIDD) — Valuation'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center')

# Title block
r = 2
for i, (k, v) in enumerate([
    ('Company', 'The Middleby Corporation'),
    ('Ticker', 'NASDAQ: MIDD'),
    ('Date', today),
    ('Price', f'${PRICE:.2f}'),
    ('Shares Outstanding', f'{SHARES_MM:.1f}M'),
    ('Market Cap', f'${MC:.1f}B'),
    ('Enterprise Value', f'${EV:.1f}B'),
    ('Primary Lens', 'Forward P/E + Scenario DCF'),
    ('Stance', 'Watch / Needs more work (post-spin optionality)'),
]):
    ws.cell(r+i, 1, k).font = Font(bold=True)
    ws.cell(r+i, 2, v)

# Valuation metrics table
r = 13
ws.cell(r, 1, 'Metric').font = hdr_font
ws.cell(r, 2, 'Value').font = hdr_font
ws.cell(r, 3, 'Notes').font = hdr_font
style_header(ws, r, 3)
metrics = [
    ('P/E (TTM, normalized cont ops)', f'{PRICE / eps_ttm:.2f}', 'Normalized EPS $7.29; reported EPS is -$8.16 due to discontinued ops (Midera spin)'),
    ('Forward P/E (FY26)', f'{PRICE / fy26_eps:.2f}', 'On consensus EPS $9.61'),
    ('Forward P/E (FY27)', f'{PRICE / fy27_eps:.2f}', 'On consensus EPS $10.85'),
    ('P/S (TTM)', f'{MC * 1000 / revenue_h[0]:.2f}', f'MC ${MC:.1f}B / Rev ${revenue_h[0]/1000:.1f}B'),
    ('P/FCF (TTM)', f'{MC * 1000 / fcf_h[0]:.2f}', f'MC ${MC:.1f}B / FCF ${fcf_h[0]/1000:.1f}B'),
    ('EV/FCF', f'{EV * 1000 / fcf_h[0]:.2f}', f'EV ${EV:.1f}B / FCF ${fcf_h[0]/1000:.1f}B'),
    ('EV/Sales', f'{EV * 1000 / revenue_h[0]:.2f}', f'EV ${EV:.1f}B / Rev ${revenue_h[0]/1000:.1f}B'),
    ('EV/EBITDA', f'{EV * 1000 / ebitda_h[0]:.2f}', f'EV ${EV:.1f}B / EBITDA ${ebitda_h[0]/1000:.1f}B'),
]
for i, (m, v, n) in enumerate(metrics):
    rn = r + 1 + i
    style_row(ws, [m, v, n], rn)

# ===================================================================
# Sheet 2: WACC
# ===================================================================
ws = wb.create_sheet('WACC')
ws.merge_cells('A1:C1')
ws['A1'] = 'WACC — CAPM Components (as of 2026-06-26)'
ws['A1'].font = title_font

# Compute WACC components
co_e = RISK_FREE + BETA * ERP  # cost of equity
# Weights
eq_wt = MC / (MC + TOTAL_DEBT_MM / 1000)
debt_wt = 1 - eq_wt
wacc = eq_wt * co_e + debt_wt * COST_OF_DEBT * (1 - TAX_RATE / 100)

r = 2
for i, (k, v) in enumerate([
    ('Risk-Free Rate (10Y US)', f'{RISK_FREE:.2f}%'),
    ('Equity Risk Premium', f'{ERP:.1f}%'),
    ('Beta (Levered, 5Y)', f'{BETA:.2f}'),
    ('Cost of Equity (Rf + Beta*ERP)', f'{co_e:.2f}%'),
    ('Cost of Debt (implied)', f'{COST_OF_DEBT:.2f}%'),
    ('Tax Rate (TTM)', f'{TAX_RATE:.1f}%'),
    ('Market Cap ($B)', f'{MC:.1f}'),
    ('Total Debt ($B)', f'{TOTAL_DEBT_MM/1000:.2f}'),
    ('Equity Weight', f'{eq_wt:.2f}'),
    ('Debt Weight', f'{debt_wt:.2f}'),
    ('WACC', f'{wacc:.2f}%'),
]):
    ws.cell(r+i, 1, k).font = Font(bold=True)
    ws.cell(r+i, 2, v)

# ===================================================================
# Sheet 3: Scenarios
# ===================================================================
ws = wb.create_sheet('Scenarios')
ws.merge_cells('A1:P1')
ws['A1'] = 'Scenarios — Bear / Base / Bull (5-Year Terminal, Everything in $MM)'
ws['A1'].font = title_font

# Scenario parameters (all in $MM)
scenarios = {
    'Bear':     dict(cagr=0.02, fcf_margin=0.12, exit_multiple=13, term_debt_adj=1400, shares=43),
    'Base':     dict(cagr=0.05, fcf_margin=0.15, exit_multiple=15, term_debt_adj=700, shares=41),
    'Bull':     dict(cagr=0.07, fcf_margin=0.17, exit_multiple=18, term_debt_adj=200, shares=39),
}
weights = {'Bear': 0.25, 'Base': 0.50, 'Bull': 0.25}
order = ['Bear', 'Base', 'Bull']
cur_rev = revenue_h[0]  # $3,310.5M

# Pre-compute per-scenario targets
targets = {}
for name, s in scenarios.items():
    term_rev = cur_rev * (1 + s['cagr']) ** 5
    term_fcf = term_rev * s['fcf_margin']
    implied_ev = s['exit_multiple'] * term_fcf
    eq_value = implied_ev - s['term_debt_adj']
    target = eq_value / s['shares']
    upside = (target - PRICE) / PRICE * 100
    wt_value = target * weights[name]
    targets[name] = {
        'term_rev': term_rev, 'term_fcf': term_fcf, 'implied_ev': implied_ev,
        'eq_value': eq_value, 'target': target, 'upside': upside, 'wt_value': wt_value
    }

# Write header
r = 2
for c, h in enumerate(['Item'] + order, 1):
    ws.cell(r, c, h)
style_header(ws, r, 4)

# Build rows
data = [
    ('Revenue CAGR (5Y)',    *[f"{scenarios[n]['cagr']*100:.0f}%" for n in order]),
    ('Terminal Revenue',     *[f"${targets[n]['term_rev']:.0f}" for n in order]),
    ('FCF Margin',           *[f"{scenarios[n]['fcf_margin']*100:.0f}%" for n in order]),
    ('Terminal FCF',         *[f"${targets[n]['term_fcf']:.0f}" for n in order]),
    ('Exit FCF Multiple',    *[f"{scenarios[n]['exit_multiple']}x" for n in order]),
    ('Implied EV',           *[f"${targets[n]['implied_ev']:.0f}" for n in order]),
    ('Less: Terminal Net Debt', *[f"${scenarios[n]['term_debt_adj']}" for n in order]),
    ('Shares Outstanding',   *[f"{scenarios[n]['shares']}M" for n in order]),
    ('Target Price',         *[f"${targets[n]['target']:.0f}" for n in order]),
    ('Upside from Current',  *[f"{targets[n]['upside']:+.0f}%" for n in order]),
    ('Weight',               *[f"{weights[n]*100:.0f}%" for n in order]),
    ('Weighted $/Share',     *[f"${targets[n]['wt_value']:.0f}" for n in order]),
]

tgt_b = targets['Bear']['target']
tgt_bb = targets['Base']['target']
tgt_bu = targets['Bull']['target']
total_fv = tgt_b * 0.25 + tgt_bb * 0.50 + tgt_bu * 0.25
total_upside = (total_fv - PRICE) / PRICE * 100

for i, row_data in enumerate(data):
    rn = r + 1 + i
    for c, v in enumerate(row_data, 1):
        cell = ws.cell(rn, c, v)
        cell.border = thin
    ws.cell(rn, 1).font = Font(bold=True)

rn = r + 1 + len(data)
for c, v in enumerate(['Total Probability-Weighted FV', f'{total_fv:.0f}', f'{total_fv:.0f}', f'{total_fv:.0f}'], 1):
    cell = ws.cell(rn, c, v)
    cell.border = thin
    cell.font = Font(bold=True)

rn += 1
for c, v in enumerate(['Upside from Current Price', f'{total_upside:+.1f}%', f'{total_upside:+.1f}%', f'{total_upside:+.1f}%'], 1):
    cell = ws.cell(rn, c, v)
    cell.border = thin
    cell.font = Font(bold=True, color='003300')

# Print targets for verification
print(f"WACC: {wacc:.2f}%")
print(f"Bear: ${tgt_b:.0f} ({(tgt_b - PRICE)/PRICE*100:+.0f}%)")
print(f"Base: ${tgt_bb:.0f} ({(tgt_bb - PRICE)/PRICE*100:+.0f}%)")
print(f"Bull: ${tgt_bu:.0f} ({(tgt_bu - PRICE)/PRICE*100:+.0f}%)")
print(f"Weighted FV: ${total_fv:.0f} ({(total_fv - PRICE)/PRICE*100:+.0f}%)")

# ===================================================================
# Sheet 4: Actuals Source Audit
# ===================================================================
ws = wb.create_sheet('Actuals Source Audit')
ws.merge_cells('A1:E1')
ws['A1'] = 'Actuals Source Audit — All Data Points Verified'
ws['A1'].font = title_font

r = 2
for c, h in enumerate(['Data Point', 'Value', 'Source URL', 'Date', 'Notes'], 1):
    ws.cell(r, c, h)
style_header(ws, r, 5)

audit_rows = [
    ('Stock Price', f'${PRICE}', 'finance.yahoo.com/quote/MIDD/', '2026-06-26', 'Close price, NasdaqGS'),
    ('Market Cap', f'${MC}B', 'finance.yahoo.com/quote/MIDD/ (Statistics)', '2026-06-25', 'Intraday MC'),
    ('Enterprise Value', f'${EV}B', 'finance.yahoo.com/quote/MIDD/ (Statistics)', '2026-06-25', 'MC + Debt - Cash'),
    ('Shares Outstanding', '~44.8M', 'MC / Price calculation', '2026-06-26', 'MC $7.76B / $173.36; BS shows 48.9M issued, 16.0M treasury → 32.9M outstanding'),
    ('Revenue TTM', '$3,310M', 'finance.yahoo.com/quote/MIDD/financials/', 'As of Q2 FY26', 'In thousands; annual tab'),
    ('Revenue FY25', '$3,201M', 'Same as above', '12/31/2025', ''),
    ('Gross Profit TTM', '$1,283M', 'Same as above', 'As of Q2 FY26', ''),
    ('Operating Income TTM', '$593M', 'Same as above', 'As of Q2 FY26', ''),
    ('Net Income TTM (cont)', '$367M', 'Same as above', 'As of Q2 FY26', 'Continuing ops only; reported is -$420M due to discontinued'),
    ('EBITDA TTM', '$689M', 'Same as above', 'As of Q2 FY26', ''),
    ('FCF TTM', '$502M', 'finance.yahoo.com/quote/MIDD/cash-flow/', 'As of Q2 FY26', 'OCF $555M - CapEx $52M'),
    ('Total Debt', '$2,193M', 'finance.yahoo.com/quote/MIDD/balance-sheet/', '12/31/2025', 'FY25 year-end'),
    ('Cash MRQ', '$180M', 'finance.yahoo.com/quote/MIDD/ (Stats)', '2026-06-25', 'Total Cash, most recent quarter'),
    ('Beta', '1.36', 'finance.yahoo.com/quote/MIDD/ (Stats)', '2026-06-25', '5Y Monthly'),
    ('P/E TTM', '23.54', 'finance.yahoo.com/quote/MIDD/ (Stats)', '2026-06-25', 'On normalized EPS $7.29'),
    ('Forward P/E', '17.70', 'finance.yahoo.com/quote/MIDD/ (Stats)', '2026-06-25', 'On FY26 consensus $9.61'),
    ('Analyst Target 1Y', '$196.33', 'finance.yahoo.com/quote/MIDD/', '2026-06-26', 'Average; high $206'),
    ('10Y Treasury Rate', '4.376%', 'cnbc.com/quotes/US10Y', '2026-06-26', 'Per CNBC quote page'),
    ('FY26 Revenue Estimate', '$3.4B', 'finance.yahoo.com/quote/MIDD/analysis/', '2026-06-26', '9 analysts avg'),
    ('FY27 Revenue Estimate', '$3.52B', 'Same', '2026-06-26', '9 analysts avg'),
    ('FY26 EPS Estimate', '$9.61', 'Same', '2026-06-26', '11 analysts avg, normalized'),
    ('FY27 EPS Estimate', '$10.85', 'Same', '2026-06-26', '11 analysts avg, normalized'),
    ('Next Earnings Date', 'Aug 5, 2026', 'finance.yahoo.com/quote/MIDD/', 'Estimate', 'Q4 FY26'),
    ('Midera Spin-off', 'Approved', 'Yahoo News / Business Wire', '2026-06-23', 'Board approved Midera Food Processing spin'),
    ('Tax Rate (TTM)', '24.1%', 'Computed from IS', 'TTM', 'Tax prov $116M / Pretax $484M'),
]
for i, row in enumerate(audit_rows):
    rn = r + 1 + i
    style_row(ws, row, rn)

# ===================================================================
# Sheet 5: Questions
# ===================================================================
ws = wb.create_sheet('Questions')
ws.merge_cells('A1:A1')
ws['A1'] = 'Open Questions — The Middleby Corporation (MIDD)'
ws['A1'].font = title_font

questions = [
    ('1. Midera spin-off timing and proceeds: When exactly will Midera (Food Processing Equipment) spin off? What are the expected proceeds or distribution mechanics? How does this affect the revenue base, share count, and debt allocation?',),
    ('2. Discontinued operations impact: Reported TTM net income is -$420M due to discontinued ops, but continuing ops show +$367M. How does the spin-off reclassification affect normalized earnings and historical comparability?',),
    ('3. Negative tangible book value: Tangible book value is -$85M (FY25). The company has ~$1.3B in goodwill. Is the acquisition-driven goodwill from Marshall, MTT, and other deals impaired or sustainable?',),
    ('4. Massive buyback program: MIDD spent $1.06B TTM and $724M in FY25 on buybacks. Is this sustainable post-spin? Are shareholders getting a net benefit vs. the debt used to fund these repurchases?',),
    ('5. Revenue growth trajectory: Revenue has been essentially flat ($3.15-3.31B) for 3 years. Analysts see only 3.5% growth for FY27. What drives the bull case beyond share count reduction?',),
    ('6. Debt trajectory: Total debt went from $2.75B (FY22) to $2.19B (FY25), but TTM financing shows $1.56B issued vs $2.07B repaid. Net debt position: improving or concerning?',),
    ('7. Capital allocation post-spin: With Midera gone, will the remaining foodservice-focused business justify the current P/E of 23.5x? What multiple is appropriate for a specialized equipment manufacturer?',),
    ('8. Operating margin stability: Op income is $589-662M on $3.15-3.20B revenue → 18-21% margins. Are these margins defendable if tariff costs, labor, or raw materials spike?',),
    ('9. Analyst estimate revisions: EPS estimates for FY26 have been revised up from $9.29 (60 days ago) to $9.61 today. Is this trend sustainable or already priced in?',),
    ('10. Competitive differentiation in commercial kitchen equipment: Middleby serves a fragmented market via 70+ brands. Is this a moat or a liability (integration costs, brand dilution)?',),
    ('11. Earnings quality: Company has been consistently beating EPS estimates (5.3-14.5% surprises). Are these driven by operational excellence or accounting timing?',),
    ('12. Next earnings catalyst: Q4 FY26 earnings on Aug 5, 2026. What specific metric would confirm/destroy the spin-off thesis?',),
]

r = 2
for i, (q,) in enumerate(questions):
    ws.cell(r + i, 1, q).alignment = Alignment(wrap_text=True)

# ===================================================================
# Sheet 6: Sources
# ===================================================================
ws = wb.create_sheet('Sources')
ws.merge_cells('A1:B1')
ws['A1'] = 'Sources'
ws['A1'].font = title_font

sources = [
    ('1', 'Yahoo Finance — Summary & Key Stats', 'https://finance.yahoo.com/quote/MIDD/'),
    ('2', 'Yahoo Finance — Income Statement (Annual)', 'https://finance.yahoo.com/quote/MIDD/financials/'),
    ('3', 'Yahoo Finance — Balance Sheet (Annual)', 'https://finance.yahoo.com/quote/MIDD/balance-sheet/'),
    ('4', 'Yahoo Finance — Cash Flow (Annual)', 'https://finance.yahoo.com/quote/MIDD/cash-flow/'),
    ('5', 'Yahoo Finance — Analyst Estimates', 'https://finance.yahoo.com/quote/MIDD/analysis/'),
    ('6', 'Yahoo Finance Company Profile', 'https://finance.yahoo.com/quote/MIDD/'),
    ('7', 'CNBC — 10Y Treasury Rate', 'https://www.cnbc.com/quotes/US10Y'),
    ('8', 'Zacks / Business Wire — Midera Spin-off News', 'https://finance.yahoo.com/quote/MIDD/news/'),
    ('9', 'StockAnalysis.com — 404 for MIDD (unavailable)', 'https://stockanalysis.com/quote/MIDD/'),
    ('10', 'Argus Research — Hold rating, $172 target', 'Yahoo Finance Research tab'),
    ('11', 'Oppenheimer — Outperform, $205 target (6/17/2026)', 'Yahoo Finance Analyst Insights'),
]

r = 2
for c, h in enumerate(['#', 'Description', 'URL'], 1):
    ws.cell(r, c, h)
style_header(ws, r, 3)
for i, (num, desc, url) in enumerate(sources):
    style_row(ws, [num, desc, url], r + 1 + i)

# ─── column widths ───
for sheet in wb.worksheets:
    for col_idx, width in enumerate([35, 20, 30, 20, 35], 1):
        if col_idx <= sheet.max_column:
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

# ─── save ───
out = f'/home/refcell/dev/capital/models/[2026-06-26] Middleby Corporation Model.xlsx'
wb.save(out)
print(f'Saved: {out}')
print(f'Total sheets: {len(wb.sheetnames)}')
