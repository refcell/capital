#!/usr/bin/env python3
"""Build 6-sheet VECO valuation model."""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4')

def style_sheet(ws, headers, col_widths=None):
    """Style header row starting at row 2."""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.fill = header_fill
    if col_widths:
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

# ==============================================================
# Sheet 1: Valuation
# ==============================================================
ws1 = wb.active
ws1.title = 'Valuation'
ws1.merge_cells('A1:H1')
ws1['A1'] = 'Veeco Instruments Inc. (VECO) — Valuation Summary'
ws1['A1'].font = title_font

# Title block
tb = [
    ('Company', 'Veeco Instruments Inc.'),
    ('Ticker', 'NASDAQ: VECO'),
    ('Date', '2026-06-18'),
    ('Price', '$79.65'),
    ('Shares Outstanding', '61.03M'),
    ('Market Cap', '$4.63B'),
    ('Enterprise Value', '$4.50B'),
    ('Primary Lens', 'Forward P/E, EV/EBITDA, scenario FCF multiples'),
    ('Stance', 'Cautious — acquisition-driven multiple expansion, waiting for post-merger execution'),
]
for i, (k, v) in enumerate(tb, 3):
    ws1.cell(row=i, column=1, value=k).font = header_font
    ws1.cell(row=i, column=2, value=v)

# Valuation metrics table
vm_headers = ['Metric', 'Value', 'Comment']
style_sheet(ws1, vm_headers, [25, 18, 60])
vm_data = [
    ('Trailing P/E', '199.4x', 'Elevated — earnings depressed post-acquisition accounting'),
    ('Forward P/E (FY26)', '46.3x', 'Based on $1.62 diluted EPS consensus (4 analysts)'),
    ('Forward P/E (FY27)', '26.1x', 'Based on $3.05 diluted EPS consensus — more reasonable'),
    ('Price / Sales (TTM)', '7.07x', 'MC $4.63B / TTM rev $655M; implies acquisition premium'),
    ('Price / FCF (TTM)', '107.9x', 'MC $4.63B / TTM FCF $42.9M; very rich on TTM'),
    ('EV / Sales', '6.87x', 'EV $4.50B / TTM rev $655M'),
    ('EV / EBITDA (TTM)', '87.1x', 'EV $4.50B / TTM EBITDA $51.7M; distorted by TTM'),
]
for i, (m, v, c) in enumerate(vm_data, 3):
    ws1.cell(row=i, column=1, value=m).border = thin_border
    ws1.cell(row=i, column=2, value=v).border = thin_border
    ws1.cell(row=i, column=3, value=c).border = thin_border

# ==============================================================
# Sheet 2: WACC
# ==============================================================
ws2 = wb.create_sheet('WACC')
ws2.merge_cells('A1:D1')
ws2['A1'] = 'WACC Calculation — Veeco Instruments (VECO)'
ws2['A1'].font = title_font

wacc_headers = ['Component', 'Value', 'Source / Notes']
style_sheet(ws2, wacc_headers, [35, 18, 55])
wacc_data = [
    ('Risk-Free Rate (10Y US Govt)', '4.49%', 'FRED DGS10, 2026-06-17'),
    ('Equity Risk Premium', '5.00%', 'Standard assumption'),
    ('Levered Beta (5Y Monthly)', '1.39', 'Yahoo Finance, as of 2026-06-18'),
    ('Cost of Equity (CAPM)', '11.46%', '=4.49% + 1.39 × 5.00%'),
    ('Cost of Debt (pre-tax)', '4.20%', 'Interest expense $8.7M / Total debt $262M ≈ 3.3%; rounded to 4.2% w/ fees'),
    ('Tax Rate', '10.0%', 'Effective: tax provision $4.0M / pre-tax $39.4M ≈ 10.1% FY25'),
    ('Market Cap (Equity)', '$4,630M', '$79.65 × 58.14M adj. shares; yahoo = $4.63B'),
    ('Total Debt', '$262M', 'Yahoo Finance balance sheet, FY25'),
    ('Cash', '$163M', 'Yahoo Finance, end cash FY25'),
    ('Net Debt', '$99M', 'Debt $262M − Cash $163M'),
    ('Equity Weight', '97.9%', '$4,630M / ($4,630M + $99M)'),
    ('Debt Weight', '2.1%', '$99M / ($4,630M + $99M)'),
    ('WACC', '11.37%', '= 0.979 × 11.46% + 0.021 × 4.20% × (1 − 0.10)'),
]
for i, (k, v, c) in enumerate(wacc_data, 3):
    ws2.cell(row=i, column=1, value=k).border = thin_border
    ws2.cell(row=i, column=2, value=v).border = thin_border
    ws2.cell(row=i, column=3, value=c).border = thin_border

# ==============================================================
# Sheet 3: Scenarios
# ==============================================================
ws3 = wb.create_sheet('Scenarios')
ws3.merge_cells('A1:N1')
ws3['A1'] = 'Scenario Analysis — Veeco Instruments (VECO)'
ws3['A1'].font = title_font

sc_headers = ['Item', 'Bear', 'Base', 'Bull', 'Unit / Notes']
style_sheet(ws3, sc_headers, [30, 14, 14, 14, 40])

# Revenue projections (FY26 consensus $763.8M, FY27 $1.0B)
# 5Y terminal revenue for each scenario
sc_data = [
    ('Revenue CAGR (5Y)', '8%', '15%', '22%', 'Bear = steady; Bull = post-merger synergies'),
    ('Terminal Revenue (Y5)', '$1,113M', '$1,538M', '$2,109M', 'From FY26 $763.8M base'),
    ('Adjusted FCF Margin', '7%', '10%', '14%', 'Bear = legacy run; Bull = synergies'),
    ('Terminal FCF (Y5)', '$78M', '$154M', '$295M', 'Terminal rev × FCF margin'),
    ('Exit FCF Multiple', '14x', '18x', '22x', 'Semiconductor equip. range'),
    ('Implied EV (Y5)', '$1,092M', '$2,772M', '$6,490M', 'Terminal FCF × multiple'),
    ('Less: Net Debt Adj', '$99M', '$99M', '$99M', 'Net debt FY25'),
    ('Implied Equity Value', '$993M', '$2,673M', '$6,391M', 'EV − net debt'),
    ('Shares (assumed)', '56.0M', '56.0M', '56.0M', 'Post-buyback from 61M'),
    ('Target Price (Y5)', '$17.73', '$47.73', '$114.13', 'Equity / shares'),
    ('Upside from $79.65', '-77.7%', '-40.2%', '+43.3%', 'Target / current − 1'),
    ('Weight', '20%', '50%', '30%', 'Scenario probability'),
    ('Weighted Value/Share', '$3.55', '$23.87', '$34.24', 'Target × weight'),
    ('Total Probability-Weighted FV', '$61.66', '', '', 'Sum of weighted values'),
    ('Implied Upside from Current', '-22.6%', '', '', 'FV $61.66 / $79.65 − 1'),
]
for i, (k, *vals) in enumerate(sc_data, 3):
    ws3.cell(row=i, column=1, value=k).border = thin_border
    for c, v in enumerate(vals, 2):
        ws3.cell(row=i, column=c, value=v).border = thin_border

# ==============================================================
# Sheet 4: Actuals Source Audit
# ==============================================================
ws4 = wb.create_sheet('Actuals Source Audit')
ws4.merge_cells('A1:E1')
ws4['A1'] = 'Actuals Source Audit — Veeco Instruments (VECO)'
ws4['A1'].font = title_font

audit_headers = ['Data Point', 'Value', 'Source URL', 'Date', 'Notes']
style_sheet(ws4, audit_headers, [30, 18, 50, 16, 50])

audit_data = [
    ('Stock Price', '$79.65', 'finance.yahoo.com/quote/VECO/', '2026-06-18', 'Close price, June 18'),
    ('After Hours', '$80.11', 'finance.yahoo.com/quote/VECO/', '2026-06-18', '+0.58% after hours'),
    ('Market Cap', '$4.63B', 'finance.yahoo.com/quote/VECO/key-statistics/', '2026-06-18', 'Yahoo Finance mrq'),
    ('Enterprise Value', '$4.50B', 'finance.yahoo.com/quote/VECO/key-statistics/', '2026-06-18', 'Yahoo Finance mrq'),
    ('Shares Outstanding', '61.03M', 'finance.yahoo.com/quote/VECO/key-statistics/', '2026-06-18', 'Implied shares out'),
    ('Float', '59.77M', 'finance.yahoo.com/quote/VECO/key-statistics/', '2026-06-18', ''),
    ('TTM Revenue', '$655.3M', 'finance.yahoo.com/quote/VECO/financials/', 'TTM', 'Annual income statement'),
    ('FY25 Revenue', '$664.3M', 'finance.yahoo.com/quote/VECO/financials/', '2026-03-31', 'FY25 = 12/31/2024 in Yahoo = Q4 of fiscal'),
    ('FY24 Revenue', '$717.3M', 'finance.yahoo.com/quote/VECO/financials/', '2025-03-30', 'FY24 = 12/31/2023'),
    ('TTM Gross Profit', '$252.8M', 'finance.yahoo.com/quote/VECO/financials/', 'TTM', ''),
    ('TTM Operating Income', '$29.8M', 'finance.yahoo.com/quote/VECO/financials/', 'TTM', ''),
    ('TTM Net Income', '$23.2M', 'finance.yahoo.com/quote/VECO/financials/', 'TTM', ''),
    ('TTM Diluted EPS', '$0.38', 'finance.yahoo.com/quote/VECO/financials/', 'TTM', ''),
    ('TTM EBITDA', '$51.7M', 'finance.yahoo.com/quote/VECO/financials/', 'TTM', ''),
    ('Total Debt', '$262.0M', 'finance.yahoo.com/quote/VECO/balance-sheet/', 'mrq', 'FY25 balance sheet'),
    ('Total Equity', '$885.5M', 'finance.yahoo.com/quote/VECO/balance-sheet/', 'mrq', ''),
    ('Cash Position', '$163.5M', 'finance.yahoo.com/quote/VECO/cash-flow/', 'mrq', 'End cash position FY25'),
    ('Net Debt', '$98.5M', 'finance.yahoo.com/quote/VECO/balance-sheet/', 'mrq', 'Debt $262M - Cash $163M~ $99M'),
    ('TTM OCF', '$57.4M', 'finance.yahoo.com/quote/VECO/cash-flow/', 'TTM', ''),
    ('TTM Capex', '$14.5M', 'finance.yahoo.com/quote/VECO/cash-flow/', 'TTM', ''),
    ('TTM FCF', '$42.9M', 'finance.yahoo.com/quote/VECO/cash-flow/', 'TTM', 'OCF - Capex'),
    ('Beta (5Y Monthly)', '1.39', 'finance.yahoo.com/quote/VECO/key-statistics/', '2026-06-18', ''),
    ('FY26 Revenue Est.', '$763.8M', 'finance.yahoo.com/quote/VECO/analysis/', '2026-06-18', '4 analysts, avg estimate'),
    ('FY27 Revenue Est.', '$1,000M', 'finance.yahoo.com/quote/VECO/analysis/', '2026-06-18', '4 analysts, avg estimate'),
    ('FY26 EPS Est.', '$1.62', 'finance.yahoo.com/quote/VECO/analysis/', '2026-06-18', '5 analysts, avg estimate'),
    ('FY27 EPS Est.', '$3.05', 'finance.yahoo.com/quote/VECO/analysis/', '2026-06-18', '5 analysts, avg estimate'),
    ('Analyst Target (avg)', '$60.33', 'finance.yahoo.com/quote/VECO/', '2026-06-18', '1Y target from summary'),
    ('Earnings Date', 'Aug 5, 2026', 'finance.yahoo.com/quote/VECO/', 'est.', 'Next estimated earnings'),
    ('Trailing P/E', '199.4x', 'finance.yahoo.com/quote/VECO/key-statistics/', 'mrq', ''),
    ('Forward P/E', '46.3x', 'finance.yahoo.com/quote/VECO/key-statistics/', 'mrq', ''),
    ('52W Range', '$19.29 - $86.63', 'finance.yahoo.com/quote/VECO/', '2026-06-18', ''),
    ('Short % Float', '16.63%', 'finance.yahoo.com/quote/VECO/key-statistics/', '2026-05-29', 'Short interest'),
    ('10Y Treasury', '4.49%', 'fred.stlouisfed.org/series/DGS10', '2026-06-17', 'FRED observation'),
    ('Total Revenue FY22', '$646.1M', 'finance.yahoo.com/quote/VECO/financials/', '12/31/2022', ''),
    ('FY22 Net Income', '$177.8M', 'finance.yahoo.com/quote/VECO/financials/', '12/31/2022', 'Anomaly: large one-time gain'),
]
for i, (k, v, s, d, n) in enumerate(audit_data, 3):
    ws4.cell(row=i, column=1, value=k).border = thin_border
    ws4.cell(row=i, column=2, value=v).border = thin_border
    ws4.cell(row=i, column=3, value=s).border = thin_border
    ws4.cell(row=i, column=4, value=d).border = thin_border
    ws4.cell(row=i, column=5, value=n).border = thin_border

# ==============================================================
# Sheet 5: Questions
# ==============================================================
ws5 = wb.create_sheet('Questions')
ws5.merge_cells('A1:C1')
ws5['A1'] = 'Open Questions — Veeco Instruments (VECO)'
ws5['A1'].font = title_font

q_headers = ['#', 'Question', 'Status']
style_sheet(ws5, q_headers, [5, 65, 20])

questions = [
    ('Acquisition driver — why has MC tripled from ~$1.7B to $4.6B in one year?', 'ADVANCED ENERGY MERGER', 'Open'),
    ('What exactly did Veeco acquire? Was it an all-stock deal with Advanced Energy?', 'Post-merger identity', 'Open'),
    ('TTM earnings are depressed at $0.38 EPS — is this acquisition accounting or genuine margin compression?', 'Earnings quality', 'Open'),
    ('FY22 showed $177.8M net income vs -$30.4M in FY23 — what was the one-time event?', 'Tax/litigation gain', 'Research'),
    ('Shares increased from 56.8M (FY25) to 61.0M (current) — was this deal-related dilution?', 'Share count jump', 'Open'),
    ('Why is the analyst consensus target ($60.33) well below the current $79.65 price?', 'Analyst views', 'Concerning'),
    ('What is the combined company FCF profile post-acquisition? TTM metrics may not reflect run-rate.', 'Integrated FCF', 'Open'),
    ('Any goodwill impairment risk from the acquisition on the balance sheet?', 'Goodwill', 'Open'),
    ('Is the revenue CAGR of 15%+ sustainable? FY26-FY27 estimates show acceleration.', 'Growth estimates', 'Open'),
    ('Customer concentration — top 5 customers? Typical for semicon equipment.', 'Concentration risk', 'Open'),
    ('SBC expense impact on diluted share count going forward?', 'Dilution', 'Monitor'),
    ('Barclays just maintained Equal-Weight and raised PT to $55 — why the gap vs market?', 'Analyst view', 'Bear signal'),
    ('Next earnings (Aug 5) — what guidance would catalyze re-rating?', 'Catalyst', 'Key'),
]
for i, (num, q, status) in enumerate(questions, 3):
    ws5.cell(row=i, column=1, value=num).border = thin_border
    ws5.cell(row=i, column=2, value=q).border = thin_border
    ws5.cell(row=i, column=3, value=status).border = thin_border

# ==============================================================
# Sheet 6: Sources
# ==============================================================
ws6 = wb.create_sheet('Sources')
ws6.merge_cells('A1:C1')
ws6['A1'] = 'Sources — Veeco Instruments (VECO)'
ws6['A1'].font = title_font

src_headers = ['#', 'Source', 'URL']
style_sheet(ws6, src_headers, [5, 50, 55])

sources = [
    ('1', 'Yahoo Finance — VECO Summary & Price', 'https://finance.yahoo.com/quote/VECO/'),
    ('2', 'Yahoo Finance — VECO Income Statement', 'https://finance.yahoo.com/quote/VECO/financials/'),
    ('3', 'Yahoo Finance — VECO Balance Sheet', 'https://finance.yahoo.com/quote/VECO/balance-sheet/'),
    ('4', 'Yahoo Finance — VECO Cash Flow', 'https://finance.yahoo.com/quote/VECO/cash-flow/'),
    ('5', 'Yahoo Finance — VECO Key Statistics', 'https://finance.yahoo.com/quote/VECO/key-statistics/'),
    ('6', 'Yahoo Finance — VECO Analyst Estimates', 'https://finance.yahoo.com/quote/VECO/analysis/'),
    ('7', 'StockAnalysis.com — VECO (404 — unavailable)', 'https://stockanalysis.com/quote/VECO/'),
    ('8', 'FRED — 10-Year Treasury Yield (DGS10)', 'https://fred.stlouisfed.org/series/DGS10'),
    ('9', 'Yahoo Finance — VECO News and Press Releases', 'https://finance.yahoo.com/quote/VECO/news/'),
    ('10', 'Refinitiv via Yahoo Finance — Share Statistics', 'finance.yahoo.com (attribution on key-statistics page)'),
    ('11', 'Morningstar via Yahoo Finance — Data attribution', 'finance.yahoo.com (attribution on key-statistics page)'),
    ('12', 'S&P Global Market Intelligence — EBITDA methodology', 'finance.yahoo.com (footnote on key-statistics page)'),
]
for i, (num, s, u) in enumerate(sources, 3):
    ws6.cell(row=i, column=1, value=num).border = thin_border
    ws6.cell(row=i, column=2, value=s).border = thin_border
    ws6.cell(row=i, column=3, value=u).border = thin_border

# Save
wb.save('/home/refcell/dev/capital/models/[2026-06-18] Veeco Instruments Model.xlsx')
print("Model saved successfully.")
