import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

DARK_BG = "1F2937"; LIGHT_BG = "F3F4F6"; HEADER_BG = "374151"
ACCT_BLUE = "3B82F6"; GREEN = "10B981"; WHITE = "FFFFFF"

def sc(cell, bg=None, fg=WHITE, bold=False, size=11, align="center", fmt=None):
    if bg: cell.fill = PatternFill("solid", bg)
    if fg: cell.font = Font(color=fg, bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fmt: cell.number_format = fmt

# === DATA ===
revenue = [251.01, 242.69, 216.93, 180.84, 142.91, 104.24]
rev_growth = [12.65, 11.87, 19.96, 26.54, 37.10, 65.68]
gross_profit = [200.20, 194.84, 175.64, 146.34, 114.73, 83.46]
gross_margin = [79.76, 80.28, 80.96, 80.92, 80.28, 80.06]
op_income = [20.04, 23.24, 21.29, -3.71, 4.19, -13.45]
op_margin = [7.98, 9.57, 9.82, -2.05, 2.93, -12.90]
net_income = [34.42, 37.51, 26.74, 2.84, 7.02, -13.60]
net_margin = [13.71, 15.46, 12.32, 1.57, 4.91, -13.05]
eps_dil = [1.16, 1.28, 0.86, 0.08, 0.21, -0.41]
sbc = [6.49, 6.00, 7.33, 6.05, 4.71, 2.26]
ocf = [45.03, 28.17, 29.25, 15.96, 2.29, -3.25]
capex_abs = [0.79, 0.98, 1.25, 0.64, 1.08, 1.15]
fcf = [44.24, 27.19, 28.00, 15.33, 1.21, -4.40]
fcf_margin = [17.63, 11.20, 12.91, 8.48, 0.84, -4.22]
ebitda = [24.35, 26.42, 24.68, -0.57, 6.52, -11.43]
ebitda_margin = [9.70, 10.89, 11.38, -0.31, 4.56, -10.96]
shares_dil = [29, 29, 31, 34, 33, 33]
share_count = [25.38, 28.75, 30.26, 30.31, 32.91, 32.86]
cash = [63.19, 74.04, 92.54, 71.95, 216.29, 215.32]
total_debt = [83.80, 2.56, 1.50, 2.11, 3.07, 4.00]
net_cash = [-20.61, 71.48, 91.05, 69.84, 213.23, 211.32]
fcf_per_share = [round(f/x, 2) if x else 0 for f, x in zip(fcf, share_count)]

price = 18.28; mkt_cap = 456.29; ev = 477.16

# SHEET 1: Valuation
ws = wb.active; ws.title = "Valuation"
ws.merge_cells("A1:G1"); ws["A1"].value = "Docebo Inc. (NASDAQ: DCBO) — Valuation"; sc(ws["A1"], bg=DARK_BG, size=14); ws.row_dimensions[1].height = 36

info = [("Date:","2026-05-29"),("Ticker:","NASDAQ: DCBO"),("Close Price:","$18.28"),
        ("Diluted Shares:","25.38M"),("Market Cap:","$456.3M"),("Enterprise Value:","$477.2M"),
        ("Primary Lens:","FCF multiple, owner earnings, reverse DCF"),("Stance:","Own / attractively valued cash-flow compounder")]
for i,(k,v) in enumerate(info,2):
    ws.cell(i,1,k); sc(ws.cell(i,1), bold=True); ws.cell(i,2,v); sc(ws.cell(i,2), align="left")

r=12; ws.merge_cells(f"A{r}:G{r}"); ws[f"A{r}"].value="Key Valuation Metrics"; sc(ws[f"A{r}"], bg=ACCT_BLUE, size=12)
metrics = [("Trailing P/E","13.26x","Reasonable for profitable SaaS"),("Forward P/E","8.89x","Very cheap forward basis"),
    ("P/S","1.82x","Below 2x even at growth rates"),("Forward P/S","1.62x","Attractive for software"),
    ("P/FCF","10.31x","Cheap for high-margin software"),("EV/FCF","10.79x","Still cheap after debt adj"),
    ("EV/Sales","1.90x","Compressed SaaS multiple"),("EV/EBITDA","16.19x","Looks normal but EBITDA is small"),
    ("ROE (TTM)","132.85%","Distorted by negative equity post-buyback"),("ROIC","245.93%","Very high — low capital base")]
for i,(k,v,n) in enumerate(metrics,13):
    ws.cell(i,1,k); sc(ws.cell(i,1), bg=LIGHT_BG); ws.cell(i,2,v); ws.cell(i,3,n); sc(ws.cell(i,3), align="left")

r=25; ws.merge_cells(f"A{r}:G{r}"); ws[f"A{r}"].value="Historical Financial Summary ($M)"; sc(ws[f"A{r}"], bg=GREEN, size=12)
r+=1
for ci,label in enumerate(["","TTM","FY2025","FY2024","FY2023","FY2022","FY2021"],1):
    ws.cell(r,ci,label); sc(ws.cell(r,ci), bg=HEADER_BG)
r+=1
for label,vals in [("Revenue",revenue),("Revenue Growth (%)",rev_growth),("Gross Profit",gross_profit),
    ("Gross Margin (%)",gross_margin),("Op. Income",op_income),("Op. Margin (%)",op_margin),
    ("EBITDA",ebitda),("EBITDA Margin (%)",ebitda_margin),("Net Income",net_income),
    ("Net Margin (%)",net_margin),("EPS Diluted",eps_dil),("Op. Cash Flow",ocf),
    ("Capex",capex_abs),("Free Cash Flow",fcf),("FCF Margin (%)",fcf_margin),
    ("FCF Per Share",fcf_per_share),("Diluted Shares (M)",shares_dil),
    ("Shares Outstanding (M)",share_count),("SBC ($M)",sbc),("Cash ($M)",cash),
    ("Total Debt ($M)",total_debt),("Net Cash ($M)",net_cash)]:
    ws.cell(r,1,label); sc(ws.cell(r,1), bold=True, align="left")
    for j,v in enumerate(vals,2):
        ws.cell(r,j,round(v,2) if isinstance(v,float) else v); sc(ws.cell(r,j))
    r+=1
for col in range(1,8): ws.column_dimensions[get_column_letter(col)].width=16
ws.column_dimensions["A"].width=24

# SHEET 2: WACC
ws2 = wb.create_sheet("WACC")
ws2.merge_cells("A1:C1"); ws2["A1"].value="Docebo (DCBO) — WACC Calculation"; sc(ws2["A1"], bg=DARK_BG, size=14)
wacc_items = [("Component","Value","Rationale"),("Risk-free Rate (10y US)","4.25%","Current 10y note yield"),
    ("Equity Risk Premium","5.00%","Standard ERP"),("Beta (Levered)","0.90","Moderate - SaaS but volatile"),
    ("Cost of Equity (CAPM)","8.75%","=Rfr + Beta x ERP"),("Cost of Debt","5.50%","Est. blended rate"),
    ("Tax Rate","15.00%","GAAP effective rate"),("Market Cap ($M)","456.29","As of 2026-05-29"),
    ("Total Debt ($M)","83.80","Q1 FY2026"),("Equity Weight","84.4%","=E/M"),("Debt Weight","15.6%","=D/M"),
    ("WACC","7.9%","=we x Re + wd x (1-t) x Rd")]
for i,(k,v,n) in enumerate(wacc_items,3):
    ws2.cell(i,1,k); ws2.cell(i,2,v); ws2.cell(i,3,n)
    sc(ws2.cell(i,1), bold=True, align="left"); sc(ws2.cell(i,2)); sc(ws2.cell(i,3), align="left")
ws2.column_dimensions["A"].width=38; ws2.column_dimensions["B"].width=18; ws2.column_dimensions["C"].width=42

# SHEET 3: Scenarios
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells("A1:D1"); ws3["A1"].value="Docebo (DCBO) — Scenario Analysis"; sc(ws3["A1"], bg=DARK_BG, size=14)
for j,h in enumerate(["Metric","Bear","Base","Bull"],1):
    ws3.cell(3,j,h); sc(ws3.cell(3,j), bg=HEADER_BG)
scen = [("Revenue CAGR (5Y)","6%","10%","14%"),("Terminal Revenue ($M)","333","395","480"),
    ("FCF Margin","13%","17%","22%"),("Terminal FCF ($M)","43","67","106"),
    ("Exit Multiple (FCF)","10x","13x","18x"),("Implied EV","$433M","$871M","$1.91B"),
    ("Less: Net Debt","$20.6M","$20.6M","$20.6M"),("Implied Equity Value","$412M","$850M","$1.89B"),
    ("Shares (M)","25.0","25.0","25.0"),("Value/Share","$16.50","$34.00","$75.60"),
    ("Return from $18.28","-10%","+86%","+314%"),("Weight","20%","50%","30%"),
    ("Weighted Value/Share","$3.30","$17.00","$22.68"),("","Total Probability-Weighted FV","","$42.98"),
    ("","Upside from $18.28","","+135%")]
for i,row in enumerate(scen,4):
    for j,v in enumerate(row,1):
        ws3.cell(i,j,v)
        sc(ws3.cell(i,j), bg=LIGHT_BG if i%2==0 else None)
    if "Total" in str(row):
        for j in range(1,5): sc(ws3.cell(i,j), bg=GREEN, bold=True)

for col in range(1,5): ws3.column_dimensions[get_column_letter(col)].width=22
ws3.column_dimensions["A"].width=32

# SHEET 4: Actuals Source Audit
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells("A1:D1"); ws4["A1"].value="Docebo (DCBO) — Data Sources & Audit"; sc(ws4["A1"], bg=DARK_BG, size=14)
for j,h in enumerate(["Data Point","Source","Date","Notes"],1):
    ws4.cell(2,j,h); sc(ws4.cell(2,j), bg=HEADER_BG)
audit = [("Stock Price","StockAnalysis/Yahoo","2026-05-29","Closed $18.28"),
    ("Mkt Cap / EV / Shares","StockAnalysis Stats","2026-05-29","25.38M dil, $456.3M mc, $477.2M EV"),
    ("Income Statement","StockAnalysis","2026-05-30","TTM through Mar 2026"),
    ("Balance Sheet","StockAnalysis","2026-05-30","New debt $83.8M, goodwill $49.4M"),
    ("Cash Flow","StockAnalysis","2026-05-30","TTM FCF $44.2M"),
    ("Analyst Estimates","StockAnalysis (S&P Global)","2026-05-30","12 anlsysts, Buy, avg $28.55"),
    ("Valuation Ratios","StockAnalysis Stats","2026-05-29","P/E 13.26, Fwd 8.89, P/FCF 10.31"),
    ("Last Earnings","StockAnalysis Stats","2026-05-29","May 8, 2026, before open")]
for i,(dp,src,dt,note) in enumerate(audit,3):
    ws4.cell(i,1,dp); ws4.cell(i,2,src); ws4.cell(i,3,dt); ws4.cell(i,4,note)
    sc(ws4.cell(i,1), bold=True, align="left")
    for c in range(2,5): sc(ws4.cell(i,c), align="left")
ws4.column_dimensions["A"].width=28; ws4.column_dimensions["B"].width=32
ws4.column_dimensions["C"].width=14; ws4.column_dimensions["D"].width=56

# SHEET 5: Questions
ws5 = wb.create_sheet("Questions")
ws5.merge_cells("A1:C1"); ws5["A1"].value="Docebo (DCBO) — Open Questions"; sc(ws5["A1"], bg=DARK_BG, size=14)
qs = ["What caused Q1 FY2026 debt jump to $83.8M and goodwill to $49.4M — what acquisition?",
    "Negative equity ($-0.62M TTM) — buybacks + net income drag? Is this sustainable?",
    "TTM tax rate -70.5% — deferred tax releases or credit carryforwards?",
    "Massive buybacks: $101M TTM, $159M FY2023 — value creation or hiding weak growth?",
    "Is deferred revenue growth ($+13.6M TTM) enough to support $340M annualized rev?",
    "SBC only $6.5/2.6% rev — high owner earnings quality?",
    "How much rev is truly recurring vs project-based?",
    "AI learning platform vs Cornerstone/360Learning/Litmos?",
    "Customer concentration risk? Any single customer >5%?",
    "NRR/GRR not disclosed — can we infer from deferred rev growth vs revenue?",
    "When is next earnings report?",
    "Why did stock drop this much vs peers if growth is solid?"]
for i,q in enumerate(qs,3):
    ws5.cell(i,1,i-2); ws5.cell(i,2,q); sc(ws5.cell(i,2), align="left")

# SHEET 6: Sources
ws6 = wb.create_sheet("Sources")
ws6.merge_cells("A1:B1"); ws6["A1"].value="Docebo (DCBO) — Sources"; sc(ws6["A1"], bg=DARK_BG, size=14)
srcs = ["StockAnalysis DCBO Overview", "https://stockanalysis.com/stocks/dcbo/",
    "Financials (Income/Cash Flow/Balance Sheet)", "https://stockanalysis.com/stocks/dcbo/financials/",
    "Statistics", "https://stockanalysis.com/stocks/dcbo/statistics/",
    "Forecast/Analyst Estimates", "https://stockanalysis.com/stocks/dcbo/forecast/",
    "Price History", "https://stockanalysis.com/stocks/dcbo/history/",
    "SEC EDGAR (CIK 0001786988)", "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=0001786988"]
for i,s in enumerate(srcs,3):
    ws6.cell(i,1,i-2); ws6.cell(i,2,s); sc(ws6.cell(i,2), align="left")

ws6.column_dimensions["A"].width=5; ws6.column_dimensions["B"].width=90

path = "/home/refcell/dev/capital/models/[2026-05-30] Docebo Model.xlsx"
wb.save(path)
print(f"Saved: {path}")
