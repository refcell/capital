"""Add a CEG (Constellation Energy) tab to projections/Projections.xlsx.

Clones the last existing tab (QXO) to preserve styling/formulas, then
overwrites the input cells with Constellation Energy numbers.

Constellation is the largest U.S. nuclear operator and (after closing the
Calpine acquisition on Jan 7, 2026) the largest private-sector power producer in
the world.  The company guides and is covered by the Street on an Adjusted
(non-GAAP) Operating Earnings basis, so this tab uses adjusted operating
earnings / EPS (not GAAP, which is distorted by large mark-to-market hedge and
nuclear-decommissioning-trust swings).  All $ figures in millions.

Sources (Q4/FY2025 results Feb 2026; 2026 Business & Earnings Outlook Mar 2026;
Q1 2026 results May 2026; analyst consensus):
  - FY2025 revenue ~$25.5B; GAAP net income $2,319M ($7.40/sh);
    Adjusted Operating Earnings $2,944M ($9.39/sh); diluted shares ~314M
  - FY2024 Adjusted Operating EPS $8.67; GAAP EPS $11.89
  - Calpine acquisition closed Jan 7, 2026 (stock-for-stock, ~$11.9B equity
    issued + ~$12.7B debt assumed); shares rose to ~361M
  - FY2026 guidance: Adjusted Operating EPS $11.00-$12.00 on ~361M avg diluted
    shares; consensus revenue ~$34.5B; Q1 2026 revenue $11.1B, adj. EPS $2.74
  - Growth framework: Base EPS* growth of 20%+ from 2026-2029, long-term rolling
    3-yr 10%+; upside from nuclear PPAs/data-center deals, gas contracts, PTC
    inflation floors, capital allocation could add 30-35% to 2029 earnings.
    Revenue ~flat post-Calpine (consensus 2027 ~$35B), so EPS growth is driven by
    margin expansion (capacity prices, PTC floors, premium contracts) + buybacks.
  - Analyst consensus adj. EPS: 2026 ~$11.71, 2027 ~$13.55
  - Price ~$274-300 (Jun 2026); market cap ~$110B; forward P/E ~24-26x
"""

from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "projections" / "Projections.xlsx"

# Shared inputs
SHARES = 361          # diluted shares (M), post-Calpine; per 2026 guidance basis
REV_2026 = 34500      # FY2026E revenue (~$34.5B consensus, full-year Calpine)
NI_2026 = 4150        # FY2026E adjusted operating earnings (~$11.50/sh x 361)
NI_G_2026 = 0.41      # adj. earnings growth 2025->2026 ($2,944M -> $4,150M, Calpine)

# Analyst adjusted operating earnings estimates 2026-2029
# (Street adj. EPS ~$11.71/$13.55/$15.60/$17.90 x 361 shares; EPS growth is
#  margin-/contract-driven as revenue is ~flat after the Calpine consolidation.)
ANALYST = {"C": 4227, "D": 4892, "E": 5630, "F": 6460}

# Per-case assumptions.  Each template block is offset by 24 rows:
#   Bull base row 4, Base 28, Bear 52.  Within a block, relative rows:
#     +2 revenue, +3 revenue growth, +5 net income, +6 NI growth,
#     +8 analyst est, +10 NI margins, +14 PE low, +15 PE high, H(+0) shares.
# Revenue grows only modestly post-Calpine (organic load growth, electrification,
# data-center demand).  Adjusted net margin expands materially on nuclear PTC
# inflation floors (step-up in 2028 as state ZEC/CMC programs roll off), rising
# PJM capacity prices and premium long-term nuclear/data-center contracts
# (bear: capacity/PTC upside disappoints and premium deals stall).
CASES = {
    "bull": {
        "base": 4,
        # more data-center offtake + gas utilization lift the top line
        "rev_g": {"C": 0.06, "D": 0.06, "E": 0.07, "F": 0.06, "G": 0.05, "H": 0.05},
        "margins": {"D": 0.14, "E": 0.16, "F": 0.175, "G": 0.185, "H": 0.19},
        "pe_low": {c: 26 for c in "CDEFGH"},
        "pe_high": {c: 34 for c in "CDEFGH"},
    },
    "base": {
        "base": 28,
        # steady organic growth; margins expand on PTC floors + capacity prices
        "rev_g": {"C": 0.04, "D": 0.04, "E": 0.05, "F": 0.05, "G": 0.04, "H": 0.04},
        "margins": {"D": 0.136, "E": 0.151, "F": 0.165, "G": 0.168, "H": 0.17},
        "pe_low": {c: 22 for c in "CDEFGH"},
        "pe_high": {c: 28 for c in "CDEFGH"},
    },
    "bear": {
        "base": 52,
        # soft power/capacity prices, premium deals slow, flatter margins
        "rev_g": {"C": 0.02, "D": 0.02, "E": 0.03, "F": 0.03, "G": 0.02, "H": 0.02},
        "margins": {"D": 0.125, "E": 0.13, "F": 0.135, "G": 0.135, "H": 0.135},
        "pe_low": {c: 16 for c in "CDEFGH"},
        "pe_high": {c: 22 for c in "CDEFGH"},
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    src = wb[wb.sheetnames[-1]]          # clone the last tab (QXO)
    ws = wb.copy_worksheet(src)
    ws.title = "CEG"

    # Header
    ws["B2"] = "CEG"
    ws["C2"] = datetime(2026, 6, 21)

    for cfg in CASES.values():
        b = cfg["base"]
        ws[f"H{b}"] = SHARES                      # share count

        # Revenue base (2026E) + growth row
        ws[f"C{b + 2}"] = REV_2026
        for col, g in cfg["rev_g"].items():
            ws[f"{col}{b + 3}"] = g

        # Adjusted operating earnings base (2026E) + base-year growth
        ws[f"C{b + 5}"] = NI_2026
        ws[f"C{b + 6}"] = NI_G_2026

        # Analyst estimates (2026-2029)
        for col, v in ANALYST.items():
            ws[f"{col}{b + 8}"] = v

        # NI margins (D..H; C stays as formula =NI/Rev)
        for col, m in cfg["margins"].items():
            ws[f"{col}{b + 10}"] = m

        # PE low / high
        for col, v in cfg["pe_low"].items():
            ws[f"{col}{b + 14}"] = v
        for col, v in cfg["pe_high"].items():
            ws[f"{col}{b + 15}"] = v

    wb.save(XLSX)
    print(f"Added CEG tab. Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
