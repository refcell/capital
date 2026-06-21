"""Add a ZETA (Zeta Global) tab to projections/Projections.xlsx.

Clones the last existing tab (FISV) to preserve styling/formulas, then
overwrites the input cells with Zeta Global numbers.

Zeta runs GAAP net losses (FY2025 net loss -$31.5M) but is FCF-positive and
guiding to positive GAAP net income in 2026.  Consistent with the house view in
projections/ZETA.md, this tab is built on a NORMALIZED / ADJUSTED net income
basis (adjusted EPS x diluted shares), which is how the Street values ZETA.
All $ figures in millions.

Sources (FY2025 results Feb 24 2026; Q1 2026 results Apr 30 2026 + 10-Q;
projections/ZETA.md):
  - FY2025 revenue $1,304.7M (+30% Y/Y); GAAP net loss -$31.5M
  - FY2026 guide: revenue ~$1,785M (~37% growth), adj EBITDA ~$397M, FCF ~$235M,
    positive GAAP net income
  - Zeta 2028 target: revenue >=$2.3B, adj EBITDA >=$573M (~25% margin)
  - Diluted shares ~249M (Q1'26 10-Q, Apr 24 2026) -> use 250M
  - Analyst adj EPS consensus: FY2026 ~$0.95, FY2027 ~$1.19, FY2028 ~$1.37
  - Price ~$18.5 (Jun 2026); normalized P/E ~19x FY26, ~15x FY27
  - Revenue / normalized-net-income paths mirror projections/ZETA.md bear/base/bull
"""

from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "projections" / "Projections.xlsx"

# Shared inputs
SHARES = 250          # diluted shares (M); ~249M Q1'26, buybacks ~offset SBC
REV_2026 = 1785       # FY2026E revenue (company guide midpoint, ~37% growth)
NI_2026 = 238         # FY2026E normalized net income (= adj EPS ~$0.95 x 250)
NI_G_2026 = 0.30      # normalized NI growth 2025->2026 (~$0.73 -> $0.95 adj EPS)

# Analyst normalized net income estimates 2026-2029 (adj EPS x shares; consensus
# thins past FY2028 so FY2029 is a consensus-trend extrapolation).
ANALYST = {"C": 238, "D": 298, "E": 343, "F": 400}

# Per-case assumptions.  Each template block is offset by 24 rows:
#   Bull base row 4, Base 28, Bear 52.  Within a block, relative rows:
#     +6 revenue, +7 revenue growth, +9 net income, +10 NI growth,
#     +12 analyst est, +14 NI margins, +18 PE low, +19 PE high, H(+0) shares.
# Revenue growth (D..H) and normalized net margins (D..H) are calibrated to hit
# the revenue / normalized-net-income paths in projections/ZETA.md.
CASES = {
    "bull": {
        "base": 4,
        # 2027 2150, 2028 2650, 2029 3350, ->2031 5000
        "rev_g": {"C": 0.37, "D": 0.204, "E": 0.233, "F": 0.264, "G": 0.22, "H": 0.22},
        # normalized net margins -> NI ~387/466/630/.../1065
        "margins": {"D": 0.18, "E": 0.185, "F": 0.19, "G": 0.20, "H": 0.213},
        "pe_low": {c: 18 for c in "CDEFGH"},
        "pe_high": {c: 24 for c in "CDEFGH"},
    },
    "base": {
        "base": 28,
        # 2027 2080, 2028 2410, 2029 2950, ->2031 4050
        "rev_g": {"C": 0.37, "D": 0.165, "E": 0.159, "F": 0.224, "G": 0.17, "H": 0.17},
        # normalized net margins -> NI ~301/370/475/.../753
        "margins": {"D": 0.145, "E": 0.154, "F": 0.161, "G": 0.173, "H": 0.186},
        "pe_low": {c: 15 for c in "CDEFGH"},
        "pe_high": {c: 20 for c in "CDEFGH"},
    },
    "bear": {
        "base": 52,
        # 2027 2040, 2028 2300, 2029 2650, ->2031 3400
        "rev_g": {"C": 0.37, "D": 0.143, "E": 0.127, "F": 0.152, "G": 0.13, "H": 0.13},
        # normalized net margins -> NI ~237/286/331/.../469
        "margins": {"D": 0.116, "E": 0.124, "F": 0.125, "G": 0.131, "H": 0.138},
        "pe_low": {c: 12 for c in "CDEFGH"},
        "pe_high": {c: 16 for c in "CDEFGH"},
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    src = wb[wb.sheetnames[-1]]          # clone the last tab (FISV)
    ws = wb.copy_worksheet(src)
    ws.title = "ZETA"

    # Header
    ws["B2"] = "ZETA"
    ws["C2"] = datetime(2026, 6, 21)

    for cfg in CASES.values():
        b = cfg["base"]
        ws[f"H{b}"] = SHARES                      # share count

        # Revenue base (2026E) + growth row
        ws[f"C{b + 2}"] = REV_2026
        for col, g in cfg["rev_g"].items():
            ws[f"{col}{b + 3}"] = g

        # Normalized net income base (2026E) + base-year growth
        ws[f"C{b + 5}"] = NI_2026
        ws[f"C{b + 6}"] = NI_G_2026

        # Analyst estimates (2026-2029)
        for col, v in ANALYST.items():
            ws[f"{col}{b + 8}"] = v

        # NI margins (D..H; C stays as formula =NI/Rev)
        for col, m in cfg["margins"].items():
            ws[f"{col}{b + 10}"] = m

        # PE low / high
        for col, v in cfg["pe_low"].items():
            ws[f"{col}{b + 14}"] = v
        for col, v in cfg["pe_high"].items():
            ws[f"{col}{b + 15}"] = v

    wb.save(XLSX)
    print(f"Added ZETA tab. Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
