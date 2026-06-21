"""Add a NOW (ServiceNow) tab to projections/Projections.xlsx.

Clones the last existing tab (NFLX) to preserve styling/formulas, then
overwrites the input cells with ServiceNow numbers.

ServiceNow is GAAP-profitable but, like other large-cap enterprise software,
is valued and covered by the Street on a non-GAAP (adjusted) basis, so this tab
uses non-GAAP net income / EPS.  All share/EPS figures are split-adjusted for
the 5-for-1 forward stock split effective December 17, 2025.  All $ in millions.

Sources (Q4/FY2025 results Jan 2026; Q1 2026 results Apr 22 2026; 2025 10-K):
  - FY2025 total revenue $13,278M (+20.9% Y/Y); subscription $12,883M
    GAAP net income $1,748M (diluted EPS $1.67); diluted shares 1,047M
    Non-GAAP net income $3,669M (~27.6% margin); non-GAAP diluted EPS $3.51
    Non-GAAP operating margin 31%
  - FY2024 total revenue $10,984M; non-GAAP NI $2,902M; non-GAAP EPS $13.92
    (pre-split; $2.78 split-adjusted)
  - FY2026 guidance: subscription revenue $15.74B-$15.78B (raised at Q1),
    implying ~total revenue ~$16.2B (+22%); non-GAAP operating margin ~31.5%;
    GAAP diluted weighted-average shares ~1.05B
  - Q1 2026: revenue $3,770M (+22% Y/Y); non-GAAP EPS $0.97
  - Analyst consensus (non-GAAP): FY2026 EPS ~$4.19 (+~19% Y/Y), growing
    high-teens to ~$7 by 2029; BofA models 18%-22% revenue growth through 2028;
    $30B+ subscription revenue target by 2030
  - Price ~$100 (Jun 2026), down ~50% from 52-wk high $211; forward P/E ~24x
    (vs. 10-yr average ~60-70x); 45 analysts, consensus Strong Buy, avg target ~$165
"""

from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "projections" / "Projections.xlsx"

# Shared inputs
SHARES = 1050         # diluted shares (M), split-adjusted; ~flat (buybacks offset dilution)
REV_2026 = 16200      # FY2026E total revenue (~22% growth off $13,278M)
NI_2026 = 4400        # FY2026E non-GAAP net income (~27% margin, EPS ~$4.19)
NI_G_2026 = 0.20      # net income growth 2025->2026 ($3,669M -> $4,400M)

# Analyst non-GAAP net income estimates 2026-2029
# (EPS ~$4.19/$5.05/$6.00/$7.05 x 1050 shares).
ANALYST = {"C": 4400, "D": 5300, "E": 6300, "F": 7400}

# Per-case assumptions.  Each template block is offset by 24 rows:
#   Bull base row 4, Base 28, Bear 52.  Within a block, relative rows:
#     +2 revenue, +3 revenue growth, +5 net income, +6 NI growth,
#     +8 analyst est, +10 NI margins, +14 PE low, +15 PE high, H(+0) shares.
# Revenue decelerates from the low-20s 2026 guide toward mid-teens as the base
# scales ($30B+ subscription target by 2030).  Non-GAAP net margin expands on
# operating leverage / AI efficiencies, partly offset by lower interest income
# and Armis acquisition dilution (bear: AI disruption + integration drag).
CASES = {
    "bull": {
        "base": 4,
        # AI/Now Assist monetization sustains low-20s growth longer
        "rev_g": {"C": 0.22, "D": 0.20, "E": 0.19, "F": 0.18, "G": 0.17, "H": 0.16},
        "margins": {"D": 0.28, "E": 0.30, "F": 0.31, "G": 0.32, "H": 0.33},
        "pe_low": {"C": 30, "D": 29, "E": 28, "F": 27, "G": 26, "H": 25},
        "pe_high": {"C": 37, "D": 36, "E": 35, "F": 33, "G": 32, "H": 30},
    },
    "base": {
        "base": 28,
        # growth decelerates toward mid-teens as revenue base scales
        "rev_g": {"C": 0.22, "D": 0.18, "E": 0.17, "F": 0.15, "G": 0.14, "H": 0.13},
        "margins": {"D": 0.27, "E": 0.28, "F": 0.29, "G": 0.29, "H": 0.30},
        "pe_low": {"C": 24, "D": 23, "E": 22, "F": 21, "G": 20, "H": 19},
        "pe_high": {"C": 30, "D": 29, "E": 28, "F": 26, "G": 25, "H": 24},
    },
    "bear": {
        "base": 52,
        # AI disruption to legacy workflows + competition slow growth
        "rev_g": {"C": 0.22, "D": 0.15, "E": 0.13, "F": 0.11, "G": 0.10, "H": 0.09},
        "margins": {"D": 0.26, "E": 0.26, "F": 0.27, "G": 0.27, "H": 0.27},
        "pe_low": {"C": 18, "D": 17, "E": 16, "F": 15, "G": 14, "H": 13},
        "pe_high": {"C": 24, "D": 23, "E": 22, "F": 20, "G": 19, "H": 18},
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    src = wb[wb.sheetnames[-1]]          # clone the last tab (NFLX)
    ws = wb.copy_worksheet(src)
    ws.title = "NOW"

    # Header
    ws["B2"] = "NOW"
    ws["C2"] = datetime(2026, 6, 21)

    for cfg in CASES.values():
        b = cfg["base"]
        ws[f"H{b}"] = SHARES                      # share count

        # Revenue base (2026E) + growth row
        ws[f"C{b + 2}"] = REV_2026
        for col, g in cfg["rev_g"].items():
            ws[f"{col}{b + 3}"] = g

        # Non-GAAP net income base (2026E) + base-year growth
        ws[f"C{b + 5}"] = NI_2026
        ws[f"C{b + 6}"] = NI_G_2026

        # Analyst estimates (2026-2029)
        for col, v in ANALYST.items():
            ws[f"{col}{b + 8}"] = v

        # NI margins (D..H; C stays as formula =NI/Rev)
        for col, m in cfg["margins"].items():
            ws[f"{col}{b + 10}"] = m

        # PE low / high
        for col, v in cfg["pe_low"].items():
            ws[f"{col}{b + 14}"] = v
        for col, v in cfg["pe_high"].items():
            ws[f"{col}{b + 15}"] = v

    wb.save(XLSX)
    print(f"Added NOW tab. Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
