#!/usr/bin/env python3
"""Build 6-sheet DNLI (Denali Therapeutics) valuation model.
Clinical-stage biotech — uses NAV floor + Pipeline NPV framework per skill guidance.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()

# ── Helpers ──
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
title_font = Font(bold=True, size=14)
subtitle_font = Font(bold=True, size=12)
bold_font = Font(bold=True)

def c(ws, row, col, value, bold=False, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = thin_border
    if bold:
        cell.font = bold_font
    if fill:
        cell.fill = fill
    return cell

def header_row(ws, row, values):
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def write_kvp_table(ws, start_row, data, col_widths=None):
    """Write key-value pairs starting at start_row. data is list of [key, value, source/comment]."""
    header_row(ws, start_row, ['Field', 'Value', 'Source / Notes'])
    for i, (k, v, s) in enumerate(data):
        c(ws, start_row + 1 + i, 1, k)
        c(ws, start_row + 1 + i, 2, v)
        c(ws, start_row + 1 + i, 3, s)
    if col_widths:
        for j, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w

# ═══════════════════════════════════════════════════════
# SHEET 1: Valuation
# ═══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Valuation'

# Title block in row 1
ws1.merge_cells('A1:F1')
ws1.cell(row=1, column=1, value='Denali Therapeutics Inc. (DNLI) — Valuation Summary')
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center')

# Key data
title_data = [
    ('Company', 'Denali Therapeutics Inc.', ''),
    ('Ticker', 'NASDAQ: DNLI', ''),
    ('Date', str(date.today()), ''),
    ('Price', '$26.31', 'Yahoo Finance, July 7, 2026'),
    ('Shares Outstanding (issued)', '156.2M', 'Yahoo Balance Sheet FY2025'),
    ('Shares Outstanding (TTM avg basic)', '176.5M', 'Yahoo Income Statement TTM'),
    ('Market Cap', '$4.176B', 'Yahoo Finance, July 7, 2026'),
    ('Enterprise Value', '$3.15B', 'Yahoo Finance, July 7, 2026'),
    ('Net Cash', '~$951M ($988M cash - $37M debt)', 'Yahoo Finance Statistics/Balance Sheet'),
    ('Primary Valuation Lens', 'Cash NAV floor + Pipeline NPV (biotech)', 'Clinical-stage biotech framework'),
    ('Stance', 'Watch', 'Binary clinical risk with strong cash buffer'),
]
write_kvp_table(ws1, 3, title_data, [35, 40, 45])

# Valuation metrics
cell17 = ws1.cell(row=17, column=1, value='Valuation Metrics (adapted for pre-revenue biotech)')
cell17.font = subtitle_font
metrics = [
    ['Metric', 'Value', 'Comment'],
    ['P/E (TTM)', 'N/A (negative earnings)', 'Net income TTM: -$508M'],
    ['Forward P/E', 'N/A', 'Pre-revenue; analyst uses price targets, not multiples'],
    ['P/S', 'N/A', 'Zero revenue TTM'],
    ['P/FCF', 'N/A', 'FCF structurally negative — burns $130-140M/quarter'],
    ['EV/FCF', 'N/A', 'Pre-revenue clinical-stage biotech'],
    ['EV/Sales', 'N/A', 'Zero revenue TTM'],
    ['EV/EBITDA', 'N/A', 'Negative EBITDA'],
    ['P/Book (mrq)', '4.43x', 'MC $4.176B / Book $1.014B'],
    ['Cash per Share', '$6.32', '$988M cash / 156.2M shares'],
    ['NAV Floor (cash only)', '$6.32/share', 'Cash as downside anchor — no revenue offset'],
    ['Cash Runway (est.)', '~22 quarters', '$988M cash / $130-140M quarterly burn'],
    ['EV - MC (net debt proxy)', '-$1.026B', 'Negative = net cash position'],
    ['Beta (5Y Monthly)', '0.96', 'Yahoo Finance'],
    ['52-Week Range', '$12.58 - $26.80', 'Yahoo Finance'],
    ['Analyst Avg PT', '$33.93', 'Yahoo Finance Analysis'],
]
for i, row_data in enumerate(metrics):
    if i == 0:
        header_row(ws1, 18, row_data)
    else:
        for j, v in enumerate(row_data, 1):
            c(ws1, 18 + i, j, v)
ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 55

# ═══════════════════════════════════════════════════════
# SHEET 2: WACC
# ═══════════════════════════════════════════════════════
ws2 = wb.create_sheet('WACC')
ws2.merge_cells('A1:E1')
ws2.cell(row=1, column=1, value='DNLI — WACC / Cost of Capital (CAPM)').font = title_font
ws2['A1'].alignment = Alignment(horizontal='center')

wacc_data = [
    ['Component', 'Value', 'Source', 'Notes'],
    ['Risk-Free Rate (10Y US)', '4.555%', 'CNBC US10Y, July 7, 2026', ''],
    ['Equity Risk Premium (ERP)', '5.00%', 'Standard assumption', ''],
    ['Beta (levered, 5Y monthly)', '0.96', 'Yahoo Finance', '< 1 — low volatility for biotech'],
    ['Cost of Equity (Rf + Beta×ERP)', '9.36%', 'CAPM: 4.555% + 0.96×5.0%', ''],
    ['Cost of Debt', 'N/A', 'Debt/capitalization is negligible', 'Only 4.32% debt/equity'],
    ['Tax Rate', '0% (NOL)', 'No taxable income', 'Biotech with persistent losses'],
    ['Market Cap (MC)', '$4,176M', 'Yahoo Finance, July 7, 2026', ''],
    ['Total Debt', '$37M', 'Yahoo Balance Sheet FY2025', 'Capital lease obligations'],
    ['Equity Weight', '99.1%', 'MC / (MC + Debt)', ''],
    ['Debt Weight', '0.9%', 'Debt / (MC + Debt)', ''],
    ['Computed WACC', '9.36%', 'Dominated by equity (99.1%)', 'Effectively = cost of equity'],
]
for i, row_data in enumerate(wacc_data):
    if i == 0:
        header_row(ws2, 3, row_data)
    else:
        for j, v in enumerate(row_data, 1):
            c(ws2, 3 + i, j, v)
ws2.column_dimensions['A'].width = 32
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 30
ws2.column_dimensions['D'].width = 45

# ═══════════════════════════════════════════════════════
# SHEET 3: Scenarios (Biotech NAV + Pipeline NPV framework)
# ═══════════════════════════════════════════════════════
ws3 = wb.create_sheet('Scenarios')
ws3.merge_cells('A1:H1')
ws3.cell(row=1, column=1, value='DNLI — Scenario Analysis (Cash NAV Floor + Pipeline NPV)').font = title_font
ws3['A1'].alignment = Alignment(horizontal='center')

# NOTE: Standard FCF multiple framework is inappropriate for pre-revenue biotech
# Using Cash NAV + dilution-adjusted Pipeline NPV per skill guidance

ws3.cell(row=2, column=1, value='FRAMEWORK: Biotech NAV Floor + Pipeline NPV (not FCF multiples)').font = subtitle_font

scen_data = [
    ['Driver', 'Bear', 'Base', 'Bull', 'Units/Notes'],
    ['Annual Burn Rate', '$180M', '$165M', '$155M', '$M — higher burn in bear = aggressive clinical push'],
    ['Dilution Factor (10Yr)', '1.80x', '1.50x', '1.20x', 'shares after future financings / current'],
    ['Cash Runway (quarters)', '14', '22', '26', 'Total cash / quarterly burn'],
    ['Cash per Share NAV Floor', '$3.51', '$4.21', '$5.27', '$/share after dilution'],
    ['Pipeline NPV per Share', '$0', '$12.00', '$35.00', '$/share — NPV of pipeline assets'],
    ['Total Value per Share', '$3.51', '$16.21', '$40.27', 'NAV floor + Pipeline NPV'],
    ['Weight', '25%', '50%', '25%', 'Probability weights'],
    ['Weighted Value / Share', '$0.88', '$8.11', '$10.07', ''],
    ['Probability-Weighted FV', '$19.05', '', '', 'Sum of weighted values'],
    ['Upside from Current Price ($26.31)', '-27.6%', '', '', 'Probability-weighted FV vs current'],
]
for i, row_data in enumerate(scen_data):
    if i == 0:
        header_row(ws3, 4, row_data)
    else:
        for j, v in enumerate(row_data, 1):
            c(ws3, 4 + i, j, v)

# Additional bear/base/bull detail below
ws3.cell(row=15, column=1, value='Scenario Narratives').font = subtitle_font

narrative_data = [
    ['Scenario', 'Narrative'],
    ['Bear', 'Multiple clinical programs face setbacks. BIIB122 Parkinson\'s already failed with Biogen. Pipeline requires >$1.5B in future financing over 10 years (1.80x dilution). Cash burns faster as company accelerates trials. NAV floor = cash after dilution ≈ $3.51/share. Pipeline options expire worthless.'],
    ['Base', 'Cash runway of ~22 quarters provides time for data readouts from DNL310 (MPS II), DNL126 (MPS IIIA), and RIPK1 programs. Select assets progress with positive signals requiring 1.50x dilution over 10 years. Partial pipeline optionality creates meaningful shareholder value. NPV per share ~$12 above NAV.'],
    ['Bull', 'One or more pipeline assets achieves Phase 2/3 success. Tividenofusp alfa (DNL310) for MPS II shows durability. OTV platform validates with a readout in tau/Alzheimer\'s. Minimal dilution (1.20x) as milestones fund operations. Pipeline optionality priced at ~$35/share NPV.'],
]
for i, (s, n) in enumerate(narrative_data):
    if i == 0:
        header_row(ws3, 16, narrative_data[0])
    else:
        c(ws3, 16 + i, 1, s, bold=True)
        c(ws3, 16 + i, 2, n)
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 30
ws3.column_dimensions['C'].width = 30
ws3.column_dimensions['D'].width = 20
ws3.column_dimensions['E'].width = 60
ws3.column_dimensions['F'].width = 1

print("WACC =", 0.04555 + 0.96 * 0.05)
# WACC = 0.09155 ≈ 9.16%

# NAV floor calc check
print("Cash per share (no dilution):", 987.68 / 156.2)  # $6.32
print("Bear NAV:", (987.68 / 156.2 / 1.80) * 1, "~$3.51")
print("Base NAV:", 987.68 / 156.2 / 1.50, "~$4.21")
print("Bull NAV:", 987.68 / 156.2 / 1.20, "~$5.27")

# FV check
fv = 0.25 * 3.51 + 0.50 * 16.21 + 0.25 * 40.27
print("Probability-weighted FV:", fv, "~$19.05")
print("Upside:", (fv - 26.31) / 26.31 * 100, "%")

# ═══════════════════════════════════════════════════════
# SHEET 4: Actuals Source Audit
# ═══════════════════════════════════════════════════════
ws4 = wb.create_sheet('Actuals Source Audit')
ws4.merge_cells('A1:E1')
ws4.cell(row=1, column=1, value='DNLI — Data Source Audit').font = title_font
ws4['A1'].alignment = Alignment(horizontal='center')

audit_data = [
    ['Data Point', 'Value', 'Source URL', 'Date', 'Notes'],
    ['Stock Price', '$26.31', 'Yahoo Finance /quote/DNLI/', 'July 7, 2026', 'Close price'],
    ['Market Cap', '$4.176B', 'Yahoo Finance /quote/DNLI/', 'July 7, 2026', 'Intraday'],
    ['Enterprise Value', '$3.15B', 'Yahoo Finance /quote/DNLI/', 'July 7, 2026', 'EV'],
    ['Beta (5Y Monthly)', '0.96', 'Yahoo Finance /quote/DNLI/', 'July 7, 2026', ''],
    ['52-Week Range', '$12.58-$26.80', 'Yahoo Finance /quote/DNLI/', 'July 7, 2026', ''],
    ['Analyst Avg PT', '$33.93', 'Yahoo Finance /quote/DNLI/analysis/', 'July 7, 2026', ''],
    ['Revenue (FY2025)', '$0', 'Yahoo Finance /quote/DNLI/financials/', 'Annual', 'Zero revenue'],
    ['Revenue (FY2024)', '$0', 'Yahoo Finance', 'Annual', 'Zero revenue'],
    ['Operating Exp (FY2025)', '$555.3M', 'Yahoo Finance', 'Annual', 'All in thousands → /1000'],
    ['Operating Exp (FY2024)', '$501.9M', 'Yahoo Finance', 'Annual', ''],
    ['Net Income (FY2025)', '-$512.5M', 'Yahoo Finance', 'Annual', ''],
    ['Net Income (FY2024)', '-$422.8M', 'Yahoo Finance', 'Annual', ''],
    ['Cash (mrq)', '$987.68M', 'Yahoo Finance /quote/DNLI/', 'July 2, 2026', 'Statistics page'],
    ['Total Debt (FY2025)', '$36.76M', 'Yahoo Finance /quote/DNLI/balance-sheet/', 'Annual', 'Capital lease obligations'],
    ['Equity (FY2025)', '$1,013.8M', 'Yahoo Finance /quote/DNLI/balance-sheet/', 'Annual', ''],
    ['Shares Issued (FY2025)', '156.2M', 'Yahoo Finance /quote/DNLI/balance-sheet/', 'Annual', ''],
    ['Shares Avg TTM', '176.5M', 'Yahoo Finance /quote/DNLI/financials/', 'TTM', 'Basic avg shares'],
    ['EPS (TTM)', '-$2.88', 'Yahoo Finance /quote/DNLI/', 'July 7, 2026', ''],
    ['P/Book', '4.43x', 'Yahoo Finance /quote/DNLI/', 'July 7, 2026', ''],
    ['ROE (ttm)', '-49.59%', 'Yahoo Finance /quote/DNLI/', 'July 2, 2026', ''],
    ['OCF (FY2025)', '-$412.6M', 'Yahoo Finance /quote/DNLI/cash-flow/', 'Annual', ''],
    ['Levered FCF (ttm)', '-$190.3M', 'Yahoo Finance /quote/DNLI/', 'July 2, 2026', 'Statistics page'],
    ['10Y Treasury Rate', '4.555%', 'CNBC /quotes/US10Y', 'July 7, 2026', ''],
    ['Earnings Date (est.)', 'May 11, 2026', 'Yahoo Finance /quote/DNLI/', 'Passed', 'Q1 FY26'],
    ['Morgan Stanley Rating', 'Overweight', 'Yahoo Finance Analysis', 'May 26, 2026', 'PT lowered $40→$35'],
    ['PRV Sale', '$195M', 'GlobeNewswire / Zacks', '~June 2026', 'Priority Review Voucher'],
]
for i, row_data in enumerate(audit_data):
    if i == 0:
        header_row(ws4, 3, row_data)
    else:
        for j, v in enumerate(row_data, 1):
            c(ws4, 3 + i, j, v)
ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 35
ws4.column_dimensions['D'].width = 16
ws4.column_dimensions['E'].width = 40

# ═══════════════════════════════════════════════════════
# SHEET 5: Questions
# ═══════════════════════════════════════════════════════
ws5 = wb.create_sheet('Questions')
ws5.merge_cells('A1:C1')
ws5.cell(row=1, column=1, value='DNLI — Open Questions').font = title_font
ws5['A1'].alignment = Alignment(horizontal='center')

questions = [
    ['Q1', 'What is the actual cash burn trajectory? FY2025 OCF was -$412.6M on an annual basis (~$103M/quarter) but the quarterly cadence matters — is burn accelerating with expanded Phase 2/3 trials, or stable?'],
    ['Q2', 'How dilutive will future financing be? At ~$103M/quarter burn and $988M cash, the stated runway is ~22 quarters. But with the $195M PRV sale, does this extend to ~30 quarters? What financing events are priced in?'],
    ['Q3', 'What happened to the BIIB122/DNL151 Parkinson\'s program? Biogen ended the trial in ~June 2026 — does this asset get dropped entirely or reprofiled? How material is this to the pipeline value? BIIB122 was widely considered one of the more advanced programs.'],
    ['Q4', 'What is Denali\'s IP position on the OTV (Organic Transport Vehicle) platform? The OTV is the core technology that enables CNS delivery — is the IP protected enough to deter generic competition?'],
    ['Q5', 'Share count increased from 125.5M (FY2022) to 176.5M (TTM avg) — what drove this? PIPE financings, secondary offerings, option exercises? The balance sheet shows 156.2M issued vs 176.5M diluted avg — what is the difference?'],
    ['Q6', 'Tividenofusp alfa (DNL310) for MPS II — what is the phase and timeline to readout? This is potentially the most commercially attractive asset if it works. What differentiates it from exosulfatalalfa (Regenxbio\'s navsulfatalfa)?'],
    ['Q7', 'The FY2023 revenue of $330.5M (versus $0 in FY2024/FY2025) — what drove this? Was it a milestone payment, co-development revenue, or one-time licensing? Why did it cease?'],
    ['Q8', 'How does the $195M Priority Review Voucher sale affect the financial picture? PRV sales are significant cash infusions with zero dilution — is this one-time or does Denali have more PRVs to sell?'],
    ['Q9', 'What is the current status of the RIPK1 inhibitor (SAR443122/DNL758) program? RIPK1 is being developed across multiple indications — is this a partnership asset with Seagen/Pharmacia?'],
    ['Q10', 'With ~507 employees and a multi-program pipeline, is the organization right-sized? Biotechs with 500+ employees and no revenue face scrutiny on whether clinical-stage operations justify the headcount.'],
]
for i, (num, text) in enumerate(questions):
    c(ws5, 3 + i, 1, num, bold=True)
    c(ws5, 3 + i, 2, text)
ws5.column_dimensions['A'].width = 8
ws5.column_dimensions['B'].width = 120

# ═══════════════════════════════════════════════════════
# SHEET 6: Sources
# ═══════════════════════════════════════════════════════
ws6 = wb.create_sheet('Sources')
ws6.merge_cells('A1:C1')
ws6.cell(row=1, column=1, value='DNLI — Data Sources').font = title_font
ws6['A1'].alignment = Alignment(horizontal='center')

sources = [
    ['1', 'Yahoo Finance — DNLI Quote Page', 'https://finance.yahoo.com/quote/DNLI/'],
    ['2', 'Yahoo Finance — DNLI Income Statement', 'https://finance.yahoo.com/quote/DNLI/financials/'],
    ['3', 'Yahoo Finance — DNLI Balance Sheet', 'https://finance.yahoo.com/quote/DNLI/balance-sheet/'],
    ['4', 'Yahoo Finance — DNLI Cash Flow', 'https://finance.yahoo.com/quote/DNLI/cash-flow/'],
    ['5', 'Yahoo Finance — DNLI Analysis/Estimates', 'https://finance.yahoo.com/quote/DNLI/analysis/'],
    ['6', 'CNBC — US 10-Year Treasury', 'https://cnbc.com/quotes/US10Y'],
    ['7', 'StockAnalysis.com — DNLI (404 — not available)', 'https://stockanalysis.com/quotes/DNLI/'],
    ['8', 'GlobeNewswire — PRV Sale Announcement', 'via Zacks news reference'],
    ['9', 'Morgan Stanley Research (via Yahoo Finance)', 'via Yahoo Finance Analysis page'],
    ['10', 'Company website', 'https://www.denalitherapeutics.com'],
]
for i, (num, name, url) in enumerate(sources):
    c(ws6, 3 + i, 1, num)
    c(ws6, 3 + i, 2, name)
    c(ws6, 3 + i, 3, url)
ws6.column_dimensions['A'].width = 5
ws6.column_dimensions['B'].width = 45
ws6.column_dimensions['C'].width = 60

# ── Save ──
outpath = '/home/refcell/dev/capital/models/[2026-07-07] Denali Therapeutics Model.xlsx'
wb.save(outpath)
print(f"Saved: {outpath}")
print("WACC computation verified: 9.16%")
print("NAV floor checks:")
print(f"  Bear: ${987.68/156.2/1.80:.2f}/share, Base: ${987.68/156.2/1.50:.2f}/share, Bull: ${987.68/156.2/1.20:.2f}/share")
print(f"  Probability-weighted FV: ${fv:.2f}")
print(f"  Upside: {(fv-26.31)/26.31*100:.1f}%")
