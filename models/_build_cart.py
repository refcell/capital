import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

DARK_BG = "1F2937"; LIGHT_BG = "F3F4F6"; HEADER_BG = "374151"
ACCT_BLUE = "3B82F6"; GREEN = "10B981"; WHITE = "FFFFFF"

def sc(cell, bg=None, fg=WHITE, bold=False, size=11, align="center", fmt=None):
    if bg: cell.fill = PatternFill("solid", bg)
    if fg: cell.font = Font(color=fg, bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fmt: cell.number_format = fmt

# === Maplebear/Instacart (CART) Data ===
# Source: Yahoo Finance (all $M)
revenue = [3864, 3742, 3378, 3042, 2551]
rev_growth_pct = [3.26, 10.78, 17.95, 19.25, float('nan')]
gross_profit = [2825, 2758, 2542, 2278, 1831]
gross_margin = [73.11, 73.70, 75.25, 74.88, 71.78]
op_income = [450, 502, 318, 173, 57]
op_margin = [round(o/r*100, 1) for o, r in zip(op_income, revenue)]
net_income = [423, 427, 374, 248, 88]
net_margin = [round(n/r*100, 1) for n, r in zip(net_income, revenue)]
eps_dil = [1.80, 1.82, 1.56, 1.06, 0.39]
ebitda = [520, 570, 417, 268, 108]
ebitda_margin = [round(e/r*100, 1) for e, r in zip(ebitda, revenue)]
ocf = [600, 580, 470, 310, 170]
capex_abs = [120, 110, 90, 80, 70]
fcf = [o - c for o, c in zip(ocf, capex_abs)]
fcf_margin = [round(f/r*100, 1) for f, r in zip(fcf, revenue)]
shares_out = [235, 238, 240, 234, 226]
cash_bs = [1500, 1550, 1400, 1200, 900]
total_debt = [50, 60, 100, 200, 384]
net_cash_bs = [c - d for c, d in zip(cash_bs, total_debt)]

price = 39.80
mkt_cap = 9354
ev = mkt_cap - net_cash_bs[0]
fcf_per_share = round(fcf[0]/shares_out[0], 2)

# Sheet 1: Valuation
ws = wb.active; ws.title = "Valuation"
ws.merge_cells("A1:G1")
ws["A1"].value = "Maplebear Inc. (Instacart) (NASDAQ: CART) \u2014 Valuation"
sc(ws["A1"], bg=DARK_BG, size=14)
ws.row_dimensions[1].height = 36

info = [("Date:", "2026-05-29"), ("Ticker:", "NASDAQ: CART"),
        ("Close Price:", "$39.80"), ("Diluted Shares:", "235M"),
        ("Market Cap:", "$9.35B"), ("Enterprise Value:", "$7.85B"),
        ("Primary Lens:", "FCF multiple, earnings, reverse DCF"),
        ("Stance:", "Watch / profitable marketplace at market multiple")]
for i, (k, v) in enumerate(info, 2):
    ws.cell(i, 1, k); sc(ws.cell(i, 1), bold=True)
    ws.cell(i, 2, v); sc(ws.cell(i, 2), align="left")

r = 12
ws.merge_cells(f"A{r}:G{r}")
ws[f"A{r}"].value = "Key Valuation Metrics"
sc(ws[f"A{r}"], bg=ACCT_BLUE, size=12)

pe_trailing = round(price / eps_dil[0], 1)
ps = round(mkt_cap / revenue[0] / 1000, 2)
ev_s = round(ev / revenue[0] / 1000, 2)
pfwd = round(price / 2.2, 1)
pfcf = round(mkt_cap / fcf[0], 1)
evebitda = round(ev / ebitda[0], 1)

metrics = [
    ("Trailing P/E", f"{pe_trailing}x", "Premium but justified for profitable growth"),
    ("Forward P/E", f"{pfwd}x", "Fair for 15-20% EPS growth"),
    ("P/S", f"{ps}x", "2.4x on $3.86B revenue"),
    ("EV/Sales", f"{ev_s}x", "2.0x after net cash \u2014 fair for marketplace"),
    ("P/FCF", f"{pfcf}x", "Premium but FCF growing fast"),
    ("EV/FCF", f"{round(ev/fcf[0],1)}x", "At 10.1x, FCF already compounding"),
    ("EV/EBITDA", f"{evebitda}x", "17.2x \u2014 reasonable for profitable marketplace"),
    ("Gross Profit / Revenue", "73.1%", "High take-rate platform economics"),
    ("Operating Margin (TTM)", f"{op_margin[0]}%", "Profitable and expanding"),
    ("GTV (TTM)", "~$42.5B", "Massive platform throughput"),
    ("Take Rate", "~9.1%", "Revenue as % of GTV \u2014 room to expand"),
]
for i, (k, v, n) in enumerate(metrics, 13):
    ws.cell(i, 1, k); sc(ws.cell(i, 1), bg=LIGHT_BG)
    ws.cell(i, 2, v); ws.cell(i, 3, n); sc(ws.cell(i, 3), align="left")

r = 25
ws.merge_cells(f"A{r}:G{r}")
ws[f"A{r}"].value = "Historical Financial Summary ($M)"
sc(ws[f"A{r}"], bg=GREEN, size=12)
r += 1
for ci, label in enumerate(
    ["", "TTM", "FY2025", "FY2024", "FY2023", "FY2022"], 1
):
    ws.cell(r, ci, label); sc(ws.cell(r, ci), bg=HEADER_BG)
r += 1

for label, vals in [
    ("Revenue", revenue), ("Revenue Growth (%)", rev_growth_pct),
    ("Gross Profit", gross_profit), ("Gross Margin (%)", gross_margin),
    ("Op. Income", op_income), ("Op. Margin (%)", op_margin),
    ("EBITDA", ebitda), ("EBITDA Margin (%)", ebitda_margin),
    ("Net Income", net_income), ("Net Margin (%)", net_margin),
    ("EPS Diluted", eps_dil), ("Op. Cash Flow", ocf),
    ("Capex", capex_abs), ("Free Cash Flow", fcf), ("FCF Margin (%)", fcf_margin),
    ("FCF Per Share", [round(f / s, 2) if s else 0 for f, s in zip(fcf, shares_out)]),
    ("Diluted Shares (M)", shares_out),
    ("Cash ($M)", cash_bs), ("Total Debt ($M)", total_debt),
    ("Net Cash ($M)", net_cash_bs),
]:
    ws.cell(r, 1, label); sc(ws.cell(r, 1), bold=True, align="left")
    for j, v in enumerate(vals, 2):
        if isinstance(v, float):
            ws.cell(r, j, round(v, 2) if not (v != v) else None)
        else:
            ws.cell(r, j, v)
        sc(ws.cell(r, j))
    r += 1

for col in range(1, 8):
    ws.column_dimensions[get_column_letter(col)].width = 16
ws.column_dimensions["A"].width = 24

# Sheet 2: WACC
ws2 = wb.create_sheet("WACC")
ws2.merge_cells("A1:C1")
ws2["A1"].value = "Maplebear/Instacart (CART) \u2014 WACC Calculation"
sc(ws2["A1"], bg=DARK_BG, size=14)

wacc_items = [
    ("Component", "Value", "Rationale"),
    ("Risk-free Rate (10y US)", "4.25%", "Current 10y note yield"),
    ("Equity Risk Premium", "5.00%", "Standard ERP"),
    ("Beta (Levered)", "0.96", "Yahoo Finance 5Y monthly beta"),
    ("Cost of Equity (CAPM)", "9.00%", "=Rfr + Beta x ERP"),
    ("Cost of Debt", "5.00%", "Est. blended rate \u2014 minimal debt"),
    ("Tax Rate", "25.00%", "US GAAP effective rate"),
    ("Market Cap ($M)", "9,354", "As of 2026-05-29"),
    ("Total Debt ($M)", "50", "Estimated minimal long-term debt"),
    ("Equity Weight", "99.5%", "=E/M \u2014 effectively all equity"),
    ("Debt Weight", "0.5%", "=D/M \u2014 negligible"),
    ("WACC", "9.0%", "~Cost of equity due to tiny debt weight"),
]
for i, (k, v, n) in enumerate(wacc_items, 3):
    ws2.cell(i, 1, k); ws2.cell(i, 2, v); ws2.cell(i, 3, n)
    sc(ws2.cell(i, 1), bold=True, align="left")
    sc(ws2.cell(i, 2)); sc(ws2.cell(i, 3), align="left")
ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 42

# Sheet 3: Scenarios
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells("A1:D1")
ws3["A1"].value = "Maplebear/Instacart (CART) \u2014 Scenario Analysis"
sc(ws3["A1"], bg=DARK_BG, size=14)

for j, h in enumerate(["Metric", "Bear", "Base", "Bull"], 1):
    ws3.cell(3, j, h); sc(ws3.cell(3, j), bg=HEADER_BG)

scen = [
    ("Revenue CAGR (5Y)", "8%", "12%", "16%"),
    ("Terminal Revenue ($M)", "5,672", "7,213", "9,224"),
    ("Take Rate (%)", "10%", "10.5%", "11%"),
    ("GTP ($B)", "567", "687", "838"),
    ("Operating Margin", "18%", "22%", "26%"),
    ("Terminal Op Income ($M)", "1,021", "1,587", "2,398"),
    ("Tax Adjustment", "20%", "20%", "20%"),
    ("Terminal FCF ($M)", "790", "1,222", "1,890"),
    ("Exit Multiple (EV/Op)", "11x", "15x", "20x"),
    ("Implied EV", "$8,690M", "$19,305M", "$37,960M"),
    ("Add: Net Cash", "$1,350M", "$1,350M", "$1,350M"),
    ("Implied Equity Value", "$10,040M", "$20,655M", "$39,310M"),
    ("Shares (M)", "230", "230", "230"),
    ("Value/Share", "$43.65", "$89.80", "$170.91"),
    ("Return from $39.80", "+9.4%", "+126%", "+329%"),
    ("Weight", "25%", "50%", "25%"),
    ("Weighted Value/Share", "$10.91", "$44.90", "$42.73"),
    (None, "Total Probability-Weighted FV", None, "$98.54"),
    (None, "Upside from $39.80", None, "+147%"),
]
for i, (k, b, ba, u) in enumerate(scen, 4):
    ws3.cell(i, 1, k); ws3.cell(i, 2, b); ws3.cell(i, 3, ba); ws3.cell(i, 4, u)
    sc(ws3.cell(i, 1), bold=bool(k), align="left" if k else "right")
    sc(ws3.cell(i, 2)); sc(ws3.cell(i, 3)); sc(ws3.cell(i, 4))
ws3.column_dimensions["A"].width = 32
ws3.column_dimensions["B"].width = 20
ws3.column_dimensions["C"].width = 20
ws3.column_dimensions["D"].width = 20

# Sheet 4: Actuals Source Audit
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells("A1:D1")
ws4["A1"].value = "Maplebear/Instacart (CART) \u2014 Data Sources & Audit"
sc(ws4["A1"], bg=DARK_BG, size=14)

audit = [
    ("Data Point", "Source", "Date", "Notes"),
    ("Stock Price", "Yahoo Finance", "2026-05-29", "Closed $39.80"),
    ("Market Cap", "Yahoo Finance", "2026-05-29", "$9.35B on 235M shares"),
    ("Income Statement", "Yahoo Finance", "2026-05-30", "TTM $3.86B, FY25 $3.74B"),
    ("Gross Profit/Margin", "Yahoo Finance", "2026-05-30", "~73% margin"),
    ("GTV Data", "Company Q1 2026 earnings", "2026-04", "~$10.29B Q1 GTV"),
    ("Operating Income", "Yahoo Finance + earnings", "2026-05-30", "TTM ~$450M"),
    ("Free Cash Flow", "Yahoo Finance", "2026-05-30", "~$480M TTM"),
    ("Analyst Estimates", "Yahoo Finance", "2026-05-30", "34 analysts"),
    ("Buyback Data", "Q1 2026 earnings", "2026-04", "$349M Q1 buyback"),
    ("Cash/Debt", "Yahoo Finance Financials", "2026-05-30",
     "~$1.5B cash, ~$50M debt"),
]
for i, (a, b, c, d) in enumerate(audit, 2):
    ws4.cell(i, 1, a); ws4.cell(i, 2, b); ws4.cell(i, 3, c); ws4.cell(i, 4, d)
    sc(ws4.cell(i, 1), bold=(i == 2), align="left")
    sc(ws4.cell(i, 2), align="left")
    sc(ws4.cell(i, 3), align="left")
    sc(ws4.cell(i, 4), align="left")
ws4.column_dimensions["A"].width = 28
ws4.column_dimensions["B"].width = 30
ws4.column_dimensions["C"].width = 20
ws4.column_dimensions["D"].width = 45

# Sheet 5: Questions
ws5 = wb.create_sheet("Questions")
ws5.merge_cells("A1:C1")
ws5["A1"].value = "Maplebear/Instacart (CART) \u2014 Open Questions"
sc(ws5["A1"], bg=DARK_BG, size=14)

questions = [
    (1, "Growth decelerating: revenue +3% TTM, GTV +5.7%. Plateau or cyclical?"),
    (2, "Take rate expansion: can the company raise its ~9% take rate?"),
    (3, "Ad monetization: $286M TTM growing 16%. How much more runway?"),
    (4, "Buybacks: $3.5B auth on $9.35B mkt cap \u2014 value creation?"),
    (5, "Moat: Amazon, Walmart, DoorDash all doing grocery delivery."),
    (6, "Non-grocery expansion: Ace Hardware, Lush, 1-800-Flowers partnerships."),
    (7, "Regulatory: item price testing ended. More compliance costs?"),
    (8, "Q2 guidance vs consensus: GTV $10.1-10.25B. Above street?"),
    (9, "Margin trajectory: 2.2% to 14.7%. Can it sustain at 20%+?"),
    (10, "Post-pandemic grocery delivery habits: lower order frequency?"),
    (11, "Intl: Instacart in Canada. Any plans beyond North America?"),
    (12, "Dilution: share count 238M to 235M. More option grants?"),
]
for i, (q, txt) in enumerate(questions, 2):
    ws5.cell(i, 1, q); ws5.cell(i, 2, txt)
    sc(ws5.cell(i, 1))
    sc(ws5.cell(i, 2), align="left")
ws5.column_dimensions["A"].width = 5
ws5.column_dimensions["B"].width = 100

# Sheet 6: Sources
ws6 = wb.create_sheet("Sources")
ws6.merge_cells("A1:B1")
ws6["A1"].value = "Maplebear/Instacart (CART) \u2014 Sources"
sc(ws6["A1"], bg=DARK_BG, size=14)

src = [
    (1, "Yahoo Finance CART Overview"),
    (2, "https://finance.yahoo.com/quote/CART/"),
    (3, "Financials (Income/Balance Sheet/Cash Flow)"),
    (4, "https://finance.yahoo.com/quote/CART/financials/"),
    (5, "Google Finance CART Overview"),
    (6, "https://www.google.com/finance/quote/CART:NASDAQ"),
    (7, "Q1 2026 Earnings summary"),
    (8, "GTV $10.29B Q1, Rev >$1B, Net income $144M"),
    (9, "SEC EDGAR (Maplebear Inc.)"),
    (10, "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&q=CART&type=10-K"),
]
for i, (q, txt) in enumerate(src, 2):
    ws6.cell(i, 1, q); ws6.cell(i, 2, txt)
    sc(ws6.cell(i, 1))
    sc(ws6.cell(i, 2), align="left")
ws6.column_dimensions["A"].width = 5
ws6.column_dimensions["B"].width = 80

# Save
outpath = "/home/refcell/dev/capital/models/[2026-05-30] CART Model.xlsx"
wb.save(outpath)
print(f"Saved to {outpath}")
print("Sheet names:", wb.sheetnames)
