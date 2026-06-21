"""Add a UBER (Uber Technologies) tab to projections/Projections.xlsx.

Clones the last existing tab (CEG) to preserve styling/formulas, then
overwrites the input cells with Uber numbers.

Uber's GAAP net income is heavily distorted by large one-time tax valuation
release benefits ($6.4B in 2024, $5.0B in 2025) and by mark-to-market
revaluations of its equity investments (Aurora, Didi, Grab, etc.).  Uber guides
on Non-GAAP EPS / Adjusted EBITDA, and the Street's forward multiples are quoted
on Non-GAAP earnings, so this tab uses Non-GAAP Net Income / Non-GAAP EPS.  All
$ figures in millions.

Sources (Q4/FY2025 results Feb 2026; Q1 2026 guidance; analyst consensus):
  - FY2025 revenue $52,017M (+18% Y/Y); Gross Bookings $193,454M (+19%)
    GAAP net income $10,053M / GAAP diluted EPS $4.73 (incl. $5.0B tax-release)
    Non-GAAP Net Income $5,237M; Non-GAAP EPS $2.45 (+35% Y/Y)
    Adjusted EBITDA $8,730M (+35%); diluted shares ~2,120M
  - FY2024 revenue $43,978M; Non-GAAP NI $3,970M; Non-GAAP EPS $1.82
  - Q1 2026 guidance: Gross Bookings $52.0B-$53.5B (+17%-21% cc);
    Non-GAAP EPS $0.65-$0.72 (+37% YoY midpoint); adj. EBITDA $2.37B-$2.47B
  - Analyst consensus (Non-GAAP basis): 2026 EPS ~$3.30, 2027 ~$4.30,
    2028 ~$5.40, 2029 ~$6.60; revenue ~$58-60B in 2026 rising to ~$83B by 2029
  - Price ~$71-73 (Jun 2026); market cap ~$146B; forward P/E ~21x;
    59 analysts, consensus Strong Buy, avg 12-mo PT ~$106
  - Optionality (not modeled): autonomous-vehicle scaling, high-margin
    advertising, continued buybacks
"""

from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "projections" / "Projections.xlsx"

# Shared inputs
SHARES = 2080         # diluted shares (M); declining modestly on buybacks
REV_2026 = 59500      # FY2026E revenue (~14% growth off $52,017M)
NI_2026 = 6900        # FY2026E Non-GAAP net income (~11.6% margin, EPS ~$3.32)
NI_G_2026 = 0.32      # Non-GAAP NI growth 2025->2026 ($5,237M -> $6,900M)

# Analyst Non-GAAP net income estimates 2026-2029
# (Street Non-GAAP EPS ~$3.32/$4.28/$5.38/$6.59 x ~2,080 shares.)
ANALYST = {"C": 6900, "D": 8900, "E": 11200, "F": 13700}

# Per-case assumptions.  Each template block is offset by 24 rows:
#   Bull base row 4, Base 28, Bear 52.  Within a block, relative rows:
#     +2 revenue, +3 revenue growth, +5 net income, +6 NI growth,
#     +8 analyst est, +10 NI margins, +14 PE low, +15 PE high, H(+0) shares.
# Revenue (slower than Gross Bookings as take rate matures) decelerates from the
# mid-teens.  Non-GAAP net margin expands on operating leverage, scaling
# high-margin advertising and mobility/delivery cost discipline (bear: AV/robotaxi
# competition, pricing pressure and slower margin expansion).
CASES = {
    "bull": {
        "base": 4,
        # AV partnerships + advertising + membership sustain mid-teens growth
        "rev_g": {"C": 0.15, "D": 0.14, "E": 0.13, "F": 0.12, "G": 0.11, "H": 0.11},
        "margins": {"D": 0.13, "E": 0.145, "F": 0.16, "G": 0.17, "H": 0.18},
        "pe_low": {c: 24 for c in "CDEFGH"},
        "pe_high": {c: 32 for c in "CDEFGH"},
    },
    "base": {
        "base": 28,
        # growth decelerates toward low-teens; steady margin expansion
        "rev_g": {"C": 0.13, "D": 0.12, "E": 0.11, "F": 0.10, "G": 0.10, "H": 0.09},
        "margins": {"D": 0.125, "E": 0.135, "F": 0.145, "G": 0.15, "H": 0.155},
        "pe_low": {c: 20 for c in "CDEFGH"},
        "pe_high": {c: 26 for c in "CDEFGH"},
    },
    "bear": {
        "base": 52,
        # robotaxi/AV competition + pricing pressure cap growth and margins
        "rev_g": {"C": 0.11, "D": 0.09, "E": 0.08, "F": 0.07, "G": 0.07, "H": 0.06},
        "margins": {"D": 0.115, "E": 0.12, "F": 0.125, "G": 0.125, "H": 0.13},
        "pe_low": {c: 14 for c in "CDEFGH"},
        "pe_high": {c: 20 for c in "CDEFGH"},
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    src = wb[wb.sheetnames[-1]]          # clone the last tab (CEG)
    ws = wb.copy_worksheet(src)
    ws.title = "UBER"

    # Header
    ws["B2"] = "UBER"
    ws["C2"] = datetime(2026, 6, 21)

    for cfg in CASES.values():
        b = cfg["base"]
        ws[f"H{b}"] = SHARES                      # share count

        # Revenue base (2026E) + growth row
        ws[f"C{b + 2}"] = REV_2026
        for col, g in cfg["rev_g"].items():
            ws[f"{col}{b + 3}"] = g

        # Non-GAAP net income base (2026E) + base-year growth
        ws[f"C{b + 5}"] = NI_2026
        ws[f"C{b + 6}"] = NI_G_2026

        # Analyst estimates (2026-2029)
        for col, v in ANALYST.items():
            ws[f"{col}{b + 8}"] = v

        # NI margins (D..H; C stays as formula =NI/Rev)
        for col, m in cfg["margins"].items():
            ws[f"{col}{b + 10}"] = m

        # PE low / high
        for col, v in cfg["pe_low"].items():
            ws[f"{col}{b + 14}"] = v
        for col, v in cfg["pe_high"].items():
            ws[f"{col}{b + 15}"] = v

    wb.save(XLSX)
    print(f"Added UBER tab. Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
