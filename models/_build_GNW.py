#!/usr/bin/env python3
"""Build 6-sheet Genworth Financial (GNW) valuation model."""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()

# Styles
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2F5496")
title_font = Font(bold=True, size=14, color="2F5496")
bold_font = Font(bold=True)
border_all = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border_all

def sc(ws, row, col, value=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = border_all
    return cell

# Data anchors
PRICE = 9.74
SHARES_MM = 383.01
MC_B = 3.72
EV_B = 3.11
BETA = 0.86
RFR = 4.56
ERP = 5.0
KE = RFR + BETA * ERP  # 12.86
KD = 4.2
TAX = 0.20
TOTAL_DEBT_B = 2.31
CASH_B = 2.16
DEBT_EQ_PCT = 23.43

eq_wt = MC_B / (MC_B + TOTAL_DEBT_B)
debt_wt = TOTAL_DEBT_B / (MC_B + TOTAL_DEBT_B)
WACC = eq_wt * KE + debt_wt * KD * (1 - TAX)

print(f"WACC: {WACC:.2f}%")

# Scenario assumptions (insurance-specific: P/B + Forward P/E framework)
# Revenue ~$7B flat. EPS FY26 est $1.08, FY27 $1.10.
# P/B: peers LNC ~0.6x, CNO ~0.8x, JXN ~1.6x, GL ~2.0x
# Bear: P/B stays 0.4x, P/E 12x on $0.90 EPS
# Base: P/B 0.5x, P/E 15x on $1.15 EPS  
# Bull: P/B 0.7x, P/E 18x on $1.40 EPS

bear_5y_rev = 7.0  # flat
base_5y_rev = 7.5  # slight growth
bull_5y_rev = 8.5  # LTCLIB unwind tailwinds

bear_5y_eps = 0.90
base_5y_eps = 1.15
bull_5y_eps = 1.40

bear_pb = 0.40
base_pb = 0.50
bull_pb = 0.70

BOOK_PER_SHARE = 22.88  # MRQ

bear_target = bear_pb * BOOK_PER_SHARE
base_target = base_pb * BOOK_PER_SHARE
bull_target = bull_pb * BOOK_PER_SHARE

bear_upside = (bear_target - PRICE) / PRICE * 100
base_upside = (base_target - PRICE) / PRICE * 100
bull_upside = (bull_target - PRICE) / PRICE * 100

bear_pv = bear_target * 0.25
base_pv = base_target * 0.50
bull_pv = bull_target * 0.25
weighted_fv = bear_pv + base_pv + bull_pv
weighted_upside = (weighted_fv - PRICE) / PRICE * 100

print(f"Bear target: ${bear_target:.2f} ({bear_upside:.1f}%)")
print(f"Base target: ${base_target:.2f} ({base_upside:.1f}%)")
print(f"Bull target: ${bull_target:.2f} ({bull_upside:.1f}%)")
print(f"Weighted FV: ${weighted_fv:.2f} ({weighted_upside:.1f}%)")

# ════════════════════════════════════════
# Sheet 1: Valuation
# ════════════════════════════════════════
ws = wb.active
ws.title = "Valuation"

ws.merge_cells("A1:D1")
ws["A1"] = "Genworth Financial, Inc. (NYSE: GNW) — Valuation Model"
ws["A1"].font = Font(bold=True, size=16, color="2F5496")
ws["A1"].alignment = Alignment(horizontal="center")

title_rows = [
    ("Company:", "Genworth Financial, Inc.", "", "Sector: Financial Services / Insurance"),
    ("Date:", str(date.today()), "", "Status: Initial"),
    ("Ticker:", "NYSE: GNW", "", "Analysts: 1 (K.B.W.)"),
    ("Price:", f"${PRICE}", "", "Analyst PT: $12.00"),
    ("Shares Outstanding:", f"{SHARES_MM:.0f}M", "", "Float: 374M"),
    ("Market Cap:", f"${MC_B:.2f}B", "", "% Inst: 91.95%"),
    ("Enterprise Value:", f"${EV_B:.2f}B", "", "Beta: 0.86"),
    ("Primary Lens:", "P/B and Forward P/E", "", "Stance: Watch"),
]

for i, (l1, v1, l2, v2) in enumerate(title_rows):
    r = i + 3
    sc(ws, r, 1, l1).font = bold_font
    sc(ws, r, 2, v1).font = bold_font
    sc(ws, r, 3, l2).font = bold_font
    sc(ws, r, 4, v2).font = bold_font

r = 12
for c, h in enumerate(["Metric", "Value", "Comment"], 1):
    sc(ws, r, c, h)
style_header_row(ws, r, 3)

metrics = [
    ("P/E (TTM)", f"{18.67:.2f}x", "TTM EPS $0.52 transitional; FY25 includes LTCLIB unwind one-timer"),
    ("Forward P/E (FY26)", f"{PRICE/1.08:.1f}x", "Based on consensus EPS $1.08 for FY26"),
    ("P/S (TTM)", f"{MC_B/7.07:.2f}x", "Low; stable ~$7B revenue; typical for mature insurers"),
    ("P/B (MRQ)", f"{PRICE/22.88:.2f}x", "Deep discount to book — LTC/closed block overhang priced in"),
    ("EV/Revenue", f"{EV_B/7.07:.2f}x", "EV below MC due to net cash ($0.61B)"),
    ("FCF Yield (TTM)", f"{384/MC_B*1000/100:.1f}%", "$384M OCF TTM / $3.72B MC"),
    ("52-Week Range", "$7.13 - $9.99", "Near 52-wk high; +34% YTD from low"),
    ("Profit Margin", "2.96%", "Operating margin ~7.8% healthier; interest drag on net"),
    ("ROE (TTM)", "3.47%", "Below cost of equity (12.86%) — equity rebuilding"),
    ("ROA (TTM)", "0.38%", "Typical for life insurers — $88B asset-heavy balance sheet"),
]

for i, (m, v, c) in enumerate(metrics):
    rr = r + 1 + i
    sc(ws, rr, 1, m)
    sc(ws, rr, 2, v)
    sc(ws, rr, 3, c)

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 55

# ════════════════════════════════════════
# Sheet 2: WACC
# ════════════════════════════════════════
ws2 = wb.create_sheet("WACC")

ws2.merge_cells("A1:C1")
ws2["A1"] = "WACC Calculation — Genworth Financial (GNW)"
ws2["A1"].font = Font(bold=True, size=14, color="2F5496")
ws2["A1"].alignment = Alignment(horizontal="center")

wacc_lines = [
    ("CAPM Components", ""),
    ("Risk-Free Rate (10Y US Treasury)", f"{RFR:.2f}%"),
    ("Equity Risk Premium", f"{ERP:.1f}%"),
    ("Beta (Levered, 5Y Monthly)", f"{BETA:.2f}"),
    ("Cost of Equity (CAPM)", f"{KE:.2f}%"),
    ("", ""),
    ("Cost of Debt", f"{KD:.1f}%"),
    ("Tax Rate (Effective)", f"{TAX*100:.0f}%"),
    ("", ""),
    ("Capital Structure", ""),
    ("Market Cap", f"${MC_B:.2f}B"),
    ("Total Debt", f"${TOTAL_DEBT_B:.2f}B"),
    ("Total Capital", f"${MC_B + TOTAL_DEBT_B:.2f}B"),
    ("Equity Weight", f"{eq_wt*100:.1f}%"),
    ("Debt Weight", f"{debt_wt*100:.1f}%"),
    ("Debt/Equity Ratio", f"{DEBT_EQ_PCT:.1f}%"),
    ("", ""),
    ("WACC", f"{WACC:.2f}%"),
]

for i, (label, val) in enumerate(wacc_lines):
    rr = i + 3
    c1 = sc(ws2, rr, 1, label)
    c2 = sc(ws2, rr, 2, val)
    if label in ("", "CAPM Components", "Cost of Debt", "Capital Structure",):
        c1.font = bold_font if label else None
        c2.font = None if label == "" else bold_font
    if label == "WACC":
        c1.font = Font(bold=True, size=13)
        c2.font = Font(bold=True, size=13, color="2F5496")

ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 16

# ════════════════════════════════════════
# Sheet 3: Scenarios
# ════════════════════════════════════════
ws3 = wb.create_sheet("Scenarios")

ws3.merge_cells("A1:I1")
ws3["A1"] = "Scenario Analysis — Genworth Financial (GNW)"
ws3["A1"].font = Font(bold=True, size=14, color="2F5496")
ws3["A1"].alignment = Alignment(horizontal="center")

# Note row about framework choice
ws3.merge_cells("A2:I2")
ws3["A2"] = "Framework: P/B-based (insurance-specific). Standard FCF multiple not primary due to balance sheet scale; P/B and Forward P/E anchor valuation."
ws3["A2"].font = Font(size=9, italic=True)

headers3 = ["Scenario", "5Y Rev ($B)", "5Y EPS", "Exit P/B", "Target Price",
            "Upside%", "Weight", "Weighted $/share", "Notes"]
r3 = 4
for ci, h in enumerate(headers3, 1):
    sc(ws3, r3, ci, h)
style_header_row(ws3, r3, len(headers3))

scenarios = [
    ("Bear", bear_5y_rev, bear_5y_eps, bear_pb, bear_target, bear_upside, 0.25,
     bear_pv, "Multiple compression stays; LTC reserves held; flat growth"),
    ("Base", base_5y_rev, base_5y_eps, base_pb, base_target, base_upside, 0.50,
     base_pv, "LTCLIB unwind tailwinds; P/B recovers to 0.5x; modest growth"),
    ("Bull", bull_5y_rev, bull_5y_eps, bull_pb, bull_target, bull_upside, 0.25,
     bull_pv, "Full LTC unwind + buybacks; P/B expands to 0.7x; earnings acceleration"),
    ("Weighted Total", "", "", "", "", weighted_upside, 1.00,
     weighted_fv, f"Probability-weighted FV: ${weighted_fv:.2f}"),
]

for i, (name, rv, ep, pb, tp, ups, wt, wv, notes) in enumerate(scenarios):
    rr = r3 + 1 + i
    vals = [name, f"{rv}" if isinstance(rv, float) else rv,
            f"${ep:.2f}" if isinstance(ep, float) else ep,
            f"{pb:.2f}x" if isinstance(pb, float) else pb,
            f"${tp:.2f}" if isinstance(tp, float) else tp,
            f"{ups:.1f}%" if isinstance(ups, float) else ups,
            f"{wt:.0%}" if isinstance(wt, float) else wt,
            f"${wv:.2f}" if isinstance(wv, float) else wv,
            notes]
    for ci, v in enumerate(vals, 1):
        c = sc(ws3, rr, ci, v)
        if name == "Weighted Total":
            c.font = Font(bold=True)

ws3.column_dimensions["A"].width = 16
ws3.column_dimensions["J"].width = 1
for ci in range(2, 10):
    ws3.column_dimensions[get_column_letter(ci)].width = 14

# ════════════════════════════════════════
# Sheet 4: Actuals Source Audit
# ════════════════════════════════════════
ws4 = wb.create_sheet("Actuals Source Audit")

ws4.merge_cells("A1:D1")
ws4["A1"] = "Data Source Audit — Genworth Financial (GNW)"
ws4["A1"].font = Font(bold=True, size=14, color="2F5496")
ws4["A1"].alignment = Alignment(horizontal="center")

audit_header = ["Data Point", "Value", "Source URL", "Date / Notes"]
r4 = 3
for ci, h in enumerate(audit_header, 1):
    sc(ws4, r4, ci, h)
style_header_row(ws4, r4, 4)

audit_rows = [
    ("Stock Price", "$9.74", "finance.yahoo.com/quote/GNW/", "Jul 15, 2026 close"),
    ("Market Cap", "$3.72B", "finance.yahoo.com/quote/GNW/key-statistics/", "Jul 14, 2026"),
    ("Enterprise Value", "$3.11B", "finance.yahoo.com/quote/GNW/key-statistics/", "Jul 14, 2026"),
    ("Shares Outstanding", "383.0M", "finance.yahoo.com/quote/GNW/key-statistics/", "Yahoo implied shares"),
    ("TTM Revenue", "$7.07B", "finance.yahoo.com/quote/GNW/financials/", "TTM as of Jul 2026"),
    ("FY2025 Revenue", "$7.11B", "finance.yahoo.com/quote/GNW/financials/", "Annual income statement"),
    ("FY2024 Revenue", "$7.14B", "finance.yahoo.com/quote/GNW/financials/", "Annual income statement"),
    ("FY2022 Net Income", "$916M", "finance.yahoo.com/quote/GNW/financials/", "Annual — includes LTCLIB"),
    ("FY2023 Net Income", "$76M", "finance.yahoo.com/quote/GNW/financials/", "Annual — trough year"),
    ("FY2025 Net Income", "$223M", "finance.yahoo.com/quote/GNW/financials/", "Annual"),
    ("FY2025 OCF", "$327M", "finance.yahoo.com/quote/GNW/cash-flow/", "Annual"),
    ("TTM OCF", "$384M", "finance.yahoo.com/quote/GNW/cash-flow/", "TTM"),
    ("Total Cash", "$2.16B", "finance.yahoo.com/quote/GNW/key-statistics/", "MRQ Q1 FY26"),
    ("Total Debt", "$2.31B", "finance.yahoo.com/quote/GNW/key-statistics/", "MRQ Q1 FY26"),
    ("Book Value/Share", "$22.88", "finance.yahoo.com/quote/GNW/key-statistics/", "MRQ Q1 FY26"),
    ("P/B Ratio", "0.42x", "finance.yahoo.com/quote/GNW/key-statistics/", "Current"),
    ("Beta", "0.86", "finance.yahoo.com/quote/GNW/key-statistics/", "5Y monthly"),
    ("FY26 EPS Estimate", "$1.08", "finance.yahoo.com/quote/GNW/analysis/", "1 analyst — K.B.W."),
    ("FY27 EPS Estimate", "$1.10", "finance.yahoo.com/quote/GNW/analysis/", "1 analyst"),
    ("Analyst Target", "$12.00", "finance.yahoo.com/quote/GNW/analysis/", "K.B.W. Jul 13, 2026"),
    ("Next Earnings", "Aug 5, 2026", "finance.yahoo.com/quote/GNW/profile/", "8-K scheduled Aug 6"),
    ("Tax Rate", "20%", "finance.yahoo.com/quote/GNW/financials/", "Effective: $79k/$418k TTM"),
    ("Debt/Equity", "23.43%", "finance.yahoo.com/quote/GNW/key-statistics/", "MRQ"),
]

for i, (dp, val, src, note) in enumerate(audit_rows):
    rr = r4 + 1 + i
    sc(ws4, rr, 1, dp)
    sc(ws4, rr, 2, val)
    sc(ws4, rr, 3, src)
    sc(ws4, rr, 4, note)

ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 12
ws4.column_dimensions["C"].width = 35
ws4.column_dimensions["D"].width = 25

# ════════════════════════════════════════
# Sheet 5: Questions
# ════════════════════════════════════════
ws5 = wb.create_sheet("Questions")

ws5.merge_cells("A1:B1")
ws5["A1"] = "Open Questions — Genworth Financial (GNW)"
ws5["A1"].font = Font(bold=True, size=14, color="2F5496")
ws5["A1"].alignment = Alignment(horizontal="center")

questions = [
    "LTCLIB Wind-Down: How much of FY25 EPS ($0.54) is attributable to the long-term care liability transfer to LTC Insurance Benefit Holdings (LTCLIB) vs. organic operations? The unwind is a multi-year process — how much residual reserve benefit remains in future periods?",
    "Closed Block Legacy: The Closed Block segment (post-crisis legacy book) still generates reserves and claims costs. How much does it drag on ROE? Is it fully reserved or are there residual claim volatility risks?",
    "Share Buyback Sustainability: GNW repurchased $269M in FY26 TTM with OCF of only $384M — that's 70% of OCF going to buybacks. Is this sustainable? What happens if OCF dips?",
    "CEO Transition: CEO Thomas McInerney on leave as of July 2026; Jerome Upton (CFO) named interim CEO. What are the governance implications? Is this a planned succession or a health issue?",
    "LTC Reserve Adequacy: Has Genworth fully transferred its LTC liabilities? What residual obligations remain on the balance sheet post-LTCLIB?",
    "Mortgage Insurance Segment Cyclical Risk: The Enact mortgage segment is sensitive to housing markets. With mortgage rates elevated, is origination volume depressed, or are margins benefiting from the higher-rate environment?",
    "Capital Return vs. Regulatory Requirements: As an SIFI-designated entity, does Genworth face capital retention requirements that limit buyback/dividend capacity?",
    "Analyst Coverage: Only 1 analyst (K.B.W.) covers GNW. Does this reflect low institutional interest, or just niche coverage? Low coverage can mean stale estimates.",
    "Book Value Components: With $9.77B total equity vs. $8.75B common equity, what sits in the $1.02B difference? Preferred stock? AOCI? This matters for the P/B denominator.",
    "Investing Cash Flow: FY22-23 showed enormous positive ICashFlow ($1.26B, $861M) — likely asset sales/liquidations. Does the investing CF represent recurring income or one-time portfolio shifts?",
]

for i, q in enumerate(questions):
    rr = i + 3
    sc(ws5, rr, 1, f"{i+1}.", ).font = bold_font
    sc(ws5, rr, 2, q)

ws5.column_dimensions["A"].width = 5
ws5.column_dimensions["B"].width = 90

# ════════════════════════════════════════
# Sheet 6: Sources
# ════════════════════════════════════════
ws6 = wb.create_sheet("Sources")

ws6.merge_cells("A1:B1")
ws6["A1"] = "Sources — Genworth Financial (GNW)"
ws6["A1"].font = Font(bold=True, size=14, color="2F5496")
ws6["A1"].alignment = Alignment(horizontal="center")

sources = [
    ("finance.yahoo.com/quote/GNW/", "Yahoo Finance — price, market cap, statistics, trading data"),
    ("finance.yahoo.com/quote/GNW/financials/", "Yahoo Finance — income statement (annual), all numbers in thousands"),
    ("finance.yahoo.com/quote/GNW/balance-sheet/", "Yahoo Finance — balance sheet (annual)"),
    ("finance.yahoo.com/quote/GNW/cash-flow/", "Yahoo Finance — cash flow statement (annual)"),
    ("finance.yahoo.com/quote/GNW/key-statistics/", "Yahoo Finance — valuation multiples, shares, beta, dividends"),
    ("finance.yahoo.com/quote/GNW/analysis/", "Yahoo Finance — analyst estimates and targets"),
    ("finance.yahoo.com/quote/GNW/profile/", "Yahoo Finance — company profile, earnings dates, executives"),
    ("cnbc.com/quotes/US10Y", "CNBC — 10Y US Treasury yield (~4.56% as of Jul 15, 2026)"),
    ("stockanalysis.com/quote/GNW/", "StockAnalysis — 404, not available; all data from Yahoo Finance"),
    ("Business Wire — Jul 8, 2026", "CEO McInerney leave of absence; CFO Upton named interim"),
    ("finance.yahoo.com/quote/GNW/", "Analyst: K.B.W. maintains Outperform, raises PT from $11 to $12 (Jul 13, 2026)"),
]

for i, (src, desc) in enumerate(sources):
    rr = i + 3
    sc(ws6, rr, 1, f"{i+1}.", ).font = bold_font
    sc(ws6, rr, 2, f"{src} — {desc}")

ws6.column_dimensions["A"].width = 5
ws6.column_dimensions["B"].width = 90

# Save
today = date.today().strftime("%Y-%m-%d")
path = f"/home/refcell/dev/capital/models/[{today}] Genworth Financial Model.xlsx"
wb.save(path)
print(f"\nSaved: {path}")

# Verify
wb2 = openpyxl.load_workbook(path)
print(f"Sheets: {wb2.sheetnames}")
for sn in wb2.sheetnames:
    ws_check = wb2[sn]
    print(f"  {sn}: {ws_check.max_row} rows x {ws_check.max_column} cols")
