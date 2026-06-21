"""Add a HOOD (Robinhood Markets) tab to projections/Projections.xlsx.

Clones the last existing tab (ZETA) to preserve styling/formulas, then
overwrites the input cells with Robinhood numbers.

Robinhood is solidly GAAP-profitable with very high net margins (FY2025 net
margin ~42%), so this tab is built on a straightforward GAAP net income basis.
All $ figures in millions.

Sources (FY2025 results Feb 10 2026; Q1 2026 results Apr 28 2026):
  - FY2025 total net revenues $4,473M (+52% Y/Y); GAAP net income $1,883M
    (~42% margin); diluted EPS $2.05; adjusted EBITDA $2,522M (56% margin)
  - FY2024 revenue $2,951M; net income $1,411M (incl. $424M one-time tax benefit)
  - Q1 2026: revenue $1,070M (+15%), net income $350M, diluted EPS $0.38
  - Diluted shares ~915M (Q1'26 915.0M); buybacks roughly offset SBC -> use 915M
  - Revenue is transaction-heavy (crypto/options/equities) + net interest income,
    so growth is the big cyclical swing; margins are high and fairly resilient.
  - Analyst consensus: ~13-15% forward revenue growth; FY2026 net income ~flat-to-
    +10% off 2025 (2024/2025 had tailwinds); EPS path ~$2.30 -> $3.80 (2026-2029).
  - Price ~$108 (Jun 18 2026); trailing P/E ~47x, forward P/E ~35x (premium multiple)
"""

from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "projections" / "Projections.xlsx"

# Shared inputs
SHARES = 915          # diluted shares (M); ~915M Q1'26, buybacks ~offset SBC
REV_2026 = 5100       # FY2026E total net revenues (~14% on 2025's $4,473M)
NI_2026 = 2100        # FY2026E GAAP net income (~41% margin, EPS ~$2.30)
NI_G_2026 = 0.12      # net income growth 2025->2026 ($1,883M -> $2,100M)

# Analyst GAAP net income estimates 2026-2029 (EPS ~$2.30/$2.70/$3.25/$3.80 x 915).
ANALYST = {"C": 2100, "D": 2500, "E": 3000, "F": 3500}

# Per-case assumptions.  Each template block is offset by 24 rows:
#   Bull base row 4, Base 28, Bear 52.  Within a block, relative rows:
#     +6 revenue, +7 revenue growth, +9 net income, +10 NI growth,
#     +12 analyst est, +14 NI margins, +18 PE low, +19 PE high, H(+0) shares.
# Revenue growth is the main swing for a trading-driven platform; net margins are
# high but compress in the bear (lower rates hit net interest income).
CASES = {
    "bull": {
        "base": 4,
        # crypto/options boom + product velocity -> sustained ~20%+ growth
        "rev_g": {"C": 0.14, "D": 0.22, "E": 0.22, "F": 0.20, "G": 0.18, "H": 0.18},
        "margins": {"D": 0.42, "E": 0.43, "F": 0.44, "G": 0.45, "H": 0.46},
        "pe_low": {c: 32 for c in "CDEFGH"},
        "pe_high": {c: 42 for c in "CDEFGH"},
    },
    "base": {
        "base": 28,
        # consensus ~13-15% growth, normalizing from hypergrowth
        "rev_g": {"C": 0.14, "D": 0.14, "E": 0.14, "F": 0.13, "G": 0.12, "H": 0.12},
        "margins": {"D": 0.41, "E": 0.42, "F": 0.42, "G": 0.43, "H": 0.43},
        "pe_low": {c: 25 for c in "CDEFGH"},
        "pe_high": {c: 33 for c in "CDEFGH"},
    },
    "bear": {
        "base": 52,
        # trading volumes normalize, rate cuts pressure net interest revenue
        "rev_g": {"C": 0.14, "D": 0.08, "E": 0.07, "F": 0.06, "G": 0.06, "H": 0.06},
        "margins": {"D": 0.39, "E": 0.38, "F": 0.37, "G": 0.36, "H": 0.36},
        "pe_low": {c: 16 for c in "CDEFGH"},
        "pe_high": {c: 22 for c in "CDEFGH"},
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    src = wb[wb.sheetnames[-1]]          # clone the last tab (ZETA)
    ws = wb.copy_worksheet(src)
    ws.title = "HOOD"

    # Header
    ws["B2"] = "HOOD"
    ws["C2"] = datetime(2026, 6, 21)

    for cfg in CASES.values():
        b = cfg["base"]
        ws[f"H{b}"] = SHARES                      # share count

        # Revenue base (2026E) + growth row
        ws[f"C{b + 2}"] = REV_2026
        for col, g in cfg["rev_g"].items():
            ws[f"{col}{b + 3}"] = g

        # GAAP net income base (2026E) + base-year growth
        ws[f"C{b + 5}"] = NI_2026
        ws[f"C{b + 6}"] = NI_G_2026

        # Analyst estimates (2026-2029)
        for col, v in ANALYST.items():
            ws[f"{col}{b + 8}"] = v

        # NI margins (D..H; C stays as formula =NI/Rev)
        for col, m in cfg["margins"].items():
            ws[f"{col}{b + 10}"] = m

        # PE low / high
        for col, v in cfg["pe_low"].items():
            ws[f"{col}{b + 14}"] = v
        for col, v in cfg["pe_high"].items():
            ws[f"{col}{b + 15}"] = v

    wb.save(XLSX)
    print(f"Added HOOD tab. Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
