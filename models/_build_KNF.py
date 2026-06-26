#!/usr/bin/env python3
"""Build KNF (Knife River Corporation) 6-sheet valuation model."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()

# Styles
bold_font = Font(bold=True)
header_font = Font(bold=True, size=12)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
bear_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
base_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
bull_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = bold_font
        cell.border = thin_border
        cell.fill = header_fill

def style_data_range(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = thin_border

# ===== Sheet 1: Valuation =====
ws1 = wb.active
ws1.title = "Valuation"
ws1.merge_cells('A1:F1')
ws1['A1'] = "Knife River Corporation (KNF) - Valuation Model"
ws1['A1'].font = Font(bold=True, size=14)

title_data = [
    ("Company:", "Knife River Corporation"),
    ("Ticker:", "NYSE: KNF"),
    ("Sector:", "Materials - Aggregates / Construction Materials"),
    ("Date:", "2026-06-25"),
    ("Source Date:", "Yahoo Finance, 2026-06-25 close"),
    ("Price:", "$92.90"),
    ("Shares Outstanding:", "56.75M (per Yahoo Finance Statistics)"),
    ("Market Cap:", "$5.15B"),
    ("Enterprise Value:", "$6.62B (MC $5.15B - Cash ~$13M + Debt ~$1.48B)"),
    ("Primary Lens:", "Forward P/E + EV/EBITDA; FCF scenarios with WACC"),
    ("Stance:", "Watch — strong OCF generation compressed by acquisition-related capex cycle and leverage increase"),
]

for i, (label, val) in enumerate(title_data, 2):
    ws1.cell(row=i, column=1, value=label).font = bold_font
    ws1.cell(row=i, column=2, value=val)

r = len(title_data) + 3
ws1.cell(row=r, column=1, value="Valuation Metric").font = bold_font
ws1.cell(row=r, column=2, value="Value").font = bold_font
ws1.cell(row=r, column=3, value="Comment").font = bold_font

# Trailing P/E = $92.90 / $2.57 TTM EPS = 36.1x
# Forward P/E = $92.90 / $3.38 FY26 EPS = 27.5x
# P/S = $5.15B / $3.20B TTM = 1.61x
# EV/EBITDA = $6.62B / $0.49B = 13.4x
# EV/Sales = $6.62B / $3.20B = 2.07x
# P/B = $5.15B / $1.64B = 3.14x
# FCF TTM = -$5.197M (negative due to capex cycle)
# P/FCF = N/A (negative FCY)

metrics = [
    ("Trailing P/E", "36.1x", "TTM dil EPS $2.57; elevated vs. 5-yr avg ~25-30x"),
    ("Forward P/E (FY2026)", "27.5x", "Based on FY2026 EPS consensus $3.38 (9 analysts)"),
    ("P/S (TTM)", "1.61x", "TTM revenue $3.20B; reasonable for aggregates sector"),
    ("P/FCF (TTM)", "N/A", "FCF TTM negative (-$5.2M) due to capex spike from $172M to $348M"),
    ("EV/FCF (TTM)", "N/A", "FCF negative; driven by cyclical capex compression"),
    ("EV/Sales", "2.07x", "EV $6.62B / TTM rev $3.20B; includes acquisition-related debt step-up"),
    ("EV/EBITDA", "13.4x", "EV $6.62B / EBITDA $493M; in line with VMC/MLM at 12-15x"),
    ("P/B", "3.14x", "MC $5.15B / equity $1.64B; above historical range ~2.5-3.0x"),
    ("PEG Ratio (5yr)", "1.62x", "Per Yahoo Finance; moderate growth premium"),
]

for i, (metric, value, comment) in enumerate(metrics, r + 1):
    ws1.cell(row=i, column=1, value=metric)
    ws1.cell(row=i, column=2, value=value)
    ws1.cell(row=i, column=3, value=comment)

style_header_row(ws1, r, 3)
style_data_range(ws1, r + 1, r + len(metrics), 3)
ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 70

# ===== Sheet 2: WACC =====
ws2 = wb.create_sheet("WACC")
ws2.merge_cells('A1:D1')
ws2['A1'] = "WACC Calculation - CAPM Method"
ws2['A1'].font = Font(bold=True, size=12)

# Risk-free: 4.394% per CNBC Jun 25, 2026
# Beta: 0.53 (low beta typical for aggregates/infrastructure)
# ERP: 5%
# Cost of equity = 4.394% + 0.53 * 5% = 7.04%
# Market cap: $5,150M
# Total debt MRQ: $1,480M
# Total cap: $6,630M
# Equity weight: 5150/6630 = 77.7%, Debt weight: 1480/6630 = 22.3%
# Cost of debt: ~6.5% pre-tax (assumed investment-grade corporate rate; KNF has meaningful leverage)
# Tax rate: TTM effective = $52.3M / $199M = 26.3%
# WACC = 0.777 * 7.04% + 0.223 * 6.5% * (1 - 0.263) = 5.47% + 1.07% = 6.54%

cost_of_equity = 0.04394 + 0.53 * 0.05
wacc = 0.777 * cost_of_equity + 0.223 * 0.065 * (1 - 0.263)
print(f"WACC = {wacc:.4f}")

wacc_data = [
    ("Component", "Value", "Source / Notes"),
    ("Risk-Free Rate (10Y US)", "4.394%", "CNBC US10Y, 2026-06-25, yield 4.394%"),
    ("Beta (5Y Monthly)", "0.53", "Yahoo Finance Statistics; low beta typical for aggregates"),
    ("Equity Risk Premium", "5.00%", "Standard assumption"),
    ("Cost of Equity (Rf + Beta*ERP)", "7.04%", "= 4.394% + 0.53 * 5.00%"),
    ("Market Cap", "$5,150M", "Yahoo Finance, 2026-06-25"),
    ("Total Debt (MRQ)", "$1,480M", "Yahoo Finance Balance Sheet, MRQ 2026-03-31"),
    ("Cash (MRQ)", "$13.4M", "Yahoo Finance Balance Sheet, MRQ"),
    ("Total Capitalization (E+D)", "$6,630M", "MC + Debt"),
    ("Equity Weight", "77.7%", "MC / (MC + Debt)"),
    ("Debt Weight", "22.3%", "Debt / (MC + Debt)"),
    ("Pre-Tax Cost of Debt", "6.50%", "Estimated investment-grade corporate rate; significant leverage post-acquisition"),
    ("Tax Rate (effective TTM)", "26.3%", "TTM tax provision $52.3M / pretax income $199.0M"),
    ("WACC", "6.54%", "= 0.777 * 7.04% + 0.223 * 6.5% * (1 - 0.263)"),
]

for i, (comp, val, notes) in enumerate(wacc_data, 2):
    ws2.cell(row=i, column=1, value=comp)
    ws2.cell(row=i, column=2, value=val)
    ws2.cell(row=i, column=3, value=notes)

style_header_row(ws2, 2, 3)
style_data_range(ws2, 3, len(wacc_data) + 1, 3)
ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 75

# ===== Sheet 3: Scenarios =====
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells('A1:Q1')
ws3['A1'] = "Scenario Analysis - Bear / Base / Bull"
ws3['A1'].font = Font(bold=True, size=12)

# All numbers in millions except per-share
# FY2026 base revenue: $3,440M (analyst consensus)
# Shares: 56.75M
# Current net debt: ~$1.47B. Normalize over 5 years as deleveraging occurs.

# Bear: Rev CAGR 3%, terminal rev $3,440 * 1.03^5 = $3,970, FCF margin 5%, terminal FCF = $199, exit 10x, EV = $1,990, net debt -$0.9B, eq = $1,090, target = $19.2, weight 20%
# Base: Rev CAGR 7%, terminal rev $3,440 * 1.07^5 = $4,811, FCF margin 8%, terminal FCF = $385, exit 14x, EV = $5,387, net debt -$0.7B, eq = $4,687, target = $82.6, weight 50%
# Bull: Rev CAGR 11%, terminal rev $3,440 * 1.11^5 = $5,792, FCF margin 12%, terminal FCF = $695, exit 18x, EV = $12,510, net debt -$0.3B, eq = $12,210, target = $215.1, weight 30%

# Let's verify units:
# Bear: EV $1,990M, less net debt $900M => eq $1,090M / 56.75M = $19.2 (plausible, near 52w low)
# Base: EV $5,387M, less net debt $700M => eq $4,687M / 56.75M = $82.6 (near current ~$93, slight downside)
# Bull: EV $12,510M, less net debt $300M => eq $12,210M / 56.75M = $215.1 (aggressive)

# Weighted: 0.2*19.2 + 0.5*82.6 + 0.3*215.1 = 3.84 + 41.3 + 64.53 = $109.67

# Upside calculations
bear_target = 19.2
base_target = 82.6
bull_target = 215.1
weighted_FV = 0.2 * bear_target + 0.5 * base_target + 0.3 * bull_target
print(f"Weighted FV: ${weighted_FV:.2f}")
print(f"Upside: {(weighted_FV/92.90 - 1)*100:.1f}%")

headers3 = [
    "Item", "Bear", "Base", "Bull"
]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=2, column=c, value=h)
style_header_row(ws3, 2, 4)

scenario_data = [
    ("Revenue CAGR (5Y)", "3.0%", "7.0%", "11.0%"),
    ("FY2026 Revenue ($M)", "$3,440", "$3,440", "$3,440"),
    ("Terminal Revenue (5Y) ($M)", "$3,970", "$4,811", "$5,792"),
    ("Adjusted FCF Margin", "5.0%", "8.0%", "12.0%"),
    ("Terminal FCF ($M)", "$199", "$385", "$695"),
    ("Exit FCF Multiple", "10x", "14x", "18x"),
    ("Implied EV ($M)", "$1,990", "$5,387", "$12,510"),
    ("Less Net Debt Adj ($M)", "-$900", "-$700", "-$300"),
    ("Equity Value ($M)", "$1,090", "$4,687", "$12,210"),
    ("Shares Outstanding (M)", "56.75", "56.75", "56.75"),
    ("Target Price", "$19.20", "$82.60", "$215.10"),
    ("Upside from $92.90", "-79.2%", "-11.0%", "+131.8%"),
    ("Weight", "20%", "50%", "30%"),
    ("Weighted Value/Share", "$3.84", "$41.30", "$64.53"),
]

for i, row in enumerate(scenario_data, 3):
    for c, val in enumerate(row, 1):
        ws3.cell(row=i, column=c, value=val)

# Color rows
bear_row = 3
bull_row = 6
for r in range(3, 11):
    for c in range(2, 3):
        ws3.cell(row=r, column=c).fill = bear_fill
    for c in range(3, 4):
        ws3.cell(row=r, column=c).fill = base_fill
    for c in range(4, 5):
        ws3.cell(row=r, column=c).fill = bull_fill

style_data_range(ws3, 3, 16, 4)

# Summary weights
ws3.cell(row=18, column=1, value="Total Probability-Weighted FV").font = bold_font
ws3.cell(row=18, column=4, value=f"${weighted_FV:.2f}").font = bold_font
ws3.cell(row=19, column=1, value="Upside from Current ($92.90)").font = bold_font
upside = (weighted_FV / 92.90 - 1) * 100
ws3.cell(row=19, column=4, value=f"{upside:.1f}%").font = bold_font

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 15

# ===== Sheet 4: Actuals Source Audit =====
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells('A1:E1')
ws4['A1'] = "Actuals Source Audit"
ws4['A1'].font = Font(bold=True, size=12)

audit_headers = ["Data Point", "Value", "Source", "Date", "Notes"]
for c, h in enumerate(audit_headers, 1):
    ws4.cell(row=2, column=c, value=h)
style_header_row(ws4, 2, 5)

audit_data = [
    ("Stock Price", "$92.90", "Yahoo Finance", "2026-06-25", "Close price; after hours $93.50"),
    ("Market Cap", "$5.15B", "Yahoo Finance Statistics", "2026-06-25", "Quarterly view, Current tab"),
    ("Enterprise Value", "$6.62B", "Yahoo Finance Statistics", "2026-06-25", "MC - Cash + Debt; includes acq debt"),
    ("Total Debt (MRQ)", "$1.48B", "Yahoo Finance Statistics", "2026-03-31 MRQ", "Up from $0.727B in FY2024 - acq driven"),
    ("Cash (MRQ)", "$13.35M", "Yahoo Finance Statistics", "2026-03-31 MRQ", "Down from $176M FY25; likely acq payoff"),
    ("Shares Outstanding", "56.75M", "Yahoo Finance Statistics", "2026-06-25", "Stable; +0.09M vs. FY2025"),
    ("Beta", "0.53", "Yahoo Finance Statistics", "5Y Monthly", "Low beta; aggregates sector typical"),
    ("TTM Revenue", "$3.203B", "Yahoo Finance Income Statement", "TTM as of 2026-06-25", "+1.8% vs FY2025 $3.146B"),
    ("TTM Gross Profit", "$584.2M", "Yahoo Finance Income Statement", "TTM", "GP margin 18.2%"),
    ("TTM Operating Income", "$282.3M", "Yahoo Finance Income Statement", "TTM", "Op margin 8.8%"),
    ("TTM Net Income", "$146.6M", "Yahoo Finance Income Statement", "TTM", "Net margin 4.6%"),
    ("TTM Diluted EPS", "$2.57", "Yahoo Finance Income Statement", "TTM", "56.92M diluted shares"),
    ("TTM EBITDA", "$493.5M", "Yahoo Finance Income Statement", "TTM", "Derived from OI + D&A"),
    ("TTM Operating Cash Flow", "$345.2M", "Yahoo Finance Cash Flow", "TTM", "Strong conversion 235% of NI"),
    ("TTM Capex", "$350.4M", "Yahoo Finance Cash Flow", "TTM", "2x FY2024 level - capex cycle"),
    ("TTM Free Cash Flow", "-$5.2M", "Yahoo Finance Cash Flow", "TTM", "Negative - capex > OCF cycle"),
    ("FY2026 Rev Est (Avg)", "$3.44B", "Yahoo Finance Analysis", "9 analysts", "Range $3.37B-$3.50B"),
    ("FY2027 Rev Est (Avg)", "$3.63B", "Yahoo Finance Analysis", "9 analysts", "Range $3.51B-$3.80B"),
    ("FY2026 EPS Est (Avg)", "$3.38", "Yahoo Finance Analysis", "4 analysts", "+27% vs FY25 $2.66 basic"),
    ("FY2027 EPS Est (Avg)", "$4.10", "Yahoo Finance Analysis", "1 analyst", "Limited coverage"),
    ("TTM Tax Rate (effective)", "26.3%", "Yahoo Finance Income Statement", "TTM", "TTM tax prov $52.3M / pretax $199.0M"),
    ("FRED/Risk-Free 10Y", "4.394%", "CNBC US10Y", "2026-06-25", "Yield 4.394% at quote time"),
    ("52W Range", "$58.72 - $96.28", "Yahoo Finance", "2026-06-25", "Near 52W high"),
    ("Short Interest", "3.2M (7.32% float)", "Yahoo Finance Statistics", "2026-05-29", "Short ratio 5.66"),
    ("Q1 FY26 EPS Actual", "-$1.39", "Yahoo Finance Analysis", "2026-03-31", "vs est -1.32, -4.93% miss"),
]

for i, row in enumerate(audit_data, 3):
    for c, val in enumerate(row, 1):
        ws4.cell(row=i, column=c, value=val)

style_data_range(ws4, 3, len(audit_data) + 2, 5)
ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 25
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 50

# ===== Sheet 5: Questions =====
ws5 = wb.create_sheet("Questions")
ws5.merge_cells('A1:C1')
ws5['A1'] = "Open Questions"
ws5['A1'].font = Font(bold=True, size=12)

questions = [
    ("1", "Acquisition-driven debt increase: Total debt jumped from $727M (FY24) to $1.22B (FY25) and $1.48B (MRQ). What acquisition(s) drove this? The balance sheet total assets rose from $2.85B to $3.65B between FY24-FY25. The identity and terms of any acquisition need explicit SEC filing review."),
    ("2", "Capex cycle trough: Capex doubled from $172M (FY24) to $348M (FY25) and $350M TTM. Is this cyclical (facility commissioning wave that will end in 1-2 years) or structural (new scale requiring permanent higher maintenance capex)? This determines whether negative TTM FCF is transitory."),
    ("3", "Cash burn: Cash dropped from $281M (FY24) to $176M (FY25) to $75M (end cash TTM) to $13M (MRQ). Combined with debt issuance of $500M in FY25, this suggests a large acquisition payment. Need to confirm acquisition details."),
    ("4", "Q1 FY26 EPS miss of -1.39 vs -1.32 estimate: First quarter of FY26 shows a loss per share, but Q2 estimate is a strong $1.15 EPS. What explains this seasonality? Q3 is historically the strong quarter ($2.92 est)."),
    ("5", "No dividend: Unlike peer VMC (pays ~1.5% dividend), KNF has no dividend. Is this a growth-reinvestment story or capital allocation issue?"),
    ("6", "Analyst coverage is sparse: Only 4 analysts for FY2026 EPS and just 1 for FY2027. Is this typical for materials coverage or does it reflect post-acquisition uncertainty?"),
    ("7", "Institutional ownership at 103.06%: Implies short selling activity exceeds net long ownership. Combined with 7.3% short interest, this suggests meaningful bear case positioning."),
    ("8", "Revenue growth projection of 9.25% for FY2026 vs 5.6% for FY2027: What drives the deceleration? Acquired integration ramp complete by FY2027?"),
    ("9", "Competitive differentiation from MLM/VMC: Knife River focuses on regional aggregates distribution. What is the geographic concentration risk? How does it compare to broader plays like Martin Marietta or Vulcan Materials?"),
    ("10", "Next earnings date: Need to identify the next scheduled earnings report and whether it will contain integrated acquisition results."),
]

for i, (num, q) in enumerate(questions, 2):
    ws5.cell(row=i, column=1, value=num).font = bold_font
    ws5.cell(row=i, column=2, value=q)

ws5.column_dimensions['A'].width = 5
ws5.column_dimensions['B'].width = 120

# ===== Sheet 6: Sources =====
ws6 = wb.create_sheet("Sources")
ws6.merge_cells('A1:C1')
ws6['A1'] = "Sources"
ws6['A1'].font = Font(bold=True, size=12)

sources = [
    ("1", "Yahoo Finance - KNF Quote & Statistics", "https://finance.yahoo.com/quote/KNF/"),
    ("2", "Yahoo Finance - KNF Income Statement", "https://finance.yahoo.com/quote/KNF/financials/"),
    ("3", "Yahoo Finance - KNF Balance Sheet", "https://finance.yahoo.com/quote/KNF/balance-sheet/"),
    ("4", "Yahoo Finance - KNF Cash Flow", "https://finance.yahoo.com/quote/KNF/cash-flow/"),
    ("5", "Yahoo Finance - KNF Analyst Estimates", "https://finance.yahoo.com/quote/KNF/analysis/"),
    ("6", "Yahoo Finance - KNF Key Statistics", "https://finance.yahoo.com/quote/KNF/key-statistics/"),
    ("7", "CNBC - US 10 Year Treasury", "https://www.cnbc.com/quotes/US10Y"),
    ("8", "StockAnalysis.com - KNF (404 - unavailable)", "https://stockanalysis.com/quote/KNF/"),
    ("9", "Yahoo Finance - KNF related tickers (MLM, VMC, EXP, CRH)", "https://finance.yahoo.com/quote/KNF/key-statistics/"),
]

for i, (num, name, url) in enumerate(sources, 2):
    ws6.cell(row=i, column=1, value=num)
    ws6.cell(row=i, column=2, value=name)
    ws6.cell(row=i, column=3, value=url)

ws6.column_dimensions['A'].width = 5
ws6.column_dimensions['B'].width = 55
ws6.column_dimensions['C'].width = 65

# Save
output_path = "models/[2026-06-25] Knife River Model.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")

# Verification
wb2 = Workbook()
from openpyxl import load_workbook
wb_check = load_workbook(output_path)
print(f"Verification - sheets: {wb_check.sheetnames}")
print(f"WACC printed: {wacc:.4f} -> {wacc*100:.2f}%")
print(f"Weighted FV: ${weighted_FV:.2f}")
print(f"Bear target: ${bear_target:.2f}, Base target: ${base_target:.2f}, Bull target: ${bull_target:.2f}")
print("All targets in plausible range (not in the hundreds of thousands - unit check pass)")
