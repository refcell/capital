"""Build a 6-sheet Palantir Technologies (PLTR) valuation model.
Price snapshot: $113.50 on 2026-06-24. Data source: Yahoo Finance.
"""
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# ── Data ──
TICKER = "PLTR"
DATE = datetime(2026, 6, 24)
PRICE = 113.50
MC = 279.77        # $B
EV = 271.95       # $B
CASH = 8.03
DEBT = 0.229
NET_CASH = CASH - DEBT
SHARES_M = round(MC * 1000 / PRICE)   # 2464M
BETA = 1.51
RISK_FREE = 0.04402
ERP = 0.05
COE = RISK_FREE + BETA * ERP         # ~0.119
COD = 0.04
TAX = 0.15
EQ_W = MC / (MC + DEBT)
DEB_W = DEBT / (MC + DEBT)
WACC = round(EQ_W * COE + DEB_W * COD * (1 - TAX), 4)

# Historical ($M from Yahoo Income Statement — all numbers in thousands → /1000)
REV = {"TTM": 5224.2, "2025": 4475.4, "2024": 2865.5, "2023": 2225.0, "2022": 1905.9}
GP  = {"TTM": 4392.2, "2025": 3686.3, "2024": 2299.5, "2023": 1793.9, "2022": 1497.3}
OP  = {"TTM": 1992.0, "2025": 1414.0, "2024": 310.4,  "2023": 120.0,  "2022": -161.2}
NI  = {"TTM": 2281.5,"2025": 1625.0, "2024": 462.2,  "2023": 209.8,  "2022": -373.7}
EPS = {"TTM": 0.89,  "2025": 0.63,   "2024": 0.19,   "2023": 0.09,   "2022": -0.18}
OCF = {"TTM": 2723.4,"2025": 2134.5, "2024": 1153.9, "2023": 712.2,  "2022": 223.7}
CPX = {"TTM": 35.1,  "2025": 33.9,   "2024": 12.6,   "2023": 15.1,   "2022": 40.0}
FCF = {"TTM": 2688.3,"2025": 2100.6, "2024": 1141.2, "2023": 697.1,  "2022": 183.7}
EBD = {"TTM": 2018.3,"2025": 1440.2, "2024": 342.0,  "2023": 153.3,  "2022": -138.7}

# Multiples (Yahoo Statistics)
PE_TTM = 131.12; PE_FWD = 79.37; PS_TTM = 57.40; PB = 33.11
EV_REV = 52.06; EV_EB = 134.75
P_FCF = round(MC / (FCF["TTM"] / 1000), 1)
EV_FCF = round(EV / (FCF["TTM"] / 1000), 1)

# ── Scenarios ──────────────────────────────────────────────────────────────
SCENARIOS = {}
for name, cagr, fcf_m, mult, wt in [
    ("Bear", 0.35, 0.40, 45, 0.25),
    ("Base", 0.43, 0.50, 55, 0.50),
    ("Bull", 0.52, 0.55, 65, 0.25),
]:
    tr = REV["TTM"] * (1 + cagr) ** 5        # $ millions
    tf = tr * fcf_m                          # $ millions
    ie = tf * mult                           # $ millions
    eq = ie + NET_CASH * 1000               # add net cash in millions too
    tp = round(eq / SHARES_M, 2)            # both in millions → $ / share
    up = round((tp / PRICE - 1) * 100, 1)
    SCENARIOS[name] = {
        "cagr": cagr, "tr": round(tr, 1), "fcf_m": fcf_m,
        "tf": round(tf, 1), "mult": mult, "ie": round(ie, 1),
        "eq": round(eq, 1), "tp": tp, "up": up, "wt": wt,
        "wv": round(tp * wt, 2),
    }

tot_wv = round(sum(s["wv"] for s in SCENARIOS.values()), 2)
tot_up = round((tot_wv / PRICE - 1) * 100, 1)

# ── Styles ─────────────────────────────────────────────────────────────────
BF = Font(bold=True)
B14 = Font(bold=True, size=14)
B12 = Font(bold=True, size=12)
BD = Border(left=Side("thin"), right=Side("thin"),
            top=Side("thin"), bottom=Side("thin"))
D2 = "0.00"
D4 = "0.0000"
PP = "0.0%"
PP2 = "0.00%"
DB = "$#,##0.0\"B\""
DM = "$#,##0.0"
DP = "$#,##0.00"


def c(ws, r, co, v, fmt=None, font=None, border=None):
    """Write a cell."""
    cell = ws.cell(row=r, column=co, value=v)
    if fmt is not None:
        cell.number_format = fmt
    if font is not None:
        cell.font = font
    if border is not None:
        cell.border = border
    return cell


def title(ws, text):
    ws.merge_cells("A1:F1")
    c(ws, 1, 1, text, font=B14)


def hdr_row(ws, r, cols):
    for j, h in enumerate(cols, 1):
        c(ws, r, j, h, font=B12, border=BD)


def data_row(ws, r, vals):
    for j, v in enumerate(vals, 1):
        c(ws, r, j, v, border=BD)


# ── Build ──────────────────────────────────────────────────────────────────
OUT = Path("/home/refcell/dev/capital/models/[2026-06-24] Palantir Model.xlsx")
wb = openpyxl.Workbook()

# ── Sheet 1: Valuation ────────────────────────────────────────────────────
ws = wb.active
ws.title = "Valuation"
title(ws, f"{TICKER} \u2014 {DATE.strftime('%Y-%m-%d')} Valuation Model")

items = [
    ("Company", "Palantir Technologies Inc."),
    ("Ticker", "NASDAQ: PLTR"),
    ("Date", DATE.strftime("%Y-%m-%d")),
    ("Price", PRICE),
    ("Shares Outstanding (M)", SHARES_M),
    ("Market Cap ($B)", MC),
    ("Enterprise Value ($B)", EV),
    ("Cash ($B)", CASH),
    ("Debt ($B)", DEBT),
    ("Net Cash ($B)", NET_CASH),
    ("Primary Valuation Lens", "Forward P/E + Scenario Analysis"),
    ("Stance", "Watch \u2014 extreme multiple, waiting for growth validation"),
]
for i, (lab, val) in enumerate(items, 2):
    c(ws, i, 1, lab, font=BF, border=BD)
    c(ws, i, 2, val, border=BD)

c(ws, 10, 1, "Key Valuation Metrics", font=B12)

metrics = [
    ("P/E (TTM)", PE_TTM, "TTM; reflects $0.89 diluted EPS on $5.22B revenue"),
    ("Forward P/E", PE_FWD, "FY26E EPS $1.47; ~55x on FY27E"),
    ("P/S (TTM)", PS_TTM, "57x; only defensible if AI premium sustains"),
    ("P/FCF (TTM)", P_FCF, "FCF ~$2.69B; ~104x; extreme for software"),
    ("EV/FCF", EV_FCF, "Enterprise lens; ~101x on TTM FCF"),
    ("EV/Revenue", EV_REV, "EV at 52x revenue"),
    ("EV/EBITDA", EV_EB, "135x; needs 4x compression to normalize"),
]
for i, (nm, val, txt) in enumerate(metrics, 11):
    c(ws, i, 1, nm, font=BF, border=BD)
    c(ws, i, 2, val, fmt=D2, border=BD)
    c(ws, i, 3, txt, border=BD)
    ws.merge_cells(f"C{i}:F{i}")

# ── Sheet 2: WACC ─────────────────────────────────────────────────────────
w2 = wb.create_sheet("WACC")
title(w2, "WACC Calculation \u2014 CAPM")

wd = [
    ("Risk-Free Rate (10Y US)", RISK_FREE, PP2, "CNBC 6/24/26: 4.402%"),
    ("Equity Risk Premium", ERP, PP, "Standard 5%"),
    ("Beta (5Y Monthly)", BETA, D2, "Levered beta"),
    ("Cost of Equity", COE, PP2, f"{RISK_FREE*100:.2f}% + 1.51 x 5.00%"),
    ("Cost of Debt", COD, PP, f"~4.0%"),
    ("Tax Rate", TAX, PP, f"Forward est 15%"),
    ("", None, None, ""),
    ("Market Cap ($B)", MC, DB, f"Intraday 6/24/26"),
    ("Total Debt ($B)", DEBT, DB, "Capital leases only"),
    ("Equity Weight", round(EQ_W, 4), D4, f"{EQ_W*100:.2f}%"),
    ("Debt Weight", round(DEB_W, 4), D4, f"{DEB_W*100:.2f}%"),
    ("", None, None, ""),
    ("WACC", WACC, PP2, f"= 0.9992 x {COE*100:.2f}% + small debt adj"),
]
for i, (lab, val, fmt, note) in enumerate(wd, 2):
    c(w2, i, 1, lab, font=BF if lab else None, border=BD)
    c(w2, i, 2, val, fmt=fmt, border=BD)
    c(w2, i, 3, note, border=BD)
    if note:
        w2.merge_cells(f"C{i}:E{i}")

# ── Sheet 3: Scenarios ────────────────────────────────────────────────────
w3 = wb.create_sheet("Scenarios")
title(w3, "Bear / Base / Bull Scenario Analysis")

cols3 = ["Scenario", "Rev CAGR 5Y", "Term Rev ($B)", "FCF Margin",
         "Term FCF ($B)", "Exit Mult", "Implied EV ($B)",
         "Less Net Debt ($B)", "Shares (M)",
         "Target Price", "Upside %", "Weight", "Wtd Value/Share"]
hdr_row(w3, 3, cols3)

for i, (nm, s) in enumerate(SCENARIOS.items(), 4):
    vals = [nm, s["cagr"], s["tr"], s["fcf_m"], s["tf"], s["mult"],
            s["ie"], -NET_DEBT if False else -NET_CASH, SHARES_M,
            s["tp"], s["up"] / 100, s["wt"], s["wv"]]
    for j, v in enumerate(vals, 1):
        c(w3, i, j, v, border=BD)
    w3.cell(row=i, column=2).number_format = PP
    w3.cell(row=i, column=4).number_format = PP
    w3.cell(row=i, column=11).number_format = PP
    w3.cell(row=i, column=12).number_format = PP

tr = 7
c(w3, tr, 1, "TOTAL", font=BF, border=BD)
c(w3, tr, 12, 1.0, fmt=PP, font=BF, border=BD)
c(w3, tr, 13, tot_wv, fmt=DP, font=BF, border=BD)
c(w3, tr + 1, 12, "Prob-Weighted FV ($)", font=BF, border=BD)
c(w3, tr + 1, 13, tot_wv, fmt=DP, font=BF, border=BD)
c(w3, tr + 2, 1, f"Upside from ${PRICE}", font=BF)
c(w3, tr + 2, 2, tot_up / 100, fmt=PP, font=BF)

notes = [
    "Notes:",
    "1. Bear: 35% CAGR, 45x exit. Terminal ~2031 (5Y fwd).",
    "2. Base: 43% CAGR, 55x exit. Consensus-aligned.",
    "3. Bull: 52% CAGR, 65x exit. AI premium sustained.",
    f"4. Net cash adj = ${NET_CASH:.2f}B. Shares = {SHARES_M}M (MC/price recon).",
]
for i, n in enumerate(notes, tr + 3):
    c(w3, i, 1, n)
    w3.merge_cells(f"A{i}:M{i}")

# ── Sheet 4: Actuals Source Audit ────────────────────────────────────────
w4 = wb.create_sheet("Actuals Source Audit")
title(w4, "Data Point Source Audit")

hdr_row(w4, 2, ["Data Point", "Value", "Source", "Date", "Notes"])

audit = [
    ("Price", f"${PRICE}", "yahoo.com/quote/PLTR", "2026-06-24", "Close"),
    ("Market Cap", f"${MC}B", "yahoo.com/key-statistics", "2026-06-24", ""),
    ("Enterprise Value", f"${EV}B", "yahoo.com/key-statistics", "2026-06-24", ""),
    ("Shares (M)", SHARES_M, "MC/price recon", "2026-06-24", "Basic ~2391M; ~2464M implied"),
    ("Rev TTM", f"${REV['TTM']}M", "yahoo.com/financials", "2026-06-24", ""),
    ("Rev FY25", f"${REV['2025']}M", "yahoo.com/financials", "2026-06-24", ""),
    ("Rev FY24", f"${REV['2024']}M", "yahoo.com/financials", "2026-06-24", ""),
    ("GP TTM", f"${GP['TTM']}M", "yahoo.com/financials", "2026-06-24", ""),
    ("OP TTM", f"${OP['TTM']}M", "yahoo.com/financials", "2026-06-24", ""),
    ("NI TTM", f"${NI['TTM']}M", "yahoo.com/financials", "2026-06-24", ""),
    ("Diluted EPS TTM", f"${EPS['TTM']}", "yahoo.com/financials", "2026-06-24", ""),
    ("OCF TTM", f"${OCF['TTM']}M", "yahoo.com/cash-flow", "2026-06-24", ""),
    ("Capex TTM", f"${CPX['TTM']}M", "yahoo.com/cash-flow", "2026-06-24", ""),
    ("FCF TTM", f"${FCF['TTM']}M", "yahoo.com/cash-flow", "2026-06-24", ""),
    ("EBITDA TTM", f"${EBD['TTM']}M", "yahoo.com/financials", "2026-06-24", ""),
    ("Cash MRQ", f"${CASH}B", "yahoo.com/key-statistics", "2026-06-24", "Q1 FY26"),
    ("Debt MRQ", f"${DEBT}B", "yahoo.com/balance-sheet", "2026-06-24", "Leases"),
    ("Beta", str(BETA), "yahoo.com/key-statistics", "2026-06-24", ""),
    ("P/E TTM", str(PE_TTM), "yahoo.com/key-statistics", "2026-06-24", ""),
    ("Fwd P/E", str(PE_FWD), "yahoo.com/key-statistics", "2026-06-24", ""),
    ("10Y Treasury", f"{RISK_FREE*100:.3f}%", "cnbc.com/quotes/US10Y", "2026-06-24", ""),
    ("Analyst Est.", "See analysis", "yahoo.com/analysis", "2026-06-24", "30 analysts"),
    ("Earn Date", "Aug 3, 2026", "yahoo.com/quote/PLTR", "2026-06-24", "Q2 FY26"),
]
for i, (dp, val, src, dt, note) in enumerate(audit, 3):
    for j, v in enumerate([dp, val, src, dt, note], 1):
        c(w4, i, j, v, border=BD)

# ── Sheet 5: Questions ───────────────────────────────────────────────────
w5 = wb.create_sheet("Questions")
title(w5, "Open Questions")

questions = [
    ("1", "SBC Dilution", "Diluted shares ~2570M vs basic ~2391M. Annual SBC ~$800M+. Per-share drag is structural."),
    ("2", "International Revenue", "U.S. +104% Q1'26 vs intl lag. What % non-U.S.? Regulatory risk (UK, Germany, etc.)?"),
    ("3", "Gov vs Commercial Mix", "Gotham orig product. What % government? Termination risk? AIP diversification?"),
    ("4", "Valuation Sustainability", "131x P/E, 57x P/S. What 5Y growth sustains this? Multiple compression risk."),
    ("5", "Interest Income Dep", "~$245M/yr interest income = 10-15% of pre-tax NI. Rate-cut sensitive."),
    ("6", "AIP Land-and-Expand", "U.S. commercial +133% Q1'26. One-time migration or sustainable platform?"),
    ("7", "Tax Rate Norm", "TTM eff rate ~1.3%. Normalization to 15-21% could compress NI materially."),
    ("8", "Drawdown from 52W High", "$207.52 \u2192 $112.25 = 46% drop. Multiple compression or growth concern?"),
    ("9", "Competitive AI Landscape", "MSFT, Databricks, Snowflake compete. Genuine edge or table-stakes?"),
    ("10", "Next Earnings Aug 3", "Q2 consensus: rev ~$1.81B (+80%), EPS ~$0.35. Guidance critical."),
]
hdr_row(w5, 2, ["#", "Question", "Detail"])
for i, (num, q, d) in enumerate(questions, 3):
    c(w5, i, 1, num, border=BD)
    c(w5, i, 2, q, font=BF, border=BD)
    c(w5, i, 3, d, border=BD)
    w5.cell(row=i, column=3).alignment = Alignment(wrap_text=True)

w5.column_dimensions["C"].width = 80

# ── Sheet 6: Sources ─────────────────────────────────────────────────────
w6 = wb.create_sheet("Sources")
title(w6, "Data Sources")

hdr_row(w6, 2, ["#", "Source", "URL"])
srcs = [
    ("1", "Yahoo Finance \u2014 Quote", "finance.yahoo.com/quote/PLTR"),
    ("2", "Yahoo Finance \u2014 Income Stmt", "finance.yahoo.com/quote/PLTR/financials"),
    ("3", "Yahoo Finance \u2014 Balance Sheet", "finance.yahoo.com/quote/PLTR/balance-sheet"),
    ("4", "Yahoo Finance \u2014 Cash Flow", "finance.yahoo.com/quote/PLTR/cash-flow"),
    ("5", "Yahoo Finance \u2014 Statistics", "finance.yahoo.com/quote/PLTR/key-statistics"),
    ("6", "Yahoo Finance \u2014 Analyst Estimates", "finance.yahoo.com/quote/PLTR/analysis"),
    ("7", "CNBC \u2014 10Y Treasury", "cnbc.com/quotes/US10Y"),
    ("8", "StockAnalysis (404)", "stockanalysis.com/quote/PLTR"),
    ("9", "Prior PLTR research (internal)", "~/dev/capital/models/_build_pltr_tab.py"),
]
for i, (nm, des, url) in enumerate(srcs, 3):
    c(w6, i, 1, nm, border=BD)
    c(w6, i, 2, des, border=BD)
    c(w6, i, 3, url, border=BD)

w6.column_dimensions["B"].width = 40
w6.column_dimensions["C"].width = 50

# ── Save ──────────────────────────────────────────────────────────────────
wb.save(str(OUT))
print(f"Saved {OUT}")
print(f"Sheets: {wb.sheetnames}")
print(f"WACC = {WACC*100:.2f}%, Weighted FV = ${tot_wv} ({tot_up}% upside)")
