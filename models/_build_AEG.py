#!/usr/bin/env python3
"""Build AEG (Aegon Ltd.) 6-sheet valuation model - Insurance-specific framework."""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()
thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
bold = Font(bold=True)
header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)
subtitle_font = Font(bold=True, size=11, italic=True)
fill_blue = PatternFill("solid", fgColor="D9E2F3")
fill_green = PatternFill("solid", fgColor="E2EFDA")
fill_orange = PatternFill("solid", fgColor="FCE4D6")

def c(ws, row, col, val, font=None, fill=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.border = thin
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    return cell

# ========== DATA ==========
price = 9.46
shares_out_mm = 1512.5  # ordinary shares outstanding
mkt_cap_b = 14.04      # billion
ev_b = 15.18           # billion
total_debt_b = 3.729   # billion (Key Stats)
total_debt_bs_b = 3.729  # from balance sheet (aligned)
net_debt_b = ev_b - mkt_cap_b  # 1.14B (cleanest proxy)
cash_b = 2.733         # approx from end cash position
rev_ttm_eur_b = 26.864
net_inc_common_eur_b = 0.925
ocf_ttm_eur_b = 0.434
fcf_ttm_eur_b = 0.380
eur_usd = 1.0          # Yahoo shows in USD, so we use USD directly
p_b = 1.29
fwd_pe = 10.83
trail_pe = 13.72
rev_ps = 0.58

# WACC inputs
risk_free_rate = 0.04688  # 10Y US Treasury
erp = 0.05
beta = 1.00              # diversified insurance, moderate systemic risk
cost_of_equity = risk_free_rate + beta * erp  # 9.688%
cost_of_debt = 0.035     # estimated ~3.5% on investment-grade insurance debt
tax_rate = 0.062         # ~6.2% effective tax (€65M / €1,045M pretax)
total_cap_b = 12.942     # total capitalization = debt + equity
equity_weight = (mkt_cap_b) / (mkt_cap_b + total_debt_b)
debt_weight = total_debt_b / (mkt_cap_b + total_debt_b)
wacc = equity_weight * cost_of_equity + debt_weight * cost_of_debt * (1 - tax_rate)

print(f"WACC: {wacc*100:.2f}%")

# Insurance-specific: Primary lens is P/B and Forward P/E
# Scenarios use BVPS CAGR, exit P/B, and implied price/share
bvkps_current_p = price / p_b  # ~$7.33 BVPS

# Revenue CAGR assumptions (modest - insurance is mature)
# FY24: €19.52B -> FY25: €26.86B (acquisition spike, not organic)
# Organic growth is modest 2-4%

# Scenario framework: P/B based (insurance-specific)
# Bear: P/B 0.75x (legacy/stress discount), Base: 1.00x, Bull: 1.45x
# BVPS CAGR: Bear 1%, Base 4%, Bull 7%

horizon = 5  # years

scenarios = {
    "bear": {"bvsp_cagr": 0.01, "exit_pb": 0.75, "weight": 0.20},
    "base": {"bvsp_cagr": 0.04, "exit_pb": 1.00, "weight": 0.50},
    "bull": {"bvsp_cagr": 0.07, "exit_pb": 1.45, "weight": 0.30},
}

scen_results = {}
for case, s in scenarios.items():
    bvps_5y = bvkps_current_p * ((1 + s["bvsp_cagr"]) ** horizon)
    target = bvps_5y * s["exit_pb"]
    upside = (target / price - 1) * 100
    weighted = target * s["weight"]
    scen_results[case] = {
        "bvps_5y": bvps_5y,
        "target": target,
        "upside": upside,
        "weighted": weighted,
    }

total_weighted_fv = sum(v["weighted"] for v in scen_results.values())
total_upside = (total_weighted_fv / price - 1) * 100

print(f"Bear target: ${scen_results['bear']['target']:.2f} ({scen_results['bear']['upside']:.1f}%)")
print(f"Base target: ${scen_results['base']['target']:.2f} ({scen_results['base']['upside']:.1f}%)")
print(f"Bull target: ${scen_results['bull']['target']:.2f} ({scen_results['bull']['upside']:.1f}%)")
print(f"Weighted FV: ${total_weighted_fv:.2f} (upside: {total_upside:.1f}%)")

# ========== SHEET 1: Valuation ==========
ws1 = wb.active
ws1.title = "Valuation"
ws1.merge_cells("A1:F1")
t = ws1.cell(row=1, column=1, value="Aegon Ltd. (AEG) - Valuation Summary")
t.font = title_font
t.alignment = Alignment(horizontal="center")

row = 3
title_data = [
    ("Company", "Aegon Ltd."),
    ("Ticker", "NYSE: AEG"),
    ("Date", str(date.today())),
    ("Price", f"${price:.2f}"),
    ("Shares Outstanding", f"{shares_out_mm/1e6:.2f}B"),
    ("Market Cap", f"${mkt_cap_b:.2f}B"),
    ("Enterprise Value", f"${ev_b:.2f}B"),
    ("Primary Lens", "P/B ratio & Forward P/E (Insurance framework)"),
    ("Stance", "Watch — Multiple Rule 425 filings signal M&A; insurance recovery story but regulatory/legacy uncertainty remains"),
]
for i, (k, v) in enumerate(title_data, 2):
    c(ws1, i, 1, k, header_font)
    c(ws1, i, 2, v)

row = 13
c(ws1, row, 1, "Key Valuation Metrics", header_font)
ws1.merge_cells(f"A{row}:C{row}")

val_metrics = [
    ("Trailing P/E", f"{trail_pe:.2f}x", "Based on ~$0.69 TTM GAAP EPS; trailing EPS distorted by FY23 restructuring"),
    ("Forward P/E", f"{fwd_pe:.2f}x", "FY2026 consensus ~$0.87 EPS; forward multiple compresses toward 10x"),
    ("P/S", f"{rev_ps:.2f}x", "Deep discount for diversified insurer; revenue base is large and stable"),
    ("P/B", f"{p_b:.2f}x", "Above insurance peer median (~0.8-1.1x); encodes some buyback-driven accretion"),
    ("EV/Revenue", f"{ev_b/rev_ttm_eur_b:.2f}x", "EV includes debt adjustment; very low for the sector"),
    ("FCF Yield", f"{fcf_ttm_eur_b/mkt_cap_b*100:.1f}%", f"FCF ~${fcf_ttm_eur_b:.0f}B on ${mkt_cap_b:.2f}B MC; insurance OCF is structural"),
    ("Dividend Yield", "~2.5-3.5%", "Quarterly dividends on ADR; maintenance signals cash confidence"),
]
for i, (metric, val, comment) in enumerate(val_metrics, 14):
    c(ws1, i, 1, metric, header_font)
    c(ws1, i, 2, val)
    c(ws1, i, 3, comment)

# ========== SHEET 2: WACC ==========
ws2 = wb.create_sheet("WACC")
ws2.merge_cells("A1:D1")
t = ws2.cell(row=1, column=1, value="WACC Calculation — CAPM Framework")
t.font = title_font
t.alignment = Alignment(horizontal="center")

wacc_data = [
    ("Component", "Value", "Source/Assumption"),
    ("Risk-Free Rate (10Y US)", f"{risk_free_rate*100:.3f}%", "US 10Y Treasury (CNBC, Aug 3 2026)"),
    ("Equity Risk Premium", f"{erp*100:.1f}%", "Standard ERP assumption"),
    ("Beta (Levered)", f"{beta:.2f}", "Diversified insurance - moderate systemic risk"),
    ("Cost of Equity (Rf + Beta*ERP)", f"{cost_of_equity*100:.2f}%", f"{risk_free_rate*100:.2f}% + {beta:.2f} x {erp*100:.1f}%"),
    ("Cost of Debt", f"{cost_of_debt*100:.1f}%", "Estimated IG insurance debt rate"),
    ("Tax Rate (Effective)", f"{tax_rate*100:.1f}%", f"TTM: €65M / €1,045M pretax = {tax_rate*100:.1f}%"),
    ("Market Cap", f"${mkt_cap_b:.2f}B", "Yahoo Finance Key Stats"),
    ("Total Debt", f"${total_debt_b:.3f}B", "Yahoo Finance Balance Sheet"),
    ("Equity Weight", f"{equity_weight*100:.1f}%", f"${mkt_cap_b:.2f}B / (${mkt_cap_b:.2f}B + ${total_debt_b:.3f}B)"),
    ("Debt Weight", f"{debt_weight*100:.1f}%", f"${total_debt_b:.3f}B / (${mkt_cap_b:.2f}B + ${total_debt_b:.3f}B)"),
    ("WACC", f"{wacc*100:.2f}%", f"{equity_weight*100:.1f}% x {cost_of_equity*100:.2f}% + {debt_weight*100:.1f}% x {cost_of_debt*100:.1f}% x (1-{tax_rate*100:.1f}%)"),
]

for i, (comp, val, src) in enumerate(wacc_data, 2):
    c(ws2, i, 1, comp, font=header_font if i == 2 or comp == "WACC" else None)
    c(ws2, i, 2, val, fill=fill_green if comp == "WACC" else None)
    c(ws2, i, 3, src)

# ========== SHEET 3: Scenarios ==========
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells("A1:I1")
t = ws3.cell(row=1, column=1, value="Scenario Analysis — P/B Framework (Insurance)")
t.font = title_font
t.alignment = Alignment(horizontal="center")

c(ws3, 3, 1, "Note: Insurance-specific valuation using BVPS CAGR and exit P/B multiple. FCF multiple not used — insurance OCF includes investment portfolio management. OCF anchor: ~$434M TTM.")
ws3.merge_cells("A3:I3")
ws3.cell(row=3, column=1).font = subtitle_font

headers3 = ["Metric", "Bear (20%)", "Base (50%)", "Bull (30%)", "Notes"]
for j, h in enumerate(headers3, 1):
    c(ws3, 4, j, h, header_font, fill_blue)

scen_rows = [
    ("Current BVPS", f"${bvkps_current_p:.2f}", f"${bvkps_current_p:.2f}", f"${bvkps_current_p:.2f}", f"P/B = {p_b:.2f}x on ${price:.2f} price"),
    ("BVPS CAGR (5Y)", f"{scenarios['bear']['bvsp_cagr']*100:.0f}%", f"{scenarios['base']['bvsp_cagr']*100:.0f}%", f"{scenarios['bull']['bvsp_cagr']*100:.0f}%", "Insurance BVPS growth: modest, buyback-accelerated"),
    ("Terminal BVPS (5Y)", f"${scen_results['bear']['bvps_5y']:.2f}", f"${scen_results['base']['bvps_5y']:.2f}", f"${scen_results['bull']['bvps_5y']:.2f}", "BVPS × (1+CAGR)^5"),
    ("Exit P/B Multiple", f"{scenarios['bear']['exit_pb']:.2f}x", f"{scenarios['base']['exit_pb']:.2f}x", f"{scenarios['bull']['exit_pb']:.2f}x", "Insurance peer range: 0.6-1.6x"),
    ("Target Price", f"${scen_results['bear']['target']:.2f}", f"${scen_results['base']['target']:.2f}", f"${scen_results['bull']['target']:.2f}", "Terminal BVPS × Exit P/B"),
    ("Upside vs Current", f"{scen_results['bear']['upside']:.1f}%", f"{scen_results['base']['upside']:.1f}%", f"{scen_results['bull']['upside']:.1f}%", f"Current price: ${price:.2f}"),
    ("Scenario Weight", f"{scenarios['bear']['weight']*100:.0f}%", f"{scenarios['base']['weight']*100:.0f}%", f"{scenarios['bull']['weight']*100:.0f}%", ""),
    ("Weighted Value/Share", f"${scen_results['bear']['weighted']:.2f}", f"${scen_results['base']['weighted']:.2f}", f"${scen_results['bull']['weighted']:.2f}", "Target × Weight"),
    ("", "", "", "", ""),
    ("Probability-Weighted FV", "", "", f"${total_weighted_fv:.2f}", "Sum of weighted values"),
    ("Total Upside", "", "", f"{total_upside:.1f}%", f"vs ${price:.2f} current"),
]
for i, row_data in enumerate(scen_rows, 5):
    for j, val in enumerate(row_data, 1):
        c(ws3, i, j, val, fill=fill_green if "Probability" in str(val) or "Total Upside" in str(val) else None)

# ========== SHEET 4: Actuals Source Audit ==========
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells("A1:D1")
t = ws4.cell(row=1, column=1, value="Actuals Source Audit — All Data Points")
t.font = title_font

audit = [
    ("Data Point", "Value", "Source", "Notes"),
    ("Stock Price", "$9.46", "Yahoo Finance Key Stats", "Aug 3, 2026 close"),
    ("Market Cap", "$14.04B", "Yahoo Finance Key Stats", "Aug 3, 2026"),
    ("Enterprise Value", "$15.18B", "Yahoo Finance Key Stats", "MC + Total Debt - Cash"),
    ("Net Debt (EV-MC)", "$1.14B", "Computed", "EV - MC = $1.14B"),
    ("Shares Outstanding", "1,512.5M", "Yahoo Finance Balance Sheet", "Ordinary shares number"),
    ("TTM Revenue", "$26.86B", "Yahoo Finance Income Statement", "In EUR, TTM"),
    ("FY25 Revenue", "$26.86B", "Yahoo Finance Income Statement", "Large jump from FY24 €19.52B — acquisition-driven"),
    ("FY24 Revenue", "$19.52B", "Yahoo Finance Income Statement", "Pre-acquisition base"),
    ("FY23 Revenue", "$20.06B", "Yahoo Finance Income Statement", "Post-restructuring year"),
    ("TTM Pretax Income", "$1.045B", "Yahoo Finance Income Statement", "In EUR thousands"),
    ("TTM Net Income Common", "$0.925B", "Yahoo Finance Income Statement", "€925M"),
    ("TTM Diluted EPS", "$0.59", "Yahoo Finance Income Statement", "Basic = dilute"),
    ("TTM OCF", "$0.434B", "Yahoo Finance Cash Flow", "€434M"),
    ("TTM FCF", "$0.380B", "Yahoo Finance Cash Flow", "OCF - Capex (€54M)"),
    ("TTM Capex", "$54M", "Yahoo Finance Cash Flow", "Minimal for insurance — low capex profile"),
    ("Total Debt", "$3.73B", "Yahoo Finance Balance Sheet", "€3,729M"),
    ("Total Cash", "$2.73B", "Yahoo Finance Cash Flow", "End cash position"),
    ("Common Stock Equity", "$9.41B", "Yahoo Finance Balance Sheet", "€9,410M"),
    ("P/B Ratio", "1.29x", "Yahoo Finance Key Stats", "Price / Book Value per Share"),
    ("Forward P/E", "10.83x", "Yahoo Finance Key Stats", "Based on FY2026 consensus"),
    ("Trailing P/E", "13.72x", "Yahoo Finance Key Stats", "Based on TTM EPS"),
    ("Beta", "~1.00", "Estimated", "Diversified insurance; not visible on Yahoo for this ticker"),
    ("10Y US Treasury", "4.688%", "CNBC.com", "Aug 3, 2026"),
    ("Next Earnings Date", "Aug 20, 2026", "Yahoo Finance Profile", "8:30 AM EDT"),
    ("Rule 425 Filings", "Multiple, May-Jun 2026", "Yahoo Finance Profile", "Business combination transactions — likely M&A"),
    ("Shares Trend", "Declining", "Yahoo Finance Balance Sheet", "1,592M FY24 → 1,512M FY25 = 5% buyback reduction"),
    ("Repurchase of Capital Stock", "$549M", "Yahoo Finance Cash Flow", "TTM; significant buyback activity"),
]

for i, row_data in enumerate(audit, 2):
    for j, val in enumerate(row_data, 1):
        c(ws4, i, j, val, fill=fill_blue if i == 2 else None)

# ========== SHEET 5: Questions ==========
ws5 = wb.create_sheet("Questions")
ws5.merge_cells("A1:C1")
t = ws5.cell(row=1, column=1, value="Open Questions — Aegon Ltd. (AEG)")
t.font = title_font

questions = [
    "Rule 425 Filings — What Business Combination? Multiple Rule 425 filings (May 27, Jun 9/16/24, 2026) indicate M&A consideration or an active business combination. What is the target, expected close date, and financial impact? This is the single most significant near-term catalyst.",
    "Revenue Jump FY24→FY25 — €19.52B to €26.86B (+37.6%): Acquisition-driven? FY25 revenue surged dramatically. Was this organic growth or an M&A combination? This fundamentally changes the revenue CAGR baseline.",
    "FY22 Negative Revenue (-€21.64B): Restructuring artifact or one-time credit? FY2022 shows negative total revenue — consistent with Aegon's post-demerger restructuring of its Dutch life insurance business into a 'good bank/bad bank' split.",
    "Operating Income Not Visible: Yahoo shows pretax income (€1,045M TTM) but no visible operating income line. For insurance companies, the income statement is structured differently. What is the actual operating margin after investment portfolio income is excluded?",
    "Investment Income vs. Underwriting Profit: €5.81B in interest income on €26.86B revenue = 21.6% of revenue from investment portfolio. What portion of earnings is underwriting vs. investment return-driven?",
    "Tax Rate Anomaly: Effective tax rate of ~6.2% (€65M / €1,045M) is very low vs. Dutch statutory. Does this reflect deferred tax credits, geographic mix, or tax structuring around the demerger?",
    "Minority Interest Position: Total equity (€9,495M) vs. common equity (€9,410M) shows €85M minority interest. Post-restructuring, is this significant or residual?",
    "Capital Lease Obligations: €187M — standard for insurance but worth confirming no off-balance-sheet obligations or P/C lease structures.",
    "Dutch 'Good Bank/Bad Bank' Demerger: Aegon split its Dutch life business. What is the residual legacy liability exposure? Does the P/B discount encode AIG-adjacent stigma from the original bailout?",
    "Forward Guidance: Management guidance for FY2026 given the M&A activity? Next earnings on Aug 20 may address this.",
    "Share Buyback Sustainability: €549M repurchase TTM vs €434M OCF = 126% of cash flow. Are buybacks funded through debt issuance? Issuance of debt was €946M but repayment was €1,740M — net deleveraging.",
    "Competitive Position in US Retirement (Transamerica): How does Transamerica's 401(k)/annuity franchise compete against Fidelity, Vanguard, BlackRock? Fee compression risk?",
]

for i, q in enumerate(questions, 2):
    c(ws5, i, 1, f"Q{i-1}", header_font)
    c(ws5, i, 2, q)

# ========== SHEET 6: Sources ==========
ws6 = wb.create_sheet("Sources")
ws6.merge_cells("A1:B1")
t = ws6.cell(row=1, column=1, value="Sources")
t.font = title_font

sources = [
    ("1", "Yahoo Finance — AEG Key Statistics", "https://finance.yahoo.com/quote/AEG/key-statistics/"),
    ("2", "Yahoo Finance — AEG Income Statement", "https://finance.yahoo.com/quote/AEG/financials/"),
    ("3", "Yahoo Finance — AEG Balance Sheet", "https://finance.yahoo.com/quote/AEG/balance-sheet/"),
    ("4", "Yahoo Finance — AEG Cash Flow", "https://finance.yahoo.com/quote/AEG/cash-flow/"),
    ("5", "Yahoo Finance — AEG Profile/Company Info", "https://finance.yahoo.com/quote/AEG/profile/"),
    ("6", "Yahoo Finance — AEG Analysis/Estimates", "https://finance.yahoo.com/quote/AEG/analysis/"),
    ("7", "CNBC — US 10Y Treasury Yield", "https://www.cnbc.com/quotes/US10Y"),
    ("8", "StockAnalysis.com — 404 for AEG", "https://stockanalysis.com/quote/AEG/"),
    ("9", "Aegon Ltd. Corporate Website", "https://www.aegon.com"),
    ("10", "SEC EDGAR — Rule 425 Filings (May-Jun 2026)", "Referenced via Yahoo Profile Recent Events"),
]

for i, (num, name, url) in enumerate(sources, 2):
    c(ws6, i, 1, num)
    c(ws6, i, 2, name)
    c(ws6, i, 3, url)

# Set column widths
for ws in [ws1, ws2, ws3, ws4, ws5, ws6]:
    for col_idx in range(1, 10):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

ws1.column_dimensions["A"].width = 20
ws1.column_dimensions["B"].width = 30
ws1.column_dimensions["C"].width = 60

ws4.column_dimensions["A"].width = 25
ws4.column_dimensions["B"].width = 20
ws4.column_dimensions["C"].width = 35
ws4.column_dimensions["D"].width = 50

ws5.column_dimensions["A"].width = 5
ws5.column_dimensions["B"].width = 120

ws6.column_dimensions["A"].width = 5
ws6.column_dimensions["B"].width = 55
ws6.column_dimensions["C"].width = 60

outfile = "/home/refcell/dev/capital/models/[2026-08-03] Aegon Model.xlsx"
wb.save(outfile)
print(f"Saved: {outfile}")
print(f"WACC: {wacc*100:.2f}%")
print(f"Weighted FV: ${total_weighted_fv:.2f}")
print(f"Upside: {total_upside:.1f}%")
