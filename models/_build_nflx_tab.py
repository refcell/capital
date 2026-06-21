"""Add a NFLX (Netflix) tab to projections/Projections.xlsx.

Clones the last existing tab (HOOD) to preserve styling/formulas, then
overwrites the input cells with Netflix numbers.

Netflix is highly GAAP-profitable, so this tab uses a straightforward GAAP net
income basis.  All share/EPS figures are split-adjusted for the 10-for-1 forward
stock split effected November 14, 2025.  All $ figures in millions.

Sources (Q4 2025 shareholder letter Jan 2026; Q1 2026 results Apr 2026):
  - FY2025 revenue $45,183M (+16% Y/Y); GAAP net income $10,981M (~24% margin);
    diluted EPS $2.53; full-year operating margin ~29.5%
  - FY2024 revenue $39,001M; net income $8,712M; diluted EPS $1.98 (split-adj)
  - 2026 guidance: revenue $50.7B-$51.7B (+12%-14%), operating margin 31.5%,
    ad revenue roughly doubling to ~$3B; FCF guide ~$12.5B
  - Diluted shares ~4,317M (Q4'25 FD), declining on buybacks ($31.8B auth) -> 4,300M
  - Analyst consensus: ~12-15% EPS growth; FY2026 EPS ~$3.10 rising to ~$4.70 by 2029
  - Price ~$80 (Jun 2026, post-split); forward P/E ~26x (vs historical 30-45x)
"""

from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "projections" / "Projections.xlsx"

# Shared inputs
SHARES = 4300         # diluted shares (M), split-adjusted; declining on buybacks
REV_2026 = 51200      # FY2026E revenue (guidance midpoint, ~13% growth)
NI_2026 = 13300       # FY2026E GAAP net income (~26% margin, EPS ~$3.09)
NI_G_2026 = 0.21      # net income growth 2025->2026 ($10,981M -> $13,300M)

# Analyst GAAP net income estimates 2026-2029 (EPS ~$3.10/$3.55/$4.10/$4.70 x 4300).
ANALYST = {"C": 13300, "D": 15300, "E": 17600, "F": 20200}

# Per-case assumptions.  Each template block is offset by 24 rows:
#   Bull base row 4, Base 28, Bear 52.  Within a block, relative rows:
#     +6 revenue, +7 revenue growth, +9 net income, +10 NI growth,
#     +12 analyst est, +14 NI margins, +18 PE low, +19 PE high, H(+0) shares.
# Revenue growth decelerates from the low-teens 2026 guide; net margins expand on
# operating leverage + scaling high-margin ad revenue (bear: content/competition pressure).
CASES = {
    "bull": {
        "base": 4,
        # ads + pricing + members + games sustain low-teens growth
        "rev_g": {"C": 0.13, "D": 0.14, "E": 0.14, "F": 0.13, "G": 0.12, "H": 0.12},
        "margins": {"D": 0.27, "E": 0.28, "F": 0.30, "G": 0.31, "H": 0.32},
        "pe_low": {c: 30 for c in "CDEFGH"},
        "pe_high": {c: 38 for c in "CDEFGH"},
    },
    "base": {
        "base": 28,
        # growth decelerates toward high-single digits as base scales
        "rev_g": {"C": 0.13, "D": 0.12, "E": 0.11, "F": 0.10, "G": 0.10, "H": 0.09},
        "margins": {"D": 0.27, "E": 0.27, "F": 0.28, "G": 0.28, "H": 0.29},
        "pe_low": {c: 24 for c in "CDEFGH"},
        "pe_high": {c: 30 for c in "CDEFGH"},
    },
    "bear": {
        "base": 52,
        # member growth slows, content costs/competition cap margin expansion
        "rev_g": {"C": 0.13, "D": 0.09, "E": 0.08, "F": 0.07, "G": 0.07, "H": 0.06},
        "margins": {"D": 0.25, "E": 0.25, "F": 0.26, "G": 0.26, "H": 0.26},
        "pe_low": {c: 18 for c in "CDEFGH"},
        "pe_high": {c: 24 for c in "CDEFGH"},
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    src = wb[wb.sheetnames[-1]]          # clone the last tab (HOOD)
    ws = wb.copy_worksheet(src)
    ws.title = "NFLX"

    # Header
    ws["B2"] = "NFLX"
    ws["C2"] = datetime(2026, 6, 21)

    for cfg in CASES.values():
        b = cfg["base"]
        ws[f"H{b}"] = SHARES                      # share count

        # Revenue base (2026E) + growth row
        ws[f"C{b + 2}"] = REV_2026
        for col, g in cfg["rev_g"].items():
            ws[f"{col}{b + 3}"] = g

        # GAAP net income base (2026E) + base-year growth
        ws[f"C{b + 5}"] = NI_2026
        ws[f"C{b + 6}"] = NI_G_2026

        # Analyst estimates (2026-2029)
        for col, v in ANALYST.items():
            ws[f"{col}{b + 8}"] = v

        # NI margins (D..H; C stays as formula =NI/Rev)
        for col, m in cfg["margins"].items():
            ws[f"{col}{b + 10}"] = m

        # PE low / high
        for col, v in cfg["pe_low"].items():
            ws[f"{col}{b + 14}"] = v
        for col, v in cfg["pe_high"].items():
            ws[f"{col}{b + 15}"] = v

    wb.save(XLSX)
    print(f"Added NFLX tab. Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
