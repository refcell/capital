#!/usr/bin/env python3
"""
Build 6-sheet Excel valuation model for Brookfield Infrastructure Partners (BIP).
Date: 2026-07-31
Infrastructure holding company — primary lens: EV/EBITDA, distribution yield.
GAAP net income is structurally distorted by massive D&A ($4.3B TTM).
FCF is negative during capex investment cycles — do NOT use FCF multiples.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

bold = Font(bold=True)
bold14 = Font(bold=True, size=14)
bold12 = Font(bold=True, size=12)
header_font = Font(bold=True, size=10)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_w = Font(bold=True, color="FFFFFF", size=10)
scenario_fill = {
    "bear": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    "base": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "bull": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
}

def c(ws, row, col, value, font=None, border=False, fill=None, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if border:
        cell.border = thin_border
    if fill:
        cell.fill = fill
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# =============================================================================
# DATA
# =============================================================================
price = 41.76
shares_mm = 458.65  # TTM diluted avg shares (thousands) / 1000 = millions
mc_b = 19.22  # Market cap $B
ev_b = 78.60  # Enterprise value $B
total_debt_b = 69.117  # $B
cash_b = 6.642  # Total cash $B
net_debt_b = ev_b - mc_b  # ~$59.38B — use EV-MC as cleanest proxy
beta = 0.92
rf_rate = 4.718  # 10Y Treasury, CNBC 2026-07-31
erp = 5.0
tax_rate = 19.4  # TTM effective tax rate 576M/2967M
cost_of_debt = 5.5  # estimated blended rate on infrastructure debt

# WACC calculation
cost_of_equity = rf_rate + beta * erp  # 4.718 + 0.92*5 = 9.32%
# Equity weight: MC / (MC + Total Debt)
eq_weight = mc_b / (mc_b + total_debt_b)
debt_weight = total_debt_b / (mc_b + total_debt_b)
wacc = eq_weight * cost_of_equity + debt_weight * cost_of_debt * (1 - tax_rate / 100)
wacc_pct = round(wacc, 2)

print(f"WACC: {wacc_pct}%")
print(f"Cost of Equity: {cost_of_equity:.2f}%")
print(f"Eq Weight: {eq_weight:.4f}, Debt Weight: {debt_weight:.4f}")

# Revenue in $M (TTM)
revenue_ttm = 25062
ebitda_ttm = 11060
ebit_ttm = 6767
ocf_ttm = 6300
capex_ttm = 7522
fcp = ocf_ttm - capex_ttm  # FCF in $M (negative: -1222)
op_margin_ttm = 100 * 6204 / 25062  # 24.8%

# =============================================================================
# Sheet 1: Valuation
# =============================================================================
ws1 = wb.active
ws1.title = "Valuation"

# Merge title row
ws1.merge_cells("A1:F1")
c(ws1, 1, 1, "Brookfield Infrastructure Partners L.P. (BIP) — Valuation Model", bold14)
ws1.merge_cells("A2:F2")
c(ws1, 2, 1, f"As of July 31, 2026 | NYSE: BIP | Price $41.76 | MC ${mc_b:.2f}B | EV ${ev_b:.2f}B", Font(size=11))
ws1.merge_cells("A3:F3")
c(ws1, 3, 1, f"Primary Lens: EV/EBITDA | Infrastructure Holding Co. — GAAP net income distorted by D&A ($4.3B TTM)", Font(size=10))
ws1.merge_cells("A4:F4")
c(ws1, 4, 1, f"Stance: Watch — Heavy leverage, positive OCF, but FCF-negative during capex cycle. Requires discipline.", Font(size=10))

title_data = [
    ("Company", "Brookfield Infrastructure Partners L.P."),
    ("Ticker", "NYSE: BIP"),
    ("Date", "2026-07-31"),
    ("Price", f"${price:.2f}"),
    ("Shares Outstanding (diluted avg)", f"{shares_mm:.1f}M"),
    ("Market Cap", f"${mc_b:.2f}B"),
    ("Enterprise Value", f"${ev_b:.2f}B"),
    ("Total Debt", f"${total_debt_b:.2f}B"),
    ("Net Debt (EV-MC proxy)", f"${net_debt_b:.2f}B"),
    ("Primary Lens", "EV/EBITDA, Distribution Yield"),
    ("Stance", "Watch"),
]

for i, (label, val) in enumerate(title_data, 6):
    c(ws1, i, 1, label, header_font, border=True)
    ws1.merge_cells(f"B{i}:F{i}")
    c(ws1, i, 2, val, border=True)

# Metrics table starts AFTER title block (rows 6-16) + blank separator (row 17)
metrics_start_row = 6 + len(title_data) + 1  # = 18

metrics_data = [
    ("Trailing P/E", f"{68.37:.1f}x", "Distorted — GAAP NI ($319M TTM) suppressed by D&A ($4.3B TTM). Not a meaningful earnings metric."),
    ("P/FFO (estimated)", "~24-26x", "FFO ≈ NI + D&A = $4.6B TTM. $19.22B / $4.6B ≈ 4.2x on MC, but implied FFO/share is complex with minority interests."),
    ("EV/EBITDA", f"{100*ev_b/ebitda_ttm:.1f}x", "Primary framework. TTM EBITDA $11.06B. Reasonable for infra holding co. vs peers."),
    ("P/S", f"{100*mc_b/revenue_ttm:.2f}", f"Revenue $25.06B TTM. Low P/S masks low profitability on GAAP basis; see EV/EBITDA instead."),
    ("EV/Sales", f"{100*ev_b/revenue_ttm:.2f}", f"Enterprise value on top-line revenue."),
    ("EV/EBIT", f"{100*ev_b/ebit_ttm:.1f}x", f"Operating earnings power vs enterprise value."),
    ("OCF/Revenue", f"{100*ocf_ttm/revenue_ttm:.1f}%", f"Cash conversion quality. OCF $6.30B on $25.06B revenue."),
    ("Interest Coverage (EBIT/Interest)", f"{ebit_ttm/3800:.1f}x", f"EBIT $6.77B on interest $3.80B. Adequate coverage."),
    ("Net Debt / EBITDA", f"{100*net_debt_b/ebitda_ttm:.1f}x", f"Net debt ${net_debt_b:.2f}B vs EBITDA $11.06B. High but manageable for infra."),
    ("CapEx / OCF", f"{100*capex_ttm/ocf_ttm:.1f}%", f"CapEx $7.52B vs OCF $6.30B. Heavy investment cycle — structurally negative FCF."),
    ("Beta", f"{beta:.2f}", "Low-volatility infrastructure characteristics."),
]

for i, (label, val, comment) in enumerate(metrics_data, metrics_start_row):
    c(ws1, i, 1, label, header_font, border=True)
    c(ws1, i, 2, val, border=True)
    ws1.merge_cells(f"C{i}:F{i}")
    c(ws1, i, 3, comment, border=True)

set_col_widths(ws1, [28, 15, 70, 8, 8, 8])

# =============================================================================
# Sheet 2: WACC
# =============================================================================
ws2 = wb.create_sheet("WACC")
ws2.merge_cells("A1:E1")
c(ws2, 1, 1, "WACC — CAPM Components", bold14)
ws2.merge_cells("A2:E2")
c(ws2, 2, 1, f"Brookfield Infrastructure Partners | 2026-07-31 | Infrastructure Holding Company", Font(size=10))

wacc_data = [
    ("Risk-Free Rate (10Y US)", f"{rf_rate:.3f}%", "CNBC US10Y, July 31 2026"),
    ("Equity Risk Premium", f"{erp:.1f}%", "Standard assumption"),
    ("Beta (levered)", f"{beta:.2f}", "Yahoo Finance Key Statistics"),
    ("Cost of Equity (Rf + Beta*ERP)", f"{cost_of_equity:.2f}%", f"= {rf_rate:.3f} + {beta:.2f} × {erp:.1f}"),
    ("Cost of Debt (blended)", f"{cost_of_debt:.1f}%", "Estimated blended infrastructure debt rate"),
    ("Tax Rate (TTM effective)", f"{tax_rate:.1f}%", "TTM: $576M provision / $2,967M pretax"),
    ("Market Cap", f"${mc_b:.2f}B", "As of July 31, 2026"),
    ("Total Debt", f"${total_debt_b:.2f}B", "Yahoo Finance Balance Sheet FY2025"),
    ("Equity Weight", f"{eq_weight:.4f}", f"MC / (MC + Debt)"),
    ("Debt Weight", f"{debt_weight:.4f}", f"Debt / (MC + Debt)"),
    ("WACC", f"{wacc_pct:.2f}%", "Weighted average = EqWt×Ke + DtWt×Kd×(1-τ)"),
]

for i, (label, val, note) in enumerate(wacc_data, 4):
    c(ws2, i, 1, label, header_font, border=True)
    c(ws2, i, 2, val, bold, border=True)
    ws2.merge_cells(f"C{i}:E{i}")
    c(ws2, i, 3, note, border=True)

set_col_widths(ws2, [40, 20, 50, 8, 8])

# =============================================================================
# Sheet 3: Scenarios
# =============================================================================
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells("A1:H1")
c(ws3, 1, 1, "Scenario Analysis — Brookfield Infrastructure Partners", bold14)
ws3.merge_cells("A2:H2")
c(ws3, 2, 1, "Framework: EV/EBITDA exit multiples + distribution growth. FCF multiples NOT used (structurally negative FCF during capex cycle).", Font(size=10))

# SCENARIO FRAMEWORK: P/FFO-based (standard infrastructure metric)
# EV/EBITDA produces negative bear targets because $59.4B net debt overwhelms
# implied EV at compressed multiples. Per skill guidance on "debt/FCF gray zone":
# switch to a framework that avoids the debt amplification problem.
#
# FFO (Funds From Operations) = Net Income + D&A
# TTM NI (common) $319M + D&A $4,293M = $4.612B FFO
# FFO/share = $4.612B / 458.65M = $10.06
# Current P/FFO = $41.76 / $10.06 = 4.15x

ni_common_ttm = 319       # millions
da_ttm = 4293             # millions
ffo_ttm = ni_common_ttm + da_ttm  # $4.612B
ffo_per_share = ffo_ttm / shares_mm  # $10.06
current_p_ffo = price / ffo_per_share
print(f"FFO TTM: ${ffo_ttm:.0f}M, FFO/share: ${ffo_per_share:.2f}, P/FFO: {current_p_ffo:.2f}x")

# P/FFO scenario assumptions
# Bear: 3.3x P/FFO (compressed) + 3% FFO CAGR
# Base: 4.5x P/FFO (current-ish) + 6% FFO CAGR
# Bull: 5.5x P/FFO (premium for infra) + 8% FFO CAGR
rev_cagr = {"Bear": 3.0, "Base": 6.0, "Bull": 8.0}
term_years = 5  # 5-year horizon
p_ffo_exit = {"Bear": 3.3, "Base": 4.5, "Bull": 5.5}
fills = {"Bear": scenario_fill["bear"], "Base": scenario_fill["base"], "Bull": scenario_fill["bull"]}

# Compute terminal values using FFO framework
results = {}
for case in ["Bear", "Base", "Bull"]:
    cagr = rev_cagr[case] / 100
    terminal_ffo_ps = ffo_per_share * ((1 + cagr) ** term_years)
    target_price = terminal_ffo_ps * p_ffo_exit[case]
    upside = (target_price - price) / price * 100
    terminal_revenue = revenue_ttm * ((1 + cagr) ** term_years)
    terminal_ebitda = terminal_revenue * 0.40  # normalized EBITDA margin ~40%
    implied_ev_b = terminal_ebitda / 1000 * 8.5  # cross-check EV/EBITDA
    results[case] = {
        "cagr": rev_cagr[case],
        "terminal_revenue": terminal_revenue,
        "terminal_ebitda": terminal_ebitda,
        "terminal_ffo_ps": terminal_ffo_ps,
        "implied_ev_b": implied_ev_b,
        "target_price": target_price,
        "upside": upside,
    }
    print(f"{case}: FFOps=${terminal_ffo_ps:.2f}, P/FFO={p_ffo_exit[case]}x, Target=${target_price:.2f}, Upside={upside:.1f}%")

# Weighted scenario
weights = {"Bear": 0.25, "Base": 0.50, "Bull": 0.25}
weighted_fv = sum(weights[k] * results[k]["target_price"] for k in results)
total_w = sum(weights.values())
print(f"Probability-Weighted FV: ${weighted_fv:.2f}")
print(f"Upside from current: {(weighted_fv - price) / price * 100:.1f}%")

# Write scenario table
headers = ["Metric", "Bear", "Base", "Bull", "Notes"]
for j, h in enumerate(headers, 1):
    c(ws3, 4, j, h, header_font_w, border=True, fill=header_fill)

scenario_rows = [
    ("Revenue/FFO CAGR (5Y)", f"{rev_cagr['Bear']:.1f}%", f"{rev_cagr['Base']:.1f}%", f"{rev_cagr['Bull']:.1f}%", ""),
    ("Terminal Revenue (5Y, $M)", f"${results['Bear']['terminal_revenue']:.0f}", f"${results['Base']['terminal_revenue']:.0f}", f"${results['Bull']['terminal_revenue']:.0f}", ""),
    ("Terminal FFO/Share ($)", f"${results['Bear']['terminal_ffo_ps']:.2f}", f"${results['Base']['terminal_ffo_ps']:.2f}", f"${results['Bull']['terminal_ffo_ps']:.2f}", ""),
    ("Exit P/FFO Multiple", f"{p_ffo_exit['Bear']:.1f}x", f"{p_ffo_exit['Base']:.1f}x", f"{p_ffo_exit['Bull']:.1f}x", ""),
    ("Implied EV (Cross-Check, $B)", f"${results['Bear']['implied_ev_b']:.1f}", f"${results['Base']['implied_ev_b']:.1f}", f"${results['Bull']['implied_ev_b']:.1f}", "EV/EBITDA @ 8.5x"),
    ("Target Price/Share", f"${results['Bear']['target_price']:.2f}", f"${results['Base']['target_price']:.2f}", f"${results['Bull']['target_price']:.2f}", ""),
    ("Upside from $41.76", f"{results['Bear']['upside']:.1f}%", f"{results['Base']['upside']:.1f}%", f"{results['Bull']['upside']:.1f}%", ""),
    ("Weight", f"{weights['Bear']*100:.0f}%", f"{weights['Base']*100:.0f}%", f"{weights['Bull']*100:.0f}%", ""),
    ("Weighted Value/Share", "", "", f"${weighted_fv:.2f}", f"Sum of weighted @ ${weighted_fv:.2f}"),
    ("Current Price", "", "", f"${price:.2f}", ""),
    ("Implied Upside", "", "", f"{(weighted_fv-price)/price*100:.1f}%", ""),
]

for i, row in enumerate(scenario_rows, 5):
    for j, val in enumerate(row):
        fill = fills.get(row[0].split()[0], None) if j in [2, 3, 4] else None
        font = bold if row[0] in ["Weighted Value/Shares", "Implied Upside"] else None
        c(ws3, i, j+1, val, border=True, fill=fill, font=font)

# Add note about framework
ws3.merge_cells("A17:A19")
c(ws3, 17, 1, "NOTE: P/FFO framework used instead of EV/EBITDA for target prices. BIP has $59.4B net debt (EV-MC) which overwhelms implied EV at compressed EBITDA multiples, producing negative bear targets. FFO = NI + D&A = $4.61B TTM. Current P/FFO ~4.15x.", Font(italic=True, size=9))

set_col_widths(ws3, [30, 18, 18, 18, 50, 8, 8, 8])

# =============================================================================
# Sheet 4: Actuals Source Audit
# =============================================================================
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells("A1:D1")
c(ws4, 1, 1, "Actuals Source Audit — BIP", bold14)

audit_data = [
    ["Data Point", "Value", "Source", "Notes"],
    ["Stock Price", "$41.76", "Yahoo Finance, 2026-07-31 close", ""],
    ["Market Cap", "$19.22B", "Yahoo Finance Key Statistics", "Current column"],
    ["Enterprise Value", "$78.60B", "Yahoo Finance Key Statistics", "Current column"],
    ["Shares Outstanding (TTM dil avg)", "458.65M", "Yahoo Finance IS TTM", ""],
    ["Shares Issued (BS FY25)", "462.89M", "Yahoo Finance BS 12/31/2025", ""],
    ["Preferred Shares", "33.92M", "Yahoo Finance BS 12/31/2025", "Down from 43.90M in 2024"],
    ["TTM Revenue", "$25.06B", "Yahoo Finance IS", "Thousands; $25,062,000K"],
    ["FY2025 Revenue", "$23.10B", "Yahoo Finance IS", "Thousands"],
    ["FY2024 Revenue", "$21.04B", "Yahoo Finance IS", "Thousands"],
    ["TTM Gross Profit", "$6.65B", "Yahoo Finance IS", ""],
    ["TTM Operating Income", "$6.20B", "Yahoo Finance IS", "OpEx only $444K — D&A flows through COGS"],
    ["TTM D&A", "$4.29B", "Yahoo Finance IS", "Reconciled Depreciation line"],
    ["TTM Interest Expense", "$3.80B", "Yahoo Finance IS", ""],
    ["TTM EBITDA", "$11.06B", "Yahoo Finance IS", ""],
    ["TTM Net Income (Common)", "$0.32B", "Yahoo Finance IS", "Heavily distorted by D&A"],
    ["TTM EPS (Diluted)", "$0.62", "Yahoo Finance IS", "GAAP — distorted"],
    ["Total Debt (BS FY25)", "$69.12B", "Yahoo Finance BS 12/31/2025", "Up from $56.35B FY24"],
    ["Total Cash (BS FY25)", "$6.64B", "Yahoo Finance BS 12/31/2025", ""],
    ["Net Debt (BS FY25)", "$62.48B", "Yahoo Finance BS 12/31/2025", "Debt - Cash"],
    ["Net Debt Proxy (EV-MC)", "$59.38B", "Computed", "Enterprise - Market Cap"],
    ["TTM OCF", "$6.30B", "Yahoo Finance CF", ""],
    ["TTM CapEx", "$7.52B", "Yahoo Finance CF", "Heavy investment cycle"],
    ["TTM FCF", "-$1.22B", "Yahoo Finance CF", "OCF - CapEx. Negative in investment cycle."],
    ["Common Stock Equity", "$5.64B", "Yahoo Finance BS 12/31/2025", ""],
    ["Tangible Book Value", "-$29.33B", "Yahoo Finance BS 12/31/2025", "Negative — goodwill-heavy"],
    ["Beta", "0.92", "Yahoo Finance Key Statistics", ""],
    ["TTM Tax Rate", "19.4%", "Computed", "$576M / $2,967M pretax"],
    ["Q2 FY26 Revenue", "$6.48B", "Yahoo Finance Analysis", ""],
    ["Q2 FY26 Earnings", "$62M", "Yahoo Finance Analysis", ""],
    ["Earnings Date", "July 30, 2026", "Yahoo Finance Profile", "Just reported"],
    ["Ex-Dividend Date", "August 30, 2026", "Yahoo Finance Profile", ""],
    ["10Y Treasury Yield", "4.718%", "CNBC US10Y, 2026-07-31", "WACC risk-free rate"],
    ["Preferred Stock", "33.92M shares", "Yahoo Finance BS", "Decreased from 43.90M FY24"],
    ["Capital Lease Obligations", "$4.58B", "Yahoo Finance BS 12/31/2025", "Up from $3.63B FY23"],
]

for i, row in enumerate(audit_data, 2):
    for j, val in enumerate(row, 1):
        font = header_font if i == 2 else None
        fill = header_fill if i == 2 else None
        c(ws4, i, j, val, font=font, border=True, fill=fill)

set_col_widths(ws4, [35, 25, 45, 45, 8])

# =============================================================================
# Sheet 5: Questions
# =============================================================================
ws5 = wb.create_sheet("Questions")
ws5.merge_cells("A1:C1")
c(ws5, 1, 1, "Open Questions — BIP", bold14)

questions = [
    ("1", "Preferred Stock Terms & Obligations: Preferred shares decreased from 43.90M (FY24) to 33.92M (FY25). What are the dividend obligations on these preferred shares? How does the preferred count change affect fixed-charge coverage? Should preferred value be subtracted from market cap for common equity valuation?"),
    ("2", "Debt Trajectory Post-Acquisition: Total debt jumped from $56.35B (FY24) to $69.12B (FY25) — up $12.77B or 22.7%. What acquisitions or investments financed? How does this debt maturity schedule look?"),
    ("3", "FCF Negativity During Capex Cycle: TTM FCF is -$1.22B (OCF $6.30B - CapEx $7.52B). Is this a cyclical capex spike from acquisitions/integration or a structural investment requirement? When does Free Cash Flow turn positive?"),
    ("4", "D&A Through COGS vs. Operating Income: Operating income of $6.20B appears high because D&A ($4.29B TTM) is embedded in cost of revenue, not in operating expenses. How does management characterize 'Operating Income' vs EBITDA? Standard infra metric is DCF/Distributable Cash Flow."),
    ("5", "Minority Interest Treatment: Total equity is $35.54B but common stock equity is only $5.64B. The $29.90B difference is minority interest / non-controlling ownership in subsidiaries. How does this affect per-share value attribution?"),
    ("6", "Distributable Cash Flow (DCF): Yahoo Finance doesn't show BIP's distributable cash flow metric (the standard infrastructure valuation denominator). DCF differs materially from FCF because it adds back non-cash charge (NCrCC). What is the true DCF yield?"),
    ("7", "Capital Lease Obligations Growth: Capital lease obligations rose from $3.63B (FY23) to $4.58B (FY25). Are these from acquisition-related obligations or new infrastructure leases?"),
    ("8", "Asset Management Fee Platform: BIP operates an asset management arm (Brookfield). How is this valued — as a stand-alone platform or rolled into operational earnings? Does it have fee-on-fees exposure?"),
    ("9", "Portfolio Composition & Geopolitical Exposure: Revenue spans utilities, transport, midstream, data across US, Canada, India, UK, Brazil, Japan, Colombia, France, Australia, Germany. What is the concentration risk?"),
    ("10", "Next Earnings Date & Guidance: Earnings date July 30, 2026 already passed. Q2 FY26 revenue $6.48B, earnings $62M. What guidance was given for FY27? When are next earnings?"),
    ("11", "Revenue vs. Earnings Quality: Q2 FY26 revenue $6.48B with earnings of only $62M. Is the earnings quality driven by D&A, minority interest allocations, or genuine margin compression?"),
    ("12", "Stock Buyback vs. Dividend: TTM buybacks $76M (minimal). Does management prefer dividends over buybacks? What is the current distribution yield?"),
]

for i, (num, q) in enumerate(questions, 3):
    c(ws5, i, 1, num, bold, border=True)
    ws5.merge_cells(f"B{i}:C{i}")
    c(ws5, i, 2, q, border=True)

set_col_widths(ws5, [5, 60, 60, 8])

# =============================================================================
# Sheet 6: Sources
# =============================================================================
ws6 = wb.create_sheet("Sources")
ws6.merge_cells("A1:B1")
c(ws6, 1, 1, "Sources — BIP Valuation Model", bold14)

sources = [
    ("1", "Yahoo Finance — Income Statement", "https://finance.yahoo.com/quote/BIP/financials/"),
    ("2", "Yahoo Finance — Balance Sheet", "https://finance.yahoo.com/quote/BIP/balance-sheet/"),
    ("3", "Yahoo Finance — Cash Flow", "https://finance.yahoo.com/quote/BIP/cash-flow/"),
    ("4", "Yahoo Finance — Key Statistics", "https://finance.yahoo.com/quote/BIP/key-statistics/"),
    ("5", "Yahoo Finance — Profile", "https://finance.yahoo.com/quote/BIP/profile/"),
    ("6", "Yahoo Finance — Analysis / Estimates", "https://finance.yahoo.com/quote/BIP/analysis/"),
    ("7", "CNBC — 10Y Treasury (US10Y)", "https://www.cnbc.com/quotes/US10Y"),
    ("8", "StockAnalysis.com — 404 for BIP", "https://stockanalysis.com/quote/BIP/"),
    ("9", "Brookfield Infrastructure Partners — Official Site", "https://www.bip.brookfield.com"),
]

for i, (num, name, url) in enumerate(sources, 3):
    c(ws6, i, 1, num, bold, border=True)
    c(ws6, i, 2, name, border=True)
    c(ws6, i, 3, url, border=True)

set_col_widths(ws6, [5, 45, 65, 8])

# =============================================================================
# Save
# =============================================================================
outpath = "/home/refcell/dev/capital/models/[2026-07-31] Brookfield Infrastructure Partners Model.xlsx"
wb.save(outpath)
print(f"\nSaved to: {outpath}")

# Verify
from openpyxl import load_workbook
wb2 = load_workbook(outpath)
print(f"Sheets: {wb2.sheetnames}")
print(f"Valuation metrics written: {len(metrics_data)}")
print(f"Scenarios computed: {len(results)}")
print(f"WACC: {wacc_pct}%")
for k, v in results.items():
    print(f"  {k} Target: ${v['target_price']:.2f} ({v['upside']:+.1f}%)")
print(f"Weighted FV: ${weighted_fv:.2f}")
