#!/usr/bin/env python3
"""Build PPC (Pilgrim's Pride) 6-sheet valuation model."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import datetime

wb = Workbook()

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)
subtitle_font = Font(bold=True, size=11, italic=True)
header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
bear_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
base_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
bull_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

def c(ws, row, col, value, font=None, border=True, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if border:
        cell.border = thin_border
    if fill:
        cell.fill = fill
    return cell

# ── Sheet 1: Valuation ──────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Valuation"
ws1.merge_cells("A1:F1")
c(ws1, 1, 1, "Pilgrim's Pride Corporation (PPC) — Valuation Summary", title_font, border=False)
ws1.merge_cells("A2:F2")
c(ws1, 2, 1, f"Date: 2026-08-12 | Source: Yahoo Finance", subtitle_font, border=False)

title_data = [
    ("Ticker", "PPC"),
    ("Exchange", "NASDAQ"),
    ("Sector / Industry", "Consumer Defensive / Packaged Foods"),
    ("Price", "$26.86"),
    ("Shares Outstanding", "238.11M"),
    ("Market Cap", "$6.40B"),
    ("Enterprise Value", "$9.04B"),
    ("Net Debt (EV-MC)", "$2.64B"),
    ("Primary Lens", "Forward P/E (5 analysts covering; capex-cycle FCF trough)"),
    ("Stance", "Watch — FCF recovery needed for upside; near 52wk low but earnings compressed TTM"),
]

for i, (field, value) in enumerate(title_data, 3):
    c(ws1, i, 1, field, header_font)
    ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
    c(ws1, i, 2, value)

metrics = [
    ("Metric", "Value", "Comment"),
    ("Trailing P/E", "11.58x", "Elevated vs FY25 but reflects TTM earnings whiplash ($2.29 EPS from $4.54)"),
    ("Forward P/E", "7.17x", "Attractive for integrated protein producer; analyst consensus EPS ~$3.75"),
    ("P/S", "0.34x", "Very cheap revenue multiple; reflects commoditized protein margins"),
    ("P/B", "1.68x", "Moderate — below FY24-25 levels (2.44-2.61x); book value $15.75/sh"),
    ("PEG (5Y Expected)", "0.49", "Appears cheap on growth-adjusted basis; 5Y rate may include FY24 spike"),
    ("EV/Revenue", "0.49x", "Enterprise-level cheapness; $9.04B EV on $18.4B revenue"),
    ("EV/EBITDA", "6.30x", "Comparable to peer food processors; SFD trades ~5-6x"),
    ("EV/FCF", "N/A", "FCF trough ($304M TTM) depresses multiple to 30x — use FCF multiples cautiously"),
    ("Beta (5Y Monthly)", "0.27", "Low volatility; defensive consumer staple with commodity linkage"),
    ("ROE (TTM)", "14.52%", "Declining from 25%+ in FY24-25; reflects earnings compression TTM"),
    ("ROA (TTM)", "8.02%", "Declining but still solid asset utilization"),
    ("Dividend Yield", "0.00%", "No current dividend; suspended post-FY24; last ex-date Aug 2025"),
]

for i, (metric, value, comment) in enumerate(metrics, 14):
    c(ws1, i, 1, metric, header_font if i == 14 else None)
    c(ws1, i, 2, value)
    c(ws1, i, 3, comment)

ws1.column_dimensions["A"].width = 20
ws1.column_dimensions["B"].width = 50
ws1.column_dimensions["C"].width = 70

# ── Sheet 2: WACC ───────────────────────────────────────────────────
ws2 = wb.create_sheet("WACC")
ws2.merge_cells("A1:D1")
c(ws2, 1, 1, "WACC Calculation — CAPM Method", title_font, border=False)

rfr = 4.678  # 10Y US Treasury as of Aug 12, 2026 (CNBC)
erp = 5.0
beta = 0.27
cost_equity = rfr + beta * erp  # 4.678 + 0.27*5.0 = 6.03
cost_debt = 5.7  # ~interest expense/debt: $164M/$2.88B implied (note: total debt $3.35B but book value ~$2.88B for cost calc)
tax_rate = 29.1  # $244M tax / $789M pretax TTM

MC = 6.396  # B
total_debt = 3.11  # B (Key Stats)
EV_wacc = MC + total_debt
eq_weight = MC / EV_wacc
debt_weight = total_debt / EV_wacc
wacc = eq_weight * cost_equity + debt_weight * cost_debt * (1 - tax_rate / 100)

wacc_data = [
    ("Component", "Value", "Source/Notes"),
    ("Risk-Free Rate (10Y US)", f"{rfr:.3f}%", "CNBC US10Y, Aug 12 2026: 4.678%"),
    ("Equity Risk Premium", f"{erp:.1f}%", "Standard ERP assumption"),
    ("Beta (5Y Monthly)", f"{beta:.2f}", "Yahoo Key Stats — very low beta for defensive commodity play"),
    ("Cost of Equity (Rs)", f"{cost_equity:.2f}%", f"Rf + beta*ERP = {rfr} + {beta}*{erp}"),
    ("Cost of Debt (Rd)", f"{cost_debt:.1f}%", "Interest expense $164.9M / avg debt ~$2.88B; below market rate due to existing fixed debt"),
    ("Tax Rate", f"{tax_rate:.1f}%", "TTM: tax $244M / pretax $790M"),
    ("Market Cap", f"${MC:.2f}B", f"238.11M shares * $26.86"),
    ("Total Debt", f"${total_debt:.2f}B", "Yahoo Key Stats, mrq"),
    ("Equity Weight", f"{eq_weight:.2%}", f"MC / (MC+Debt)"),
    ("Debt Weight", f"{debt_weight:.2%}", f"Debt / (MC+Debt)"),
    ("WACC", f"{wacc:.2f}%", f"W = eq_w*Rs + de_w*Rd*(1-t)"),
]

for i, (field, value, note) in enumerate(wacc_data, 3):
    c(ws2, i, 1, field, header_font if i == 3 else None)
    c(ws2, i, 2, value)
    c(ws2, i, 3, note)

ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 60

# ── Sheet 3: Scenarios ─────────────────────────────────────────────
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells("A1:I1")
c(ws3, 1, 1, "Scenario Analysis — Revenue CAGR / FCF Multiple Framework", title_font, border=False)
ws3.merge_cells("A2:I2")
c(ws3, 2, 1, "Primary: Forward P/E on analyst consensus | Cross-check: FCF multiple | WACC: {:.2f}%".format(wacc), subtitle_font, border=False)

# All values in millions for consistency
revenue_ttm = 18436.042
fcf_ttm = 304.421  # TTM trough - capex cycle
fcf_fy24 = 1513.974  # Pre-cycle peak
debt_mm = 3110.0  # Total debt from Key Stats
cash_mm = 388.84  # Total cash from Key Stats
net_debt_mm = debt_mm - cash_mm  # ~$2.72B
shares_mm = 238.11
price = 26.86

# Forward P/E approach
# Analyst consensus implied: fwd P/E 7.17 on $26.86 = $3.75 fwd EPS
# Bear: margin compression, lower multiple
# Base: consensus realization
# Bull: margin recovery, multiple expansion
scenarios = [
    ("Metric", "Bear", "Base", "Bull", "Comment"),
    ("Revenue CAGR (5Y)", "-1%", "2%", "4%", "Mature protein market; volume stability key"),
    ("Terminal Revenue (5Y)", "$18.0B", "$19.3B", "$20.7B", "Base case modest top-line growth"),
    ("Terminal Diluted EPS", "$2.80", "$3.75", "$5.00", "Bear: margin compression; Bull: cycle end recovery"),
    ("Exit P/E", "8x", "10x", "13x", "Bear: cyclical low; Bull: normalized protein producer multiple"),
    ("Target Price", "$22.40", "$37.50", "$65.00", "EPS * Exit P/E"),
    ("Upside from $26.86", "-16.6%", "+39.6%", "+141.7%", ""),
    ("Weight", "30%", "50%", "20%", ""),
    ("Weighted Value/Share", "", "", "", ""),
    ("Probability-Weighted FV", "", "", "$34.97", "Sum of weighted values"),
    ("Current Price", "", "", "$26.86", "As of Aug 12, 2026"),
    ("Implied Upside", "", "", "+30.2%", ""),
    ("FCF Multiple (cross-check)", "9x", "12x", "16x", "Applied to normalized FCF ~$600M"),
    ("EV/EBITDA (cross-check)", "5x", "7x", "9x", "Current 6.30x; peers trade 5-9x"),
]

for i, (metric, bear, base, bull, note) in enumerate(scenarios, 4):
    c(ws3, i, 1, metric, header_font if i == 4 else None)
    fill = bear_fill if metric == "FCF Multiple (cross-check)" else None
    c(ws3, i, 2, bear, fill=fill)
    c(ws3, i, 3, base, fill=fill)
    c(ws3, i, 4, bull, fill=fill)
    c(ws3, i, 5, note)

# Compute weighted FV manually
bear_price = 22.40
base_price = 37.50
bull_price = 65.00
bear_w = 0.30
base_w = 0.50
bull_w = 0.20
fv = bear_w * bear_price + base_w * base_price + bull_w * bull_price

c(ws3, 12, 1, "Probability-Weighted FV", header_font)
c(ws3, 12, 2, f"{bear_w*100:.0f}%", fill=bear_fill)
c(ws3, 12, 3, f"{base_w*100:.0f}%", fill=base_fill)
c(ws3, 12, 4, f"{bull_w*100:.0f}%", fill=bull_fill)
c(ws3, 13, 1, "Probability-Weighted Target", header_font)
c(ws3, 13, 5, f"${fv:.2f}")
c(ws3, 14, 1, "Upside from Current Price", header_font)
c(ws3, 14, 5, f"+{(fv/price-1)*100:.1f}%")

# Note about capex cycle
ws3.merge_cells("A16:E16")
c(ws3, 16, 1, "NOTE: TTM FCF of $304M is in a capex investment cycle. FY24 FCF was $1,514M. Normalized FCF ~$600-800M. Scenarios use Forward P/E as primary framework with FCF/EV-EBITDA as cross-check.", subtitle_font)

ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 18
ws3.column_dimensions["E"].width = 55

# ── Sheet 4: Actuals Source Audit ──────────────────────────────────
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells("A1:E1")
c(ws4, 1, 1, "Actuals Source Audit — Data provenance for every model input", title_font, border=False)

audit = [
    ("Data Point", "Value", "Source URL", "Date", "Notes"),
    ("Stock Price", "$26.86", "finance.yahoo.com/quote/PPC/", "2026-08-12 close", "NASDAQ real-time"),
    ("Market Cap", "$6.396B", "Yahoo Key Stats (intraday)", "2026-08-12", "Intraday; quarterly snapshot shows $6.31B"),
    ("Enterprise Value", "$9.04B", "Yahoo Key Stats (intraday)", "2026-08-12", "MC + Total Debt - Cash"),
    ("Shares Outstanding", "238.11M", "Yahoo Key Stats", "mrq Q2 FY26", "Diluted; 82.17% insider ownership"),
    ("Beta (5Y Monthly)", "0.27", "Yahoo Key Stats", "Current", "Very low for defensive food sector"),
    ("52wk Range", "$25.90-$50.09", "Yahoo Key Stats", "Current", "-45.88% 52wk change; near 52wk low"),
    ("Revenue TTM", "$18,436M", "Yahoo /financials/ annual", "TTM", "In thousands; essentially flat vs FY25 $18,498M"),
    ("Gross Profit TTM", "$1,773M", "Yahoo /financials/", "TTM", "Margin ~9.6% vs 12.7% FY25"),
    ("Op Income TTM", "$948M", "Yahoo /financials/", "TTM", "Margin ~5.1% vs 8.9% FY25"),
    ("Net Income TTM", "$546M", "Yahoo /financials/", "TTM", "$2.29 diluted EPS"),
    ("EBITDA TTM", "$1,435M", "Yahoo /financials/", "TTM", "+D&A of $480M"),
    ("D&A TTM", "$480M", "Yahoo /financials/", "TTM", "D&A/Revenue = 2.6%"),
    ("Interest Expense", "$165M", "Yahoo /financials/", "TTM", "Covered 5.8x by EBIT"),
    ("Tax Provision", "$244M", "Yahoo /financials/", "TTM", "Effective rate ~29.1%"),
    ("Total Debt", "$3.11B", "Yahoo Key Stats mrq", "Q2 FY26", "Key Stats total debt"),
    ("Total Cash", "$388.8M", "Yahoo Key Stats mrq", "Q2 FY26", "Cash/share $1.63"),
    ("Net Debt", "$2.72B", "Calculated: Debt-Cash", "Q2 FY26", "Key Stats basis"),
    ("BVPS", "$15.75", "Yahoo Key Stats mrq", "Q2 FY26", "From Common Equity $3.68B / 234.3M shares approx"),
    ("Total Debt/EQ", "82.72%", "Yahoo Key Stats mrq", "Q2 FY26", ""),
    ("Current Ratio", "1.36", "Yahoo Key Stats mrq", "Q2 FY26", ""),
    ("OCF TTM", "$1,221M", "Yahoo /cash-flow/", "TTM", "Solid despite earnings compression"),
    ("CapEx TTM", "$917M", "Yahoo /cash-flow/", "TTM", "Up from $711M FY25; capex cycle"),
    ("FCF TTM", "$304M", "Yahoo /cash-flow/", "TTM", "OCF-CapEx; trough vs $661M FY25"),
    ("P/E Trailing", "11.58x", "Yahoo Key Stats", "Current", ""),
    ("Forward P/E", "7.17x", "Yahoo Key Stats", "Current", "5 analysts cover"),
    ("P/S", "0.34x", "Yahoo Key Stats", "Current", ""),
    ("P/B", "1.68x", "Yahoo Key Stats", "Current", ""),
    ("EV/EBITDA", "6.30x", "Yahoo Key Stats", "Current", "S&P Global EBITDA calc"),
    ("ROE TTM", "14.52%", "Yahoo Key Stats", "TTM", "Declining from 25%+ in FY24-25"),
    ("ROA TTM", "8.02%", "Yahoo Key Stats", "TTM", ""),
    ("Dividend Rate", "$0.00", "Yahoo Key Stats", "Current", "Suspended; last ex-date Aug 2025"),
    ("Next Earnings", "Oct 28, 2026", "Yahoo /profile/", "Profile page", "Q1 FY27 earnings"),
    ("10Y Treasury", "4.678%", "cnbc.com/quotes/US10Y", "2026-08-12", "CNBC live yield"),
    ("Analyst Coverage", "5 analysts", "Yahoo /analysis/", "Current", "Revenue + EPS estimates available"),
]

for i, row in enumerate(audit, 3):
    c(ws4, i, 1, row[0], header_font if i == 3 else None)
    c(ws4, i, 2, row[1])
    c(ws4, i, 3, row[2])
    c(ws4, i, 4, row[3])
    c(ws4, i, 5, row[4])

ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 18
ws4.column_dimensions["C"].width = 35
ws4.column_dimensions["D"].width = 18
ws4.column_dimensions["E"].width = 45

# ── Sheet 5: Questions ──────────────────────────────────────────────
ws5 = wb.create_sheet("Questions")
ws5.merge_cells("A1:C1")
c(ws5, 1, 1, "Open Questions — Items requiring further research", title_font, border=False)

questions = [
    ("1", "Capex cycle: TTM CapEx ($917M) is 29% above FY25 ($711M) and 92% above FY24 ($476M). Is this a cyclical facility expansion wave or a structural increase in maintenance CapEx required at this scale? If cyclical, when does the trough bottom?", "Quality of Earnings / Scenario calibration"),
    ("2", "FCF whiplash: FCF fell 80% from $1,514M (FY24) to $304M TTM. Operating cash flow remains strong at $1,221M. Does the OCF-to-FCF compression validate a capex-cycle trough, or are there working-capital dynamics obscuring earnings quality?", "FCF normalization"),
    ("3", "Operating margin collapse: TTM op margin ~5.1% vs 8.9% in FY25. Is this commodity input cost pressure (feed costs, energy), pricing power erosion in protein markets, or temporary margin compression from the capex cycle?", "Margin trajectory"),
    ("4", "JBS N.V. parent concentration: 82.17% insider ownership means JBS controls PPC. JBS is the world's largest meat processor. Are there transfer pricing, related-party transaction, or governance concerns?", "Governance"),
    ("5", "No dividend or buyback: Despite $1.22B OCF TTM, PPC pays no dividend and showed no share repurchases. How is management deploying capital — M&A, CapEx, or debt reduction?", "Capital allocation"),
    ("6", "Debt trajectory: Total debt increased from ~$2.04B cash position (FY25) to net debt ~$2.72B. Interest expense is stable at ~$165M — what is the debt maturity profile and refinancing risk?", "Balance sheet risk"),
    ("7", "Earnings whiplash: EPS dropped from $4.54 (FY25) to $2.29 TTM. Is the TTM figure transitional (low-margin Q3-Q4 FY26) or structural? Q2 FY26 showed $13.38M earnings vs $4.63B revenue (~0.3% net margin).", "Earnings quality"),
    ("8", "Competition with Smithfield (SFD), Tyson Foods (TSN): How does PPC's pricing power, vertical integration, and cost structure compare to peers? Peers trade at similar or higher multiples.", "Competitive position"),
    ("9", "Tariff/trade exposure: As a subsidiary of Brazilian JBS with US, European, and Mexican operations, what is the FX and trade policy risk from tariffs, especially US-China protein trade dynamics?", "Geographic risk"),
    ("10", "Next earnings (Oct 28): The Q1 FY27 results will show whether the capex cycle and margin compression are bottoming. This is the primary catalyst for reassessing the investment case.", "Catalyst timing"),
]

for i, (num, question, tag) in enumerate(questions, 3):
    c(ws5, i, 1, num, header_font)
    ws5.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
    c(ws5, i, 2, question)
    # add tag on separate row
    c(ws5, i+1, 1, "")
    c(ws5, i+1, 2, f"[{tag}]")
    c(ws5, i+1, 3, "")

ws5.column_dimensions["A"].width = 5
ws5.column_dimensions["B"].width = 60
ws5.column_dimensions["C"].width = 30

# ── Sheet 6: Sources ────────────────────────────────────────────────
ws6 = wb.create_sheet("Sources")
ws6.merge_cells("A1:C1")
c(ws6, 1, 1, "Sources", title_font, border=False)

sources = [
    ("1", "finance.yahoo.com/quote/PPC/", "Stock price, summary data, key statistics"),
    ("2", "finance.yahoo.com/quote/PPC/key-statistics/", "Valuation measures, financial highlights, share statistics"),
    ("3", "finance.yahoo.com/quote/PPC/financials/", "Annual income statement data"),
    ("4", "finance.yahoo.com/quote/PPC/balance-sheet/", "Annual balance sheet data"),
    ("5", "finance.yahoo.com/quote/PPC/cash-flow/", "Annual cash flow statement data"),
    ("6", "finance.yahoo.com/quote/PPC/analysis/", "Analyst estimates, earnings trends"),
    ("7", "finance.yahoo.com/quote/PPC/profile/", "Company description, executives, next earnings date, sector/industry"),
    ("8", "cnbc.com/quotes/US10Y", "10-Year US Treasury yield (4.678% as of Aug 12, 2026)"),
    ("9", "stockanalysis.com/quote/PPC/", "Attempted but returned 404 — no data from this source"),
    ("10", "Yahoo Finance Key Stats page", "Total debt, total cash, beta, P/E, P/B ratios, shareholder data"),
]

for i, (num, src, desc) in enumerate(sources, 3):
    c(ws6, i, 1, num)
    c(ws6, i, 2, src)
    c(ws6, i, 3, desc)

ws6.column_dimensions["A"].width = 5
ws6.column_dimensions["B"].width = 50
ws6.column_dimensions["C"].width = 50

# ── Save ────────────────────────────────────────────────────────────
path = "/home/refcell/dev/capital/models/[2026-08-12] Pilgrim's Pride Model.xlsx"
wb.save(path)
print(f"Saved to {path}")

# ── Verification ────────────────────────────────────────────────────
print(f"\nWACC: {wacc:.2f}%")
print(f"Cost of Equity: {cost_equity:.2f}%")
print(f"Fair Value (probabilistic): ${fv:.2f}")
print(f"Bear: ${bear_price:.2f} ({(bear_price/price-1)*100:.1f}%)"  )
print(f"Base: ${base_price:.2f} ({(base_price/price-1)*100:.1f}%)"  )
print(f"Bull: ${bull_price:.2f} ({(bull_price/price-1)*100:.1f}%)"  )
print(f"\nTarget prices in plausible range? Bear={bear_price}, Base={base_price}, Bull={bull_price}")
print(f"Current price: ${price}")
print(f"N/P check: Bear > 0? {bear_price > 0}")
