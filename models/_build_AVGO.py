#!/usr/bin/env python3
"""Build Broadcom's six-sheet valuation workbook."""
from math import isclose
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).with_name("2026-09-05 Broadcom Model.xlsx")
PRICE = 357.90
SHARES = 4_760.0  # millions
MARKET_CAP = 1_703_000.0  # USD millions
ENTERPRISE_VALUE = 1_738_444.0  # USD millions
NET_DEBT = 35_444.0  # USD millions
FY27_REVENUE = 173_100.0  # visible StockAnalysis headline consensus, USD millions

SCENARIOS = {
    "Bear": {"growth": 0.08, "margin": 0.35, "multiple": 18.0, "net_debt": 10_000.0, "shares": 4_900.0, "discount": 0.115, "weight": 0.20},
    "Base": {"growth": 0.13, "margin": 0.40, "multiple": 24.0, "net_debt": -10_000.0, "shares": 4_800.0, "discount": 0.105, "weight": 0.55},
    "Bull": {"growth": 0.18, "margin": 0.44, "multiple": 30.0, "net_debt": -30_000.0, "shares": 4_700.0, "discount": 0.095, "weight": 0.25},
}
for case in SCENARIOS.values():
    case["terminal_revenue"] = FY27_REVENUE * (1 + case["growth"]) ** 5
    case["terminal_fcf"] = case["terminal_revenue"] * case["margin"]
    case["implied_ev"] = case["terminal_fcf"] * case["multiple"]
    case["terminal_equity"] = case["implied_ev"] - case["net_debt"]
    case["terminal_price"] = case["terminal_equity"] / case["shares"]
    case["target"] = case["terminal_price"] / (1 + case["discount"]) ** 5
    case["upside"] = case["target"] / PRICE - 1
weighted_value = sum(case["target"] * case["weight"] for case in SCENARIOS.values())

assert SCENARIOS["Bear"]["target"] < PRICE
assert 250 < SCENARIOS["Base"]["target"] < 550
assert 200 < weighted_value < 700

rf, erp, beta = 0.0479, 0.05, 1.46
cost_equity = rf + beta * erp
pretax_debt_cost = 2_687.0 / 59_419.0
tax_rate = 0.14
equity_weight = MARKET_CAP / (MARKET_CAP + 59_419.0)
debt_weight = 1 - equity_weight
wacc = equity_weight * cost_equity + debt_weight * pretax_debt_cost * (1 - tax_rate)

wb = Workbook()
navy, blue, gold, gray, white = "17365D", "D9EAF7", "D8B34B", "E7E6E6", "FFFFFF"
thin = Side(style="thin", color="A6A6A6")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def put(ws, row, col, value, *, bold=False, fill=None, color="000000", size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Aptos", size=size, bold=bold, color=color)
    cell.fill = PatternFill("solid", fgColor=fill) if fill else PatternFill(fill_type=None)
    cell.border = border
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    return cell


def title(ws, text, end_col=5, subtitle=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    put(ws, 1, 1, text, bold=True, fill=navy, color=white, size=15)
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        put(ws, 2, 1, subtitle, bold=True, fill=blue)


def header(ws, row, values):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=True, fill=navy, color=white)


def widths(ws, values):
    for col, width in enumerate(values, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


# 1. Valuation
ws = wb.active
ws.title = "Valuation"
title(ws, "Broadcom Inc. (AVGO) — Valuation", 4, "Quote: September 4, 2026 close | Model date: September 5, 2026 | USD")
facts = [
    ("Company", "Broadcom Inc.", "Custom AI accelerators, networking semiconductors, connectivity and infrastructure software"),
    ("Ticker", "NASDAQ: AVGO", "Fiscal year ends in early November"),
    ("Price", PRICE, "StockAnalysis September 4 close"),
    ("Shares outstanding (M)", SHARES, "StockAnalysis filing-date/current shares"),
    ("Market capitalization ($M)", MARKET_CAP, "StockAnalysis rounded $1.70T; price × shares implies $1.704T"),
    ("Enterprise value ($M)", ENTERPRISE_VALUE, "Market cap plus $35.44B net debt"),
    ("Net debt ($M)", NET_DEBT, "$59.42B debt less $23.98B cash"),
    ("Primary valuation lens", "Discounted terminal FCF", "Forward non-GAAP P/E is the principal cross-check"),
    ("Stance", "Watch / selective buy", "Exceptional AI growth and cash conversion; concentration, expectations and valuation create a wide range"),
]
header(ws, 3, ["Field", "Value", "Comment"])
for row, values in enumerate(facts, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
header(ws, 14, ["Valuation metric", "Value", "Interpretation"])
metrics = [
    ("Trailing P/E", 45.59, "GAAP TTM EPS; acquisition amortization and SBC make it less useful than forward adjusted earnings"),
    ("Forward P/E", 20.64, "Based on provider non-GAAP adjusted estimates"),
    ("P/S", 19.11, "Very high on TTM revenue; rapid AI growth must persist"),
    ("P/FCF", 43.21, "2.31% TTM FCF yield"),
    ("EV/FCF", 44.11, "Debt-adjusted cash valuation"),
    ("EV/Sales", 19.51, "Premium reflects AI growth and software-like margins"),
    ("EV/EBITDA", 33.39, "High despite 58.42% EBITDA margin"),
    ("Interest coverage", 14.09, "Debt service is currently comfortable"),
    ("Analyst average target", 533.41, "49 analysts; not intrinsic value"),
]
for row, values in enumerate(metrics, 15):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
widths(ws, [31, 25, 96, 14])

# 2. WACC
ws = wb.create_sheet("WACC")
title(ws, "Broadcom — Weighted Average Cost of Capital", 4, "CAPM with market-value capital weights")
header(ws, 3, ["Component", "Value", "Source / formula"])
wacc_rows = [
    ("Risk-free rate", rf, "FRED DGS10, September 2, 2026"),
    ("Equity risk premium", erp, "Model assumption"),
    ("Levered beta", beta, "StockAnalysis five-year beta"),
    ("Cost of equity", cost_equity, "Risk-free rate + beta × ERP"),
    ("Pre-tax cost of debt", pretax_debt_cost, "TTM cash interest / total debt"),
    ("Normalized tax rate", tax_rate, "Model rate; TTM 5.45% and Q3 YTD 11.47% are unusually low"),
    ("Market capitalization", MARKET_CAP, "USD millions"),
    ("Total debt", 59_419.0, "USD millions"),
    ("Equity weight", equity_weight, "Equity / (equity + debt)"),
    ("Debt weight", debt_weight, "Debt / (equity + debt)"),
    ("WACC", wacc, "E/V × Ke + D/V × Kd × (1 − tax rate)"),
]
for row, values in enumerate(wacc_rows, 4):
    for col, value in enumerate(values, 1):
        cell = put(ws, row, col, value, bold=(col == 1 or values[0] == "WACC"), fill=(gold if values[0] == "WACC" else None))
        if col == 2 and values[0] not in {"Levered beta", "Market capitalization", "Total debt"}:
            cell.number_format = "0.00%"
widths(ws, [31, 22, 96, 14])

# 3. Scenarios
ws = wb.create_sheet("Scenarios")
title(ws, "Broadcom — Discounted FCF Scenarios", 6, "Five years from FY2027 consensus; financial values in USD millions")
header(ws, 3, ["Metric", "Bear", "Base", "Bull", "Notes"])
scenario_rows = [
    ("FY2027 revenue anchor", FY27_REVENUE, FY27_REVENUE, FY27_REVENUE, "Visible StockAnalysis headline estimate"),
    ("Revenue CAGR (5Y)", *(SCENARIOS[k]["growth"] for k in ("Bear", "Base", "Bull")), "AI custom silicon, networking and VMware growth after FY2027"),
    ("Terminal revenue", *(SCENARIOS[k]["terminal_revenue"] for k in ("Bear", "Base", "Bull")), "FY2027 revenue compounded five years"),
    ("Adjusted FCF margin", *(SCENARIOS[k]["margin"] for k in ("Bear", "Base", "Bull")), "TTM FCF margin is 44.22%"),
    ("Terminal FCF", *(SCENARIOS[k]["terminal_fcf"] for k in ("Bear", "Base", "Bull")), "Terminal revenue × FCF margin"),
    ("Exit FCF multiple", *(SCENARIOS[k]["multiple"] for k in ("Bear", "Base", "Bull")), "Compressed, durable-growth and leadership outcomes"),
    ("Implied enterprise value", *(SCENARIOS[k]["implied_ev"] for k in ("Bear", "Base", "Bull")), "Terminal FCF × multiple"),
    ("Net debt / (cash)", *(SCENARIOS[k]["net_debt"] for k in ("Bear", "Base", "Bull")), "Debt reduction and future capital allocation"),
    ("Shares outstanding", *(SCENARIOS[k]["shares"] for k in ("Bear", "Base", "Bull")), "SBC may offset repurchases"),
    ("Terminal price", *(SCENARIOS[k]["terminal_price"] for k in ("Bear", "Base", "Bull")), "Terminal equity value / shares"),
    ("Discount rate", *(SCENARIOS[k]["discount"] for k in ("Bear", "Base", "Bull")), "Scenario risk adjustment"),
    ("Present target price", *(SCENARIOS[k]["target"] for k in ("Bear", "Base", "Bull")), "Five-year terminal value discounted to present"),
    ("Upside / (downside)", *(SCENARIOS[k]["upside"] for k in ("Bear", "Base", "Bull")), "Versus September 4 close"),
    ("Probability", *(SCENARIOS[k]["weight"] for k in ("Bear", "Base", "Bull")), "20% / 55% / 25%"),
    ("Weighted value/share", *(SCENARIOS[k]["target"] * SCENARIOS[k]["weight"] for k in ("Bear", "Base", "Bull")), "Contribution to weighted value"),
]
for row, values in enumerate(scenario_rows, 4):
    for col, value in enumerate(values, 1):
        cell = put(ws, row, col, value, bold=(col == 1))
        if col in (2, 3, 4) and values[0] in {"Revenue CAGR (5Y)", "Adjusted FCF margin", "Discount rate", "Upside / (downside)", "Probability"}:
            cell.number_format = "0.0%"
        elif col in (2, 3, 4) and values[0] in {"Terminal price", "Present target price", "Weighted value/share"}:
            cell.number_format = "$0.00"
put(ws, 20, 1, "Probability-weighted fair value", bold=True, fill=gold)
put(ws, 20, 2, weighted_value, bold=True, fill=gold).number_format = "$0.00"
put(ws, 21, 1, "Upside from current price", bold=True, fill=gold)
put(ws, 21, 2, weighted_value / PRICE - 1, bold=True, fill=gold).number_format = "0.0%"
widths(ws, [31, 18, 18, 18, 96, 14])

# 4. Actuals Source Audit
ws = wb.create_sheet("Actuals Source Audit")
title(ws, "Broadcom — Actuals Source Audit", 5, "Figures in USD millions unless stated")
header(ws, 3, ["Data point", "Value", "Source URL", "Source date", "Notes"])
sa = "https://stockanalysis.com/stocks/avgo"
ir = "https://investors.broadcom.com/news-releases/news-release-details/broadcom-inc-announces-third-quarter-fiscal-year-2026-financial"
audit = [
    ("Stock price", "$357.90", f"{sa}/", "2026-09-04", "Closing price"),
    ("Market capitalization", "$1.70T", f"{sa}/statistics/", "2026-09-04", "Rounded provider value"),
    ("Enterprise value", "$1.74T", f"{sa}/statistics/", "2026-09-04", "Includes net debt"),
    ("Shares outstanding", "4.76B", f"{sa}/statistics/", "2026-09-04", "+1.00% YoY"),
    ("TTM revenue", "$89.104B", f"{sa}/financials/", "2026-08-02", "+48.69%"),
    ("TTM gross profit", "$67.287B", f"{sa}/financials/", "2026-08-02", "75.52% margin"),
    ("TTM operating income", "$43.294B", f"{sa}/financials/", "2026-08-02", "48.59% GAAP margin"),
    ("TTM net income", "$38.265B", f"{sa}/financials/", "2026-08-02", "42.94% margin"),
    ("TTM operating cash flow", "$40.653B", f"{sa}/financials/cash-flow-statement/", "2026-08-02", "Includes material working-capital use"),
    ("TTM capex", "$1.250B", f"{sa}/financials/cash-flow-statement/", "2026-08-02", "Provider reports cash outflow as negative"),
    ("TTM free cash flow", "$39.403B", f"{sa}/financials/cash-flow-statement/", "2026-08-02", "44.22% margin"),
    ("TTM stock compensation", "$8.482B", f"{sa}/financials/cash-flow-statement/", "2026-08-02", "21.5% of reported FCF"),
    ("Cash and investments", "$23.975B", f"{sa}/financials/balance-sheet/", "2026-08-02", "Official Q3 balance sheet agrees"),
    ("Total debt", "$59.419B", f"{sa}/financials/balance-sheet/", "2026-08-02", "$2.252B short-term plus $57.167B long-term"),
    ("Goodwill", "$97.801B", f"{sa}/financials/balance-sheet/", "2026-08-02", "Primarily acquisition-created; equals 52% of assets"),
    ("Q3 revenue", "$29.591B", ir, "2026-09-02", "+86% YoY"),
    ("Q3 AI semiconductor revenue", "$16.7B", ir, "2026-09-02", "+221% YoY and +54% QoQ"),
    ("Q3 semiconductor revenue", "$20.839B", ir, "2026-09-02", "70% of revenue; +127% YoY"),
    ("Q3 infrastructure software revenue", "$8.752B", ir, "2026-09-02", "30% of revenue; +29% YoY"),
    ("Q3 GAAP / non-GAAP EPS", "$2.68 / $3.32", ir, "2026-09-02", "Non-GAAP excludes amortization, SBC and other items"),
    ("Q4 revenue guidance", "$34.8B", ir, "2026-09-02", "+93% YoY; non-GAAP operating margin 66%"),
    ("FY2026 consensus revenue", "$105.87B", f"{sa}/forecast/", "2026-09-04", "20 analysts; +65.72%"),
    ("FY2026 adjusted EPS", "$11.60", f"{sa}/forecast/", "2026-09-04", "Non-GAAP; +70.12%"),
    ("FY2027 headline revenue", "$173.10B", f"{sa}/forecast/", "2026-09-04", "+63.50%; detailed row gated"),
    ("FY2027 headline adjusted EPS", "$19.49", f"{sa}/forecast/", "2026-09-04", "+68.00%; detailed row gated"),
    ("Analyst average target", "$533.41", f"{sa}/forecast/", "2026-09-04", "49 analysts; median $537.50"),
    ("Beta", "1.46", f"{sa}/statistics/", "2026-09-04", "Five-year beta"),
    ("Latest earnings", "September 2, 2026", ir, "2026-09-02", "Already released; next quarter is the next catalyst"),
]
for row, values in enumerate(audit, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
widths(ws, [35, 24, 100, 16, 78])

# 5. Questions
ws = wb.create_sheet("Questions")
title(ws, "Broadcom — Open Questions", 4, "Items that can change valuation or conviction")
header(ws, 3, ["#", "Question", "Why it matters", "Best evidence / next check"])
questions = [
    (1, "How concentrated is AI accelerator revenue by customer and program?", "A small number of hyperscalers can create timing, pricing and insourcing risk.", "Revenue and backlog concentration; customer-specific tape-out cadence."),
    (2, "Can management's expectation that AI revenue doubles in FY2027 and FY2028 be met without vendor financing?", "The valuation assumes extraordinary growth and capital availability.", "Customer capex, contracted backlog, financing guarantees and cancellation terms."),
    (3, "How much Q3/Q4 AI growth reflects memory and system content pass-through rather than higher-value silicon?", "Mix can raise revenue while compressing gross margin.", "AI revenue bridge, gross margin and content per accelerator."),
    (4, "Can consolidated gross margin hold near the mid-70s as lower-margin AI systems become a larger mix?", "A few margin points have large FCF implications at this scale.", "Quarterly GAAP and non-GAAP gross margins by mix."),
    (5, "What is the normalized cash tax rate after acquisition-related benefits?", "TTM 5.45% is unlikely to be a perpetual tax rate.", "Cash taxes, deferred tax asset use and jurisdiction mix."),
    (6, "How durable is VMware revenue after pricing, bundling and channel changes?", "Software margins fund cash flow, but customer disruption can create renewal risk.", "Renewal rates, bookings, RPO, workload churn and partner channel trends."),
    (7, "Can VMware's subscription transition sustain 29% infrastructure-software growth?", "Current growth may include repricing and portfolio rationalization that cannot repeat indefinitely.", "Recurring revenue, seats/workloads and net retention."),
    (8, "What returns justify $97.8B of goodwill and $26.3B of intangibles?", "Acquisition assets exceed tangible equity and create impairment risk.", "VMware synergy realization and segment cash returns."),
    (9, "Will debt repayment remain the priority after $8.45B of TTM repurchases?", "Net debt is manageable but capital allocation affects equity duration.", "Quarterly debt, buybacks, dividends and acquisition spending."),
    (10, "Why did receivables use $7.2B and inventory use $2.3B of TTM cash?", "Hypergrowth can consume working capital and make quarterly FCF timing volatile.", "DSO, inventory days, customer acceptance and payment terms."),
    (11, "Can supply partners support the HBM, advanced packaging and networking ramp?", "Broadcom is fabless and exposed to constrained external manufacturing capacity.", "Foundry/package commitments, lead times and prepayments."),
    (12, "How vulnerable are Google TPU and other custom programs to internal redesign or Marvell competition?", "Winning each generation is not automatic even with incumbent IP.", "Design-win disclosures and next-generation program timing."),
    (13, "Does the emerging AI infrastructure financing model create guarantees or lease backstops?", "Off-balance-sheet commitments could change the apparent capital-light model.", "Contractual commitments and credit-rating disclosures."),
    (14, "Can SBC fall from $8.48B as a percentage of revenue?", "FCF is strong, but compensation remains a real owner cost and shares grew 1% YoY.", "SBC/revenue, diluted shares and repurchase effectiveness."),
    (15, "When is the fiscal Q4 earnings release?", "It will test $34.8B guidance, $21.7B AI revenue and 66% non-GAAP operating margin.", "Broadcom investor-relations calendar."),
]
for row, values in enumerate(questions, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
widths(ws, [7, 84, 74, 72])

# 6. Sources
ws = wb.create_sheet("Sources")
title(ws, "Broadcom — Sources", 4, "Public sources accessed September 5, 2026")
header(ws, 3, ["#", "Source", "URL", "Use"])
sources = [
    (1, "StockAnalysis overview", f"{sa}/", "Price, identity, market data and analyst summary"),
    (2, "StockAnalysis financials", f"{sa}/financials/", "Historical statements, segments, margins and cash flow"),
    (3, "StockAnalysis balance sheet", f"{sa}/financials/balance-sheet/", "Cash, debt, goodwill, intangibles and equity"),
    (4, "StockAnalysis cash flow", f"{sa}/financials/cash-flow-statement/", "OCF, capex, FCF, SBC, buybacks and working capital"),
    (5, "StockAnalysis statistics", f"{sa}/statistics/", "Valuation ratios, shares, beta, returns and earnings date"),
    (6, "StockAnalysis forecast", f"{sa}/forecast/", "Consensus revenue, adjusted EPS and targets"),
    (7, "Broadcom Q3 FY2026 results", ir, "Official quarter, segment, cash flow and Q4 guidance"),
    (8, "FRED DGS10", "https://fred.stlouisfed.org/series/DGS10", "Risk-free-rate reference"),
    (9, "Broadcom company profile", "https://www.broadcom.com/company/about-us", "Products, strategy and corporate identity"),
]
for row, values in enumerate(sources, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
widths(ws, [7, 40, 112, 74])

wb.save(OUT)
check = load_workbook(OUT, data_only=False)
assert check.sheetnames == ["Valuation", "WACC", "Scenarios", "Actuals Source Audit", "Questions", "Sources"]
assert check["Valuation"]["B6"].value == PRICE
assert isclose(check["Scenarios"]["B20"].value, weighted_value, rel_tol=1e-12)
print(f"WACC: {wacc:.2%}")
for name, case in SCENARIOS.items():
    print(f"{name}: ${case['target']:.2f} ({case['upside']:+.1%})")
print(f"Probability-weighted fair value: ${weighted_value:.2f} ({weighted_value / PRICE - 1:+.1%})")
print(f"Created {OUT}")
