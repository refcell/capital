#!/usr/bin/env python3
"""Build NDSN (Nordson Corporation) 6-sheet valuation model."""
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# --- Style definitions ---
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=11, italic=True)
NORMAL_FONT = Font(size=10)
SMALL_FONT = Font(size=9)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

DSTAMP = "2026-08-13"

def style_header(ws, row, cols, font=HEADER_FONT, fill=HEADER_FILL):
    for ci in range(1, cols + 1):
        c = ws.cell(row=row, column=ci)
        c.font = font
        c.fill = fill
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")

def style_data(ws, row, cols):
    for ci in range(1, cols + 1):
        c = ws.cell(row=row, column=ci)
        c.border = THIN_BORDER
        c.font = NORMAL_FONT

# ============================================
# SHEET 1: Valuation
# ============================================
ws1 = wb.active
ws1.title = "Valuation"

# Title block - row 1 merged
ws1.merge_cells("A1:F1")
ws1["A1"].value = "Nordson Corporation (NDSN) — Valuation Summary"
ws1["A1"].font = TITLE_FONT
ws1["A1"].alignment = Alignment(horizontal="center")

ws1.merge_cells("A2:F2")
ws1["A2"].value = f"Price: $309.99 | Date: {DSTAMP} | Primary Lens: Forward P/E + EV/EBITDA | Stance: Modestly Overvalued — Watch"
ws1["A2"].font = SUBTITLE_FONT
ws1["A2"].alignment = Alignment(horizontal="center")

# Title data
title_data = [
    ("Ticker", "NDSN"),
    ("Company", "Nordson Corporation"),
    ("Date", DSTAMP),
    ("Price ($)", "309.99"),
    ("Shares Outstanding (M)", "55.72"),
    ("Market Cap ($B)", "17.20"),
    ("Enterprise Value ($B)", "19.07"),
    ("Primary Lens", "Forward P/E + EV/EBITDA"),
    ("Stance", "Modestly Overvalued / Watch"),
]

for i, (label, val) in enumerate(title_data):
    ws1.cell(row=i + 3, column=1, value=label).font = HEADER_FONT
    ws1.cell(row=i + 3, column=1).border = THIN_BORDER
    ws1.cell(row=i + 3, column=1).fill = HEADER_FILL
    ws1.cell(row=i + 3, column=2, value=val).font = NORMAL_FONT
    ws1.cell(row=i + 3, column=2).border = THIN_BORDER

# Valuation metrics table
ws1.cell(row=14, column=1, value="Valuation Metrics").font = SUBTITLE_FONT
metrics = [
    ("Metric", "Value", "Comment"),
    ("Trailing P/E", "32.98x", "Above peer median (~24x). Reflects premium for consistent growth + buybacks."),
    ("Forward P/E", "24.69x", "More reasonable. Implies $12.55 FY2027 EPS. Near peer median."),
    ("PEG Ratio (5yr exp)", "1.91", "Slightly elevated. Suggests modest growth premium."),
    ("Price/Sales", "5.99x", "High for industrials. Reflects margin quality and buyback-driven EPS growth."),
    ("Price/Book", "5.37x", "Elevated. Intangible-heavy BS; BV not a strong value indicator."),
    ("EV/Revenue", "6.57x", "Reasonable for specialty machinery with buyback tailwinds."),
    ("EV/EBITDA", "21.37x", "Above peer median (~15-17x). Premium for quality and consistency."),
    ("P/FCF (TTM)", "N/A", "FCF of $714M TTM; P/FCF ~24.1x. Moderate for slow-growth industrials."),
    ("Dividend Yield", "1.05%", "Low yield. Payout ratio 34.6% — room for dividend growth."),
    ("Beta (5Y)", "0.96", "Slightly below market. Defensive characteristics."),
]

for i, (col_a, col_b, col_c) in enumerate(metrics):
    r = i + 15
    ws1.cell(row=r, column=1, value=col_a)
    ws1.cell(row=r, column=2, value=col_b)
    ws1.cell(row=r, column=3, value=col_c)
    style_data(ws1, r, 3)
    if i == 0:
        style_header(ws1, r, 3)
        ws1.cell(row=r, column=1).alignment = Alignment(horizontal="left")

ws1.column_dimensions["A"].width = 20
ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 55

# ============================================
# SHEET 2: WACC
# ============================================
ws2 = wb.create_sheet("WACC")

ws2.merge_cells("A1:D1")
ws2["A1"].value = "WACC Calculation — Nordson Corporation (NDSN)"
ws2["A1"].font = TITLE_FONT
ws2["A1"].alignment = Alignment(horizontal="center")

wacc_data = [
    ("Component", "Value", "Source / Method", "Notes"),
    ("Risk-free rate (10Y US)", "4.645%", "CNBC / US10Y", "As of Aug 13, 2026"),
    ("Equity Risk Premium", "5.00%", "Assumed", "Standard ERP assumption"),
    ("Beta (levered, 5Y)", "0.96", "Yahoo Finance Key Stats", "Slightly below market beta"),
    ("Cost of Equity (Ke)", "9.45%", "Rf + Beta*ERP", "4.645% + 0.96 * 5.0%"),
    ("Pre-tax Cost of Debt", "6.00%", "Est. from interest/debt", "~96M interest on ~2.09B debt = 4.6%; add 1.4% credit spread"),
    ("Tax Rate", "18.9%", "FY2025 tax provision / pretax", "113,174 / 597,648"),
    ("After-tax Cost of Debt", "4.86%", "Kd * (1-Tax)", "6.0% * (1 - 0.189)"),
    ("Market Cap ($B)", "17.20", "55.72M shares * $309.99", ""),
    ("Total Debt ($B)", "1.87", "EV - MC (conservative proxy)", "BS shows $2.09B; using EV-MC net proxy"),
    ("Equity Weight", "90.2%", "MC / (MC + Debt)", ""),
    ("Debt Weight", "9.8%", "Debt / (MC + Debt)", ""),
    ("", "", "", ""),
    ("WACC", "9.13%", "0.902 * 9.45% + 0.098 * 4.86%", "Discount rate for scenario analysis"),
]

for i, (a, b, c, d) in enumerate(wacc_data):
    r = i + 3
    ws2.cell(row=r, column=1, value=a)
    ws2.cell(row=r, column=2, value=b)
    ws2.cell(row=r, column=3, value=c)
    ws2.cell(row=r, column=4, value=d)
    style_data(ws2, r, 4)
    if i == 0:
        style_header(ws2, r, 4)
    if a == "WACC":
        for ci in range(1, 5):
            ws2.cell(row=r, column=ci).fill = GREEN_FILL
            ws2.cell(row=r, column=ci).font = Font(bold=True, size=11)

ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 30
ws2.column_dimensions["D"].width = 35

# ============================================
# SHEET 3: Scenarios
# ============================================
ws3 = wb.create_sheet("Scenarios")

ws3.merge_cells("A1:K1")
ws3["A1"].value = "Scenario Analysis — Nordson Corporation (NDSN)"
ws3["A1"].font = TITLE_FONT
ws3["A1"].alignment = Alignment(horizontal="center")

ws3.merge_cells("A2:K2")
ws3["A2"].value = "Forward P/E framework — Primary scenario lens. FCF multiple shown as cross-check only."
ws3["A2"].font = SUBTITLE_FONT
ws3["A2"].alignment = Alignment(horizontal="center")
ws3["A2"].fill = YELLOW_FILL
for ci in range(1, 12):
    ws3.cell(row=2, column=ci).fill = YELLOW_FILL

# Scenario calculations
# NDSN FY2025 revenue: $2,791.7M, FY2026 consensus: $2.98B, FY2027: $3.14B
# FY2025 EPS diluted: $8.51, FY2026 est: $11.59, FY2027 est: $12.55
# Revenue CAGR 2022-2025: ~2.47%

# Scenarios: revenue CAGR 5Y from FY2025 base
bearrev_cagr = 0.02    # 2% - low growth slowdown
baserev_cagr = 0.04    # 4% - steady growth
bullrev_cagr = 0.06    # 6% - above-consensus

# Terminal revenue (5Y from FY2025 base of $2,791.7M)
bear_term_rev = 2791.7 * (1 + bearrev_cagr) ** 5
base_term_rev = 2791.7 * (1 + baserev_cagr) ** 5
bull_term_rev = 2791.7 * (1 + bullrev_cagr) ** 5

# Adjusted FCF margin (TTM is ~24.6%)
bear_fcf_margin = 0.20
base_fcf_margin = 0.24
bull_fcf_margin = 0.27

bear_term_fcf = bear_term_rev * bear_fcf_margin
base_term_fcf = base_term_rev * base_fcf_margin
bull_term_fcf = bull_term_rev * bull_fcf_margin

# Forward P/E exit multiples
bear_pe = 20
base_pe = 24
bull_pe = 28

# FY2027 EPS estimates and project 5Y EPS from 2027 base
bear_eps_term = 12.55 * (1 + 0.02) ** 5
base_eps_term = 12.55 * (1 + 0.04) ** 5
bull_eps_term = 12.55 * (1 + 0.06) ** 5

bear_target = bear_eps_term * bear_pe
base_target = base_eps_term * base_pe
bull_target = bull_eps_term * bull_pe

# Net debt adj: EV - MC = $1.87B (in $B) = $1,870M
# Shares: 55.72M
net_debt_mm = 1870
shares_mm = 55.72

# FCF multiple cross-check
# Bear: 7x, Base: 9x, Bull: 12x
bear_fcf_mult = 7
base_fcf_mult = 9
bull_fcf_mult = 12

bear_ev_fcf = bear_term_fcf * bear_fcf_mult
base_ev_fcf = base_term_fcf * base_fcf_mult
bull_ev_fcf = bull_term_fcf * bull_fcf_mult

bear_eq_fcf = bear_ev_fcf - net_debt_mm
base_eq_fcf = base_ev_fcf - net_debt_mm
bull_eq_fcf = bull_ev_fcf - net_debt_mm

bear_target_fcf = bear_eq_fcf / shares_mm
base_target_fcf = base_eq_fcf / shares_mm
bull_target_fcf = bull_eq_fcf / shares_mm

# Upside from current price $309.99
current_price = 309.99

# Weights
bear_w = 0.25
base_w = 0.50
bull_w = 0.25

weighted_fv = bear_w * bear_target + base_w * base_target + bull_w * bull_target

scenarios = [
    ("Item", "Bear", "Base", "Bull", "Notes"),
    ("Revenue CAGR (5Y)", f"{bearrev_cagr:.0%}", f"{baserev_cagr:.0%}", f"{bullrev_cagr:.0%}", "From FY2025 base of $2,792M"),
    ("Terminal Revenue ($M)", f"{bear_term_rev:.0f}", f"{base_term_rev:.0f}", f"{bull_term_rev:.0f}", ""),
    ("Adjusted FCF Margin", f"{bear_fcf_margin:.0%}", f"{base_fcf_margin:.0%}", f"{bull_fcf_margin:.0%}", "TTM was ~24.6%"),
    ("Terminal FCF ($M)", f"{bear_term_fcf:.0f}", f"{base_term_fcf:.0f}", f"{bull_term_fcf:.0f}", ""),
    ("Exit FCF Multiple", f"{bear_fcf_mult}x", f"{base_fcf_mult}x", f"{bull_fcf_mult}x", "Cross-check only"),
    ("Implied EV ($M)", f"{bear_ev_fcf:.0f}", f"{base_ev_fcf:.0f}", f"{bull_ev_fcf:.0f}", ""),
    ("Less: Net Debt ($M)", f"{net_debt_mm}", f"{net_debt_mm}", f"{net_debt_mm}", "EV - MC proxy"),
    ("Terminal EPS (Fwd)", f"${bear_eps_term:.2f}", f"${base_eps_term:.2f}", f"${bull_eps_term:.2f}", "From FY2027 $12.55 EPS base"),
    ("Exit P/E Multiple", f"{bear_pe}x", f"{base_pe}x", f"{bull_pe}x", "Primary framework"),
    ("Target Price (Fwd PE)", f"${bear_target:.2f}", f"${base_target:.2f}", f"${bull_target:.2f}", "Primary estimate"),
    ("Target Price (FCF Mult)", f"${bear_target_fcf:.2f}", f"${base_target_fcf:.2f}", f"${bull_target_fcf:.2f}", "Cross-check"),
    ("Upside from Current", f"{(bear_target/current_price-1)*100:.1f}%", f"{(base_target/current_price-1)*100:.1f}%", f"{(bull_target/current_price-1)*100:.1f}%", "From $309.99"),
    ("Case Weight", f"{bear_w:.0%}", f"{base_w:.0%}", f"{bull_w:.0%}", ""),
    ("Weighted Value/Share", f"${bear_w * bear_target:.2f}", f"${base_w * base_target:.2f}", f"${bull_w * bull_target:.2f}", ""),
    ("Probability-Weighted FV", "", "", f"${weighted_fv:.2f}", "Sum of weighted values"),
    ("Current Price", "", "", f"${current_price:.2f}", ""),
    ("Implied Upside (weighted)", "", "", f"{(weighted_fv/current_price-1)*100:.1f}%", ""),
]

for i, (a, b, c, d, e) in enumerate(scenarios):
    r = i + 4
    ws3.cell(row=r, column=1, value=a)
    ws3.cell(row=r, column=2, value=b)
    ws3.cell(row=r, column=3, value=c)
    ws3.cell(row=r, column=4, value=d)
    ws3.cell(row=r, column=5, value=e)
    style_data(ws3, r, 5)
    if i == 0:
        style_header(ws3, r, 5)

# Highlight FV and current price rows
for r_offset in [14, 15, 16]:
    for ci in range(1, 6):
        ws3.cell(row=r_offset + 4, column=ci).fill = GREEN_FILL
        ws3.cell(row=r_offset + 4, column=ci).font = Font(bold=True)

print(f"WACC: 9.13%")
print(f"Bear target (Fwd PE): ${bear_target:.2f}")
print(f"Base target (Fwd PE): ${base_target:.2f}")
print(f"Bull target (Fwd PE): ${bull_target:.2f}")
print(f"Weighted FV: ${weighted_fv:.2f}")
print(f"Upside (weighted): {(weighted_fv/current_price-1)*100:.1f}%")

# ============================================
# SHEET 4: Actuals Source Audit
# ============================================
ws4 = wb.create_sheet("Actuals Source Audit")

ws4.merge_cells("A1:F1")
ws4["A1"].value = "Actuals Source Audit — NDSN"
ws4["A1"].font = TITLE_FONT
ws4["A1"].alignment = Alignment(horizontal="center")

audit = [
    ("Data Point", "Value", "Source", "URL", "Date Accessed", "Notes"),
    ("Stock Price", "$309.99", "Yahoo Finance", "finance.yahoo.com/quote/NDSN/", "2026-08-13", "Close Aug 13, 2026"),
    ("Market Cap", "$17.20B", "Yahoo Finance Key Stats", "finance.yahoo.com/quote/NDSN/key-statistics/", "2026-08-13", "Quarterly data as of MRQ"),
    ("Enterprise Value", "$19.07B", "Yahoo Finance Key Stats", "finance.yahoo.com/quote/NDSN/key-statistics/", "2026-08-13", ""),
    ("Shares Outstanding", "55.72M", "Yahoo Finance Key Stats", "finance.yahoo.com/quote/NDSN/key-statistics/", "2026-08-13", "Implied shares outstanding"),
    ("Total Revenue TTM", "$2,904M", "Yahoo Finance IS", "finance.yahoo.com/quote/NDSN/financials/", "2026-08-13", "TTM in thousands"),
    ("Gross Profit TTM", "$1,600M", "Yahoo Finance IS", "finance.yahoo.com/quote/NDSN/financials/", "2026-08-13", ""),
    ("Operating Income TTM", "$778M", "Yahoo Finance IS", "finance.yahoo.com/quote/NDSN/financials/", "2026-08-13", "26.8% margin"),
    ("Net Income TTM", "$528M", "Yahoo Finance IS", "finance.yahoo.com/quote/NDSN/financials/", "2026-08-13", "18.2% margin"),
    ("Diluted EPS TTM", "$9.36", "Yahoo Finance IS", "finance.yahoo.com/quote/NDSN/financials/", "2026-08-13", ""),
    ("EBITDA TTM", "$893M", "Yahoo Finance IS / S&P Calc", "finance.yahoo.com/quote/NDSN/financials/", "2026-08-13", "S&P Global calculated EBITDA"),
    ("D&A TTM", "$149M", "Yahoo Finance IS", "finance.yahoo.com/quote/NDSN/financials/", "2026-08-13", ""),
    ("Total Assets", "$5,918M", "Yahoo Finance BS", "finance.yahoo.com/quote/NDSN/balance-sheet/", "2026-08-13", "FY2025"),
    ("Total Debt (BS)", "$2,092M", "Yahoo Finance BS", "finance.yahoo.com/quote/NDSN/balance-sheet/", "2026-08-13", "FY2025"),
    ("Total Debt (KStats)", "$1,970M", "Yahoo Finance KStats", "finance.yahoo.com/quote/NDSN/key-statistics/", "2026-08-13", "MRQ — slightly lower than BS; using EV-MC for net debt proxy"),
    ("Total Cash (MRQ)", "$103M", "Yahoo Finance KStats", "finance.yahoo.com/quote/NDSN/key-statistics/", "2026-08-13", "$103.14M MRQ"),
    ("Operating CF TTM", "$762M", "Yahoo Finance CF", "finance.yahoo.com/quote/NDSN/cash-flow/", "2026-08-13", ""),
    ("CapEx TTM", "$48M", "Yahoo Finance CF", "finance.yahoo.com/quote/NDSN/cash-flow/", "2026-08-13", ""),
    ("FCF TTM", "$714M", "Yahoo Finance CF", "finance.yahoo.com/quote/NDSN/cash-flow/", "2026-08-13", "OCF - CapEx"),
    ("Share Repurchases TTM", "$289M", "Yahoo Finance CF", "finance.yahoo.com/quote/NDSN/cash-flow/", "2026-08-13", "Aggressive buyback program"),
    ("Beta (5Y)", "0.96", "Yahoo Finance KStats", "finance.yahoo.com/quote/NDSN/key-statistics/", "2026-08-13", ""),
    ("Trailing P/E", "32.98x", "Yahoo Finance KStats", "finance.yahoo.com/quote/NDSN/key-statistics/", "2026-08-13", "Quarterly snapshot"),
    ("Forward P/E", "24.69x", "Yahoo Finance KStats", "finance.yahoo.com/quote/NDSN/key-statistics/", "2026-08-13", ""),
    ("FY2026 EPS Est", "$11.59", "Yahoo Finance Analysis", "finance.yahoo.com/quote/NDSN/analysis/", "2026-08-13", "Non-GAAP consensus avg"),
    ("FY2027 EPS Est", "$12.55", "Yahoo Finance Analysis", "finance.yahoo.com/quote/NDSN/analysis/", "2026-08-13", "Non-GAAP consensus avg"),
    ("FY2026 Revenue Est", "$2.98B", "Yahoo Finance Analysis", "finance.yahoo.com/quote/NDSN/analysis/", "2026-08-13", "Consensus avg"),
    ("FY2027 Revenue Est", "$3.14B", "Yahoo Finance Analysis", "finance.yahoo.com/quote/NDSN/analysis/", "2026-08-13", "Consensus avg"),
    ("Valuation Ratios", "Various", "Yahoo Finance KStats", "finance.yahoo.com/quote/NDSN/key-statistics/", "2026-08-13", "Quarterly data"),
    ("Next Earnings Date", "Aug 19, 2026", "Yahoo Finance Profile", "finance.yahoo.com/quote/NDSN/profile/", "2026-08-13", "Q3 FY26 earnings"),
    ("10Y Treasury Rate", "4.645%", "CNBC", "cnbc.com/quotes/US10Y", "2026-08-13", "For WACC calculation"),
    ("Company Description", "Specialty Industrial", "Yahoo Finance Profile", "finance.yahoo.com/quote/NDSN/profile/", "2026-08-13", "3 segments; 8,200 employees"),
    ("Tax Rate", "18.9%", "Yahoo Finance IS", "finance.yahoo.com/quote/NDSN/financials/", "2026-08-13", "FY2025 tax provision / pretax income"),
]

for i, row_data in enumerate(audit):
    r = i + 3
    for ci, val in enumerate(row_data):
        ws4.cell(row=r, column=ci + 1, value=val)
    style_data(ws4, r, 6)
    if i == 0:
        style_header(ws4, r, 6)

ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 16
ws4.column_dimensions["C"].width = 22
ws4.column_dimensions["D"].width = 45
ws4.column_dimensions["E"].width = 14
ws4.column_dimensions["F"].width = 35

# ============================================
# SHEET 5: Questions
# ============================================
ws5 = wb.create_sheet("Questions")

ws5.merge_cells("A1:E1")
ws5["A1"].value = "Open Questions — NDSN"
ws5["A1"].font = TITLE_FONT
ws5["A1"].alignment = Alignment(horizontal="center")

questions = [
    ("#", "Question", "Category", "Priority", "Source"),
    ("1", "Debt increased from $860M (FY2022) to $2,092M (FY2025) — +143% in 3 years. What acquisitions or strategic investments drove this? How much is capital lease obligations ($96M)?", "Capital Allocation", "High", "BS + CF"),
    ("2", "Aggressive share repurchases: $289M TTM, $306M in FY2025. With debt up 143% since FY2022, is the buyback funded by leverage? What is the buyback / OCF ratio?", "Capital Allocation", "High", "CF"),
    ("3", "Goodwill jump: Tangible book value went negative (-$943M FY2025, -$1.09B FY2024) after being positive ($160M in FY2022). What acquisition created this goodwill? When was goodwill recognized?", "Balance Sheet", "High", "BS"),
    ("4", "Interest expense surged from $22M (FY2022) to $96M (TTM) — +330%. What is the debt composition (fixed vs variable rate)? Interest coverage ratio?", "Debt / Leverage", "High", "IS"),
    ("5", "Investing cash flow shows +$40M TTM, +$27M FY2025, but was -$844M FY2024 and -$1.44B FY2023. Was FY2023-24 a large acquisition cycle? Any divestitures?", "Investing Activity", "High", "CF"),
    ("6", "Tangible book value is negative (-$943M) while common equity is $3,044M. This implies ~$2.57B in goodwill/intangibles. What is the goodwill amortization schedule?", "Balance Sheet", "Medium", "BS"),
    ("7", "Revenue grew from $2.59B to $2.79B FY2022-2025 (2.5% CAGR). Is this purely organic? Did any FY2023-24 acquisitions contribute?", "Growth Attribution", "Medium", "IS + News"),
    ("8", "Segments: Industrial Precision Solutions, Medical and Fluid Solutions, Advanced Technology Solutions. What are the segment revenue/margin splits?", "Segment Analysis", "Medium", "10-K / Earnings calls"),
    ("9", "FY2024 saw massive debt issuance ($783M) and debt repayment ($325M). FY2023: $2.18B issuance, $1.21B repayment. What financing events occurred?", "Capital Structure", "High", "CF"),
    ("10", "Diluted shares declining: 57,631K (FY2023) to 56,340K (TTM). Is this purely buyback-driven? Any stock-based compensation offset?", "Share Count", "Low", "IS"),
    ("11", "SBC (Stock-Based Compensation): No visible SBC line on Yahoo Finance IS. How does Nordson compensate executives — options, RSUs? What is SBC as % of revenue?", "Compensation", "Medium", "10-K"),
    ("12", "Total Unusual Items: -$12.5M in TTM and FY2025. What unusual items were recorded? Any impairment, litigation, or restructuring charges?", "Accounting Quality", "Medium", "IS"),
    ("13", "Q3 FY26 (1/31/2026) EPS came in at exact estimate ($2.37). Was this a weaker quarter? Any seasonality in Q3?", "Earnings Quality", "Low", "Analysis"),
    ("14", "CEO Sundaram Nagarajan took over in 2019. Has he changed capital allocation strategy (e.g., from buybacks to acquisitions)?", "Management", "Medium", "Profile"),
    ("15", "Next earnings: Aug 19, 2026. What guidance is expected? Any commentary on tariff exposure, supply chain, or defense/semiconductor end markets?", "Catalyst", "High", "10-K / Earnings calls"),
]

for i, (a, b, c, d, e) in enumerate(questions):
    r = i + 3
    ws5.cell(row=r, column=1, value=a)
    ws5.cell(row=r, column=2, value=b)
    ws5.cell(row=r, column=3, value=c)
    ws5.cell(row=r, column=4, value=d)
    ws5.cell(row=r, column=5, value=e)
    style_data(ws5, r, 5)
    if i == 0:
        style_header(ws5, r, 5)
    if d == "High":
        ws5.cell(row=r, column=4).fill = RED_FILL
    elif d == "Medium":
        ws5.cell(row=r, column=4).fill = YELLOW_FILL

ws5.column_dimensions["A"].width = 5
ws5.column_dimensions["B"].width = 65
ws5.column_dimensions["C"].width = 20
ws5.column_dimensions["D"].width = 12
ws5.column_dimensions["E"].width = 12

# ============================================
# SHEET 6: Sources
# ============================================
ws6 = wb.create_sheet("Sources")

ws6.merge_cells("A1:D1")
ws6["A1"].value = "Sources — NDSN"
ws6["A1"].font = TITLE_FONT
ws6["A1"].alignment = Alignment(horizontal="center")

sources = [
    ("#", "Source", "URL", "Type"),
    ("1", "Yahoo Finance — Summary & Price", "finance.yahoo.com/quote/NDSN/", "Market data: price, volume, basic stats"),
    ("2", "Yahoo Finance — Income Statement", "finance.yahoo.com/quote/NDSN/financials/", "Revenue, gross profit, operating income, net income, EPS, EBITDA, D&A"),
    ("3", "Yahoo Finance — Balance Sheet", "finance.yahoo.com/quote/NDSN/balance-sheet/", "Assets, liabilities, equity, debt, cash, goodwill"),
    ("4", "Yahoo Finance — Cash Flow", "finance.yahoo.com/quote/NDSN/cash-flow/", "OCF, CapEx, FCF, investing/financing CF, share repurchases"),
    ("5", "Yahoo Finance — Key Statistics", "finance.yahoo.com/quote/NDSN/key-statistics/", "Valuation ratios, beta, shares, short interest, dividend data"),
    ("6", "Yahoo Finance — Analysis / Estimates", "finance.yahoo.com/quote/NDSN/analysis/", "Analyst estimates: EPS, revenue, revisions, growth rates"),
    ("7", "Yahoo Finance — Profile", "finance.yahoo.com/quote/NDSN/profile/", "Company description, employees, sector/industry, executives, SEC filings"),
    ("8", "CNBC — 10Y Treasury", "cnbc.com/quotes/US10Y", "Risk-free rate for WACC: 4.645% as of Aug 13, 2026"),
    ("9", "StockAnalysis", "stockanalysis.com/quote/NDSN/", "Attempted but returned 404 — Yahoo Finance used instead"),
    ("10", "Company Website", "nordson.com", "Official company information"),
]

for i, (a, b, c, d) in enumerate(sources):
    r = i + 3
    ws6.cell(row=r, column=1, value=a)
    ws6.cell(row=r, column=2, value=b)
    ws6.cell(row=r, column=3, value=c)
    ws6.cell(row=r, column=4, value=d)
    style_data(ws6, r, 4)
    if i == 0:
        style_header(ws6, r, 4)

ws6.column_dimensions["A"].width = 4
ws6.column_dimensions["B"].width = 35
ws6.column_dimensions["C"].width = 45
ws6.column_dimensions["D"].width = 35

# Save
filename = f"[{DSTAMP}] Nordson Corporation Model.xlsx"
filepath = f"/home/refcell/dev/capital/models/{filename}"
wb.save(filepath)
print(f"\nSaved: {filepath}")
print("DONE")
