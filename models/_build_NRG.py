"""Build 6-sheet valuation model for NRG Energy, Inc. (NRG).

Sources:
  - Yahoo Finance Income Statement / Balance Sheet / Cash Flow / Statistics / Analysis / Profile
  - StockAnalysis.com returned 404; Yahoo Finance is the primary source.
  - CNBC US10Y: 4.649% (2026-08-07)
  - Quote date: 2026-08-07, close $118.13

All $ figures in millions unless noted otherwise.

Key characteristics:
  - Utilities - Independent Power Producer (IPP), Houston TX
  - Revenue $33.1B TTM, +11% QoQ YoY; +93M annualized dividend, 1.60% yield
  - D&A-heavy: $1.054B TTM D&A buried in COGS; GAAP net income $782M on $33.1B revenue (2.4% margin)
  - High-leverage: Total debt $23.47B (Key Stats) vs $16.62B (BS FY25). Net debt ~$23.96B via EV-MC.
  - Massive capex cycle: -$8.265B investing CF TTM (TX data center gas plant, renewable build-out).
  - OCF $1.56B TTM but FCF only $316M TTM — FCF/Revenue = 0.96%, well below asset manager norms.
  - Interest expense $1.025B TTM on $2.08B op income = 49.3% ratio — rate-sensitive.
  - Primary lens: Forward P/E (High-Leverage / FCF-Insufficient pattern). FCF multiple framework invalid.
  - Analyst consensus: $8.82 EPS FY26, $11.16 EPS FY27
  - Beta 1.20, P/B 5.87, P/S 0.72, EV/EBITDA 13.03x
"""

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Paths
HERE = Path(__file__).resolve().parent
XLSX = HERE / "[2026-08-07] NRG Energy Model.xlsx"

# Raw inputs
PRICE = 118.13
SHARES_M = 210.21            # Yahoo Stats Aug 7, 2026 shares outstanding
MC_B = 24.69                # $B market cap
EV_B = 48.65                # $B enterprise value
NET_DEBT_M = (EV_B - MC_B) * 1000  # $23,960 = $23.96B — cleanest proxy avoids BS cash ambiguity
TOTAL_DEBT_KEYSTATS_M = 23470  # from Yahoo Key Stats mrq (Aug 7)
TOTAL_DEBT_BS_M = 16622       # from Yahoo Balance Sheet FY2025
CASH_BS_M = 4860             # end cash from BS FY2025
CASH_MRQ_M = 162             # total cash from Yahoo Key Stats mrq (Q2 FY26 = Jun 2026)
PREFERRED_STOCK_M = 650      # from Yahoo Balance Sheet FY2025 — CONSTANT since FY2023

# Income statement ($M, from Yahoo /financials/ Annual + TTM)
#                     TTM       FY2025    FY2024    FY2023    FY2022
REV_TTM = 33125
REV_F25 = 30713
REV_F24 = 28130
REV_F23 = 28823
REV_F22 = 31543

GP_TTM = 6226
GP_F25 = 5952
GP_F24 = 6030
GP_F23 = 2340
GP_F22 = 4100

OP_IN_TTM = 2080
OP_IN_F25 = 1944
OP_IN_F24 = 2282
OP_IN_F23 = -1049
OP_IN_F22 = 2224

NI_TTM = 782
NI_F25 = 797
NI_F24 = 1058
NI_F23 = -256
NI_F22 = 1221

EPS_D_TTM = 3.84
EPS_D_F25 = 4.01
EPS_D_F24 = 4.99
EPS_D_F23 = -1.12
EPS_D_F22 = 5.17

EBITDA_TTM = 3735
EBITDA_F25 = 3281
EBITDA_F24 = 3502
EBITDA_F23 = 1749
EBITDA_F22 = 2800

EBIT_TTM = 2073
EBIT_F25 = 1875

INTEREST_TTM = 1025
INTEREST_F25 = 741
INTEREST_F24 = 651
INTEREST_F23 = 667
INTEREST_F22 = 417

# Cash flow ($M, from Yahoo /cash-flow/)
#                    TTM       FY2025    FY2024    FY2023    FY2022
OCF_TTM = 1555
OCF_F25 = 1913
OCF_F24 = 2306
OCF_F23 = -221
OCF_F22 = 360

CAPEX_TTM = 1239
CAPEX_F25 = 1148
CAPEX_F24 = 490
CAPEX_F23 = 622
CAPEX_F22 = 373

INV_CF_TTM = -8265      # massive investing outflow TTM — data center gas plant, renewables
INV_CF_F25 = -1638

FCF_TTM = 316
FCF_F25 = 765
FCF_F24 = 1816
FCF_F23 = -843
FCF_F22 = -13

BUYBACK_TTM = 1753
BUYBACK_F25 = 1403

# Balance sheet ($M, FY2025)
TOTAL_ASSETS_M = 29140
TOTAL_LIAB_M = 27459
TOTAL_EQUITY_M = 1681
COMMON_EQUITY_M = 1031
NET_TANGIBLE_M = -5645

# Analyst estimates (Yahoo Finance /analysis/, Aug 7 2026)
REV_F26_C = 36640        # consensus FY2026 revenue $36.64B, 10 analysts
REV_F27_C = 38520        # consensus FY2027 revenue $38.52B, 11 analysts
EPS_F26_C = 8.82         # consensus FY2026 EPS, 10 analysts
EPS_F27_C = 11.16        # consensus FY2027 EPS, 11 analysts

# Earnings history (Normalized EPS)
# Q3 FY25: est $2.13, actual $2.78, +30.73%
# Q4 FY25: est $0.89, actual $1.04, +16.38%
# Q1 FY26: est $1.73, actual $1.49, -14.04%
# Q2 FY26: est $1.74, actual $1.49, -14.37%

# Tax rate, beta, rates
TAX_RATE = 0.190           # TTM: $199M / $1,048M pretax = 19.0%
BETA = 1.20                # 5Y monthly beta (Yahoo Stats)
RISK_FREE = 0.04649        # 10Y US Treasury, CNBC Aug 7, 2026
ERP = 0.05                 # equity risk premium

# Computed
COE = RISK_FREE + BETA * ERP  # CAPM cost of equity = 0.04649 + 1.20*0.05 = 0.10649 = 10.65%
COST_OF_DEBT = RISK_FREE + 0.035  # ~350bps spread over RF for utility-grade debt = ~8.15%

# For WACC weights, use EV components
TOTAL_CAP_M = EV_B * 1000  # enterprise value = $48,650M
EQUITY_WEIGHT = MC_B / EV_B  # equity weight in capital structure
DEBT_WEIGHT = 1 - EQUITY_WEIGHT  # debt weight
WACC = EQUITY_WEIGHT * COE + DEBT_WEIGHT * COST_OF_DEBT * (1 - TAX_RATE)

# Valuation multiples
PE_TRAIL_TTM = PRICE / EPS_D_TTM            # ~30.76x
PE_FWD_F26 = PRICE / EPS_F26_C              # ~13.40x
PE_FWD_F27 = PRICE / EPS_F27_C              # ~10.05x
PS = MC_B / (REV_TTM / 1000)                # ~0.75x
PB = MC_B / (COMMON_EQUITY_M / 1000)        # ~24.0x on common equity
PB_TOTAL = MC_B / (TOTAL_EQUITY_M / 1000)   # ~14.7x on total equity
EV_REVENUE = EV_B / (REV_TTM / 1000)        # ~1.47x
EV_EBITDA = EV_B / (EBITDA_TTM / 1000)      # ~13.03x
P_FCF = MC_B / (FCF_TTM / 1000)             # ~78.1x — meaningless with $316M FCF
EV_FCF = EV_B / (FCF_TTM / 1000)            # ~154x — N/A (high-leverage FCF insufficient)

# Interest coverage
INTEREST_COVERAGE = EBIT_TTM / INTEREST_TTM  # ~2.02x — manageable but thin buffer

# D&A
DA_TTM = 1054  # TTM reconciled depreciation ($M)
DA_F25 = 896
DA_F24 = 1071
DA_F23 = 1174
DA_F22 = 688

# FFO (= NI + D&A) — useful cross-check for D&A-heavy utilities
FFO_TTM = NI_TTM + DA_TTM  # $1,836M
FFO_PER_SHARE_TTM = FFO_TTM / SHARES_M  # ~$8.57
P_FFO = PRICE / FFO_PER_SHARE_TTM  # ~13.79x P/FFO

# Net debt / FCF check
NET_DEBT_RATIO_TO_FCF = NET_DEBT_M / FCF_TTM if FCF_TTM > 0 else float('inf')
# = 23960 / 316 = 75.8x — FCF framework INVALID

# Scenario framework: Forward P/E (Primary) due to FCF insufficiency
# EBITDA-based cross-check only

# Check: Is EV/EBITDA bear target plausible?
# Bear EBITDA at 7.5x → implied EV = 7.5 * 3,500 = ~$26.25B → equity = $26.25B - $24.0B = $2.25B
# At 210M shares: $10.7/share — plausible as lower bound
# OK for cross-check.

# Scenario assumptions (Forward P/E framework)
# Revenue growth: FY26 consensus $36.64B = +11.3% from TTM $33.13B, FY27 $38.52B = +5.1% → mature utility growth
# 5Y CAGR decelerates from there

scenarios = {
    'bear': {
        'rev_cagr_5y': 0.02,      # 2% — slow utility growth; bear = stagnation, capex drag
        'terminal_rev_5y': REV_F27_C * (1.02)**3,  # ~$40.8B
        'op_margin': 0.06,         # 6% — rate risk, integration costs, margin compression
        'terminal_FCF_margin': 0.02,
        'exit_PE': 11.0,          # compressed to regulated-utility norms
    },
    'base': {
        'rev_cagr_5y': 0.035,     # 3.5% — utilities sector average; data center tailwind moderate
        'terminal_rev_5y': REV_F27_C * (1.035)**3,  # ~$42.8B
        'op_margin': 0.08,        # 8% — normalization after capex cycle ends; FY24-like
        'terminal_FCF_margin': 0.04,
        'exit_PE': 14.0,          # utility-growth premium; analyst consensus near
    },
    'bull': {
        'rev_cagr_5y': 0.05,      # 5% — data center tailwind + renewable build-out accelerates
        'terminal_rev_5y': REV_F27_C * (1.05)**3,  # ~$44.9B
        'op_margin': 0.10,        # 10% — capex cycle ends, OCF normalizes
        'terminal_FCF_margin': 0.06,
        'exit_PE': 17.0,          # premium IPP; Vistra trades high 15-18x fwd
    }
}

# Weights
weights = {'bear': 0.20, 'base': 0.55, 'bull': 0.25}

# Style helpers
title_font = Font(name='Calibri', bold=True, size=16)
subtitle_font = Font(name='Calibri', bold=True, size=11, color='666666')
header_font = Font(name='Calibri', bold=True, size=10)
data_font = Font(name='Calibri', size=10)
bold_font = Font(name='Calibri', bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

def ws_title(ws, text):
    """Write merged title at row 1."""
    ws.merge_cells('A1:F1')
    ws['A1'] = text
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    for col in range(1, 7):
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

def ws_subtitle(ws, row, text):
    """Write subtitle at given row."""
    c = ws.cell(row=row, column=1, value=text)
    c.font = subtitle_font

def write_table(ws, headers, data, start_row=3):
    """Write header row + data rows with borders and formatting."""
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    for ri, row_data in enumerate(data, start_row + 1):
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = data_font
            c.border = thin_border
            if ci == 1:
                c.font = bold_font
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18


# Build workbook
wb = openpyxl.Workbook()

# Sheet 1: Valuation
ws1 = wb.active
ws1.title = 'Valuation'
ws_title(ws1, 'NRG Energy, Inc. (NYSE: NRG)')
ws_subtitle(ws1, 2, 'Valuation Model — Data as of $118.13 on 2026-08-07 (close)')

val_headers = ['Field', 'Value', 'Comment']
val_data = [
    ['Company', 'NRG Energy, Inc.', 'Utilities - Independent Power Producers'],
    ['Ticker', 'NYSE: NRG', ''],
    ['Date', '2026-08-07', ''],
    ['Close Price', f'${PRICE}', 'After hours $118.70 (+0.48%)'],
    ['Shares Outstanding', f'{SHARES_M}M', 'Yahoo Stats Aug 7, 2026'],
    ['Market Cap', f'${MC_B:.2f}B', ''],
    ['Enterprise Value', f'${EV_B:.2f}B', 'Net debt ~$23.96B via EV-MC proxy'],
    ['Primary Valuation Lens', 'Forward P/E', 'High-Leverage FCF-Insufficient pattern: net debt/'
                                              + f'FCF = {NET_DEBT_RATIO_TO_FCF:.0f}x — FCF framework invalid'],
    ['Stance', 'Watch', 'Capex cycle trough; data-center power catalyst; debt trajectory is primary risk'],
    ['', '', ''],
    ['Trailing P/E (TTM)', f'{PE_TRAIL_TTM:.2f}x', f'On dil EPS ${EPS_D_TTM:.2f} — distorted by D&A'],
    ['Fwd P/E FY26', f'{PE_FWD_F26:.2f}x', f'On consensus EPS ${EPS_F26_C:.2f} (10 analysts)'],
    ['Fwd P/E FY27', f'{PE_FWD_F27:.2f}x', f'On consensus EPS ${EPS_F27_C:.2f} (11 analysts)'],
    ['P/S', f'{PS:.2f}x', f'On TTM revenue ${REV_TTM/1000:.1f}B'],
    ['P/B (Common Equity)', f'{PB:.0f}x', f'Common equity ${COMMON_EQUITY_M}M FY25'],
    ['P/B (Total Equity)', f'{PB_TOTAL:.1f}x', f'Total equity ${TOTAL_EQUITY_M}M FY25'],
    ['P/FFO', f'{P_FFO:.1f}x', f'FFO = NI+DA = ${FFO_TTM}M; FFO/Share = ${FFO_PER_SHARE_TTM:.2f}'],
    ['EV/Sales', f'{EV_REVENUE:.2f}x', f''],
    ['EV/EBITDA', f'{EV_EBITDA:.2f}x', f'EBITDA ${EBITDA_TTM/1000:.1f}B TTM'],
    ['EV/FCF', 'N/A', f'FCF only ${FCF_TTM}M — {NET_DEBT_RATIO_TO_FCF:.0f}x net debt/FCF; framework invalid'],
    ['P/FCF', 'N/A', f'${MC_B:.2f}B / $0.316B = ~78x — meaningless'],
    ['D&A (TTM)', f'${DA_TTM}M', f'{DA_TTM/REV_TTM*100:.1f}% of revenue — D&A distortion significant'],
    ['Interest Coverage (EBIT/Int)', f'{INTEREST_COVERAGE:.2f}x', f'EBIT ${EBIT_TTM}M / Int ${INTEREST_TTM}M'],
    ['Beta (5Y Monthly)', f'{BETA}', 'Yahoo Stats'],
    ['52W High', '$189.96', f'Stock down {((189.96-PRICE)/189.96)*100:.0f}% from peak'],
    ['52W Low', '$112.50', f'Stock at {((PRICE-112.50)/112.50)*100:.0f}% above absolute low'],
    ['Dividend Yield', '1.60%', 'Forward annual $1.90; payout ratio 47.66%'],
    ['Preferred Stock', '$650M', 'Constant since FY2023; must flag in Questions'],
]
write_table(ws1, val_headers, val_data, 3)

# Sheet 2: WACC
ws2 = wb.create_sheet('WACC')
ws_title(ws2, 'WACC Calculation — CAPM')
ws_subtitle(ws2, 2, 'Components as of 2026-08-07')

wacc_data = [
    ['Risk-Free Rate (10Y US)', f'{RISK_FREE*100:.3f}%', 'CNBC US10Y, Aug 7 2026: 4.649%'],
    ['Equity Risk Premium', '5.00%', 'Assumed'],
    ['Beta (Levered, 5Y Monthly)', f'{BETA}', 'Yahoo Stats'],
    ['Cost of Equity (CAPM)', f'{COE*100:.2f}%', f'{RISK_FREE:.4f} + {BETA} x 0.05'],
    ['Cost of Debt (pre-tax)', f'{COST_OF_DEBT*100:.2f}%', 'RF + 350bps utility-grade spread'],
    ['Tax Rate', f'{TAX_RATE*100:.1f}%', f'TTM effective: ${199}M / ${1048}M'],
    ['Market Cap', f'${MC_B:.2f}B', ''],
    ['Total Debt (Key Stats)', f'${TOTAL_DEBT_KEYSTATS_M/1000:.2f}B', 'Yahoo Key Stats Q2 FY26'],
    ['Total Debt (BS FY25)', f'${TOTAL_DEBT_BS_M/1000:.2f}B', 'Yahoo BS FY2025 — $6.85B lower'],
    ['Net Debt (EV-MC proxy)', f'${NET_DEBT_M/1000:.2f}B', 'EV - MC = $48.65B - $24.69B'],
    ['Equity Weight', f'{EQUITY_WEIGHT*100:.2f}%', f'MC / EV = {MC_B} / {EV_B}'],
    ['Debt Weight', f'{DEBT_WEIGHT*100:.2f}%', f'1 - equity weight'],
    ['WACC', f'{WACC*100:.2f}%', ''],
]
write_table(ws2, ['Component', 'Value', 'Note'], wacc_data, 3)

print(f"WACC: {WACC*100:.2f}%")

# Sheet 3: Scenarios (Forward P/E framework, EV/EBITDA cross-check)
ws3 = wb.create_sheet('Scenarios')
ws_title(ws3, 'Scenario Analysis — Forward P/E Framework')
ws_subtitle(ws3, 2, 'High-Leverage / FCF-Insufficient pattern: net debt/FCF = ~76x → FCF multiples invalid. Primary lens = Forward P/E.')

# Compute scenario targets
# Terminal EPS derived from: terminal revenue × terminal op margin × (1-tax) / shares
# Revenue 5Y from FY27 = FY27 rev × (1+cagr)^3

bear_term_rev = scenarios['bear']['terminal_rev_5y']
base_term_rev = scenarios['base']['terminal_rev_5y']
bull_term_rev = scenarios['bull']['terminal_rev_5y']

def calc_target(term_rev_M, op_margin, tax_rate, exit_pe, shares_M):
    term_op_in = term_rev_M * op_margin
    term_ni = term_op_in * (1 - tax_rate)
    term_eps = term_ni / shares_M
    target = term_eps * exit_pe
    return term_eps, target, term_op_in

bear_eps, bear_target, bear_opin = calc_target(
    bear_term_rev, scenarios['bear']['op_margin'], TAX_RATE, scenarios['bear']['exit_PE'], SHARES_M)
base_eps, base_target, base_opin = calc_target(
    base_term_rev, scenarios['base']['op_margin'], TAX_RATE, scenarios['base']['exit_PE'], SHARES_M)
bull_eps, bull_target, bull_opin = calc_target(
    bull_term_rev, scenarios['bull']['op_margin'], TAX_RATE, scenarios['bull']['exit_PE'], SHARES_M)

# EBITDA-based cross-check
# Bear: EBITDA ~3,500M, exit 7.5x → $26.25B EV → equity = $26.25B - $24B = $2.25B → ~$10.7/share
# Base: EBITDA ~4,000M, exit 10x → $40B EV → equity = $40B - $24B = $16B → ~$76/share
# Bull: EBITDA ~4,500M, exit 12x → $54B EV → equity = $54B - $24B = $30B → ~$143/share
bear_ev_ebitda = (bear_term_rev * 0.11) * 7.5  # ~11% EBITDA margin
base_ev_ebitda = (base_term_rev * 0.12) * 10.0
bull_ev_ebitda = (bull_term_rev * 0.14) * 12.0

# Weights and FV
bear_w, base_w, bull_w = weights['bear'], weights['base'], weights['bull']
weighted_fv = bear_w * bear_target + base_w * base_target + bull_w * bull_target

bear_upside = (bear_target - PRICE) / PRICE * 100
base_upside = (base_target - PRICE) / PRICE * 100
bull_upside = (bull_target - PRICE) / PRICE * 100
weighted_upside = (weighted_fv - PRICE) / PRICE * 100

scen_headers = ['Scenario Driver', 'Bear', 'Base', 'Bull', 'Note']
scen_data = [
    ['Revenue CAGR (5Y)', f'{scenarios["bear"]["rev_cagr_5y"]*100:.0f}%',
     f'{scenarios["base"]["rev_cagr_5y"]*100:.0f}%',
     f'{scenarios["bull"]["rev_cagr_5y"]*100:.0f}%', 'From FY27 baseline $38.52B'],
    ['Terminal Revenue (5Y, $MM)', f'{bear_term_rev:.0f}',
     f'{base_term_rev:.0f}', f'{bull_term_rev:.0f}', ''],
    ['Op Margin', f'{scenarios["bear"]["op_margin"]*100:.1f}%',
     f'{scenarios["base"]["op_margin"]*100:.1f}%',
     f'{scenarios["bull"]["op_margin"]*100:.1f}%', ''],
    ['Terminal Operating Income ($MM)', f'{bear_opin:.0f}',
     f'{base_opin:.0f}', f'{bull_opin:.0f}', ''],
    ['Terminal EPS (5Y)', f'${bear_eps:.2f}', f'${base_eps:.2f}', f'${bull_eps:.2f}', ''],
    ['Exit P/E', f'{scenarios["bear"]["exit_PE"]:.0f}x',
     f'{scenarios["base"]["exit_PE"]:.0f}x',
     f'{scenarios["bull"]["exit_PE"]:.0f}x', ''],
    ['Target Price', f'${bear_target:.2f}', f'${base_target:.2f}', f'${bull_target:.2f}', ''],
    ['Upside %', f'{bear_upside:.1f}%', f'{base_upside:.1f}%', f'{bull_upside:.1f}%',
     f'From ${PRICE}'],
    ['Weight', f'{bear_w*100:.0f}%', f'{base_w*100:.0f}%', f'{bull_w*100:.0f}%', ''],
    ['Weighted Value/Share', f'${bear_w*bear_target:.2f}',
     f'${base_w*base_target:.2f}', f'${bull_w*bull_target:.2f}', ''],
    ['', '', '', '', ''],
    ['Probability-Weighted FV', '', '', f'${weighted_fv:.2f}', 'Sum of weighted'],
    ['Current Price', '', '', f'${PRICE}', ''],
    ['Weighted Upside', '', '', f'{weighted_upside:.1f}%', ''],
    ['', '', '', '', ''],
    ['Framework Note', '', '', '', ''],
    ['Primary Lens', '', '', '', 'Forward P/E (high-leverage / FCF-insufficient)'],
    ['FCF Framework', '', '', '', 'INVALID — net debt/FCF = ~76x'],
    ['', '', '', '', f'Net debt ${NET_DEBT_M/1000:.1f}B / FCF ${FCF_TTM}M'],
    ['D&A Year TTM', f'${DA_TTM}M', '', '',
     'GAAP net income structurally suppressed by D&A'],
    ['EV/EBITDA Cross-Check', '', '', '', ''],
    ['EV/EBITDA Implied EV ($MM)', f'{bear_ev_ebitda:.0f}',
     f'{base_ev_ebitda:.0f}', f'{bull_ev_ebitda:.0f}',
     'Bear: 7.5x, Base: 10x, Bull: 12x'],
    ['EV/EBITDA Implied Equity ($MM)', f'{(bear_ev_ebitda-NET_DEBT_M):.0f}',
     f'{(base_ev_ebitda-NET_DEBT_M):.0f}',
     f'{(bull_ev_ebitda-NET_DEBT_M):.0f}',
     f'Less net debt ${NET_DEBT_M:.0f}M'],
    ['EV/EBITDA Price/Share Cross-Check', f'${(bear_ev_ebitda-NET_DEBT_M)/SHARES_M:.2f}',
     f'${(base_ev_ebitda-NET_DEBT_M)/SHARES_M:.2f}',
     f'${(bull_ev_ebitda-NET_DEBT_M)/SHARES_M:.2f}', ''],
]

write_table(ws3, scen_headers, scen_data, 3)

# Sheet 4: Actuals Source Audit
ws4 = wb.create_sheet('Actuals Source Audit')
ws_title(ws4, 'Actuals Source Audit')
ws_subtitle(ws4, 2, 'Every data point with source URL, date, and verification notes')

audit_headers = ['Data Point', 'Value', 'Source URL', 'Date', 'Notes']
audit_data = [
    ['Stock Price', f'${PRICE}', 'Yahoo Finance /quote/NRG/', '2026-08-07', 'Close price'],
    ['After Hours', '$118.70', 'Yahoo Finance /quote/NRG/', '2026-08-07', '+0.48%'],
    ['Market Cap', f'${MC_B:.2f}B', 'Yahoo Finance Statistics', '2026-08-07', ''],
    ['Enterprise Value', f'${EV_B:.2f}B', 'Yahoo Finance Statistics', '2026-08-07', ''],
    ['Shares Outstanding (Key Stats)', f'{SHARES_M}M', 'Yahoo Finance Key Stats', '2026-08-07', 'Implied shares outstanding'],
    ['Preferred Stock Equity', '$650M', 'Yahoo Finance Balance Sheet', 'FY2025', 'Constant since FY2023'],
    ['52W High', '$189.96', 'Yahoo Finance Statistics', '2026-08-07', ''],
    ['52W Low', '$112.50', 'Yahoo Finance Statistics', '2026-08-07', ''],
    ['Beta (5Y Month)', f'{BETA}', 'Yahoo Finance Key Stats', '2026-08-07', ''],
    ['Dividend Rate Forward', '$1.90', 'Yahoo Finance Key Stats', '2026-08-07', '1.60% yield'],
    ['Next Earnings', 'Nov 5, 2026 7:30 AM EST', 'Yahoo Finance /profile/', '2026-08-07', ''],
    ['TTM Revenue', f'${REV_TTM}M', 'Yahoo Finance /financials/', 'TTM', 'All numbers in $000s → M'],
    ['FY25 Revenue', f'${REV_F25}M', 'Yahoo Finance /financials/', 'FY2025', ''],
    ['FY24 Revenue', f'${REV_F24}M', 'Yahoo Finance /financials/', 'FY2024', ''],
    ['FY23 Revenue', f'${REV_F23}M', 'Yahoo Finance /financials/', 'FY2023', ''],
    ['FY22 Revenue', f'${REV_F22}M', 'Yahoo Finance /financials/', 'FY2022', ''],
    ['TTM Gross Profit', f'${GP_TTM}M', 'Yahoo Finance /financials/', 'TTM', ''],
    ['TTM Operating Income', f'${OP_IN_TTM}M', 'Yahoo Finance /financials/', 'TTM', ''],
    ['TTM Net Income', f'${NI_TTM}M', 'Yahoo Finance /financials/', 'TTM', ''],
    ['TTM EPS (Dil)', f'${EPS_D_TTM:.2f}', 'Yahoo Finance /financials/', 'TTM', ''],
    ['TTM EBITDA', f'${EBITDA_TTM}M', 'Yahoo Finance /financials/', 'TTM', ''],
    ['TTM EBIT', f'${EBIT_TTM}M', 'Yahoo Finance /financials/', 'TTM', ''],
    ['TTM Interest Expense', f'${INTEREST_TTM}M', 'Yahoo Finance /financials/', 'TTM', ''],
    ['TTM Operating CF', f'${OCF_TTM}M', 'Yahoo Finance /cash-flow/', 'TTM', 'Cash operating margin 4.7%'],
    ['TTM CapEx', f'${CAPEX_TTM}M', 'Yahoo Finance /cash-flow/', 'TTM', 'Down from $1,148M FY25'],
    ['TTM Investing CF', f'${INV_CF_TTM}M', 'Yahoo Finance /cash-flow/', 'TTM',
     'MASSIVE: data-center gas plant, renewables — capex cycle'],
    ['TTM D&A', f'${DA_TTM}M', 'Yahoo Finance /financials/', 'TTM', 'Depreciation distortion in COGS'],
    ['TTM Free Cash Flow', f'${FCF_TTM}M', 'Yahoo Finance /cash-flow/', 'TTM', 'Only 0.96% of revenue'],
    ['Total Debt (Key Stats)', f'${TOTAL_DEBT_KEYSTATS_M}M', 'Yahoo Finance Key Stats', 'Q2 FY26', ''],
    ['Total Debt (BS FY25)', f'${TOTAL_DEBT_BS_M}M', 'Yahoo Finance Balance Sheet', 'FY2025',
     '$6.85B LOWER than Key Stats'],
    ['Total Cash (Key Stats)', f'${CASH_MRQ_M}M', 'Yahoo Finance Key Stats', 'Q2 FY26', ''],
    ['Total Cash (BS FY25 End)', f'${CASH_BS_M}M', 'Yahoo Finance /cash-flow/', 'FY2025', 'End cash position'],
    ['Net Debt (EV-MC proxy)', f'${NET_DEBT_M:.0f}M', 'EV - MC calculation', '2026-08-07',
     'Preferred over BS cash due to debt discrepancy'],
    ['TTM Repurchase', f'${BUYBACK_TTM}M', 'Yahoo Finance /cash-flow/', 'TTM', ''],
    ['EPS F26 Consensus', f'${EPS_F26_C}', 'Yahoo Finance /analysis/', '2026-08-07', '10 analysts'],
    ['EPS F27 Consensus', f'${EPS_F27_C}', 'Yahoo Finance /analysis/', '2026-08-07', '11 analysts'],
    ['Rev F26 Consensus', f'${REV_F26_C}M', 'Yahoo Finance /analysis/', '2026-08-07', '10 analysts'],
    ['Rev F27 Consensus', f'${REV_F27_C}M', 'Yahoo Finance /analysis/', '2026-08-07', '11 analysts'],
    ['Q1 FY26 EPS Actual', '$1.49', 'Yahoo Finance /analysis/', 'Mar 2026', 'Miss est $1.73 by -14.04%'],
    ['Q2 FY26 EPS Actual', '$1.49', 'Yahoo Finance /analysis/', 'Jun 2026', 'Miss est $1.74 by -14.37%'],
    ['Trailing P/E', f'{PE_TRAIL_TTM:.2f}x', 'Yahoo Finance Statistics', '2026-08-07', 'Distorted by D&A'],
    ['Forward P/E (Key Stats)', '12.48x', 'Yahoo Finance Statistics', '2026-08-07', 'Close to our calc'],
    ['P/FCF (Key Stats)', 'P/FCF ~78x', 'Calculated', '2026-08-07', 'N/A — FCF framework invalid'],
    ['EBITDA (S&P Calc)', f'${EBITDA_TTM}M', 'Yahoo Finance Statistics', 'Calculated by S&P Global', ''],
    ['10Y US Treasury', f'{RISK_FREE*100:.3f}%', 'CNBC /quotes/US10Y', '2026-08-07', '4.649%'],
]
write_table(ws4, audit_headers, audit_data, 3)

# Sheet 5: Questions
ws5 = wb.create_sheet('Questions')
ws_title(ws5, 'Open Questions')
ws_subtitle(ws5, 2, 'Issues requiring follow-up or monitoring')

quest_headers = ['#', 'Question', 'Category', 'Priority']
quest_data = [
    [1, 'Key Stats debt $23.47B vs BS debt $16.62B — $6.85B discrepancy: What accounts for the difference? Does Key Stats include capital lease obligations, convertible debt, or subordinated notes classified outside permanent debt on the statutory BS?', 'Debt', 'High'],
    [2, 'Investing CF of -$8.27B TTM vs capex of -$1.24B: The $7B spread is acquisitions, asset purchases, or other investing outflows. What major acquisitions or build-outs drove this? Is the 1.2-GW Texas gas plant (mentioned in Aug 2026 news) accounting for the bulk?', 'Capex Cycle', 'High'],
    [3, 'Preferred stock of $650M is constant since FY2023: What are the terms? What is the dividend obligation? At what rate does it pay? Should it be subtracted from MC for common equity value?', 'Preferred Stock', 'High'],
    [4, 'FCF of only $316M on $33.1B TTM revenue (0.96% margin) while OCF is $1.56B: Is the negative FCF purely capex-cycle-driven (data center power infrastructure build) or are there structural issues? When does the cycle end?', 'FCF Insufficiency', 'High'],
    [5, 'Interest coverage at 2.02x (EBIT/Int) is thin for an IPP. What is the debt maturity schedule? How much debt comes due in 2027-2030? What is the fixed vs. floating rate mix?', 'Debt Maturity', 'High'],
    [6, 'Operating margin TTM of 6.3% ($2.08B/$33.1B) — this is lower than FY24 (8.1% on $28.1B revenue). The revenue increase is NOT translating to proportional OpEx compression. Is this a transition year due to integration/acquisition costs?', 'Operating Margin', 'Medium'],
    [7, 'GAAP net income $782M on revenue of $33.1B — the 2.4% net margin is structurally distorted by $1.054B in D&A flowing through COGS. What is the FFO and FFO margin? Is FFO a better normalization metric?', 'Earnings Quality', 'High'],
    [8, 'Q1 and Q2 FY26 both missed EPS estimates by ~14% after 2 quarters of beats. Has guidance been lowered? What is the next earnings date (Nov 5, 2026)? Will Q3 also miss?', 'Earnings', 'High'],
    [9, 'EPS revisions trending DOWN over 30 days: FY26 EPS $9.14→$8.82, FY27 EPS $11.80→$11.16. Five analysts cut estimates. What is driving the revision trend?', 'Analyst Revisions', 'Medium'],
    [10, 'Buyback of $1.753B TTM: Are buybacks deployed at attractive prices? Current price of $118 is near 52-week low ($112.50) so the buyback may be at favorable valuation. Buyback/OCF ratio = $1.753B / $1.555B = 113% — EXCEEDS operating cash generation, meaning buybacks require debt issuance.', 'Capital Allocation', 'High'],
    [11, 'Revenue jumped $32 billion in FY25 (from $28.1B to $30.7B) but then OCF dropped from $2.3B (FY24) to $1.91B (FY25) to $1.555B TTM. Is the revenue growth higher-margin or lower-margin business?', 'Revenue Quality', 'Medium'],
    [12, 'Data center power demand: NRG is positioned to benefit from AI/data center power demand in Texas. Has management specifically quantified the TAM or announced specific data center power contracts?', 'Growth Catalyst', 'Medium'],
    [13, 'Working capital swing from $151M (FY24) to $5.12B (FY25): What drove this $5B swing? Is it temporary or structural?', 'Working Capital', 'Medium'],
    [14, 'Tax rate TTM 19.0%: Consistent with FY25 (23.8% on pretax) but lower than FY24 (22.3%). Geographic mix? R&D credits? International exposures?', 'Tax', 'Low'],
    [15, 'Employee count: 16,702 for an IPP. Is this labor-intensive? What is the FTE trajectory vs. revenue growth?', 'Efficiency', 'Low'],
    [16, 'Net tangible assets are -$5.6B: Does NRG have significant goodwill or intangibles from acquisitions? What is the total goodwill and amortization schedule?', 'Goodwill', 'Medium'],
    [17, 'Stock down 38% from 52W high ($189.96) to current $118.13, and only 5% above 52W low ($112.50): Near absolute bottom. Is this a capitulation or a justified repricing given the capex cycle?', 'Valuation', 'Medium'],
    [18, 'Dividend: Is it covered by FFO? FFO/Share = $8.57, annualized dividend = $1.90, so coverage = 44.8%. But is FFO the right denominator for coverage?', 'Dividend', 'Low'],
]
write_table(ws5, quest_headers, quest_data, 3)
ws5.column_dimensions[get_column_letter(2)].width = 55

# Sheet 6: Sources
ws6 = wb.create_sheet('Sources')
ws_title(ws6, 'Data Sources')
ws_subtitle(ws6, 2, 'All data accessed on 2026-08-07')

src_headers = ['#', 'Source', 'URL', 'Content']
src_data = [
    [1, 'Yahoo Finance — Quote/Summary', 'https://finance.yahoo.com/quote/NRG/', 'Price, MC, EV shares'],
    [2, 'Yahoo Finance — Income Statement', 'https://finance.yahoo.com/quote/NRG/financials/', 'Revenue, GP, OpIn, NI, EPS, EBITDA'],
    [3, 'Yahoo Finance — Balance Sheet', 'https://finance.yahoo.com/quote/NRG/balance-sheet/', 'Assets, Liabilities, Debt, Equity, Preferred'],
    [4, 'Yahoo Finance — Cash Flow', 'https://finance.yahoo.com/quote/NRG/cash-flow/', 'OCF, Inv CF, Fin CF, FCF, CapEx, Buybacks'],
    [5, 'Yahoo Finance — Key Statistics', 'https://finance.yahoo.com/quote/NRG/key-statistics/', 'Shares, MC, EV, P/E, P/S, PB, Beta, 52W range, total debt'],
    [6, 'Yahoo Finance — Analysis', 'https://finance.yahoo.com/quote/NRG/analysis/', 'Analyst estimates, EPS, revisions, earnings history'],
    [7, 'Yahoo Finance — Profile', 'https://finance.yahoo.com/quote/NRG/profile/', 'Sector, industry, employees, next earnings, news'],
    [8, 'CNBC — US10Y', 'https://www.cnbc.com/quotes/US10Y', '10Y Treasury yield: 4.649%'],
    [9, 'StockAnalysis.com', 'https://stockanalysis.com/quote/NRG/', 'Returned 404 — not available for this ticker'],
]
write_table(ws6, src_headers, src_data, 3)
ws6.column_dimensions[get_column_letter(3)].width = 45
ws6.column_dimensions[get_column_letter(4)].width = 35

# Save
wb.save(str(XLSX))
print(f"Saved: {XLSX}")

# Verification
print(f"\n=== VERIFICATION ===")
print(f"WACC: {WACC*100:.4f}%")
print(f"COE: {COE*100:.2f}%; Cost of Debt: {COST_OF_DEBT*100:.2f}%")
print(f"Equity Weight: {EQUITY_WEIGHT*100:.1f}%; Debt Weight: {DEBT_WEIGHT*100:.1f}%")
print(f"\nNet Debt / FCF = {NET_DEBT_RATIO_TO_FCF:.0f}x — FCF framework INVALID")
print(f"Net debt: ${NET_DEBT_M/1000:.1f}B; FCF: ${FCF_TTM}M")
print(f"\nScenario Targets (Forward P/E):")
print(f"  Bear:  ${bear_target:.2f} (EPS ${bear_eps:.2f}, {scenarios['bear']['exit_PE']}x P/E) — {bear_upside:.1f}% upside")
print(f"  Base:  ${base_target:.2f} (EPS ${base_eps:.2f}, {scenarios['base']['exit_PE']}x P/E) — {base_upside:.1f}% upside")
print(f"  Bull:  ${bull_target:.2f} (EPS ${bull_eps:.2f}, {scenarios['bull']['exit_PE']}x P/E) — {bull_upside:.1f}% upside")
print(f"\nProbability-Weighted FV: ${weighted_fv:.2f} — {weighted_upside:.1f}% upside from ${PRICE}")
print(f"\nEV/EBITDA Cross-Check targets:")
print(f"  Bear:  ${(bear_ev_ebitda-NET_DEBT_M)/SHARES_M:.2f}/share (implied EV ${bear_ev_ebitda/1000:.1f}B - ${NET_DEBT_M/1000:.1f}B net debt)")
print(f"  Base:  ${(base_ev_ebitda-NET_DEBT_M)/SHARES_M:.2f}/share (implied EV ${base_ev_ebitda/1000:.1f}B - ${NET_DEBT_M/1000:.1f}B net debt)")
print(f"  Bull:  ${(bull_ev_ebitda-NET_DEBT_M)/SHARES_M:.2f}/share (implied EV ${bull_ev_ebitda/1000:.1f}B - ${NET_DEBT_M/1000:.1f}B net debt)")
print(f"\nD&A distortion: D&A/Revenue TTM = {DA_TTM/REV_TTM*100:.1f}%")
print(f"FFO TTM: ${FFO_TTM}M; FFO/Share: ${FFO_PER_SHARE_TTM:.2f}; P/FFO: {P_FFO:.1f}x")
print(f"Buyback/OCF ratio: ${BUYBACK_TTM}M / ${OCF_TTM}M = {BUYBACK_TTM/OCF_TTM*100:.0f}%")
print(f"Interest coverage (EBIT/Int): {INTEREST_COVERAGE:.2f}x — thin but serviceable")
print(f"\nAnalyst PT sanity check: Base case ${base_target:.2f} vs implied from FY26 consensus ${EPS_F26_C} * 15x = ${EPS_F26_C*15:.2f}, FY27 ${EPS_F27_C} * 15x = ${EPS_F27_C*15:.2f}")
