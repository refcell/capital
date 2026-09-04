#!/usr/bin/env python3
"""Build the Atlassian six-sheet valuation workbook."""
from math import isclose
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT = Path(__file__).with_name("2026-09-04 Atlassian Model.xlsx")
PRICE = 194.68
SHARES = 253.14  # millions
MARKET_CAP = 49_280.0  # USD millions
ENTERPRISE_VALUE = 49_270.0  # USD millions
NET_CASH = 6.6  # USD millions
FY27_REVENUE = 7_480.0  # USD millions, public consensus
TTM_FCF = 1_319.0  # USD millions

SCENARIOS = {
    "Bear": {"growth": 0.07, "margin": 0.16, "multiple": 18.0, "net_cash": 500.0, "shares": 245.0, "discount": 0.115, "weight": 0.20},
    "Base": {"growth": 0.12, "margin": 0.20, "multiple": 24.0, "net_cash": 1_500.0, "shares": 235.0, "discount": 0.101, "weight": 0.55},
    "Bull": {"growth": 0.16, "margin": 0.23, "multiple": 30.0, "net_cash": 2_500.0, "shares": 225.0, "discount": 0.095, "weight": 0.25},
}
for case in SCENARIOS.values():
    case["terminal_revenue"] = FY27_REVENUE * (1 + case["growth"]) ** 5
    case["terminal_fcf"] = case["terminal_revenue"] * case["margin"]
    case["implied_ev"] = case["terminal_fcf"] * case["multiple"]
    case["terminal_price"] = (case["implied_ev"] + case["net_cash"]) / case["shares"]
    case["target"] = case["terminal_price"] / (1 + case["discount"]) ** 5
    case["upside"] = case["target"] / PRICE - 1
weighted_value = sum(case["target"] * case["weight"] for case in SCENARIOS.values())

assert SCENARIOS["Bear"]["target"] < PRICE
assert abs(SCENARIOS["Base"]["target"] - 192.52) / 192.52 < 0.20
assert 50 < weighted_value < 400

rf = 0.0478
erp = 0.05
beta = 1.16
cost_equity = rf + beta * erp
pretax_debt_cost = 0.05432 / 1.233
tax_rate = 0.21
equity_weight = MARKET_CAP / (MARKET_CAP + 1_233.0)
debt_weight = 1 - equity_weight
wacc = equity_weight * cost_equity + debt_weight * pretax_debt_cost * (1 - tax_rate)

wb = Workbook()
navy, blue, gold, gray, white = "17365D", "D9EAF7", "D8B34B", "E7E6E6", "FFFFFF"
thin = Side(style="thin", color="A6A6A6")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def put(ws, row, col, value, *, bold=False, fill=None, color="000000", size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Aptos", size=size, bold=bold, color=color)
    cell.fill = PatternFill("solid", fgColor=fill) if fill else PatternFill(fill_type=None)
    cell.border = border
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    return cell


def title(ws, text, end_col=5, subtitle=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    put(ws, 1, 1, text, bold=True, fill=navy, color=white, size=15)
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        put(ws, 2, 1, subtitle, bold=True, fill=blue)


def header(ws, row, values):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=True, fill=navy, color=white)


def widths(ws, values):
    for col, width in enumerate(values, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A3"


# 1. Valuation
ws = wb.active
ws.title = "Valuation"
title(ws, "Atlassian Corporation (TEAM) — Valuation", 4, "Quote: September 3, 2026 close | Model date: September 4, 2026 | USD")
facts = [
    ("Company", "Atlassian Corporation", "AI-powered collaboration and team productivity software"),
    ("Ticker", "NASDAQ: TEAM", "Jira, Confluence, Loom, Jira Service Management, Rovo and related products"),
    ("Price", PRICE, "StockAnalysis September 3 close"),
    ("Shares outstanding (M)", SHARES, "Filing-date shares; down 0.62% YoY"),
    ("Market capitalization ($M)", MARKET_CAP, "StockAnalysis"),
    ("Enterprise value ($M)", ENTERPRISE_VALUE, "Near market cap because cash and debt are approximately equal"),
    ("Net cash ($M)", NET_CASH, "$1.240B cash and investments less $1.233B debt"),
    ("Primary valuation lens", "Discounted terminal FCF", "Forward adjusted P/E and EV/FCF are cross-checks"),
    ("Stance", "Watch", "Excellent product momentum, but the post-earnings price already exceeds average analyst target and FCF quality is diluted by SBC"),
]
header(ws, 3, ["Field", "Value", "Comment"])
for row, values in enumerate(facts, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
header(ws, 14, ["Valuation metric", "Value", "Interpretation"])
metrics = [
    ("Trailing P/E", "N/A", "GAAP net income is negative; trailing P/E is not meaningful"),
    ("Forward P/E", 35.91, "Based on non-GAAP adjusted FY2027 EPS consensus"),
    ("P/S", 7.50, "Elevated for 13.8% FY2027 revenue growth"),
    ("P/FCF", 37.36, "2.68% reported FCF yield before treating SBC as an owner cost"),
    ("EV/FCF", 37.36, "Cash and debt nearly offset"),
    ("EV/Sales", 7.50, "Requires durable cloud growth and margin expansion"),
    ("EV/EBITDA", 112.91, "GAAP EBITDA remains depressed by SBC and investment"),
    ("FCF margin", 0.2007, "Down from 27.1% in FY2025 as receivables rose and SBC increased"),
    ("Analyst average target", 192.52, "Slightly below the September 3 close"),
]
for row, values in enumerate(metrics, 15):
    for col, value in enumerate(values, 1):
        cell = put(ws, row, col, value, bold=(col == 1))
        if col == 2 and values[0] == "FCF margin":
            cell.number_format = "0.0%"
widths(ws, [30, 24, 90, 14])

# 2. WACC
ws = wb.create_sheet("WACC")
title(ws, "Atlassian — Weighted Average Cost of Capital", 4, "CAPM and market-value capital weights")
header(ws, 3, ["Component", "Value", "Source / formula"])
wacc_rows = [
    ("Risk-free rate", rf, "U.S. 10-year Treasury, September 3, 2026"),
    ("Equity risk premium", erp, "Model assumption"),
    ("Levered beta", beta, "StockAnalysis five-year beta"),
    ("Cost of equity", cost_equity, "Risk-free rate + beta × ERP"),
    ("Pre-tax cost of debt", pretax_debt_cost, "FY2026 cash interest / total debt; conservative proxy"),
    ("Tax rate", tax_rate, "Normalized model rate; TTM effective rate is distorted by low pretax income"),
    ("Market capitalization", MARKET_CAP, "USD millions"),
    ("Total debt", 1_233.0, "USD millions"),
    ("Equity weight", equity_weight, "Equity / (equity + debt)"),
    ("Debt weight", debt_weight, "Debt / (equity + debt)"),
    ("WACC", wacc, "E/V × Ke + D/V × Kd × (1 − tax rate)"),
]
for row, values in enumerate(wacc_rows, 4):
    for col, value in enumerate(values, 1):
        cell = put(ws, row, col, value, bold=(col == 1 or values[0] == "WACC"), fill=(gold if values[0] == "WACC" else None))
        if col == 2 and values[0] not in {"Levered beta", "Market capitalization", "Total debt"}:
            cell.number_format = "0.00%"
widths(ws, [31, 22, 90, 14])

# 3. Scenarios
ws = wb.create_sheet("Scenarios")
title(ws, "Atlassian — Discounted FCF Scenarios", 6, "Five-year terminal value from FY2027 consensus revenue; all financial values USD millions")
header(ws, 3, ["Metric", "Bear", "Base", "Bull", "Notes"])
scenario_rows = [
    ("FY2027 revenue anchor", FY27_REVENUE, FY27_REVENUE, FY27_REVENUE, "Visible StockAnalysis consensus"),
    ("Revenue CAGR (5Y)", *(SCENARIOS[k]["growth"] for k in ("Bear", "Base", "Bull")), "Cloud migration, enterprise penetration and AI monetization"),
    ("Terminal revenue", *(SCENARIOS[k]["terminal_revenue"] for k in ("Bear", "Base", "Bull")), "FY2027 consensus compounded five years"),
    ("Adjusted FCF margin", *(SCENARIOS[k]["margin"] for k in ("Bear", "Base", "Bull")), "Reported FCF; interpretation must account for SBC"),
    ("Terminal FCF", *(SCENARIOS[k]["terminal_fcf"] for k in ("Bear", "Base", "Bull")), "Terminal revenue × FCF margin"),
    ("Exit FCF multiple", *(SCENARIOS[k]["multiple"] for k in ("Bear", "Base", "Bull")), "Multiple reflects terminal growth and owner-quality of cash flow"),
    ("Implied enterprise value", *(SCENARIOS[k]["implied_ev"] for k in ("Bear", "Base", "Bull")), "Terminal FCF × exit multiple"),
    ("Net cash adjustment", *(SCENARIOS[k]["net_cash"] for k in ("Bear", "Base", "Bull")), "Year-five assumption after buybacks and acquisitions"),
    ("Shares outstanding", *(SCENARIOS[k]["shares"] for k in ("Bear", "Base", "Bull")), "Repurchases partly offset SBC"),
    ("Terminal price", *(SCENARIOS[k]["terminal_price"] for k in ("Bear", "Base", "Bull")), "Equity value / shares"),
    ("Discount rate", *(SCENARIOS[k]["discount"] for k in ("Bear", "Base", "Bull")), "Scenario risk adjustment around WACC"),
    ("Present target price", *(SCENARIOS[k]["target"] for k in ("Bear", "Base", "Bull")), "Five-year terminal value discounted to present"),
    ("Upside / (downside)", *(SCENARIOS[k]["upside"] for k in ("Bear", "Base", "Bull")), "Versus $194.68"),
    ("Probability", *(SCENARIOS[k]["weight"] for k in ("Bear", "Base", "Bull")), "20% / 55% / 25%"),
    ("Weighted value/share", *(SCENARIOS[k]["target"] * SCENARIOS[k]["weight"] for k in ("Bear", "Base", "Bull")), "Contribution to probability-weighted value"),
]
for row, values in enumerate(scenario_rows, 4):
    for col, value in enumerate(values, 1):
        cell = put(ws, row, col, value, bold=(col == 1))
        if col in (2, 3, 4) and values[0] in {"Revenue CAGR (5Y)", "Adjusted FCF margin", "Discount rate", "Upside / (downside)", "Probability"}:
            cell.number_format = "0.0%"
        elif col in (2, 3, 4) and values[0] in {"Terminal price", "Present target price", "Weighted value/share"}:
            cell.number_format = "$0.00"
put(ws, 20, 1, "Probability-weighted fair value", bold=True, fill=gold)
put(ws, 20, 2, weighted_value, bold=True, fill=gold).number_format = "$0.00"
put(ws, 21, 1, "Upside from current price", bold=True, fill=gold)
put(ws, 21, 2, weighted_value / PRICE - 1, bold=True, fill=gold).number_format = "0.0%"
widths(ws, [31, 18, 18, 18, 90, 14])

# 4. Actuals Source Audit
ws = wb.create_sheet("Actuals Source Audit")
title(ws, "Atlassian — Actuals Source Audit", 5, "Figures in USD millions unless stated")
header(ws, 3, ["Data point", "Value", "Source URL", "Source date", "Notes"])
sa = "https://stockanalysis.com/stocks/team"
audit = [
    ("Stock price", "$194.68", f"{sa}/", "2026-09-03", "Closing price"),
    ("Market capitalization", "$49.28B", f"{sa}/statistics/", "2026-09-03", "Price × current shares"),
    ("Enterprise value", "$49.27B", f"{sa}/statistics/", "2026-09-03", "Cash and debt nearly offset"),
    ("Shares outstanding", "253.14M", f"{sa}/statistics/", "2026-09-03", "Down 0.62% YoY and 2.90% QoQ"),
    ("FY2026 revenue", "$6.572B", f"{sa}/financials/", "2026-06-30", "+26.02%"),
    ("FY2026 gross profit", "$5.629B", f"{sa}/financials/", "2026-06-30", "85.64% margin"),
    ("FY2026 operating income", "$295.73M", f"{sa}/financials/", "2026-06-30", "4.50% GAAP margin"),
    ("FY2026 net income", "-$53.83M", f"{sa}/financials/", "2026-06-30", "GAAP"),
    ("FY2026 operating cash flow", "$1.353B", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "Down 7.34%"),
    ("FY2026 capex", "$34.06M", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "Cash outflow shown as negative by provider"),
    ("FY2026 free cash flow", "$1.319B", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "20.07% margin; down 6.82%"),
    ("FY2026 stock compensation", "$1.607B", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "Exceeds reported FCF and is an owner-cost warning"),
    ("FY2026 repurchases", "$1.800B", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "Exceeded FCF; reduced cash balance"),
    ("FY2026 cash acquisitions", "$1.229B", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "Browser Company and other transaction activity"),
    ("Cash and investments", "$1.240B", f"{sa}/financials/balance-sheet/", "2026-06-30", "Down 57.8% YoY"),
    ("Total debt", "$1.233B", f"{sa}/financials/balance-sheet/", "2026-06-30", "Includes leases and long-term debt"),
    ("Goodwill", "$2.303B", f"{sa}/financials/balance-sheet/", "2026-06-30", "Up from $1.304B after acquisition activity"),
    ("Current deferred revenue", "$2.495B", f"{sa}/financials/balance-sheet/", "2026-06-30", "Subscription billing support"),
    ("FY2027 revenue consensus", "$7.48B", f"{sa}/forecast/", "2026-09-03", "32 analysts; +13.79%"),
    ("FY2027 adjusted EPS consensus", "$5.42", f"{sa}/forecast/", "2026-09-03", "Non-GAAP; range $4.81-$6.11"),
    ("FY2028 headline revenue", "$8.57B", f"{sa}/forecast/", "2026-09-03", "+14.59%; detailed table gated"),
    ("FY2028 headline adjusted EPS", "$6.74", f"{sa}/forecast/", "2026-09-03", "+24.37%; detailed table gated"),
    ("Analyst average target", "$192.52", f"{sa}/forecast/", "2026-09-03", "33 analysts; median $180"),
    ("Beta", "1.16", f"{sa}/statistics/", "2026-09-03", "Five-year beta"),
    ("Last earnings", "August 6, 2026", f"{sa}/statistics/", "2026-09-03", "Already released; next quarterly report is forward catalyst"),
]
for row, values in enumerate(audit, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
widths(ws, [34, 22, 96, 16, 72])

# 5. Questions
ws = wb.create_sheet("Questions")
title(ws, "Atlassian — Open Questions", 4, "Items that can change valuation or conviction")
header(ws, 3, ["#", "Question", "Why it matters", "Best evidence / next check"])
questions = [
    (1, "How much of FY2026 growth came from cloud migration timing versus durable seat and price expansion?", "Data Center end-of-life can accelerate recognition and pull demand forward.", "Cloud ARR, new-logo growth, seat expansion and cohort retention."),
    (2, "Can cloud revenue sustain growth above 20% while Data Center declines roughly 17% in FY2027?", "Consensus total growth is only 13.8%; mix determines durability and gross margin.", "Quarterly cloud and Data Center revenue."),
    (3, "What revenue is directly attributable to Rovo and AI credits?", "Usage is strong, but valuation requires paid monetization rather than engagement alone.", "AI ARR, attach, credit consumption and renewal uplift."),
    (4, "Why did stock-based compensation reach $1.607B, exceeding reported FCF?", "Reported FCF overstates owner cash if grants remain this large.", "SBC/revenue, grant-date value, dilution and repurchase offset."),
    (5, "Are $1.8B of repurchases economically accretive when cash acquisitions also consumed $1.229B?", "Both uses exceeded annual FCF and reduced cash 58%.", "Average repurchase price, authorization, minimum cash policy and acquisition returns."),
    (6, "What explains the $999M goodwill increase and $187M intangible increase?", "Acquisition success is now material to capital efficiency and earnings quality.", "Purchase-price allocation and acquired revenue/retention."),
    (7, "How much of the FY2027 non-GAAP margin decline to roughly 25% reflects intentional AI investment?", "A lower adjusted margin despite growth raises the hurdle for monetization.", "R&D, inference cost and incremental gross profit from AI."),
    (8, "Can GAAP operating margin reach the 4.5% FY2027 target while adjusted margin contracts?", "The bridge depends on SBC, amortization and restructuring discipline.", "Quarterly GAAP/non-GAAP reconciliation."),
    (9, "How durable is Service Collection's >$1B ARR and >30% growth?", "It is a central enterprise and AI proof point.", "Enterprise ARR, customer adds, renewal and cross-sell."),
    (10, "Do Rovo users growing ARR at 2x non-Rovo users reflect selection bias?", "Larger, faster-growing customers may adopt Rovo first without causation.", "Matched cohorts and pre/post adoption expansion."),
    (11, "How will The Browser Company fit the System of Work without distracting product investment?", "Integration could create a new AI interface or become an expensive adjacency.", "Product roadmap, customer adoption and acquired-team retention."),
    (12, "What is the long-term fate of Bitbucket as developer workflows consolidate around GitHub and AI coding tools?", "A weakening developer entry point could reduce platform distribution.", "Bitbucket cloud usage, migrations and developer tool attach."),
    (13, "Can current deferred revenue cover the working-capital deficit without creating renewal risk?", "Current liabilities exceed current assets, largely because subscriptions are billed ahead.", "RPO conversion, churn and contract duration."),
    (14, "Will management continue reducing the share count despite SBC exceeding FCF?", "Repurchases currently protect per-share economics but consume owner cash.", "Fully diluted share trend and repurchases/FCF."),
    (15, "When is the next quarterly earnings release after August 6?", "It should test FY2027 guidance, cloud durability, Data Center decline and AI monetization.", "Atlassian investor-relations calendar."),
]
for row, values in enumerate(questions, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
widths(ws, [7, 82, 72, 70])

# 6. Sources
ws = wb.create_sheet("Sources")
title(ws, "Atlassian — Sources", 4, "Public sources accessed September 4, 2026")
header(ws, 3, ["#", "Source", "URL", "Use"])
sources = [
    (1, "StockAnalysis overview", f"{sa}/", "Price, identity, market data and analyst summary"),
    (2, "StockAnalysis financials", f"{sa}/financials/", "Historical revenue, gross profit, operating income and net income"),
    (3, "StockAnalysis balance sheet", f"{sa}/financials/balance-sheet/", "Cash, debt, deferred revenue, goodwill and equity"),
    (4, "StockAnalysis cash flow", f"{sa}/financials/cash-flow-statement/", "OCF, capex, FCF, SBC, acquisitions and repurchases"),
    (5, "StockAnalysis statistics", f"{sa}/statistics/", "Valuation ratios, shares, beta, margins and earnings date"),
    (6, "StockAnalysis forecast", f"{sa}/forecast/", "Consensus revenue, adjusted EPS and price targets"),
    (7, "Atlassian Q4 FY2026 shareholder letter", "https://www.atlassian.com/blog/company-news/shareholder-letter-q4fy26", "Strategy, AI, enterprise and fiscal-year commentary"),
    (8, "Atlassian FY2026 earnings release", "https://investors.atlassian.com/news/news-details/2026/Atlassian-Announces-Fourth-Quarter-and-Fiscal-Year-2026-Results/default.aspx", "Official results, non-GAAP reconciliation and guidance"),
    (9, "Atlassian Q3 FY2026 shareholder letter", "https://www.atlassian.com/blog/company-news/shareholder-letter-q3fy26", "Service Collection and Rovo operating metrics"),
    (10, "U.S. Treasury yield curve", "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/TextView?type=daily_treasury_yield_curve&field_tdr_date_value=2026", "Risk-free-rate reference"),
]
for row, values in enumerate(sources, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
widths(ws, [7, 38, 110, 72])

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

wb.save(OUT)
check = load_workbook(OUT, data_only=False)
assert check.sheetnames == ["Valuation", "WACC", "Scenarios", "Actuals Source Audit", "Questions", "Sources"]
assert check["Valuation"]["B6"].value == PRICE
assert isclose(check["Scenarios"]["B20"].value, weighted_value, rel_tol=1e-12)
print(f"WACC: {wacc:.2%}")
for name, case in SCENARIOS.items():
    print(f"{name}: ${case['target']:.2f} ({case['upside']:+.1%})")
print(f"Probability-weighted fair value: ${weighted_value:.2f} ({weighted_value / PRICE - 1:+.1%})")
print(f"Created {OUT}")
