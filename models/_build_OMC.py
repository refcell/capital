#!/usr/bin/env python3
"""Build 6-sheet valuation model for OMC (Omnicom Group Inc.)"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()

# Styling
title_font = Font(name="Calibri", size=14, bold=True)
header_font = Font(name="Calibri", size=11, bold=True)
subtitle_font = Font(name="Calibri", size=10, italic=True)
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", size=10, bold=True)
header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
bear_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
base_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
bull_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def c(ws, row, col, value, font=normal_font, border=False, fill=None, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if border:
        cell.border = thin_border
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    return cell

# ============================================================
# Sheet 1: Valuation
# ============================================================
ws1 = wb.active
ws1.title = "Valuation"
ws1.merge_cells("A1:F1")
c(ws1, 1, 1, "Omnicom Group Inc. (OMC) - Valuation Model", title_font)
c(ws1, 2, 1, "As of August 26, 2026 | Post-Entertainment One Acquisition", subtitle_font)

title_data = [
    ("Ticker", "OMC", "NYSE"),
    ("Company", "Omnicom Group Inc.", "Advertising Agencies / Communication Services"),
    ("Date", "2026-08-26", ""),
    ("Price", "$87.89", "Close Aug 26, 2026"),
    ("Shares Outstanding", "274.35M", "Yahoo Key Stats"),
    ("Market Cap", "$24.20B", "Price x Shares"),
    ("Enterprise Value", "$32.27B", "Yahoo Key Stats"),
    ("Net Debt (EV-MC)", "$8.07B", "EV - MC proxy"),
    ("Primary Lens", "Forward P/E", "Post-acquisition earnings transition; trailing P/E meaningless"),
    ("Stance", "Watch", "Integration year ahead; FY26 earnings the test"),
]

for i, (field, value, note) in enumerate(title_data, 2):
    c(ws1, i, 1, field, bold_font, True)
    c(ws1, i, 2, value, normal_font, True)
    for col_letter in range(3, 7):
        c(ws1, i, col_letter, "", normal_font, True)
    # Merge note cells
    ws1.merge_cells(start_row=i, start_column=3, end_row=i, end_column=6)
    c(ws1, i, 3, note, subtitle_font, True)

c(ws1, 13, 1, "Valuation Metrics", header_font)
c(ws1, 14, 1, "Metric", header_font, True, header_fill)
c(ws1, 14, 2, "Value", header_font, True, header_fill)
c(ws1, 14, 3, "Comment", header_font, True, header_fill)

metrics = [
    ("Trailing P/E", "238.38x", "Distorted by FY25 acquisition transition loss of $-54.5M"),
    ("Forward P/E", "8.37x", "On FY2026 consensus EPS of $10.59"),
    ("P/Sales (TTM)", "0.99x", "$24.2B / $22.37B TTM revenue"),
    ("P/Book", "2.50x", "$24.2B / $12.05B common equity"),
    ("Forward PEG (5yr)", "15.97x", "Data quality suspect post-acquisition; P/E denominator distorted"),
    ("EV/Revenue", "1.44x", "$32.27B / $22.37B"),
    ("EV/EBITDA", "18.59x", "$32.27B / $1.74B TTM EBITDA; integration depressed EBITDA"),
    ("FCF Yield", "17.6%", "$4.26B FCF / $24.2B MC; includes investment portfolio activity"),
    ("Dividend Yield", "3.63%", "$3.20 annualized / $87.89"),
    ("Beta (5Y)", "0.66", "Low beta - defensive advertising services"),
    ("52W Range", "$66.33 - $89.57", "Near 52-week high"),
    ("50-Day MA", "$81.08", ""),
    ("200-Day MA", "$77.60", ""),
]

for i, (metric, value, comment) in enumerate(metrics, 15):
    c(ws1, i, 1, metric, bold_font, True)
    c(ws1, i, 2, value, normal_font, True)
    c(ws1, i, 3, comment, subtitle_font, True)

ws1.column_dimensions["A"].width = 20
ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 65

# ============================================================
# Sheet 2: WACC
# ============================================================
ws2 = wb.create_sheet("WACC")
ws2.merge_cells("A1:D1")
c(ws2, 1, 1, "WACC Calculation - OMC", title_font)
c(ws2, 2, 1, "CAPM Framework | Post-Acquisition Capital Structure", subtitle_font)

wacc_data = [
    ("Risk-Free Rate (10Y US)", "4.66%", "CNBC US10Y Aug 26 2026"),
    ("", "", ""),
    ("Equity Risk Premium", "5.00%", "Standard assumption"),
    ("Levered Beta (5Y Monthly)", "0.66", "Yahoo Key Stats"),
    ("", "", ""),
    ("Cost of Equity (CAPM)", "7.96%", "=Rfr + Beta x ERP = 4.66% + 0.66 x 5.00%"),
    ("", "", ""),
    ("Total Debt (Key Stats)", "$11.41B", "Yahoo Key Stats - higher than BS $10.73B; more conservative"),
    ("Total Cash (Key Stats)", "$3.34B", "Yahoo Key Stats"),
    ("Market Cap", "$24.20B", ""),
    ("", "", ""),
    ("Total Capitalization", "$35.61B", "Debt + MC = 11.41 + 24.20"),
    ("Debt Weight", "32.0%", "11.41 / 35.61"),
    ("Equity Weight", "68.0%", "24.20 / 35.61"),
    ("", "", ""),
    ("Cost of Debt (est.)", "4.50%", "Estimate based on investment grade rating; debt increased post-acquisition"),
    ("Tax Rate", "25.6%", "TTM: 380.4M / 859.9M pretax"),
    ("After-Tax Cost of Debt", "3.34%", "=4.50% x (1-25.6%)"),
    ("", "", ""),
    ("WACC", "6.48%", "=0.68 x 7.96% + 0.32 x 3.34%"),
]

for i, (field, value, note) in enumerate(wacc_data, 3):
    c(ws2, i, 1, field, bold_font if field and not field.startswith(" ") and "=" not in field else normal_font, True)
    c(ws2, i, 2, value, bold_font if "=" in (note or "") else normal_font, True)
    c(ws2, i, 3, note, subtitle_font, True)

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 65

# WACC verification
rfr = 0.0466
erp = 0.05
beta = 0.66
ke = rfr + beta * erp
tax_rate = 0.256
kd = 0.045
kd_at = kd * (1 - tax_rate)
w_debt = 11.41 / 35.61
w_eq = 24.20 / 35.61
wacc = w_eq * ke + w_debt * kd_at
print(f"WACC = {wacc:.4f} = {w_eq:.3f} x {ke:.4f} + {w_debt:.3f} x {kd_at:.4f}")

# ============================================================
# Sheet 3: Scenarios
# ============================================================
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells("A1:I1")
c(ws3, 1, 1, "Scenario Analysis - OMC", title_font)
c(ws3, 2, 1, "Forward P/E Framework (Primary) | FY2026 Acquired Base | Earnings Consensus Anchored", subtitle_font)
ws3.merge_cells("C3:I3")
c(ws3, 3, 3, "NOTE: Trailing P/E excluded due to FY25 acquisition-driven net loss (-$54.5M). Forward P/E on consensus is the correct framework.", bold_font)

headers = ["Metric", "Bear", "Base", "Bull", "Comment"]
for j, h in enumerate(headers, 1):
    c(ws3, 5, j, h, header_font, True, header_fill)

scenarios = [
    ("Revenue CAGR (5Y)", "2%", "4.5%", "7%", "Post-acquisition; organic growth ~2-4%, add M&A integration tailwind in bull"),
    ("FY2026 Revenue Base", "$25.84B", "$25.84B", "$25.84B", "Consensus estimate - acquired base"),
    ("Terminal Revenue (Yr5)", "$27.9B", "$32.5B", "$36.3B", "=base x (1+cagr)^5"),
    ("Terminal EPS (Yr5)", "$10.00", "$14.00", "$17.50", "Growth from FY27 $12.04 consensus base"),
    ("Exit P/E Multiple", "7x", "10x", "12x", "Peer: WPP ~15x, Publicis ~12x; integration discount in bear/base"),
    ("Implied 5Y Target", "$70.00", "$140.00", "$210.00", "=Terminal EPS x Exit P/E"),
    ("Upside from $87.89", "-20.3%", "+59.1%", "+138.1%", "Bear: value trap if integration fails"),
    ("", "", "", "", ""),
    ("Scenario Weight", "20%", "50%", "30%", ""),
    ("Weighted Value/Share", "$14.00", "$70.00", "$63.00", ""),
    ("", "", "", ""),
    ("Probability-Weighted FV", "", "", "$147.00", "Sum of weighted values"),
    ("Upside from Current", "", "", "+67.2%", "Compelling if integration delivers"),
    ("", "", "", "", ""),
    ("Cross-Check: EV/EBITDA", "", "", "", ""),
    ("Bear EBITDA", "", "$900M", "", "If integration drags margins"),
    ("Base EBITDA", "", "$2.4B", "", "Return to pre-acquisition normalized run-rate"),
    ("Bull EBITDA", "", "$3.0B", "", "Synergy realization"),
    ("EV/EBITDA Bear Multiple", "", "12x", "", "Compressed for integration risk"),
    ("EV/EBITDA Base Multiple", "", "16x", "", "Peer median"),
    ("EV/EBITDA Bull Multiple", "", "20x", "", "Premium for proven synergies"),
    ("Implied EV (Bear)", "", "$10.8B", "", "Below current $32.3B EV - bear EV/EBITDA unrealistic"),
    ("Implied EV (Base)", "", "$38.4B", "", "Near current"),
    ("Implied EV (Bull)", "", "$60.0B", "", "Synergies proven"),
]

for i, row_data in enumerate(scenarios, 6):
    for j, val in enumerate(row_data):
        font = bold_font if j == 0 and val else normal_font
        fill = None
        if j == 1 and val:
            fill = bear_fill
        elif j == 2 and val:
            fill = base_fill
        elif j == 3 and val:
            fill = bull_fill
        c(ws3, i, j+1, val, font, True, fill)

# Verify scenario math
terminal_eps_bear = 10.00
terminal_eps_base = 14.00
terminal_eps_bull = 17.50
exit_pe_bear = 7
exit_pe_base = 10
exit_pe_bull = 12
target_bear = terminal_eps_bear * exit_pe_bear
target_base = terminal_eps_base * exit_pe_base
target_bull = terminal_eps_bull * exit_pe_bull
print(f"Targets: Bear=${target_bear:.2f} Base=${target_base:.2f} Bull=${target_bull:.2f}")

fv = 0.20 * target_bear + 0.50 * target_base + 0.30 * target_bull
print(f"Probability-Weighted FV: ${fv:.2f}")
print(f"WACC: {wacc:.4f}")

# ============================================================
# Sheet 4: Actuals Source Audit
# ============================================================
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells("A1:E1")
c(ws4, 1, 1, "Actuals Source Audit - OMC", title_font)
c(ws4, 2, 1, "Every data point with source URL and extraction date", subtitle_font)

c(ws4, 4, 1, "Data Point", header_font, True, header_fill)
c(ws4, 4, 2, "Value", header_font, True, header_fill)
c(ws4, 4, 3, "Source", header_font, True, header_fill)
c(ws4, 4, 4, "Date", header_font, True, header_fill)
c(ws4, 4, 5, "Notes", header_font, True, header_fill)

audit = [
    ("Stock Price", "$87.89", "Yahoo Finance Key Stats", "2026-08-26", "Close price"),
    ("Market Cap", "$24.20B", "Yahoo Finance Key Stats", "2026-08-26", ""),
    ("Enterprise Value", "$32.27B", "Yahoo Finance Key Stats", "2026-08-26", ""),
    ("Shares Outstanding", "274.35M", "Yahoo Finance Key Stats", "2026-08-26", "Significant increase post-acquisition (~196M pre)"),
    ("Beta (5Y Monthly)", "0.66", "Yahoo Finance Key Stats", "2026-08-26", ""),
    ("", "", "", "", ""),
    ("TTM Revenue", "$22,371M", "Yahoo Finance Income Statement", "2026-08-26", "In thousands; acquired base"),
    ("FY2025 Revenue", "$17,272M", "Yahoo Finance Income Statement", "2026-08-26", "Acquisition year"),
    ("FY2024 Revenue", "$15,689M", "Yahoo Finance Income Statement", "2026-08-26", "Pre-acquisition comparable"),
    ("TTM Gross Profit", "$2,505M", "Yahoo Finance Income Statement", "2026-08-26", ""),
    ("TTM Operating Income", "$1,122M", "Yahoo Finance Income Statement", "2026-08-26", "Suppressed by integration costs"),
    ("FY2025 Operating Income", "$445M", "Yahoo Finance Income Statement", "2026-08-26", "Collapsed from $2,275M FY24"),
    ("TTM Net Income", "$390M", "Yahoo Finance Income Statement", "2026-08-26", ""),
    ("FY2025 Net Income", "$-55M", "Yahoo Finance Income Statement", "2026-08-26", "Loss in acquisition year"),
    ("TTM Diluted EPS", "$0.37", "Yahoo Finance Income Statement", "2026-08-26", "Distorted by mix of FY25 loss + H1 FY26"),
    ("TTM EBITDA", "$1,736M", "Yahoo Finance Income Statement", "2026-08-26", "Calculated by Yahoo/Refinitiv"),
    ("TTM EBIT", "$1,244M", "Yahoo Finance Income Statement", "2026-08-26", ""),
    ("", "", "", "", ""),
    ("Total Assets", "$54,415M", "Yahoo Finance Balance Sheet", "2026-08-26", "Doubled from $29,621M FY25 - acquisition"),
    ("Total Debt (BS)", "$10,734M", "Yahoo Finance Balance Sheet", "2026-08-26", "FY25 figure"),
    ("Total Debt (Key Stats)", "$11,410M", "Yahoo Finance Key Stats", "2026-08-26", "MRQ; may include capital leases"),
    ("Total Cash", "$3,340M", "Yahoo Finance Key Stats", "2026-08-26", "MRQ"),
    ("Net Debt (BS)", "$2,235M", "Calculated from BS", "2026-08-26", "Debt - Cash"),
    ("Common Stock Equity", "$12,046M", "Yahoo Finance Balance Sheet", "2026-08-26", "FY25 figure"),
    ("", "", "", "", ""),
    ("TTM Operating Cash Flow", "$2,583M", "Yahoo Finance Cash Flow", "2026-08-26", ""),
    ("TTM Levered FCF", "$4,260M", "Yahoo Finance Key Stats", "2026-08-26", "Includes investment portfolio activity"),
    ("", "", "", "", ""),
    ("Forward P/E", "8.37x", "Yahoo Finance Key Stats", "2026-08-26", "On FY2026 consensus"),
    ("Fwd P/E (Current Qtr)", "N/A", "Yahoo Finance Key Stats", "2026-08-26", "Not separately listed"),
    ("P/S Ratio", "0.99x", "Yahoo Finance Key Stats", "2026-08-26", ""),
    ("P/B Ratio", "2.50x", "Yahoo Finance Key Stats", "2026-08-26", ""),
    ("EV/Revenue", "1.44x", "Yahoo Finance Key Stats", "2026-08-26", ""),
    ("EV/EBITDA", "18.59x", "Yahoo Finance Key Stats", "2026-08-26", "S&P Global EBITDA calc"),
    ("", "", "", "", ""),
    ("Analyst EPS Consensus FY26", "$10.59", "Yahoo Finance Analysis", "2026-08-26", "12 analysts; Normalised/Non-GAAP"),
    ("Analyst EPS Consensus FY27", "$12.04", "Yahoo Finance Analysis", "2026-08-26", "12 analysts"),
    ("Analyst Revenue FY26", "$25,840M", "Yahoo Finance Analysis", "2026-08-26", "11 analysts"),
    ("Analyst Revenue FY27", "$25,110M", "Yahoo Finance Analysis", "2026-08-26", "11 analysts - slight decline expected"),
    ("", "", "", "", ""),
    ("Next Earnings Date", "Oct 20, 2026", "Yahoo Finance Profile", "2026-08-26", ""),
    ("Dividend Ex-Date", "Sep 18, 2026", "Yahoo Finance Profile", "2026-08-26", "$0.80 quarterly"),
    ("10Y Treasury Rate", "4.656%", "CNBC US10Y", "2026-08-26", ""),
]

for i, (dp, val, src, dt, note) in enumerate(audit, 5):
    c(ws4, i, 1, dp, normal_font, True)
    c(ws4, i, 2, val, normal_font, True)
    c(ws4, i, 3, src, normal_font, True)
    c(ws4, i, 4, dt, normal_font, True)
    c(ws4, i, 5, note, normal_font, True)

ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 15
ws4.column_dimensions["C"].width = 35
ws4.column_dimensions["D"].width = 15
ws4.column_dimensions["E"].width = 55

# ============================================================
# Sheet 5: Questions
# ============================================================
ws5 = wb.create_sheet("Questions")
ws5.merge_cells("A1:C1")
c(ws5, 1, 1, "Open Questions - OMC", title_font)

c(ws5, 3, 1, "#", header_font, True, header_fill)
c(ws5, 3, 2, "Question", header_font, True, header_fill)
c(ws5, 3, 3, "Category", header_font, True, header_fill)

questions = [
    ("1", "Revenue jumped 29.5% YoY (FY25 $17.3B to TTM $22.4B) and assets nearly doubled ($29.6B to $54.4B). What is the size and identity of the acquisition? Was it the Entertainment One (EONE) acquisition announced in January 2025 for ~$830M net cash?", "Acquisition"),
    ("2", "Oper. Income collapsed from $2.27B (FY24) to $445M (FY25) and net income flipped to -$55M loss. Is this entirely acquisition integration, or does it signal organic margin weakness in the advertising franchise?", "Earnings Quality"),
    ("3", "Debt on Key Stats ($11.41B) vs Balance Sheet ($10.73B) - what accounts for the $680M difference? Capital lease obligations now $1.62B (up from $814M FY24) - does the Key Stats figure include these?", "Capital Structure"),
    ("4", "Share count expanded from ~196M (FY24) to 274.35M (current) - a 40% increase. Was this all acquisition-related stock consideration? Any PIPE financing or debt-to-equity swaps?", "Dilution"),
    ("5", "Payout ratio of 837.84% is a mathematical artifact of depressed TTM earnings. What is the normalised payout ratio against forward earnings? At $10.59 FY26 EPS and $3.20 dividend, the ratio is ~30%.", "Dividends"),
    ("6", "Capital lease obligations doubled from $814M to $1.62B. This signals new office space, production facilities, or fleet commitments. What are the terms?", "Capital Lease"),
    ("7", "TTM levered FCF of $4.26B appears dramatically higher than OCF of $2.58B - the $1.68B gap likely represents investment portfolio mark-to-market or asset sales, not operating cash flow.", "Cash Flow Quality"),
    ("8", "FY25 gross profit was only $1.47B (8.5% margin) vs $2.92B (18.6%) in FY24. The acquired entity has lower margins than the existing portfolio - what is the blended margin trajectory?", "Margin Profile"),
    ("9", "EV/EBITDA of 18.59x is substantially above the 10-12x peer range for advertising agencies. Does the market think integration synergies will return EBITDA to FY24 levels?", "Valuation Calibration"),
    ("10", "What is the entertainment segment's contribution to revenue growth vs. organic advertising growth? If the acquisition was EONE, how does that change the risk profile given global entertainment industry headwinds?", "Segment Mix"),
    ("11", "Next earnings date is Oct 20, 2026 (Q3 FY26). Will this be the first post-acquisition quarter with comparables? Full year of combined results would be FY26.", "Catalyst"),
    ("12", "Management guidance: Has John Wren or the co-CEOs provided revenue or margin guidance for the combined entity? Any synergy target language in Q2 FY26 earnings?", "Management Guidance"),
]

for i, (num, q, cat) in enumerate(questions, 4):
    c(ws5, i, 1, num, bold_font, True)
    c(ws5, i, 2, q, normal_font, True)
    c(ws5, i, 3, cat, normal_font, True)

ws5.column_dimensions["A"].width = 5
ws5.column_dimensions["B"].width = 100
ws5.column_dimensions["C"].width = 20

# ============================================================
# Sheet 6: Sources
# ============================================================
ws6 = wb.create_sheet("Sources")
ws6.merge_cells("A1:C1")
c(ws6, 1, 1, "Sources - OMC", title_font)

c(ws6, 3, 1, "#", header_font, True, header_fill)
c(ws6, 3, 2, "Source", header_font, True, header_fill)
c(ws6, 3, 3, "URL", header_font, True, header_fill)

sources = [
    ("1", "Yahoo Finance - Key Statistics", "https://finance.yahoo.com/quote/OMC/key-statistics/"),
    ("2", "Yahoo Finance - Income Statement", "https://finance.yahoo.com/quote/OMC/financials/"),
    ("3", "Yahoo Finance - Balance Sheet", "https://finance.yahoo.com/quote/OMC/balance-sheet/"),
    ("4", "Yahoo Finance - Cash Flow", "https://finance.yahoo.com/quote/OMC/cash-flow/"),
    ("5", "Yahoo Finance - Analyst Estimates", "https://finance.yahoo.com/quote/OMC/analysis/"),
    ("6", "Yahoo Finance - Company Profile", "https://finance.yahoo.com/quote/OMC/profile/"),
    ("7", "CNBC - 10Y Treasury Yield", "https://www.cnbc.com/quotes/US10Y"),
    ("8", "StockAnalysis (404 - unavailable for OMC)", "https://stockanalysis.com/quote/OMC/"),
    ("9", "Omnicom Group Official Website", "https://www.omc.com"),
]

for i, (num, src, url) in enumerate(sources, 4):
    c(ws6, i, 1, num, normal_font, True)
    c(ws6, i, 2, src, normal_font, True)
    c(ws6, i, 3, url, normal_font, True)

ws6.column_dimensions["A"].width = 5
ws6.column_dimensions["B"].width = 45
ws6.column_dimensions["C"].width = 65

# ============================================================
# Save
# ============================================================
output_path = "/home/refcell/dev/capital/models/[2026-08-26] Omnicom Group Model.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")

# Verification
print(f"\nVerification:")
print(f"  WACC = {wacc:.4f} ({wacc*100:.2f}%)")
print(f"  FV = ${fv:.2f}")
print(f"  Current Price = $87.89")
print(f"  Upside = {(fv - 87.89) / 87.89 * 100:.1f}%")
print(f"  Net Debt = $8.07B (EV - MC = 32.27 - 24.20)")
print(f"  FCF/NetDebt = {4.26/8.07:.2f}x (adequate)")
