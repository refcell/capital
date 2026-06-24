#!/usr/bin/env python3
"""Build PAY (Paymentus Holdings, Inc.) 6-sheet valuation model."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()

# Styles
bold_font = Font(bold=True)
header_font = Font(bold=True, size=12)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
bear_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
base_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
bull_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = bold_font
        cell.border = thin_border
        cell.fill = header_fill

def style_data_range(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = thin_border

# ============================================================
# Sheet 1: Valuation
# ============================================================
ws1 = wb.active
ws1.title = "Valuation"
ws1.merge_cells('A1:F1')
ws1['A1'] = "Paymentus Holdings, Inc. (PAY) — Valuation Model"
ws1['A1'].font = Font(bold=True, size=14)

title_data = [
    ("Company:", "Paymentus Holdings, Inc."),
    ("Ticker:", "NYSE: PAY"),
    ("Sector:", "Technology — Fintech / Payments Infrastructure"),
    ("Date:", "2026-06-24"),
    ("Source Date:", "2026-06-23 close / pre-market 2026-06-24"),
    ("Price:", "$20.56"),
    ("Shares Outstanding:", "125.79M (implied, per Yahoo Finance Statistics)"),
    ("Market Cap:", "$2.59B"),
    ("Enterprise Value:", "$2.25B (MC $2.59B - Cash $324.5M + Debt $6.9M)"),
    ("Primary Lens:", "Forward P/E + EV/EBITDA; DCF scenarios with WACC"),
    ("Stance:", "Watch — strong cash flow, clean balance sheet, but valuation at a premium to near-term earnings growth"),
]

for i, (label, val) in enumerate(title_data, 2):
    ws1.cell(row=i, column=1, value=label).font = bold_font
    ws1.cell(row=i, column=2, value=val)

r = len(title_data) + 3
ws1.cell(row=r, column=1, value="Valuation Metric").font = bold_font
ws1.cell(row=r, column=2, value="Value").font = bold_font
ws1.cell(row=r, column=3, value="Comment").font = bold_font

metrics = [
    ("Trailing P/E", "36.1x", "TTM EPS $0.57; stock declined 40% from 52W high of $39.38"),
    ("Forward P/E (est.)", "24.8x", "Based on FY2026 EPS consensus $0.83 (non-GAAP/normalized)"),
    ("P/S (TTM)", "2.0x", "Revenue $1.28B TTM, declining from peak ~4.4x"),
    ("P/FCF (TTM)", "2.5x", "FCF $104.8M TTM on strong OCF $142M minus capex $37.3M"),
    ("EV/EBITDA (TTM)", "17.8x", "Enterprise value $2.25B; EBITDA ~$92.8M TTM (Yahoo Finance)"),
    ("P/B", "4.4x", "Book value $345.7M tangible; total equity $560.4M"),
    ("EV/Sales", "1.8x", "EV $2.25B / TTM revenue $1.28B"),
]

for i, (metric, value, comment) in enumerate(metrics, r + 1):
    ws1.cell(row=i, column=1, value=metric)
    ws1.cell(row=i, column=2, value=value)
    ws1.cell(row=i, column=3, value=comment)

ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 65

# ============================================================
# Sheet 2: WACC
# ============================================================
ws2 = wb.create_sheet("WACC")
ws2.merge_cells('A1:D1')
ws2['A1'] = "WACC Calculation — CAPM Method"
ws2['A1'].font = Font(bold=True, size=12)

wacc_data = [
    ("Component", "Value", "Source / Notes"),
    ("Risk-Free Rate (10Y US)", "4.48%", "CNBC US10Y, 2026-06-23, yield 4.479%"),
    ("Beta (5Y Monthly)", "1.31", "Yahoo Finance Statistics"),
    ("Equity Risk Premium", "5.00%", "Standard assumption"),
    ("Cost of Equity (Rf + Beta*ERP)", "11.02%", "= 4.48% + 1.31 * 5.00%"),
    ("Market Cap", "$2,586M", "Yahoo Finance, 2026-06-23"),
    ("Total Debt (capital leases)", "$6.9M", "Yahoo Finance Balance Sheet, FY2025"),
    ("Total Capitalization (E+D)", "$2,593M", "MC + Debt"),
    ("Equity Weight", "99.74%", "= 2586 / 2593"),
    ("Debt Weight", "0.26%", "= 6.9 / 2593"),
    ("Cost of Debt (pre-tax)", "5.00%", "Estimated; minimal debt, near-risk-free for lease obligations"),
    ("Effective Tax Rate", "22.17%", "TTM tax provision $22.6M / pretax income $96.6M"),
    ("After-Tax Cost of Debt", "3.89%", "= 5.00% * (1 - 22.17%)"),
    ("", "", ""),
    ("WACC", "11.01%", "= 0.9974 * 11.02% + 0.0026 * 3.89% (debt weight negligible)"),
]

for i, row_data in enumerate(wacc_data, 2):
    for j, val in enumerate(row_data, 1):
        cell = ws2.cell(row=i, column=j, value=val)
        if row_data[0] in ("Component", "",):
            cell.font = bold_font if row_data[0] else None

style_header_row(ws2, 2, 3)
style_data_range(ws2, 3, 15, 3)
ws2.column_dimensions['A'].width = 32
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 55

# ============================================================
# Sheet 3: Scenarios
# ============================================================
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells('A1:L1')
ws3['A1'] = "Bear / Base / Bull Scenario Analysis — 5-Year Projections (WACC: 11.01%)"
ws3['A1'].font = Font(bold=True, size=12)

headers = [
    "Scenario", "Revenue CAGR (5Y)", "Terminal Revenue ($M)",
    "Adjusted FCF Margin", "Terminal FCF ($M)",
    "Exit FCF Multiple", "Implied EV ($M)",
    "Less Net Debt Adj.", "Shares (M)",
    "Target Price", "Upside (%)",
    "Weight", "Weighted Value/Share",
]

for j, h in enumerate(headers, 1):
    ws3.cell(row=2, column=j, value=h)
style_header_row(ws3, 2, len(headers))

# Current anchors (base for CAGR): TTM revenue $1,279.7M
# Analyst estimates: FY2026 rev $1.43B/EPS $0.83; FY2027 rev $1.69B/EPS $1.02
# FCF TTM $104.8M on rev $1,279.7M → TTM FCF margin ~8.2%
# FY2025 FCF $125.0M on rev $1,196.5M → FY2025 FCF margin ~10.4%
# Net cash: cash $338.8M Q1-2026 - debt $6.6M = $332.2M net cash
# Shares: 125.79M implied outstanding

# Scenario math:
# Bear:  7% CAGR → term_rev = 1279.7*(1.07)^5 = 1795M, FCF margin 6.5% → FCF 116.7M, EV/FCF 10x → EV 1167M, equity = EV+cash = 1499M, share = 1499/125.79 = $11.92
# Base:  9% CAGR → term_rev = 1279.7*(1.09)^5 = 1971M, FCF margin 7.5% → FCF 147.8M, EV/FCF 12x → EV 1774M, equity = 2106M, share = 2106/125.79 = $16.74
# Bull: 14% CAGR → term_rev = 1279.7*(1.14)^5 = 2476M, FCF margin 8.5% → FCF 210.5M, EV/FCF 15x → EV 3157M, equity = 3489M, share = 3489/125.79 = $27.74

NET_CASH = 332.2
SHARES = 125.79
CURRENT_PRICE = 20.56

scenarios = [
    ("Bear", 0.07, 1795, 0.065, 116.7, 10, 1167, -NET_CASH, SHARES, 11.92, -42.2, 0.20),
    ("Base", 0.09, 1971, 0.075, 147.8, 12, 1774, -NET_CASH, SHARES, 16.74, -18.6, 0.50),
    ("Bull", 0.14, 2476, 0.085, 210.5, 15, 3157, -NET_CASH, SHARES, 27.74,  35.0, 0.30),
]

for i, (name, cagr, term_rev, fcf_margin, term_fcf, exit_mult, impl_ev, net_debt_adj, shares, tgt, upside, weight) in enumerate(scenarios, 3):
    weighted_val = weight * tgt
    ws3.cell(row=i, column=1, value=name)
    ws3.cell(row=i, column=2, value=round(cagr, 1))
    ws3.cell(row=i, column=3, value=round(term_rev, 1))
    ws3.cell(row=i, column=4, value=round(fcf_margin, 1))
    ws3.cell(row=i, column=5, value=round(term_fcf, 1))
    ws3.cell(row=i, column=6, value=exit_mult)
    ws3.cell(row=i, column=7, value=round(impl_ev, 1))
    ws3.cell(row=i, column=8, value=round(net_debt_adj, 1))
    ws3.cell(row=i, column=9, value=round(shares, 2))
    ws3.cell(row=i, column=10, value=round(tgt, 2))
    ws3.cell(row=i, column=11, value=round(upside, 2))
    ws3.cell(row=i, column=12, value=weight)
    ws3.cell(row=i, column=13, value=round(weighted_val, 2))
    if name == "Bear": fill = bear_fill
    elif name == "Base": fill = base_fill
    else: fill = bull_fill
    for c in range(1, 14):
        ws3.cell(row=i, column=c).fill = fill
    style_data_range(ws3, i, i, len(headers))

# Summary rows — weighted FV = 0.2*11.92 + 0.5*16.74 + 0.3*27.74 = 2.384 + 8.37 + 8.322 = 19.076
r = len(scenarios) + 4
ws3.cell(row=r, column=1, value="Probability-Weighted Fair Value").font = bold_font
ws3.cell(row=r, column=13, value="$19.08").font = bold_font
ws3.cell(row=r+1, column=1, value="Weighted Upside from $20.56").font = bold_font
ws3.cell(row=r+1, column=13, value="-7.2%").font = bold_font
ws3.cell(row=r+2, column=1, value="Current Price").font = bold_font
ws3.cell(row=r+2, column=13, value="$20.56").font = bold_font

ws3.column_dimensions['A'].width = 18
for j in range(2, 14):
    ws3.column_dimensions[get_column_letter(j)].width = 14

# ============================================================
# Sheet 4: Actuals Source Audit
# ============================================================
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells('A1:E1')
ws4['A1'] = "Data Source Audit — Every Data Point Traced"
ws4['A1'].font = Font(bold=True, size=12)

audit_headers = ["Data Point", "Value", "Source URL", "Date Accessed", "Notes"]
for j, h in enumerate(audit_headers, 1):
    ws4.cell(row=2, column=j, value=h)
style_header_row(ws4, 2, 5)

audit_data = [
    ("Stock Price", "$20.56", "finance.yahoo.com/quote/PAY/", "2026-06-23", "Close at 4:00 PM EDT; pre-market $20.72"),
    ("Market Cap", "$2.59B", "Yahoo Finance Statistics", "2026-06-23", "As reported; shares 125.79M implied"),
    ("Enterprise Value", "$2.25B", "Yahoo Finance Statistics", "2026-06-23", "MC $2.59B less net cash ~$332M"),
    ("Shares Outstanding", "125.79M", "Yahoo Finance Statistics", "2026-06-23", "Implied shares outstanding (incl. convertible subsidiary equity)"),
    ("Revenue TTM", "$1,279.7M", "Yahoo Finance /PAY/financials/", "2026-06-24", "TTM annual income statement"),
    ("Revenue FY2025", "$1,196.5M", "Yahoo Finance /PAY/financials/", "2026-06-24", "FY2025 annual"),
    ("Revenue FY2024", "$871.7M", "Yahoo Finance /PAY/financials/", "2026-06-24", "FY2024 annual"),
    ("Gross Profit FY2025", "$296.3M", "Yahoo Finance /PAY/financials/", "2026-06-24", "Gross margin ~24.8%"),
    ("Op Income FY2025", "$75.5M", "Yahoo Finance /PAY/financials/", "2026-06-24", "Operating margin ~6.3%"),
    ("Net Income FY2025", "$66.9M", "Yahoo Finance /PAY/financials/", "2026-06-24", "Diluted EPS $0.52"),
    ("EBITDA FY2025", "$116.6M", "Yahoo Finance /PAY/financials/", "2026-06-24", "From income statement"),
    ("Cash Flow TTM", "$142.1M OCF", "Yahoo Finance /PAY/cash-flow/", "2026-06-24", "Operating cash flow TTM"),
    ("Capex TTM", "$37.3M", "Yahoo Finance /PAY/cash-flow/", "2026-06-24", "Capital expenditure TTM"),
    ("FCF TTM", "$104.8M", "Yahoo Finance /PAY/cash-flow/", "2026-06-24", "Operating CF $142.1M less Capex $37.3M"),
    ("Total Assets FY2025", "$576.2M", "Yahoo Finance /PAY/balance-sheet/", "2026-06-24", "Balance sheet annual"),
    ("Total Debt FY2025", "$6.9M", "Yahoo Finance /PAY/balance-sheet/", "2026-06-24", "Capital lease obligations, declining"),
    ("Total Equity FY2025", "$560.4M", "Yahoo Finance /PAY/balance-sheet/", "2026-06-24", "Book value per share $4.64"),
    ("Beta (5Y)", "1.31", "Yahoo Finance Statistics", "2026-06-23", "5-year monthly beta"),
    ("52W High/Low", "$39.38 / $20.11", "Yahoo Finance Statistics", "2026-06-23", "Near 52W low on current day"),
    ("Trailing P/E", "36.1x", "Yahoo Finance Statistics", "2026-06-23", "TTM EPS $0.57"),
    ("P/B Ratio", "4.4x", "Yahoo Finance Statistics", "2026-06-23", "Per share basis"),
    ("FY2026 Rev Estimate", "$1.43B", "Yahoo Finance /PAY/analysis/", "2026-06-24", "7 analyst avg; low $1.43B, high $1.44B"),
    ("FY2026 EPS Estimate", "$0.83", "Yahoo Finance /PAY/analysis/", "2026-06-24", "7 analyst avg; low $0.69, high $0.89; non-GAAP/normalized"),
    ("FY2027 Rev Estimate", "$1.69B", "Yahoo Finance /PAY/analysis/", "2026-06-24", "Year/est growth 18.1%"),
    ("FY2027 EPS Estimate", "$1.02", "Yahoo Finance /PAY/analysis/", "2026-06-24", "Year/est growth 23.2%"),
    ("10Y Treasury Rate", "4.48%", "CNBC US10Y", "2026-06-23", "Yield 4.479% at 4:23 AM EDT"),
]

for i, (dp, val, src, dt, note) in enumerate(audit_data, 3):
    ws4.cell(row=i, column=1, value=dp)
    ws4.cell(row=i, column=2, value=val)
    ws4.cell(row=i, column=3, value=src)
    ws4.cell(row=i, column=4, value=dt)
    ws4.cell(row=i, column=5, value=note)

style_data_range(ws4, 3, 2 + len(audit_data), 5)
ws4.column_dimensions['A'].width = 25
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 40
ws4.column_dimensions['D'].width = 15
ws4.column_dimensions['E'].width = 50

# ============================================================
# Sheet 5: Questions
# ============================================================
ws5 = wb.create_sheet("Questions")
ws5.merge_cells('A1:C1')
ws5['A1'] = "Open Questions And Due Diligence Items"
ws5['A1'].font = Font(bold=True, size=12)

questions = [
    ("1.", "Revenue Recognition Model", "PAY processes payments for other companies (not a card network). Understand the pass-through nature — how much of the $1.28B revenue is billings vs net revenue. If most is passed through to merchants/banks, the gross margin of 24.8% may include significant third-party clearing costs."),
    ("2.", "Customer Concentration", "PAY works with processors, banks, and fintechs. What is the concentration risk if one or two major clients left? Revenue grew 30% YoY TTM — is this organic or acquisition-driven?"),
    ("3.", "Share Count Dilution", "Shares grew from 123.5M (FY2023) to 125.6M (Q1 FY2026) — modest dilution via stock-based compensation. No buybacks or significant share repurchases observed. Is SBC diluting owners at a meaningful rate?"),
    ("4.", "Operating Leverage Sustainability", "Operating income surged from $18.1M (FY2023) to $75.5M (FY2025) — a 4x improvement on ~37% revenue growth. Much of this is real operating leverage in the processing network, but how much is one-time (SBC write-offs, restructuring)?"),
    ("5.", "Valuation Premium", "At 36x trailing and 25x forward P/E, PAY trades at a premium to the market. The stock fell 40% from the 52-week high — has the multiple already compressed to fair value, or is there more downside?"),
    ("6.", "Competitive Position", "PAY operates in B2B payment processing infrastructure (ISO/processor enablement). Does it face threat from Stripe, Adyen, or larger processors moving downstream? What is the switching cost for its clients?"),
    ("7.", "Cash Position and Capital Allocation", "Cash grew from $149.7M (FY2022) to $324.5M (FY2025) — a strong accumulation with minimal debt. Management has not been returning capital. Will they buy back shares, pay dividends, or pursue acquisitions?"),
    ("8.", "Q1 2026 EPS Beat", "EPS actual $0.21 vs estimate $0.18, a 19.8% beat. This is the 4th consecutive quarter of beating estimates. Is the market fully repricing in continued beats, or are estimates still too conservative?"),
    ("9.", "Guidance Reliability", "PAY has not issued formal public guidance in visible sources. Analysts are deriving estimates independently. This creates a revision risk if management's internal projections differ materially."),
    ("10.", "EBITDA vs Operating Income Gap", "EBITDA $116.6M vs Op Income $75.5M in FY2025 implies $41.1M in D&A on $1.2B revenue. Verify the capitalization policy for software/technology investments — is the company capitalizing development costs that drive higher operating income?"),
]

for i, (num, topic, detail) in enumerate(questions, 2):
    ws5.cell(row=i, column=1, value=num).font = bold_font
    ws5.cell(row=i, column=2, value=topic).font = bold_font
    ws5.cell(row=i, column=3, value=detail)

ws5.column_dimensions['A'].width = 5
ws5.column_dimensions['B'].width = 30
ws5.column_dimensions['C'].width = 100

# ============================================================
# Sheet 6: Sources
# ============================================================
ws6 = wb.create_sheet("Sources")
ws6.merge_cells('A1:C1')
ws6['A1'] = "Data Sources"
ws6['A1'].font = Font(bold=True, size=12)

sources = [
    ("1.", "Yahoo Finance — PAY Quote and Statistics", "https://finance.yahoo.com/quote/PAY/"),
    ("2.", "Yahoo Finance — PAY Income Statement", "https://finance.yahoo.com/quote/PAY/financials/"),
    ("3.", "Yahoo Finance — PAY Balance Sheet", "https://finance.yahoo.com/quote/PAY/balance-sheet/"),
    ("4.", "Yahoo Finance — PAY Cash Flow Statement", "https://finance.yahoo.com/quote/PAY/cash-flow/"),
    ("5.", "Yahoo Finance — PAY Analyst Estimates", "https://finance.yahoo.com/quote/PAY/analysis/"),
    ("6.", "CNBC — US10Y Treasury Yield", "https://www.cnbc.com/quotes/US10Y"),
    ("7.", "StockAnalysis — PAY (404 — unavailable)", "https://stockanalysis.com/quotes/PAY/"),
    ("8.", "Yahoo Finance — PAY Q1 FY2026 Earnings Beat", "https://finance.yahoo.com/quote/PAY/analysis/"),
]

for i, (num, desc, url) in enumerate(sources, 2):
    ws6.cell(row=i, column=1, value=num)
    ws6.cell(row=i, column=2, value=desc)
    ws6.cell(row=i, column=3, value=url)

ws6.column_dimensions['A'].width = 5
ws6.column_dimensions['B'].width = 55
ws6.column_dimensions['C'].width = 65

# ============================================================
# Save
# ============================================================
outpath = "/home/refcell/dev/capital/models/[2026-06-24] Paymentus Holdings Model.xlsx"
wb.save(outpath)
print(f"Saved to {outpath}")

# Verify
from openpyxl import load_workbook
wb2 = load_workbook(outpath)
print(f"Sheets: {wb2.sheetnames}")
for sn in wb2.sheetnames:
    ws = wb2[sn]
    print(f"  {sn}: {ws.max_row} rows x {ws.max_column} cols")
wb2.close()
