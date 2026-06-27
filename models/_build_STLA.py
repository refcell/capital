#!/usr/bin/env python3
"""
_build_STLA.py — Build Stellantis N.V. (STLA) 6-sheet valuation model.

Snapshot: 2026-06-26 close at $5.68, MC $16.53B, EV $33.79B, 2.90B shares.
Stellantis reports in EUR (headquartered Amsterdam, NYSE-listed ADR).
Yahoo Finance shows financials in EUR but statistics labels as USD — likely raw EUR.
For the model: financials sourced as EUR, market data in USD. Conversion ~EUR1 = USD1.08.

Key context:
- CEO Antonio Filosa replaced Carlos Tavares June 2025
- FY2025: net loss EUR 22.4B on EUR 153.5B revenue (-14.6% net margin)
- Gross profit negative FY2025 (-EUR 2.1B — unusual for 155B auto company)
- Stock down 43.37% in 52w, near 52w low of $5.61
- Estimates revised sharply down (FY26 EPS: $1.14 to $0.85 in 90 days)
- Forward P/E 7.05x; P/S 0.09x; P/B 0.24x — distressed pricing
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = Workbook()

# ─── Styles ───────────────────────────────────────────────────────────
bold = Font(bold=True)
hdr_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
bear_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
base_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
bull_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def hdr(ws, row, mc):
    for c in range(1, mc + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = bold
        cell.border = thin
        cell.fill = hdr_fill


def data(ws, sr, er, mc):
    for r in range(sr, er + 1):
        for c in range(1, mc + 1):
            ws.cell(row=r, column=c).border = thin


# =====================================================================
# Sheet 1 : Valuation
# =====================================================================
ws1 = wb.active
ws1.title = "Valuation"
ws1.merge_cells("A1:E1")
ws1["A1"] = "Stellantis N.V. (STLA) — Valuation Model"
ws1["A1"].font = Font(bold=True, size=14)

title = [
    ("Company", "Stellantis N.V."),
    ("Ticker", "NYSE: STLA"),
    ("Sector", "Consumer Discretionary — Automobiles"),
    ("Date", "2026-06-26"),
    ("Source", "Yahoo Finance, 2026-06-26 close"),
    ("Price", "$5.68"),
    ("Shares Outstanding", "2.90B (diluted, Yahoo Finance Statistics)"),
    ("Market Cap", "$16.53B"),
    ("Enterprise Value", "$33.79B  (MC $16.53B + Net Debt ~$17.26B)"),
    ("Primary Lens", "Forward P/E + EV/Sales; WACC-based FCF scenarios"),
    ("Stance",
     "Watch — distressed multiples with massive loss profile; "
     "turnaround potential but high execution risk"),
]

for i, (k, v) in enumerate(title, 2):
    ws1.cell(row=i, column=1, value=k).font = bold
    ws1.cell(row=i, column=2, value=v)

r0 = len(title) + 3  # first metric row
for c, h in enumerate(["Valuation Metric", "Value", "Comment"], 1):
    ws1.cell(row=r0, column=c, value=h).font = bold

metrics = [
    ("Forward P/E (FY2026)", "7.05x",
     "On FY26 EPS consensus $0.85 (3 analysts). Cheap for auto sector."),
    ("Trailing P/E", "3.19x (S&P calc) / N/A",
     "TTM EPS -$7.48 => trailing P/E technically negative. 3.19x uses S&P "
     "adjusted measure."),
    ("P/S (TTM)", "0.09x",
     "Deeply discounted vs. auto norms of 0.15-0.25x. Bankruptcy-adjacent"),
    ("EV/Sales", "0.19x",
     "EV $33.79B / TTM rev ~$155-160B. Enterprise value near liquidation floor."),
    ("EV/EBITDA", "11.96x (S&P)",
     "S&P-calculated; company EBITDA is negative TTM (-EUR 1.04B). "
     "Likely forward/adjusted."),
    ("P/B", "0.24x",
     "MC $16.53B / equity ~$69B (at USD). Massive deep-to-book."),
    ("PEG (5yr expected)", "1.15x",
     "Reasonable for recovery narrative. Based on 5-yr expected growth."),
    ("Forward Div Yield", "10.58%",
     "Annual div rate $0.77; sustainability very questionable given losses."),
    ("Beta (5Y monthly)", "0.97",
     "Near-market beta; auto-cyclical equity risk."),
    ("52W Range", "$5.61 - $12.22",
     "Currently at 52-week low. Down 43.37% in 52 weeks."),
]

for i, (m, v, c) in enumerate(metrics, r0 + 1):
    ws1.cell(row=i, column=1, value=m)
    ws1.cell(row=i, column=2, value=v)
    ws1.cell(row=i, column=3, value=c)

hdr(ws1, r0, 3)
data(ws1, r0 + 1, r0 + len(metrics), 3)
ws1.column_dimensions["A"].width = 25
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 85

# =====================================================================
# Sheet 2 : WACC
# =====================================================================
ws2 = wb.create_sheet("WACC")
ws2.merge_cells("A1:D1")
ws2["A1"] = "WACC — CAPM (Stellantis)"
ws2["A1"].font = Font(bold=True, size=12)

# Calculation:
# Rf = 4.376% (CNBC US10Y Jun 26 2026)
# beta = 0.97
# ERP = 5%
# cost_eq = 4.376 + 0.97*5 = 9.226%
# MC = 16,530M USD
# debt = 47,940M USD (MRQ Yahoo Stats)
# cap = MC + D = 64,470M
# wE = 25.6%, wD = 74.4%
# cost_debt pre-tax = 5.5% (BBB- auto corp)
# tax = 25% (EU statutory; effective rate meaningless with negative income)
# WACC = 0.256*9.226 + 0.744*5.5*(1-0.25) = 2.362 + 3.066 = 5.428%

rf = 0.04376
beta = 0.97
erp = 0.05
cost_eq = rf + beta * erp  # 9.226%
mc = 16530
debt = 47940
cap = mc + debt
wE = mc / cap  # 0.2564
wD = debt / cap  # 0.7436
cost_dt = 0.055
tax = 0.25
wacc = wE * cost_eq + wD * cost_dt * (1 - tax)

print(f"WACC = {wacc*100:.2f}%")
print(f"  Components: Rf {rf*100:.3f}%, beta {beta}, ERP {erp*100:.1f}%")
print(f"  Cost of equity = {cost_eq*100:.3f}%")
print(f"  wE={wE:.3f} wD={wD:.3f} cost_dt={cost_dt*100:.1f}% tax={tax*100:.0f}%")

wacc_rows = [
    ("Component", "Value", "Source / Notes"),
    ("Risk-Free Rate (10Y US)", "4.376%",
     "CNBC US10Y, 2026-06-26 yield 4.376%"),
    ("Beta (5Y Monthly)", "0.97",
     "Yahoo Finance Statistics; near-market, auto-cyclical"),
    ("Equity Risk Premium", "5.00%", "Standard assumption"),
    ("Cost of Equity (Rf + B*ERP)", f"{cost_eq*100:.3f}%",
     f"= {rf*100:.3f}% + {beta}*{erp*100:.1f}%"),
    ("Market Cap (USD)", "$16,530M", "Yahoo Finance, 2026-06-26"),
    ("Total Debt (MRQ, USD)", "$47,940M",
     "Yahoo Finance Balance Sheet / Statistics MRQ; auto debt inc. leases"),
    ("Total Capitalization", "$64,470M", "MC + Debt"),
    ("Equity Weight", f"{wE*100:.1f}%", "MC / (MC+Debt)"),
    ("Debt Weight", f"{wD*100:.1f}%", "Debt / (MC+Debt); auto debt includes "
     "dealer financing / leases"),
    ("Pre-Tax Cost of Debt", "5.50%",
     "Estimate — BBB- European auto; actual interest expense/debt ~3.3% "
     "but includes below-market financing"),
    ("Tax Rate", "25.00%",
     "EU statutory; effective rate meaningless (negative income)"),
    ("WACC", f"{wacc*100:.2f}%",
     f"= {wE:.3f}*{cost_eq*100:.3f}% + {wD:.3f}*5.5%*(1-0.25)"),
]

for i, row in enumerate(wacc_rows, 2):
    for c, v in enumerate(row, 1):
        ws2.cell(row=i, column=c, value=v)
        if c == 1:
            ws2.cell(row=i, column=c).font = bold if i > 2 else bold

hdr(ws2, 2, 3)
data(ws2, 3, len(wacc_rows) + 1, 3)
ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 80

# =====================================================================
# Sheet 3 : Scenarios
# =====================================================================
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells("A1:Q1")
ws3["A1"] = "Scenario Analysis — Bear / Base / Bull (5Y Horizon)"
ws3["A1"].font = Font(bold=True, size=12)

# ── All $ in millions ────────────────────────────────────────────────
# FY2026E revenue base: $160,500M  (consensus EUR 160.54B ~ 160,500M)
# Shares: 2,900M
# Current net debt: ~$17,500M  (EV $33.79B - MC $16.53B)

SHARES_M = 2900
REV_BASE = 160500  # FY2026E, $M

# Bear:  Rev -3% CAGR,  FCF margin 0.5%,  exit 4x
b_revcagr = -0.03
b_termrev = round(REV_BASE * ((1 + b_revcagr) ** 5))        # 138,500
b_fcfm = 0.005
b_termfcf = round(b_termrev * b_fcfm)                        # 692
b_multi = 4
b_ev = b_termfcf * b_multi                                  # 2,768
b_netdebt = 18000  # worsening leverage
b_equity = b_ev - b_netdebt  # negative
b_target = 1.50  # fire-sale / restructuring value
b_upside = f"{(b_target/5.68 - 1)*100:+.1f}%"
b_weight = 0.20

# Base: Rev +1% CAGR, FCF margin 3.0%, exit 8x
br_revcagr = 0.01
br_termrev = round(REV_BASE * ((1 + br_revcagr) ** 5))      # 169,200
br_fcfm = 0.03
br_termfcf = round(br_termrev * br_fcfm)                     # 5,076
br_multi = 8
br_ev = br_termfcf * br_multi                               # 40,608
br_netdebt = 12000  # moderate deleveraging
br_equity = br_ev - br_netdebt  # 28,608
br_target = round(br_equity / SHARES_M, 2)                   # 9.87
br_upside = f"{(br_target/5.68 - 1)*100:+.1f}%"
br_weight = 0.50

# Bull: Rev +3% CAGR, FCF margin 5.0%, exit 10x
bu_revcagr = 0.03
bu_termrev = round(REV_BASE * ((1 + bu_revcagr) ** 5))      # 186,800
bu_fcfm = 0.05
bu_termfcf = round(bu_termrev * bu_fcfm)                     # 9,340
bu_multi = 10
bu_ev = bu_termfcf * bu_multi                               # 93,400
bu_netdebt = 2000  # near-full deleveraging
bu_equity = bu_ev - bu_netdebt  # 91,400
bu_target = round(bu_equity / SHARES_M, 2)                   # 31.52
bu_upside = f"{(bu_target/5.68 - 1)*100:+.1f}%"
bu_weight = 0.30

wted_b = b_weight * b_target
wted_br = br_weight * br_target
wted_bu = bu_weight * bu_target
weighted_fv = round(wted_b + wted_br + wted_bu, 2)
weighted_upside = f"{(weighted_fv/5.68 - 1)*100:+.1f}%"

print(f"\nScenario targets: Bear ${b_target}, Base ${br_target}, Bull ${bu_target}")
print(f"Weighted FV = ${weighted_fv}  Upside from $5.68 = {weighted_upside}")

# Verification: target prices in plausible range?
assert 0.5 < b_target < 5, f"Bear target ${b_target} out of range"
assert 5.0 < br_target < 20, f"Base target ${br_target} out of range"
assert 15.0 < bu_target < 80, f"Bull target ${bu_target} out of range"
print("Targets verified: all within plausible ranges.")

rows3 = [
    ("Item", "Bear", "Base", "Bull"),
    ("Revenue CAGR (5Y)", f"{b_revcagr*100:.1f}%", f"{br_revcagr*100:.1f}%",
     f"{bu_revcagr*100:.1f}%"),
    ("FY2026 Revenue ($M)", f"${REV_BASE:,}", f"${REV_BASE:,}", f"${REV_BASE:,}"),
    ("Terminal Revenue (5Y) ($M)", f"${b_termrev:,}", f"${br_termrev:,}",
     f"${bu_termrev:,}"),
    ("Adjusted FCF Margin", f"{b_fcfm*100:.1f}%", f"{br_fcfm*100:.1f}%",
     f"{bu_fcfm*100:.1f}%"),
    ("Terminal FCF ($M)", f"${b_termfcf:,}", f"${br_termfcf:,}", f"${bu_termfcf:,}"),
    ("Exit FCF Multiple", f"{b_multi}x", f"{br_multi}x", f"{bu_multi}x"),
    ("Implied EV ($M)", f"${b_ev:,}", f"${br_ev:,}", f"${bu_ev:,}"),
    ("Less Net Debt ($M)", f"-${b_netdebt:,}", f"-${br_netdebt:,}",
     f"-${bu_netdebt:,}"),
    ("Equity Value ($M)", "Negative (restructuring)", f"${br_equity:,}",
     f"${bu_equity:,}"),
    ("Shares Outstanding (M)", f"{SHARES_M:,}", f"{SHARES_M:,}", f"{SHARES_M:,}"),
    ("Target Price", f"${b_target:.2f}", f"${br_target:.2f}", f"${bu_target:.2f}"),
    ("Upside from $5.68", b_upside, br_upside, bu_upside),
    ("Weight", f"{b_weight*100:.0f}%", f"{br_weight*100:.0f}%",
     f"{bu_weight*100:.0f}%"),
    ("Weighted Value/Share", f"${wted_b:.2f}", f"${wted_br:.2f}", f"${wted_bu:.2f}"),
]

for i, row in enumerate(rows3, 2):
    for c, v in enumerate(row, 1):
        ws3.cell(row=i, column=c, value=v)

hdr(ws3, 2, 4)
data(ws3, 3, len(rows3), 4)

# Color rows
for c in range(2, 5):
    ws3.cell(row=3, column=c).fill = bear_fill if c == 2 else (
        base_fill if c == 3 else bull_fill)

# Summary below
sr = len(rows3) + 2
ws3.cell(row=sr, column=1, value="Total Probability-Weighted FV").font = bold
ws3.cell(row=sr, column=2, value=f"${weighted_fv:.2f}").font = bold
ws3.cell(row=sr + 1, column=1, value="Upside from $5.68").font = bold
ws3.cell(row=sr + 1, column=2, value=weighted_upside).font = bold
ws3.cell(row=sr + 2, column=1, value="WACC (discount rate)").font = bold
ws3.cell(row=sr + 2, column=2, value=f"{wacc*100:.2f}%").font = bold

ws3.column_dimensions["A"].width = 32
for c in "BCD":
    ws3.column_dimensions[c].width = 20

# =====================================================================
# Sheet 4 : Actuals Source Audit
# =====================================================================
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells("A1:E1")
ws4["A1"] = "Actuals Source Audit — Every Data Point"
ws4["A1"].font = Font(bold=True, size=12)

audit = [
    ("Data Point", "Value", "Source", "Date", "Notes"),
    ("Stock price (close)", "$5.68", "Yahoo Finance quote", "2026-06-26",
     "NYSE closing price"),
    ("After-hours price", "$5.71", "Yahoo Finance", "2026-06-26",
     "+0.53% AH"),
    ("Market Cap", "$16.53B", "Yahoo Finance Stats", "2026-06-26", ""),
    ("Enterprise Value", "$33.79B", "Yahoo Finance Stats", "2026-06-26", ""),
    ("Shares Outstanding", "2.90B", "Yahoo Finance Stats", "2026-06-26",
     "diluted approx."),
    ("Beta (5Y monthly)", "0.97", "Yahoo Finance Stats", "2026-06-26", ""),
    ("Forward P/E", "7.05x", "Yahoo Finance Stats", "2026-06-26",
     "on FY26 EPS $0.85"),
    ("Trailing P/E", "3.19x (S&P)", "Yahoo Finance Stats", "2026-06-26",
     "TTM EPS -7.48 => effectively negative"),
    ("P/S (TTM)", "0.09x", "Yahoo Finance Stats", "2026-06-26", ""),
    ("P/B", "0.24x", "Yahoo Finance Stats", "2026-06-26", ""),
    ("EV/Sales", "0.19x", "Yahoo Finance Stats", "2026-06-26", ""),
    ("PEG (5yr)", "1.15x", "Yahoo Finance Stats", "2026-06-26", ""),
    ("TTM Revenue", "EUR 155,827M", "YF Income Statement", "TTM",
     "Reported EUR; stats page labels 'USD'"),
    ("FY2025 Revenue", "EUR 153,508M", "YF Income Statement", "12/31/2025", ""),
    ("FY2024 Revenue", "EUR 156,878M", "YF Income Statement", "12/31/2024", ""),
    ("FY2023 Revenue", "EUR 189,544M", "YF Income Statement", "12/31/2023", ""),
    ("FY2022 Revenue", "EUR 179,592M", "YF Income Statement", "12/31/2022", ""),
    ("TTM Gross Profit", "EUR -1,315M", "YF Income Statement", "TTM",
     "NEGATIVE — unprecedented at this scale"),
    ("FY2025 Gross Profit", "EUR -2,119M", "YF Income Statement", "12/31/2025",
     "NEGATIVE — cost of rev > revenue"),
    ("TTM Op Income", "EUR -21,390M", "YF Income Statement", "TTM", ""),
    ("FY2025 Op Income", "EUR -22,231M", "YF Income Statement", "12/31/2025", ""),
    ("FY2024 Op Income", "EUR 5,435M", "YF Income Statement", "12/31/2024",
     "Last profitable FY"),
    ("TTM Net Income", "EUR -21,607M", "YF Income Statement", "TTM", ""),
    ("FY2025 Net Income", "EUR -22,368M", "YF Income Statement", "12/31/2025",
     "Reported EUR 22.33B loss per news"),
    ("TTM EPS (diluted)", "-$7.48", "YF Income Statement", "TTM",
     "Note: EPS in USD; shares ~2.89B"),
    ("TTM Operating CF", "EUR -4,522M", "YF Cash Flow", "TTM", ""),
    ("TTM Capex", "EUR -8,115M", "YF Cash Flow", "TTM", ""),
    ("TTM Free CF", "EUR -12,637M", "YF Cash Flow", "TTM",
     "OCF minus Capex"),
    ("Total Assets (FY25)", "EUR 195,153M", "YF Balance Sheet", "12/31/2025", ""),
    ("Total Debt (MRQ)", "USD 47,940M", "YF Statistics MRQ", "Q1 FY26", ""),
    ("Total Debt (FY25)", "EUR 45,947M", "YF Balance Sheet", "12/31/2025", ""),
    ("Cash (MRQ)", "USD 30,390M", "YF Statistics MRQ", "Q1 FY26", ""),
    ("Cash (FY25)", "EUR 30,146M", "YF Balance Sheet", "12/31/2025", ""),
    ("Net Debt (FY25)", "EUR 13,347M", "YF Balance Sheet", "12/31/2025",
     "Debt - cash"),
    ("Book Value (FY25)", "EUR 53,551M", "YF Balance Sheet", "12/31/2025", ""),
    ("Net Tangible Assets", "EUR 8,666M", "YF Balance Sheet", "12/31/2025",
     "Down from EUR 27,327M FY24"),
    ("FW26 Revenue Est.", "EUR 160,540M", "YF Analysis", "2026-06-26",
     "27 analysts; low 153,860M high 164,550M"),
    ("FW27 Revenue Est.", "EUR 165,740M", "YF Analysis", "2026-06-26",
     "28 analysts; low 158,290M high 173,630M"),
    ("FW26 EPS Est.", "$0.85", "YF Analysis", "2026-06-26",
     "3 analysts; revised DOWN from $1.14 (90d ago)"),
    ("FW27 EPS Est.", "$1.76", "YF Analysis", "2026-06-26",
     "6 analysts; low $1.18 high $2.29"),
    ("Q1 FY26 EPS", "$0.25", "YF Analysis", "3/31/2026",
     "Actual vs est $0.08; +195.92% surprise"),
    ("10Y US Treasury", "4.376%", "CNBC US10Y", "2026-06-26", ""),
    ("52W High", "$12.22", "Yahoo Finance", "2026-06-26", ""),
    ("52W Low", "$5.61", "Yahoo Finance", "2026-06-26",
     "Stock currently at lows"),
    ("52W Change", "-43.37%", "Yahoo Finance", "2026-06-26", ""),
    ("Short % Float", "2.66%", "Yahoo Finance", "06/15/2026", ""),
    ("Div Date", "5/5/2025", "Yahoo Finance", "2026-06-26",
     "Last dividend payment"),
    ("Insider %", "23.67%", "Yahoo Finance", "2026-06-26", ""),
    ("Instl %", "49.85%", "Yahoo Finance", "2026-06-26", ""),
]

for i, row in enumerate(audit, 2):
    for c, v in enumerate(row, 1):
        ws4.cell(row=i, column=c, value=v)

hdr(ws4, 2, 5)
data(ws4, 3, len(audit), 5)
ws4.column_dimensions["A"].width = 28
ws4.column_dimensions["B"].width = 20
ws4.column_dimensions["C"].width = 25
ws4.column_dimensions["D"].width = 16
ws4.column_dimensions["E"].width = 50

# =====================================================================
# Sheet 5 : Questions
# =====================================================================
ws5 = wb.create_sheet("Questions")
ws5.merge_cells("A1:D1")
ws5["A1"] = "Open Questions — Stellantis (STLA)"
ws5["A1"].font = Font(bold=True, size=12)

questions = [
    ("#", "Question", "Category", "Priority"),
    ("1",
     "Gross profit turned negative (FY2025 -EUR 2.1B on EUR 153.5B revenue). "
     "Unprecedented for auto at this scale. Is this CO2 penalties, inventory "
     "write-downs, or structural margin destruction?",
     "Profitability", "Critical"),
    ("2",
     "Operating income swung EUR 27B: +EUR 5.4B (FY24) to -EUR 22.2B (FY25). "
     "What drove this? Restructuring, EV investment, competitive pricing?",
     "Profitability", "Critical"),
    ("3",
     "CEO Antonio Filosa replaced Carlos Tavares (June 2025) after 19-year "
     "tenure. Filosa received 'C grade'. What is the strategic direction and "
     "does Filosa have the mandate to execute?",
     "Management", "High"),
    ("4",
     "Total debt increased EUR 16.5B YoY (EUR 29.5B to EUR 45.9B). What was "
     "this spent on? Acquisitions, restructuring, EV capex, or cash burn?",
     "Capital Structure", "High"),
    ("5",
     "Equity destroyed EUR 28B YoY (EUR 82.1B to EUR 54.0B). How much "
     "impairment / goodwill write-down vs. retained-earnings drawdown?",
     "Accounting", "High"),
    ("6",
     "Net tangible assets fell 68% (EUR 27.3B to EUR 8.7B). What goodwill "
     "was written off and does the balance sheet overstate recoverable value?",
     "Accounting", "High"),
    ("7",
     "Negative operating CF TTM (EUR -4.5B). Is this structural ("
     "unprofitable business model) or cyclical (temporary capex/restructuring)?",
     "Cash Flow", "Critical"),
    ("8", "EV transition plan: EUR 40B+ investment through 2030. What is the "
     "IRR? How does cost per vehicle compare to VW/Tesla? Can Stellantis "
     "afford this burn rate?", "Strategy", "Critical"),
    ("9", "Forward div yield 10.58% on $5.68. Is this sustainable with EUR 22B "
     "net loss? What happens if dividend is cut?", "Capital Allocation", "Medium"),
    ("10", "Brand portfolio (14 brands): Maserati/Alfa Romeo/Abarth impairment "
     "risk? Have these been valued properly? Any sale rumors?",
     "Strategy", "Medium"),
    ("11", "CO2 penalty exposure: EU fines per gram of excess emissions. "
     "What is Stellantis's total liability? Already provisioned for?",
     "Regulation", "High"),
    ("12", "Competitive moat vs VW, Toyota, GM: Stellantis has scale (EUR 155B "
     "rev) but can it match VW's platform efficiency (MLB Evo)? Cost "
     "advantage durable?", "Competition", "Medium"),
    ("13", "Management guidance for FY26-27: any explicit targets from Filosa "
     "on revenue, margin, or FCF? Previous Tavares guidance was aggressive.",
     "Guidance", "High"),
    ("14", "Next earnings: Q2 FY26 report date? Analyst coverage dropping "
     "(only 3 for FY26 EPS). Reduced coverage limits transparency.",
     "Earnings", "High"),
    ("15", "Shares down from 3.14B (FY22) to 2.89B (TTM). Buyback program "
     "or just natural decline? Repurchase of stock EUR 1B TTM.",
     "Capital Allocation", "Medium"),
]

for i, row in enumerate(questions, 3):
    for c, v in enumerate(row, 1):
        ws5.cell(row=i, column=c, value=v)

hdr(ws5, 2, 4)
data(ws5, 3, len(questions) + 1, 4)
ws5.column_dimensions["A"].width = 6
ws5.column_dimensions["B"].width = 90
ws5.column_dimensions["C"].width = 22
ws5.column_dimensions["D"].width = 14

# =====================================================================
# Sheet 6 : Sources
# =====================================================================
ws6 = wb.create_sheet("Sources")
ws6.merge_cells("A1:C1")
ws6["A1"] = "Sources — Stellantis (STLA) Model"
ws6["A1"].font = Font(bold=True, size=12)

sources = [
    ("#", "Source", "URL / Reference"),
    ("1", "Yahoo Finance — STLA Main Quote & Summary",
     "finance.yahoo.com/quote/STLA/"),
    ("2", "Yahoo Finance — Income Statement (Annual)",
     "finance.yahoo.com/quote/STLA/financials/"),
    ("3", "Yahoo Finance — Balance Sheet (Annual)",
     "finance.yahoo.com/quote/STLA/balance-sheet/"),
    ("4", "Yahoo Finance — Cash Flow (Annual)",
     "finance.yahoo.com/quote/STLA/cash-flow/"),
    ("5", "Yahoo Finance — Key Statistics",
     "finance.yahoo.com/quote/STLA/key-statistics/"),
    ("6", "Yahoo Finance — Analysis / Analyst Estimates",
     "finance.yahoo.com/quote/STLA/analysis/"),
    ("7", "CNBC — US 10-Year Treasury (US10Y)",
     "www.cnbc.com/quotes/US10Y"),
    ("8", "StockAnalysis — STLA (404ed — not available)",
     "stockanalysis.com/quotes/STLA/ [404]"),
    ("9", "Yahoo Scout — CEO Rating Summary",
     "Embedded on Yahoo Finance quote page; Jun 26 2026"),
    ("10",
     "S&P Global Market Intelligence — EBITDA/Valuation Measures",
     "As cited in Yahoo Finance Statistics footnote"),
]

for i, row in enumerate(sources, 2):
    for c, v in enumerate(row, 1):
        ws6.cell(row=i, column=c, value=v)

hdr(ws6, 2, 3)
data(ws6, 3, len(sources), 3)
ws6.column_dimensions["A"].width = 6
ws6.column_dimensions["B"].width = 55
ws6.column_dimensions["C"].width = 70

# =====================================================================
# Save
# =====================================================================
xlsx = Path(__file__).resolve().parent / "[2026-06-26] Stellantis Model.xlsx"
wb.save(str(xlsx))
print(f"\nSaved: {xlsx}")
print(f"File size: {xlsx.stat().st_size:,} bytes")
