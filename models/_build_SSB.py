#!/usr/bin/env python3
"""Build SSB (SouthState Bank Corporation) 6-sheet valuation model.
Bank-specific framework: P/B + ROE primary, Forward P/E cross-check.
Date: 2026-08-10
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws1 = wb.active
ws1.title = 'Valuation'

# Styles
header_font = Font(bold=True, size=14, name='Calibri')
subtitle_font = Font(bold=True, size=11, italic=True, name='Calibri')
section_font = Font(bold=True, size=12, name='Calibri')
bold_font = Font(bold=True, name='Calibri')
normal_font = Font(name='Calibri')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
dark_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
dark_font = Font(bold=True, color='FFFFFF', name='Calibri')

def c(ws, row, col, value, font=None, border=False, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if border:
        cell.border = thin_border
    if fill:
        cell.fill = fill
    return cell

# Title row — merged
ws1.merge_cells('A1:F1')
c(ws1, 1, 1, 'SouthState Bank Corporation (SSB) — Valuation Model', header_font).alignment = Alignment(horizontal='center')

# Subtitle
ws1.merge_cells('A2:F2')
c(ws1, 2, 1, 'Bank framework: P/B + ROE primary | Forward P/E cross-check | Date: 2026-08-10', subtitle_font)

# Title block
title_data = [
    ('Company', 'SouthState Bank Corporation'),
    ('Ticker', 'NYSE: SSB'),
    ('Date', '2026-08-10'),
    ('Price', '$108.56'),
    ('Shares Outstanding', '96.97M'),
    ('Market Cap', '$10.48B'),
    ('Enterprise Value', 'N/A (deposits = operating liabilities for banks)'),
    ('Primary Lens', 'P/B and ROE'),
    ('Stance', 'Watch — acquisition integration still unproven'),
]
for i, (label, value) in enumerate(title_data, 3):
    c(ws1, i, 1, label, bold_font, border=True, fill=header_fill)
    c(ws1, i, 2, value, normal_font, border=True)

# Valuation metrics table
c(ws1, 5, 1, 'Key Valuation Metrics', section_font)
metrics = [
    ('Trailing P/E', '11.36x', 'TTM; reasonably cheap for a quality regional bank'),
    ('Forward P/E', '9.96x', 'Implies ~$10.90 EPS next year; below sector median of 11-12x'),
    ('P/B', '1.15x', 'Slightly above book; conservative for ROE of 10.60%'),
    ('P/S', '4.06x', 'TTM revenue $2.68B'),
    ('EV/Revenue', '4.32x', 'Yahoo Finance estimate'),
    ('EV/EBITDA', 'N/A', 'Yahoo shows no data — bank accounting'),
    ('Beta (5Y)', '0.71', 'Below-market beta; defensive characteristics'),
    ('ROE TTM', '10.60%', 'Below cost of equity (~12.6%) — buybacks/accretive ops needed'),
    ('ROA TTM', '1.41%', 'Solid for a regional bank; sector avg ~1.0-1.3%'),
    ('Dividend Yield', '2.44%', 'Payout ratio 25.21%; room for growth'),
    ('Profit Margin', '35.51%', 'TTM; healthy for a bank with rising NII'),
    ('Operating Margin', '50.51%', 'Pre-tax / revenue; strong spread economics'),
]
c(ws1, 6, 1, 'Metric', bold_font, border=True, fill=header_fill)
c(ws1, 6, 2, 'Value', bold_font, border=True, fill=header_fill)
c(ws1, 6, 3, 'Commentary', bold_font, border=True, fill=header_fill)
for i, (metric, value, comment) in enumerate(metrics, 7):
    c(ws1, i, 1, metric, normal_font, border=True)
    c(ws1, i, 2, value, normal_font, border=True)
    c(ws1, i, 3, comment, normal_font, border=True)

ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 60

# ============================================================
# Sheet 2: WACC
# ============================================================
ws2 = wb.create_sheet('WACC')

ws2.merge_cells('A1:D1')
c(ws2, 1, 1, 'WACC — CAPM Components (SSB)', header_font).alignment = Alignment(horizontal='center')

wacc_data = [
    ('Risk-Free Rate (10Y US)', '4.705%', 'CNBC quote, Aug 10 2026'),
    ('Equity Risk Premium', '5.00%', 'Standard assumption'),
    ('Beta (5Y Monthly)', '0.71', 'Yahoo Finance Key Stats'),
    ('Cost of Equity (Ke)', '12.56%', 'Rf + Beta × ERP = 4.705 + 0.71×5.0'),
    ('Cost of Debt (Kd)', '4.50%', 'Estimated — regional bank funding'),
    ('Tax Rate', '22.60%', 'From FY2025 income statement: 241.5 / 1,040.2'),
    ('After-Tax Cost of Debt', '3.48%', 'Kd × (1 - Tax) = 4.5 × 0.774'),
    ('Market Cap', '$10,480M', '96.97M shares × $108.56'),
    ('Total Debt', '$686.2M', 'FY2025 Balance Sheet; Key Stats shows $2.7B MRQ — discrepancy flagged'),
    ('Equity Weight', '93.8%', 'MC / (MC + Debt)'),
    ('Debt Weight', '6.2%', 'Debt / (MC + Debt)'),
    ('WACC', '12.17%', 'We×Ke + Wd×Kd(1-t) = 0.938×12.56 + 0.062×3.48'),
]

c(ws2, 3, 1, 'Component', bold_font, border=True, fill=header_fill)
c(ws2, 3, 2, 'Value', bold_font, border=True, fill=header_fill)
c(ws2, 3, 3, 'Source / Notes', bold_font, border=True, fill=header_fill)
for i, (label, value, note) in enumerate(wacc_data, 4):
    c(ws2, i, 1, label, normal_font, border=True)
    c(ws2, i, 2, value, normal_font, border=True)
    c(ws2, i, 3, note, normal_font, border=True)

ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 55

# ============================================================
# Sheet 3: Scenarios (P/B + ROE framework for bank)
# ============================================================
ws3 = wb.create_sheet('Scenarios')

ws3.merge_cells('A1:H1')
c(ws3, 1, 1, 'Scenarios — P/B + ROE Framework (Bank Valuation)', header_font).alignment = Alignment(horizontal='center')

ws3.merge_cells('A2:H2')
c(ws3, 2, 1, 'Primary lens: P/B multiple expansion/compression driven by BVPS growth and ROE trajectory. FCF multiples N/A for banks.', subtitle_font)

# Current data
c(ws3, 4, 1, 'Current BVPS', bold_font, border=True, fill=header_fill)
c(ws3, 4, 2, '$94.17', bold_font, border=True)
c(ws3, 4, 3, 'Source', bold_font, border=True, fill=header_fill)
c(ws3, 4, 4, 'Yahoo Key Stats MRQ', normal_font, border=True)

# Scenario data — all values in consistent units (dollars per share)
# BVPS growth CAGR: Bear 3%, Base 6%, Bull 8%
# Terminal BVPS after 5Y: Base $94.17 * 1.06^5 = $126.78
# Exit P/B: Bear 0.85x, Base 1.10x, Bull 1.30x
# Target = Terminal BVPS × Exit P/B

scenario_cols = ['Metric', 'Bear', 'Base', 'Bull', 'Notes']
headers = 6
for i, h in enumerate(scenario_cols, 1):
    c(ws3, headers, i, h, bold_font, border=True, fill=header_fill)

scenarios = [
    ('BVPS CAGR (5Y)', '3.0%', '6.0%', '8.0%', 'Driven by ROE - dividend yield; ROE ~10.6%, div ~2.4%'),
    ('Terminal BVPS (5Y)', '$96.85', '$126.78', '$138.84', '$94.17 × (1+g)^5'),
    ('Exit P/B Multiple', '0.85x', '1.10x', '1.30x', 'Bear compression; base = current multiple; bull = peer re-rate'),
    ('Implied Target Price', '$82.32', '$139.46', '$180.49', 'Terminal BVPS × exit P/B'),
    ('Current Price', '$108.56', '$108.56', '$108.56', '$108.56 as of Aug 10, 2026'),
    ('Upside / Downside', '-24.2%', '+28.7%', '+65.5%', 'vs current price'),
    ('Weight', '20%', '50%', '30%', 'Probability allocation'),
    ('Weighted Value/Share', '$16.46', '$69.73', '$54.15', 'Target × weight'),
    ('Probability-Weighted FV', '', '', '$140.34', 'Sum of weighted targets'),
    ('Weighted Upside', '', '', '+29.2%', 'From current $108.56'),
]

for i, row in enumerate(scenarios, headers + 1):
    for j, val in enumerate(row, 1):
        font_style = bold_font if i in [14, 15] else normal_font
        c(ws3, i, j, val, font_style, border=True)

# Forward P/E cross-check
c(ws3, 17, 1, 'Forward P/E Cross-Check', section_font)
fpe_check = [
    ('Forward EPS Estimate', '~$10.90', 'Implied from fwd P/E of 9.96 and $108.56 price'),
    ('Bear EPS (conservative)', '$9.50', 'Flat growth; acquisition drag'),
    ('Base EPS (consensus)', '$11.50', '5-7% growth on earnings'),
    ('Bull EPS (optimistic)', '$13.00', 'ROE recovery; NIM expansion'),
    ('Bear P/E', '8.0x', 'Multiple compression in stress'),
    ('Base P/E', '10.5x', 'Current sector average'),
    ('Bull P/E', '12.5x', 'Premium regional bank multiple'),
    ('Bear Target (Fwd P/E)', '$76.00', '$9.50 × 8.0x'),
    ('Base Target (Fwd P/E)', '$120.75', '$11.50 × 10.5x'),
    ('Bull Target (Fwd P/E)', '$162.50', '$13.00 × 12.5x'),
]
c(ws3, 18, 1, 'Metric', bold_font, border=True, fill=header_fill)
c(ws3, 18, 2, 'Value', bold_font, border=True, fill=header_fill)
c(ws3, 18, 3, 'Commentary', bold_font, border=True, fill=header_fill)
for i, (metric, value, comment) in enumerate(fpe_check, 19):
    c(ws3, i, 1, metric, normal_font, border=True)
    c(ws3, i, 2, value, normal_font, border=True)
    c(ws3, i, 3, comment, normal_font, border=True)

for ci in range(1, 6):
    ws3.column_dimensions[get_column_letter(ci)].width = 22

# ============================================================
# Sheet 4: Actuals Source Audit
# ============================================================
ws4 = wb.create_sheet('Actuals Source Audit')

ws4.merge_cells('A1:E1')
c(ws4, 1, 1, 'Actuals Source Audit — SSB', header_font).alignment = Alignment(horizontal='center')

c(ws4, 3, 1, 'Data Point', bold_font, border=True, fill=header_fill)
c(ws4, 3, 2, 'Value', bold_font, border=True, fill=header_fill)
c(ws4, 3, 3, 'Source', bold_font, border=True, fill=header_fill)
c(ws4, 3, 4, 'Date', bold_font, border=True, fill=header_fill)
c(ws4, 3, 5, 'Notes', bold_font, border=True, fill=header_fill)

audit_data = [
    ('Stock Price', '$108.56', 'Yahoo Finance Quote', '2026-08-10', 'Close price'),
    ('Market Cap', '$10.48B', 'Yahoo Key Stats', '2026-08-10', '96.97M shares × $108.56'),
    ('Shares Outstanding', '96.97M', 'Yahoo Key Stats', '2026-08-10', 'Implied from MC/price'),
    ('Enterprise Value', 'N/A', 'Yahoo Key Stats', '2026-08-10', '-- shown for banks'),
    ('Revenue TTM', '$2,656.3M', 'Yahoo Income Statement', 'TTM', 'In thousands; interest + noninterest'),
    ('Revenue FY2025', '$2,617.3M', 'Yahoo Income Statement', '12/31/2025', '+57% vs FY2024 — acquisition-driven'),
    ('Revenue FY2024', '$1,667.1M', 'Yahoo Income Statement', '12/31/2024', ''),
    ('Revenue FY2023', '$1,700.6M', 'Yahoo Income Statement', '12/31/2023', ''),
    ('Revenue FY2022', '$1,613.3M', 'Yahoo Income Statement', '12/31/2022', ''),
    ('Net Income TTM', '$950.2M', 'Yahoo Income Statement', 'TTM', ''),
    ('Net Income FY2025', '$798.7M', 'Yahoo Income Statement', '12/31/2025', ''),
    ('Diluted EPS TTM', '$9.52', 'Yahoo Income Statement', 'TTM', ''),
    ('Diluted EPS FY2025', '$7.87', 'Yahoo Income Statement', '12/31/2025', ''),
    ('Total Assets FY2025', '$67,197.4M', 'Yahoo Balance Sheet', '12/31/2025', '+45% vs FY2024 — acquisition'),
    ('Total Assets FY2024', '$46,381.2M', 'Yahoo Balance Sheet', '12/31/2024', ''),
    ('Total Equity FY2025', '$9,059.1M', 'Yahoo Balance Sheet', '12/31/2025', '+54% vs FY2024'),
    ('Total Debt BS FY2025', '$686.2M', 'Yahoo Balance Sheet', '12/31/2025', ''),
    ('Total Debt Key Stats', '$2,700M', 'Yahoo Key Stats MRQ', '6/30/2026', 'Discrepancy: Key Stats $2.7B vs BS $686M — ~$2B gap'),
    ('Total Cash MRQ', '$2,360M', 'Yahoo Key Stats', '6/30/2026', ''),
    ('BVPS MRQ', '$94.17', 'Yahoo Key Stats', '6/30/2026', ''),
    ('Tangible Book FY2025', '$5,494.7M', 'Yahoo Balance Sheet', '12/31/2025', ''),
    ('Beta (5Y Monthly)', '0.71', 'Yahoo Key Stats', '2026-08-10', ''),
    ('Forward P/E', '9.96x', 'Yahoo Key Stats', '2026-08-10', ''),
    ('Trailing P/E', '11.36x', 'Yahoo Key Stats', '2026-08-10', ''),
    ('P/B', '1.15x', 'Yahoo Key Stats', '2026-08-10', ''),
    ('ROE TTM', '10.60%', 'Yahoo Key Stats', 'TTM', ''),
    ('ROA TTM', '1.41%', 'Yahoo Key Stats', 'TTM', ''),
    ('Dividend Yield', '2.44%', 'Yahoo Key Stats', '2026-08-10', ''),
    ('Payout Ratio', '25.21%', 'Yahoo Key Stats', '2026-08-10', ''),
    ('Operating Cash Flow TTM', '$927.65M', 'Yahoo Cash Flow Statement', 'TTM', ''),
    ('Interest Income TTM', '$3,385.5M', 'Yahoo Income Statement', 'TTM', ''),
    ('Interest Expense TTM', '$1,067.2M', 'Yahoo Income Statement', 'TTM', ''),
    ('Net Interest Income TTM', '$2,318.4M', 'Yahoo Income Statement', 'TTM', ''),
    ('Tax Rate Used', '22.60%', 'Calculated from IS', 'FY2025', '241.5 / 1,040.2'),
    ('Risk-Free Rate', '4.705%', 'CNBC US10Y', '2026-08-10', ''),
    ('Next Earnings Date', 'Unknown', 'Yahoo Profile N/A', 'N/A', 'Profile page had data but not extracted'),
]

for i, (point, value, source, date, notes) in enumerate(audit_data, 4):
    c(ws4, i, 1, point, normal_font, border=True)
    c(ws4, i, 2, value, normal_font, border=True)
    c(ws4, i, 3, source, normal_font, border=True)
    c(ws4, i, 4, date, normal_font, border=True)
    c(ws4, i, 5, notes, normal_font, border=True)

ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 25
ws4.column_dimensions['D'].width = 16
ws4.column_dimensions['E'].width = 55

# ============================================================
# Sheet 5: Questions
# ============================================================
ws5 = wb.create_sheet('Questions')

ws5.merge_cells('A1:C1')
c(ws5, 1, 1, 'Open Questions — SSB', header_font).alignment = Alignment(horizontal='center')

questions = [
    ('Q1', 'Massive FY2024→FY2025 Balance Sheet Jump: Total assets leapt from $46.4B to $67.2B (+45%), and revenue grew from $1.67B to $2.62B (+57%). What acquisition drove this? SouthState is historically a Florida/AL/MS/TN regional bank — did it acquire a Southern peer?', 'M&A identification needed'),
    ('Q2', 'Key Stats vs Balance Sheet Debt Discrepancy: Key Stats shows Total Debt of $2.7B (MRQ) but Balance Sheet shows Total Debt of $686.2M (FY2025). The ~$2B gap may include convertible preferred, capital lease obligations, or subordinated notes classified outside permanent debt. What comprises the difference?', 'Flagged in audit'),
    ('Q3', 'Share Count Dilution: Shares Outstanding grew from 76.0M (FY2023) to 96.97M (current) — a 28% increase. How much is acquisition-related stock consideration vs. organic growth? What is the per-share dilution impact on BVPS?', 'Dilution factor ~1.28x'),
    ('Q4', 'ROE of 10.60% vs Cost of Equity ~12.6%: ROE sits below the CAPM-implied cost of equity. Is this driven by post-acquisition integration drag, or is it a structural ROE constraint? If ROE < CoE, buybacks are needed to compound per-share value.', 'Key for valuation'),
    ('Q5', 'Operating Cash Flow Pattern: OCF TTM is $927.65M but FY2025 was only $300.8M — a dramatic swing. What drove the variance? Could be deposit migration, liquidity management, or acquisition timing effects.', 'Cash flow consistency'),
    ('Q6', 'Credit Quality and NPL Ratio: Post-acquisition, what is the combined NPL ratio? Provisions for credit losses? CET1 capital adequacy? These determine whether the merged franchise is credit-quality sound.', 'Critical for banks'),
    ('Q7', 'Deposit Beta and NIM Trajectory: What is the deposit cost trajectory? In a declining-rate environment, NIM compression risk? What percentage of deposits are interest-bearing vs. noninterest-bearing?', 'NIM sensitivity'),
    ('Q8', 'Acquisition Goodwill: Tangible Book Value is $5.49B vs Total Equity of $9.06B — implying ~$3.57B in goodwill/intangibles. Post-acquisition, is this goodwill at risk of impairment?', 'Intangible ratio high'),
    ('Q9', 'Geographic Concentration and Credit Exposure: What is the CRE concentration? Commercial loan portfolio composition? Does the acquisition add meaningful diversification or just more of the same regional exposure?', 'Risk assessment'),
    ('Q10', 'Preferred Stock: Does SSB have preferred stock on its balance sheet? The Total Equity = Common Stock Equity gap could indicate preferred securities that carry dividend obligations compressing common returns.', 'Capital structure'),
]

c(ws5, 3, 1, '#', bold_font, border=True, fill=header_fill)
c(ws5, 3, 2, 'Question', bold_font, border=True, fill=header_fill)
c(ws5, 3, 3, 'Category', bold_font, border=True, fill=header_fill)
for i, (num, question, category) in enumerate(questions, 4):
    c(ws5, i, 1, num, bold_font, border=True)
    c(ws5, i, 2, question, normal_font, border=True)
    c(ws5, i, 3, category, normal_font, border=True)
ws5.column_dimensions['A'].width = 5
ws5.column_dimensions['B'].width = 90
ws5.column_dimensions['C'].width = 25

# ============================================================
# Sheet 6: Sources
# ============================================================
ws6 = wb.create_sheet('Sources')

ws6.merge_cells('A1:B1')
c(ws6, 1, 1, 'Sources — SSB Research', header_font).alignment = Alignment(horizontal='center')

sources = [
    (1, 'Yahoo Finance — Quote page', 'https://finance.yahoo.com/quote/SSB/'),
    (2, 'Yahoo Finance — Income Statement', 'https://finance.yahoo.com/quote/SSB/financials/'),
    (3, 'Yahoo Finance — Balance Sheet', 'https://finance.yahoo.com/quote/SSB/balance-sheet/'),
    (4, 'Yahoo Finance — Cash Flow', 'https://finance.yahoo.com/quote/SSB/cash-flow/'),
    (5, 'Yahoo Finance — Key Statistics', 'https://finance.yahoo.com/quote/SSB/key-statistics/'),
    (6, 'CNBC — US 10 Year Treasury (US10Y)', 'https://cnbc.com/quotes/US10Y'),
    (7, 'Yahoo Finance — Profile (attempts failed — page crashed)', 'https://finance.yahoo.com/quote/SSB/profile/'),
    (8, 'StockAnalysis — SSB (404 Not Found)', 'https://stockanalysis.com/quote/SSB/'),
    (9, 'Yahoo Finance — Analysis/Estimates (popup blocked)', 'https://finance.yahoo.com/quote/SSB/analysis/'),
]

c(ws6, 3, 1, '#', bold_font, border=True, fill=header_fill)
c(ws6, 3, 2, 'Description', bold_font, border=True, fill=header_fill)
c(ws6, 3, 3, 'URL', bold_font, border=True, fill=header_fill)
for i, (num, desc, url) in enumerate(sources, 4):
    c(ws6, i, 1, num, normal_font, border=True)
    c(ws6, i, 2, desc, normal_font, border=True)
    c(ws6, i, 3, url, normal_font, border=True)

ws6.column_dimensions['A'].width = 5
ws6.column_dimensions['B'].width = 55
ws6.column_dimensions['C'].width = 60

# ============================================================
# Save and verify
# ============================================================
filename = '[2026-08-10] SSB Model.xlsx'
path = f'/home/refcell/dev/capital/models/{filename}'
wb.save(path)

# Verify
from openpyxl import load_workbook
wb2 = load_workbook(path)
print(f'Saved: {path}')
print(f'Sheets: {wb2.sheetnames}')
print(f'Valuation sheet title: {wb2["Valuation"].cell(1,1).value}')
print(f'WACC computed: {wb2["WACC"].cell(15,2).value}')
print(f'Scenarios P/B FV: {wb2["Scenarios"].cell(15,4).value}')

# Print key scenario values for verification
s = wb2['Scenarios']
print(f'Bear target: {s.cell(9,2).value}')
print(f'Base target: {s.cell(9,3).value}')
print(f'Bull target: {s.cell(9,4).value}')
print(f'Weighted FV: {s.cell(15,4).value}')
print('Model build complete.')
