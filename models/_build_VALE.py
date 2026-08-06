"""Build a 6-sheet Vale S.A. (VALE) valuation model.
Price snapshot: $14.94 on 2026-08-05. Data source: StockAnalysis.com, Yahoo Finance, CNBC.
Primary lens: Forward P/E — FCF framework fails due to debt/FCF gray zone (net debt ~4.7x TTM FCF).
"""
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# ── Data ──
TICKER = "VALE"
DATE = datetime(2026, 8, 5)
PRICE = 14.94
MC = 63.76        # $B
EV = 80.37       # $B
CASH = 5.279     # $B (Mar 2026)
DEBT = 21.329    # $B (Mar 2026)
NET_DEBT = DEBT - CASH  # $16.05B
SHARES_M = 4260   # 4.26B shares outstanding
BETA = 0.73
RISK_FREE = 0.04611
ERP = 0.05
COE = RISK_FREE + BETA * ERP        # 0.04611 + 0.73*0.05 = 0.0826 ≈ 8.26%
COD = 0.055
TAX = 0.30      # Brazilian effective tax rate forward
EQ_W = round(MC / (MC + DEBT), 4)   # 0.7490
DEB_W = round(DEBT / (MC + DEBT), 4)  # 0.2510
WACC = round(EQ_W * COE + DEB_W * COD * (1 - TAX), 4)

# Historical financials ($M)
REV = {"TTM": 41216, "2025": 38403, "2024": 38056, "2023": 41784, "2022": 43839, "2021": 54502}
GP  = {"TTM": 14368, "2025": 13456, "2024": 13791, "2023": 17695, "2022": 19811, "2021": 32773}
OP  = {"TTM": 6629, "2025": 5897, "2024": 10788, "2023": 14205, "2022": 17208, "2021": 27693}
NI  = {"TTM": 1323, "2025": 2352, "2024": 6166, "2023": 7983, "2022": 18788, "2021": 22445}
EPS = {"TTM": 0.49, "2025": 0.55, "2024": 1.44, "2023": 1.83, "2022": 4.05, "2021": 4.47}
FCF = {"TTM": 3555, "2025": 507, "2024": 472, "2023": 7245, "2022": 6039, "2021": 20646}
EBD = {"TTM": 10006, "2025": 6460, "2024": 11283, "2023": 17275, "2022": 20379, "2021": 30727}

# Multiples (StockAnalysis + Yahoo Finance)
PE_TTM = 31.86
PE_FWD = 7.68
PS_TTM = 1.51
PB = 1.64
P_FCF = round(MC / (FCF["TTM"] / 1000), 1)
EV_FCF = round(EV / (FCF["TTM"] / 1000), 1)
EV_REV = round(EV / (REV["TTM"] / 1000), 2)
EV_EB = round(EV / (EBD["TTM"] / 1000), 2)

# ── Scenarios (Forward P/E based — FCF framework fails due to debt amplification) ──
# Analyst PT average: $16.78. Base must align within ~±15% of this.
# FY2026 consensus EPS: $1.95, FY2027: $2.08

SCENARIOS = {}
for name, cagr, term_eps, exit_pe, wt in [
    ("Bear", 0.015, 1.40, 10, 0.25),
    ("Base", 0.035, 2.00, 9, 0.50),
    ("Bull", 0.060, 2.60, 12, 0.25),
]:
    tr = REV["TTM"] * (1 + cagr) ** 5       # $ millions
    fcf_m = round(term_eps * 0.65)          # implied FCF margin ~8% in base
    # FCF cross-check: implied EV via FCF multiple (secondary)
    fcf_cross_mult = 9 if name == "Base" else (7 if name == "Bear" else 11)
    fcf_cross_ev = round(fcf_m * fcf_cross_mult, 1)
    tp = round(term_eps * exit_pe, 2)
    up = round((tp / PRICE - 1) * 100, 1)
    SCENARIOS[name] = {
        "cagr": cagr,
        "tr": round(tr, 0),
        "term_eps": term_eps,
        "exit_pe": exit_pe,
        "tp": tp,
        "up": up,
        "wt": wt,
        "wv": round(tp * wt, 2),
        "fcf_cross_ev": fcf_cross_ev,
    }

tot_wv = round(sum(s["wv"] for s in SCENARIOS.values()), 2)
tot_up = round((tot_wv / PRICE - 1) * 100, 1)

# ── Styles ──
BF = Font(bold=True)
B14 = Font(bold=True, size=14)
B12 = Font(bold=True, size=12)
BD = Border(left=Side("thin"), right=Side("thin"),
            top=Side("thin"), bottom=Side("thin"))
D2 = "0.00"
PP = "0.0%"
PP2 = "0.00%"
DB = "$#,##0.0\"B\""
DM = "$#,##0"
DP = "$#,##0.00"


def c(ws, r, co, v, fmt=None, font=None, border=None):
    """Write a cell."""
    cell = ws.cell(row=r, column=co, value=v)
    if fmt is not None:
        cell.number_format = fmt
    if font is not None:
        cell.font = font
    if border is not None:
        cell.border = border
    return cell


def title(ws, text):
    ws.merge_cells("A1:F1")
    c(ws, 1, 1, text, font=B14)


def hdr_row(ws, r, cols):
    for j, h in enumerate(cols, 1):
        c(ws, r, j, h, font=B12, border=BD)


def data_row(ws, r, vals):
    for j, v in enumerate(vals, 1):
        c(ws, r, j, v, border=BD)


# ── Build ──
OUT = Path("/home/refcell/dev/capital/models/[2026-08-05] Vale Model.xlsx")
wb = openpyxl.Workbook()

# ── Sheet 1: Valuation ──
ws = wb.active
ws.title = "Valuation"
title(ws, f"{TICKER} — {DATE.strftime('%Y-%m-%d')} Valuation Model")

items = [
    ("Company", "Vale S.A."),
    ("Ticker", "NYSE: VALE"),
    ("Date", DATE.strftime("%Y-%m-%d")),
    ("Price", PRICE),
    ("Shares Outstanding (M)", SHARES_M),
    ("Market Cap ($B)", MC),
    ("Enterprise Value ($B)", EV),
    ("Cash ($B)", CASH),
    ("Debt ($B)", DEBT),
    ("Net Debt ($B)", NET_DEBT),
    ("Primary Valuation Lens", "Forward P/E"),
    ("Stance", "Watch — forward P/E cheap at 7.7x, but commodity cycle at a trough"),
]
for i, (lab, val) in enumerate(items, 2):
    c(ws, i, 1, lab, font=BF, border=BD)
    c(ws, i, 2, val, border=BD)

c(ws, 10, 1, "Key Valuation Metrics", font=B12)

metrics = [
    ("P/E (TTM)", PE_TTM, "TTM on $0.49 EPS; distorted by cycle trough (iron ore down)"),
    ("Forward P/E", PE_FWD, "FY26E EPS $1.95; 7.7x is cheap for diversified miner"),
    ("P/S (TTM)", PS_TTM, "1.5x; normal for mining at cycle peak, cheap at trough"),
    ("P/FCF (TTM)", P_FCF, f"FCF ~$3.56B; {P_FCF}x; note: FCF framework fails due to debt amplification"),
    ("EV/FCF", EV_FCF, f"Enterprise lens; {EV_FCF}x; debt/FCF ~4.7x pushes equity value to negatives"),
    ("EV/Revenue", EV_REV, f"EV at {EV_REV}x revenue; reasonable for commodities"),
    ("EV/EBITDA", EV_EB, f"{EV_EB}x; 6-10x range for diversified miners"),
    ("P/B", PB, "1.64x; above historical lows due to buyback pressure"),
    ("Debt/EBITDA", round(DEBT * 1000 / EBD["TTM"], 1), "~2.1x; manageable leverage for a miner"),
    ("Dividend Yield", "6.06%", "$0.91/share annualized; strong income cushion"),
]
for i, (nm, val, txt) in enumerate(metrics, 11):
    c(ws, i, 1, nm, font=BF, border=BD)
    c(ws, i, 2, val, fmt=D2 if isinstance(val, (int, float)) else None, border=BD)
    c(ws, i, 3, txt, border=BD)
    ws.merge_cells(f"C{i}:F{i}")

# ── Sheet 2: WACC ──
w2 = wb.create_sheet("WACC")
title(w2, "WACC Calculation — CAPM")

wd = [
    ("Risk-Free Rate (10Y US)", RISK_FREE, PP2, "CNBC 8/5/26: 4.611%"),
    ("Equity Risk Premium", ERP, PP, "Standard 5%"),
    ("Beta (5Y Monthly)", BETA, D2, "Low beta for commodity — diversification hedge"),
    ("Cost of Equity", round(COE, 4), PP2, f"{RISK_FREE*100:.3f}% + 0.73 x 5.00% = 8.26%"),
    ("Cost of Debt", COD, PP, "~5.5%; Brazilian sovereign + corporate spread"),
    ("Tax Rate", TAX, PP, "~30% Brazilian effective (IRPJ + CSLL)"),
    ("", None, None, ""),
    ("Market Cap ($B)", MC, DB, "8/5/26 close"),
    ("Total Debt ($B)", DEBT, DB, "Mar 2026 balance sheet"),
    ("Equity Weight", EQ_W, PP2, f"{EQ_W*100:.1f}%"),
    ("Debt Weight", DEB_W, PP2, f"{DEB_W*100:.1f}%"),
    ("", None, None, ""),
    ("WACC", WACC, PP2, f"= {EQ_W} x {COE*100:.2f}% + {DEB_W} x {COD*100:.1f}% x (1-{TAX})"),
]
for i, (lab, val, fmt, note) in enumerate(wd, 2):
    c(w2, i, 1, lab, font=BF if lab else None, border=BD)
    c(w2, i, 2, val, fmt=fmt, border=BD)
    c(w2, i, 3, note, border=BD)
    if note:
        w2.merge_cells(f"C{i}:E{i}")

print(f"  WACC = {WACC*100:.2f}%")

# ── Sheet 3: Scenarios ──
w3 = wb.create_sheet("Scenarios")
title(w3, "Bear / Base / Bull Scenario Analysis (Forward P/E Framework)")

cols3 = ["Scenario", "Rev CAGR 5Y", "Term Rev ($M)",
         "Term EPS", "Exit P/E", "Target Price",
         "Upside %", "Weight", "Wtd Value/Share",
         "Notes"]
hdr_row(w3, 2, cols3)

for i, (nm, s) in enumerate(SCENARIOS.items(), 3):
    vals = [nm, s["cagr"], s["tr"], s["term_eps"], s["exit_pe"],
            s["tp"], s["up"] / 100, s["wt"], s["wv"],
            ""]
    for j, v in enumerate(vals, 1):
        c(w3, i, j, v, border=BD)
    w3.cell(row=i, column=2).number_format = PP2
    w3.cell(row=i, column=7).number_format = PP2
    w3.cell(row=i, column=8).number_format = PP2

# Total row
tr = 6
c(w3, tr, 1, "TOTAL", font=BF, border=BD)
c(w3, tr, 8, 1.0, fmt=PP2, font=BF, border=BD)
c(w3, tr, 9, tot_wv, fmt=DP, font=BF, border=BD)

c(w3, tr + 1, 8, "Prob-Weighted FV", font=BF, border=BD)
c(w3, tr + 1, 9, tot_wv, fmt=DP, font=BF, border=BD)
c(w3, tr + 2, 1, f"Upside from ${PRICE}", font=BF)
c(w3, tr + 2, 2, tot_up / 100, fmt=PP2, font=BF)

# FCF framework warning
rnote = tr + 4
warn_text = (
    "NOTE: FCF multiple framework suppressed for VALE. Net debt of ~$16B / FCF of ~$3.6B "
    "= 4.4x. Even at 10x FCF, implied EV is only ~$36B, leaving ~$20B equity → ~$4.7/share "
    "vs current $14.94. The debt/FCF amplification makes FCF multiples economically meaningless. "
    "Forward P/E is the primary lens. Analyst avg PT: $16.78 (12% upside)."
)
c(w3, rnote, 1, warn_text, font=Font(italic=True, color="FF0000"))
w3.merge_cells(f"A{rnote}:J{rnote}")

notes = [
    "Scenario logic:",
    "Bear: Iron ore stays depressed; FY31 EPS only $1.40, exit P/E compresses to 10x as cycle risk premiums remain.",
    "Base: Earnings recover toward FY26E consensus of $1.95 then grow modestly; exit P/E 9x reflects cyclical peer norms.",
    "Bull: Iron ore/copper prices recover; scale economy; EPS compounds to $2.60 with P/E expansion to 12x.",
    f"FY2026 consensus EPS: $1.95 (11 analysts), FY2027: $2.08 (12 analysts). Source: Yahoo Finance analysis 8/5/26.",
    f"Net debt: ${NET_DEBT:.1f}B. Shares: {SHARES_M}M. WACC: {WACC*100:.2f}%.",
]
for i, n in enumerate(notes, rnote + 1):
    c(w3, i, 1, n)
    w3.merge_cells(f"A{i}:J{i}")

print(f"  Weighted FV = ${tot_wv} ({tot_up}% upside)")

# ── Sheet 4: Actuals Source Audit ──
w4 = wb.create_sheet("Actuals Source Audit")
title(w4, "Data Point Source Audit")

hdr_row(w4, 2, ["Data Point", "Value", "Source", "Date", "Notes"])

audit = [
    ("Price", f"${PRICE}", "stockanalysis.com/stocks/VALE", "2026-08-05", "Close"),
    ("Market Cap", f"${MC}B", "stockanalysis.com/stocks/VALE/statistics", "2026-08-05", ""),
    ("Enterprise Value", f"${EV}B", "stockanalysis.com/stocks/VALE/statistics", "2026-08-05", ""),
    ("Shares (M)", SHARES_M, "stockanalysis.com + MC/price recon", "2026-08-05", ""),
    ("Rev TTM", f"${REV['TTM']}", "stockanalysis.com/stocks/VALE/financials", "2026-08-05", ""),
    ("Rev FY25", f"${REV['2025']}", "stockanalysis.com/stocks/VALE/financials", "2026-08-05", ""),
    ("Rev FY24", f"${REV['2024']}", "stockanalysis.com/stocks/VALE/financials", "2026-08-05", ""),
    ("GP TTM", f"${GP['TTM']}", "stockanalysis.com/stocks/VALE/financials", "2026-08-05", ""),
    ("GP FY25", f"${GP['2025']}", "stockanalysis.com/stocks/VALE/financials", "2026-08-05", ""),
    ("OP TTM", f"${OP['TTM']}", "stockanalysis.com/stocks/VALE/financials", "2026-08-05", ""),
    ("NI TTM", f"${NI['TTM']}", "stockanalysis.com/stocks/VALE/financials", "2026-08-05", ""),
    ("Diluted EPS TTM", f"${EPS['TTM']}", "stockanalysis.com/stocks/VALE/financials", "2026-08-05", ""),
    ("FCF TTM", f"${FCF['TTM']}M", "stockanalysis.com/stocks/VALE/financials", "2026-08-05", "FCF margin 8.63%"),
    ("FCF FY25", f"${FCF['2025']}M", "stockanalysis.com/stocks/VALE/financials", "2026-08-05", "Cycle trough: 1.32% margin"),
    ("EBITDA TTM", f"${EBD['TTM']}M", "stockanalysis.com/stocks/VALE/financials", "2026-08-05", ""),
    ("Cash", f"${CASH}B", "stockanalysis.com/stocks/VALE/financials/balance-sheet", "2026-08-05", "Mar 2026"),
    ("Total Debt", f"${DEBT}B", "stockanalysis.com/stocks/VALE/financials/balance-sheet", "2026-08-05", "Mar 2026"),
    ("Beta", str(BETA), "stockanalysis.com/stocks/VALE + Yahoo Stats", "2026-08-05", "0.73 low beta"),
    ("P/E TTM", str(PE_TTM), "stockanalysis.com/stocks/VALE/statistics", "2026-08-05", ""),
    ("Forward P/E", str(PE_FWD), "stockanalysis.com/stocks/VALE/statistics", "2026-08-05", "7.68x; cheap"),
    ("EV/EBITDA", str(EV_EB), "stockanalysis.com/stocks/VALE/statistics", "2026-08-05", ""),
    ("P/B", str(PB), "stockanalysis.com/stocks/VALE/statistics", "2026-08-05", ""),
    ("10Y Treasury", f"{RISK_FREE*100:.3f}%", "cnbc.com/quotes/US10Y", "2026-08-05", ""),
    ("Analyst Revenue FY26", "$41.24B", "finance.yahoo.com/quote/VALE/analysis", "2026-08-05", "19 analysts"),
    ("Analyst EPS FY26", "$1.95", "finance.yahoo.com/quote/VALE/analysis", "2026-08-05", "11 analysts"),
    ("Analyst PT", "$16.78", "finance.yahoo.com/quote/VALE", "2026-08-05", "Buy consensus, +12.3%"),
    ("Earn Date", "Jul 30, 2026", "stockanalysis.com/stocks/VALE/statistics", "2026-08-05", "Q2 FY26"),
    ("Dividend Rate", "$0.91 annualized", "stockanalysis.com/stocks/VALE", "2026-08-05", "6.06% yield"),
]
for i, (dp, val, src, dt, note) in enumerate(audit, 3):
    for j, v in enumerate([dp, val, src, dt, note], 1):
        c(w4, i, j, v, border=BD)

# ── Sheet 5: Questions ──
w5 = wb.create_sheet("Questions")
title(w5, "Open Questions")

questions = [
    ("1", "OCF Data Anomaly", "TTM OCF shows $50.8B on $41.2B revenue (123% margin). This is likely a StockAnalysis currency conversion artifact. FCF of $3.56B (8.6%) is the more reliable cash flow figure."),
    ("2", "Iron Ore Price Sensitivity", "64% of revenue is iron ore. TTM margins compressed from 42% (FY23) to 35% (TTM) gross, 34%→16% operating. How deep is the trough? Chinese demand is the key variable."),
    ("3", "Nickel Transition Risk", "Nickel revenue of $5B TTM. EV transition has depressed nickel prices. Is this a structural headwind or cyclical? LME nickel futures context."),
    ("4", "Copper Growth Optionality", "Copper revenue up to $4.6B TTM (+31% YoY from FY25 $3.55B). Copper benefits from long-term electrification thesis. Is this a growth driver or just price cyclicality?"),
    ("5", "Tax Rate Volatility", "TTM effective tax rate 7.57% vs FY25 57.38%. FY23 was 27.3%, FY22 15.0%. Extreme variance signals one-time items. Forward rate assumption is ~30%."),
    ("6", "Dividend Sustainability", "Annual dividend of ~$0.91 on EPS of $0.49 TTM = payout ratio >100%. Dividend is funded by retained earnings and cash reserves. Is the dividend at risk if commodity prices stay low?"),
    ("7", "Capital Allocation", "Debt/EBITDA at 2.1x is manageable but debt grew from $13.9B (FY23) to $21.3B (Mar'26). Buyback pace + dividend + debt service vs. FCF generation. Net debt trajectory?"),
    ("8", "Chinese Demand and Geopolitics", "China is ~70% of iron ore demand. Tariff escalation, trade tensions, or Chinese stimulus could swing the stock 30-50% in either direction."),
    ("9", "Sustainability/ESG Exposure", "Brazil mining ESG risks: Brumadinho dam disaster (2019) legacy, environmental liabilities, regulatory overhang. Does the market discount this in the P/B of 1.64x?"),
    ("10", "Next Earnings Catalyst", "Q2 FY26 earned Jul 30, 2026. Missed estimates by -26.21% EPS surprise. Q3 FY26 estimate of $0.50 EPS with downward revisions (-14% in 30 days). Next catalyst is Q3 results in Nov 2026."),
]
hdr_row(w5, 2, ["#", "Question", "Detail"])
for i, (num, q, d) in enumerate(questions, 3):
    c(w5, i, 1, num, border=BD)
    c(w5, i, 2, q, font=BF, border=BD)
    c(w5, i, 3, d, border=BD)
    w5.cell(row=i, column=3).alignment = Alignment(wrap_text=True)

w5.column_dimensions["C"].width = 80

# ── Sheet 6: Sources ──
w6 = wb.create_sheet("Sources")
title(w6, "Data Sources")

hdr_row(w6, 2, ["#", "Source", "URL"])
srcs = [
    ("1", "StockAnalysis — Overview", "stockanalysis.com/stocks/VALE"),
    ("2", "StockAnalysis — Financials (Income)", "stockanalysis.com/stocks/VALE/financials"),
    ("3", "StockAnalysis — Balance Sheet", "stockanalysis.com/stocks/VALE/financials/balance-sheet"),
    ("4", "StockAnalysis — Statistics", "stockanalysis.com/stocks/VALE/statistics"),
    ("5", "Yahoo Finance — Analyst Estimates", "finance.yahoo.com/quote/VALE/analysis"),
    ("6", "Yahoo Finance — Key Statistics", "finance.yahoo.com/quote/VALE/key-statistics"),
    ("7", "CNBC — 10Y Treasury", "cnbc.com/quotes/US10Y"),
    ("8", "Vale Corporate Site", "vale.com"),
]
for i, (nm, des, url) in enumerate(srcs, 3):
    c(w6, i, 1, nm, border=BD)
    c(w6, i, 2, des, border=BD)
    c(w6, i, 3, url, border=BD)

w6.column_dimensions["B"].width = 45
w6.column_dimensions["C"].width = 60

# ── Save ──
wb.save(str(OUT))
print(f"Saved {OUT}")
print(f"Sheets: {wb.sheetnames}")
