#!/usr/bin/env python3
"""Build 6-sheet PEGA valuation model. Data as of June 18, 2026."""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = openpyxl.Workbook()

hdr_font = Font(bold=True, size=12)
sub_font = Font(bold=True, size=10)
title_font = Font(bold=True, size=14)
tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
hf = PatternFill('solid', fgColor='D9E2F3')

def fmt(ws, r, c, v, b=False, f=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.border = tb
    if b: cell.font = sub_font
    if f: cell.number_format = f
    return cell

def hdr(ws, r, cols):
    for c, v in enumerate(cols, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = sub_font
        cell.fill = hf
        cell.border = tb

# ── Sheet 1: Valuation ──
ws = wb.active; ws.title = "Valuation"
ws.merge_cells('A1:D1')
ws.cell(1,1).value = "Pegasystems Inc. (PEGA) - Valuation Summary"
ws.cell(1,1).font = title_font

td = [("Company","Pegasystems Inc."),("Ticker","NASDAQ: PEGA"),("Date","2026-06-19"),
      ("Price","30.09"),("Shares Outstanding (M)","167.12"),("Market Cap ($M)","5,030"),
      ("Enterprise Value ($M)","4,620"),("Primary Lens","Forward P/E, scenario-based FCF multiple"),
      ("Stance","Watch — compelling FCF but recent earnings miss")]
for i,(k,v) in enumerate(td):
    fmt(ws, i+2, 1, k, True); fmt(ws, i+2, 2, v)

hr = 13
hdr(ws, hr, ["Valuation Metric","Value","Commentary"])
vm = [
    ("Trailing P/E","16.25x","Down from 62.63x at Q1; TTM earnings still depressed by Q4 miss"),
    ("Forward P/E (FY26)","10.28x","Very cheap on consensus $2.93 EPS"),
    ("Forward P/E (FY27)","9.18x","Extremely cheap on $3.04 EPS consensus"),
    ("P/S","3.22x","Halved from 6.57x at year-end 2025"),
    ("P/FCF (TTM)","10.16x","$495M FCF on $5.03B price; compelling"),
    ("P/B","7.12x","Book value at $705M tangible; elevated but justified by ROE"),
    ("EV/EBITDA","20.95x","From 27.53x at FY25; multiple compression real"),
    ("EV/Sales","2.72x","From 5.71x; dramatic contraction"),
    ("PEG (5Y)","3.45x","Still high — legacy of prior 60x multiple"),
    ("Div Yield","0.37%","$0.12/yr; token income; payout ratio 8%"),
]
for i,(m,v,c) in enumerate(vm):
    fmt(ws, hr+1+i, 1, m); fmt(ws, hr+1+i, 2, v); fmt(ws, hr+1+i, 3, c)
ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 16; ws.column_dimensions['C'].width = 55

# ── Sheet 2: WACC ──
ws = wb.create_sheet("WACC")
ws.merge_cells('A1:C1')
ws.cell(1,1).value = "WACC - CAPM"
ws.cell(1,1).font = title_font

wd = [
    ("Risk-Free Rate (10Y)","4.49%","DGS10 2026-06-17"),
    ("Equity Risk Premium","5.00%","Standard"),
    ("Beta (5Y Monthly)","0.85","Yahoo; below-market vol"),
    ("Cost of Equity","8.79%","=4.49% + 0.85*5%"),
    ("After-Tax Cost of Debt","2.78%","~3.5% pre-tax * (1-20.4%)"),
    ("Effective Tax Rate","20.38%","TTM tax provision / pretax"),
    ("Market Cap ($M)","5,030","Yahoo"),
    ("Total Debt ($M)","76","Capital lease obligations"),
    ("Total Cash ($M)","474","Yahoo MRQ"),
    ("Equity Weight","98.5%","MC / (MC + Debt - Cash adj)"),
    ("Debt Weight","1.5%","Debt / (MC + Debt)"),
    ("WACC","8.74%","=0.985*8.79% + 0.015*2.78%"),
]
hdr(ws, 3, ["Component","Value","Source/Notes"])
for i,(c,v,n) in enumerate(wd):
    fmt(ws, 4+i, 1, c, True); fmt(ws, 4+i, 2, v); fmt(ws, 4+i, 3, n)
ws.column_dimensions['A'].width = 28; ws.column_dimensions['B'].width = 16; ws.column_dimensions['C'].width = 45

# ── Sheet 3: Scenarios ──
ws = wb.create_sheet("Scenarios")
ws.merge_cells('A1:K1')
ws.cell(1,1).value = "Scenario Analysis — Bear / Base / Bull"
ws.cell(1,1).font = title_font

sc_hdr = ["Metric","Bear (20%)","Base (50%)","Bull (30%)","Notes"]
hr = 3; hdr(ws, hr, sc_hdr)

# FY25 revenue = $1,745.8M; FY26 cons = $1.98B; FY27 cons = $2.22B
# Revenue CAGR to 5Y terminal:
# Bear: 7% CAGR (slower than consensus, misses growth inflection)
# Base: 11% CAGR (aligns with consensus ~12-13% growth next 2 yrs then normalizes)
# Bull: 15% CAGR (AI platform expansion accelerates)
# FCF margin TTM = $495M/$1,700M = 29.1%. Very strong.
# Bear FCF margin: 22% (margin compression from competition/investment)
# Base FCF margin: 27% (sustained strong FCF conversion)
# Bull FCF margin: 32% (operating leverage + margin expansion)

sc = [
    ("Revenue CAGR (5Y)","7.0%","11.0%","15.0%","Base aligns with consensus 12-13% growth rate"),
    ("Terminal Revenue (Y5)","$2,448M","$2,962M","$3,550M","From FY25 $1,746M base"),
    ("Adjusted FCF Margin","22.0%","27.0%","32.0%","TTM FCF margin 29.1%"),
    ("Terminal FCF","$539M","$800M","$1,136M","Rev * margin"),
    ("Exit Multiple (EV/FCF)","14x","17x","21x","SaaS/AI platform comps"),
    ("Implied EV","$7,539M","$13,593M","$23,844M","FCF * exit multiple"),
    ("Less Net Debt Adj","-$398M","-$398M","-$398M","Cash $474M - Debt $76M; assumes constant"),
    ("Implied Equity Value","$7,938M","$13,991M","$24,241M","EV + net cash"),
    ("Shares (M)","167.1","160.0","155.0","Buybacks reduce count; bullish on share reduction"),
    ("Target Price / Share","$47.49","$87.44","$156.40","Equity value / shares"),
    ("Upside from $30.09","+57.8%","+190.9%","+420.7%","Target vs current"),
    ("Weight","20%","50%","30%","Probability weights"),
    ("Weighted Value / Share","$9.50","$43.72","$46.92","Target * weight"),
    ("Total Probability-Weighted FV","","","$100.14","Sum of weighted values"),
    ("Upside from Current","","","+233%","$100.14 vs $30.09"),
]
for i,(m,b,ba,bu,n) in enumerate(sc):
    fmt(ws, hr+1+i, 1, m, True if m in ["Total Probability-Weighted FV","Upside from Current"] else False)
    fmt(ws, hr+1+i, 2, b); fmt(ws, hr+1+i, 3, ba); fmt(ws, hr+1+i, 4, bu); fmt(ws, hr+1+i, 5, n)
ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 14; ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14; ws.column_dimensions['E'].width = 45

# ── Sheet 4: Actuals Source Audit ──
ws = wb.create_sheet("Actuals Source Audit")
ws.merge_cells('A1:E1')
ws.cell(1,1).value = "Actuals Source Audit"
ws.cell(1,1).font = title_font

sa_hdr = ["Data Point","Value","Source URL","Date","Notes"]
hr = 3; hdr(ws, hr, sa_hdr)
sa = [
    ("Stock Price","$30.09","Yahoo Finance /finance.yahoo.com/quote/PEGA","2026-06-18","NasdaqGS close"),
    ("After Hours","$30.00","Yahoo Finance","2026-06-18 19:58 EDT","Min -0.09"),
    ("Market Cap","$5.03B","Yahoo Statistics","2026-06-18","Current"),
    ("Enterprise Value","$4.62B","Yahoo Statistics","2026-06-18","Current"),
    ("Shares Outstanding","167.12M","Yahoo Statistics","2026-06-18","Basic"),
    ("Beta","0.85","Yahoo Statistics","5Y Monthly","Below market"),
    ("TTM Revenue","$1.700B","Yahoo Income Statement","TTM 12/31/2025","Annual + TTM"),
    ("FY25 Revenue","$1,745.8M","Yahoo Income Statement","12/31/2025","Fiscal year"),
    ("FY24 Revenue","$1,497.2M","Yahoo Income Statement","12/31/2024","Fiscal year"),
    ("TTM Gross Profit","$1.274B","Yahoo Income Statement","TTM","Gross margin 75.0%"),
    ("TTM Operating Income","$194.4M","Yahoo Income Statement","TTM","Op margin 11.4%"),
    ("TTM Net Income","$340.8M","Yahoo Income Statement","TTM","Profit margin 20.0%"),
    ("TTM Diluted EPS","$1.85","Yahoo Income Statement","TTM","From $341.5M on 182.3M diluted"),
    ("TTM EBITDA","$220.6M","Yahoo Income Statement","TTM","Calculated"),
    ("TTM FCF","$494.9M","Yahoo Cash Flow","TTM","OCF $513.3M - CapEx $18.4M"),
    ("Total Cash","$473.95M","Yahoo Statistics MRQ","Q1 2026","MRQ per share $2.83"),
    ("Total Debt","$72.13M","Yahoo Statistics MRQ","Q1 2026","Debt/equity 10.2%"),
    ("Capital Lease Obligations","$76.0M","Yahoo Balance Sheet","FY2025","Total debt proxy"),
    ("10Y Treasury Rate","4.49%","FRED DGS10","2026-06-17","Risk-free rate"),
    ("FY26 Revenue Consensus","$1.98B","Yahoo Analysis","2026-06-19","11 analysts"),
    ("FY27 Revenue Consensus","$2.22B","Yahoo Analysis","2026-06-19","11 analysts"),
    ("FY26 EPS Consensus","$2.71","Yahoo Analysis (non-GAAP)","2026-06-19","12 analysts"),
    ("FY27 EPS Consensus","$3.04","Yahoo Analysis (non-GAAP)","2026-06-19","12 analysts"),
    ("Trailing P/E","16.25x","Yahoo Statistics","Current","Down from 62.63x Q1"),
    ("Forward P/E","10.28x","Yahoo Statistics","Current","On FY26 EPS"),
    ("Q4 EPS Miss","-33.60%","Yahoo Analysis","2026-03-31","Est $0.69, actual $0.46"),
    ("Dividend","$0.03/qtr","Yahoo Finance","2026-06-19","Ex-date Jul 1"),
]
for i,(dp,v,s,d,n) in enumerate(sa):
    fmt(ws, hr+1+i, 1, dp, True); fmt(ws, hr+1+i, 2, v); fmt(ws, hr+1+i, 3, s)
    fmt(ws, hr+1+i, 4, d); fmt(ws, hr+1+i, 5, n)
ws.column_dimensions['A'].width = 28; ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 35; ws.column_dimensions['D'].width = 16; ws.column_dimensions['E'].width = 40

# ── Sheet 5: Questions ──
ws = wb.create_sheet("Questions")
ws.merge_cells('A1:B1')
ws.cell(1,1).value = "Open Questions"
ws.cell(1,1).font = title_font

q = [
    "Why did Q4 FY26 (Mar 2026) EPS miss by 33.6%? Was this a one-time item or a sign of structural headwinds? What drove the $0.23 miss?",
    "Revenue grew from $1,318M (FY22) to $1,746M (FY25) — 9.6% CAGR. Is post-pandemic normalization at risk? TTM revenue of $1.70B is slightly below FY25 peak of $1.75B.",
    "Net income swung wildly: -$346M (FY22) to $393M (FY25). FY22 losses were driven by negative tax provision of $184M (massive deferred tax asset release?). Check deferred tax treatment.",
    "Huge share buyback acceleration: $567M in TTM buybacks vs. $73M in FY24. Is the company buying back at inflated prices? Shares fell from 170.8M (FY25) to 167.1M (current).",
    "Total debt dropped dramatically from $688M (FY22) to $72M (current). Debt retirement of $468M in TTM. What drove this? Organic paydown or deal-related restructuring?",
    "SBC (stock-based comp) in operating expense of $1.08B — what portion is SBC vs. actual cash opex? This matters for FCF quality assessment.",
    "The company recently announced AI-focused product initiatives. How differentiated is Pegasystems' AI positioning vs. competition (ServiceNow, Salesforce, MuleSoft)?",
    "Customer concentration risk? Largest customers by % of revenue? Contract renewal rates?",
    "Management guidance vs. market expectations? Has the company guided for FY26 revenue of $1.98B? What is the official Q2 guidance?",
    "Next earnings date? Q2 FY26 results should appear in late August / early September 2026. This is a critical data point for the thesis.",
    "Why did StockAnalysis.com return 404 for PEGA? Is this a coverage issue or data-provider exclusion?",
    "Pegasystems' PEG ratio of 3.45x is very high despite declining stock price. Does this reflect stretched multi-year growth expectations?",
]
for i, txt in enumerate(q, 1):
    fmt(ws, i+2, 1, f"{i}.", True)
    ws.cell(row=i+2, column=1).fill = hf
    ws.merge_cells(f'B{i+2}:F{i+2}')
    fmt(ws, i+2, 2, txt)
ws.column_dimensions['A'].width = 5; ws.column_dimensions['B'].width = 100

# ── Sheet 6: Sources ──
ws = wb.create_sheet("Sources")
ws.merge_cells('A1:B1')
ws.cell(1,1).value = "Research Sources"
ws.cell(1,1).font = title_font

src = [
    ("Yahoo Finance — Financial Statements","https://finance.yahoo.com/quote/PEGA/financials/","Income statement, balance sheet, cash flow"),
    ("Yahoo Finance — Key Statistics","https://finance.yahoo.com/quote/PEGA/key-statistics/","Valuation multiples, beta, share count, short data"),
    ("Yahoo Finance — Analyst Estimates","https://finance.yahoo.com/quote/PEGA/analysis/","Revenue/EPS consensus, revisions, surprise history"),
    ("FRED — DGS10 (10Y Treasury)","https://fred.stlouisfed.org/series/DGS10","Risk-free rate: 4.49% as of 2026-06-17"),
    ("Yahoo Finance — Summary","https://finance.yahoo.com/quote/PEGA/","Price, market cap, enterprise value, dividend"),
]
for i,(s,u,n) in enumerate(src, 1):
    fmt(ws, i+2, 1, f"{i}. {s}", True); fmt(ws, i+2, 2, u + " | " + n)
ws.column_dimensions['A'].width = 40; ws.column_dimensions['B'].width = 100

# ── Save ──
path = "/home/refcell/dev/capital/models/[2026-06-19] Pegasystems Model.xlsx"
wb.save(path)
print(f"Saved: {path}")
# Verify
wb2 = openpyxl.load_workbook(path)
print(f"Sheets: {wb2.sheetnames}")
for name in wb2.sheetnames:
    ws = wb2[name]
    print(f"  {name}: {ws.max_row} rows x {ws.max_column} cols")
print("OK — workbook verified.")
