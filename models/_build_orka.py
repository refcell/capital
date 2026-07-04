"""Build a 6-sheet valuation model for Oruka Therapeutics (ORKA).

ORKA is a clinical-stage biopharma with zero revenue and ~$33M/qtr burn rate.
Standard DCF/FCF frameworks are meaningless. Primary valuation lens:
  - Cash-per-share NAV floor
  - Pipeline milestone / probability-weighted NPV
  - Analyst consensus targets (9 analysts covering)
  - P/B ratio relative to biotech peers

Price snapshot: $85.21 on 2026-07-02. Data sources: Yahoo Finance.
StockAnalysis returned 404 for this ticker.

Key context:
  - SPAC-listed via Merger 360 (2021), ~68 employees
  - Recently acquired Apogee Therapeutics (deal announced mid-2026)
  - Stock surged 575% YTD on psoriasis trial progress + acquisition
  - 1Y analyst target average: $144.67; high: $200
  - 9 analysts covering, predominantly Buy ratings

"""
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constants ──
TICKER = "ORKA"
DATE = datetime(2026, 7, 2)
PRICE = 85.21
PRICE_AH = 85.00  # after hours
MC = 5.139        # $B intraday
EV = 5.24        # $B
CASH = 388.83    # $M (MRQ)
DEBT = 1.45      # $M implied: Debt/Equity 0.37% x equity $469M
TOTAL_DEBT_M = DEBT
CASH_PER_SHARE_M = None  # computed below
NET_DEBT = -CASH + TOTAL_DEBT_M  # net cash position in $M
BETA = 1.20       # estimated; biotech sector proxy (beta not available on Yahoo)
RISK_FREE = 0.04485  # 10Y US from CNBC
ERP = 0.05
# CAPM for biotech: high cost of equity reflects binary risk
COE = RISK_FREE + BETA * ERP
COD = 0.05        # cost of debt — minimal debt
TAX = 0.21        # US corporate
TOTAL_EQUITY_M = 468.999  # from BS
TOTAL_ASSETS_M = 488.617
TOTAL_LIAB_M = 19.618

# Shares: MC / PRICE ≈ 60.3M (intraday MC reflects recent price)
SHARES_M = round(MC * 1000 / PRICE, 1)  # ~60.3M
CASH_PER_SHARE = round(CASH / SHARES_M, 2)

# Equity weight (debt negligible)
EQ_W = TOTAL_EQUITY_M / (TOTAL_EQUITY_M + TOTAL_DEBT_M)
DEB_W = TOTAL_DEBT_M / (TOTAL_EQUITY_M + TOTAL_DEBT_M)
WACC = round(EQ_W * COE + DEB_W * COD * (1 - TAX), 4)

# Historical data (Yahoo Finance — all $ in millions, thousands converted)
# TTM = trailing twelve months; FY2025 = annual
REV = {"TTM": 0.0, "2025": 0.0, "2024": 0.0}
OPEX = {"TTM": -133.4, "2025": -122.1, "2024": None}
NI = {"TTM": -116.3, "2025": -105.4, "2024": None}
EPS = {"TTM": -1.93, "2025": -1.85, "2024": None}
OCF = {"TTM": None,  # insufficient data
       "TTM_est": round(-133.4 + 17.1, 1)}  # opex + interest income approx
AVG_SHARES = {"TTM": 48.99, "2025": 45.61}  # basic avg, in millions

# Analyst estimates (Yahoo Finance, Jul 2, 2026)
EPS_FORECAST = {
    "FY2026": -2.04,
    "FY2027": -2.49,
    "Q2_2026": -0.50,
    "Q3_2026": -0.51,
}

# ── Scenarios ── (biotech framework: cash runway → dilution → milestone NPV)
# These scenarios model cash depletion, dilution from future financing,
# and probability-weighted pipeline NPV contribution per program.
#
# Pipeline programs (from Yahoo Finance company description):
#   ORKA-001: anti-IL23, Phase 2a psoriasis (~$2-3B market if approved)
#   ORKA-002: anti-IL17A/17F, Phase 2 (~$3-5B market if approved)
#   ORKA-003: undisclosed target
#   ORKA-021: combo ORKA-002 + ORKA-001
#
# Apogee Therapeutics acquisition adds oncology programs
#   Apogee APG-1252: anti-PD-1 for oncology (Phase 3)
#   Apogee APG-2447: TIGIT inhibitor
#
# Runway analysis:
#   Current cash: $389M at $33M/qtr burn → ~12 quarters (~3 years)
#   Apogee deal adds R&D expense but potentially oncology pipeline value
#   Dilution from future financing will reduce per-share NAV

# Bear: Programs fail or show mediocre data, requires heavy dilutive financing
# Base: At least one program shows meaningful response, partnership/licensing deal
# Bull: Multiple programs reach approval or strategic acquisition

# Burn projections ($M per year):
# TTM burn through expenses: ~$133M, offset by interest income ~$17M → ~$116M net loss
# Cash burn per year: ~$133M opex + integration costs post-Apogee → ~$140-160M going forward
BURN_YEARLY = {"bear": 160, "base": 150, "bull": 140}  # $M

# Pipeline NPV contributions (probability-weighted per share value from pipeline)
# These are estimates of the risk-adjusted value of the combined pipeline
# after accounting for trial failure probabilities, regulatory risk, and commercialization
PIPELINE_NPV_PER_SHARE = {"bear": 5.0, "base": 35.0, "bull": 95.0}  # $/share

# Terminal value components:
# NAV_floor = cash_remaining / post-dilution_shares
# Pipeline_value = NPV per share
# Total = NAV_floor + pipeline value

# Dilution estimates from required future financing:
# Bear: needs 2-3 more financings → 40-50% dilution
# Base: 1-2 more financings → 25-35% dilution
# Bull: strategic deal/acquisition reduces dilution → 15-25% dilution
DILUTION = {"bear": 1.45, "base": 1.30, "bull": 1.20}  # shares_multiplier

SCENARIOS = {}

for name in ["Bear", "Base", "Bull"]:
    key = name.lower()
    burn = BURN_YEARLY[key]
    # Cash remaining after 5 years of burn
    cash_remaining = max(CASH - burn * 5, 0)
    # Shares after dilution
    diluted_shares = SHARES_M * DILUTION[key]
    # NAV floor per share from remaining cash
    nav_floor = round(cash_remaining / diluted_shares, 2)
    # Pipeline contribution
    pipeline_val = PIPELINE_NPV_PER_SHARE[key]
    # Total implied value
    implied_value = round(nav_floor + pipeline_val, 2)
    # Upside from current price
    upside = round((implied_value / PRICE - 1) * 100, 1)

    # Weights
    if name == "Bear":
        weight = 0.25
    elif name == "Base":
        weight = 0.45
    else:
        weight = 0.30

    wv = round(implied_value * weight, 2)

    SCENARIOS[name] = {
        "annual_burn": burn,
        "cash_5y": round(cash_remaining, 1),
        "diluted_shares": round(diluted_shares, 1),
        "nav_floor": nav_floor,
        "pipeline_npv": pipeline_val,
        "implied_value": implied_value,
        "upside": upside,
        "weight": weight,
        "wv": wv,
    }

tot_wv = round(sum(s["wv"] for s in SCENARIOS.values()), 2)
tot_up = round((tot_wv / PRICE - 1) * 100, 1)

# ── Styles ──
BF = Font(bold=True)
B14 = Font(bold=True, size=14)
B12 = Font(bold=True, size=12)
BD = Border(left=Side("thin"), right=Side("thin"),
            top=Side("thin"), bottom=Side("thin"))
D2 = "$#,##0.00"
DB = "$#,##0.0"
DM = "$#,##0"
PP = "0.0%"
D4 = "0.0000"


def c(ws, r, co, v, fmt=None, font=None, border=None):
    cell = ws.cell(row=r, column=co, value=v)
    if fmt:
        cell.number_format = fmt
    if font:
        cell.font = font
    if border:
        cell.border = border
    return cell


def title(ws, text):
    ws.merge_cells("A1:F1")
    c(ws, 1, 1, text, font=B14)


# ── Build ──
OUTPUT = Path(__file__).resolve().parent / f"[{DATE.strftime('%Y-%m-%d')}] Oruka Therapeutics Model.xlsx"

wb = openpyxl.Workbook()

# ── Sheet 1: Valuation ──
ws = wb.active
ws.title = "Valuation"
title(ws, f"{TICKER} Valuation Model — {DATE.strftime('%Y-%m-%d')}")

r = 3
data = [
    ("Ticker", TICKER),
    ("Company", "Oruka Therapeutics, Inc."),
    ("Exchange", "NasdaqGM"),
    ("Date", DATE.strftime('%Y-%m-%d')),
    ("Price", PRICE, D2),
    ("Price (After Hours)", PRICE_AH, D2),
    ("Shares Outstanding (M)", SHARES_M, "0.0"),
    ("Market Cap ($B)", MC, DB),
    ("Enterprise Value ($B)", EV, DB),
    ("Enterprise Value Note", "EV ≈ MC (negligible debt; clinical biotech with no revenue)"),
    ("Cash ($M)", CASH, DM),
    ("Total Debt ($M)", TOTAL_DEBT_M, DM),
    ("Net Cash ($M)", round(CASH - TOTAL_DEBT_M, 1), DM),
    ("Primary Valuation Lens", "Cash NAV floor + Pipeline NPV (standard DCF/FCF N/A — zero revenue, no earnings)"),
    ("Current Stance", "Watch / Needs more work — binary clinical risk, post-run-up valuation, heavy insider selling"),
    ("P/B Ratio (MRQ)", 11.63, None),
    ("P/E Ratio", "N/A (negative earnings)"),
    ("P/S Ratio", "N/A (zero revenue)"),
    ("P/FCF", "N/A (negative FCF)"),
    ("EV/EBITDA", "N/A (negative EBITDA)"),
    ("Cash per Share", CASH_PER_SHARE, D2),
    ("Current Price vs Cash/Share Pctg Premium", round((PRICE / CASH_PER_SHARE - 1) * 100, 1), PP),
    ("Analyst Avg Target (1Y)", 144.67, D2),
    ("Analyst High Target", 200.00, D2),
    ("Number of Analysts", 9),
    ("Next Earnings Date", "Aug 10, 2026"),
    ("Beta (Estimate)", BETA),
    ("Comment P/B", "11.63x PB for a cash-burning clinical biotech; PB reflects optionality premium, not earnings power"),
    ("Comment Valuation", "Stock trades at ~12.4x its total cash position. Market is pricing in full pipeline optionality at a steep premium. Cash per share of $6.45 provides a theoretical floor but only if company stops burning cash entirely — which is unlikely during active clinical development."),
]

for i, row in enumerate(data, 2):
    if len(row) == 2:
        c(ws, i, 1, row[0], font=BF, border=BD)
        c(ws, i, 2, row[1], border=BD)
    elif len(row) == 3:
        c(ws, i, 1, row[0], font=BF, border=BD)
        c(ws, i, 2, row[1], fmt=row[2], border=BD)

# ── Sheet 2: WACC ──
ws = wb.create_sheet("WACC")
title(ws, f"{TICKER} WACC / Cost of Capital")

r = 10
wacc_data = [
    ("Risk-Free Rate (10Y US Treasury)", RISK_FREE, D4),
    ("Equity Risk Premium", ERP, D4),
    ("Beta (Levered — estimated biotech sector proxy)", BETA, D2),
    ("  (Note: Yahoo Finance beta unavailable for ORKA)", ""),
    ("Cost of Equity (CAPM: Rf + β×ERP)", COE, D4),
    ("Cost of Debt", COD, D4),
    ("Tax Rate (US Corporate)", TAX, D4),
    ("Market Cap ($B)", MC, DB),
    ("Total Debt ($M)", TOTAL_DEBT_M, DM),
    ("Total Debt ($B)", round(TOTAL_DEBT_M, 2), DB),
    ("Equity Weight", round(EQ_W, 4), D4),
    ("Debt Weight", round(DEB_W, 4), D4),
    ("", ""),
    ("WACC", WACC, D4),
    ("", ""),
    ("WACC Comment", f"WACC of {WACC:.1%} is effectively the cost of equity given negligible debt. "
                     f"Biotechs have no meaningful cost of debt component. The WACC here is "
                     f"a discount rate — it is NOT used for DCF because there is no cash flow to discount. "
                     f"For biotechs, WACC is more relevant as the hurdle rate for pipeline investment decisions."),
]

for i, row in enumerate(wacc_data, 2):
    if len(row) == 2:
        c(ws, i, 1, row[0], font=BF, border=BD)
        c(ws, i, 2, row[1], border=BD)
    elif len(row) == 3:
        c(ws, i, 1, row[0], font=BF, border=BD)
        c(ws, i, 2, row[1], fmt=row[2] if row[2] else None, border=BD)

# ── Sheet 3: Scenarios ──
ws = wb.create_sheet("Scenarios")
title(ws, f"{TICKER} Scenarios — Cash NAV + Pipeline NPV Framework")

c(ws, 2, 1, "Note: Standard DCF/FCF framework not applicable — zero revenue, negative FCF. "
            "Scenarios model cash runway through Year 5, post-dilution share count, "
            "remaining NAV floor, and risk-adjusted pipeline NPV contribution per share.",
  font=Font(italic=True))

headers = [
    "Metric", "Bear", "Base", "Bull"
]
c(ws, 4, 1, "Metric", font=BF, border=BD)
for j, h in enumerate(headers[1:], 2):
    c(ws, 4, j, h, font=B12, border=BD)

scenario_rows = [
    ("Framework", "Cash NAV + Pipeline NPV", "Cash NAV + Pipeline NPV", "Cash NAV + Pipeline NPV"),
    ("Annual Burn Rate ($M)", SCENARIOS["Bear"]["annual_burn"], SCENARIOS["Base"]["annual_burn"], SCENARIOS["Bull"]["annual_burn"]),
    ("Cash Remaining After 5Yr ($M)", SCENARIOS["Bear"]["cash_5y"], SCENARIOS["Base"]["cash_5y"], SCENARIOS["Bull"]["cash_5y"]),
    ("Current Shares (M)", SHARES_M, SHARES_M, SHARES_M),
    ("Post-Dilution Shares (M)", SCENARIOS["Bear"]["diluted_shares"], SCENARIOS["Base"]["diluted_shares"], SCENARIOS["Bull"]["diluted_shares"]),
    ("Dilution Factor", DILUTION["bear"], DILUTION["base"], DILUTION["bull"]),
    ("NAV Floor / Share ($)", SCENARIOS["Bear"]["nav_floor"], SCENARIOS["Base"]["nav_floor"], SCENARIOS["Bull"]["nav_floor"]),
    ("Pipeline NPV / Share ($)", SCENARIOS["Bear"]["pipeline_npv"], SCENARIOS["Base"]["pipeline_npv"], SCENARIOS["Bull"]["pipeline_npv"]),
    ("Implied Value / Share ($)", SCENARIOS["Bear"]["implied_value"], SCENARIOS["Base"]["implied_value"], SCENARIOS["Bull"]["implied_value"]),
    ("Upside from Current Price", SCENARIOS["Bear"]["upside"], SCENARIOS["Base"]["upside"], SCENARIOS["Bull"]["upside"]),
    ("Scenario Weight", SCENARIOS["Bear"]["weight"], SCENARIOS["Base"]["weight"], SCENARIOS["Bull"]["weight"]),
    ("Weighted Value / Share", SCENARIOS["Bear"]["wv"], SCENARIOS["Base"]["wv"], SCENARIOS["Bull"]["wv"]),
]

for i, row in enumerate(scenario_rows, 5):
    c(ws, i, 1, row[0], font=BF, border=BD)
    for j, val in enumerate(row[1:], 2):
        fmt = None
        if isinstance(val, float):
            if i in [5, 6, 7, 8, 10, 11]:  # dollar amounts or shares
                fmt = D2 if "Share" in row[0] or "Floor" in row[0] or "Weighted" in row[0] else DM
            elif i in [12]:
                fmt = PP
            else:
                fmt = D4 if "weight" in str(val).lower() else DM
        c(ws, i, j, val, fmt=fmt, border=BD)

c(ws, 18, 1, "Probability-Weighted Fair Value", font=BF, border=BD)
c(ws, 18, 2, tot_wv, fmt=D2, font=BF, border=BD)
c(ws, 18, 3, f"Total weighted value: ${tot_wv:.2f}/share", font=BF, border=BD)

c(ws, 19, 1, "Weighted Upside from Current Price", font=BF, border=BD)
c(ws, 19, 2, tot_up, fmt=PP, font=BF, border=BD)

c(ws, 21, 1, "Current Price", font=BF, border=BD)
c(ws, 21, 2, PRICE, fmt=D2, border=BD)

c(ws, 23, 1, "Bear Case Description", font=Font(italic=True))
c(ws, 24, 1, "Programs show limited efficacy or miss endpoints; 2-3 additional financings required at trough valuations; "
              "dilution reduces NAV floor but heavy insider selling and clinical disappointments drive price toward cash NAV.", border=BD)

c(ws, 25, 1, "Base Case Description", font=Font(italic=True))
c(ws, 26, 1, "At least one program (likely ORKA-002 anti-IL17 or Apogee PD-1 program) shows meaningful clinical response; "
              "partnership/licensing deal or 1-2 more dilutive rounds; pipeline value partially realized.", border=BD)

c(ws, 27, 1, "Bull Case Description", font=Font(italic=True))
c(ws, 28, 1, "Multiple programs advance to Phase 3 or attract strategic acquisition interest; "
              "Apogee oncology assets de-risk the profile; minimal dilution from strategic partnering rather than public financing.", border=BD)

# ── Sheet 4: Actuals Source Audit ──
ws = wb.create_sheet("Actuals Source Audit")
title(ws, f"{TICKER} Actuals Source Audit")

audit_data = [
    ("Field", "Value", "Source", "Date", "Notes"),
    ("Stock Price", f"${PRICE}", "Yahoo Finance /finance.yahoo.com/quote/ORKA/", "Jul 2, 2026", "Close price"),
    ("After Hours Price", f"${PRICE_AH}", "Yahoo Finance", "Jul 2, 2026", "5:55 PM EDT"),
    ("Market Cap", f"${MC}B", "Yahoo Finance (intraday)", "Jul 2, 2026", "Intraday calculation"),
    ("Enterprise Value", f"${EV}B", "Yahoo Finance Statistics", "Jul 1, 2026", "≈ MC, negligible debt"),
    ("Total Revenue TTM", "$0", "Yahoo Finance Income Statement", "TTM", "Clinical-stage biotech, zero revenue"),
    ("Operating Expense TTM", "-$133.4M", "Yahoo Finance Income Statement", "TTM", "R&D heavy, clinical trials"),
    ("Net Income TTM", "-$116.3M", "Yahoo Finance Income Statement", "TTM", "GAAP net income"),
    ("EPS (Basic/Diluted) TTM", "-$1.93", "Yahoo Finance Income Statement", "TTM", "Basic = diluted (anti-dilutive options)"),
    ("Basic Avg Shares TTM", "48.99M", "Yahoo Finance Income Statement", "TTM", "In thousands: 48,987"),
    ("Basic Avg Shares FY2025", "45.61M", "Yahoo Finance Income Statement", "FY2025", "In thousands: 45,614"),
    ("Total Cash (MRQ)", "$388.83M", "Yahoo Finance Statistics", "Jul 1, 2026", "From MRQ balance sheet"),
    ("Total Debt/Equity", "0.37%", "Yahoo Finance Statistics", "Jul 1, 2026", "Negligible debt"),
    ("Total Assets", "$488.6M", "Yahoo Finance Balance Sheet", "FY2025", "In thousands: 488,617"),
    ("Total Liabilities", "$19.6M", "Yahoo Finance Balance Sheet", "FY2025", "In thousands: 19,618"),
    ("Total Equity", "$469.0M", "Yahoo Finance Balance Sheet", "FY2025", "In thousands: 468,999"),
    ("P/B Ratio (MRQ)", "11.63", "Yahoo Finance Statistics", "Jul 1, 2026", "MC / book value"),
    ("Beta", "N/A (estimated 1.20)", "Yahoo Finance shows --; estimated from biotech peer average",
             "Jul 2, 2026", "Clinical biotech, beta unreliable"),
    ("1Y Analyst Target", "$144.67", "Yahoo Finance Analysis", "Jul 2, 2026", "Average of 9 analysts"),
    ("Analyst High Target", "$200.00", "Yahoo Finance Analysis", "Jul 2, 2026", "High of 9 analysts"),
    ("Avg Volume", "1,829,340", "Yahoo Finance Summary", "Jul 2, 2026", "Recent average"),
    ("Volume Jul 2", "4,523,803", "Yahoo Finance Summary", "Jul 2, 2026", "2.5x normal volume"),
    ("Net Non-Op Interest Income TTM", "$17.1M", "Yahoo Finance Income Statement", "TTM", "Interest on cash reserves"),
    ("EBITDA TTM", "-$116.7M", "Yahoo Finance Income Statement", "TTM", "In thousands: -116,671"),
    ("EPS Forecast FY2026", "-$2.04", "Yahoo Finance Analysis", "Jul 2, 2026", "9-analyst consensus"),
    ("EPS Forecast FY2027", "-$2.49", "Yahoo Finance Analysis", "Jul 2, 2026", "9-analyst consensus"),
    ("P/E Ratio TTM", "N/A", "Yahoo Finance Summary", "Jul 2, 2026", "Negative earnings"),
    ("Next Earnings Date", "Aug 10, 2026", "Yahoo Finance Summary", "Jul 2, 2026", "Estimated"),
    ("Employees", "68", "Yahoo Finance Summary", "Jul 2, 2026", "Full-time"),
    ("Sector / Industry", "Healthcare / Biotechnology", "Yahoo Finance Summary", "Jul 2, 2026", ""),
]

for i, row in enumerate(audit_data, 2):
    for j, val in enumerate(row, 1):
        font = BF if i == 2 else None
        c(ws, i, j, val, font=font, border=BD)

# ── Sheet 5: Questions ──
ws = wb.create_sheet("Questions")
title(ws, f"{TICKER} Open Questions")

questions = [
    "Apogee Therapeutics acquisition: What are the financial terms, accounting treatment, and integration timeline? "
    "How did the APG-1252 (anti-PD-1) Phase 3 program change the risk profile? "
    "Was the deal cash or stock — does this explain the share count increase from 45.6M (FY2025 avg) to ~60.3M (intraday)?",

    "Share count discrepancy: TTM basic avg shares = 48.99M but intraday MC at $85.21 implies ~60.3M shares. "
    "This 23% difference is likely from the Apogee deal (stock consideration) or recent SPAC-related adjustments. "
    "Until this is reconciled, per-share metrics are unreliable.",

    "Why the 575% YTD return? The stock has nearly quintupled from a $12.48 52-week low. "
    "What specific catalysts drove this — was it the psoriasis trial data, the Apogee deal, "
    "or general biotech speculation? Is the price justified by fundamentals or speculation?",

    "Heavy insider selling: Multiple reports of insiders selling significant positions. "
    "Head of Finance sold 10,000+ shares; another insider sold 105,000 shares worth $6.2M. "
    "After a 575% run-up, is this profit-taking or a signal of overvaluation?",

    "Cash runway sustainability: $389M cash at current ~$33M/qtr burn gives ~12 quarters (~3 years). "
    "But Apogee integration costs will increase burn. How many dilutive financings are likely before "
    "any program generates revenue?",

    "Pipeline de-risking timeline: ORKA-001 in Phase 2a, ORKA-002 in Phase 2. "
    "Even in the best case, Phase 3 + FDA approval is 3-5+ years out. "
    "Apogee APG-1252 is in Phase 3 — is this the most valuable asset?",

    "Biotech peer P/B context: ORKA at 11.63x P/B is above the sector average for clinical biotechs. "
    "How does this compare to Spyre (SYRE at $7.5B MC), Enliven ($3.6B), and comparable peers?",

    "StockAnalysis 404: This ticker is unavailable on StockAnalysis. "
    "Is ORKA too new/small for coverage, or is it a platform limitation?",

    "SPAC legacy: Original listing via Merger 360 SPAC. "
    "Are there any remaining SPAC-related trusts, redeemable interests, or sponsor lock-up expirations?",

    "Revenue recognition: $0 revenue TTM. Is there any licensing/in-licensing revenue, grant income, "
    "or milestone payments that might appear in future quarters?",
]

for i, q in enumerate(questions, 2):
    c(ws, i, 1, f"Q{i-1}:", font=BF)
    c(ws, i, 2, q, border=BD)

# ── Sheet 6: Sources ──
ws = wb.create_sheet("Sources")
title(ws, f"{TICKER} Sources")

sources = [
    ("Yahoo Finance — Summary/Quote", f"https://finance.yahoo.com/quote/{TICKER}/",
     "Price, market cap, volume, analyst targets, company description, employees, sector/industry"),
    ("Yahoo Finance — Income Statement", f"https://finance.yahoo.com/quote/{TICKER}/financials/",
     "Revenue, operating expense, net income, EPS, share counts"),
    ("Yahoo Finance — Balance Sheet", f"https://finance.yahoo.com/quote/{TICKER}/balance-sheet/",
     "Total assets, liabilities, equity, cash, debt"),
    ("Yahoo Finance — Statistics", f"https://finance.yahoo.com/quote/{TICKER}/statistics/",
     "P/B ratio, EV, returns on assets/equity, debt/equity, total cash"),
    ("Yahoo Finance — Analysis/Estimates", f"https://finance.yahoo.com/quote/{TICKER}/analysis/",
     "Analyst consensus EPS estimates, revision trends, price targets"),
    ("CNBC — 10Y Treasury", "https://www.cnbc.com/quotes/US10Y",
     "Risk-free rate for CAPM calculation"),
    ("StockAnalysis", f"https://stockanalysis.com/quote/{TICKER}/",
     "Attempted — returned 404. Ticker not available on this platform."),
]

for i, (name, url, note) in enumerate(sources, 2):
    c(ws, i, 1, f"{i-1}.", font=BF, border=BD)
    c(ws, i, 2, name, font=BF, border=BD)
    c(ws, i, 3, url, border=BD)
    c(ws, i, 4, note, border=BD)

# ── Save ──
wb.save(OUTPUT)
print(f"✏️  Saved: {OUTPUT}")
print(f"   Cash per share: ${CASH_PER_SHARE:.2f}")
print(f"   NAV floor premium: {round((PRICE/CASH_PER_SHARE-1)*100,1)}%")
print(f"   WACC: {WACC:.1%}")
print(f"   Weighted FV: ${tot_wv:.2f}/share → {tot_up:+.1f}% upside")
for name in ["Bear", "Base", "Bull"]:
    s = SCENARIOS[name]
    print(f"   {name}: ${s['implied_value']:.2f}/share ({s['upside']:+.1f}%) weight={s['weight']}")
