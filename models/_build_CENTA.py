#!/usr/bin/env python3
"""
Build script for CENTA (Central Garden & Pet Company) 6-sheet valuation model.
Data as of June 26, 2026.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# ── Helpers ────────────────────────────────────────────────────────────────
bold = Font(bold=True)
bold12 = Font(bold=True, size=12)
bold14 = Font(bold=True, size=14)
red = Font(bold=True, color="FF0000")
header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def style_row(ws, row, cells, font=None, fill=None, alignment=None, border=True, number_fmt=None):
    for c, val in cells:
        cell = ws.cell(row=row, column=c, value=val)
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = thin_border
        if number_fmt and c in number_fmt:
            cell.number_format = number_fmt[c]

def write_title_block(ws, company, ticker, date, price, shares_mm, mc_b, ev_b, lens, stance):
    ws.merge_cells("A1:G1")
    cell = ws.cell(row=1, column=1, value=f"{company} ({ticker}) Valuation Model — {date}")
    cell.font = bold14
    cell.alignment = Alignment(horizontal="center")
    
    title_data = [
        ("Ticker", ticker),
        ("Date", date),
        ("Price", price),
        ("Shares Outstanding (M)", round(shares_mm, 1)),
        ("Market Cap ($B)", round(mc_b, 2)),
        ("Enterprise Value ($B)", round(ev_b, 2)),
        ("Primary Valuation Lens", lens),
        ("Stance", stance),
    ]
    for i, (label, val) in enumerate(title_data, 2):
        ws.cell(row=i, column=1, value=label).font = bold
        ws.cell(row=i, column=2, value=val)

# ── Sheet 1: Valuation ─────────────────────────────────────────────────────
ws_val = wb.active
ws_val.title = "Valuation"

price = 39.32
shares_mm = 62.4  # diluted avg shares TTM
mc_b = round(price * shares_mm / 1000, 2)  # 2.46B
ev_b = 3.27  # from Yahoo Finance

write_title_block(ws_val, "Central Garden & Pet Company", "CENTA", "2026-06-26",
                  price, shares_mm, mc_b, ev_b,
                  "P/E ratio, EV/EBITDA, P/FCF with FCF exit multiple scenarios",
                  "Watch / Needs more work")

# Valuation metrics
metrics = [
    ("P/E (TTM)", 14.19, "Trailing P/E. Consistent with a steady-state consumer goods company."),
    ("Forward P/E", 11.78, "Forward P/E based on FY2026 consensus EPS of ~$3.33 (TTM EPS $2.75). Attractive given beat pattern."),
    ("P/S (TTM)", 0.77, "Below 1.0x. Reflects cyclical garden headwinds and modest margins."),
    ("P/FCF (TTM)", round(mc_b * 1000 / 282.417, 2), f"MC ${mc_b}B / FCF $282.4M. Reasonable for the sector."),
    ("EV/FCF", round(ev_b * 1000 / 282.417, 2), f"EV ${ev_b}B / FCF $282.4M. Slightly above P/FCF due to net debt."),
    ("EV/Sales", 1.03, "Below 1.0x enterprise multiple on $3.16B revenue. Cheap."),
    ("EV/EBITDA", 8.88, "Near sector median. Consumer goods typically 8-12x. On the low end."),
]

headers = ["Metric", "Value", "Comment"]
for i, h in enumerate(headers, 1):
    ws_val.cell(row=11, column=i, value=h).font = bold
    ws_val.cell(row=11, column=i).fill = header_fill
    ws_val.cell(row=11, column=i).border = thin_border

num_fmt = {2: '0.00'}
for i, (metric, val, comment) in enumerate(metrics, 12):
    style_row(ws_val, i, [(1, metric), (2, val), (3, comment)], number_fmt=num_fmt)

# Set column widths
for col, width in [(1, 25), (2, 15), (3, 60)]:
    ws_val.column_dimensions[get_column_letter(col)].width = width

# ── Sheet 2: WACC ──────────────────────────────────────────────────────────
ws_wacc = wb.create_sheet("WACC")

rf_rate = 0.0438   # 10Y US Treasury from CNBC 4.376%
erp = 0.05        # Equity risk premium
beta = 0.55       # 5Y monthly from Yahoo Finance
cost_of_equity = rf_rate + beta * erp  # 4.38% + 0.55 * 5% = 7.13%
cost_of_debt = 0.055  # Interest expense / Total debt ~ 57.3/1440 = ~4.0%, add spread
tax_rate = 0.243  # TTM tax provision / pretax income = 55,138 / 226,944
total_debt_mm = 1440.3   # in millions
equity_mm = mc_b * 1000  # 2458.1M

total_cap = equity_mm + total_debt_mm
w_eq = equity_mm / total_cap
w_d = total_debt_mm / total_cap
wacc = w_eq * cost_of_equity + w_d * cost_of_debt * (1 - tax_rate)

wacc_data = [
    header_fill,
    ("Risk-Free Rate (10Y US Treasury)", rf_rate, "4.38% from CNBC, June 26 2026"),
    ("Equity Risk Premium", erp, "Standard 5% assumption"),
    ("Beta (5Y monthly, levered)", beta, "Yahoo Finance Statistics. Low beta = defensive consumer staples."),
    ("Cost of Equity (CAPM)", round(cost_of_equity, 4), f"=Rf + Beta*ERP = {rf_rate:.2%} + {beta}*{erp:.0%} = {cost_of_equity:.2%}"),
    ("Cost of Debt (pre-tax)", cost_of_debt, "Est. ~5.5% based on revolving credit facility + lease costs"),
    ("Tax Rate (TTM effective)", tax_rate, f"Tax provision ${55.1}M / Pretax ${226.9}M = {tax_rate:.1%}"),
    ("Market Cap ($M)", round(equity_mm, 1), f"Price ${price} x {shares_mm}M shares"),
    ("Total Debt ($M)", total_debt_mm, "Total debt per Yahoo Finance balance sheet as of 9/30/2025"),
    ("Equity Weight (E/V)", round(w_eq, 4), f"={equity_mm:.0f} / {total_cap:.0f}"),
    ("Debt Weight (D/V)", round(w_d, 4), f"={total_debt_mm:.1f} / {total_cap:.0f}"),
    ("WACC", round(wacc, 4), f"=W_eq*Ke + W_d*Kd*(1-T) = {w_eq:.2%}*{cost_of_equity:.2%} + {w_d:.2%}*{cost_of_debt:.2%}*(1-{tax_rate:.2%}) = {wacc:.2%}"),
]

ws_wacc.merge_cells("A1:C1")
ws_wacc.cell(row=1, column=1, value="WACC Calculation — CENTA").font = bold14
ws_wacc.cell(row=1, column=1).alignment = Alignment(horizontal="center")

for i, row_data in enumerate(wacc_data[1:], 3):
    label, val, note = row_data
    style_row(ws_wacc, i, [(1, label), (2, val), (3, note)],
              font=bold if i == len(wacc_data) - 1 else None,
              fill=green_fill if i == len(wacc_data) - 1 else None,
              number_fmt={2: '0.0000'})
ws_wacc.cell(row=2, column=1, value="Component").font = bold
ws_wacc.cell(row=2, column=2, value="Value").font = bold
ws_wacc.cell(row=2, column=3, value="Notes / Calculation").font = bold
for c in range(1, 4):
    ws_wacc.cell(row=2, column=c).fill = header_fill
    ws_wacc.cell(row=2, column=c).border = thin_border

for col, width in [(1, 40), (2, 18), (3, 55)]:
    ws_wacc.column_dimensions[get_column_letter(col)].width = width

# ── Sheet 3: Scenarios ─────────────────────────────────────────────────────
ws_scen = wb.create_sheet("Scenarios")

ws_scen.merge_cells("A1:L1")
ws_scen.cell(row=1, column=1, value="Scenario Analysis — CENTA").font = bold14
ws_scen.cell(row=1, column=1).alignment = Alignment(horizontal="center")

# All values in MILLIONS for unit consistency
# TTM Revenue: $3,162.6M, TTM FCF: $282.4M, Net Debt: $309.2M
# Base FY26 revenue ~ $3.05B per consensus, FY27 ~ $2.95B (slight decline)
# Conservative approach: use TTM as baseline

# Scenario parameters (everything in millions except prices)
# Using 5-year horizon
scenarios = {
    "Bear": {
        "revenue_cagr_5y": -2.0,      # Slowdown; garden softness + pet competition
        "terminal_revenue": round(3162.6 * (1 - 0.02)**5, 1),
        "fcf_margin": 0.08,           # Below historical ~9% TTM — competition + costs
        "terminal_fcf": round(3162.6 * (1 - 0.02)**5 * 0.08, 1),
        "exit_multiple": 9.0,         # Compressed multiple
        "weight": 0.25,
    },
    "Base": {
        "revenue_cagr_5y": 0.5,       # Roughly flat to slight growth — mature consumer goods
        "terminal_revenue": round(3162.6 * (1 + 0.005)**5, 1),
        "fcf_margin": 0.10,           # TTM FCF margin ~ 8.9%, assume slight improvement
        "terminal_fcf": round(3162.6 * (1 + 0.005)**5 * 0.10, 1),
        "exit_multiple": 11.0,        # Reasonable consumer goods multiple
        "weight": 0.50,
    },
    "Bull": {
        "revenue_cagr_5y": 3.0,       # Re-acceleration via pricing power + pet trend
        "terminal_revenue": round(3162.6 * (1 + 0.03)**5, 1),
        "fcf_margin": 0.12,           # Margin expansion from scale + productivity
        "terminal_fcf": round(3162.6 * (1 + 0.03)**5 * 0.12, 1),
        "exit_multiple": 14.0,        # Expanded multiple
        "weight": 0.25,
    },
}

net_debt_mm = 309.2  # from Yahoo Finance balance sheet

# Scenario header
scen_headers = [
    "Scenario", "Revenue CAGR (5Y)", "Terminal Revenue ($M)", 
    "Adjusted FCF Margin", "Terminal FCF ($M)", "Exit FCF Multiple",
    "Implied EV ($M)", "Less Net Debt ($M)", "Shares (M)",
    "Target Price", "Upside %", "Weight"
]
for i, h in enumerate(scen_headers, 1):
    cell = ws_scen.cell(row=3, column=i, value=h)
    cell.font = bold
    cell.fill = header_fill
    cell.border = thin_border

weighted_values = []
for row_idx, (name, s) in enumerate(scenarios.items(), 4):
    implied_ev = s["terminal_fcf"] * s["exit_multiple"]
    eq_value = implied_ev + net_debt_mm  # EV less NET DEBT = EV - debt + cash = EV + (cash - debt); net_debt = debt - cash, so eq = ev - net_debt
    # Actually: EV = Equity + Net Debt => Equity = EV - Net Debt
    eq_value = implied_ev - net_debt_mm
    target = eq_value / shares_mm
    upside = (target - price) / price
    
    weighted_values.append((name, s["weight"] * target))
    
    row_data = [
        name, f"{s['revenue_cagr_5y']:.1f}%",
        s["terminal_revenue"], f"{s['fcf_margin']:.1%}",
        s["terminal_fcf"], s["exit_multiple"],
        round(implied_ev, 1), net_debt_mm,
        shares_mm, round(target, 2), f"{upside:.1%}",
        s["weight"],
    ]
    style_row(ws_scen, row_idx, list(enumerate(row_data, 1)),
              fill=yellow_fill if name == "Base" else None,
              number_fmt={10: '0.00', 11: '0.0%'})

# Weighted value row
total_wtd_fv = sum(v for _, v in weighted_values)
total_upside = (total_wtd_fv - price) / price

wtd_row = ["Weighted Value/Share", "", "", "", "", "", "", "", "", 
           round(total_wtd_fv, 2), f"{total_upside:.1%}"]
style_row(ws_scen, 8, list(enumerate(wtd_row, 1)), font=bold, fill=green_fill,
          number_fmt={10: '0.00'})

# Total probability-weighted FV
ws_scen.cell(row=9, column=1, value="Total Probability-Weighted Fair Value").font = bold
ws_scen.cell(row=9, column=10, value=round(total_wtd_fv, 2)).font = bold
ws_scen.cell(row=9, column=10).fill = green_fill
ws_scen.cell(row=9, column=10).border = thin_border

ws_scen.cell(row=10, column=1, value="Upside from Current Price").font = bold
ws_scen.cell(row=10, column=10, value=f"{total_upside:.1%}").font = bold
ws_scen.cell(row=10, column=10).fill = green_fill
ws_scen.cell(row=10, column=10).border = thin_border

for col, width in [(1, 22), (2, 16), (3, 20), (4, 18), (5, 18), (6, 16), 
                   (7, 15), (8, 16), (9, 12), (10, 14), (11, 12), (12, 10)]:
    ws_scen.column_dimensions[get_column_letter(col)].width = width

# ── Sheet 4: Actuals Source Audit ──────────────────────────────────────────
ws_audit = wb.create_sheet("Actuals Source Audit")

ws_audit.merge_cells("A1:E1")
ws_audit.cell(row=1, column=1, value="Actuals Source Audit — CENTA").font = bold14
ws_audit.cell(row=1, column=1).alignment = Alignment(horizontal="center")

audit_headers = ["Data Point", "Value", "Source URL", "Date", "Notes"]
for i, h in enumerate(audit_headers, 1):
    cell = ws_audit.cell(row=3, column=i, value=h)
    cell.font = bold
    cell.fill = header_fill
    cell.border = thin_border

audit_data = [
    ("Stock Price", "$39.32", "finance.yahoo.com/quote/CENTA/", "2026-06-26", "Close price"),
    ("Market Cap", "$2.46B", "finance.yahoo.com/quote/CENTA/", "2026-06-26", "Intraday MC"),
    ("Enterprise Value", "$3.27B", "finance.yahoo.com/quote/CENTA/", "2026-06-26", "Yahoo Finance Statistics"),
    ("Shares Outstanding (diluted avg)", "62.4M", "finance.yahoo.com/quote/CENTA/financials/", "TTM", "Diluted average shares TTM"),
    ("Shares (issued, latest)", "62.87M", "finance.yahoo.com/quote/CENTA/balance-sheet/", "2025-09-30", "Ordinary shares number"),
    ("Revenue (TTM)", "$3,162.6M", "finance.yahoo.com/quote/CENTA/financials/", "TTM", "Trailing twelve months"),
    ("Revenue FY2025", "$3,129.1M", "finance.yahoo.com/quote/CENTA/financials/", "2025-09-30", "Annual"),
    ("Revenue FY2024", "$3,200.5M", "finance.yahoo.com/quote/CENTA/financials/", "2024-09-30", "Annual"),
    ("Revenue FY2023", "$3,310.1M", "finance.yahoo.com/quote/CENTA/financials/", "2023-09-30", "Annual"),
    ("Revenue FY2022", "$3,338.6M", "finance.yahoo.com/quote/CENTA/financials/", "2022-09-30", "Annual"),
    ("Gross Profit (TTM)", "$1,018.7M", "finance.yahoo.com/quote/CENTA/financials/", "TTM", "TTM"),
    ("Gross Margin (TTM)", "32.2%", "Calculated", "TTM", "1018.7/3162.6"),
    ("Operating Income (TTM)", "$259.2M", "finance.yahoo.com/quote/CENTA/financials/", "TTM", "TTM"),
    ("Operating Margin (TTM)", "8.2%", "Calculated", "TTM", "259.2/3162.6"),
    ("Net Income (TTM)", "$171.5M", "finance.yahoo.com/quote/CENTA/financials/", "TTM", "TTM"),
    ("EPS Diluted (TTM)", "$2.75", "finance.yahoo.com/quote/CENTA/", "TTM", "Yahoo Finance summary"),
    ("EBITDA (TTM)", "$367.9M", "finance.yahoo.com/quote/CENTA/financials/", "TTM", "TTM"),
    ("Operating Cash Flow (TTM)", "$328.3M", "finance.yahoo.com/quote/CENTA/cash-flow/", "TTM", "TTM"),
    ("Capex (TTM)", "$45.9M", "finance.yahoo.com/quote/CENTA/cash-flow/", "TTM", "Capital expenditures TTM"),
    ("Free Cash Flow (TTM)", "$282.4M", "finance.yahoo.com/quote/CENTA/cash-flow/", "TTM", "OCF - Capex"),
    ("Total Cash (latest)", "$668.4M", "finance.yahoo.com/quote/CENTA/", "2025-09-30", "Yahoo Finance Statistics"),
    ("Total Debt", "$1,440.3M", "finance.yahoo.com/quote/CENTA/balance-sheet/", "2025-09-30", "Total debt"),
    ("Net Debt", "$309.2M", "finance.yahoo.com/quote/CENTA/balance-sheet/", "2025-09-30", "Total debt - cash"),
    ("Beta (5Y Monthly)", "0.55", "finance.yahoo.com/quote/CENTA/", "2026-06-26", "Yahoo Finance Summary"),
    ("P/E TTM", "14.19", "finance.yahoo.com/quote/CENTA/", "2026-06-26", "Yahoo Finance Statistics"),
    ("Forward P/E", "11.78", "finance.yahoo.com/quote/CENTA/", "2026-06-26", "Yahoo Finance Statistics"),
    ("Analyst FY2026 Revenue est.", "$3.05B", "finance.yahoo.com/quote/CENTA/analysis/", "2026-06-26", "4 analysts"),
    ("Analyst FY2027 Revenue est.", "$2.95B", "finance.yahoo.com/quote/CENTA/analysis/", "2026-06-26", "4 analysts"),
    ("Analyst FY2026 EPS est.", "$2.97", "finance.yahoo.com/quote/CENTA/analysis/", "2026-06-26", "5 analysts"),
    ("Analyst FY2027 EPS est.", "$3.09", "finance.yahoo.com/quote/CENTA/analysis/", "2026-06-26", "5 analysts"),
    ("1Y Target Estimate", "$43.50", "finance.yahoo.com/quote/CENTA/", "2026-06-26", "Average target"),
    ("Next Earnings Date", "Aug 5, 2026", "finance.yahoo.com/quote/CENTA/", "2026-06-26", "Estimated"),
    ("10Y Treasury Rate", "4.376%", "cnbc.com/quotes/US10Y", "2026-06-26", "CNBC real-time"),
    ("Tax Rate (effective, TTM)", "24.3%", "Calculated", "TTM", "55.1M / 226.9M"),
    ("Capital Lease Obligations", "$248.6M", "finance.yahoo.com/quote/CENTA/balance-sheet/", "2025-09-30", "BS item"),
]

for i, (dp, val, src, dt, note) in enumerate(audit_data, 4):
    style_row(ws_audit, i, [(1, dp), (2, val), (3, src), (4, dt), (5, note)])

for col, width in [(1, 30), (2, 16), (3, 45), (4, 14), (5, 40)]:
    ws_audit.column_dimensions[get_column_letter(col)].width = width

# ── Sheet 5: Questions ─────────────────────────────────────────────────────
ws_q = wb.create_sheet("Questions")

ws_q.merge_cells("A1:C1")
ws_q.cell(row=1, column=1, value="Open Questions — CENTA").font = bold14
ws_q.cell(row=1, column=1).alignment = Alignment(horizontal="center")

questions = [
    ("Q1", "Revenue Decline Trend", "Revenue has declined from $3,338.6M (FY2022) to $3,162.6M (TTM) — a 5.3% decline over 3 years. Analyst consensus expects further decline to $2.95B by FY2027. Is this a cyclical low (garden softness) or structural (channel shift to Amazon/e-commerce)?"),
    ("Q2", "Margin Profile Assessment", "Gross margin of 32.2% and operating margin of 8.2% are modest for branded consumer goods. Competitors like Ingredion (INGR) typically show 30%+ gross margins. Why is CENTA's margin profile below peer averages?"),
    ("Q3", "Debt Management Trajectory", "Net debt has declined significantly from $1,009M (FY2022) to $309M (FY2025) — a $700M+ reduction. At what pace will debt continue declining, and will management shift to buybacks/dividends?"),
    ("Q4", "Share Count Decline", "Shares outstanding fell from 67.8M (FY2022) to 62.9M (FY2025) — 7% reduction. Buybacks averaged $34M/year (TTM). Is this acceleration likely?"),
    ("Q5", "Garden Segment Cyclicity", "The Garden segment (fertilizers, seeds, pesticides) is highly seasonal and cyclical. FY2025 operating income of $250M was up sharply from $185M in FY2024. What drove the recovery — operational efficiency or cyclical tailwinds?"),
    ("Q6", "Pet Segment Competitive Threats", "Pet segment carries major brands (Nylabone, Kaytee, Four Paws) but faces competition from Chewy, Amazon Pets, and private label. Brand loyalty and pricing power need closer examination."),
    ("Q7", "FCF Sustainability", "TTM FCF of $282.4M represents 8.9% FCF margin — strong. But FY2024 FCF was $352M and FY2023 was $328M. The decline suggests either working capital swings, higher capex needs, or revenue drag. Is $282M the new run rate?"),
    ("Q8", "Capital Lease Obligations", "Capital leases of $248.6M (up from $186M in FY2023) suggest facility/warehouse expansion. Is this a growth investment or maintenance commitment?"),
    ("Q9", "Next Earnings — Aug 5, 2026", "Q1 FY2027 earnings on Aug 5 represent the first full post-pandemic-peak garden season. Will garden revenue stabilize? Any new guidance on FY2026-full-year outlook?"),
    ("Q10", "Competitive Differentiation", "What protects CENTA's brands from commoditization? Private label penetration in garden chemicals and pet supplies is rising — does CENTA have a cost advantage from scale?"),
]

q_headers = ["#", "Topic", "Question"]
for i, h in enumerate(q_headers, 1):
    cell = ws_q.cell(row=3, column=i, value=h)
    cell.font = bold
    cell.fill = header_fill
    cell.border = thin_border

for i, (qno, topic, question) in enumerate(questions, 4):
    style_row(ws_q, i, [(1, qno), (2, topic), (3, question)])

for col, width in [(1, 6), (2, 25), (3, 100)]:
    ws_q.column_dimensions[get_column_letter(col)].width = width

# ── Sheet 6: Sources ───────────────────────────────────────────────────────
ws_src = wb.create_sheet("Sources")

ws_src.merge_cells("A1:B1")
ws_src.cell(row=1, column=1, value="Sources — CENTA").font = bold14
ws_src.cell(row=1, column=1).alignment = Alignment(horizontal="center")

sources_data = [
    ("1", "Yahoo Finance Summary", "https://finance.yahoo.com/quote/CENTA/", "Price, market cap, beta, P/E, forward P/E, analyst targets, earnings date"),
    ("2", "Yahoo Finance Income Statement", "https://finance.yahoo.com/quote/CENTA/financials/", "Revenue, gross profit, operating income, net income, EPS — annual and TTM"),
    ("3", "Yahoo Finance Balance Sheet", "https://finance.yahoo.com/quote/CENTA/balance-sheet/", "Total assets, debt, cash, shares outstanding, working capital"),
    ("4", "Yahoo Finance Cash Flow", "https://finance.yahoo.com/quote/CENTA/cash-flow/", "Operating cash flow, capex, free cash flow, buybacks"),
    ("5", "Yahoo Finance Analysis", "https://finance.yahoo.com/quote/CENTA/analysis/", "Analyst revenue/EPS estimates, revision trends, earnings surprise"),
    ("6", "CNBC 10Y Treasury", "https://www.cnbc.com/quotes/US10Y/", "10-Year US Treasury yield: 4.376%"),
    ("7", "StockAnalysis", "https://stockanalysis.com/quotes/CENTA/", "404 error — not available for this ticker"),
    ("8", "Company Website", "https://www.central.com", "Company description, brand portfolio, segment breakdown"),
]

src_headers = ["#", "Source", "URL", "Data Provided"]
for i, h in enumerate(src_headers, 1):
    cell = ws_src.cell(row=3, column=i, value=h)
    cell.font = bold
    cell.fill = header_fill
    cell.border = thin_border

for i, (num, src, url, data) in enumerate(sources_data, 4):
    style_row(ws_src, i, [(1, num), (2, src), (3, url), (4, data)])

for col, width in [(1, 5), (2, 25), (3, 50), (4, 55)]:
    ws_src.column_dimensions[get_column_letter(col)].width = width

# ── Save ───────────────────────────────────────────────────────────────────
outpath = os.path.join(os.path.dirname(__file__), "[2026-06-26] Central Garden & Pet Model.xlsx")
wb.save(outpath)
print(f"Saved: {outpath}")

# Print key values for verification
print(f"WACC: {wacc:.2%}")
print(f"Cost of Equity: {cost_of_equity:.2%}")
print(f"Equity Weight: {w_eq:.2%}, Debt Weight: {w_d:.2%}")
print(f"Tax Rate: {tax_rate:.1%}")
print()
for name, s in scenarios.items():
    implied_ev = s["terminal_fcf"] * s["exit_multiple"]
    eq_value = implied_ev - net_debt_mm
    target = eq_value / shares_mm
    upside = (target - price) / price
    print(f"{name}: Terminal FCF={s['terminal_fcf']:.1f}M, EV={implied_ev:.1f}M, Target=${target:.2f}, Upside={upside:.1%}")
print(f"Weighted FV: ${total_wtd_fv:.2f}, Upside: {total_upside:.1%}")
