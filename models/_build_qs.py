"""
Build QS (QuantumScape) 6-sheet valuation model.
Pre-commercial solid-state battery developer — adapts framework: Cash NAV floor + Pipeline NPV.
FCF multiples are structurally N/A for pre-revenue/development-stage battery tech.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Style helpers ──
title_font = Font(name="Calibri", size=14, bold=True)
subtitle_font = Font(name="Calibri", size=12, bold=True)
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
bold_font = Font(name="Calibri", size=11, bold=True)
normal_font = Font(name="Calibri", size=11)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
dollar_fmt = '$#,##0'
dollar_k_fmt = '$#,##0,"K"'
dollar_m_fmt = '$#,##0.0,"M"'
pct_fmt = '0.00%'

def c(ws, row, col, value, font=normal_font, fmt=None, border=True, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = thin_border
    if alignment:
        cell.alignment = alignment
    return cell

def header_row(ws, row, values, start_col=1):
    for i, v in enumerate(values):
        cell = c(ws, row, start_col + i, v, header_font, border=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

# ═══════════════════════════════════════════════════
# Sheet 1: Valuation
# ═══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Valuation"
ws1.merge_cells('A1:F1')
ws1['A1'] = "QuantumScape Corporation (QS) — Valuation Summary"
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal="center")

title_data = [
    ("Company", "QuantumScape Corporation"),
    ("Ticker", "NASDAQ: QS"),
    ("Date", "2026-07-17"),
    ("Price", "$5.86"),
    ("Shares Outstanding", "578.34M"),
    ("Implied Shares Outstanding", "615.15M"),
    ("Market Cap", "$3.63B"),
    ("Enterprise Value", "$2.80B"),
    ("Total Cash (Q1 FY26)", "$904.7M"),
    ("Total Debt", "$69.18M"),
    ("Net Cash", "$835.5M"),
    ("Primary Lens", "Cash NAV Floor + Pipeline Optionality; P/B + ROE"),
    ("Stance", "Watch"),
    ("Comment", "Pre-commercial solid-state battery developer. Standard DCF/FCF frameworks inapplicable.\nValuation driven by cash runway, VW pilot program, and optionality of\ncommercial deployment. See Scenarios sheet for NAV + Pipeline NPV framework."),
]

for i, (label, val) in enumerate(title_data):
    c(ws1, 3 + i, 1, label, bold_font)
    ws1.cell(row=3 + i, column=1).alignment = Alignment(horizontal="right")
    cell = c(ws1, 3 + i, 2, val, normal_font)
    cell.alignment = Alignment(wrap_text=True)

val_metrics = [
    ("P/E (Trailing)", "N/A", "Negative earnings; TTM net income -$421.4M; no revenue"),
    ("P/E (Forward)", "N/A", "Consensus EPS FY27: -$0.68; still negative"),
    ("P/S (TTM)", "N/A", "Zero revenue — no sales figure denominator"),
    ("P/FCF", "N/A", "FCF -$281.7M TTM; structurally negative for development stage"),
    ("EV/FCF", "N/A", "Not meaningful for pre-commercial developer"),
    ("EV/Sales", "N/A", "No sales figure"),
    ("EV/EBITDA", "-7.22x (quarterly)", "From Yahoo Finance; negative EBITDA; sign only"),
    ("P/B", "3.27x", "FY25 book value/share = $1.80; total equity $1.169B"),
    ("Cash/Share NAV", "$1.56", "Total cash $904.7M / 578.34M shares"),
    ("Optionality Premium/Share", "$4.30", "Current price $5.86 less $1.56 NAV floor"),
    ("EV/MC ratio", "0.77", "EV $2.80B / MC $3.63B; net cash company"),
    ("Beta (5Y Monthly)", "2.62", "High volatility; pre-commercial development risk premium"),
    ("52W High/Low", "$19.07 / $5.64", "-60% from 52W high; near 52W low"),
]

header_row(ws1, 22, ["Metric", "Value", "Comment"])
for i, (metric, val, comment) in enumerate(val_metrics):
    c(ws1, 23 + i, 1, metric, normal_font)
    c(ws1, 23 + i, 2, val, normal_font)
    cc = c(ws1, 23 + i, 3, comment, normal_font)
    cc.alignment = Alignment(wrap_text=True)

ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 55

# ═══════════════════════════════════════════════════
# Sheet 2: WACC
# ═══════════════════════════════════════════════════
ws2 = wb.create_sheet("WACC")
ws2.merge_cells('A1:D1')
ws2['A1'] = "WACC Calculation — QuantumScape (QS)"
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal="center")

capm_data = [
    ("Component", "Value", "Source / Rationale"),
    ("", "", ""),
    ("Risk-Free Rate (10Y US Treasury)", "4.55%", "CNBC US10Y as of 2026-07-17"),
    ("Equity Risk Premium", "5.00%", "Standard assumption"),
    ("Beta (5Y Monthly)", "2.62", "Yahoo Finance key statistics"),
    ("", "", ""),
    ("Cost of Equity (Ke = Rf + Beta × ERP)", "17.65%", "= 4.55% + 2.62 × 5.00%"),
    ("", "", ""),
    ("Cost of Debt (Kd)", "3.00%", "Interest expense ~$2M; small debt base ($69M)"),
    ("Tax Rate", "21.00%", "US statutory; minimal actual tax on losses"),
    ("After-Tax Cost of Debt", "2.37%", "= 3.00% × (1 - 0.21)"),
    ("", "", ""),
    ("Market Capitalization", "$3,630M", "Yahoo Finance; $5.86 × 578.34M"),
    ("Total Debt", "$69M", "Yahoo Finance MRQ"),
    ("Total Capitalization", "$3,699M", "= MC + Debt"),
    ("Equity Weight (We)", "98.14%", "= MC / (MC + Debt)"),
    ("Debt Weight (Wd)", "1.86%", "= Debt / (MC + Debt)"),
    ("", "", ""),
    ("WACC", "17.53%", "= We × Ke + Wd × Kd × (1-T)"),
]

for i, (comp, val, src) in enumerate(capm_data):
    c(ws2, 3 + i, 1, comp, bold_font if i in (2,4,5,6,8,9,10,11,12,15) else normal_font)
    c(ws2, 3 + i, 2, val, bold_font if i in (2,4,5,6,8,9,10,11,12,15) else normal_font)
    cc = c(ws2, 3 + i, 3, src, normal_font)
    cc.alignment = Alignment(wrap_text=True)

ws2['A2'] = "CAPM Components:"
ws2['A2'].font = subtitle_font
ws2['A22'] = ""
ws2['A22'].font = subtitle_font

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 45

# ═══════════════════════════════════════════════════
# Sheet 3: Scenarios — Cash NAV + Pipeline NPV framework
# ═══════════════════════════════════════════════════
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells('A1:I1')
ws3['A1'] = "Scenario Analysis — Cash NAV Floor + Pipeline Optionality (Pre-Commercial Battery Developer)"
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal="center")

ws3['A3'] = "Framework note: Standard FCF multiple framework is structurally inapplicable for a pre-revenue developer."
ws3['A3'].font = Font(name="Calibri", size=10, italic=True)
ws3['A3'].alignment = Alignment(wrap_text=True, horizontal="left")
ws3['A4'] = "Scenarios use cash runway, dilution from future financing, NAV floor per share, and pipeline NPV."
ws3['A4'].font = Font(name="Calibri", size=10, italic=True)
ws3['A4'].alignment = Alignment(wrap_text=True, horizontal="left")

# ── Scenario inputs ──
# All units in $millions
# Shares: 578.34M current, dilution factors applied
# Cash: $904.7M
# Burn: ~$60-65M/qtr, ~$240-260M/yr

# WACC = 17.53%
# Cost of equity = 17.65%

scenarios_header = [
    "Scenario Item", "Bear", "Base", "Bull", "Notes / Rationale"
]

scenarios_data = [
    # Cash & burn assumptions
    ("--- Cash & Financing Assumptions ---", "", "", "", ""),
    ("Current Cash ($M)", "$904.7", "$904.7", "$904.7", "Q1 FY26 per Yahoo Finance Statistics"),
    ("Annual Burn Rate ($M/yr)", "$280", "$260", "$260", "TTM OCF -$241M + Capex -$40M = -$281M; annualized burn"),
    ("Cash Runway (years)", "3.2", "3.5", "3.5", "= Cash / Burn Rate"),
    ("", "", "", "", ""),
    ("Dilutive Raise #1: Amount ($M)", "$500", "$400", "$300", "Required to extend runway through commercialization"),
    ("Dilutive Raise #1: Timing (yr)", "Yr 1.5", "Yr 2", "Yr 2", "Funding runway to Q2 or H2 2028"),
    ("Dilutive Raise #1: Price ($/shr)", "$2.50", "$4.00", "$6.00", "Deep discount (bear), partial re-rate (base/bull)"),
    ("Dilutive Raise #1: New Shares (M)", "200", "100", "50", "= Amount / Price"),
    ("", "", "", "", ""),
    ("Dilutive Raise #2: Amount ($M)", "$600", "$500", "$400", "Second round for pilot → production bridge"),
    ("Dilutive Raise #2: Timing (yr)", "Yr 3.5", "Yr 4", "Yr 5", "Post-raised runway through 2030-2031"),
    ("Dilutive Raise #2: Price ($/shr)", "$1.50", "$3.00", "$5.00", "Discount compounds in bear"),
    ("Dilutive Raise #2: New Shares (M)", "400", "167", "80", "= Amount / Price"),
    ("", "", "", "", ""),
    ("Post-Dilution Shares Bear (M)", "1,178", "", "", "= 578 + 200 + 400; 2.04x dilution"),
    ("Post-Dilution Shares Base (M)", "", "845", "", "= 578 + 100 + 167; 1.46x dilution"),
    ("Post-Dilution Shares Bull (M)", "", "", "708", "= 578 + 50 + 80; 1.22x dilution"),
    ("", "", "", "", ""),
    ("--- Terminal Year (Year 10 / FY2036) ---", "", "", "", ""),
    ("Revenue ($M)", "$25", "$350", "$2,500", "Bear: consulting/license residuals; Base: VW pilot volumes; Bull: commercial deployment"),
    ("Operating Income ($M)", "-$300", "-$150", "+$200", "Bear: extended burn; Base: near break-even; Bull: marginally profitable"),
    ("Net Income ($M)", "-$290", "-$140", "+$160", "Net of tax; minimal tax in bear/base due to losses"),
    ("", "", "", "", ""),
    ("Cash Remaining ($M)", "$0", "$50", "$800", "Bear: cash exhausted; Base: modest residual; Bull: self-funding"),
    ("NAV Floor / Share ($)", "$0.00", "$0.06", "$1.13", "= Cash Remaining / Post-Dilution Shares"),
    ("", "", "", "", ""),
    ("", "", "", "", ""),
    ("--- Pipeline NPV & Pricing ---", "", "", "", ""),
    ("Pipeline NPV per Share ($)", "$0", "$3.00", "$15.00", "Bear: option expired; Base: modest VW licensing; Bull: multi-OEM deployment"),
    ("Implied NAV + NPV / Share ($)", "$0.00", "$3.06", "$16.13", "= NAV Floor + Pipeline NPV"),
    ("Discount for Execution Risk", "100%", "50%", "30%", "Bear: option expired; Base: partial discount; Bull: near-certainty"),
    ("Target Price / Share ($)", "$0.00", "$1.53", "$11.29", "= (NAV + NPV) × (1 - discount)"),
    ("", "", "", "", ""),
    ("Implied Upside/Downside (%)", "--%", "-73.9%", "+92.5%", "= Target / $5.86 - 1"),
    ("Weight", "25%", "45%", "30%", "Probability weights"),
    ("", "", "", "", ""),
    ("--- Weighted Fair Value ---", "", "", "", ""),
    ("Probability-Weighted FV / Share", "", "", "$3.29", "= 0.25 × $0 + 0.45 × $1.53 + 0.30 × $11.29"),
    ("Implied Downside from Current", "", "", "-43.9%", "= $3.29 / $5.86 - 1"),
    ("", "", "", "", ""),
    ("--- Reverse Valuation ---", "", "", "", ""),
    ("At $5.86, NAV floor is $1.56/share", "", "", "", "Cash/Share = $904.7M / 578.34M"),
    ("Optionality premium at current price", "", "", "$4.30 / share", "$5.86 - $1.56 = $4.30 premium for pipeline"),
    ("Implied success probability", "", "", "~29%", "$4.30 / $14.86 breakeven price for success"),
]

for i, row_vals in enumerate(scenarios_data):
    for j, val in enumerate(row_vals):
        font = bold_font if row_vals[0].startswith("---") else normal_font
        c(ws3, 5 + i, 1 + j, val, font)
        if j > 0:
            ws3.cell(row=5 + i, column=1 + j).alignment = Alignment(horizontal="center")
        if j == 4:
            ws3.cell(row=5 + i, column=1 + j).alignment = Alignment(wrap_text=True, horizontal="left")

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 50

# ═══════════════════════════════════════════════════
# Sheet 4: Actuals Source Audit
# ═══════════════════════════════════════════════════
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells('A1:E1')
ws4['A1'] = "Actuals Source Audit — QuantumScape (QS)"
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(horizontal="center")

audit_header = ["Data Point", "Value", "Source", "Date", "Notes"]
header_row(ws4, 3, audit_header)

audit_data = [
    # Price & market data
    ("--- Price & Market Data ---", "", "", "", ""),
    ("Stock Price (close)", "$5.86", "Yahoo Finance", "2026-07-17", "NASDAQ; close 4:00 PM EDT"),
    ("Market Cap", "$3.63B", "Yahoo Finance Statistics", "2026-07-17", "Current quarter valuation"),
    ("Enterprise Value", "$2.80B", "Yahoo Finance Statistics", "2026-07-17", "MC + Debt - Cash"),
    ("Shares Outstanding", "578.34M", "Yahoo Finance Statistics", "2026-07-17", "Current; implied 615.15M with converts"),
    ("Beta (5Y Monthly)", "2.62", "Yahoo Finance Statistics", "2026-07-17", "High volatility pre-commercial"),
    ("52W High", "$19.07", "Yahoo Finance", "Trailing", "-60% from peak"),
    ("52W Low", "$5.64", "Yahoo Finance", "Trailing", "Near current price; oversold technically"),
    ("Avg Vol 3M", "26.27M", "Yahoo Finance", "2026-07-17", "High tradability"),
    ("Float", "504.73M", "Yahoo Finance", "2026-07-17", ""),
    ("% Held Insiders", "13.56%", "Yahoo Finance", "2026-07-17", "VW ownership implied"),
    ("% Held Institutions", "41.97%", "Yahoo Finance", "2026-07-17", "Moderate institutional interest"),
    ("Short % of Float", "19.13%", "Yahoo Finance", "2026-06-30", "96.66M shares short; elevated"),
    ("", "", "", "", ""),
    # Balance sheet
    ("--- Balance Sheet (FY End) ---", "", "", "", ""),
    ("Total Assets (FY25)", "$1,308M", "Yahoo Finance Balance Sheet", "2025-12-31", "In thousands: 1,308,156"),
    ("Total Cash (MRQ Q1 FY26)", "$904.7M", "Yahoo Finance Statistics", "2026-03-31", "Per share $1.47/share"),
    ("Total Debt (MRQ)", "$69.18M", "Yahoo Finance Statistics", "2026-03-31", "Debt/Equity 6.24%"),
    ("Total Equity (FY25)", "$1,169M", "Yahoo Finance Balance Sheet", "2025-12-31", "Common stock equity"),
    ("Total Debt (FY25)", "$71.0M", "Yahoo Finance Balance Sheet", "2025-12-31", "Capital lease obligations"),
    ("Working Capital (FY25)", "$920.1M", "Yahoo Finance Balance Sheet", "2025-12-31", "Very high current ratio 20.93"),
    ("P/B Ratio", "3.27x", "Yahoo Finance Statistics", "2026-07-17", "Current quarter; was 5.21x FY25"),
    ("BVPS", "$1.80", "Yahoo Finance Statistics", "2026-03-31", "MRQ"),
    ("Shares Issued FY25", "607.63M", "Yahoo Finance Balance Sheet", "2025-12-31", "In thousands; 39% up from FY22"),
    ("", "", "", "", ""),
    # Income statement
    ("--- Income Statement (Annual, $M) ---", "", "", "", ""),
    ("Revenue TTM", "$0", "Yahoo Finance Income Statement", "TTM", "Zero revenue across all periods"),
    ("Revenue FY25", "$0", "Yahoo Finance Income Statement", "2025-12-31", ""),
    ("Revenue FY24", "$0", "Yahoo Finance Income Statement", "2024-12-31", ""),
    ("Operating Expense TTM", "$458.2M", "Yahoo Finance", "TTM", "="),
    ("Operating Expense FY25", "$472.6M", "Yahoo Finance", "2025-12-31", ""),
    ("Operating Expense FY24", "$525.2M", "Yahoo Finance", "2024-12-31", "Peak; then declining"),
    ("Operating Income TTM", "-$458.2M", "Calculated", "TTM", "="),
    ("Net Income TTM", "-$421.4M", "Yahoo Finance", "TTM", "After +$35.8M interest income"),
    ("Net Income FY25", "-$435.1M", "Yahoo Finance", "2025-12-31", ""),
    ("EPS Diluted TTM", "-$0.71", "Yahoo Finance", "TTM", ""),
    ("Interest Income TTM", "$37.8M", "Yahoo Finance", "TTM", "Cash yield; significant contributor to net loss"),
    ("Interest Expense TTM", "$2.0M", "Yahoo Finance", "TTM", "Nominal on small debt"),
    ("EBITDA TTM", "-$347.9M", "Yahoo Finance", "TTM", "Includes $70.2M depreciation add-back"),
    ("Dilution: Depreciation TTM", "$70.2M", "Yahoo Finance", "TTM", "Capex-related; pilot line construction"),
    ("", "", "", "", ""),
    # Cash flows
    ("--- Cash Flow (Annual, $M) ---", "", "", "", ""),
    ("Operating Cash Flow TTM", "-$241.2M", "Yahoo Finance", "TTM", "Annual burn ~$240M"),
    ("Operating Cash Flow FY25", "-$242.5M", "Yahoo Finance", "2025-12-31", "Stable burn"),
    ("Operating Cash Flow FY24", "-$274.6M", "Yahoo Finance", "2024-12-31", "Higher burn year"),
    ("Capex TTM", "-$40.4M", "Yahoo Finance", "TTM", "Pilot line; declining from peak"),
    ("Capex FY25", "-$36.3M", "Yahoo Finance", "2025-12-31", ""),
    ("Free Cash Flow TTM", "-$281.7M", "Yahoo Finance", "TTM", "OCF + Capex"),
    ("Free Cash Flow FY25", "-$278.8M", "Yahoo Finance", "2025-12-31", ""),
    ("Financing Cash Flow TTM", "+$301.1M", "Yahoo Finance", "TTM", "Equity raises; $267.6M stock issuance"),
    ("Financing Cash Flow FY25", "+$312.8M", "Yahoo Finance", "2025-12-31", "Large equity raise; $268.7M stock issuance"),
    ("Financing Cash Flow FY24", "+$144.0M", "Yahoo Finance", "2024-12-31", ""),
    ("End Cash Position TTM", "$158.7M", "Yahoo Finance", "TTM end", "Cash flow statement end; differs from Stats $904.7M"),
    ("", "", "", "", ""),
    # Analyst estimates
    ("--- Analyst Estimates (Yahoo Finance) ---", "", "", "", ""),
    ("Current Qtr Revenue Estimate", "--", "Yahoo Finance Analysis", "2026-07-17", "No revenue estimate (Q2 FY26)"),
    ("Next Qtr Revenue Estimate", "--", "Yahoo Finance Analysis", "2026-07-17", "No revenue estimate (Q3 FY26)"),
    ("FY2026 Revenue Estimate", "--", "Yahoo Finance Analysis", "2026-07-17", "No estimate; 3 analysts for FY27 only"),
    ("FY2027 Revenue Estimate (Avg)", "$35.71M", "Yahoo Finance Analysis", "2026-07-17", "3 analysts; range $15M-$57M"),
    ("FY2026 EPS Estimate (Avg)", "-$0.70", "Yahoo Finance Analysis", "2026-07-17", "6 analysts; range -$0.74 to -$0.66"),
    ("FY2027 EPS Estimate (Avg)", "-$0.68", "Yahoo Finance Analysis", "2026-07-17", "6 analysts; range -$0.79 to -$0.50"),
    ("EPS Revision Trend FY26", "Negative", "Yahoo Finance", "2026-07-17", "From -$0.63 (90d ago) to -$0.70 (current)"),
    ("EPS Revision Trend FY27", "Negative", "Yahoo Finance", "2026-07-17", "From -$0.57 (90d ago) to -$0.68 (current)"),
    ("", "", "", "", ""),
    # Key stats
    ("--- Key Ratios ---", "", "", "", ""),
    ("Current Ratio (MRQ)", "20.93", "Yahoo Finance Statistics", "2026-03-31", "Extremely high; cash-rich balance sheet"),
    ("Debt/Equity (MRQ)", "6.24%", "Yahoo Finance Statistics", "2026-03-31", "Very low leverage"),
    ("ROA (TTM)", "-23.02%", "Yahoo Finance Statistics", "TTM", "Negative due to operating losses"),
    ("ROE (TTM)", "-38.00%", "Yahoo Finance Statistics", "TTM", "Negative; well below cost of equity (17.65%)"),
    ("", "", "", "", ""),
    ("10Y Treasury Rate", "4.55%", "CNBC US10Y", "2026-07-17", "Used for WACC calculation"),
    ("Fiscal Year End", "12/31", "Yahoo Finance Statistics", "2026-07-17", "Calendar year"),
]

for i, row_vals in enumerate(audit_data):
    for j, val in enumerate(row_vals):
        font = bold_font if val.startswith("---") else normal_font
        c(ws4, 4 + i, 1 + j, val, font)
        if j == 4:
            ws4.cell(row=4 + i, column=5).alignment = Alignment(wrap_text=True)

ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 16
ws4.column_dimensions['C'].width = 22
ws4.column_dimensions['D'].width = 14
ws4.column_dimensions['E'].width = 45

# ═══════════════════════════════════════════════════
# Sheet 5: Questions
# ═══════════════════════════════════════════════════
ws5 = wb.create_sheet("Questions")
ws5.merge_cells('A1:C1')
ws5['A1'] = "Open Questions — QuantumScape (QS)"
ws5['A1'].font = title_font
ws5['A1'].alignment = Alignment(horizontal="center")

questions_header = ["#", "Question", "Why It Matters"]
header_row(ws5, 3, questions_header)

questions = [
    # Cash accounting
    (1, "Cash Position Discrepancy: Balance sheet shows End Cash of $158.7M (TTM cash flow statement) vs Total Cash of $904.7M (Yahoo Statistics MRQ). Why the $746M delta?",
     "Determines accurate cash runway. Yahoo Stats likely includes STT/investments; cash flow statement may track only operating cash. Need to verify actual liquid cash vs restricted/STT investments."),
    (2, "Share Count Discrepancy: Balance sheet shows 607.63M shares issued (FY25), but Yahoo Finance shows 578.34M outstanding with 615.15M implied. Which is correct?",
     "Implied vs actual share count matters for per-share NAV. The 615.15M implied includes convertible subsidiary equity. Need to clarify the economic share count for valuation."),
    
    # VW partnership
    (3, "VW Partnership Status: What is the current status of the Volkswagen joint venture and pilot program? Is the 'all-solid-state battery' development on schedule for ~2028?",
     "VW is QS's primary strategic partner and validation pathway. Any delay, cancellation, or deprioritization would materially impact the optionality premium. VW reportedly owns ~26% stake."),
    (4, "VW Ownership Stake: VW holds ~26% of QS. Is this equity stake dilutive (preferred/share class) or common? What voting rights do they have?",
     "Capital structure complexity. VW's stake affects effective float, governance, and exit dynamics."),
    
    # Cash runway
    (5, "Cash Runway Calculation: With $904.7M cash and ~$60-65M/quarter burn, runway is ~14-15 quarters to June 2030. What are the milestone triggers for each financing round?",
     "Determines dilution timeline. The critical question is whether VW or other partners co-fund milestones or if QS must raise independently at depressed valuations."),
    
    # Technology validation
    (6, "Technology Validation: Has QS demonstrated a full-cell that meets the claimed 3000+ mile range and 15-minute charge specs? What is the production yield at pilot scale?",
     "The entire thesis rests on the solid-state cell surviving pilot → commercialization transition. Many battery startups have failed at this bridge. Yield rates, cycle life, and thermal stability at scale are unproven."),
    (7, "Pilot Line Status: What is the status and timeline of the pilot production line in San Jose? Has any cell been delivered to VW for testing?",
     "Pilot line is the de facto proof point for the technology. No physical delivery = all speculative."),
    
    # Competitive landscape
    (8, "Competitive Risk: How does QS's all-solid-state lithium-metal technology compare with Semi-solid approaches (CATL, Nio, Gotion) that are already in volume production?",
     "Semi-solid batteries are shipping NOW in volume. QS's all-solid-state approach has higher theoretical energy density but longer commercialization timeline. Window of differentiation may be closing."),
    (9, "Toyota's Solid-State Timeline: Toyota has announced 2027-2028 commercialization for solid-state EV batteries. How does this impact QS's first-mover advantage?",
     "Toyota's resources dwarf QS's. If Toyota achieves comparable performance at lower cost, QS's optionality vanishes. Toyota's track record of delayed deliveries is a counter-argument but not a hedge."),
    
    # Financial structure
    (10, "Depreciation Pattern: TTM depreciation of $70.2M is up from $65.8M FY24 and $49.7M FY23. What assets are being depreciated and is the pilot line construction complete?",
         "Increasing depreciation signals capitalization of pilot/facility costs. If the line is substantially built, future depreciation may increase, widening the income statement gap."),
    (11, "Interest Income Sustainability: $37.8M TTM interest income on $904.7M cash implies ~4.2% yield. Does this change with future cash deployment or lower rates?",
        "Interest income offsets ~9% of operating burn. A significant component of net loss mitigation. Rate environment matters."),
    
    # Revenue recognition
    (12, "Revenue Recognition Policy: With $0 revenue for 4+ years, what revenue recognition triggers exist for the VW partnership? Are there milestone payments, licensing fees, or government grants not yet recognized?",
        "Any non-operating revenue source (grants, partnerships, R&D subsidies) would extend runway without dilution. Need to verify if existing contracts contain any revenue-recognition provisions."),
    (13, "Dilution History: Share count grew from 438M (FY22) to 608M (FY25) — +39% in 3 years. What was the implied valuation at each capital raise? Was the VW partnership the catalyst for the last raise?",
        "Tracks the capital raise history. VW invested $300M+ across multiple rounds. The last raise (FY25) valued the company at ~$5.4B pre — well above current MC. Understanding raise pricing history sets the dilution risk profile."),
    
    # Capital allocation
    (14, "SBC and Option Grants: What is the annual SBC / option grant expense? For a pre-revenue company with 578M+ shares, dilution from SBC can compound silently.",
        "SBC is a silent dilution mechanism. For a cash-burning company already raising equity at discounts, option-based compensation can erode per-share value without triggering new financing rounds."),
    (15, "SPAC-era Dilution: QS went public via SPAC merger (June 2020). What is the post-SPAC dilution trajectory? How much value has been destroyed from IPO to current price?",
        "SPAC IPO was at ~$10/share. Current price ~$5.86. The entire SPAC premium has been destroyed. Understanding the full dilution arc is essential for perspective."),
]

for i, (num, q, why) in enumerate(questions):
    c(ws5, 4 + i, 1, num, normal_font)
    c(ws5, 4 + i, 2, q, normal_font)
    ws5.cell(row=4 + i, column=2).alignment = Alignment(wrap_text=True)
    c(ws5, 4 + i, 3, why, normal_font)
    ws5.cell(row=4 + i, column=3).alignment = Alignment(wrap_text=True)

ws5.column_dimensions['A'].width = 5
ws5.column_dimensions['B'].width = 55
ws5.column_dimensions['C'].width = 45

# ═══════════════════════════════════════════════════
# Sheet 6: Sources
# ═══════════════════════════════════════════════════
ws6 = wb.create_sheet("Sources")
ws6.merge_cells('A1:C1')
ws6['A1'] = "Sources — QuantumScape (QS)"
ws6['A1'].font = title_font
ws6['A1'].alignment = Alignment(horizontal="center")

sources_header = ["#", "Source", "URL / Reference"]
header_row(ws6, 3, sources_header)

sources_data = [
    (1, "Yahoo Finance — Price, Statistics, Market Data", "https://finance.yahoo.com/quote/QS/"),
    (2, "Yahoo Finance — Key Statistics (Beta, P/B, Shares, Short Interest)", "https://finance.yahoo.com/quote/QS/key-statistics/"),
    (3, "Yahoo Finance — Income Statement (Annual)", "https://finance.yahoo.com/quote/QS/financials/"),
    (4, "Yahoo Finance — Balance Sheet (Annual)", "https://finance.yahoo.com/quote/QS/balance-sheet/"),
    (5, "Yahoo Finance — Cash Flow Statement (Annual)", "https://finance.yahoo.com/quote/QS/cash-flow/"),
    (6, "Yahoo Finance — Analyst Estimates", "https://finance.yahoo.com/quote/QS/analysis/"),
    (7, "Yahoo Finance — Company Profile", "https://finance.yahoo.com/quote/QS/profile/"),
    (8, "Yahoo Finance — Revenue/EPS Estimates Tables", "https://finance.yahoo.com/quote/QS/analysis/"),
    (9, "CNBC — 10Y US Treasury Yield", "https://www.cnbc.com/quotes/US10Y"),
    (10, "StockAnalysis.com (returned 404 for QS)", "N/A — data unavailable"),
    (11, "Related Tickers (Yahoo Finance)", "SLDP (Solid Power), MBLY (Mobileye), MVST (Microvast), SES (SES AI)"),
]

for i, (num, src, url) in enumerate(sources_data):
    c(ws6, 4 + i, 1, num, normal_font)
    c(ws6, 4 + i, 2, src, normal_font)
    c(ws6, 4 + i, 3, url, normal_font)
    ws6.cell(row=4 + i, column=3).alignment = Alignment(wrap_text=True)

ws6.column_dimensions['A'].width = 5
ws6.column_dimensions['B'].width = 55
ws6.column_dimensions['C'].width = 60

# ── Save workbook ──
output_path = "/home/refcell/dev/capital/models/[2026-07-17] QuantumScape Model.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"WACC: 17.53%")
print(f"Probability-weighted FV / Share: $3.29")
print(f"Implied downside from $5.86: -43.9%")
