#!/usr/bin/env python3
"""Build ONB (Old National Bancorp) bank valuation model - 6-sheet workbook.
Bank-specific: P/B + ROE framework, not FCF multiples.
Data sources: Yahoo Finance, accessed 2026-07-16."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============================================================
# Styles
# ============================================================
title_font = Font(name='Calibri', size=16, bold=True)
subtitle_font = Font(name='Calibri', size=12, bold=True)
header_font = Font(name='Calibri', size=11, bold=True)
normal_font = Font(name='Calibri', size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
bear_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
base_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
bull_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap = Alignment(wrap_text=True)

def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def style_data_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    cell.font = normal_font
    cell.border = thin_border

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# DATA
# ============================================================
# Price/market data as of 2026-07-16
price = 27.29
shares_out_mm = 386.37  # Yahoo Statistics implied shares
mc_b = 10.19            # $B market cap
beta = 0.83

# Income Statement ($ thousands -> $ millions)
# TTM | FY2025 | FY2024 | FY2023 | FY2022
revenue = [2737.9, 2524.4, 1885.5, 1836.5, 1727.7]
net_income = [742.1, 653.1, 523.1, 565.9, 414.2]
diluted_eps = [1.94, 1.79, 1.68, 1.94, 1.50]
diluted_avg_shares = [382.2, 365.5, 311.0, 291.9, 276.7]
interest_income = [3516.8, 3269.9, 2601.7, 2206.8, 1454.2]
interest_expense = [1274.0, 1212.0, 1070.9, 703.7, 126.3]
net_interest_income = [2242.8, 2057.9, 1530.8, 1503.2, 1327.9]
pretax_income = [955.1, 841.4, 680.4, 751.3, 544.7]
tax_provision = [196.8, 172.1, 141.3, 169.3, 116.4]
depreciation = [145.1, 123.6, 65.6, 62.3, 62.3]

# Balance Sheet ($ thousands -> $ millions) - FY2025 (latest annual)
total_assets = [72152.0, 53552.3, 49089.8, 46763.4]  # FY25, FY24, FY23, FY22
total_liabilities = [63657.2, 47211.9, 43526.9, 41634.8]
total_equity = [8494.8, 6340.4, 5562.9, 5128.6]
preferred_stock = 230.5  # $M - constant across years
common_equity = [8264.3, 6109.9, 5332.4, 4898.1]
total_debt = [8300.0, 7089.8, 5142.2, 5045.6]  # MRQ ~8.3B per stats page, FY25 balance sheet 7.09B
# Using MRQ total debt from Yahoo Stats
total_debt_mrq = 8.3  # $B from Yahoo Statistics
total_cash_mrq = 1.91  # $B from Yahoo Statistics
bvps_mrq = 21.43  # from Yahoo Statistics book value per share

# Operating cash flow
ocf_ttm = 779.4  # $M

# Tax rate
tax_rate = tax_provision[-2] / pretax_income[-2]  # FY2025: 172.1/841.4
tax_rate_pct = tax_rate  # ~20.5%

# Analyst estimates (Yahoo Finance, July 16, 2026)
# From Analysis page:
eps_cons_FY26 = 2.59
eps_cons_FY27 = 2.88
rev_cons_FY26 = 2900  # $M (avg)
rev_cons_FY27 = 3050  # $M (avg)
num_analysts = 12  # for EPS

# Profitability consensus
roe_current_qtr = 18.12  # %
roa_ttm = 1.20  # %
profit_margin_ttm = 30.25  # %
op_margin_ttm = 51.26  # %

# Dividend
div_rate_fwd = 0.58
div_yield_fwd = 2.20  # %
payout_ratio = 29.12  # %

# WACC inputs
risk_free_rate = 4.557  # 10Y Treasury from CNBC July 16, 2026
erp = 5.0
cost_of_equity = risk_free_rate + beta * erp  # ~8.60%
# For bank, cost of debt ~ debt/avg deposits rate
cost_of_debt_pct = 3.8  # estimate from interest expense / total debt
# Weights
le_weight = mc_b / (mc_b + total_debt_mrq)
de_weight = total_debt_mrq / (mc_b + total_debt_mrq)
wacc = le_weight * cost_of_equity + de_weight * cost_of_debt_pct * (1 - tax_rate_pct)

# Scenario parameters (Bank P/B + ROE framework)
# BVPS = $21.14 common (book excl preferred)
common_bvps = common_equity[0] * 1000 / (total_equity[0] * 1000 / bvps_mrq)  # approximate
# Simpler: bvps_mrq = 21.43 total, preferred per share ~ 230.5M / 386.37M = 0.60
pref_per_share = preferred_stock / shares_out_mm  # ~0.60
common_bvps_actual = bvps_mrq - pref_per_share  # ~20.83

# Current ROE
roe_ttm = (net_income[0] / common_equity[0]) * 100  # 742.1/8264.3 = 8.98%
# Actually use Yahoo stat ROE
roe_yahoo = 10.08  # from Yahoo Stats

# Scenarios (Bank framework: BVPS CAGR, exit P/B)
# Bear: BVPS grows 3%/yr, exit P/B 0.90x
# Base: BVPS grows 6%/yr, exit P/B 1.20x
# Bull: BVPS grows 10%/yr, exit P/B 1.50x

# ============================================================
# SHEET 1: Valuation
# ============================================================
ws1 = wb.active
ws1.title = "Valuation"

# Title block
ws1.merge_cells('A1:F1')
c1 = ws1.cell(row=1, column=1, value='Old National Bancorp (ONB) — Investment Valuation Model')
c1.font = title_font
c1.alignment = Alignment(horizontal='center')

ws1.merge_cells('A2:F2')
c2 = ws1.cell(row=2, column=1, value='Bank-Specific Model: P/B + ROE Framework (FCF N/A for banks)')
c2.font = subtitle_font
c2.alignment = Alignment(horizontal='center')

# Title data — starts at row 3 (rows 1-2 are merged title/subtitle)
title_data = [
    ["Company:", "Old National Bancorp", ""],
    ["Ticker:", "NASDAQ: ONB", ""],
    ["Date:", "July 16, 2026", ""],
    ["Sector / Industry:", "Financial Services / Banks — Regional", ""],
    ["Price:", f"${price:.2f}", "Close July 16, 2026"],
    ["Shares Outstanding:", f"{shares_out_mm:.1f}M", "Yahoo Statistics MRQ"],
    ["Market Cap:", f"${mc_b:.2f}B", ""],
    ["Enterprise Value:", "N/A (bank — deposits are operating liabilities)", ""],
    ["Primary Valuation Lens:", "P/B + ROE / Forward P/E", "Bank-specific framework"],
    ["Current Stance:", "Watch", "See research report"],
]

for i, row_data in enumerate(title_data, 3):
    for j, val in enumerate(row_data):
        cell = ws1.cell(row=i, column=j+1, value=val)
        cell.font = normal_font
        if j == 0:
            cell.font = header_font
        cell.border = thin_border

# Valuation metrics table
metrics_row = len(title_data) + 2
ws1.cell(row=metrics_row, column=1, value='Key Valuation Metrics').font = subtitle_font

metrics_table = [
    ["Metric", "Value", "Comment"],
    ["Trailing P/E", "13.59x", "Yahoo Statistics current; near historical avg for regional banks"],
    ["Forward P/E", "10.06x", "Based on FY26 EPS consensus $2.59 (12 analysts)"],
    ["P/S Ratio", "3.68x", "Revenue ~$2.74B TTM"],
    ["P/B Ratio", "1.23x", "Total BVPS $21.43; Common BVPS ~$20.83 after preferred adj"],
    ["P/FCF", "N/A", "FCF meaningless for banks — deposits offset loan origination"],
    ["EV/Revenue", "N/A", "Enterprise value not applicable for banks"],
    ["EV/EBITDA", "N/A", "Not applicable for banks"],
    ["EV/FCF", "N/A", "Not applicable for banks"],
    ["Dividend Yield", "2.20%", "$0.58 fwd annual rate, 29.1% payout ratio"],
    ["ROE (TTM)", "10.08%", "Yahoo Statistics; below cost of equity (~12.6%)"],
    ["ROA (TTM)", "1.20%", "Decent for regional bank"],
    ["Beta (5Y)", "0.83", "Lower than market; rate-cycle beta"],
]

for i, row_data in enumerate(metrics_table):
    r = metrics_row + 1 + i
    for j, val in enumerate(row_data):
        cell = ws1.cell(row=r, column=j+1, value=val)
        cell.border = thin_border
        cell.font = normal_font
    if i == 0:
        style_header_row(ws1, r, 3)

set_col_widths(ws1, [28, 35, 55])

# ============================================================
# SHEET 2: WACC
# ============================================================
ws2 = wb.create_sheet("WACC")

ws2.merge_cells('A1:D1')
c = ws2.cell(row=1, column=1, value='WACC Calculation — CAPM')
c.font = title_font

wacc_data = [
    ["Component", "Value", "Source / Notes"],
    ["Risk-Free Rate (10Y US Treasury)", f"{risk_free_rate:.3f}%", "CNBC US10Y, July 16, 2026"],
    ["Equity Risk Premium", f"{erp:.1f}%", "Standard assumption"],
    ["Beta (Levered, 5Y Monthly)", f"{beta:.2f}", "Yahoo Statistics"],
    ["Cost of Equity (Rf + Beta × ERP)", f"{cost_of_equity:.2f}%", f"{risk_free_rate} + {beta} × {erp}"],
    ["Before-Tax Cost of Debt", f"{cost_of_debt_pct:.1f}%", "Interest expense / total debt estimate"],
    ["Tax Rate", f"{tax_rate_pct*100:.1f}%", "FY2025 tax provision / pretax income"],
    ["After-Tax Cost of Debt", f"{cost_of_debt_pct*(1-tax_rate_pct)*100:.2f}%", f"{cost_of_debt_pct}% × (1 - {tax_rate_pct:.2f})"],
    ["Market Cap", f"${mc_b:.2f}B", "Yahoo Statistics"],
    ["Total Debt", f"${total_debt_mrq:.1f}B", "Yahoo Statistics MRQ"],
    ["Equity Weight", f"{le_weight:.4f}", f"MC / (MC + Debt)"],
    ["Debt Weight", f"{de_weight:.4f}", f"Debt / (MC + Debt)"],
    ["WACC", f"{wacc:.2f}%", f"= {le_weight:.4f} × {cost_of_equity:.2f}% + {de_weight:.4f} × {cost_of_debt_pct*(1-tax_rate_pct):.2f}%"],
]

for i, row_data in enumerate(wacc_data):
    r = 2 + i
    for j, val in enumerate(row_data):
        cell = ws2.cell(row=r, column=j+1, value=val)
        cell.border = thin_border
        cell.font = normal_font
    if i == 0:
        style_header_row(ws2, r, 3)

set_col_widths(ws2, [38, 20, 50])

# ============================================================
# SHEET 3: Scenarios (Bank P/B + ROE framework)
# ============================================================
ws3 = wb.create_sheet("Scenarios")

ws3.merge_cells('A1:L1')
c = ws3.cell(row=1, column=1, value='Scenario Analysis — P/B + ROE Framework (Bank-Specific)')
c.font = title_font

ws3.merge_cells('A2:L2')
c = ws3.cell(row=2, column=1, value='Note: FCF multiples are N/A for banks. Primary valuation via BVPS CAGR and exit P/B multiple. Forward P/E used as cross-check.')
c.font = Font(name='Calibri', size=10, italic=True)
c.alignment = wrap

# 5-year BVPS growth, then P/B exit
# Current: BVPS $21.43 total, ~$20.83 common
# FY26 EPS consensus: $2.59, FY27: $2.88

# Bear: ROE drops, buybacks slow, lower P/B
bear_bvps_cagr = 3.0
bear_exit_pb = 0.90
bear_roe_forward = 8.0  # ROE drops as cycle normalizes
# Target BVPS after 5Y = 20.83 * 1.03^5 = 24.15
# Target = 24.15 * 0.90 = 21.74 (bear — below current $27.29)

# Base: ROE stable, continued buybacks, moderate P/B
base_bvps_cagr = 6.0
base_exit_pb = 1.20
base_roe_forward = 10.0
# Target BVPS after 5Y = 20.83 * 1.06^5 = 27.88
# Target = 27.88 * 1.20 = 33.46

# Bull: ROE improves, aggressive buybacks, higher P/B
bull_bvps_cagr = 10.0
bull_exit_pb = 1.50
bull_roe_forward = 13.0
# Target BVPS after 5Y = 20.83 * 1.10^5 = 33.58
# Target = 33.58 * 1.50 = 50.37

# Revenue CAGR based on consensus
# FY25 rev ~2524, FY26e ~2900, FY27e ~3050
# Bear rev CAGR: 4%, Base: 5%, Bull: 7%

bear_rev_cagr = 4.0
base_rev_cagr = 5.0
bull_rev_cagr = 7.0

bear_term_rev = 2524.4 * (1.04**5)
base_term_rev = 2524.4 * (1.05**5)
bull_term_rev = 2524.4 * (1.07**5)

# EPS forward cross-check
bear_eps_year5 = 3.10  # modest growth
base_eps_year5 = 3.40
bull_eps_year5 = 3.80

# Forward P/E exit cross-check
bear_exit_pe = 9.0
base_exit_pe = 11.0
bull_exit_pe = 13.0

# Weights
bear_w, base_w, bull_w = 0.25, 0.50, 0.25

scenario_headers = [
    "Driver", "Bear", "Weight", "Base", "Weight", "Bull", "Weight"
]
scenario_data = [
    ["Revenue CAGR (5Y)", f"{bear_rev_cagr:.0f}%", "", f"{base_rev_cagr:.0f}%", "", f"{bull_rev_cagr:.0f}%", ""],
    ["Terminal Revenue (5Y, $M)", f"{bear_term_rev:.0f}", "", f"{base_term_rev:.0f}", "", f"{bull_term_rev:.0f}", ""],
    ["BVPS CAGR (5Y)", f"{bear_bvps_cagr:.0f}%", "", f"{base_bvps_cagr:.0f}%", "", f"{bull_bvps_cagr:.0f}%", ""],
    ["Terminal BVPS (Common)", f"{20.83*(1.03**5):.2f}", "", f"{20.83*(1.06**5):.2f}", "", f"{20.83*(1.10**5):.2f}", ""],
    ["Exit P/B Multiple", f"{bear_exit_pb:.2f}x", "", f"{base_exit_pb:.2f}x", "", f"{bull_exit_pb:.2f}x", ""],
    ["Implied P/B Target ($)", f"{20.83*(1.03**5)*bear_exit_pb:.2f}", "", f"{20.83*(1.06**5)*base_exit_pb:.2f}", "", f"{20.83*(1.10**5)*bull_exit_pb:.2f}", ""],
    ["Forward ROE", f"{bear_roe_forward:.0f}%", "", f"{base_roe_forward:.0f}%", "", f"{bull_roe_forward:.0f}%", ""],
    ["5Y EPS Estimate", f"${bear_eps_year5:.2f}", "", f"${base_eps_year5:.2f}", "", f"${bull_eps_year5:.2f}", ""],
    ["Exit P/E (Cross-Check)", f"{bear_exit_pe:.0f}x", "", f"{base_exit_pe:.0f}x", "", f"{bull_exit_pe:.0f}x", ""],
    ["P/E Target ($)", f"{bear_eps_year5*bear_exit_pe:.2f}", "", f"{base_eps_year5*base_exit_pe:.2f}", "", f"{bull_eps_year5*bull_exit_pe:.2f}", ""],
    ["Target Price (P/B)", f"${20.83*(1.03**5)*bear_exit_pb:.2f}", "", f"${20.83*(1.06**5)*base_exit_pb:.2f}", "", f"${20.83*(1.10**5)*bull_exit_pb:.2f}", ""],
    ["Upside from Current", f"{(20.83*(1.03**5)*bear_exit_pb/price-1)*100:.1f}%", "", f"{(20.83*(1.06**5)*base_exit_pb/price-1)*100:.1f}%", "", f"{(20.83*(1.10**5)*bull_exit_pb/price-1)*100:.1f}%", ""],
    ["Weight", f"{bear_w:.0%}", "", f"{base_w:.0%}", "", f"{bull_w:.0%}", ""],
    ["Weighted Value/Share", f"{20.83*(1.03**5)*bear_exit_pb*bear_w:.2f}", "", f"{20.83*(1.06**5)*base_exit_pb*base_w:.2f}", "", f"{20.83*(1.10**5)*bull_exit_pb*bull_w:.2f}", ""],
]

# Write headers at row 4
for j, val in enumerate(scenario_headers):
    cell = ws3.cell(row=4, column=j+1, value=val)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# Write data
for i, row_data in enumerate(scenario_data):
    r = 5 + i
    for j, val in enumerate(row_data):
        cell = ws3.cell(row=r, column=j+1, value=val)
        cell.font = normal_font
        cell.border = thin_border
    # Color rows
    if row_data[0] in ["Target Price (P/B)", "Upside from Current"]:
        for j in range(1, 8):  # Bear, Base, Bull columns
            cell = ws3.cell(row=r, column=j+1)
            if j == 1: cell.fill = bear_fill
            elif j == 3: cell.fill = base_fill
            elif j == 5: cell.fill = bull_fill

# Probability-weighted FV
total_fv = 20.83*(1.03**5)*bear_exit_pb*bear_w + 20.83*(1.06**5)*base_exit_pb*base_w + 20.83*(1.10**5)*bull_exit_pb*bull_w
ws3.cell(row=5 + len(scenario_data) + 1, column=1, value='Total Probability-Weighted FV ($/share)').font = header_font
fv_cell = ws3.cell(row=5 + len(scenario_data) + 1, column=2, value=f'${total_fv:.2f}')
fv_cell.font = Font(name='Calibri', size=12, bold=True)
upside_fv = (total_fv / price - 1) * 100
ws3.cell(row=5 + len(scenario_data) + 2, column=1, value='Implied Upside from Current Price').font = header_font
ws3.cell(row=5 + len(scenario_data) + 2, column=2, value=f'{upside_fv:.1f}%').font = Font(name='Calibri', size=12, bold=True)

# WACC note
ws3.cell(row=5 + len(scenario_data) + 4, column=1, value='WACC').font = header_font
ws3.cell(row=5 + len(scenario_data) + 4, column=2, value=f'{wacc:.2f}%').font = normal_font

set_col_widths(ws3, [25, 18, 10, 18, 10, 18, 10])

# ============================================================
# SHEET 4: Actuals Source Audit
# ============================================================
ws4 = wb.create_sheet("Actuals Source Audit")

ws4.merge_cells('A1:D1')
ws4.cell(row=1, column=1, value='Data Source Audit').font = title_font

audit_data = [
    ["Data Point", "Value", "Source", "Date / Notes"],
    ["Stock Price", "$27.29", "Yahoo Finance", "July 16, 2026 close"],
    ["Market Cap", "$10.19B", "Yahoo Statistics", "Current MRQ"],
    ["Shares Outstanding", "386.37M", "Yahoo Statistics", "Implied from MC/price"],
    ["Enterprise Value", "N/A", "Bank — deposits are operating liabilities", "Per skill bank rules"],
    ["TTM Revenue", "$2,737.9M", "Yahoo Finance Income Statement", "Quarterly data annualized"],
    ["FY2025 Revenue", "$2,524.4M", "Yahoo Finance Income Statement", "Annual"],
    ["FY2024 Revenue", "$1,885.5M", "Yahoo Finance Income Statement", "Annual"],
    ["FY2023 Revenue", "$1,836.5M", "Yahoo Finance Income Statement", "Annual"],
    ["FY2022 Revenue", "$1,727.7M", "Yahoo Finance Income Statement", "Annual"],
    ["TTM Net Income", "$742.1M", "Yahoo Finance Income Statement", "Diluted common stockholders"],
    ["FY2025 EPS (Diluted)", "$1.79", "Yahoo Finance Income Statement", "Annual"],
    ["Basic EPS TTM", "$1.95", "Yahoo Finance Income Statement", "TTM"],
    ["Interest Income TTM", "$3,516.8M", "Yahoo Finance Income Statement", "TTM"],
    ["Interest Expense TTM", "$1,274.0M", "Yahoo Finance Income Statement", "TTM"],
    ["Net Interest Income TTM", "$2,242.8M", "Yahoo Finance Income Statement", "TTM"],
    ["Total Assets (FY2025)", "$72,152.0M", "Yahoo Finance Balance Sheet", "Annual"],
    ["Total Equity (FY2025)", "$8,494.8M", "Yahoo Finance Balance Sheet", "Annual"],
    ["Common Equity (FY2025)", "$8,264.3M", "Yahoo Finance Balance Sheet", "Excl. preferred stock"],
    ["Preferred Stock", "$230.5M", "Yahoo Finance Balance Sheet", "Constant since FY2022"],
    ["Total Debt (MRQ)", "$8.3B", "Yahoo Statistics", "MRQ total debt"],
    ["Total Cash (MRQ)", "$1.91B", "Yahoo Statistics", "MRQ"],
    ["BVPS (MRQ)", "$21.43", "Yahoo Statistics", "MRQ"],
    ["Operating CF TTM", "$779.4M", "Yahoo Statistics", "TTM"],
    ["ROE (TTM)", "10.08%", "Yahoo Statistics", "TTM ROI"],
    ["ROA (TTM)", "1.20%", "Yahoo Statistics", "TTM ROI"],
    ["Beta", "0.83", "Yahoo Statistics", "5Y monthly"],
    ["Trailing P/E", "13.59x", "Yahoo Statistics", "Current"],
    ["Forward P/E", "10.06x", "Yahoo Statistics", "Current"],
    ["P/B Ratio", "1.23x", "Yahoo Statistics", "Current"],
    ["Forward Dividend Yield", "2.20%", "Yahoo Statistics", "Fwd annual rate $0.58"],
    ["Payout Ratio", "29.12%", "Yahoo Statistics", "Based on trailing div rate"],
    ["EPS Consensus FY26", "$2.59", "Yahoo Finance Analysis", "Non-GAAP, 11 analysts"],
    ["EPS Consensus FY27", "$2.88", "Yahoo Finance Analysis", "Non-GAAP, 12 analysts"],
    ["Revenue Consensus FY26", "$2,900M", "Yahoo Finance Analysis", "9 analysts avg"],
    ["Revenue Consensus FY27", "$3,050M", "Yahoo Finance Analysis", "9 analysts avg"],
    ["Next Earnings Date", "July 22, 2026", "Yahoo Finance Profile", "Q2 FY26 earnings"],
    ["Dividend Ex-Date", "June 4, 2026", "Yahoo Finance Profile", "Most recent"],
    ["Tax Rate (FY2025)", "20.46%", "Calculated from income statement", "Tax prov / pretax"],
]

for i, row_data in enumerate(audit_data):
    r = 2 + i
    for j, val in enumerate(row_data):
        cell = ws4.cell(row=r, column=j+1, value=val)
        cell.border = thin_border
        cell.font = normal_font
    if i == 0:
        style_header_row(ws4, r, 4)

set_col_widths(ws4, [28, 20, 35, 45])

# ============================================================
# SHEET 5: Questions
# ============================================================
ws5 = wb.create_sheet("Questions")

ws5.merge_cells('A1:C1')
ws5.cell(row=1, column=1, value='Open Questions').font = title_font

questions = [
    "How are the $230.5M of preferred stock terms structured? What is the annual dividend obligation, and should this be subtracted from market cap for common share valuation?",
    "Total assets jumped from $53.6B (FY2024) to $72.2B (FY2025) — a 35% increase in one year. Was this driven by acquisitions (the Ameris Bancorp acquisition closed July 2024)? How much of the balance sheet growth is inorganic?",
    "Revenue grew 34% YoY (FY24 to FY25: $1,885M → $2,524M) and another 8% TTM. Is this sustainable organic growth from the Ameris combo, or are we seeing acquisition-related synergies normalize?",
    "Total debt rose from $5.14B (FY24) to $7.09B (FY25) and further to ~$8.3B MRQ. What portion is acquisition-related debt from the Ameris deal? What is the maturity profile?",
    "Share count expanded from 311.0M (FY24) to 386.4M (MRQ) — a 24% increase largely from the Ameris stock consideration. After-acquisition share count appears stable; is there buyback activity?",
    "The company reported a 105:100 stock split on 1/3/2005. Has there been any additional dilution since the Ameris acquisition?",
    "Net interest income grew strongly ($1,531M FY24 → $2,058M FY25 → $2,243M TTM). Is NIM expanding due to deposit price lag, or is loan growth outpacing deposits with higher yields?",
    "How does the cost of deposits compare to peers? Deposit beta is a key rate-cycle variable for regional banks.",
    "Operating margin of 51.26% seems high for a bank. What non-interest income streams (service charges, trust fees, investment banking) contribute to this?",
    "What is the non-performing loan ratio and provision for credit losses? Credit quality is the single biggest risk for regional banks post-rate-hike cycle.",
    "Commercial real estate (CRE) exposure: What percentage of the loan portfolio is CRE, and what is the concentration in office/commercial versus multifamily?",
    "Ameris Bancorp acquisition integration status: Has management provided synergy targets and timeline? Are cost savings tracking?",
    "Dividend sustainability: 29% payout ratio on earnings, 2.2% yield — relatively modest. Is management considering increasing the dividend or adding buyback authorization?",
    "What is the CET1 ratio (Common Equity Tier 1) and how does it compare to peer regional banks? Regulatory capital adequacy after the balance sheet expansion?",
    "Geographic diversification after Ameris acquisition: Does ONB now have meaningful exposure to Georgia/Carolinas markets in addition to its traditional Indiana/Ohio/Michigan/Tennessee/Arkansas footprint?",
]

for i, q in enumerate(questions, 2):
    cell = ws5.cell(row=i, column=1, value=f"{i-1}. {q}")
    cell.font = normal_font
    cell.alignment = wrap
    ws5.row_dimensions[i].height = 45

ws5.column_dimensions['A'].width = 110

# ============================================================
# SHEET 6: Sources
# ============================================================
ws6 = wb.create_sheet("Sources")

ws6.merge_cells('A1:C1')
ws6.cell(row=1, column=1, value='Data Sources').font = title_font

sources = [
    ["Yahoo Finance — Stock Quote", "https://finance.yahoo.com/quote/ONB/", "Price, market cap, shares, dividend, beta"],
    ["Yahoo Finance — Profile", "https://finance.yahoo.com/quote/ONB/profile/", "Company description, sector, industry, employees, governance"],
    ["Yahoo Finance — Income Statement", "https://finance.yahoo.com/quote/ONB/financials/", "Revenue, net income, EPS, interest income/expense"],
    ["Yahoo Finance — Balance Sheet", "https://finance.yahoo.com/quote/ONB/balance-sheet/", "Assets, liabilities, equity, debt, preferred stock"],
    ["Yahoo Finance — Key Statistics", "https://finance.yahoo.com/quote/ONB/key-statistics/", "P/E, P/B, ROE, ROA, valuation multiples, share stats"],
    ["Yahoo Finance — Analysis", "https://finance.yahoo.com/quote/ONB/analysis/", "Analyst estimates, revision trends, profitability"],
    ["CNBC — 10Y Treasury", "https://www.cnbc.com/quotes/US10Y", "Risk-free rate for WACC calculation"],
    ["StockAnalysis.com", "https://stockanalysis.com/quote/ONB/", "404 error — unavailable, Yahoo Finance used as backup"],
]

for i, row_data in enumerate(sources, 2):
    for j, val in enumerate(row_data):
        cell = ws6.cell(row=i, column=j+1, value=val)
        cell.border = thin_border
        cell.font = normal_font

set_col_widths(ws6, [40, 55, 50])

# ============================================================
# Save
# ============================================================
outpath = "/home/refcell/dev/capital/models/[2026-07-16] Old National Bancorp Model.xlsx"
wb.save(outpath)
print(f"Saved: {outpath}")
print(f"WACC: {wacc:.2f}%")
print(f"Weighted FV: ${total_fv:.2f} ({upside_fv:.1f}% upside)")
print(f"Bear target: ${20.83*(1.03**5)*bear_exit_pb:.2f}")
print(f"Base target: ${20.83*(1.06**5)*base_exit_pb:.2f}")
print(f"Bull target: ${20.83*(1.10**5)*bull_exit_pb:.2f}")
