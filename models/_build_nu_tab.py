"""Add a NU (Nu Holdings / Nubank) tab to projections/Projections.xlsx.

Clones the last existing tab (BABA) to preserve styling/formulas, then
overwrites the input cells with Nu Holdings numbers.

Nu Holdings (NYSE: NU) is a Cayman-incorporated digital bank ("Nubank")
operating in Brazil, Mexico and Colombia, reporting in USD under IFRS with a
December fiscal year.  Consistent with the other December-FY tabs (HOOD, NFLX),
the "2026" column is a FY2026 ESTIMATE built off FY2025 actuals and the Q1'26
run-rate; the model then projects FY2027-FY2031.  Basis: IFRS (GAAP) net income
/ diluted EPS, which is how NU is quoted.  All $ figures in millions.

The thesis: hyper-growth LatAm fintech compounding revenue on customer adds
(135M+ customers, +14% Y/Y), ARPAC expansion (~$16/mo) and a low cost-to-serve
(efficiency ratio ~20%), with Mexico hitting break-even in Q1'26 and a measured
U.S. optionality.  Near-term EPS is noisy from IFRS-9 credit-loss-allowance
builds (Q1'26 provisions +33% QoQ on portfolio growth/seasonality/mix) and BRL
FX translation (FX-neutral growth runs well above reported USD growth).  Net
margin expands toward the low-20s% as scale and operating leverage compound
(ROE ~29-33%); bull assumes faster growth + a fintech re-rating, bear assumes
Brazil credit-cycle stress and FX drag.

Sources (Q4/FY2025 results 2/25/2026; Q1'26 results 5/14/2026; consensus):
  - FY2025 revenue $16.3B (+45% Y/Y); net income $2,871.7M (attrib. $2,868.9M)
    diluted EPS $0.5846; diluted shares ~4,907M; ROE 33%; 131M customers
  - FY2024 revenue ~$11.2B; net income $1,972.1M; diluted EPS $0.4034
  - Q1'26: revenue >$5.0B (+42% FXN); net income $871M (+41% FXN); EPS $0.178;
    ROE 29%; 135.2M customers; credit portfolio $37.2B (+40% FXN); NIM 21.1%,
    risk-adjusted NIM 9.5%; efficiency ratio 17.6% (FY26 target ~20%);
    Mexico break-even reached; provisions/CLA $1.79B (+33% QoQ)
  - FY2026E (this model base): revenue ~$21B (~+29%); net income ~$3.8B (~18%
    margin); EPS ~$0.78
  - Consensus / model views: revenue CAGR mid-to-high teens; net margin ~20%;
    NU compounding net income >80% Y/Y since 2022; TIKR model PT ~$25
  - Price ~$12 (May 2026); high-growth fintech; $1.0B buyback authorized Jun 2026
  - Risks (bear): Brazil household-debt/credit cycle, BRL depreciation,
    concentration in unsecured/credit-card lending, U.S./Mexico execution
"""

from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "projections" / "Projections.xlsx"

# Shared inputs (USD millions; shares = diluted shares in millions)
SHARES = 4900         # diluted shares (~4,907M FY25); ~flat ($1B buyback vs SBC)
REV_2026 = 21000      # FY2026E revenue (~+29% off $16.3B; Q1'26 ~$5.0B run-rate)
NI_2026 = 3800        # FY2026E IFRS net income (~18% margin, EPS ~$0.78)
NI_G_2026 = 0.32      # net income growth FY2025->FY2026E (~$2.87B -> ~$3.8B)

# Analyst IFRS net income estimates FY2026-FY2029 (USD M)
# (EPS ~$0.78 / $1.00 / $1.27 / $1.57 on ~4,900M shares.)
ANALYST = {"C": 3800, "D": 4900, "E": 6200, "F": 7700}

# Per-case assumptions.  Each template block is offset by 24 rows:
#   Bull base row 4, Base 28, Bear 52.  Within a block, relative rows:
#     +2 revenue, +3 revenue growth, +5 net income, +6 NI growth,
#     +8 analyst est, +10 NI margins, +14 PE low, +15 PE high, H(+0) shares.
# NOTE: revenue growth applies with a one-column lag (D6 = C6*(1+C7)), so the
# column-C growth rate drives FY2027 revenue, column-D drives FY2028, etc.
# Revenue decelerates from high-20s% as the base scales (USD < FX-neutral due to
# BRL).  Net margin expands toward low-20s% on operating leverage + Mexico
# ramp (bull: faster growth + fintech re-rating; bear: Brazil credit cycle + FX).
CASES = {
    "bull": {
        "base": 4,
        # sustained customer/ARPAC growth + Mexico/US optionality
        "rev_g": {"C": 0.32, "D": 0.28, "E": 0.24, "F": 0.20, "G": 0.17, "H": 0.15},
        "margins": {"D": 0.21, "E": 0.225, "F": 0.24, "G": 0.245, "H": 0.25},
        "pe_low": {c: 20 for c in "CDEFGH"},
        "pe_high": {c: 30 for c in "CDEFGH"},
    },
    "base": {
        "base": 28,
        # high-teens/20s% growth decelerating; steady margin expansion
        "rev_g": {"C": 0.26, "D": 0.22, "E": 0.19, "F": 0.16, "G": 0.14, "H": 0.12},
        "margins": {"D": 0.19, "E": 0.20, "F": 0.21, "G": 0.21, "H": 0.22},
        "pe_low": {c: 14 for c in "CDEFGH"},
        "pe_high": {c: 22 for c in "CDEFGH"},
    },
    "bear": {
        "base": 52,
        # Brazil credit-cycle stress + BRL drag cap growth and margins
        "rev_g": {"C": 0.18, "D": 0.15, "E": 0.12, "F": 0.10, "G": 0.09, "H": 0.08},
        "margins": {"D": 0.16, "E": 0.16, "F": 0.165, "G": 0.17, "H": 0.17},
        "pe_low": {c: 9 for c in "CDEFGH"},
        "pe_high": {c: 14 for c in "CDEFGH"},
    },
}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    src = wb[wb.sheetnames[-1]]          # clone the last tab (BABA)
    ws = wb.copy_worksheet(src)
    ws.title = "NU"

    # Header
    ws["B2"] = "NU"
    ws["C2"] = datetime(2026, 6, 21)

    for cfg in CASES.values():
        b = cfg["base"]
        ws[f"H{b}"] = SHARES                      # diluted share count

        # Revenue base (FY2026E) + growth row
        ws[f"C{b + 2}"] = REV_2026
        for col, g in cfg["rev_g"].items():
            ws[f"{col}{b + 3}"] = g

        # IFRS net income base (FY2026E) + base-year growth
        ws[f"C{b + 5}"] = NI_2026
        ws[f"C{b + 6}"] = NI_G_2026

        # Analyst estimates (FY2026-FY2029)
        for col, v in ANALYST.items():
            ws[f"{col}{b + 8}"] = v

        # NI margins (D..H; C stays as formula =NI/Rev)
        for col, m in cfg["margins"].items():
            ws[f"{col}{b + 10}"] = m

        # PE low / high
        for col, v in cfg["pe_low"].items():
            ws[f"{col}{b + 14}"] = v
        for col, v in cfg["pe_high"].items():
            ws[f"{col}{b + 15}"] = v

    wb.save(XLSX)
    print(f"Added NU tab. Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
