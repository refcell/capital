#!/usr/bin/env python3
"""Build 6-sheet Excel valuation model for Mizuho Financial Group (MFG).
Bank-specific valuation: P/B and ROE driven. FCF and EV not applicable.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = Workbook()

# ── Styles ──
header_font = Font(bold=True, size=12)
title_font = Font(bold=True, size=14)
section_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def style_data_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    cell.border = thin_border

def style_data_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        style_data_cell(ws, row, c)

# ════════════════════════════════════════════════════════════
# Sheet 1: Valuation (bank-specific — P/B & ROE focused)
# ════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Valuation'

# Row 1: title block — merged across A:F
ws1.merge_cells('A1:F1')
ws1['A1'].value = 'Mizuho Financial Group, Inc. — Valuation Model'
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center')

# Title block rows 2-10
title_data = [
    ('Company:', 'Mizuho Financial Group, Inc.'),
    ('Ticker:', 'NYSE: MFG'),
    ('Date:', '2026-06-22'),
    ('Closing Price:', '$10.15'),
    ('Shares Outstanding:', '12.18B (Yahoo Finance)'),
    ('Market Cap:', '$119.60B'),
    ('Enterprise Value:', 'N/A (bank — deposits are operating liabilities)'),
    ('Primary Lens:', 'P/B and ROE (Residual Income framework)'),
    ('Stance:', 'Fairly valued → Watch'),
    ('Note:', 'FCF and FCF multiples are N/A for banks per valuation guidelines'),
]

for i, (label, val) in enumerate(title_data, 2):
    ws1.cell(row=i, column=1, value=label).font = section_font
    ws1.cell(row=i, column=2, value=val)

# Valuation metrics table starting row 13
r = 13
ws1.cell(row=r, column=1, value='Valuation Metric').font = section_font
ws1.cell(row=r, column=2, value='Value').font = section_font
ws1.cell(row=r, column=3, value='Comment').font = section_font
style_header_row(ws1, r, 3)

metrics = [
    ('P/B (Price / Book)', '1.78', 'Current; close to fair P/B of ~1.77'),
    ('P/B at FY2026 Book', '1.38', 'As of 3/31/2026; stock up ~29% since'),
    ('Forward P/E', '15.53', 'Implies modest but positive earnings growth'),
    ('Trailing P/E', '16.43', 'TTM basis, post-earnings surge'),
    ('P/S', '4.75', 'Revenue-to-market ratio; context-dependent for banks'),
    ('ROE (TTM)', '11.44%', 'Above cost of equity (est. 6.45%) → value creation'),
    ('ROA (TTM)', '0.43%', 'Low absolute number, normal for large Japanese banks'),
    ('Beta (5Y Monthly)', '0.39', 'Low volatility; defensive financial name'),
    ('FCF / FCF Multiple', 'N/A', 'Meaningless for banks — deposits offset loan origination'),
    ('EV-Based Metrics', 'N/A', 'Enterprise value not applicable for banks'),
    ('Dividend Yield (Fwd)', '1.81%', 'Modest yield; payout ratio 28.14%'),
    ('Book Value Per Share', '$5.79', 'Yahoo Finance MRQ basis'),
]

for i, (metric, val, comment) in enumerate(metrics, 14):
    ws1.cell(row=i, column=1, value=metric).font = Font(italic=True)
    ws1.cell(row=i, column=2, value=val)
    ws1.cell(row=i, column=3, value=comment)
    style_data_row(ws1, i, 3)

# Column widths
ws1.column_dimensions['A'].width = 28
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 50

# ════════════════════════════════════════════════════════════
# Sheet 2: WACC (CAPM with bank-specific inputs)
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('WACC')

ws2.merge_cells('A1:F1')
ws2['A1'].value = 'WACC Calculation — CAPM Components'
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center')

wacc_data = [
    ('Component', 'Value', 'Notes'),
    ('', '', ''),
    ('Risk-Free Rate (10Y US Treasury)', '4.501%', 'CNBC as of 2026-06-22 ~22:00 EDT'),
    ('Equity Risk Premium', '5.00%', 'Standard assumption'),
    ('Beta (Levered, 5Y Monthly)', '0.39', 'Yahoo Finance statistics'),
    ('Cost of Equity (CAPM)', '6.45%', '= 4.50% + 0.39 × 5.00%'),
    ('', '', ''),
    ('Cost of Debt', 'N/A', 'Bank deposits are operating liabilities; not used in traditional WACC'),
    ('Tax Rate', '22.10%', 'TTM effective tax rate: 368,127 / 1,622,285'),
    ('', '', ''),
    ('Market Cap', '$119.60B', 'Yahoo Finance current'),
    ('Total Debt (accounting)', '¥26.18T', 'Yahoo Finance balance sheet — but deposits offset this'),
    ('Equity Weight', '~100%', 'For bank WACC, equity dominates'),
    ('Debt Weight', '~0%', 'Deposits not treated as financial debt'),
    ('', '', ''),
    ('WACC (bank-specific)', '6.45%', '≈ Cost of Equity given equity-dominant capital structure'),
    ('', '', ''),
    ('Residual Income Fair P/B', '~1.77', 'ROE/Ke = 11.44%/6.45% ≈ fair value multiple'),
    ('Current P/B vs. Fair P/B', '1.78 vs 1.77', 'Slightly above fair value; consistent with Watch stance'),
]

for i, (comp, val, note) in enumerate(wacc_data, 2):
    ws2.cell(row=i, column=1, value=comp)
    ws2.cell(row=i, column=2, value=val)
    ws2.cell(row=i, column=3, value=note)
    if comp:  # not blank separator
        ws2.cell(row=i, column=1).font = section_font if comp else Font()
        style_data_row(ws2, i, 3)

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 55

# ════════════════════════════════════════════════════════════
# Sheet 3: Scenarios (Bank-specific: BVPS × P/B)
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Scenarios')

ws3.merge_cells('A1:H1')
ws3['A1'].value = 'Scenario Analysis — Bank Valuation (BVPS × P/B Framework)'
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='center')

# Framework note
ws3.cell(row=2, column=1, value='Framework:').font = section_font
ws3.cell(row=3, column=1, value='For banks, value = Book Value Per Share × Price/Book Multiple.').font = Font(italic=True)
ws3.cell(row=4, column=1, value='FCF-based DCF is not appropriate — deposits offset loan origination flow.').font = Font(italic=True)
ws3.cell(row=5, column=1, value='Current BVPS: $5.79 (Yahoo Finance MRQ). Current P/B: 1.78. Current Price: $10.15.').font = Font(italic=True)

# Scenario table
r = 7
scen_headers = ('Item', 'Bear', 'Base', 'Bull')
for c, h in enumerate(scen_headers, 1):
    ws3.cell(row=r, column=c, value=h).font = Font(bold=True)
    ws3.cell(row=r, column=c).fill = header_fill
    ws3.cell(row=r, column=c).border = thin_border

scen_data = [
    ('BVPS (Current)', '$5.79', '$5.79', '$5.79'),
    ('BVPS CAGR (5Y assumption)', '6.0%', '8.2%', '10.0%'),
    ('BVPS in 5 Years', '$7.75', '$8.57', '$9.33'),
    ('Exit P/B Multiple', '1.30x', '1.70x', '2.00x'),
    ('Implied Price (5Y Target)', '$10.08', '$14.57', '$18.66'),
    ('Weight', '25%', '50%', '25%'),
    ('Weighted Value/Share', '$2.52', '$7.29', '$4.67'),
    ('Upside from $10.15', '-0.7%', '+43.5%', '+84.1%'),
]

# Probability-weighted fair value row
for i, row_data in enumerate(scen_data, 8):
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=i, column=c, value=val)
        ws3.cell(row=i, column=c).border = thin_border
    if row_data[0] in ('Probability-Weighted Fair Value',):
        for c in range(1, 5):
            ws3.cell(row=i, column=c).fill = green_fill

# Add fair value row
r_fv = 8 + len(scen_data)
ws3.cell(row=r_fv, column=1, value='Probability-Weighted Fair Value').font = Font(bold=True)
# 0.25×10.08 + 0.50×14.57 + 0.25×18.66 = 2.52 + 7.29 + 4.67 = $14.48
ws3.cell(row=r_fv, column=2, value='$14.48').font = Font(bold=True)
ws3.cell(row=r_fv, column=2).fill = green_fill
for c in range(1, 5):
    ws3.cell(row=r_fv, column=c).border = thin_border

ws3.cell(row=r_fv + 1, column=1, value='Implied Total Upside').font = Font(bold=True)
ws3.cell(row=r_fv + 1, column=2, value='+42.6%').font = Font(bold=True)

# Scenario rationale
r_rationale = r_fv + 3
ws3.cell(row=r_rationale, column=1, value='Scenario Rationale').font = section_font
ws3.cell(row=r_rationale + 1, column=1, value='Bear: ROE normalizes to ~9% as rate environment normalizes; P/B compresses to 1.3x. Japan\'s banking sector has seen P/B mean-reversion to historical average of ~1.2-1.3x.')
ws3.cell(row=r_rationale + 2, column=1, value='Base: ROE sustains at 11.4%+ as Mizuho benefits from Japan rate normalization and fee income growth; P/B stays at 1.7x, consistent with current levels.')
ws3.cell(row=r_rationale + 3, column=1, value='Bull: ROE expands to 13%+ with continued rate rises, fee compression reversals, and digital transformation gains; P/B re-rates to 2.0x as market recognizes improved profitability.')

ws3.column_dimensions['A'].width = 35
for c in ['B', 'C', 'D']:
    ws3.column_dimensions[c].width = 18

# ════════════════════════════════════════════════════════════
# Sheet 4: Actuals Source Audit
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Actuals Source Audit')

ws4.merge_cells('A1:E1')
ws4['A1'].value = 'Data Source Audit — Every Data Point Traced'
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(horizontal='center')

audit_header = ('Data Point', 'Value', 'Source URL', 'Date Accessed', 'Notes')
r = 2
for c, h in enumerate(audit_header, 1):
    ws4.cell(row=r, column=c, value=h).font = Font(bold=True)
    ws4.cell(row=r, column=c).fill = header_fill
    ws4.cell(row=r, column=c).border = thin_border

audit_data = [
    ('Stock Price (close)', '$10.15', 'Yahoo Finance quote page', '2026-06-22', 'Regular market close'),
    ('Overnight Price', '$9.97', 'Yahoo Finance quote page', '2026-06-22', 'Blue Ocean ATS session'),
    ('52-Week Range', '$5.39 - $10.39', 'Yahoo Finance statistics', '2026-06-22', 'Near 52-week high'),
    ('Market Cap (current)', '$119.60B', 'Yahoo Finance statistics', '2026-06-22', 'Intraday estimate'),
    ('Shares Outstanding', '12.18B', 'Yahoo Finance statistics', '2026-06-22', 'See Q3 — discrepancy with BS'),
    ('Beta (5Y monthly)', '0.39', 'Yahoo Finance statistics', '2026-06-22', 'Low vol — defensive name'),
    ('Revenue FY2026', '¥4,317B', 'Yahoo Finance /financials', '2026-06-22', 'JPY, in thousands'),
    ('Revenue FY2025', '¥3,809B', 'Yahoo Finance /financials', '2026-06-22', ''),
    ('Revenue FY2024', '¥2,971B', 'Yahoo Finance /financials', '2026-06-22', ''),
    ('Revenue FY2023', '¥2,670B', 'Yahoo Finance /financials', '2026-06-22', ''),
    ('Net Income FY2026', '¥1,249B', 'Yahoo Finance /financials', '2026-06-22', ''),
    ('Net Income FY2025', '¥885B', 'Yahoo Finance /financials', '2026-06-22', ''),
    ('Diluted EPS FY2026', '¥100.58', 'Yahoo Finance /financials', '2026-06-22', 'JPY per share'),
    ('Diluted EPS FY2025', '¥70.04', 'Yahoo Finance /financials', '2026-06-22', ''),
    ('Total Assets FY2026', '¥302,240B', 'Yahoo Finance /balance-sheet', '2026-06-22', ''),
    ('Total Equity FY2026', '¥11,404B', 'Yahoo Finance /balance-sheet', '2026-06-22', ''),
    ('Tangible BV FY2026', '¥10,434B', 'Yahoo Finance /balance-sheet', '2026-06-22', ''),
    ('Total Debt FY2026', '¥26,181B', 'Yahoo Finance /balance-sheet', '2026-06-22', ''),
    ('BVPS (Yahoo MRQ)', '$5.79', 'Yahoo Finance statistics', '2026-06-22', 'Most recent quarter'),
    ('ROE (TTM)', '11.44%', 'Yahoo Finance statistics', '2026-06-22', ''),
    ('ROA (TTM)', '0.43%', 'Yahoo Finance statistics', '2026-06-22', ''),
    ('P/B (current)', '1.78', 'Yahoo Finance statistics', '2026-06-22', ''),
    ('Trailing P/E', '16.43', 'Yahoo Finance statistics', '2026-06-22', ''),
    ('Forward P/E', '15.53', 'Yahoo Finance statistics', '2026-06-22', ''),
    ('Effective Tax Rate', '22.1%', 'Calculated from income statement', '2026-06-22', '368,127/1,622,285'),
    ('Dividend Yield (fwd)', '1.81%', 'Yahoo Finance statistics', '2026-06-22', ''),
    ('Payout Ratio', '28.14%', 'Yahoo Finance statistics', '2026-06-22', ''),
    ('Revenue Est FY2027', '¥3.99T avg', 'Yahoo Finance /analysis', '2026-06-22', 'vs ¥2.08T YoY prior'),
    ('Revenue Est FY2028', '¥4.39T avg', 'Yahoo Finance /analysis', '2026-06-22', 'vs ¥2.45T YoY prior'),
    ('10Y US Treasury Yield', '4.501%', 'CNBC US10Y quote', '2026-06-22', '~22:00 EDT'),
    ('StockAnalysis availability', '404', 'stockanalysis.com/quotes/mfg/', '2026-06-22', 'Not available — used Yahoo Finance'),
]

for i, (dp, val, src, dt, note) in enumerate(audit_data, 3):
    ws4.cell(row=i, column=1, value=dp)
    ws4.cell(row=i, column=2, value=val)
    ws4.cell(row=i, column=3, value=src)
    ws4.cell(row=i, column=4, value=dt)
    ws4.cell(row=i, column=5, value=note)
    style_data_row(ws4, i, 5)

ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 16
ws4.column_dimensions['C'].width = 35
ws4.column_dimensions['D'].width = 14
ws4.column_dimensions['E'].width = 40

# ════════════════════════════════════════════════════════════
# Sheet 5: Questions
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('Questions')

ws5.merge_cells('A1:C1')
ws5['A1'].value = 'Open Questions & Items Requiring Investigation'
ws5['A1'].font = title_font
ws5['A1'].alignment = Alignment(horizontal='center')

questions = [
    ('Q1', 'Share count discrepancy',
     'Yahoo Finance shows 12.18B shares outstanding, but the balance sheet shows only 2.44B ordinary shares. Mizuho likely has ADS conversions, preferred shares, or depositary receipts. Need to verify the ADS conversion ratio and confirm which share count is appropriate for per-share calculations.'),
    ('Q2', 'Revenue estimate doubling',
     'FY2027 revenue consensus (¥3.99T) shows 92%+ growth over the prior year estimate (¥2.08T). This near-doubling suggests a major accounting change, business acquisition, or merger integration. Need to understand whether this is organic growth or a restatement effect.'),
    ('Q3', 'Japanese banking regulation',
     'How do FSA (Financial Services Agency) capital requirements affect Mizuho\'s dividend sustainability and share buyback capacity? Japan\'s bank capital framework differs from Basel frameworks in key respects.'),
    ('Q4', 'FX risk for USD investors',
     'Mizuho reports in JPY but trades in USD. Currency fluctuations of 10-15% are common in USD/JPY. For a bank whose fundamentals are JPY-denominated, the USD price includes implicit FX optionality. How does the current ¥/USD rate affect valuation?'),
    ('Q5', 'Cost of funds dynamics',
     'Interest income was ¥5,852B vs interest expense of ¥4,475B → NII of ¥1,377B. But interest expense rose to ¥4,955B in FY2025 and is declining. How sustainable is the current NII trajectory as BOJ continues rate normalization?'),
    ('Q6', 'Asset growth vs ROE trajectory',
     'Total assets grew from ¥254T to ¥302T in 3 years (+19%), while equity grew from ¥9.2T to ¥11.4T (+24%). Is this driven by organic earnings compounding or capital raises? What is the organic ROE trend?'),
    ('Q7', 'Non-performing loans (NPLs)',
     'No NPL data extracted in this session. Japan\'s banking sector carries legacy NPL issues. Mizuho\'s NPL ratio and coverage ratio are critical quality-of-earnings metrics for a bank.'),
    ('Q8', 'Fee income and securities arm',
     'Mizuho has a significant securities/asset management division. How does its profitability compare to pure-play Japanese banks? Is Mizuho being under-valued because the market treats it as a traditional bank?'),
    ('Q9', 'Government ownership',
     'Japan\'s major banks historically had significant government ownership. Does the Japanese government or public institutions still hold meaningful blocks in Mizuho? If so, what are the governance implications?'),
    ('Q10', 'Tax rate anomalies',
     'Effective tax rate of 22.1% (FY2026) vs 23.8% (FY2025) — is Japan\'s corporate tax rate changing? Any deferred tax assets/liabilities that distort the effective rate?'),
]

ws5.cell(row=2, column=1, value='No.').font = Font(bold=True)
ws5.cell(row=2, column=1).fill = header_fill
ws5.cell(row=2, column=2, value='Topic').font = Font(bold=True)
ws5.cell(row=2, column=2).fill = header_fill
ws5.cell(row=2, column=3, value='Detail').font = Font(bold=True)
ws5.cell(row=2, column=3).fill = header_fill

for i, (num, topic, detail) in enumerate(questions, 3):
    ws5.cell(row=i, column=1, value=num).font = section_font
    ws5.cell(row=i, column=2, value=topic).font = Font(italic=True)
    ws5.cell(row=i, column=3, value=detail)
    ws5.cell(row=i, column=3).alignment = Alignment(wrap_text=True)
    style_data_row(ws5, i, 3)

ws5.column_dimensions['A'].width = 6
ws5.column_dimensions['B'].width = 30
ws5.column_dimensions['C'].width = 80

# ════════════════════════════════════════════════════════════
# Sheet 6: Sources
# ════════════════════════════════════════════════════════════
ws6 = wb.create_sheet('Sources')

ws6.merge_cells('A1:C1')
ws6['A1'].value = 'Data Sources'
ws6['A1'].font = title_font
ws6['A1'].alignment = Alignment(horizontal='center')

sources = [
    (1, 'Yahoo Finance — Quote & Market Data', 'https://finance.yahoo.com/quote/MFG/'),
    (2, 'Yahoo Finance — Income Statement', 'https://finance.yahoo.com/quote/MFG/financials/'),
    (3, 'Yahoo Finance — Balance Sheet', 'https://finance.yahoo.com/quote/MFG/balance-sheet/'),
    (4, 'Yahoo Finance — Cash Flow', 'https://finance.yahoo.com/quote/MFG/cash-flow/'),
    (5, 'Yahoo Finance — Key Statistics', 'https://finance.yahoo.com/quote/MFG/key-statistics/'),
    (6, 'Yahoo Finance — Analyst Estimates', 'https://finance.yahoo.com/quote/MFG/analysis/'),
    (7, 'CNBC — 10Y US Treasury Yield', 'https://www.cnbc.com/quotes/US10Y'),
    (8, 'StockAnalysis (unavailable — 404)', 'https://stockanalysis.com/quotes/mfg/'),
    (9, 'Yahoo Finance — EPS (¥84.35 TTM diluted)', 'Per income statement table, basic=diluted'),
    (10, 'Note: All financials in JPY thousands per Yahoo Finance', 'Currency note — market data in USD'),
]

ws6.cell(row=2, column=1, value='No.').font = Font(bold=True)
ws6.cell(row=2, column=1).fill = header_fill
ws6.cell(row=2, column=2, value='Source').font = Font(bold=True)
ws6.cell(row=2, column=2).fill = header_fill
ws6.cell(row=2, column=3, value='URL / Reference').font = Font(bold=True)
ws6.cell(row=2, column=3).fill = header_fill

for i, (num, src, url) in enumerate(sources, 3):
    ws6.cell(row=i, column=1, value=num)
    ws6.cell(row=i, column=2, value=src)
    ws6.cell(row=i, column=3, value=url)
    style_data_row(ws6, i, 3)

# Set wrap text on description columns
for c in range(1, 4):
    ws6.cell(row=2, column=c).alignment = Alignment(horizontal='center')

ws6.column_dimensions['A'].width = 6
ws6.column_dimensions['B'].width = 45
ws6.column_dimensions['C'].width = 55

# ════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════
output_path = '/home/refcell/dev/capital/models/2026-06-22 Mizuho Financial Group Model.xlsx'
wb.save(output_path)
print(f"Saved: {output_path}")

# Verify by reading back
wb2 = load_workbook(output_path, data_only=True)
print(f"Sheets: {wb2.sheetnames}")
print(f"Sheet 1 (Valuation) has {ws1.max_row*ws1.max_column} cells")
print("All 6 sheets created successfully.")
