#!/usr/bin/env python3
"""Build the UBS Group six-sheet bank valuation workbook."""
from math import isclose
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT = Path(__file__).with_name("2026-09-03 UBS Group Model.xlsx")
PRICE = 54.90
SHARES = 3062.0  # millions
MARKET_CAP = 168180.0  # USD millions
BOOK_EQUITY = 89430.0  # USD millions
TANGIBLE_EQUITY = 82319.0  # USD millions
BVPS = 29.12
TBVPS = 26.89
TTM_REVENUE = 53040.0
TTM_NET_INCOME = 9520.0
TTM_EPS = 2.93
FY26_EPS = 3.69
FY26_REVENUE = 53680.0

# Bank valuation: compound current book value for five years, apply a terminal
# P/B multiple, then discount at a scenario-specific cost of equity.
SCENARIOS = {
    "Bear": {"bvps_cagr": 0.03, "roe": 0.09, "exit_pb": 1.50, "discount": 0.115, "weight": 0.20},
    "Base": {"bvps_cagr": 0.08, "roe": 0.125, "exit_pb": 2.00, "discount": 0.090, "weight": 0.55},
    "Bull": {"bvps_cagr": 0.11, "roe": 0.15, "exit_pb": 2.35, "discount": 0.085, "weight": 0.25},
}
for case in SCENARIOS.values():
    case["terminal_bvps"] = BVPS * (1 + case["bvps_cagr"]) ** 5
    case["terminal_price"] = case["terminal_bvps"] * case["exit_pb"]
    case["target"] = case["terminal_price"] / (1 + case["discount"]) ** 5
    case["upside"] = case["target"] / PRICE - 1
weighted_value = sum(case["target"] * case["weight"] for case in SCENARIOS.values())

assert SCENARIOS["Bear"]["target"] < PRICE
assert abs(SCENARIOS["Base"]["target"] - 55.03) / 55.03 < 0.20

rf = 4.792
erp = 5.00
beta = 0.83
cost_equity = rf + beta * erp

wb = Workbook()
navy, blue, gold, gray, white = "17365D", "D9EAF7", "D8B34B", "E7E6E6", "FFFFFF"
thin = Side(style="thin", color="A6A6A6")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def put(ws, row, col, value, *, bold=False, fill=None, color="000000", size=10, wrap=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Aptos", size=size, bold=bold, color=color)
    cell.fill = PatternFill("solid", fgColor=fill) if fill else PatternFill(fill_type=None)
    cell.border = border
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    return cell


def title(ws, text, end_col=5, subtitle=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    put(ws, 1, 1, text, bold=True, fill=navy, color=white, size=15)
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        put(ws, 2, 1, subtitle, bold=True, fill=blue)


def header(ws, row, values):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=True, fill=navy, color=white)


def widths(ws, values):
    for col, width in enumerate(values, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A3"


# 1. Valuation
ws = wb.active
ws.title = "Valuation"
title(ws, "UBS Group AG (UBS) — Valuation", 4, "Quote and model date: September 3, 2026 | USD unless stated")
facts = [
    ("Company", "UBS Group AG", "Swiss global wealth manager and universal bank"),
    ("Ticker", "NYSE: UBS", "Swiss ordinary shares represented in New York"),
    ("Price", PRICE, "StockAnalysis September 3 snapshot"),
    ("Shares outstanding (M)", SHARES, "Filing-date common shares"),
    ("Market capitalization ($M)", MARKET_CAP, "Price × shares; StockAnalysis"),
    ("Enterprise value", "N/A", "Deposits and trading liabilities are operating funding; EV is not meaningful for banks"),
    ("Book value per share", BVPS, "$89.43B common equity / 3.062B shares"),
    ("Tangible book value per share", TBVPS, "$82.32B tangible common equity / 3.062B shares"),
    ("Primary valuation lens", "P/B, P/TBV and ROE", "Residual-income logic; FCF and EV multiples are not decision-useful"),
    ("Stance", "Watch", "Credit Suisse integration is producing earnings, but the current multiple requires durable ROE above cost of equity"),
]
header(ws, 3, ["Field", "Value", "Comment"])
for row, values in enumerate(facts, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))

header(ws, 15, ["Valuation metric", "Value", "Interpretation"])
metrics = [
    ("Trailing P/E", 17.67, "TTM EPS is still affected by Credit Suisse integration and prior purchase-accounting volatility"),
    ("Forward P/E", 13.58, "StockAnalysis current forward multiple"),
    ("P/S", 3.17, "Revenue is a weak bank valuation denominator"),
    ("P/B", 1.88, "Premium requires sustained ROE and successful integration"),
    ("P/TBV", 2.04, "Stricter capital-base multiple"),
    ("ROE", 0.1067, "TTM return remains only modestly above the model cost of equity"),
    ("ROA", 0.0057, "Normal bank-style asset return; leverage converts this into ROE"),
    ("Payout ratio", 0.1874, "Low reported payout leaves room for capital return, subject to Swiss requirements"),
    ("Buyback yield", 0.0271, "Repurchases are the larger current capital-return lever"),
    ("EV / FCF", "N/A", "FCF and EV are not meaningful for deposit-funded banks"),
]
for row, values in enumerate(metrics, 16):
    for col, value in enumerate(values, 1):
        cell = put(ws, row, col, value, bold=(col == 1))
        if col == 2 and values[0] in {"ROE", "ROA", "Payout ratio", "Buyback yield"}:
            cell.number_format = "0.0%"
widths(ws, [30, 24, 86, 14])

# 2. WACC / cost of equity
ws = wb.create_sheet("WACC")
title(ws, "UBS Group — Cost of Equity", 4, "CAPM; conventional corporate WACC is not meaningful for a deposit-funded bank")
header(ws, 3, ["Component", "Value", "Source / formula"])
wacc_rows = [
    ("Risk-free rate", rf / 100, "U.S. 10-year Treasury snapshot carried from September 1, 2026"),
    ("Equity risk premium", erp / 100, "Model assumption"),
    ("Levered beta", beta, "StockAnalysis five-year beta"),
    ("Cost of equity", cost_equity / 100, "Risk-free rate + beta × ERP"),
    ("TTM ROE", 0.1067, "StockAnalysis through June 30, 2026"),
    ("ROE spread", 0.1067 - cost_equity / 100, "TTM ROE − cost of equity"),
    ("Conventional WACC", "N/A", "Deposits are operating inputs and debt cannot be separated cleanly from bank operations"),
    ("Valuation discount rate", 0.09, "Base-case cost of equity rounded to 9.0%"),
]
for row, values in enumerate(wacc_rows, 4):
    for col, value in enumerate(values, 1):
        cell = put(ws, row, col, value, bold=(col == 1 or values[0] == "Cost of equity"), fill=(gold if values[0] == "Cost of equity" else None))
        if col == 2 and isinstance(value, float) and values[0] != "Levered beta":
            cell.number_format = "0.00%"
widths(ws, [31, 20, 87, 14])

# 3. Scenarios
ws = wb.create_sheet("Scenarios")
title(ws, "UBS Group — P/B and ROE Scenarios", 6, "Five-year terminal book value discounted to September 2026; bank-specific residual-income framework")
header(ws, 3, ["Metric", "Bear", "Base", "Bull", "Notes"])
scenario_rows = [
    ("Starting BVPS", BVPS, BVPS, BVPS, "June 30, 2026 book value per share"),
    ("BVPS CAGR (5Y)", *(SCENARIOS[k]["bvps_cagr"] for k in ("Bear", "Base", "Bull")), "Retained earnings less buybacks/dividends and regulatory capital needs"),
    ("Normalized ROE", *(SCENARIOS[k]["roe"] for k in ("Bear", "Base", "Bull")), "Integration failure / durable execution / synergy realization"),
    ("Terminal BVPS", *(SCENARIOS[k]["terminal_bvps"] for k in ("Bear", "Base", "Bull")), "Starting BVPS compounded five years"),
    ("Exit P/B", *(SCENARIOS[k]["exit_pb"] for k in ("Bear", "Base", "Bull")), "Premium varies with ROE spread, capital rules and franchise durability"),
    ("Terminal price", *(SCENARIOS[k]["terminal_price"] for k in ("Bear", "Base", "Bull")), "Terminal BVPS × exit P/B"),
    ("Cost of equity", *(SCENARIOS[k]["discount"] for k in ("Bear", "Base", "Bull")), "Scenario-specific discount rate"),
    ("Present target price", *(SCENARIOS[k]["target"] for k in ("Bear", "Base", "Bull")), "Five-year terminal value discounted to present"),
    ("Upside / (downside)", *(SCENARIOS[k]["upside"] for k in ("Bear", "Base", "Bull")), "Versus $54.90"),
    ("Probability", *(SCENARIOS[k]["weight"] for k in ("Bear", "Base", "Bull")), "20% / 55% / 25%"),
    ("Weighted value/share", *(SCENARIOS[k]["target"] * SCENARIOS[k]["weight"] for k in ("Bear", "Base", "Bull")), "Contribution to probability-weighted fair value"),
]
for row, values in enumerate(scenario_rows, 4):
    for col, value in enumerate(values, 1):
        cell = put(ws, row, col, value, bold=(col == 1))
        if col in (2, 3, 4) and values[0] in {"BVPS CAGR (5Y)", "Normalized ROE", "Cost of equity", "Upside / (downside)", "Probability"}:
            cell.number_format = "0.0%"
        elif col in (2, 3, 4) and values[0] in {"Starting BVPS", "Terminal BVPS", "Terminal price", "Present target price", "Weighted value/share"}:
            cell.number_format = "$0.00"
put(ws, 16, 1, "Probability-weighted fair value", bold=True, fill=gold)
put(ws, 16, 2, weighted_value, bold=True, fill=gold).number_format = "$0.00"
put(ws, 17, 1, "Upside from current price", bold=True, fill=gold)
put(ws, 17, 2, weighted_value / PRICE - 1, bold=True, fill=gold).number_format = "0.0%"
put(ws, 19, 1, "Framework note", bold=True, fill=gray)
put(ws, 19, 2, "UBS is a bank. Deposits fund loans and trading assets, so enterprise value and free cash flow are not economically comparable with non-financial companies. The model values common equity through book-value growth, normalized ROE, and an exit P/B multiple. The present target is a discounted terminal value, not a one-year trading forecast.", fill=gray)
ws.merge_cells("B19:E19")
widths(ws, [31, 18, 18, 18, 89, 14])

# 4. Actuals Source Audit
ws = wb.create_sheet("Actuals Source Audit")
title(ws, "UBS Group — Actuals Source Audit", 5, "Financial statement figures in USD millions unless noted")
header(ws, 3, ["Data point", "Value", "Source URL", "Source date", "Notes"])
sa = "https://stockanalysis.com/stocks/ubs"
audit = [
    ("Stock price", "$54.90", f"{sa}/", "2026-09-03", "Market snapshot"),
    ("Market cap", "$168.18B", f"{sa}/statistics/", "2026-09-03", "Price × current shares"),
    ("Enterprise value", "N/A", f"{sa}/statistics/", "2026-09-03", "Not applicable to banks"),
    ("Shares outstanding", "3.062B", f"{sa}/statistics/", "2026-09-03", "Down 2.71% YoY"),
    ("Revenue TTM", "$53.04B", f"{sa}/financials/", "2026-06-30", "+10.74%"),
    ("Net income TTM", "$9.52B", f"{sa}/financials/", "2026-06-30", "+51.5%; prior years distorted by Credit Suisse combination"),
    ("EPS TTM", "$2.93", f"{sa}/financials/", "2026-06-30", "Diluted provider figure"),
    ("Operating income TTM", "$16.52B", f"{sa}/statistics/", "2026-06-30", "31.14% provider operating margin"),
    ("Total assets", "$1.707T", f"{sa}/financials/balance-sheet/", "2026-06-30", "Includes trading assets and loans"),
    ("Gross loans", "$685.38B", f"{sa}/financials/balance-sheet/", "2026-06-30", "Allowance of $2.545B"),
    ("Total deposits", "$784.85B", f"{sa}/financials/balance-sheet/", "2026-06-30", "$526.45B interest-bearing; $258.40B non-interest-bearing"),
    ("Common equity", "$89.17B", f"{sa}/financials/balance-sheet/", "2026-06-30", "Excludes $265M minority interest"),
    ("Book value per share", "$29.12", f"{sa}/financials/balance-sheet/", "2026-06-30", "Current common shares"),
    ("Tangible book value", "$82.32B", f"{sa}/financials/balance-sheet/", "2026-06-30", "$26.89 per share"),
    ("Provision for credit losses TTM", "$452M", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "Low relative to loans; must be tested through cycle"),
    ("Repurchases TTM", "$6.294B", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "2.71% buyback yield at current market cap"),
    ("Common dividends paid TTM", "$3.404B", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "Cash-flow statement"),
    ("FY2026 revenue consensus", "$53.68B", f"{sa}/forecast/", "2026-09-03", "13 analysts; +12.88% versus provider FY2025 base"),
    ("FY2026 EPS consensus", "$3.69", f"{sa}/forecast/", "2026-09-03", "13 analysts; range $3.52-$3.80"),
    ("FY2026 net income consensus", "$11.19B", f"{sa}/forecast/", "2026-09-03", "Data-provider estimate"),
    ("Average analyst target", "$55.03", f"{sa}/forecast/", "2026-09-03", "Three displayed analysts; $50-$62.60 range"),
    ("P/B", "1.88x", f"{sa}/statistics/", "2026-09-03", "Current price / BVPS"),
    ("P/TBV", "2.04x", f"{sa}/statistics/", "2026-09-03", "Current price / TBVPS"),
    ("ROE", "10.67%", f"{sa}/statistics/", "2026-06-30", "TTM provider return"),
    ("Beta", "0.83", f"{sa}/statistics/", "2026-09-03", "Five-year beta"),
    ("Last reported earnings", "July 29, 2026", f"{sa}/statistics/", "2026-09-03", "Already released; next quarter is the forward catalyst"),
]
for row, values in enumerate(audit, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
widths(ws, [32, 21, 96, 16, 70])

# 5. Questions
ws = wb.create_sheet("Questions")
title(ws, "UBS Group — Open Questions", 4, "Items that can change valuation or conviction")
header(ws, 3, ["#", "Question", "Why it matters", "Best evidence / next check"])
questions = [
    (1, "What is the remaining Credit Suisse legal, operational and client-migration tail risk?", "Integration benefits are central to consensus earnings, while tail liabilities can consume capital unexpectedly.", "Quarterly litigation provisions, non-core runoff and migration milestones."),
    (2, "Can UBS sustain normalized ROE above its roughly 9% cost of equity after integration costs fade?", "A 1.9x book multiple requires a durable positive residual-income spread.", "Reported and underlying ROE excluding purchase-accounting effects."),
    (3, "How much additional CET1 capital will Swiss regulators require for the enlarged group?", "Higher capital requirements can reduce buybacks, lower ROE and compress justified P/B.", "Final Swiss too-big-to-fail rules and UBS capital plan."),
    (4, "What is current CET1 capital adequacy and how much management buffer is truly distributable?", "Buybacks and dividends depend on excess capital after regulatory and stress buffers.", "Quarterly CET1 ratio, risk-weighted assets and management target."),
    (5, "Is the $452M TTM provision adequate for the $685B gross loan book?", "Current credit costs are low; normalization or commercial-real-estate stress would reduce earnings.", "NPL ratio, stage 2/3 loans, charge-offs and allowance coverage."),
    (6, "What portion of acquired Credit Suisse assets still sits in Non-Core and Legacy?", "Runoff can release capital but can also create valuation and litigation losses.", "NCL risk-weighted assets, operating losses and exit timetable."),
    (7, "How much of wealth-management growth is net new money versus market appreciation and FX?", "Flows are the higher-quality source of recurring fee growth.", "Net new assets by region and mandate type."),
    (8, "Can Global Wealth Management protect fee margins as cash shifts into lower-fee products?", "Fee compression can offset asset growth and weaken operating leverage.", "Invested assets, recurring fee margin and client cash balances."),
    (9, "What is the deposit beta and funding-cost trajectory after rate cuts or curve shifts?", "Net interest income responds to both asset yields and how quickly deposit pricing resets.", "Interest-bearing deposit cost and NII sensitivity disclosure."),
    (10, "How volatile is Investment Bank revenue and risk-weighted asset usage under a weaker market?", "The segment adds earnings but can dilute wealth-management quality and consume scarce capital.", "VaR, RWA, compensation ratio and revenue mix."),
    (11, "Are the TTM $6.29B of repurchases accretive at roughly 1.9x book?", "Buying above book creates value only if sustainable ROE comfortably exceeds cost of equity.", "Average purchase price, canceled shares and pro forma BVPS impact."),
    (12, "How should the FY2023 $27.4B net income be normalized?", "Credit Suisse bargain-purchase accounting makes that year unusable as an earnings baseline.", "Purchase price allocation and underlying profit reconciliation."),
    (13, "Why did reported cash and investments rise to $575B while total debt reached $599B?", "For a bank these are operating positions, but liquidity composition and encumbrance determine resilience.", "LCR, NSFR, high-quality liquid assets and secured funding detail."),
    (14, "How concentrated is exposure to Swiss mortgages, U.S. commercial real estate and leveraged finance?", "A benign aggregate provision can hide pockets of tail risk.", "Loan-book concentrations and stress-loss disclosures."),
    (15, "When is the next quarterly earnings release after July 29?", "The next report should test integration execution, flows, credit costs and capital-return capacity.", "UBS investor-relations calendar."),
]
for row, values in enumerate(questions, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
widths(ws, [7, 82, 70, 66])

# 6. Sources
ws = wb.create_sheet("Sources")
title(ws, "UBS Group — Sources", 4, "Public sources accessed September 3, 2026")
header(ws, 3, ["#", "Source", "URL", "Use"])
sources = [
    (1, "StockAnalysis overview", f"{sa}/", "Price, company identity and headline market data"),
    (2, "StockAnalysis financials", f"{sa}/financials/", "Historical revenue, net income, EPS and segment revenue"),
    (3, "StockAnalysis balance sheet", f"{sa}/financials/balance-sheet/", "Assets, loans, deposits, equity, debt, BVPS and TBVPS"),
    (4, "StockAnalysis cash flow", f"{sa}/financials/cash-flow-statement/", "Provisions, repurchases, dividends and cash-flow volatility"),
    (5, "StockAnalysis statistics", f"{sa}/statistics/", "Shares, valuation, ROE, beta and analyst summary"),
    (6, "StockAnalysis forecast", f"{sa}/forecast/", "FY2026 revenue, net income, EPS and price-target consensus"),
    (7, "StockAnalysis ratios", f"{sa}/financials/ratios/", "Historical P/B, P/E, ROE and shareholder yield"),
    (8, "StockAnalysis company profile", f"{sa}/company/", "Sector, management, employees and recent foreign-issuer filings"),
    (9, "UBS investor relations", "https://www.ubs.com/global/en/investor-relations.html", "Primary results, presentations, capital and integration updates"),
    (10, "Swiss FINMA", "https://www.finma.ch/en/", "Capital regulation and supervisory developments"),
    (11, "Swiss Federal Council", "https://www.admin.ch/gov/en/start.html", "Too-big-to-fail policy and proposed bank-capital changes"),
]
for row, values in enumerate(sources, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
widths(ws, [7, 38, 108, 72])

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = sheet.dimensions

wb.save(OUT)
check = load_workbook(OUT, data_only=False)
expected = ["Valuation", "WACC", "Scenarios", "Actuals Source Audit", "Questions", "Sources"]
assert check.sheetnames == expected, check.sheetnames
assert isclose(check["Scenarios"]["B16"].value, weighted_value, rel_tol=1e-12)
print(f"Cost of equity: {cost_equity:.2f}%")
for name, case in SCENARIOS.items():
    print(f"{name}: ${case['target']:.2f} ({case['upside']:.1%})")
print(f"Probability-weighted fair value: ${weighted_value:.2f} ({weighted_value / PRICE - 1:.1%})")
print(f"Wrote and verified {OUT}")
