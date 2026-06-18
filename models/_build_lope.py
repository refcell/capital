"""
Build LOPE (Grand Canyon Education, Inc.) 6-sheet valuation model.
Data sourced from Yahoo Finance, June 17-18, 2026.
All financial figures in thousands unless noted.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

wb = openpyxl.Workbook()

# Styles
title_font = Font(name='Calibri', bold=True, size=16, color='1F4E79')
header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1F4E79')
data_font = Font(name='Calibri', size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

# ═══════════════════════════════════════════════════════
# SHEET 1: Valuation
# ═══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Valuation'
ws1.merge_cells('A1:F1')
ws1['A1'].font = title_font
ws1['A1'].value = 'Grand Canyon Education, Inc. (LOPE) — Valuation Model'

details = [
    ('Ticker', 'NASDAQ: LOPE'),
    ('Company', 'Grand Canyon Education, Inc.'),
    ('Sector', 'Consumer Discretionary / Online Higher Education'),
    ('Date', '2026-06-18'),
    ('Price', '$142.46 (Jun 17, 2026 close)'),
    ('Diluted Shares', '27.6M'),
    ('Market Cap', '$3.78B'),
    ('Enterprise Value', '$3.47B (net cash adjustment)'),
    ('Primary Valuation Lens', 'P/E, EV/EBITDA, P/FCF, scenario DCF'),
    ('Stance', 'WATCH — depressed by regulatory overhang; P/E 16.7x'),
]
for i, (label, val) in enumerate(details):
    r = i + 2
    ws1.cell(r, 1, label).font = Font(bold=True, size=11)
    ws1.cell(r, 2, val).font = data_font
    for c in (1, 2):
        ws1.cell(r, c).border = thin_border

r = len(details) + 4
ws1.cell(r, 1, 'Key Valuation Metrics').font = Font(bold=True, size=14, color='1F4E79')
r += 1
for c, h in enumerate(['Metric', 'Value', 'Notes'], 1):
    ws1.cell(r, c, h)
style_header(ws1, r, 3)

vm = [
    ('Trailing P/E', '16.7x', '$142.46 / $7.99 TTM EPS'),
    ('Forward P/E (est.)', '17.5x', 'Based on ~$8.13 FY26 EPS estimate'),
    ('P/Sales (TTM)', '3.35x', '$3.78B MC / $1.13B revenue'),
    ('P/FCF (TTM)', '13.9x', '$3.78B / $273M FCF (OCF - Capex est.)'),
    ('EV/EBITDA (TTM)', '9.9x', '$3.47B EV / $352M EBITDA'),
    ('EV/Sales (TTM)', '3.1x', '$3.47B / $1.13B'),
    ('Price-to-Book', '5.04x', '$3.78B / $747M equity'),
    ('Dividend Yield', 'N/A', 'No dividend; buybacks instead'),
]
for i, (m, v, n) in enumerate(vm):
    r += 1
    ws1.cell(r, 1, m).font = Font(bold=True, size=11)
    ws1.cell(r, 2, v).font = data_font
    ws1.cell(r, 3, n).font = Font(size=10, italic=True, color='666666')
    for c in (1, 2, 3):
        ws1.cell(r, c).border = thin_border

ws1.column_dimensions['A'].width = 28
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 50

# ═══════════════════════════════════════════════════════
# SHEET 2: WACC
# ═══════════════════════════════════════════════════════
ws2 = wb.create_sheet('WACC')
ws2.merge_cells('A1:E1')
ws2['A1'].font = title_font
ws2['A1'].value = 'WACC — Weighted Average Cost of Capital'

r = 3
for c, h in enumerate(['Component', 'Value', 'Formula / Notes'], 1):
    ws2.cell(r, c, h)
style_header(ws2, r, 3)

wacc = [
    ('10Y US Treasury (risk-free)', '4.15%', 'FRED DGS10 reference, Jun 2026'),
    ('Equity Risk Premium', '5.0%', 'Standard market assumption'),
    ('Beta (levered)', '0.90', 'EdTech sector; stable revenue/cash flow'),
    ('Cost of Equity', '8.65%', '4.15% + 0.90 * 5.0%'),
    ('Cost of Debt', 'N/A ~0%', 'No material interest-bearing debt'),
    ('Tax Rate (effective)', '23.4%', 'TTM: $67,030 / $286,930 pretax'),
    ('Market Cap ($M)', '~3,777', '27.6M shares * $142.46'),
    ('Total Debt ($M)', '~0', 'No significant debt'),
    ('Cash ($M)', '~311', 'Investing CF + BS estimates'),
    ('Equity Weight', '~100%', 'All-equity capital structure'),
    ('Debt Weight', '~0%', 'Negligible'),
    ('WACC', '8.65%', 'All-equity; WACC = Cost of Equity'),
]
for i, (item, val, note) in enumerate(wacc):
    r += 1
    ws2.cell(r, 1, item).font = Font(bold=True, size=11)
    ws2.cell(r, 2, val).font = data_font
    ws2.cell(r, 3, note).font = Font(size=10, italic=True, color='666666')
    for c in (1, 2, 3):
        ws2.cell(r, c).border = thin_border
        ws2.cell(r, c).alignment = Alignment(wrap_text=True)

ws2.column_dimensions['A'].width = 32
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 55

# ═══════════════════════════════════════════════════════
# SHEET 3: Scenarios
# ═══════════════════════════════════════════════════════
ws3 = wb.create_sheet('Scenarios')
ws3.merge_cells('A1:E1')
ws3['A1'].font = title_font
ws3['A1'].value = 'LOPE Scenario Analysis — Bear / Base / Bull'

base_rev = 1125.5  # $M TTM
shares = 25.5  # est 5Y shares after buybacks
nc = 311  # net cash

bear_tp = (base_rev * 1.04**5 * 0.20 * 12 + nc) / shares
base_tp = (base_rev * 1.07**5 * 0.24 * 16 + nc) / shares
bull_tp = (base_rev * 1.10**5 * 0.28 * 20 + nc) / shares
prob_wtd = bear_tp * 0.20 + base_tp * 0.50 + bull_tp * 0.30

r = 3
for c, h in enumerate(['Assumption', 'Bear', 'Base', 'Bull', 'Notes'], 1):
    ws3.cell(r, c, h)
style_header(ws3, r, 5)

sc = [
    ('Revenue CAGR (5Y)', '4.0%', '7.0%', '10.0%', 'Regulatory overhang (bear) to strong expansion'),
    ('Terminal Revenue ($M)', f'{base_rev*1.04**5:.0f}', f'{base_rev*1.07**5:.0f}', f'{base_rev*1.10**5:.0f}', f'From ${base_rev:.0f}M base'),
    ('Adjusted FCF Margin', '20.0%', '24.0%', '28.0%', 'TTM ~24%; bear adds reg/capex drag'),
    ('Terminal FCF ($M)', f'{base_rev*1.04**5*0.20:.0f}', f'{base_rev*1.07**5*0.24:.0f}', f'{base_rev*1.10**5*0.28:.0f}', 'Rev x margin'),
    ('Exit FCF Multiple', '12x', '16x', '20x', 'Current ~13.5x; range around consensus'),
    ('Implied EV ($M)', f'{base_rev*1.04**5*0.20*12:.0f}', f'{base_rev*1.07**5*0.24*16:.0f}', f'{base_rev*1.10**5*0.28*20:.0f}', 'FCF x multiple'),
    ('Net Cash Adj.', f'+${nc}', f'+${nc}', f'+${nc}', 'LOPE ~$311M net cash adds to equity'),
    ('Shares (est.)', f'{shares}M', f'{shares}M', f'{shares}M', 'Buyback decline from 27.6M'),
    ('Target Price', f'${bear_tp:.0f}', f'${base_tp:.0f}', f'${bull_tp:.0f}', '(EV+NC)/shares'),
    ('Upside from $142.46', f'{(bear_tp/142.46-1):.0%}', f'{(base_tp/142.46-1):.0%}', f'{(bull_tp/142.46-1):.0%}', 'Target/current - 1'),
    ('Scenario Weight', '20%', '50%', '30%', 'Base case most likely'),
    ('Weighted Value/Share', f'${bear_tp*0.20:.0f}', f'${base_tp*0.50:.0f}', f'${bull_tp*0.30:.0f}', 'Target x weight'),
    ('TOTAL Probability-Weighted FV', f'${prob_wtd:.0f}', '', '', 'Sum of weighted'),
    ('Implied Upside from Current', f'{(prob_wtd/142.46-1):.0%}', '', '', f'FV ${prob_wtd:.0f} vs $142.46'),
]
for i, (label, b, ba, bl, note) in enumerate(sc):
    r += 1
    ws3.cell(r, 1, label).font = Font(bold=True, size=11)
    ws3.cell(r, 2, b).font = data_font
    ws3.cell(r, 3, ba).font = data_font
    ws3.cell(r, 4, bl).font = data_font
    ws3.cell(r, 5, note).font = Font(size=10, italic=True, color='666666')
    for c in range(1, 6):
        ws3.cell(r, c).border = thin_border

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 15
ws3.column_dimensions['E'].width = 50

# ═══════════════════════════════════════════════════════
# SHEET 4: Actuals Source Audit
# ═══════════════════════════════════════════════════════
ws4 = wb.create_sheet('Actuals Source Audit')
ws4.merge_cells('A1:E1')
ws4['A1'].font = title_font
ws4['A1'].value = 'LOPE — Actuals Source Audit Trail'

r = 3
for c, h in enumerate(['Data Point', 'Value', 'Source', 'Date/Period', 'Notes'], 1):
    ws4.cell(r, c, h)
style_header(ws4, r, 5)

audit = [
    ('Stock Price', '$142.46', 'Yahoo Finance', '2026-06-17 close', 'NasdaqGS real-time'),
    ('Market Cap', '$3.777B', 'Yahoo Finance', '2026-06-17', '27.6M shares x $142.46'),
    ('Diluted Shares', '27,624K', 'Yahoo Finance IS', 'TTM', 'Diluted avg shares'),
    ('Revenue TTM', '$1,125,520K', 'Yahoo Finance IS', 'TTM', '+1.8% vs FY25'),
    ('Revenue FY25', '$1,106,070K', 'Yahoo Finance IS', '12/31/2025', '+7.1% YoY'),
    ('Revenue FY24', '$1,033,002K', 'Yahoo Finance IS', '12/31/2024', '+7.5% YoY'),
    ('Revenue FY23', '$960,899K', 'Yahoo Finance IS', '12/31/2023', '+5.4% YoY'),
    ('Revenue FY22', '$911,306K', 'Yahoo Finance IS', '12/31/2022', 'Baseline'),
    ('Gross Profit TTM', '$599,409K', 'Yahoo Finance IS', 'TTM', 'GM ~53.3%'),
    ('Operating Income TTM', '$310,760K', 'Yahoo Finance IS', 'TTM', 'OM ~27.6%'),
    ('Net Income TTM', '$219,900K', 'Yahoo Finance IS', 'TTM', 'NIM ~19.5%'),
    ('Diluted EPS TTM', '$7.99', 'Yahoo Finance IS', 'TTM', ''),
    ('EBITDA TTM', '$351,554K', 'Yahoo Finance IS', 'TTM', '+$41M depreciation'),
    ('Operating CF TTM', '$294,068K', 'Yahoo Finance CF', 'TTM', 'Good conversion'),
    ('Investing CF TTM', '-$27,625K', 'Yahoo Finance CF', 'TTM', 'Light capex'),
    ('Financing CF TTM', '-$314,807K', 'Yahoo Finance CF', 'TTM', 'Buyback-heavy'),
    ('Total Assets', '$992,305K', 'Yahoo Finance BS', '12/31/2025', 'Down from $1.018B'),
    ('Total Equity', '$746,933K', 'Yahoo Finance BS', '12/31/2025', 'Healthy'),
    ('Total Liabilities', '$245,372K', 'Yahoo Finance BS', '12/31/2025', 'Low leverage'),
    ('Tax Provision TTM', '$67,030K', 'Yahoo Finance IS', 'TTM', 'ERT ~23.4%'),
    ('Interest Income', '$13,581K', 'Yahoo Finance IS', 'TTM', 'Cash earnings'),
    ('Interest Expense', '~$0', 'Yahoo Finance IS', 'TTM', 'No debt cost'),
    ('52-Wk High', '$223.04', 'Yahoo Finance', 'Range', 'Pre-selloff peak'),
    ('52-Wk Low', '$140.94', 'Yahoo Finance', 'Range', 'Near current'),
    ('Analyst EPS Est Q1 FY26', '$2.78 est, $2.86 act', 'Yahoo Analysis', '2026-06', 'Beat 3%'),
]
for i, (dp, val, src, dt, note) in enumerate(audit):
    r += 1
    ws4.cell(r, 1, dp).font = Font(bold=True, size=11)
    ws4.cell(r, 2, val).font = data_font
    ws4.cell(r, 3, src).font = data_font
    ws4.cell(r, 4, dt).font = data_font
    ws4.cell(r, 5, note).font = Font(size=10, italic=True, color='666666')
    for c in range(1, 6):
        ws4.cell(r, c).border = thin_border
        ws4.cell(r, c).alignment = Alignment(wrap_text=True)

ws4.column_dimensions['A'].width = 25
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 22
ws4.column_dimensions['D'].width = 16
ws4.column_dimensions['E'].width = 45

# ═══════════════════════════════════════════════════════
# SHEET 5: Questions
# ═══════════════════════════════════════════════════════
ws5 = wb.create_sheet('Questions')
ws5.merge_cells('A1:D1')
ws5['A1'].font = title_font
ws5['A1'].value = 'LOPE — Open Questions & Research Gaps'

r = 3
for c, h in enumerate(['#', 'Question', 'Priority', 'Resolution Source'], 1):
    ws5.cell(r, c, h)
style_header(ws5, r, 4)

questions = [
    ('1', 'Regulatory risk trajectory: What is the specific gainsharing/bundled payment legislative risk over 12-24 months? Is the selloff driven by concrete bills or fear?', 'Critical', 'Congressional activity, GAO reports'),
    ('2', 'Gainsharing exposure: How much revenue is exposed to gainsharing reforms? Market prices this as existential.', 'Critical', 'Earnings transcripts, 10-K'),
    ('3', 'Buyback sustainability: $265M FY25, $315M TTM in buybacks. Sustainable while growing?', 'High', 'SEC filings, investor days'),
    ('4', 'Revenue growth durability: 7% YoY for years. Realistic long-term rate for online higher ed?', 'High', 'Historical trends, demographics'),
    ('5', 'Multiple rerating potential: If regulatory headwinds ease, can LOPE re-rate from 16x to 20-25x P/E?', 'High', 'Peer comps, historical multiples'),
    ('6', 'Goodwill/intangibles: Any significant goodwill from acquisitions?', 'Medium', '10-K balance sheet notes'),
    ('7', 'Student loan forgiveness impact: How would cancellation affect enrollment demand?', 'Medium', 'Academic research'),
    ('8', 'Competitive moat: What prevents other accredited programs from competing with Grand Canyon?', 'Medium', 'Peer analysis'),
    ('9', 'Accreditation risk: Any changes in accreditation standards that could affect LOPE?', 'Medium', 'Accreditation body publications'),
    ('10', 'Next earnings date and estimate revision trend?', 'Low', 'Yahoo Finance calendar'),
]
for i, (num, q, pri, src) in enumerate(questions):
    r += 1
    ws5.cell(r, 1, num).font = data_font
    ws5.cell(r, 2, q).font = Font(size=10)
    ws5.cell(r, 3, pri).font = data_font
    ws5.cell(r, 4, src).font = Font(size=10, italic=True, color='666666')
    for c in range(1, 5):
        ws5.cell(r, c).border = thin_border
        ws5.cell(r, c).alignment = Alignment(wrap_text=True)

ws5.column_dimensions['A'].width = 4
ws5.column_dimensions['B'].width = 70
ws5.column_dimensions['C'].width = 12
ws5.column_dimensions['D'].width = 30

# ═══════════════════════════════════════════════════════
# SHEET 6: Sources
# ═══════════════════════════════════════════════════════
ws6 = wb.create_sheet('Sources')
ws6.merge_cells('A1:C1')
ws6['A1'].font = title_font
ws6['A1'].value = 'LOPE — Data Sources'

r = 3
for c, h in enumerate(['#', 'Source', 'URL / Detail'], 1):
    ws6.cell(r, c, h)
style_header(ws6, r, 3)

sources = [
    ('1', 'Yahoo Finance - LOPE Income Statement', 'https://finance.yahoo.com/quote/LOPE/financials/'),
    ('2', 'Yahoo Finance - LOPE Balance Sheet', 'https://finance.yahoo.com/quote/LOPE/balance-sheet/'),
    ('3', 'Yahoo Finance - LOPE Cash Flow', 'https://finance.yahoo.com/quote/LOPE/cash-flow/'),
    ('4', 'Yahoo Finance - LOPE Analysis/Estimates', 'https://finance.yahoo.com/quote/LOPE/analysis/'),
    ('5', 'Yahoo Finance - LOPE Summary/Quote', 'https://finance.yahoo.com/quote/LOPE/'),
    ('6', '10Y Treasury Rate (FRED DGS10)', '~4.15% as of Jun 2026'),
    ('7', 'StockAnalysis - LOPE (404 - unavailable)', 'N/A'),
]
for i, (num, name, url) in enumerate(sources):
    r += 1
    ws6.cell(r, 1, num).font = data_font
    ws6.cell(r, 2, name).font = data_font
    ws6.cell(r, 3, url).font = Font(size=10, color='0563C1')
    for c in range(1, 4):
        ws6.cell(r, c).border = thin_border

ws6.column_dimensions['A'].width = 4
ws6.column_dimensions['B'].width = 50
ws6.column_dimensions['C'].width = 60

# ─── Save ───
filename = '[2026-06-18] Grand Canyon Education Model.xlsx'
wb.save(filename)
print(f'Saved: {filename}')

# Verify
wb2 = openpyxl.load_workbook(filename)
print(f'Sheets: {wb2.sheetnames}')
for sn in wb2.sheetnames:
    ws = wb2[sn]
    print(f'  {sn}: {ws.max_row} rows x {ws.max_column} cols')
