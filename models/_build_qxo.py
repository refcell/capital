#!/usr/bin/env python3
"""Build QXO, Inc. 6-sheet valuation model."""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ── Helpers ──────────────────────────────────────────────
TITLE_FONT = Font(name='Calibri', bold=True, size=14)
HEADER_FONT = Font(name='Calibri', bold=True, size=10)
SUBTITLE_FONT = Font(name='Calibri', bold=True, size=11)
BODY_FONT = Font(name='Calibri', size=10)
BOLD_FONT = Font(name='Calibri', bold=True, size=10)
BOLD_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='medium')
)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
HEADER_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

def c(ws, row, col, val, font=None, border=None, num_fmt=None, fill=None):
    cell = ws.cell(row=row, column=col, value=val)
    if font: cell.font = font
    if border: cell.border = border
    if num_fmt: cell.number_format = num_fmt
    if fill: cell.fill = fill
    return cell

def title_block(ws, company, ticker, date, price, mc, ev, lens, stance):
    ws.merge_cells('A1:G1')
    t = ws['A1']
    t.value = f"{company} ({ticker}) — Valuation Model"
    t.font = TITLE_FONT
    t.alignment = Alignment(horizontal='center')

    data = [
        ["Date:", date],
        ["Ticker:", f"NYSE: {ticker}"],
        ["Closing Price:", f"${price:.2f}"],
        ["Shares Outstanding (M):", "674.5"],
        ["Market Cap ($B):", f"${mc:.2f}"],
        ["Enterprise Value ($B):", f"${ev:.2f}"],
        ["Primary Valuation Lens:", lens],
        ["Current Stance:", stance],
    ]
    for i, (k, v) in enumerate(data):
        r = i + 3
        c(ws, r, 1, k, BOLD_FONT)
        c(ws, r, 2, v, BODY_FONT)

def write_table(ws, start_row, headers, rows, col_widths=None):
    for j, h in enumerate(headers, 1):
        cell = c(ws, start_row, j, h, HEADER_FONT, THIN_BORDER, fill=HEADER_FILL)
    for i, row_data in enumerate(rows, start_row + 1):
        for j, val in enumerate(row_data, 1):
            c(ws, i, j, val, BODY_FONT, THIN_BORDER)
    if col_widths:
        for j, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w

# ── KEY DATA (all in millions unless noted) ──────────────
price = 14.99
shares_mm = 674.55  # from BS 674,547 shares in thousands = 674.55M
mc = 15.56  # $B from Yahoo stats page
# net debt from BS: total_debt 3,914.5M, cash ~3,218.8M, net debt ~695.7M
# but we use EV-MC as cleaner proxy per skill guidance
net_debt_mm = 696  # from BS directly
ev = 16.26  # MC + net debt
# TTM financials (from Yahoo in thousands → millions)
ttm_revenue = 8558.9
ttm_gross_profit = 1976.6
ttm_operating_income = -457.8
ttm_ebitda = 182.3
ttm_net_income = -632.1
ttm_ocf = 295.5
ttm_capex = -6.7  # investing cash flow dominated by acquisitions; capex tiny
ttm_fcf = 295.5  # OCF - minimal capex
ttm_eps = -0.95
# FY2025
fy25_revenue = 6842.2
fy25_gross_profit = 1572.7
fy25_operating_income = -245.2
fy25_ebitda = 259.5
fy25_net_income = -388.3
# Analyst estimates (consensus Yahoo)
fy26_revenue_est = 13660.0  # $B → but wait: 13.66B
fy26_eps_est = 0.32
fy27_revenue_est = 18990.0
fy27_eps_est = 0.74
beta = 2.20
risk_free = 4.56  # from CNBC US10Y
tax_rate = 0.21  # standard corporate
# Cost of debt: interest expense 230M / debt 3,914.5M = 5.88%
cost_of_debt_raw = 0.0588

# ── Sheet 1: Valuation ──────────────────────────────────
ws1 = wb.active
ws1.title = "Valuation"
title_block(ws1, "QXO, Inc.", "QXO", "2026-07-10", price, mc, ev,
            "Forward P/E on FY26/FY27 consensus; P/S and EV/Revenue cross-check",
            "Watch — aggressive M&A growth in industrial distribution, negative TTM earnings, integration unproven")

# Valuation metrics table
headers = ["Metric", "Value", "Comment"]
rows = [
    ["P/E (TTM)", "N/A", f"TTM EPS negative at ${ttm_eps:.2f}. Cannot compute trailing P/E."],
    ["Forward P/E (FY26)", f"{price / fy26_eps_est:.1f}x", f"Using consensus EPS of ${fy26_eps_est}. Low-single-digit earnings on $13.7B revenue."],
    ["Forward P/E (FY27)", f"{price / fy27_eps_est:.1f}x", f"Using consensus EPS of ${fy27_eps_est}. More reasonable for hospitality."],
    ["P/S (TTM, $B)", f"{mc / (ttm_revenue / 1000):.2f}x", "P/S compressed from 2026 highs. Reasonable for hotel/casino operator."],
    ["P/FCF (TTM)", f"{mc / (ttm_fcf / 1000):.1f}x", f"FCF of ${ttm_fcf:.0f}M. FCF-driven by OCF with minimal capex."],
    ["EV/FCF", f"{ev / (ttm_fcf / 1000):.1f}x", "EV-based FCF multiple includes $0.7B net debt adjustment."],
    ["EV/Sales", f"{ev / (ttm_revenue / 1000):.2f}x", "Enterprise value scale to revenue."],
    ["EV/Revenue", f"{ev / (ttm_revenue / 1000):.2f}x", f"TTM Revenue $8.56B. Industrial distributors typically trade 0.6-1.0x EV/Revenue."],
    ["Price-to-Book (Common)", f"{mc / (8650.1 / 1000):.1f}x", "Common equity $8.65B. Tangible book value is NEGATIVE at -$280M."],
]
write_table(ws1, 3, headers, rows, [22, 18, 70])

# ── Sheet 2: WACC ──────────────────────────────────────
ws2 = wb.create_sheet("WACC")
ws2.merge_cells('A1:B1')
ws2['A1'].value = "WACC / CAPM Components"
ws2['A1'].font = TITLE_FONT

wacc_data = [
    ["Risk-Free Rate (10Y US Treasury)", f"{risk_free:.2f}%"],
    ["Equity Risk Premium", "5.00%"],
    ["Beta (5Y Monthly)", f"{beta:.2f}"],
    ["Cost of Equity (Rf + Beta × ERP)", f"{risk_free / 100 + beta * 0.05:.2%}"],
    ["Cost of Debt (pre-tax)", f"{cost_of_debt_raw:.2%}"],
    ["Cost of Debt (after-tax)", f"{cost_of_debt_raw * (1 - tax_rate):.2%}"],
    ["Tax Rate", f"{tax_rate:.0%}"],
    ["Market Cap ($B)", f"${mc:.2f}"],
    ["Total Debt ($B)", "3.91"],
    ["Enterprise Value ($B)", f"${ev:.2f}"],
    ["Equity Weight (E/EV)", f"{mc / ev:.2%}"],
    ["Debt Weight (D/EV)", f"{(ev - mc) / ev:.2%}"],
    ["WACC", ""],
]

cost_equity = risk_free / 100 + beta * 0.05
cost_debt_at = cost_of_debt_raw * (1 - tax_rate)
eq_weight = mc / ev
debt_weight = (ev - mc) / ev
wacc = eq_weight * cost_equity + debt_weight * cost_debt_at
wacc_data[-1][1] = f"{wacc:.2%}"

for i, (k, v) in enumerate(wacc_data):
    c(ws2, i + 3, 1, k, BOLD_BORDER if i == len(wacc_data) - 1 else BOLD_FONT)
    c(ws2, i + 3, 2, v, BOLD_FONT if i == len(wacc_data) - 1 else BODY_FONT)

print(f"WACC: {wacc:.2%}")

# ── Sheet 3: Scenarios ──────────────────────────────────
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells('A1:K1')
ws3['A1'].value = "Scenario Analysis — Forward P/E Framework (Ind. Distributor Peer Set)"
ws3['A1'].font = TITLE_FONT

# NOTE: Using Forward P/E as primary due to transitional earnings (acquisition-driven scale-up)
# QXO is an industrial distributor formerly known as SilverSun Technologies (renamed June 2024)
# Peers: BXC, GWW, FAST, WSO, POOL, FERG, AIT, WCC, CNM
# FY27 is base for terminal valuations (most realistic year with positive earnings)

# Revenue CAGR assumptions bridge from FY25 to Year 5 post-FY27
# FY25 revenue: $6.84B → FY27: $18.99B (consensus) then moderate growth
bear_rev_cagr = 0.12   # 12% — post-consensus deceleration
base_rev_cagr = 0.15   # 15% — consensus trajectory maintained
bull_rev_cagr = 0.18   # 18% — market share gains, expansion

# 5-year terminal revenue from FY27 base
yr5_revenue_bear = 18990 * (1 + bear_rev_cagr) ** 3  # 3 years post FY27
yr5_revenue_base = 18990 * (1 + base_rev_cagr) ** 3
yr5_revenue_bull = 18990 * (1 + bull_rev_cagr) ** 3

# Adjusted EPS (margin assumptions at terminal year)
# FY27 EPS $0.74 on $18.99B revenue = 0.0039 margin ≈ 0.39%
# Terminal margins can expand as integration completes
bear_eps = 0.40   # margin compression
base_eps = 0.85   # moderate improvement
bull_eps = 1.20   # full integration benefits

# Exit P/E multiples (industrial distributor peer range)
bear_exit_pe = 12
base_exit_pe = 16
bull_exit_pe = 20

# Target prices from P/E × EPS
bear_price = bear_eps * bear_exit_pe
base_price = base_eps * base_exit_pe
bull_price = bull_eps * bull_exit_pe

# Weights and probability
weights = [0.25, 0.50, 0.25]  # bear, base, bull
weighted_value = sum(w * p for w, p in zip(weights, [bear_price, base_price, bull_price]))

# Net debt per share (for adjustment)
nd_per_share = net_debt_mm / shares_mm  # dollars per share

scenarios_headers = ["", "Bear", "Base", "Bull"]
scenarios_data = [
    ["Revenue CAGR (5Y post-FY27)", f"{bear_rev_cagr:.0%}", f"{base_rev_cagr:.0%}", f"{bull_rev_cagr:.0%}"],
    ["Terminal Revenue ($B)", f"${yr5_revenue_bear:,.0f}", f"${yr5_revenue_base:,.0f}", f"${yr5_revenue_bull:,.0f}"],
    ["Terminal EPS", f"${bear_eps:.2f}", f"${base_eps:.2f}", f"${bull_eps:.2f}"],
    ["Exit P/E Multiple", f"{bear_exit_pe}x", f"{base_exit_pe}x", f"{bull_exit_pe}x"],
    ["Implied Value / Share", f"${bear_price:.2f}", f"${base_price:.2f}", f"${bull_price:.2f}"],
    ["Upside from Current", f"{(bear_price / price - 1):.0%}", f"{(base_price / price - 1):.0%}", f"{(bull_price / price - 1):.0%}"],
    ["Weight", f"{weights[0]:.0%}", f"{weights[1]:.0%}", f"{weights[2]:.0%}"],
    ["Weighted Value / Share", f"${weights[0]*bear_price:.2f}", f"${weights[1]*base_price:.2f}", f"${weights[2]*bull_price:.2f}"],
    ["", "", "", ""],
    ["Probability-Weighted FV / Share", "", "", f"${weighted_value:.2f}"],
    ["Upside from Current Price", "", "", f"{(weighted_value / price - 1):.0%}"],
]

write_table(ws3, 3, scenarios_headers, scenarios_data, [30, 16, 16, 16])

# Framework note
c(ws3, 15, 1, "Framework Note:", BOLD_FONT)
c(ws3, 16, 1,
    "Primary framework: Forward P/E on consensus EPS. QXO is an industrial distributor in transitional M&A-driven growth.",
    BODY_FONT)
c(ws3, 17, 1,
    "TTM earnings are negative (-$0.95 EPS) due to integration costs from acquisitions that grew revenue from $57M to $8.56B TTM.",
    BODY_FONT)
c(ws3, 18, 1,
    "Exit P/E multiples anchored to industrial distributor peers: BXC ~14x, POOL ~21x, FERG ~18x, FAST ~28x, GWW ~30x.",
    BODY_FONT)
c(ws3, 19, 1,
    "FY27 EPS of $0.74 is the consensus anchor. All terminal scenarios extrapolate from FY27, not TTM.",
    BODY_FONT)

print(f"Bear: ${bear_price:.2f}, Base: ${base_price:.2f}, Bull: ${bull_price:.2f}")
print(f"Weighted FV: ${weighted_value:.2f} (upside {(weighted_value / price - 1):.0%})")

# ── Sheet 4: Actuals Source Audit ──────────────────────
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells('A1:D1')
ws4['A1'].value = "Actuals Source Audit"
ws4['A1'].font = TITLE_FONT

audit_headers = ["Data Point", "Value", "Source URL", "Date / Notes"]
audit_data = [
    ["Stock Price (Close)", "$14.99", "finance.yahoo.com/quote/QXO/", "July 10, 2026"],
    ["Market Cap", "$15.56B", "finance.yahoo.com/quote/QXO/", "July 10, 2026 intraday"],
    ["Enterprise Value", "$16.26B", "Calculated: MC + Net Debt", "Net debt from BS"],
    ["Shares Outstanding", "674.55M", "finance.yahoo.com/quote/QXO/balance-sheet/", "BS 12/31/2025: 674,547 shares (thousands)"],
    ["Beta (5Y Monthly)", "2.20", "finance.yahoo.com/quote/QXO/", "Yahoo summary page"],
    ["52-Week Range", "$13.82 - $27.61", "finance.yahoo.com/quote/QXO/", "July 10, 2026"],
    ["TTM Revenue", "$8,558.9M", "finance.yahoo.com/quote/QXO/financials/", "TTM as of Q1 FY26"],
    ["FY2025 Revenue", "$6,842.2M", "finance.yahoo.com/quote/QXO/financials/", "Annual 12/31/2025"],
    ["FY2024 Revenue", "$56.9M", "finance.yahoo.com/quote/QXO/financials/", "Annual 12/31/2024 — pre-acquisition base"],
    ["TTM Gross Profit", "$1,976.6M", "finance.yahoo.com/quote/QXO/financials/", ""],
    ["TTM Operating Income", "-$457.8M", "finance.yahoo.com/quote/QXO/financials/", "Negative — integration costs"],
    ["TTM EBITDA", "$182.3M", "finance.yahoo.com/quote/QXO/financials/", "TTM"],
    ["TTM Net Income", "-$632.1M", "finance.yahoo.com/quote/QXO/financials/", "TTM"],
    ["TTM Diluted EPS", "-$0.95", "finance.yahoo.com/quote/QXO/financials/", "TTM"],
    ["TTM OCF", "$295.5M", "finance.yahoo.com/quote/QXO/cash-flow/", "TTM"],
    ["TTM FCF", "$295.5M", "Calculated: OCF - minimal capex", "Capex negligible post-acquisition"],
    ["Total Debt", "$3,914.5M", "finance.yahoo.com/quote/QXO/balance-sheet/", "12/31/2025"],
    ["Net Debt", "$695.7M", "finance.yahoo.com/quote/QXO/balance-sheet/", "12/31/2025"],
    ["Common Equity", "$8,650.1M", "finance.yahoo.com/quote/QXO/balance-sheet/", "12/31/2025"],
    ["Tangible Book Value", "-$280.3M", "finance.yahoo.com/quote/QXO/balance-sheet/", "12/31/2025 — NEGATIVE"],
    ["FY26 Revenue Estimate", "$13.66B", "finance.yahoo.com/quote/QXO/analysis/", "12 analysts, consensus avg"],
    ["FY26 EPS Estimate", "$0.32", "finance.yahoo.com/quote/QXO/analysis/", "12 analysts, non-GAAP"],
    ["FY27 Revenue Estimate", "$18.99B", "finance.yahoo.com/quote/QXO/analysis/", "12 analysts, consensus avg"],
    ["FY27 EPS Estimate", "$0.74", "finance.yahoo.com/quote/QXO/analysis/", "12 analysts, non-GAAP"],
    ["Q1 FY26 EPS (Actual)", "$0.11", "finance.yahoo.com/quote/QXO/analysis/", "Beat est. of $0.04 by 172%"],
    ["Q2 FY26 EPS (Est.)", "-$0.10", "finance.yahoo.com/quote/QXO/analysis/", "10 analysts"],
    ["Risk-Free Rate (10Y)", "4.56%", "cnbc.com/quotes/US10Y", "July 10, 2026"],
    ["Analyst Rev. Trend", "Mixed — 4 up, 4 down (30d on FY26 EPS)", "finance.yahoo.com/quote/QXO/analysis/", "EPS revisions heading down"],
]

write_table(ws4, 2, audit_headers, audit_data, [28, 18, 45, 35])

# ── Sheet 5: Questions ──────────────────────────────────
ws5 = wb.create_sheet("Questions")
ws5.merge_cells('A1:C1')
ws5['A1'].value = "Open Questions"
ws5['A1'].font = TITLE_FONT

questions = [
    ["1.", "QXO grew from $57M (FY2024) to $6.84B (FY2025) to $8.56B (TTM) via M&A. Which specific acquisitions drove this scale-up? The company distributes roofing, waterproofing, and building materials under GAF, CertainTeed, Owens Corning, LP SmartSide, James Hardie, and Velux. What is the acquisition integration status?"],
    ["2.", "Why is operating income negative ($-458M TTM) while OCF is positive ($296M)? What drives the $660M+ adjustment between OCF and operating income? Depreciation? Amortization? One-time charges?"],
    ["3.", "Tangible book value is NEGATIVE at -$280M on $15.9B in total assets. How much of the asset base is goodwill/intangibles from acquisition accounting?"],
    ["4.", "Total debt jumped to $3.91B (12/31/2025) from just $0.6M in FY2024. What instruments were issued? The BS shows $1.06B in preferred stock — what are the terms?"],
    ["5.", "Investing cash flow was -$10.65B TTM. Nearly all is acquisition spending. Are there follow-on acquisitions planned, or has the M&A cycle concluded?"],
    ["6.", "Share count exploded from 204M (FY2024) to 675M (current). This is ~230% dilution. Was this all stock consideration for acquisitions? What is the dilution trajectory?"],
    ["7.", "The stock's 52-week range ($13.82-$27.61) shows it has fallen ~45% from highs. Is this a cycle bottom or ongoing deterioration? What drove the decline?"],
    ["8.", "FY27 EPS estimate of $0.74 implies a P/E of 16x at current price, which is reasonable for distributors — but can QXO actually deliver normalized earnings at that rate post-integration?"],
    ["9.", "Revenue growth estimates of 68-72% QoQ in FY26 are enormous. What business model supports this? Is this driven by continued acquisitions or organic distributor operations?"],
    ["10.", "What is the next earnings date? Q1 FY26 reported $0.11 EPS (beat). Q2 FY26 estimate is NEGATIVE (-$0.10). What drives this whiplash?"],
    ["11.", "SBC (stock-based compensation) in a recently public, acquisition-heavy company is likely material. Is the positive OCF inflated by SBC add-backs?"],
    ["12.", "The company was formerly SilverSun Technologies Inc. (name changed June 2024). Why the name change and what does 'QXO' signify for the rebrand?"],
    ["13.", "With $1.06B in preferred stock on the balance sheet, what are the dividend obligations? Does this create a fixed-charge coverage risk?"],
    ["14.", "CEO Bradley Jacobs paid only $750K in FY2025 while CFO Essaid earned $910.5K. In an acquisition-driven company, is management compensation structured for retention or execution incentives?"],
]

for i, (num, q) in enumerate(questions):
    c(ws5, i + 3, 1, num, BOLD_FONT)
    c(ws5, i + 3, 2, q, BODY_FONT)
ws5.column_dimensions['A'].width = 6
ws5.column_dimensions['B'].width = 100

# ── Sheet 6: Sources ────────────────────────────────────
ws6 = wb.create_sheet("Sources")
ws6.merge_cells('A1:B1')
ws6['A1'].value = "Sources"
ws6['A1'].font = TITLE_FONT

sources = [
    ["1.", "Yahoo Finance — QXO Summary", "https://finance.yahoo.com/quote/QXO/"],
    ["2.", "Yahoo Finance — Income Statement", "https://finance.yahoo.com/quote/QXO/financials/"],
    ["3.", "Yahoo Finance — Balance Sheet", "https://finance.yahoo.com/quote/QXO/balance-sheet/"],
    ["4.", "Yahoo Finance — Cash Flow", "https://finance.yahoo.com/quote/QXO/cash-flow/"],
    ["5.", "Yahoo Finance — Analyst Estimates", "https://finance.yahoo.com/quote/QXO/analysis/"],
    ["6.", "Yahoo Finance — Company Profile", "https://finance.yahoo.com/quote/QXO/profile/"],
    ["7.", "CNBC — 10Y Treasury Yield", "https://www.cnbc.com/quotes/US10Y"],
    ["8.", "StockAnalysis.com — QXO (404 — unavailable)", "https://stockanalysis.com/quote/qxo/"],
    ["9.", "SEC EDGAR (not accessed — blocked in this environment)", "https://sec.gov/cgi-bin/browse-edgar"],
]

for i, (num, desc, url) in enumerate(sources):
    c(ws6, i + 3, 1, num, BOLD_FONT)
    c(ws6, i + 3, 2, desc, BODY_FONT)
    c(ws6, i + 3, 3, url, BODY_FONT)
ws6.column_dimensions['A'].width = 6
ws6.column_dimensions['B'].width = 50
ws6.column_dimensions['C'].width = 60

# ── Save ─────────────────────────────────────────────────
out = "/home/refcell/dev/capital/models/[2026-07-10] QXO Model.xlsx"
wb.save(out)

# ── Verify ───────────────────────────────────────────────
wb2 = openpyxl.load_workbook(out)
print(f"\nWorkbook saved to: {out}")
print(f"Sheets: {wb2.sheetnames}")
print(f"WACC: {wacc:.2%}")
print(f"Weighted FV: ${weighted_value:.2f}")
print(f"Target price sanity: Bear ${bear_price:.2f}, Base ${base_price:.2f}, Bull ${bull_price:.2f}")
print(f"Current price: ${price:.2f}")
assert all(3 < p < 50 for p in [bear_price, base_price, bull_price, weighted_value]), "Target prices out of range!"
print("All sanity checks passed.")
