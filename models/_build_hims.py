import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

wb = openpyxl.Workbook()

DARK_BG = "1F2937"
LIGHT_BG = "F3F4F6"
HEADER_BG = "374151"
ACCT_BLUE = "3B82F6"
GREEN = "10B981"
WHITE = "FFFFFF"


def sc(cell, bg=None, fg=WHITE, bold=False, size=11, align="center", fmt=None):
    if bg:
        cell.fill = PatternFill("solid", bg)
    if fg:
        cell.font = Font(color=fg, bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fmt:
        cell.number_format = fmt


# === Hims & Hers Health (HIMS) data ===
# StockAnalysis and company filings; USD millions unless noted.
revenue = [2370.00, 2348.00, 1477.00, 872.00, 526.92, 271.88]
revenue_growth = [32.81, 59.00, 69.33, 65.49, 93.81, 82.77]
gross_profit = [1699.00, 1733.00, 1173.00, 714.95, 408.72, 204.49]
gross_margin = [71.69, 73.84, 79.45, 81.99, 77.57, 75.22]
operating_income = [-30.61, 105.61, 61.90, -29.45, -68.70, -115.04]
operating_margin = [-1.29, 4.50, 4.19, -3.38, -13.04, -42.31]
net_income = [-13.24, 128.37, 126.04, -23.55, -65.68, -107.66]
net_margin = [-0.56, 5.47, 8.53, -2.70, -12.46, -39.60]
eps_diluted = [-0.09, 0.51, 0.53, -0.11, -0.32, -0.58]
ebitda = [37.57, 160.12, 78.99, -19.94, -61.22, -110.97]
ebitda_margin = [1.59, 6.82, 5.35, -2.29, -11.62, -40.81]
operating_cash_flow = [280.27, 300.01, 251.08, 73.48, -26.53, -34.41]
capex_abs = [200.56, 226.05, 41.66, 17.22, 2.71, 0.83]
free_cash_flow = [79.71, 73.96, 209.43, 56.26, -29.25, -35.24]
fcf_margin = [3.36, 3.15, 14.18, 6.45, -5.55, -12.96]
shares_outstanding = [230.70, 227.25, 220.84, 213.48, 208.43, 204.79]
diluted_shares = [245.00, 258.00, 237.00, 209.00, 205.00, 187.00]
stock_based_comp = [147.25, 135.24, 92.32, 66.08, 42.82, 67.21]

cash = [222.27, 228.62, 220.58, 96.66, 46.77, 71.78]
short_term_investments = [528.61, 348.88, 79.67, 124.32, 132.85, 175.49]
cash_and_st = [750.88, 577.49, 300.25, 220.98, 179.63, 247.27]
total_assets = [2267.00, 2155.00, 707.54, 891.71, 441.19, 347.70]
total_liabilities = [1821.00, 1614.00, 230.82, 342.45, 97.16, 54.65]
total_debt = [1132.00, 1121.00, 11.35, 63.38, 9.95, 0.00]
book_value_per_share = [1.95, 2.18, 1.98, 2.23, 1.55, 1.53]
net_debt = [381.12, 543.51, -288.90, -157.60, -169.68, -247.27]

price = 26.19
market_cap = 6060.00
enterprise_value = 6440.00
shares_current = 231.46
current_share_class = 223.08
float_shares = 210.90
beta = 2.40
risk_free_rate = 4.47
equity_risk_premium = 5.00
cost_of_equity = round(risk_free_rate + beta * equity_risk_premium, 2)
cost_of_debt = 4.50
normalized_tax_rate = 21.00
capital_base = market_cap + total_debt[0]
equity_weight = market_cap / capital_base
debt_weight = total_debt[0] / capital_base
wacc = equity_weight * cost_of_equity + debt_weight * cost_of_debt * (1 - normalized_tax_rate / 100)

fwd_sales_2026 = 2880.00
fwd_sales_2027 = 3380.00
fwd_eps_2026 = 0.93
fwd_eps_2027 = 1.29
pt_avg = 26.61
pt_low = 21.00
pt_high = 35.00

# Sheet 1: Valuation
ws = wb.active
ws.title = "Valuation"
ws.merge_cells("A1:G1")
ws["A1"].value = "Hims & Hers Health, Inc. (NYSE: HIMS) — Valuation"
sc(ws["A1"], bg=DARK_BG, size=14)
ws.row_dimensions[1].height = 36

info = [
    ("Date:", "2026-06-06"),
    ("Ticker:", "NYSE: HIMS"),
    ("Close Price:", "$26.19"),
    ("Shares Outstanding:", "231.46M"),
    ("Market Cap:", "$6.06B"),
    ("Enterprise Value:", "$6.44B"),
    ("Primary Lens:", "Scenario FCF, EV/sales, normalized cash conversion"),
    ("Stance:", "Watch / controversial consumer-health platform")
]
for i, (k, v) in enumerate(info, 2):
    ws.cell(i, 1, k)
    sc(ws.cell(i, 1), bold=True)
    ws.cell(i, 2, v)
    sc(ws.cell(i, 2), align="left")

r = 12
ws.merge_cells(f"A{r}:G{r}")
ws[f"A{r}"].value = "Key Valuation Metrics"
sc(ws[f"A{r}"], bg=ACCT_BLUE, size=12)

metrics = [
    ("P/S", "2.56x", "StockAnalysis statistics page at Jun. 5 close."),
    ("Forward P/S", "2.01x", "Provider forward-sales multiple; FY2026 forecast page implies ~2.10x."),
    ("EV/Sales", "2.72x", "Based on TTM revenue and provider EV."),
    ("EV / 2026E Sales", f"{enterprise_value / fwd_sales_2026:.2f}x", "Using 2026 revenue forecast of $2.88B."),
    ("P/FCF", "76.05x", "Current TTM FCF is too small to anchor valuation."),
    ("EV/FCF", "80.83x", "Highlights dependence on future cash conversion."),
    ("EV/EBITDA", "77.24x", "TTM EBITDA is depressed after Q1 2026 reset."),
    ("EV / 2026 Adj. EBITDA", f"{enterprise_value / 312.5:.1f}x", "Using midpoint of FY2026 adjusted EBITDA guide."),
    ("Forward EPS", "$0.93", "Financial forecast page; provider forward PE tables conflict."),
    ("Price Target Range", "$21 / $26.61 / $35", "Low / average / high S&P Global targets on StockAnalysis.")
]
for i, (k, v, n) in enumerate(metrics, 13):
    ws.cell(i, 1, k)
    sc(ws.cell(i, 1), bg=LIGHT_BG)
    ws.cell(i, 2, v)
    ws.cell(i, 3, n)
    sc(ws.cell(i, 3), align="left")

r = 25
ws.merge_cells(f"A{r}:G{r}")
ws[f"A{r}"].value = "Historical Financial Summary ($M)"
sc(ws[f"A{r}"], bg=GREEN, size=12)
r += 1
for ci, label in enumerate(["", "TTM", "FY2025", "FY2024", "FY2023", "FY2022", "FY2021"], 1):
    ws.cell(r, ci, label)
    sc(ws.cell(r, ci), bg=HEADER_BG)
r += 1
for label, vals in [
    ("Revenue", revenue),
    ("Revenue Growth (%)", revenue_growth),
    ("Gross Profit", gross_profit),
    ("Gross Margin (%)", gross_margin),
    ("Operating Income", operating_income),
    ("Operating Margin (%)", operating_margin),
    ("EBITDA", ebitda),
    ("EBITDA Margin (%)", ebitda_margin),
    ("Net Income", net_income),
    ("Net Margin (%)", net_margin),
    ("EPS Diluted", eps_diluted),
    ("Operating Cash Flow", operating_cash_flow),
    ("Capex", capex_abs),
    ("Free Cash Flow", free_cash_flow),
    ("FCF Margin (%)", fcf_margin),
    ("Shares Outstanding (M)", shares_outstanding),
    ("Diluted Shares (M)", diluted_shares),
    ("Stock-Based Comp.", stock_based_comp),
    ("Cash & ST Inv.", cash_and_st),
    ("Total Debt", total_debt),
    ("Net Debt", net_debt),
]:
    ws.cell(r, 1, label)
    sc(ws.cell(r, 1), bold=True, align="left")
    for j, v in enumerate(vals, 2):
        ws.cell(r, j, round(v, 2) if isinstance(v, float) else v)
        sc(ws.cell(r, j))
    r += 1
for col in range(1, 8):
    ws.column_dimensions[get_column_letter(col)].width = 16
ws.column_dimensions["A"].width = 24

# Sheet 2: WACC
ws2 = wb.create_sheet("WACC")
ws2.merge_cells("A1:C1")
ws2["A1"].value = "Hims & Hers Health (HIMS) — WACC Calculation"
sc(ws2["A1"], bg=DARK_BG, size=14)
wacc_items = [
    ("Component", "Value", "Rationale"),
    ("Risk-free Rate (10y US)", f"{risk_free_rate:.2f}%", "FRED DGS10 observation for 2026-06-04."),
    ("Equity Risk Premium", f"{equity_risk_premium:.2f}%", "Standard public-equity ERP."),
    ("Beta (Levered)", f"{beta:.2f}", "StockAnalysis statistics page."),
    ("Cost of Equity (CAPM)", f"{cost_of_equity:.2f}%", "Risk-free plus beta times ERP."),
    ("Cost of Debt", f"{cost_of_debt:.2f}%", "Normalized debt cost despite 0% converts."),
    ("Tax Rate", f"{normalized_tax_rate:.2f}%", "Normalized U.S. statutory cash-tax assumption."),
    ("Market Cap ($M)", f"{market_cap:,.0f}", "Jun. 5, 2026 market cap."),
    ("Total Debt ($M)", f"{total_debt[0]:,.0f}", "TTM/Mar. 31, 2026 debt from StockAnalysis."),
    ("Equity Weight", f"{equity_weight * 100:.1f}%", "Equity / (equity + debt)."),
    ("Debt Weight", f"{debt_weight * 100:.1f}%", "Debt / (equity + debt)."),
    ("WACC", f"{wacc:.1f}%", "High discount rate reflects elevated beta and controversy."),
]
for i, (k, v, n) in enumerate(wacc_items, 3):
    ws2.cell(i, 1, k)
    ws2.cell(i, 2, v)
    ws2.cell(i, 3, n)
    sc(ws2.cell(i, 1), bold=True, align="left")
    sc(ws2.cell(i, 2))
    sc(ws2.cell(i, 3), align="left")
ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 46

# Sheet 3: Scenarios
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells("A1:D1")
ws3["A1"].value = "Hims & Hers Health (HIMS) — Scenario Analysis"
sc(ws3["A1"], bg=DARK_BG, size=14)
for j, h in enumerate(["Metric", "Bear", "Base", "Bull"], 1):
    ws3.cell(3, j, h)
    sc(ws3.cell(3, j), bg=HEADER_BG)

scenario_rows = [
    ("Revenue CAGR (2025-2030)", "8%", "15%", "22%"),
    ("Terminal Revenue ($M)", "3,449", "4,722", "6,330"),
    ("Terminal FCF Margin", "5.0%", "10.5%", "15.5%"),
    ("Terminal FCF ($M)", "172", "496", "981"),
    ("Exit Multiple (FCF)", "18x", "24x", "30x"),
    ("Implied EV ($M)", "3,104", "11,904", "29,430"),
    ("Less: Net Debt Adj. ($M)", "381", "381", "381"),
    ("Implied Equity Value ($M)", "2,723", "11,523", "29,049"),
    ("Diluted Shares (M)", "244", "250", "258"),
    ("Terminal Value / Share", "$11.16", "$46.09", "$112.59"),
    ("Return vs. $26.19", "-57%", "+76%", "+330%"),
    ("Scenario Weight", "25%", "50%", "25%"),
    ("Weighted Value / Share", "$2.79", "$23.05", "$28.15"),
    ("", "Total Probability-Weighted FV", "", "$54.00"),
    ("", "Upside from Current Price", "", "+106%"),
]
for i, row in enumerate(scenario_rows, 4):
    for j, v in enumerate(row, 1):
        ws3.cell(i, j, v)
        sc(ws3.cell(i, j), bg=LIGHT_BG if i % 2 == 0 else None)
    if "Total Probability-Weighted FV" in row or "Upside from Current Price" in row:
        for j in range(1, 5):
            sc(ws3.cell(i, j), bg=GREEN, bold=True)
ws3.column_dimensions["A"].width = 32
for col in ["B", "C", "D"]:
    ws3.column_dimensions[col].width = 22

# Sheet 4: Actuals Source Audit
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells("A1:G1")
ws4["A1"].value = "Hims & Hers Health (HIMS) — Data Sources & Audit"
sc(ws4["A1"], bg=DARK_BG, size=14)
for j, h in enumerate(["Line Item", "Period", "Model Value", "Unit", "Primary Source", "Secondary Source", "Difference / Resolution"], 1):
    ws4.cell(2, j, h)
    sc(ws4.cell(2, j), bg=HEADER_BG)

audit_rows = [
    ("Share price", "2026-06-05", 26.19, "USD/share", "StockAnalysis overview", "n/a", "Jun. 5 close."),
    ("Shares outstanding", "Current", 231.46, "M", "StockAnalysis overview/statistics", "n/a", "Ties across overview and statistics."),
    ("Market cap", "Current", 6060, "USD mm", "StockAnalysis statistics", "n/a", "Rounded from $6.06B."),
    ("Enterprise value", "Current", 6440, "USD mm", "StockAnalysis statistics", "n/a", "Rounded from $6.44B."),
    ("Revenue", "FY2025", 2348, "USD mm", "StockAnalysis financials", "research/HIMS.md prior refresh", "Ties."),
    ("Revenue", "Q1 2026", 608.1, "USD mm", "StockAnalysis quarterly income statement", "Company Q1 release", "Ties."),
    ("Gross margin", "Q1 2026", 65.25, "%", "StockAnalysis quarterly income statement", "Company Q1 release", "Ties to quarterly table."),
    ("Operating cash flow", "FY2025", 300.01, "USD mm", "StockAnalysis cash flow", "research/HIMS.md", "Ties."),
    ("Free cash flow", "TTM", 79.71, "USD mm", "StockAnalysis cash flow/statistics", "n/a", "Matches provider tables."),
    ("Cash & ST investments", "Mar. 31, 2026", 750.88, "USD mm", "StockAnalysis balance sheet", "Company Q1 filing", "Ties."),
    ("Total debt", "Mar. 31, 2026", 1132, "USD mm", "StockAnalysis balance sheet", "Company Q1 filing", "Includes convert debt."),
    ("2026 revenue forecast", "FY2026E", 2880, "USD mm", "StockAnalysis forecast", "n/a", "14 analyst forecast page."),
    ("2027 revenue forecast", "FY2027E", 3380, "USD mm", "StockAnalysis forecast", "n/a", "14 analyst forecast page."),
    ("Analyst target avg", "Current", 26.61, "USD/share", "StockAnalysis forecast", "n/a", "16 analyst S&P Global target panel."),
]
for i, row in enumerate(audit_rows, 3):
    for j, v in enumerate(row, 1):
        ws4.cell(i, j, v)
        sc(ws4.cell(i, j), align="left" if j != 3 else "center")
    sc(ws4.cell(i, 1), bold=True, align="left")
for col, width in zip(["A", "B", "C", "D", "E", "F", "G"], [22, 16, 14, 12, 28, 24, 34]):
    ws4.column_dimensions[col].width = width

# Sheet 5: Questions
ws5 = wb.create_sheet("Questions")
ws5.merge_cells("A1:C1")
ws5["A1"].value = "Hims & Hers Health (HIMS) — Open Questions"
sc(ws5["A1"], bg=DARK_BG, size=14)
questions = [
    "How much of 2026-2027 growth is truly non-GLP-1 and repeatable without regulatory edge cases?",
    "Does the Q1 gross-margin reset represent transitional branded-GLP-1 mix, or a new structural margin ceiling?",
    "How much dilution should be underwritten from SBC, 2030 converts, 2032 converts, and acquisition consideration?",
    "Can Eucalyptus integration improve international growth quality, or will it mostly add acquired revenue and complexity?",
    "What portion of free cash flow is durable after working-capital normalization, pharmacy investment, and software capex?",
    "How should investors treat 0% converts in valuation when the business still has volatile owner earnings?",
    "Will branded and generic GLP-1 access commoditize the customer relationship, limiting lifetime value?",
    "Are diagnostics and personalization genuine retention drivers, or mostly narrative support for broader TAM claims?",
    "Can HIMS sustain subscriber growth while monthly revenue per average subscriber remains under pressure?",
    "What evidence would prove that international and menopause/testosterone expansions improve mix rather than just revenue?",
    "Does controlled-company governance deserve a persistent discount even if operating results improve?",
    "What does normalized 2030 cash conversion look like if revenue scales but gross margin never returns above 70%?",
]
for i, q in enumerate(questions, 3):
    ws5.cell(i, 1, i - 2)
    ws5.cell(i, 2, q)
    sc(ws5.cell(i, 1))
    sc(ws5.cell(i, 2), align="left")
ws5.column_dimensions["A"].width = 6
ws5.column_dimensions["B"].width = 110

# Sheet 6: Sources
ws6 = wb.create_sheet("Sources")
ws6.merge_cells("A1:B1")
ws6["A1"].value = "Hims & Hers Health (HIMS) — Sources"
sc(ws6["A1"], bg=DARK_BG, size=14)
sources = [
    "StockAnalysis HIMS overview", "https://stockanalysis.com/stocks/hims/",
    "StockAnalysis HIMS financials", "https://stockanalysis.com/stocks/hims/financials/",
    "StockAnalysis HIMS balance sheet", "https://stockanalysis.com/stocks/hims/financials/balance-sheet/",
    "StockAnalysis HIMS cash flow", "https://stockanalysis.com/stocks/hims/financials/cash-flow-statement/",
    "StockAnalysis HIMS forecast", "https://stockanalysis.com/stocks/hims/forecast/",
    "StockAnalysis HIMS statistics", "https://stockanalysis.com/stocks/hims/statistics/",
    "FRED DGS10", "https://fred.stlouisfed.org/series/DGS10",
    "Hims & Hers Q1 2026 Form 10-Q", "https://www.sec.gov/Archives/edgar/data/1773751/000177375126000076/hims-20260331.htm",
    "Hims & Hers Q1 2026 earnings release", "https://www.sec.gov/Archives/edgar/data/1773751/000177375126000074/hims-20260331x8xkearningsr.htm",
    "Hims & Hers investor relations", "https://investors.hims.com/overview/default.aspx",
]
for i, s in enumerate(sources, 3):
    ws6.cell(i, 1, i - 2)
    ws6.cell(i, 2, s)
    sc(ws6.cell(i, 1))
    sc(ws6.cell(i, 2), align="left")
ws6.column_dimensions["A"].width = 6
ws6.column_dimensions["B"].width = 96

path = "/home/refcell/dev/capital/models/[2026-06-06] Hims & Hers Health Model.xlsx"
wb.save(path)

# Immediate verification after write.
check = load_workbook(path, data_only=False)
print("Saved:", path)
print("Sheets:", check.sheetnames)
for sheet in check.sheetnames:
    ws_check = check[sheet]
    print(sheet, ws_check.max_row, ws_check.max_column)
    print("A1:", ws_check["A1"].value)
