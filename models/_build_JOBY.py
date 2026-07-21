#!/usr/bin/env python3
"""Build 6-sheet valuation model for JOBY (Joby Aviation) — Pre-Commercial eVTOL Developer."""

import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Styles ──
bold = Font(bold=True)
title_font = Font(bold=True, size=14)
section_font = Font(bold=True, size=12)
header_font = Font(bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')

def c(ws, row, col, value, font=None, border=True, fill=None, alignment=None):
    """Write cell with formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if border:
        cell.border = thin_border
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    return cell

def write_header_row(ws, row, headers):
    for i, h in enumerate(headers, 1):
        cell = c(ws, row, i, h, font=header_font, fill=header_fill)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

# ════════════════════════════════════════════════════════════
# Sheet 1: Valuation
# ════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Valuation"
ws1.merge_cells('A1:F1')
c(ws1, 1, 1, "JOBY (Joby Aviation, Inc.) — Valuation Summary", font=title_font)

price = 7.47
shares_outstanding_mm = 983.64
market_cap_b = price * shares_outstanding_mm / 1000  # $7.35B
total_cash_b = 2.47
total_debt_b = 0.748  # from Key Stats: $747.73M
enterprise_value_b = market_cap_b + total_debt_b - total_cash_b
date_str = "2026-07-20"

title_data = [
    ("Company", "Joby Aviation, Inc."),
    ("Ticker", "NYSE:JOBY"),
    ("Date", date_str),
    ("Price (USD)", f"${price}"),
    ("Shares Outstanding (M)", f"{shares_outstanding_mm:.2f}M"),
    ("Market Cap (USD)", f"${market_cap_b:.2f}B"),
    ("Enterprise Value (USD)", f"${enterprise_value_b:.2f}B"),
    ("Primary Valuation Lens", "Cash NAV Floor + Regulatory/Pipeline Optionality"),
    ("Stance", "Watch"),
]

for i, (label, value) in enumerate(title_data, 2):
    c(ws1, i, 1, label, font=bold)
    c(ws1, i, 2, value)

# Valuation metrics table
ws1.merge_cells('A4:F4')
c(ws1, 4, 1, "Key Valuation Metrics", font=section_font)

metric_headers = ["Metric", "Value", "Comment"]
for i, h in enumerate(metric_headers, 1):
    c(ws1, 5, i, h, font=header_font, fill=header_fill)

metrics = [
    ("P/E (TTM)", "N/A (negative earnings)", "Loss-making — no trailing P/E"),
    ("Forward P/E", "N/A (EPS still negative FY26-27)", "Consensus EPS: -$0.67 FY26, -$0.50 FY27"),
    ("P/Sales (TTM)", f"{market_cap_b*1000/77.67:.1f}x", "Revenue $77.7M TTM"),
    ("P/Sales (Fwd)", f"{market_cap_b*1000/113.88:.1f}x", "Fwd rev $113.9M (consensus)"),
    ("P/FCF", "N/A (FCF deeply negative)", "FCF: -$380.6M TTM"),
    ("EV/Revenue", f"{enterprise_value_b*1000/77.67:.1f}x", "EV-based top-line multiple"),
    ("EV/EBITDA", "N/A (EBITDA negative)", "EBITDA: -$747.9M TTM"),
    ("Cash/Share NAV", f"${total_cash_b*1000/shares_outstanding_mm:.2f}", f"${total_cash_b:.2f}B cash / {shares_outstanding_mm:.0f}M shares"),
    ("Net Cash/Share", f"${(total_cash_b-total_debt_b)*1000/shares_outstanding_mm:.2f}", f"Net cash ${(total_cash_b-total_debt_b):.2f}B"),
    ("Book Value/Share", f"${1409.713*1000/shares_outstanding_mm:.2f}", "FY2025 equity $1,409.7M"),
    ("P/B", f"{price/(1409.713*1000/shares_outstanding_mm):.2f}x", "Price vs BVPS"),
]

for i, (metric, val, comment) in enumerate(metrics, 6):
    c(ws1, i, 1, metric)
    c(ws1, i, 2, val)
    c(ws1, i, 3, comment)

for ci in range(1, 6):
    ws1.column_dimensions[get_column_letter(ci)].width = 20
ws1.column_dimensions['C'].width = 45

# ════════════════════════════════════════════════════════════
# Sheet 2: WACC
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("WACC")
ws2.merge_cells('A1:D1')
c(ws2, 1, 1, "JOBY — WACC / CAPM Components", font=title_font)

risk_free = 4.598  # 10Y US Treasury, Jul 20 2026, CNBC
erp = 5.0
beta = 2.71  # 5Y Monthly
cost_of_equity = risk_free + beta * erp
cost_of_debt = 5.5  # estimated — minimal debt
tax_rate = 1.435 / (abs(-955.951) + 1.435) * 100 if abs(-955.951 + 1.435) > 0 else 0  # effective ≈ 0.15%
# Simplified: tax rate near 0 since tax provision is tiny ($1.4M on $956M loss)
tax_rate = 0.15  # approximate
equity_weight = 1.0  # effectively all-equity
debt_weight = 0.0

wacc = cost_of_equity * equity_weight + cost_of_debt * debt_weight * (1 - tax_rate)

wacc_data = [
    ("Risk-Free Rate (10Y US)", f"{risk_free:.3f}%", "CNBC Jul 20, 2026 close"),
    ("Equity Risk Premium", f"{erp:.1f}%", "Standard assumption"),
    ("Beta (5Y Monthly, Levered)", f"{beta:.2f}", "Yahoo Finance Key Stats"),
    ("Cost of Equity (Ke)", f"{cost_of_equity:.2f}%", f"Rf + β × ERP = {risk_free} + {beta}×{erp}"),
    ("Cost of Debt (Kd)", f"{cost_of_debt:.1f}%", "Estimated — minimal interest expense"),
    ("Tax Rate", f"{tax_rate*100:.1f}%", "Effective — tiny tax provision"),
    ("Market Cap (USD)", f"${market_cap_b:.2f}B"),
    ("Total Debt (USD)", f"${total_debt_b:.3f}B"),
    ("Equity Weight", f"{equity_weight:.2f}", "Effectively all-equity"),
    ("Debt Weight", f"{debt_weight:.2f}"),
    ("WACC", f"{wacc:.2f}%", f"Ke×We + Kd×Wd×(1-T) = {cost_of_equity:.2f}×1.0 + ..."),
]

for i, (label, value, *rest) in enumerate(wacc_data, 2):
    c(ws2, i, 1, label, font=bold)
    c(ws2, i, 2, value)
    if rest:
        c(ws2, i, 3, rest[0])

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 45

# ════════════════════════════════════════════════════════════
# Sheet 3: Scenarios
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells('A1:K1')
c(ws3, 1, 1, "JOBY — Pre-Commercial eVTOL Valuation Scenarios", font=title_font)
ws3.merge_cells('A2:K2')
c(ws3, 2, 1, "Framework: Cash NAV Floor + Regulatory/Pipeline Optionality (per skill guidance for pre-commercial technology developers)", font=Font(italic=True, size=10))

# Note about framework choice
ws3.merge_cells('A3:K3')
c(ws3, 3, 1, "STANDARD FCF/P/E FRAMEWORK INAPPLICABLE: JOBY is pre-commercial — FCF is structurally negative by design ($380.6M TTM burn). Revenue is $77.7M TTM and non-commercial. Primary valuation lens is cash-per-share NAV floor + optionality premium.", font=Font(italic=True, color='FF0000', size=10))

scenario_headers = ["Metric", "Bear", "Base", "Bull", "Notes"]
write_header_row(ws3, 5, scenario_headers)

# Scenario calculations
# Current state:
# Cash: $2.47B, Shares: 983.64M, Cash/Share: $2.51
# Burn rate TTM: ~$543M operating, -$380.6M levered FCF
# Quarterly burn ~$125-160M (based on TTM patterns)
# FY2025 burn: OCF -$510M, but investing -$475M (capex), financed by +$1,027M

# Scenarios model cash runway, dilution, and optionality:
# Bear: NRC certification delayed, extended runway burn, aggressive dilution
# Base: Certification on track ~2029-2030, moderate dilution
# Bull: Accelerated certification, FAA partnerships, commercial deployment

total_cash_mm = 2470  # $2.47B in millions
shares_mm = 983.64

# Bear: Burn accelerates, heavy dilution
bear_burn_annual_mm = 700   # $700M/year
bear_runway_years = total_cash_mm / bear_burn_annual_mm  # ~3.5 years
bear_dilution_factor = 2.0  # 2x dilution from future raises
bear_post_dilution_shares = shares_mm * bear_dilution_factor

# Base: Moderate burn, moderate dilution
base_burn_annual_mm = 550
base_runway_years = total_cash_mm / base_burn_annual_mm  # ~4.5 years
base_dilution_factor = 1.5
base_post_dilution_shares = shares_mm * base_dilution_factor

# Bull: Revenue offset burn, less dilution
bull_burn_annual_mm = 400
bull_runway_years = total_cash_mm / bull_burn_annual_mm  # ~6.2 years
bull_dilution_factor = 1.2
bull_post_dilution_shares = shares_mm * bull_dilution_factor

# Terminal values (Year 10 = ~2036)
# Bear: Cash depleted by ~year 4, heavy dilution, minimal optionality
# NAV floor = essentially $0 from cash, residual optionality
bear_nav_per_share = 0.50    # Residual optionality only
bear_pipeline_npv_per_share = 0  # Certification fails or very delayed

# Base: Partial cash preservation, moderate optionality
base_remaining_cash_at_exit_mm = 200  # Some cash left at exit
base_nav_per_share = base_remaining_cash_at_exit_mm / base_post_dilution_shares
base_pipeline_npv_per_share = 4.50  # Moderate certification success probability
base_total_per_share = base_nav_per_share + base_pipeline_npv_per_share

# Bull: Revenue generation offsets burn, certification success
bull_remaining_cash_at_exit_mm = 500
bull_nav_per_share = bull_remaining_cash_at_exit_mm / bull_post_dilution_shares
bull_pipeline_npv_per_share = 18.00  # Strong certification + commercial deployment
bull_total_per_share = bull_nav_per_share + bull_pipeline_npv_per_share

# Current optionality premium: price - cash/NAV per share
cash_per_share = total_cash_mm / shares_mm  # $2.51
current_optionality_premium = price - cash_per_share  # $7.47 - $2.51 = $4.96

scenarios = [
    ("Cash (Start, $M)", f"${total_cash_mm:.0f}", f"${total_cash_mm:.0f}", f"${total_cash_mm:.0f}", "MRQ Q1 2026"),
    ("Annual Burn Rate ($M)", f"${bear_burn_annual_mm:.0f}", f"${base_burn_annual_mm:.0f}", f"${bull_burn_annual_mm:.0f}", "Pre-commercial ops + regulatory"),
    ("Cash Runway (Years)", f"{bear_runway_years:.1f}", f"{base_runway_years:.1f}", f"{bull_runway_years:.1f}", "Total cash / burn rate"),
    ("Dilution Factor", f"{bear_dilution_factor:.1f}x", f"{base_dilution_factor:.1f}x", f"{bull_dilution_factor:.1f}x", "Future financing rounds"),
    ("Post-Dilution Shares (M)", f"{bear_post_dilution_shares:.0f}", f"{base_post_dilution_shares:.0f}", f"{bull_post_dilution_shares:.0f}", "Current × dilution"),
    ("Cash/Share NAV Floor", f"${bear_nav_per_share:.2f}", f"${base_nav_per_share:.2f}", f"${bull_nav_per_share:.2f}", "Remaining cash / post-dilution shares"),
    ("Pipeline NPV/Share", f"${bear_pipeline_npv_per_share:.2f}", f"${base_pipeline_npv_per_share:.2f}", f"${bull_pipeline_npv_per_share:.2f}", "Certification + commercial optionality"),
    ("Total Value/Share", f"${bear_nav_per_share:.2f}", f"${base_total_per_share:.2f}", f"${bull_total_per_share:.2f}", "NAV + NPV"),
    ("Weight", "25%", "50%", "25%", "Probability assessment"),
    ("Weighted Value/Share", f"${bear_nav_per_share * 0.25:.2f}", f"${base_total_per_share * 0.50:.2f}", f"${bull_total_per_share * 0.25:.2f}", ""),
    ("", "", "", "", ""),
    ("Probability-Weighted FV/Share", "", "", f"${(bear_nav_per_share * 0.25 + base_total_per_share * 0.50 + bull_total_per_share * 0.25):.2f}", "Sum of weighted values"),
    ("Current Price", "", "", f"${price}", ""),
    ("Implied Upside/Downside", "", "", f"{((bear_nav_per_share * 0.25 + base_total_per_share * 0.50 + bull_total_per_share * 0.25) / price - 1) * 100:.1f}%", "FV / Price - 1"),
    ("", "", "", "", ""),
    ("Current Cash/Share NAV", "", "", f"${cash_per_share:.2f}", "Floor value"),
    ("Optionality Premium/Share", "", "", f"${current_optionality_premium:.2f}", "Price - Cash/Share NAV"),
    ("Optionality as % of Price", "", "", f"{current_optionality_premium/price*100:.1f}%", "Premium / Price"),
    ("", "", "", "", ""),
    ("NOTE", "", "", "Standard FCF multiple framework NOT used. See cell A3.", "Pre-commercial developer — per skill guidance"),
]

# Add total weighted FV line
weighted_fv = bear_nav_per_share * 0.25 + base_total_per_share * 0.50 + bull_total_per_share * 0.25

for i, (metric, bear, base, bull, note) in enumerate(scenarios, 6):
    c(ws3, i, 1, metric)
    c(ws3, i, 2, bear)
    c(ws3, i, 3, base)
    c(ws3, i, 4, bull)
    c(ws3, i, 5, note)

# Set column widths
for ci in range(1, 6):
    ws3.column_dimensions[get_column_letter(ci)].width = 22
ws3.column_dimensions['E'].width = 50

# ════════════════════════════════════════════════════════════
# Sheet 4: Actuals Source Audit
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells('A1:E1')
c(ws4, 1, 1, "JOBY — Data Source Audit", font=title_font)

audit_headers = ["Data Point", "Value", "Source", "Date", "Notes"]
write_header_row(ws4, 2, audit_headers)

audit_data = [
    ("Stock Price", "$7.47", "Yahoo Finance", "2026-07-20", "Close price"),
    ("Market Cap", f"${market_cap_b:.2f}B", "Calculated: Price × Shares", "2026-07-20", ""),
    ("Enterprise Value", f"${enterprise_value_b:.2f}B", "MC + Debt - Cash", "2026-07-20", "Uses Stats page debt figure"),
    ("Shares Outstanding", f"{shares_outstanding_mm:.2f}M", "Yahoo Finance Key Stats", "2026-07-20", "Implied shares"),
    ("Total Cash", "$2.47B", "Yahoo Finance Key Stats", "Q1 2026 MRQ", ""),
    ("Total Debt", "$747.73M", "Yahoo Finance Key Stats", "Q1 2026 MRQ", "Stats page; BS shows only $36.8M — includes convertibles/leases"),
    ("TTM Revenue", "$77.67M", "Yahoo Income Statement", "TTM Q1 2026", "in thousands: 77,671"),
    ("FY2025 Revenue", "$53.43M", "Yahoo Income Statement", "12/31/2025", "in thousands: 53,425"),
    ("FY2024 Revenue", "$0.14M", "Yahoo Income Statement", "12/31/2024", "in thousands: 136"),
    ("TTM Gross Profit", "$29.54M", "Yahoo Income Statement", "TTM Q1 2026", ""),
    ("TTM Operating Income", "-$789.89M", "Yahoo Income Statement", "TTM Q1 2026", ""),
    ("TTM Net Income", "-$957.39M", "Yahoo Income Statement", "TTM Q1 2026", ""),
    ("TTM EBITDA", "-$747.87M", "Yahoo Income Statement", "TTM Q1 2026", ""),
    ("FY2025 OpEx", "$743.69M", "Yahoo Income Statement", "12/31/2025", "in thousands: 743,688"),
    ("TTM Operating Cash Flow", "-$543.36M", "Yahoo Cash Flow", "TTM Q1 2026", ""),
    ("TTM Levered FCF", "-$380.63M", "Yahoo Key Stats", "TTM Q1 2026", ""),
    ("Beta (5Y Monthly)", "2.71", "Yahoo Finance Key Stats", "2026-07-20", ""),
    ("52W High", "$20.95", "Yahoo Finance Key Stats", "2026-07-20", ""),
    ("52W Low", "$6.89", "Yahoo Finance Key Stats", "2026-07-20", ""),
    ("Book Value/Share", "$1.76", "Calculated: Equity / Shares", "FY2025", "$1,409.7M / 983.64M"),
    ("10Y Treasury Rate", "4.598%", "CNBC", "2026-07-20", "US10Y close"),
    ("Analyst Revenue FY26", "$113.88M", "Yahoo Analysis", "2026-07-20", "Avg estimate, 8 analysts"),
    ("Analyst Revenue FY27", "$221.10M", "Yahoo Analysis", "2026-07-20", "Avg estimate, 9 analysts"),
    ("Analyst EPS FY26", "-$0.67", "Yahoo Analysis", "2026-07-20", "1 analyst"),
    ("Analyst EPS FY27", "-$0.50", "Yahoo Analysis", "2026-07-20", "2 analysts"),
    ("Total Assets", "$1,795.1M", "Yahoo Balance Sheet", "12/31/2025", "FY2025 annual"),
    ("Total Equity", "$1,409.7M", "Yahoo Balance Sheet", "12/31/2025", "FY2025 annual"),
    ("Employees", "N/A", "Yahoo Finance Profile", "2026-07-20", "Not listed prominently"),
    ("Sector", "Industrials — Aerospace & Defense", "Yahoo Finance Profile", "2026-07-20", ""),
    ("StockAnalysis", "404", "StockAnalysis.com", "2026-07-20", "Ticker not available"),
]

for i, (dp, val, src, dt, note) in enumerate(audit_data, 3):
    c(ws4, i, 1, dp)
    c(ws4, i, 2, val)
    c(ws4, i, 3, src)
    c(ws4, i, 4, dt)
    c(ws4, i, 5, note)

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 25
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 50

# ════════════════════════════════════════════════════════════
# Sheet 5: Questions
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Questions")
ws5.merge_cells('A1:D1')
c(ws5, 1, 1, "JOBY — Open Questions", font=title_font)

questions = [
    ("Q1", "Debt Reconciliation", "Key Stats shows Total Debt of $747.73M but Balance Sheet shows Total Debt of $36.8M. What accounts comprise the $711M difference? Likely convertible preferred/convertible debt sitting outside permanent equity. What are the conversion terms, maturity, and redemption features?"),
    ("Q2", "Regulatory Certification Timeline", "FAA Type Certificate for the eVTOL aircraft — what is the current certification status? When is TC application expected? NRC-equivalent timeline for FAA TC is typically 3-5 years from application. This is the primary binary."),
    ("Q3", "Cash Runway vs. Burn Rate", "At current burn rates (~$125-160M/qtr), $2.47B cash provides ~4-5 years runway. Does this runway extend past the expected certification date? What happens to dilution if runway < certification timeline?"),
    ("Q4", "Revenue Quality", "TTM revenue of $77.7M — what percentage is from government contracts vs. commercial? Q1 FY26 revenue declined 13% QoQ after 93% drop from prior quarter. What is the organic revenue floor?"),
    ("Q5", "Share Count Dilution History", "Shares expanded from 622.6M (FY2022) to 784.2M (FY2024) to 915.1M (FY2025). Implied shares at 983.6M. What was the cumulative dilution factor since IPO? At what prices did each equity round occur?"),
    ("Q6", "Uber/DHL/Toyota Partnership Dependency", "Joby has partnerships with Uber for mobility, Toyota for manufacturing, and DHL for cargo. Are these MOUs legally binding or relationship agreements? What happens if partners pivot to competing eVTOL platforms (Archer, EHang, Volocopter)?"),
    ("Q7", "Capital Expenditure Spike", "FY2025 investing cash flow was -$475.4M vs +$70.8M in FY2024 and +$80.3M in FY2023. Is the manufacturing facility investment cyclical or structural? Does this capex wave end in 1-2 years?"),
    ("Q8", "Competitive Landscape", "How does Joby differentiate from Archer Aviation (ACHR), EHang (EH), Volocopter (VOLT.DE), Lilium, and Beta Technologies? What is the first-mover advantage in U.S. eVTOL certification?"),
    ("Q9", "Manufacturing Scale Risk", "Toyota is a partner in manufacturing. What is the committed production capacity? At what unit volume does the per-aircraft cost reach the target ~$250K-300K range needed for commercial viability?"),
    ("Q10", "Management Capital Allocation", "CEO JoeBen Bevirt earns $661K salary + stock-based comp. CFO Rodrigo Brumana $681K. Is the compensation structure aligned with certification milestones or just revenue growth?"),
    ("Q11", "Interest Income Sustainability", "Interest income was $51.05M TTM ($43.2M FY2025) — this is from the $2.47B cash/short-term investments. As cash burns, this income declines. How much of the operating loss is offset by investment income?"),
    ("Q12", "Short Interest Buildout", "Short interest rose from 88.95M to 99.79M shares (15.21% of float). Short ratio of 2.55. Is this a contrarian signal or legitimate skepticism on certification timeline?"),
]

header_row = ["#", "Topic", "Question"]
write_header_row(ws5, 2, header_row)

for i, (num, topic, question) in enumerate(questions, 3):
    c(ws5, i, 1, num)
    c(ws5, i, 2, topic, font=bold)
    c(ws5, i, 3, question)
    ws5.row_dimensions[i].height = 45

ws5.column_dimensions['A'].width = 6
ws5.column_dimensions['B'].width = 25
ws5.column_dimensions['C'].width = 100

# ════════════════════════════════════════════════════════════
# Sheet 6: Sources
# ════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Sources")
ws6.merge_cells('A1:C1')
c(ws6, 1, 1, "JOBY — Data Sources", font=title_font)

sources = [
    ("1", "Yahoo Finance — Company Profile", "https://finance.yahoo.com/quote/JOBY/profile/"),
    ("2", "Yahoo Finance — Income Statement", "https://finance.yahoo.com/quote/JOBY/financials/"),
    ("3", "Yahoo Finance — Balance Sheet", "https://finance.yahoo.com/quote/JOBY/balance-sheet/"),
    ("4", "Yahoo Finance — Cash Flow", "https://finance.yahoo.com/quote/JOBY/cash-flow/"),
    ("5", "Yahoo Finance — Key Statistics", "https://finance.yahoo.com/quote/JOBY/key-statistics/"),
    ("6", "Yahoo Finance — Analyst Estimates", "https://finance.yahoo.com/quote/JOBY/analysis/"),
    ("7", "CNBC — 10Y Treasury Yield", "https://www.cnbc.com/quotes/US10Y"),
    ("8", "StockAnalysis.com (404 — ticker unavailable)", "https://stockanalysis.com/quote/JOBY/"),
    ("9", "Skill Guidance — Pre-Commercial Technology/Infrastructure Developer", "capital-research skill — references/biotech-valuation-notes.md pattern"),
]

for i, (num, name, url) in enumerate(sources, 2):
    c(ws6, i, 1, num)
    c(ws6, i, 2, name)
    c(ws6, i, 3, url)

ws6.column_dimensions['A'].width = 5
ws6.column_dimensions['B'].width = 55
ws6.column_dimensions['C'].width = 65

# ── Save ──
filename = "models/2026-07-20 Joby Aviation Model.xlsx"
wb.save(filename)
print(f"Saved: {filename}")
print(f"WACC: {wacc:.2f}%")
print(f"Cash/Share NAV: ${cash_per_share:.2f}")
print(f"Optionality Premium: ${current_optionality_premium:.2f}/share ({current_optionality_premium/price*100:.1f}% of price)")
print(f"Probability-Weighted FV: ${weighted_fv:.2f}/share")
print(f"Imp. Upside/Downside: {(weighted_fv/price - 1)*100:.1f}%")
