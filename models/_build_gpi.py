"""Build a 6-sheet valuation model for Group 1 Automotive, Inc. (GPI).

GPI is the largest group of motor vehicle dealerships in Europe with significant
US operations. Revenue ~$22.5B, operating income ~$956M TTM. Highly leveraged
but generates substantial operating cash flow. Primary valuation lens: Forward P/E
on analyst consensus, EV/EBITDA cross-check. FCF framework available but secondary
given the dealership capital cycle dynamics.

Price snapshot: $296.81 on 2026-07-06. Data sources: Yahoo Finance.
StockAnalysis returned 404 for this ticker.

Key context:
  - Largest group of motor vehicle dealers in Europe (UK primary, expanding to US)
  - ~20,452 employees; incorporated 1995, HQ Houston Texas
  - High leverage: Total debt $5.87B vs equity $2.79B (FY25)
  - Forward P/E: 6.92x on FY26 EPS $42.27; analyst avg target $434.50
  - Net tangible assets negative ($-349.7M FY25) — goodwill from acquisitions
  - Strong OCF ($628M TTM) but FCF compresses during capex cycles
  - Earnings date: Jul 23, 2026

"""
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Helper ──
def c(ws, row, col, value, bold=False, comment=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))
    if bold:
        cell.font = Font(bold=True)
    if comment:
        cell.comment = openpyxl.comments.Comment(comment, "Research")
    return cell

def bold_row(ws, row, data):
    for i, v in enumerate(data, 1):
        c(ws, row, i, v, bold=True)

# ── Constants ──
TICKER = "GPI"
DATE = datetime(2026, 7, 6)
PRICE = 296.81
MC = 3.43  # $B (stats page); intraday was $3.53B
EV = 9.00  # $B
TOTAL_DEBT_M = 5870.3  # $M (FY25)
TOTAL_CASH_M = 46.7    # $M (MRQ)
NET_DEBT_M = 5570.0  # $M: EV - MC = 9.00B - 3.43B = 5.57B
# Actually compute more carefully:
# EV = MC + Total Debt - Cash => EV - MC = Debt - Cash (net debt)
# EV 9.00B - MC 3.43B = 5.57B net debt
NET_DEBT_B = EV - MC  # 5.57 $B

BETA = 0.83
RISK_FREE = 0.04485  # 10Y US from CNBC Jul 6, 2026
ERP = 0.05
TAX = 0.21  # US corporate

# WACC components
COE = RISK_FREE + BETA * ERP  # 0.04485 + 0.83 * 0.05 = 0.08635
COD = 0.05  # cost of debt estimate
# Equity from BS FY25
TOTAL_EQUITY = 2789.0  # $M
# Total capitalization = debt + equity
# From BS FY25: Total Debt $5,870.3M, Equity $2,789.0M, Total Cap $6,229.5M
TOTAL_CAP_M = 6229.5
DEBT_W = TOTAL_DEBT_M / TOTAL_CAP_M  # 0.942
EQUITY_W = 1 - DEBT_W  # 0.058

WACC = round(EQUITY_W * COE + DEBT_W * COD * (1 - TAX), 4)

# Historical financials (Yahoo Finance — all in $ millions, converted from thousands)
# Revenue history
REV = {"TTM": 22473.2, "2025": 22571.4, "2024": 19934.4, "2023": 17873.7, "2022": 16222.1}
# Operating income
OPEX = {"TTM": 956.1, "2025": 955.2, "2024": 958.7, "2023": 1001.5, "2022": 1093.5}
# Net income
NI = {"TTM": 323.6, "2025": 321.5, "2024": 487.6, "2023": 586.8, "2022": 730.3}
# EBITDA
EBITDA = {"TTM": 865.6, "2025": 855.4, "2024": 1021.4, "2023": 1056.1, "2022": 1179.4}
# FCF (from cash flow statement — $M)
FCF = {"TTM": 326.4, "2025": 424.5, "2024": 341.2, "2023": 4.8, "2022": 430.4}
# OCF
OCF = {"TTM": 628.2}
# Interest expense
INT_EXP = {"TTM": 289.8, "2025": 284.4, "2024": 249.8, "2023": 163.9, "2022": 104.8}
# Capex
CAPEX = {"TTM": 301.8, "2025": 270.0, "2024": 245.1, "2023": 185.4, "2022": 155.5}

# Shares: MC / Price => 3.43B / 296.81 = 11.56M approximately
# Yahoo says 11.9M shares outstanding. Use that.
SHARES_M = 11.9  # millions

# Analyst estimates (Yahoo Finance Analysis, Jul 6, 2026)
EPS_FY26 = 42.27  # 12 analysts
EPS_FY27 = 47.49  # 12 analysts
REV_FY26 = 22790.0  # $M — 11 analysts
REV_FY27 = 23530.0  # $M — 11 analysts
REV_Q2Y26 = 5710.0  # $M
REV_Q3Y26 = 5840.0  # $M

# Valuation multiples (Yahoo Finance Statistics, Jul 6, 2026)
TRAIL_PE = 10.96
FWD_PE = 6.92
PS = 0.16
PB = 1.21
EV_REV = 0.40
EV_EBITDA = 10.40
PEG = 0.35

# Earnings date
EARNINGS_DATE = "Jul 23, 2026"

# ── FCF check ──
# Net debt ~$5.57B vs FCF ~$326M
# FCF * 10 = $3.26B which is less than net debt of $5.57B
# This means FCF multiple framework will produce negative target prices
# at reasonable multiples. PRIMARY FRAMEWORK: Forward P/E. Cross-check: EV/EBITDA.
print(f"WACC: {WACC}")
print(f"Net debt: ${NET_DEBT_M:.0f}M = ${NET_DEBT_B:.2f}B")
print(f"FCF * 10: ${FCF['TTM'] * 10:.0f}M")
print(f"FCF insufficient check: net_debt > FCF*10 => {NET_DEBT_M > FCF['TTM'] * 10}")
print(f"Switching to Forward P/E primary lens due to FCF insufficiency")

# ── Scenarios ── Forward P/E based (NOT FCF multiple)
# Bear: lower EPS growth, lower multiple
# Base: analyst consensus growth, reasonable multiple
# Bull: higher EPS growth, higher multiple
#
# Analyst average price target: $434.50, high $500.
# FY27 EPS consensus: $47.49.
# Bear: earnings miss consensus, P/E compresses
# Base: earnings hit consensus trajectory, P/E ~9x (near trailing P/E of 10.96)
# Bull: earnings exceed consensus, P/E ~10.5x

scenarios = {
    "Bear": {
        "rev_cagr_5y": 0.01,       # 1% revenue growth — flat
        "term_revenue": 24000.0,     # $M in year 10
        "fcf_margin": 0.012,        # 1.2% FCF/sales
        "exit_pe": 7.5,            # terminal P/E — compressed
        "term_eps": 44.0,          # EPS in year 10 — below consensus growth
        "weight": 0.25,
    },
    "Base": {
        "rev_cagr_5y": 0.025,      # 2.5% revenue growth
        "term_revenue": 28500.0,     # $M in year 10
        "fcf_margin": 0.015,        # 1.5% FCF/sales
        "exit_pe": 9.0,            # terminal P/E — normal for dealerships
        "term_eps": 52.0,          # EPS in year 10 — consensus anchored
        "weight": 0.50,
    },
    "Bull": {
        "rev_cagr_5y": 0.04,       # 4% revenue growth — M&A + organic
        "term_revenue": 32000.0,     # $M in year 10
        "fcf_margin": 0.018,        # 1.8% FCF/sales — improved conversion
        "exit_pe": 10.5,           # terminal P/E — trailing P/E anchor
        "term_eps": 62.0,          # EPS in year 10 — above consensus
        "weight": 0.25,
    },
}

# Compute implied prices per scenario
for name, s in scenarios.items():
    price = s["exit_pe"] * s["term_eps"]
    s["target_price"] = round(price, 2)
    s["upside"] = round((price / PRICE - 1) * 100, 1)
    s["weighted_value"] = round(s["weight"] * price, 2)

total_weighted = sum(s["weighted_value"] for s in scenarios.values())
total_upside = round((total_weighted / PRICE - 1) * 100, 1)
print(f"Probability-weighted FV: ${total_weighted:.2f} (upside: {total_upside}%)")
for name, s in scenarios.items():
    print(f"  {name}: ${s['target_price']:.2f}, upside {s['upside']}%, weighted ${s['weighted_value']:.2f}")

# ── Build workbook ──
out_dir = Path("/home/refcell/dev/capital/models")
out_dir.mkdir(parents=True, exist_ok=True)
wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: Valuation
# ============================================================
ws1 = wb.active
ws1.title = "Valuation"
ws1.merge_cells("A1:D1")
c(ws1, 1, 1, f"GPI — Group 1 Automotive, Inc. Valuation Model", bold=True)

title_data = [
    ["Company:", "Group 1 Automotive, Inc.", ""],
    ["Ticker:", "NYSE: GPI", ""],
    ["Date:", str(DATE.strftime("%Y-%m-%d")), ""],
    ["Price:", f"${PRICE:.2f}", "Close Jul 6, 2026"],
    ["Shares Outstanding:", f"{SHARES_M:.1f}M", "Yahoo Finance MRQ"],
    ["Market Cap:", f"${MC:.2f}B", "Yahoo Finance Stats"],
    ["Enterprise Value:", f"${EV:.2f}B", "Yahoo Finance Stats"],
    ["Primary Lens:", "Forward P/E (Analyst Consensus)", "FCF insufficient — see Scenarios note"],
    ["Stance:", "Watch / Needs more work", "High leverage, attractive Fwd P/E, earnings catalyst Jul 23"],
]

# Header at row 2
c(ws1, 2, 1, "Field", bold=True)
c(ws1, 2, 2, "Value", bold=True)
c(ws1, 2, 3, "Comment", bold=True)

for i, row_data in enumerate(title_data, 3):
    for j, val in enumerate(row_data, 1):
        c(ws1, i, j, val, bold=(j == 1))

# Valuation metrics table
r = len(title_data) + 5
ws1.merge_cells(f"A{r}:B{r}")
c(ws1, r, 1, "Key Valuation Metrics", bold=True)
r += 1
c(ws1, r, 1, "Multiple", bold=True)
c(ws1, r, 2, "Value", bold=True)
c(ws1, r, 3, "Comment", bold=True)

metrics = [
    ["P/E (Trailing)", f"{TRAIL_PE:.2f}x", "TTM EPS $26.42; 12-month average"],
    ["Forward P/E", f"{FWD_PE:.2f}x", "On FY26 EPS $42.27 (12 analysts) — looks cheap"],
    ["P/S (TTM)", f"{PS:.2f}x", "On $22.47B TTM revenue; below peer average"],
    ["P/B", f"{PB:.2f}x", "On book value; tangible book is NEGATIVE"],
    ["EV/Revenue", f"{EV_REV:.2f}x", "Below peer group average"],
    ["EV/EBITDA", f"{EV_EBITDA:.2f}x", "Reasonable for auto dealerships"],
    ["PEG (5yr)", f"{PEG:.2f}", "Below 1.0 — growth-adjusted cheapness"],
]
for i, m in enumerate(metrics, r + 1):
    for j, v in enumerate(m, 1):
        c(ws1, i, j, v, bold=(j == 1))

# ============================================================
# Sheet 2: WACC
# ============================================================
ws2 = wb.create_sheet("WACC")
ws2.merge_cells("A1:C1")
c(ws2, 1, 1, f"GPI — WACC Calculation", bold=True)

c(ws2, 2, 1, "Component", bold=True)
c(ws2, 2, 2, "Value", bold=True)
c(ws2, 2, 3, "Source / Notes", bold=True)

wacc_data = [
    ["Risk-Free Rate (10Y US)", f"{RISK_FREE:.4f} ({RISK_FREE*100:.2f}%)", "CNBC Jul 6, 2026: 4.485%"],
    ["Equity Risk Premium", f"{ERP:.2%}", "Standard assumption"],
    ["Beta (5Y Monthly)", f"{BETA:.2f}", "Yahoo Finance Stats"],
    ["Cost of Equity (CAPM)", f"{COE:.4f} ({COE*100:.2f}%)", f"Rf + Beta × ERP = {RISK_FREE:.4f} + {BETA} × {ERP}"],
    ["Cost of Debt", f"{COD:.2%}", "Estimated — dealership finance debt"],
    ["Tax Rate", f"{TAX:.0%}", "US corporate rate"],
    ["Market Cap (Equity)", f"${TOTAL_EQUITY:.1f}M", "FY25 BS equity value"],
    ["Total Debt", f"${TOTAL_DEBT_M:.1f}M", "FY25 BS total debt"],
    ["Debt Weight", f"{DEBT_W:.4f} ({DEBT_W*100:.1f}%)", f"Debt / (Debt + Equity) = {TOTAL_DEBT_M:.0f} / {TOTAL_CAP_M:.0f}"],
    ["Equity Weight", f"{EQUITY_W:.4f} ({EQUITY_W*100:.1f}%)", f"Equity / (Debt + Equity) = {TOTAL_CAP_M-TOTAL_DEBT_M:.0f} / {TOTAL_CAP_M:.0f}"],
    ["WACC", f"{WACC:.4f} ({WACC*100:.2f}%)", f"Weighted: {EQUITY_W:.3f}×{COE:.4f} + {DEBT_W:.3f}×{COD:.4f}×(1-{TAX:.2f})"],
]
for i, row_data in enumerate(wacc_data, 3):
    for j, val in enumerate(row_data, 1):
        c(ws2, i, j, val, bold=(j == 1))

# ============================================================
# Sheet 3: Scenarios
# ============================================================
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells("A1:K1")
c(ws3, 1, 1, f"GPI — Bear / Base / Bull Scenarios (Forward P/E Framework)", bold=True)

# Note about framework switch
ws3.merge_cells("A2:K2")
c(ws3, 2, 1, "NOTE: FCF multiple framework not used — net debt ($5.57B) > FCF × 10 ($3.26B). Forward P/E is the primary lens.", bold=True)

c(ws3, 3, 1, "Parameter", bold=True)
c(ws3, 3, 2, "Bear", bold=True)
c(ws3, 3, 3, "Base", bold=True)
c(ws3, 3, 4, "Bull", bold=True)

# Scenario parameters and targets
# Terminal FCF (in $M): terminal_revenue * fcf_margin
# But we use Forward P/E, so compute target price = exit_PE * terminal_EPS
# For EV cross-check: terminal_FCF = terminal_revenue * fcf_margin
# implied_EV = terminal_FCF * exit_multiple_for_FCF — but we're using P/E
# So the EV column shows: exit_PE * term_EPS * shares_M = implied market cap = EV + net cash approx

scenario_rows = [
    ["Revenue CAGR (5Y)", "1.0%", "2.5%", "4.0%"],
    ["Terminal Revenue ($M)", f"{scenarios['Bear']['term_revenue']:.0f}", f"{scenarios['Base']['term_revenue']:.0f}", f"{scenarios['Bull']['term_revenue']:.0f}"],
    ["Adj. FCF Margin", f"{scenarios['Bear']['fcf_margin']:.1%}", f"{scenarios['Base']['fcf_margin']:.1%}", f"{scenarios['Bull']['fcf_margin']:.1%}"],
    ["Terminal FCF ($M)", f"{scenarios['Bear']['term_revenue'] * scenarios['Bear']['fcf_margin']:.0f}",
     f"{scenarios['Base']['term_revenue'] * scenarios['Base']['fcf_margin']:.0f}",
     f"{scenarios['Bull']['term_revenue'] * scenarios['Bull']['fcf_margin']:.0f}"],
    ["Exit P/E Multiple", f"{scenarios['Bear']['exit_pe']:.1f}x", f"{scenarios['Base']['exit_pe']:.1f}x", f"{scenarios['Bull']['exit_pe']:.1f}x"],
    ["Terminal EPS ($)", f"{scenarios['Bear']['term_eps']:.2f}", f"{scenarios['Base']['term_eps']:.2f}", f"{scenarios['Bull']['term_eps']:.2f}"],
    ["Implied Market Cap ($B)", f"{scenarios['Bear']['exit_pe'] * scenarios['Bear']['term_eps'] * SHARES_M / 1000:.2f}",
     f"{scenarios['Base']['exit_pe'] * scenarios['Base']['term_eps'] * SHARES_M / 1000:.2f}",
     f"{scenarios['Bull']['exit_pe'] * scenarios['Bull']['term_eps'] * SHARES_M / 1000:.2f}"],
    ["Target Price ($)", f"${scenarios['Bear']['target_price']:.2f}", f"${scenarios['Base']['target_price']:.2f}", f"${scenarios['Bull']['target_price']:.2f}"],
    ["Upside from $296.81", f"{scenarios['Bear']['upside']:.1f}%", f"{scenarios['Base']['upside']:.1f}%", f"{scenarios['Bull']['upside']:.1f}%"],
    ["Scenario Weight", f"{scenarios['Bear']['weight']:.0%}", f"{scenarios['Base']['weight']:.0%}", f"{scenarios['Bull']['weight']:.0%}"],
    ["Weighted Value/Share", f"${scenarios['Bear']['weighted_value']:.2f}", f"${scenarios['Base']['weighted_value']:.2f}", f"${scenarios['Bull']['weighted_value']:.2f}"],
]

for i, row_data in enumerate(scenario_rows, 4):
    for j, val in enumerate(row_data, 1):
        c(ws3, i, j, val, bold=(j == 1))

# Weighted FV row
fv_row = len(scenario_rows) + 5
c(ws3, fv_row, 1, "Probability-Weighted Fair Value", bold=True)
c(ws3, fv_row, 2, f"${total_weighted:.2f}", bold=True)
c(ws3, fv_row, 3, f"{total_upside:.1f}%", bold=True)
c(ws3, fv_row + 1, 1, "Current Price", bold=True)
c(ws3, fv_row + 1, 2, f"${PRICE:.2f}", bold=True)

# ============================================================
# Sheet 4: Actuals Source Audit
# ============================================================
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells("A1:D1")
c(ws4, 1, 1, f"GPI — Actuals Source Audit", bold=True)

c(ws4, 2, 1, "Data Point", bold=True)
c(ws4, 2, 2, "Value", bold=True)
c(ws4, 2, 3, "Source URL", bold=True)
c(ws4, 2, 4, "Date / Notes", bold=True)

audit_rows = [
    ["Stock Price", "$296.81", "finance.yahoo.com/quote/GPI/", "Jul 6, 2026 close"],
    ["Market Cap", "$3.43B", "finance.yahoo.com/quote/GPI/key-statistics/", "Stats page as of Jul 2, 2026"],
    ["Enterprise Value", "$9.00B", "finance.yahoo.com/quote/GPI/key-statistics/", "Stats page as of Jul 2, 2026"],
    ["Shares Outstanding", "11.9M", "finance.yahoo.com/quote/GPI/key-statistics/", "Yahoo Finance MRQ"],
    ["Beta (5Y Monthly)", "0.83", "finance.yahoo.com/quote/GPI/key-statistics/", "Yahoo Finance"],
    ["TTM Revenue", "$22,473.2M", "finance.yahoo.com/quote/GPI/financials/", "Income statement TTM column"],
    ["FY25 Revenue", "$22,571.4M", "finance.yahoo.com/quote/GPI/financials/", "Income statement FY2025"],
    ["FY24 Revenue", "$19,934.4M", "finance.yahoo.com/quote/GPI/financials/", "Income statement FY2024"],
    ["FY23 Revenue", "$17,873.7M", "finance.yahoo.com/quote/GPI/financials/", "Income statement FY2023"],
    ["FY22 Revenue", "$16,222.1M", "finance.yahoo.com/quote/GPI/financials/", "Income statement FY2022"],
    ["TTM Operating Income", "$956.1M", "finance.yahoo.com/quote/GPI/financials/", "Income statement TTM"],
    ["TTM Net Income", "$323.6M", "finance.yahoo.com/quote/GPI/financials/", "Income statement TTM"],
    ["TTM EBITDA", "$865.6M", "finance.yahoo.com/quote/GPI/financials/", "Income statement TTM"],
    ["TTM Diluted EPS", "$26.42", "finance.yahoo.com/quote/GPI/financials/", "Income statement TTM"],
    ["Total Debt (FY25)", "$5,870.3M", "finance.yahoo.com/quote/GPI/balance-sheet/", "Balance sheet FY2025"],
    ["Total Equity (FY25)", "$2,789.0M", "finance.yahoo.com/quote/GPI/balance-sheet/", "Balance sheet FY2025"],
    ["Net Debt (FY25)", "$5,582.8M", "finance.yahoo.com/quote/GPI/balance-sheet/", "Balance sheet FY2025"],
    ["Total Cash (MRQ)", "$46.7M", "finance.yahoo.com/quote/GPI/key-statistics/", "Stats page"],
    ["TTM Operating CF", "$628.2M", "finance.yahoo.com/quote/GPI/cash-flow/", "Cash flow TTM"],
    ["TTM Free CF", "$326.4M", "finance.yahoo.com/quote/GPI/cash-flow/", "Cash flow TTM"],
    ["FY25 Free CF", "$424.5M", "finance.yahoo.com/quote/GPI/cash-flow/", "Cash flow FY2025"],
    ["TTM Capex", "$301.8M", "finance.yahoo.com/quote/GPI/cash-flow/", "Cash flow TTM"],
    ["Buybacks (TTM)", "$504.4M", "finance.yahoo.com/quote/GPI/cash-flow/", "Cash flow TTM repurchase"],
    ["TTM Interest Expense", "$289.8M", "finance.yahoo.com/quote/GPI/financials/", "Income statement TTM"],
    ["Analyst Avg Target", "$434.50", "finance.yahoo.com/quote/GPI/analysis/", "High $500, avg $434.50"],
    ["FY26 EPS Consensus", "$42.27", "finance.yahoo.com/quote/GPI/analysis/", "12 analysts"],
    ["FY27 EPS Consensus", "$47.49", "finance.yahoo.com/quote/GPI/analysis/", "12 analysts"],
    ["FY26 Revenue Consensus", "$22,790M", "finance.yahoo.com/quote/GPI/analysis/", "11 analysts"],
    ["FY27 Revenue Consensus", "$23,530M", "finance.yahoo.com/quote/GPI/analysis/", "11 analysts"],
    ["EBITDA (FY25)", "$855.4M", "finance.yahoo.com/quote/GPI/financials/", "Income statement FY2025"],
    ["EBITDA (FY24)", "$1,021.4M", "finance.yahoo.com/quote/GPI/financials/", "Income statement FY2024"],
    ["Earnings Date (Est.)", "Jul 23, 2026", "finance.yahoo.com/quote/GPI/", "Summary page"],
    ["10Y US Treasury", "4.485%", "cnbc.com/quotes/US10Y", "Jul 6, 2026"],
    ["P/E (Trailing)", "10.96", "finance.yahoo.com/quote/GPI/key-statistics/", "Stats page"],
    ["Forward P/E", "6.92", "finance.yahoo.com/quote/GPI/key-statistics/", "Stats page"],
    ["P/S (TTM)", "0.16", "finance.yahoo.com/quote/GPI/key-statistics/", "Stats page"],
    ["P/B (MRQ)", "1.21", "finance.yahoo.com/quote/GPI/key-statistics/", "Stats page"],
    ["EV/Revenue", "0.40", "finance.yahoo.com/quote/GPI/key-statistics/", "Stats page"],
    ["EV/EBITDA", "10.40", "finance.yahoo.com/quote/GPI/key-statistics/", "Stats page"],
]

for i, row_data in enumerate(audit_rows, 3):
    for j, val in enumerate(row_data, 1):
        c(ws4, i, j, val, bold=(j == 1))

# ============================================================
# Sheet 5: Questions
# ============================================================
ws5 = wb.create_sheet("Questions")
ws5.merge_cells("A1:B1")
c(ws5, 1, 1, f"GPI — Open Questions", bold=True)

c(ws5, 2, 1, "#", bold=True)
c(ws5, 2, 2, "Question", bold=True)

questions = [
    "FCF insufficiency: Net debt ($5.57B) exceeds 10× FCF ($3.26B). Even at 8-10x FCF multiple, implied EV ($2.6-$3.3B) is well below net debt (~$5.6B). Why does the FCF framework break, and what framework should dominate?",
    "Acquisition-driven growth: Revenue jumped from $16.2B (FY22) to $22.5B (FY25) — a 39% increase in three years. How much is organic vs. M&A? The goodwill accumulation (net tangible assets turned from $321M positive in FY23 to -$350M in FY25) signals heavy acquisition activity.",
    "TTM earnings compression: TTM net income ($323.6M) declined from $321.5M (FY25) despite revenue remaining essentially flat. Operating income was stable at ~$955M, but TTM net income dropped 17% from TTM. What drove the decline? Interest expense rose to $289.8M from $284.4M.",
    "Share count dynamics: Shares have fallen from 14.3M (FY22) to ~11.9M (MRQ) — a 17% reduction through aggressive buybacks ($504M TTM, $555M in FY25). Is this share count trend sustainable given the leverage?",
    "Leverage trajectory: Total debt grew from $3.35B (FY22) to $5.87B (FY25) — a 75% increase. Total debt/equity ratio is 197.63%. At what point does the debt service burden (currently $290M/yr interest) become problematic for a cyclical business?",
    "Dealership economics in EV transition: As the auto market shifts to EVs, dealership margins on service and parts could compress significantly. EVs require less maintenance. How exposed is GPI to this structural headwind?",
    "Geographic concentration: UK vs. US revenue split and exposure to post-Brexit trade dynamics. Does the UK market growth trajectory differ materially from the US?",
    "Brand alignment initiative: Recent PR releases show aggressive nationwide rebranding (Group 1 Kia, Group 1 Volkswagen, etc.). What is the cost of this initiative, and does it drive measurable same-store sales growth?",
    "Forward P/E anomaly: Forward P/E of 6.92x on $42.27 EPS vs. trailing P/E of 10.96x on $26.42 EPS is a massive gap. This implies analysts expect a 60%+ earnings recovery from TTM levels. Is this plausible given the stable $956M operating income base?",
    "Analyst coverage: 11-12 analysts covering with an average target of $434.50 (46% above current price). However, Argus rates it SELL with a $262 target (below current). What explains the divergence between the average target and the explicit sell rating?",
    "Next earnings catalyst: Jul 23, 2026 earnings report is imminent. Q1 FY26 showed EPS of $8.66 vs. estimate of $8.82 (miss by 1.8%). Will Q2 recover or continue the miss trend?",
    "Auto credit cycle: Dealership profitability is highly sensitive to vehicle financing conditions. Rising rates increase consumer borrowing costs and reduce deal flow. How does the current rate environment impact Q2/Q3 deal volume?",
]

for i, q in enumerate(questions, 3):
    c(ws5, i, 1, i - 2, bold=True)
    c(ws5, i, 2, q)
    ws5.row_dimensions[i].height = 40

# ============================================================
# Sheet 6: Sources
# ============================================================
ws6 = wb.create_sheet("Sources")
ws6.merge_cells("A1:B1")
c(ws6, 1, 1, f"GPI — Data Sources", bold=True)

c(ws6, 2, 1, "#", bold=True)
c(ws6, 2, 2, "Source", bold=True)

sources = [
    "Yahoo Finance Summary: finance.yahoo.com/quote/GPI/ — Price, market cap, description, earnings date, performance",
    "Yahoo Finance Income Statement: finance.yahoo.com/quote/GPI/financials/ — Revenue, operating income, net income, EPS, EBITDA",
    "Yahoo Finance Balance Sheet: finance.yahoo.com/quote/GPI/balance-sheet/ — Assets, liabilities, equity, debt, net debt",
    "Yahoo Finance Cash Flow: finance.yahoo.com/quote/GPI/cash-flow/ — OCF, FCF, capex, buybacks, debt issuance/repayment",
    "Yahoo Finance Statistics: finance.yahoo.com/quote/GPI/key-statistics/ — Valuation multiples, beta, shares, profitability, short interest",
    "Yahoo Finance Analysis: finance.yahoo.com/quote/GPI/analysis/ — Analyst estimates (EPS, revenue), revisions, profit growth",
    "Yahoo Finance Profile: finance.yahoo.com/quote/GPI/profile/ — Company description, employee count, fiscal year, executives",
    "CNBC 10Y Treasury: cnbc.com/quotes/US10Y — Risk-free rate: 4.485% as of Jul 6, 2026",
    "StockAnalysis: stockanalysis.com/quotes/GPI/ — Returned 404, not available",
    "Yahoo Finance peer comparison: AN (AutoNation), ABG (Asbury), LAD (Lithia), PAG (Penske), SAH (Sonic)",
]

for i, s in enumerate(sources, 3):
    c(ws6, i, 1, i - 2, bold=True)
    c(ws6, i, 2, s)

# ── Column widths ──
for ws in [ws1, ws2, ws4, ws6]:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 30

for ws in [ws3]:
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

ws5.column_dimensions["A"].width = 5
ws5.column_dimensions["B"].width = 120

# ── Save ──
out_path = out_dir / f"[2026-07-06] Group 1 Automotive Model.xlsx"
wb.save(out_path)

# Verify
wb2 = openpyxl.load_workbook(out_path)
print(f"\nWorkbook saved: {out_path}")
print(f"Sheets: {wb2.sheetnames}")
for s in wb2.sheetnames:
    ws = wb2[s]
    print(f"  {s}: {ws.max_row} rows × {ws.max_column} cols")
