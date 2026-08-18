#!/usr/bin/env python3
"""Build 6-sheet valuation model for CLBT (Cellebrite DI Ltd.)."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Styles
title_font = Font(bold=True, size=14)
header_font = Font(bold=True, size=12)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
italic_font = Font(italic=True, size=10)

D = "$"  # avoid f-string $ issues in some Python versions


def c(ws, row, col, value, font=None, border=False, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if border:
        cell.border = thin_border
    if alignment:
        cell.alignment = alignment
    return cell


# ============================================================================
# Sheet 1: Valuation
# ============================================================================
ws1 = wb.active
ws1.title = "Valuation"

ws1.merge_cells("A1:F1")
ws1["A1"].value = "Cellebrite DI Ltd. (CLBT) — Valuation"
ws1["A1"].font = title_font
ws1.merge_cells("A2:F2")
ws1["A2"].value = "Technology / Software — Infrastructure | NASDAQ"
ws1["A2"].font = Font(italic=True, size=11)

title_data = [
    ("Date", "2026-08-17"),
    ("Ticker", "NASDAQ: CLBT"),
    ("Price (" + D + ")", "10.44"),
    ("Shares Outstanding (M)", "250.79"),
    ("Market Cap (" + D + "B)", "2.62"),
    ("Total Cash (" + D + "M)", "442.53"),
    ("Total Debt (" + D + "M)", "23.23"),
    ("Enterprise Value (" + D + "B)", "2.20"),
    ("Net Cash Position (" + D + "M)", "419.30"),
    ("Primary Valuation Lens", "FCF Multiple"),
    ("Stance", "Watch"),
]

for i, (field, value) in enumerate(title_data, 3):
    c(ws1, i, 1, field, font=Font(bold=True))
    c(ws1, i, 2, value)

c(ws1, 15, 1, "Key Valuation Metrics", font=header_font)
for j, h in enumerate(["Metric", "Value", "Comment"], 1):
    c(ws1, 16, j, h, font=Font(bold=True, underline='single'), border=True)

metrics = [
    (
        "P/E (TTM)",
        "48.4x",
        "Key Stats EPS " + D + "0.22; suppressed by FY2024 non-op charge of -" + D + "331.5M",
    ),
    ("Forward P/E", "N/A", "No analyst coverage visible"),
    (
        "P/S (TTM)",
        "5.1x",
        "MC " + D + "2.62B / Revenue " + D + "514.3M TTM",
    ),
    (
        "P/FCF (TTM)",
        "21.0x",
        "MC " + D + "2.62B / Levered FCF " + D + "124.9M",
    ),
    ("EV/FCF", "17.6x", "EV " + D + "2.20B / Levered FCF " + D + "124.9M"),
    ("EV/Sales", "4.3x", "EV " + D + "2.20B / Revenue " + D + "514.3M"),
    (
        "EV/EBITDA",
        "28.1x",
        "EV " + D + "2.20B / EBITDA " + D + "78.4M (S&P calc)",
    ),
    ("P/B", "5.2x", "Key Stats current"),
]

for i, (metric, value, comment) in enumerate(metrics, 17):
    c(ws1, i, 1, metric, border=True)
    c(ws1, i, 2, value, border=True)
    c(ws1, i, 3, comment, border=True)

# ============================================================================
# Sheet 2: WACC
# ============================================================================
ws2 = wb.create_sheet("WACC")
ws2.merge_cells("A1:C1")
ws2["A1"].value = "WACC Calculation — CLBT"
ws2["A1"].font = title_font

wacc_data = [
    ("Risk-Free Rate (10Y US)", "4.73%", "CNBC 2026-08-17"),
    ("Equity Risk Premium", "5.00%", "Assumption"),
    ("Beta (5Y Monthly)", "1.18", "Yahoo Key Stats"),
    ("Cost of Equity (Ke)", "10.63%", "=4.73% + 1.18 * 5%"),
    ("Cost of Debt", "4.50%", "Estimate; debt is negligible"),
    ("Tax Rate", "15.00%", "FY2025 eff: 13.6%; blended forward"),
    ("", "", ""),
    ("Market Cap (" + D + "M)", "2,618.30", D + "10.44 * 250.79M"),
    ("Total Debt (" + D + "M)", "23.23", "Yahoo Key Stats MRQ"),
    ("Equity Weight", "99.1%", "=2,618.3 / (2,618.3+23.2)"),
    ("Debt Weight", "0.9%", "=23.2 / (2,618.3+23.2)"),
    ("", "", ""),
    ("WACC", "10.57%", "=0.991*10.63% + 0.009*4.50%*(1-15%)"),
]

for i, (field, value, note) in enumerate(wacc_data, 2):
    bold = Font(bold=True) if field and "=" not in value and field else None
    c(ws2, i, 1, field, font=bold)
    c(ws2, i, 2, value)
    c(ws2, i, 3, note)

# ============================================================================
# Sheet 3: Scenarios
# ============================================================================
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells("A1:J1")
ws3["A1"].value = "Scenario Analysis — CLBT (FCF Multiple Framework)"
ws3["A1"].font = title_font

# Scenario parameters
rev_mm = 514.3  # TTM revenue in $M
net_cash_mm = 419.30  # net cash in $M
shares_mm = 250.79  # shares in millions
current_price = 10.44

# Compute scenarios
bear_cagr, base_cagr, bull_cagr = 0.08, 0.12, 0.15
bear_marg, base_marg, bull_marg = 0.20, 0.25, 0.32
bear_exit, base_exit, bull_exit = 10, 15, 20

bear_rev = rev_mm * (1 + bear_cagr) ** 5
base_rev = rev_mm * (1 + base_cagr) ** 5
bull_rev = rev_mm * (1 + bull_cagr) ** 5

bear_fcf = bear_rev * bear_marg
base_fcf = base_rev * base_marg
bull_fcf = bull_rev * bull_marg

bear_implied_ev = bear_fcf * bear_exit
base_implied_ev = base_fcf * base_exit
bull_implied_ev = bull_fcf * bull_exit

bear_eq = bear_implied_ev + net_cash_mm
base_eq = base_implied_ev + net_cash_mm
bull_eq = bull_implied_ev + net_cash_mm

bear_price = bear_eq / shares_mm
base_price = base_eq / shares_mm
bull_price = bull_eq / shares_mm

bear_up = (bear_price - current_price) / current_price
base_up = (base_price - current_price) / current_price
bull_up = (bull_price - current_price) / current_price

w_bear, w_base, w_bull = 0.30, 0.50, 0.20
weighted_fv = w_bear * bear_price + w_base * base_price + w_bull * bull_price
total_up = (weighted_fv - current_price) / current_price

# Print verification
print("=== SCENARIO VERIFICATION ===")
print("  Bear: " + D + f"{bear_price:.2f} (upside {bear_up*100:.1f}%)")
print("  Base: " + D + f"{base_price:.2f} (upside {base_up*100:.1f}%)")
print("  Bull: " + D + f"{bull_price:.2f} (upside {bull_up*100:.1f}%)")
print("  Wt'd FV: " + D + f"{weighted_fv:.2f} (upside {total_up*100:.1f}%)")

# Scenario header at row 4
for j, h in enumerate(
    ["Metric", "Bear", "Base", "Bull", "Notes"], 1
):
    c(ws3, 4, j, h, font=Font(bold=True, underline='single'), border=True)

scenarios = [
    ("Revenue CAGR (5Y)", "8%", "12%", "15%", "Historical 3Y ~20%; post-guidance cut"),
    (
        "Terminal Revenue (" + D + "M)",
        f"{bear_rev:.0f}",
        f"{base_rev:.0f}",
        f"{bull_rev:.0f}",
        "= Rev * (1+CAGR)^5",
    ),
    (
        "Adjusted FCF Margin",
        "20%",
        "25%",
        "32%",
        "FCF margin hist 28-34%; bear sees compression",
    ),
    (
        "Terminal FCF (" + D + "M)",
        f"{bear_fcf:.0f}",
        f"{base_fcf:.0f}",
        f"{bull_fcf:.0f}",
        "= Term Rev * FCF Margin",
    ),
    ("Exit FCF Multiple", "10x", "15x", "20x", "SaaS range 12-25x"),
    (
        "Implied EV (" + D + "M)",
        f"{bear_implied_ev:.0f}",
        f"{base_implied_ev:.0f}",
        f"{bull_implied_ev:.0f}",
        "= Term FCF * Exit Multiple",
    ),
    (
        "Plus: Net Cash (" + D + "M)",
        f"{net_cash_mm:.1f}",
        f"{net_cash_mm:.1f}",
        f"{net_cash_mm:.1f}",
        f"{net_cash_mm:.1f} net cash",
    ),
    (
        "Equity Value (" + D + "M)",
        f"{bear_eq:.0f}",
        f"{base_eq:.0f}",
        f"{bull_eq:.0f}",
        "= EV + Net Cash",
    ),
    (
        "Shares (M)",
        f"{shares_mm:.1f}",
        f"{shares_mm:.1f}",
        f"{shares_mm:.1f}",
        str(shares_mm) + "M",
    ),
    (
        "Target Price (" + D + ")",
        f"{bear_price:.2f}",
        f"{base_price:.2f}",
        f"{bull_price:.2f}",
        "= Eq Value / Shares",
    ),
    (
        "Upside from Current",
        f"{bear_up*100:.1f}%",
        f"{base_up*100:.1f}%",
        f"{bull_up*100:.1f}%",
        "Current: " + D + f"{current_price:.2f}",
    ),
    ("Weight", "30%", "50%", "20%", "Bear-heavy: CEO change + guidance cut"),
    (
        "Weighted Value/Share",
        f"{w_bear * bear_price:.2f}",
        f"{w_base * base_price:.2f}",
        f"{w_bull * bull_price:.2f}",
        "= Target * Weight",
    ),
]

for i, row_data in enumerate(scenarios, 5):
    for j, val in enumerate(row_data, 1):
        c(ws3, i, j, val, border=True)

# Summary rows
c(ws3, 21, 1, "Probability-Weighted FV", font=Font(bold=True), border=True)
c(ws3, 21, 2, "", border=True)
c(ws3, 21, 3, "", border=True)
c(ws3, 21, 4, f"{weighted_fv:.2f}", font=Font(bold=True), border=True)
c(ws3, 21, 5, "Sum of weighted", border=True)

c(ws3, 22, 1, "Current Price", font=Font(bold=True), border=True)
c(ws3, 22, 2, "", border=True)
c(ws3, 22, 3, "", border=True)
c(ws3, 22, 4, f"{current_price:.2f}", font=Font(bold=True), border=True)
c(ws3, 22, 5, "2026-08-17 close", border=True)

c(ws3, 23, 1, "Implied Upside", font=Font(bold=True), border=True)
c(ws3, 23, 2, "", border=True)
c(ws3, 23, 3, "", border=True)
c(ws3, 23, 4, f"{total_up*100:.1f}%", font=Font(bold=True), border=True)
c(ws3, 23, 5, "= (FV - Current)" + " / Current", border=True)

# Note
c(
    ws3,
    25,
    1,
    "Framework note: FCF multiple used. Company is net-cash positive (" + D + f"{net_cash_mm:.1f}M). Standard EV -> equity math applies. No FCF insufficiency concern.",
)

# ============================================================================
# Sheet 4: Actuals Source Audit
# ============================================================================
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells("A1:E1")
ws4["A1"].value = "Actuals Source Audit — CLBT"
ws4["A1"].font = title_font

for j, h in enumerate(["Data Point", "Value", "Source URL", "Date", "Notes"], 1):
    c(ws4, 2, j, h, font=Font(bold=True, underline='single'), border=True)

audit = [
    (
        "Stock Price",
        D + "10.44",
        "finance.yahoo.com/quote/CLBT/",
        "2026-08-17",
        "Close price",
    ),
    (
        "Market Cap",
        D + "2.62B",
        "finance.yahoo.com/quote/CLBT/key-statistics/",
        "2026-08-17",
        D + "10.44 * 250.79M shares",
    ),
    (
        "Enterprise Value",
        D + "2.20B",
        "Calculated: MC + Debt - Cash",
        "2026-08-17",
        "Net cash position",
    ),
    (
        "Shares Outstanding",
        "250.79M",
        "finance.yahoo.com/quote/CLBT/key-statistics/",
        "2026-08-17",
        "Key Stats",
    ),
    (
        "Total Cash",
        D + "442.53M",
        "finance.yahoo.com/quote/CLBT/key-statistics/",
        "MRQ 6/30/2026",
        "Key Stats total cash",
    ),
    (
        "Total Debt",
        D + "23.23M",
        "finance.yahoo.com/quote/CLBT/key-statistics/",
        "MRQ 6/30/2026",
        "Key Stats total debt",
    ),
    (
        "TTM Revenue",
        D + "514.29M",
        "finance.yahoo.com/quote/CLBT/key-statistics/",
        "TTM as of MRQ",
        "Key Stats; IS TTM shows " + D + "496.4M (may lag)",
    ),
    (
        "FY2025 Revenue",
        D + "475.68M",
        "finance.yahoo.com/quote/CLBT/financials/",
        "12/31/2025",
        "Annual IS",
    ),
    (
        "FY2024 Revenue",
        D + "401.20M",
        "finance.yahoo.com/quote/CLBT/financials/",
        "12/31/2024",
        "Annual IS",
    ),
    (
        "FY2025 Operating Income",
        D + "66.48M",
        "finance.yahoo.com/quote/CLBT/financials/",
        "12/31/2025",
        "Annual IS",
    ),
    (
        "FY2025 Net Income",
        D + "77.81M",
        "finance.yahoo.com/quote/CLBT/financials/",
        "12/31/2025",
        "Annual IS",
    ),
    (
        "FY2024 Net Income",
        "-" + D + "283.01M",
        "finance.yahoo.com/quote/CLBT/financials/",
        "12/31/2024",
        "Non-op charge of -" + D + "331.5M; operating income +" + D + "56.9M",
    ),
    (
        "FY2023 Net Income",
        "-" + D + "81.10M",
        "finance.yahoo.com/quote/CLBT/financials/",
        "12/31/2023",
        "Non-op charge of -" + D + "109.0M; operating income +" + D + "33.2M",
    ),
    (
        "FY2025 FCF",
        D + "160.32M",
        "finance.yahoo.com/quote/CLBT/cash-flow/",
        "12/31/2025",
        "Cash flow statement",
    ),
    (
        "TTM Levered FCF",
        D + "124.92M",
        "finance.yahoo.com/quote/CLBT/key-statistics/",
        "TTM",
        "Key Stats levered FCF",
    ),
    (
        "TTM OCF",
        D + "157.56M",
        "finance.yahoo.com/quote/CLBT/key-statistics/",
        "TTM",
        "Key Stats operating CF",
    ),
    (
        "FY2024 Investing CF",
        "-" + D + "268.25M",
        "finance.yahoo.com/quote/CLBT/cash-flow/",
        "12/31/2024",
        "vs " + D + "13.23M CapEx; gap of -" + D + "255M = acquisition?",
    ),
    (
        "Beta (5Y Monthly)",
        "1.18",
        "finance.yahoo.com/quote/CLBT/key-statistics/",
        "2026-08-17",
        "Key Stats",
    ),
    (
        "10Y Treasury Yield",
        "4.73%",
        "cnbc.com/quotes/US10Y",
        "2026-08-17",
        "CNBC",
    ),
    (
        "Next Earnings",
        "Nov 11, 2026",
        "finance.yahoo.com/quote/CLBT/profile/",
        "2026-08-17",
        "Profile page",
    ),
]

for i, row_data in enumerate(audit, 3):
    for j, val in enumerate(row_data, 1):
        c(ws4, i, j, val, border=True)

# ============================================================================
# Sheet 5: Questions
# ============================================================================
ws5 = wb.create_sheet("Questions")
ws5.merge_cells("A1:C1")
ws5["A1"].value = "Open Questions — CLBT"
ws5["A1"].font = title_font

for j, h in enumerate(["#", "Question", "Context"], 1):
    c(ws5, 2, j, h, font=Font(bold=True, underline='single'), border=True)

questions = [
    (
        "Q1",
        "What was the -" + D + "331.5M non-op charge in FY2024?",
        "Net Non-Operating Interest Income Expense was -"
        + D
        + "331.5M vs +"
        + D
        + "56.9M operating income. Impairment? Investment loss? Acquisition accounting?",
    ),
    (
        "Q2",
        "What drove the -" + D + "255M Investing CF minus CapEx gap in FY2024?",
        "Investing CF -"
        + D
        + "268.3M but CapEx only -"
        + D
        + "13.2M. The -"
        + D
        + "255M gap suggests material acquisition or asset purchase.",
    ),
    (
        "Q3",
        "Why equity jumped from " + D + "34M (FY2023) to " + D + "336M (FY2024)?",
        "Despite -"
        + D
        + "283M loss, equity 10x'd. Share count +15% (209M to 235M). Equity issuance / capital raise of ~"
        + D
        + "500M?",
    ),
    (
        "Q4",
        "How sustainable is the 30%+ FCF margin?",
        "FCF margin 28-34% in healthy years; FY2024 FCF of "
        + D
        + "121.6M was 30.3%. Can expansion OpEx or acquisitions compress this?",
    ),
    (
        "Q5",
        "Nature of CEO change and guidance cut (Aug 2026)?",
        "New CEO Shiven Ramji named; guidance lowered for 2026. Stock dropped 6% in one day. Execution concerns?",
    ),
    (
        "Q6",
        "Is share dilution ongoing?",
        "Shares 182.7M (FY2022) -> 250.8M (current) = 37% dilution. S-8 filing Mar 2, 2026 = active equity incentives.",
    ),
    (
        "Q7",
        "FY2023 non-operating charge of -" + D + "109M?",
        "Non-operating volatility: FY2022 +D"
        + "118.9M, FY2023 -D"
        + "109M, FY2024 -D"
        + "331.5M, FY2025 +D"
        + "23.5M. What drives this?",
    ),
    (
        "Q8",
        "43% insider ownership — concentrated risk?",
        "Insiders own 43.06%. Majority owner involvement? Insider selling history?",
    ),
    (
        "Q9",
        "Israel HQ and Middle East regulatory risk?",
        "HQ in Petah Tikva, Israel. Geopolitical exposure. Government customer concentration risk.",
    ),
    (
        "Q10",
        "AI narrative — legitimate or multiple inflation?",
        "Stock at " + D + "19.98 52W high likely on AI forensics narrative. With guidance cut, is story intact?",
    ),
]

for i, (num, q, ctx) in enumerate(questions, 3):
    c(ws5, i, 1, num, border=True)
    c(ws5, i, 2, q, border=True)
    c(ws5, i, 3, ctx, border=True)

# ============================================================================
# Sheet 6: Sources
# ============================================================================
ws6 = wb.create_sheet("Sources")
ws6.merge_cells("A1:C1")
ws6["A1"].value = "Sources — CLBT"
ws6["A1"].font = title_font

for j, h in enumerate(["#", "Description", "URL"], 1):
    c(ws6, 2, j, h, font=Font(bold=True, underline='single'), border=True)

sources = [
    ("1", "Yahoo Finance — CLBT Summary", "finance.yahoo.com/quote/CLBT/"),
    ("2", "Yahoo Finance — Income Statement", "finance.yahoo.com/quote/CLBT/financials/"),
    (
        "3",
        "Yahoo Finance — Balance Sheet",
        "finance.yahoo.com/quote/CLBT/balance-sheet/",
    ),
    ("4", "Yahoo Finance — Cash Flow", "finance.yahoo.com/quote/CLBT/cash-flow/"),
    (
        "5",
        "Yahoo Finance — Key Statistics",
        "finance.yahoo.com/quote/CLBT/key-statistics/",
    ),
    ("6", "Yahoo Finance — Profile", "finance.yahoo.com/quote/CLBT/profile/"),
    ("7", "Yahoo Finance — News", "finance.yahoo.com/quote/CLBT/news/"),
    (
        "8",
        "CNBC — US10Y Treasury",
        "cnbc.com/quotes/US10Y",
    ),
    (
        "9",
        "Simply Wall St. — CLBT CEO/guidance cut",
        "SimplyWallSt.com (Aug 2026)",
    ),
    (
        "10",
        "StockAnalysis.com — 404 for CLBT",
        "N/A",
    ),
]

for i, (num, desc, url) in enumerate(sources, 3):
    c(ws6, i, 1, num, border=True)
    c(ws6, i, 2, desc, border=True)
    c(ws6, i, 3, url, border=True)

# ============================================================================
# Column widths
# ============================================================================
for ws in [ws1, ws2, ws3, ws4, ws5, ws6]:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 55

# Save
output = "models/2026-08-17 Cellebrite Model.xlsx"
wb.save(output)
print("Saved: " + output)
