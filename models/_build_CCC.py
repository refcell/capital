#!/usr/bin/env python3
"""Build CCC Intelligent Solutions 6-sheet valuation model.
Data date: 2026-07-23  Sources: Yahoo Finance, CNBC
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = Workbook()
bold = Font(bold=True)
title_font = Font(bold=True, size=14)
sub_font = Font(bold=True, italic=True, size=10)
hdr_fill = PatternFill("solid", fgColor="D9E2F3")
alt_fill = PatternFill("solid", fgColor="F2F2F2")
green_fill = PatternFill("solid", fgColor="E2EFDA")
brd = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

def wc(ws, r, c, v, f=None, fl=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.border = brd
    if f: cell.font = f
    if fl: cell.fill = fl
    return cell

# ===== Sheet 1: Valuation =====
ws = wb.active
ws.title = "Valuation"
ws.merge_cells("A1:C1")
wc(ws, 1, 1, "CCC Intelligent Solutions Holdings Inc. - Valuation Summary", title_font)
ws.merge_cells("A2:C2")
wc(ws, 2, 1, "Data as of July 23, 2026 | Sources: Yahoo Finance, CNBC", sub_font)

blk = [
    ("Ticker","CCC"), ("Date","2026-07-23"), ("Price","$5.74"),
    ("Shares Outstanding (M)","586.94"), ("Market Cap ($B)","3.36"),
    ("Enterprise Value ($B)","4.69"), ("Total Debt ($B)","1.36"),
    ("Cash ($M)","36.9"), ("Net Debt ($B)","1.33"),
    ("Primary Lens","Forward P/E"), ("Stance","Watch"),
]
for i, (k, v) in enumerate(blk, 3):
    wc(ws, i, 1, k, bold); wc(ws, i, 2, v)
    if i % 2 == 0:
        for ci in (1,2): ws.cell(row=i, column=ci).fill = alt_fill

met = [
    ("P/E Trailing","95.50","Toxic - FY25 NI $0.4M on interest expense"),
    ("Forward P/E","13.64","On consensus $0.44 FY26 EPS"),
    ("P/S (TTM)","3.09","3.36B / 1.09B"),
    ("P/B","1.95","Common equity $1.79B"),
    ("P/FCF (TTM)","13.31","7.5% FCF yield"),
    ("EV/Revenue","4.31","4.69B / 1.09B"),
    ("EV/EBITDA","15.11","4.69B / 310.2M"),
]
wc(ws, 16, 1, "Valuation Multiples", bold, hdr_fill)
ws.merge_cells("A16:C16"); ws.cell(16,1).fill = hdr_fill
for i, (m, v, c) in enumerate(met, 17):
    wc(ws, i, 1, m, bold); wc(ws, i, 2, v); wc(ws, i, 3, c)
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 60

# ===== Sheet 2: WACC =====
ws2 = wb.create_sheet("WACC")
ws2.merge_cells("A1:D1")
wc(ws2, 1, 1, "WACC - CCC Intelligent Solutions", title_font)
ws2.merge_cells("A2:D2")
wc(ws2, 2, 1, "CAPM with market-value debt/equity weights", sub_font)

mc_b, ev_b, det_b = 3.36, 4.69, 1.36
eq_w = mc_b / (mc_b + det_b)
debt_w = det_b / (mc_b + det_b)
rfr, erp = 0.0470, 0.05
beta = 0.50
ke = rfr + beta * erp
kd = 71.0 / 1362
tax = 0.31
wacc = eq_w * ke + debt_w * kd * (1 - tax)
print(f"WACC: Rf={rfr:.2%}, Beta={beta}, Ke={ke:.2%}, Kd={kd:.2%}, WACC={wacc:.2%}")

wr = [
    ("Component","Value","Source","Notes"),
    ("Risk-Free Rate (10Y US)","4.70%","CNBC US10Y Jul 23",""),
    ("Equity Risk Premium","5.00%","Assumption",""),
    ("Beta (Levered, 5Y Monthly)","0.50","Yahoo Stats","Defensive SaaS"),
    ("Cost of Equity",f"{ke:.1%}","CAPM",""),
    ("Interest Expense FY25","$71.0M","Income Stmt","Up 50% YoY"),
    ("Total Debt","$1.36B","Key Stats MRQ","Up from $824M"),
    ("Cost of Debt (Pre-Tax)",f"{kd:.1%}","71/1362",""),
    ("Tax Rate","31%","FY24 ETR",""),
    ("Market Cap ($B)",f"{mc_b:.2f}","Yahoo Finance",""),
    ("Debt ($B)",f"{det_b:.2f}","Key Stats",""),
    ("Equity Weight",f"{eq_w:.1%}","MC/(MC+D)",""),
    ("Debt Weight",f"{debt_w:.1%}","D/(MC+D)",""),
    ("WACC",f"{wacc:.1%}","Computed",""),
]
last_row = 3 + len(wr) - 1
for ri, (a, b, c, d) in enumerate(wr, 3):
    f = bold if ri == 3 or ri == last_row else None
    fl = hdr_fill if ri == 3 else (green_fill if ri == last_row else None)
    for ci, val in enumerate([a, b, c, d], 1):
        wc(ws2, ri, ci, val, f, fl)
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 28
ws2.column_dimensions["D"].width = 40

# ===== Sheet 3: Scenarios =====
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells("A1:E1")
wc(ws3, 1, 1, "Scenario Analysis - CCC", title_font)
ws3.merge_cells("A2:E2")
wc(ws3, 2, 1, "Primary: Forward P/E; FCF multiple as cross-check", sub_font)

shares = 586.94
nd = 1325  # net debt $M

bear_pe, base_pe, bull_pe = 12, 14, 18
bear_e, base_e, bull_e = 0.35, 0.48, 0.62
btp = bear_e * bear_pe    # 4.20
btm = base_e * base_pe    # 6.72
blp = bull_e * bull_pe    # 11.16

brev = 1.06 * 1.05**5
brev5 = 1.06 * 1.075**5
blrev = 1.06 * 1.10**5
bfcf = brev * 1000 * 0.16
bafcf = brev5 * 1000 * 0.22
blfcf = blrev * 1000 * 0.28
bev, baev, blev = bfcf*8, bafcf*12, blfcf*16
beq = bev - nd
baeq = baev - nd
bleq = blev - nd
bftp = beq / shares
bafcp = baeq / shares
blfcp = bleq / shares

btgt = round(btp*0.8 + bftp*0.2, 2)
batgt = round(btm*0.8 + bafcp*0.2, 2)
bltgt = round(blp*0.8 + blfcp*0.2, 2)
cur = 5.74
wb_, wba, wbl = 0.20, 0.50, 0.30
fv = round(wb_*btgt + wba*batgt + wbl*bltgt, 2)

print(f"Bear: $P/E {btp:.2f}, FCF {bftp:.2f} => {btgt:.2f}")
print(f"Base: P/E {btm:.2f}, FCF {bafcp:.2f} => {batgt:.2f}")
print(f"Bull: P/E {blp:.2f}, FCF {blfcp:.2f} => {bltgt:.2f}")
print(f"FV: ${fv:.2f} ({(fv/cur-1)*100:+.1f}%)")

sr = [
    ("Metric","Bear (20%)","Base (50%)","Bull (30%)","Notes"),
    ("Revenue CAGR (5Y)","5.0%","7.5%","10.0%","Consensus ~9.8% FY26"),
    ("Terminal Revenue (5Y, $B)",f"{brev:.2f}",f"{brev5:.2f}",f"{blrev:.2f}",""),
    ("Terminal FCF Margin","16%","22%","28%","TTM 23.7%"),
    ("Terminal FCF ($M)",f"{bfcf:.0f}",f"{bafcf:.0f}",f"{blfcf:.0f}",""),
    ("Forward P/E Exit","12x","14x","18x","Peer SaaS 12-18x"),
    ("Terminal EPS","$0.35","$0.48","$0.62",""),
    ("P/E Target Price","$4.20","$6.72","$11.16",""),
    ("FCF Cross-Check",f"{bev:.0f}",f"{baev:.0f}",f"{blev:.0f}",""),
    ("Less: Net Debt ($M)",f"{nd}",f"{nd}",f"{nd}","EV - MC"),
    ("FCF Cross-Check Price",f"${beq/shares:.2f}",f"${baeq/shares:.2f}",f"${bleq/shares:.2f}",""),
    ("Target Price",f"${btgt:.2f}",f"${batgt:.2f}",f"${bltgt:.2f}","80/20 P/E vs FCF"),
    ("Upside from $5.74",f"{(btgt/cur-1)*100:+.0%}",f"{(batgt/cur-1)*100:+.0%}",f"{(bltgt/cur-1)*100:+.0%}",""),
    ("Weight","20%","50%","30%",""),
    ("Weighted Value/Share",f"${wb_*btgt:.2f}",f"${wba*batgt:.2f}",f"${wbl*bltgt:.2f}",""),
    ("Probability-Weighted FV","","",f"${fv:.2f}",""),
    ("Current Price","","","$5.74",""),
    ("Weighted Upside","","",f"{(fv/cur-1)*100:+.1%}",""),
]
last_s = 3 + len(sr) - 1
for ri, row in enumerate(sr, 3):
    f = bold if ri == 3 or ri >= last_s - 1 else None
    fl = hdr_fill if ri == 3 else (green_fill if ri >= last_s - 1 else None)
    for ci, val in enumerate(row, 1):
        wc(ws3, ri, ci, val, f, fl)
ws3.column_dimensions["A"].width = 25
for ci in range(2, 6):
    ws3.column_dimensions[get_column_letter(ci)].width = 18

# ===== Sheet 4: Actuals Source Audit =====
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells("A1:D1")
wc(ws4, 1, 1, "Actuals Source Audit - CCC", title_font)

au = [
    ("Data Point","Value","Source","Date/Notes"),
    ("Stock Price","$5.74","Yahoo Finance","Jul 23 2026 close"),
    ("Market Cap","$3.36B","Yahoo Stats","Jul 23"),
    ("Enterprise Value","$4.69B","Yahoo Stats","Jul 23"),
    ("Shares Outstanding","586.94M","Yahoo Stats","Jul 23"),
    ("Revenue TTM","$1.087B","Yahoo Income","TTM"),
    ("Revenue FY25","$1.057B","Yahoo Income","12/31/2025"),
    ("Revenue FY24","$944.8M","Yahoo Income","12/31/2024"),
    ("Gross Profit TTM","$800.7M","Yahoo Income","TTM"),
    ("Operating Income TTM","$153.4M","Yahoo Income","TTM"),
    ("Operating Income FY25","$93.8M","Yahoo Income","FY25"),
    ("Net Income TTM","$34.5M","Yahoo Income","TTM"),
    ("Net Income FY25","$0.41M","Yahoo Income","Near zero on interest"),
    ("Diluted EPS TTM","$0.06","Yahoo Income","TTM"),
    ("EBITDA TTM","$310.2M","Yahoo Income","TTM"),
    ("Total Debt","$1.36B","Yahoo Key Stats","MRQ"),
    ("Total Cash","$36.9M","Yahoo Key Stats","MRQ"),
    ("Common Equity","$1.79B","Balance Sheet","FY2025"),
    ("OCF TTM","$314.4M","Yahoo Cash Flow","TTM"),
    ("FCF TTM","$252.4M","Yahoo Cash Flow","TTM"),
    ("Capex TTM","$62.0M","Yahoo Cash Flow","TTM"),
    ("Buybacks TTM","$628.5M","Yahoo Cash Flow","Debt-funded"),
    ("10Y Treasury","4.704%","CNBC US10Y","Jul 23 2026"),
    ("Beta (5Y Monthly)","0.50","Yahoo Stats",""),
    ("Analyst Rev FY26","$1.16B","Yahoo Analysis","10 analysts"),
    ("Analyst EPS FY26","$0.44","Yahoo Analysis","11 analysts"),
    ("Analyst Rev FY27","$1.26B","Yahoo Analysis","10 analysts"),
    ("Analyst EPS FY27","$0.51","Yahoo Analysis","11 analysts"),
    ("Forward P/E","13.64x","Yahoo Stats","On FY26 EPS"),
    ("Earnings Date","Jul 30, 2026","Yahoo Profile","8:30 AM EDT"),
    ("52-Week Range","$4.08-$10.50","Yahoo Summary",""),
]
last_a = 2 + len(au) - 1
for ri, (a, b, c, d) in enumerate(au, 2):
    f = bold if ri == 2 else None
    fl = hdr_fill if ri == 2 else (alt_fill if ri % 2 == 0 else None)
    for ci, val in enumerate([a, b, c, d], 1):
        wc(ws4, ri, ci, val, f, fl)
ws4.column_dimensions["A"].width = 28
ws4.column_dimensions["B"].width = 16
ws4.column_dimensions["C"].width = 30
ws4.column_dimensions["D"].width = 45

# ===== Sheet 5: Questions =====
ws5 = wb.create_sheet("Questions")
ws5.merge_cells("A1:B1")
wc(ws5, 1, 1, "Open Questions - CCC", title_font)

qs = [
    ("Q1", "Debt spike: $824M FY24 to $1.36B MRQ ($538M increase). Maturity structure? Covenants? Term loans vs revolvers vs bonds?"),
    ("Q2", "Net Income collapse FY25: Op income $93.8M but NI $0.41M. Interest expense $71M (up 50%). Will interest continue rising?"),
    ("Q3", "Buyback sustainability: $1.56B total repurchases (FY24 + FY25 + TTM). At $5.74 this is 46% of MC. Can this continue without further leverage?"),
    ("Q4", "Negative tangible BV (-$1.18B): Goodwill/intangibles $2.97B. Impairment risk if growth slows?"),
    ("Q5", "China exposure: What % of revenue from CCC International? Tariff/data sovereignty risk?"),
    ("Q6", "vs Guidewire (GWRE): Market share? Differentiation in P&C workflow?"),
    ("Q7", "Short interest: 48.5M shares (8.6% float), up from 36.3M. Betting on leverage unsustainability?"),
    ("Q8", "Customer concentration and churn: SaaS $1.06B+ revenue - how many insurers/repairers? Renewal rate?"),
    ("Q9", "AI displacement risk: AI damage assessment from photos could bypass CCC workflow? Or enhancement?"),
    ("Q10", "Q1 FY26 EPS beat: $0.11 actual vs $0.09 est (+16.7%). Will Q2 continue? Post-beat squeeze candidate?"),
]
for i, (q, t) in enumerate(qs, 2):
    wc(ws5, i, 1, q, bold)
    ws5.merge_cells(f"B{i}:E{i}")
    wc(ws5, i, 2, t)
ws5.column_dimensions["A"].width = 8
ws5.column_dimensions["B"].width = 120

# ===== Sheet 6: Sources =====
ws6 = wb.create_sheet("Sources")
ws6.merge_cells("A1:C1")
wc(ws6, 1, 1, "Data Sources - CCC", title_font)

srcs = [
    ("1","Yahoo Finance - Summary","finance.yahoo.com/quote/CCC/"),
    ("2","Yahoo - Income Statement","finance.yahoo.com/quote/CCC/financials/"),
    ("3","Yahoo - Balance Sheet","finance.yahoo.com/quote/CCC/balance-sheet/"),
    ("4","Yahoo - Cash Flow","finance.yahoo.com/quote/CCC/cash-flow/"),
    ("5","Yahoo - Key Statistics","finance.yahoo.com/quote/CCC/key-statistics/"),
    ("6","Yahoo - Analyst Estimates","finance.yahoo.com/quote/CCC/analysis/"),
    ("7","Yahoo - Company Profile","finance.yahoo.com/quote/CCC/profile/"),
    ("8","CNBC - US 10Y Treasury","cnbc.com/quotes/US10Y"),
    ("9","StockAnalysis.com - 404","stockanalysis.com/ticker/CCC/"),
]
for i, (n, name, url) in enumerate(srcs, 2):
    wc(ws6, i, 1, n)
    wc(ws6, i, 2, name, bold)
    ws6.merge_cells(f"C{i}:D{i}")
    wc(ws6, i, 3, url)
ws6.column_dimensions["A"].width = 6
ws6.column_dimensions["B"].width = 50
ws6.column_dimensions["C"].width = 65

# Save
path = os.path.join("models", "2026-07-23 CCC Intelligent Solutions Model.xlsx")
wb.save(path)
print(f"\nSaved: {path}")
print(f"Sheets: {wb.sheetnames}")
from openpyxl import load_workbook
wv = load_workbook(path)
print(f"Verified: {wv.sheetnames}")
print("BUILD SUCCESS")
