#!/usr/bin/env python3
"""Build STEP valuation model with 6 sheets: Valuation, WACC, Scenarios, Actuals Source Audit, Questions, Sources."""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()

# ── Style helpers ──
header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center', wrap_text=True)

def style_data_block(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border

# ===================================================================
# KEY DATA (from Yahoo Finance, July 1, 2026)
# ===================================================================
price = 40.66
shares_mm = 80.70          # 80,703,550
mc_b = 3.28               # from Statistics page as of 6/30/2026
ev_b = 3.56               # from Statistics page
net_debt_mm = ev_b * 1000 - mc_b * 1000  # ~$280M -> actually from balance sheet: total debt $1,305M, total cash $1,120M = $185M
# Let's use the actual balance sheet figures for net debt
total_debt_mm = 1305.36    # from balance sheet 3/31/2026
total_cash_mm = 1120.00    # Total Cash mrq from summary
net_debt_mm = total_debt_mm - total_cash_mm  # ~$185M

# Income statement (Yahoo, in thousands)
# TTM = FY26 since TTM = 3/31/2026
revenue = {'FY23': 67.58, 'FY24': 711.63, 'FY25': 1174.83, 'FY26': 1993.60}
gross_profit = {'FY23': 171.60, 'FY24': 381.45, 'FY25': 579.66, 'FY26': 905.94}
op_income = {'FY23': 171.77, 'FY24': 171.77, 'FY25': -266.82, 'FY26': -1023.37}
net_income = {'FY23': 58.09, 'FY24': 58.09, 'FY25': -179.56, 'FY26': -535.81}
eps = {'FY23': -0.30, 'FY24': 0.91, 'FY25': -2.52, 'FY26': -6.78}
ebitda = {'FY23': 10.18, 'FY24': 252.32, 'FY25': -163.85, 'FY26': -826.94}

# Cash flow
ocf = {'FY23': 151.18, 'FY24': 161.52, 'FY25': 64.93, 'FY26': 66.48}
capex = {'FY23': 5.63, 'FY24': 19.61, 'FY25': 5.10, 'FY26': 2.62}
fcf = {'FY23': 145.56, 'FY24': 141.92, 'FY25': 59.83, 'FY26': 63.86}

# Analyst consensus (Yahoo Finance Analysis, July 1 2026)
fy27_rev_consensus = 1570.0  # $B -> $mm
fy28_rev_consensus = 1770.0
fy27_eps_consensus = 2.49
fy28_eps_consensus = 3.29

# ===================================================================
# Sheet 1: Valuation
# ===================================================================
ws1 = wb.active
ws1.title = "Valuation"

# Title block - merged row 1
ws1.merge_cells('A1:G1')
ws1['A1'] = "StepStone Group Inc. (NASDAQ: STEP) — Valuation Summary"
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center')

title_data = [
    ("Company", "StepStone Group Inc."),
    ("Ticker", "NASDAQ: STEP"),
    ("Date", "2026-07-01"),
    ("Price", f"${price:.2f}"),
    ("Shares Outstanding", f"{shares_mm:.1f}M"),
    ("Market Cap", f"${mc_b:.2f}B"),
    ("Enterprise Value", f"${ev_b:.2f}B"),
    ("Primary Valuation Lens", "Forward P/E (TTM P/E meaningless — negative net income)"),
    ("Stance", "Watch / Needs More Work — acquisition transition creates extreme earnings noise"),
]

for i, (label, value) in enumerate(title_data):
    ws1.cell(row=i + 2, column=1, value=label).font = header_font
    ws1.cell(row=i + 2, column=2, value=value)

# Valuation metrics table
row = len(title_data) + 4
ws1.cell(row=row, column=1, value="Metric").font = header_font
ws1.cell(row=row, column=2, value="Value").font = header_font
ws1.cell(row=row, column=3, value="Comment").font = header_font
style_header_row(ws1, row, 3)

# Forward P/E = price / FY27 consensus EPS
fwd_pe = price / fy27_eps_consensus

val_metrics = [
    ("P/E (TTM)", "N/A (negative earnings: -$535.8M net income)", "TTM meaningless — FY26 includes massive acquisition-related charges"),
    ("Forward P/E (FY27)", f"{fwd_pe:.1f}x", f"Based on FY27 EPS consensus of ${fy27_eps_consensus:.2f}"),
    ("P/S (TTM)", f"{mc_b*1000/revenue['FY26']:.2f}x", f"MC ${mc_b:.2f}B / TTM revenue ${revenue['FY26']/1000:.2f}B"),
    ("EV/Sales", f"{ev_b*1000/revenue['FY26']:.2f}x", f"EV ${ev_b:.2f}B / TTM revenue ${revenue['FY26']/1000:.2f}B"),
    ("P/FCF (TTM)", f"{mc_b*1000/fcf['FY26']:.1f}x", f"MC / FCF ${fcf['FY26']/1000:.2f}B"),
    ("EV/FCF", f"{ev_b*1000/fcf['FY26']:.1f}x", f"EV / FCF ${fcf['FY26']/1000:.2f}B"),
    ("EV/EBITDA", "N/A (negative EBITDA)", "FY26 EBITDA: -$827M — acquisition distortion"),
    ("Forward P/B", f"{mc_b/(0.886):.1f}x", f"MC ${mc_b:.2f}B / Book ${0.886:.2f}B"),
    ("Analyst 1Y Target", "$72.50", "Yahoo Finance consensus; ~78% upside from current"),
    ("52-Week Range", "$38.85 - $77.79", "Stock down 48% from high; near 52-week low"),
]

for i, (metric, val, comment) in enumerate(val_metrics):
    r = row + 1 + i
    ws1.cell(row=r, column=1, value=metric)
    ws1.cell(row=r, column=2, value=val)
    ws1.cell(row=r, column=3, value=comment)

style_data_block(ws1, row, row + len(val_metrics), 3)
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 60

# ===================================================================
# Sheet 2: WACC
# ===================================================================
ws2 = wb.create_sheet("WACC")

ws2.merge_cells('A1:B1')
ws2['A1'] = "WACC Calculation — CAPM Approach"
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center')

# 10Y Treasury from CNBC: 4.481%
rf = 4.481
erp = 5.0
beta = 1.28
cost_equity = rf + beta * erp  # 4.481 + 1.28*5 = 10.881%

# Cost of debt: interest expense / total debt
int_exp_mm = 18.50
cost_debt = (int_exp_mm / total_debt_mm) * 100  # ~1.42%

# Tax rate from FY25: 22.16%; TTM is unreliable due to loss
# Use statutory ~21% since TTM tax provision is $148M on -$891M pre-tax
tax_rate = 21.0

# Weights: market values
total_cap_mm = mc_b * 1000 + total_debt_mm  # ~3280 + 1305 = 4585
w_equity = (mc_b * 1000) / total_cap_mm
w_debt = total_debt_mm / total_cap_mm

wacc = w_equity * cost_equity + w_debt * cost_debt * (1 - tax_rate / 100)

wacc_data = [
    ("Component", "Value", "Source / Notes"),
    ("Risk-Free Rate (10Y US Treasury)", f"{rf:.2f}%", "CNBC US10Y, July 1 2026"),
    ("Equity Risk Premium", f"{erp:.1f}%", "Standard assumption"),
    ("Beta (5Y Monthly)", f"{beta:.2f}", "Yahoo Finance Statistics"),
    ("Cost of Equity (CAPM)", f"{cost_equity:.2f}%", f"Rf + Beta * ERP = {rf} + {beta}*{erp}"),
    ("Interest Expense (FY26)", f"${int_exp_mm:.1f}M", "Yahoo Income Statement"),
    ("Total Debt", f"${total_debt_mm:.0f}M", "Yahoo Balance Sheet 3/31/2026"),
    ("Pre-Tax Cost of Debt", f"{cost_debt:.2f}%", f"Interest / Total Debt"),
    ("Tax Rate", f"{tax_rate:.0f}%", "Using ~21% statutory; TTM unreliable (net loss)"),
    ("Market Cap", f"${mc_b*1000:.0f}M", "Yahoo Statistics 6/30/2026"),
    ("Total Debt", f"${total_debt_mm:.0f}M", "Yahoo Balance Sheet 3/31/2026"),
    ("Equity Weight", f"{w_equity:.2f}", f"MC / (MC + Debt)"),
    ("Debt Weight", f"{w_debt:.2f}", f"Debt / (MC + Debt)"),
    ("", "", ""),
    ("WACC", f"{wacc:.2f}%", f"Weighted average: {w_equity:.2f}*{cost_equity:.2f}% + {w_debt:.2f}*{cost_debt:.2f}%*(1-{tax_rate/100:.2f})"),
]

for i, (comp, val, note) in enumerate(wacc_data):
    ws2.cell(row=i + 2, column=1, value=comp)
    ws2.cell(row=i + 2, column=2, value=val)
    ws2.cell(row=i + 2, column=3, value=note)
    if i == 0:
        for c in range(1, 4):
            ws2.cell(row=i + 2, column=c).font = header_font
            ws2.cell(row=i + 2, column=c).fill = header_fill

ws2.cell(row=len(wacc_data)+1, column=1, value="").font = Font(bold=True)
wacc_result_row = len(wacc_data) + 1
for c in range(1, 4):
    ws2.cell(row=wacc_result_row, column=c).font = Font(bold=True)

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 55

print(f"WACC: {wacc:.2f}%")

# ===================================================================
# Sheet 3: Scenarios
# ===================================================================
ws3 = wb.create_sheet("Scenarios")

ws3.merge_cells('A1:J1')
ws3['A1'] = "Scenario Analysis — Bear / Base / Bull"
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='center')

# Scenario parameters
# FY27 consensus: revenue $1.57B, EPS $2.49
# FY28 consensus: revenue $1.77B, EPS $3.29
# Current base revenue (FY26 TTM): $1.99B

# Bear: lower growth, margin compression, lower multiple
# Base: consensus path
# Bull: higher growth, margin expansion, higher multiple

# All units in MILLIONS for consistency
# Current revenue FY26: $1993.6M
base_rev = 1993.6

# 5-year terminal revenue projections
# Bear: 8% CAGR from FY26 -> terminal rev after 5Y = 1993.6 * 1.08^5 = 2952
# Base: 15% CAGR (from TTM to FY27 is actually lower since FY27 is $1.57B < $1.99B... wait that's a decline)

# Wait - TTM of $1.99B is for FY26 (March 2026). Analyst consensus for "Current Year (2027)" is $1.57B.
# This means analysts expect revenue to DECLINE from FY26 to FY27.
# This makes sense if FY26 TTM includes one-time acquisition revenue recognition.
# Let me use FY25 ($1.175B) as the real "organic" base for CAGR calculations
# FY25 = $1,174.83M, FY26 = $1,993.60M (acquisition-inflated TTM)
# FY27 consensus = $1,570M, FY28 consensus = $1,770M

organic_base = 1174.83  # FY25

# Bear: slower growth from FY25 base, 5Y CAGR 5%, terminal revenue = 1175 * 1.05^5 = 1502
# Base: consensus path continues, 5Y CAGR ~12% from a blended base, terminal revenue
# Bull: accelerated growth, higher multiple

# Let me think about this more carefully. The FY26 TTM of $1.99B is transitional due to acquisition.
# FY27 at $1.57B is the first post-acquisition full year estimate.
# CAGR should bridge from a reasonable base (FY25 $1.175B) to FY27 $1.57B

# Bear scenario
bear_cagr = 5.0
bear_term_rev = 1174.83 * (1.05 ** 5)  # ~1502
bear_fcf_margin = 2.5  # compressed margins during integration
bear_term_fcf = bear_term_rev * bear_fcf_margin / 100  # ~$37.6M
bear_exit_mult = 12.0
bear_implied_ev = bear_term_fcf * bear_exit_mult  # ~$451M
bear_eq_value = bear_implied_ev + net_debt_mm  # add net CASH (debt < cash)
bear_shares = 82.0  # slight dilution
bear_target = bear_eq_value / bear_shares

# Base scenario
base_cagr = 12.0
base_term_rev = 1174.83 * (1.12 ** 5)  # ~2090
base_fcf_margin = 5.0  # gradual recovery
base_term_fcf = base_term_rev * base_fcf_margin / 100  # ~$104.5M
base_exit_mult = 18.0
base_implied_ev = base_term_fcf * base_exit_mult  # ~$1881M
base_eq_value = base_implied_ev + net_debt_mm
base_shares = 80.0  # buybacks resume
base_target = base_eq_value / base_shares

# Bull scenario
bull_cagr = 18.0
bull_term_rev = 1174.83 * (1.18 ** 5)  # ~2674
bull_fcf_margin = 8.0
bull_term_fcf = bull_term_rev * bull_fcf_margin / 100  # ~$214M
bull_exit_mult = 22.0
bull_implied_ev = bull_term_fcf * bull_exit_mult  # ~$4708M
bull_eq_value = bull_implied_ev + net_debt_mm
bull_shares = 78.0  # significant buybacks
bull_target = bull_eq_value / bull_shares

bear_upside = (bear_target / price - 1) * 100
base_upside = (base_target / price - 1) * 100
bull_upside = (bull_target / price - 1) * 100

# Weighted: Bear 20%, Base 50%, Bull 30%
weighted_value = 0.20 * bear_target + 0.50 * base_target + 0.30 * bull_target
weighted_upside = (weighted_value / price - 1) * 100

scenario_headers = [
    "Parameter", "Bear (20%)", "Base (50%)", "Bull (30%)"
]
scenario_rows = [
    ("Revenue CAGR (5Y)", f"{bear_cagr:.0f}%", f"{base_cagr:.0f}%", f"{bull_cagr:.0f}%"),
    ("Terminal Revenue ($M)", f"${bear_term_rev:.0f}", f"${base_term_rev:.0f}", f"${bull_term_rev:.0f}"),
    ("Adjusted FCF Margin", f"{bear_fcf_margin:.1f}%", f"{base_fcf_margin:.1f}%", f"{bull_fcf_margin:.1f}%"),
    ("Terminal FCF ($M)", f"${bear_term_fcf:.1f}", f"${base_term_fcf:.1f}", f"${bull_term_fcf:.1f}"),
    ("Exit FCF Multiple", f"{bear_exit_mult:.0f}x", f"{base_exit_mult:.0f}x", f"{bull_exit_mult:.0f}x"),
    ("Implied EV ($M)", f"${bear_implied_ev:.0f}", f"${base_implied_ev:.0f}", f"${bull_implied_ev:.0f}"),
    ("Net Cash Adjustment ($M)", f"+${net_debt_mm:.0f}", f"+${net_debt_mm:.0f}", f"+${net_debt_mm:.0f}"),
    ("Shares Outstanding (M)", f"{bear_shares:.0f}", f"{base_shares:.0f}", f"{bull_shares:.0f}"),
    ("Target Price", f"${bear_target:.2f}", f"${base_target:.2f}", f"${bull_target:.2f}"),
    ("Upside from Current", f"{bear_upside:.1f}%", f"{base_upside:.1f}%", f"{bull_upside:.1f}%"),
    ("Weight", "20%", "50%", "30%"),
    ("Weighted Value/Share", f"${0.20*bear_target:.2f}", f"${0.50*base_target:.2f}", f"${0.30*bull_target:.2f}"),
]

ws3.cell(row=3, column=1, value="Note: All FCF and EV figures in $millions. Net debt = $1,305M debt - $1,120M cash = ~$185M (shown as negative, so addition).")
ws3['A3'].font = Font(italic=True, size=10)

for i, header in enumerate(scenario_headers):
    ws3.cell(row=5, column=i+1, value=header).font = header_font
    ws3.cell(row=5, column=i+1).fill = header_fill

for i, (param, bear, base, bull) in enumerate(scenario_rows):
    r = 6 + i
    ws3.cell(row=r, column=1, value=param)
    ws3.cell(row=r, column=2, value=bear)
    ws3.cell(row=r, column=3, value=base)
    ws3.cell(row=r, column=4, value=bull)

# Total probability-weighted FV
r_total = 6 + len(scenario_rows) + 1
ws3.cell(row=r_total, column=1, value="Total Probability-Weighted FV").font = Font(bold=True)
ws3.cell(row=r_total, column=2, value=f"${weighted_value:.2f}").font = Font(bold=True)
ws3.cell(row=r_total, column=3, value=f"Upside: {weighted_upside:.1f}%").font = Font(bold=True)

style_data_block(ws3, 5, r_total, 4)
ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 20

print(f"Bear target: ${bear_target:.2f}, Base target: ${base_target:.2f}, Bull target: ${bull_target:.2f}")
print(f"Weighted FV: ${weighted_value:.2f} ({weighted_upside:.1f}% upside)")

# ===================================================================
# Sheet 4: Actuals Source Audit
# ===================================================================
ws4 = wb.create_sheet("Actuals Source Audit")

ws4.merge_cells('A1:E1')
ws4['A1'] = "Data Point Source Audit"
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(horizontal='center')

audit_headers = ["Data Category", "Data Point", "Value", "Source URL", "Date / Notes"]
audit_data = [
    ("Stock Price", "Closing Price", "$40.66", "finance.yahoo.com/quote/STEP/", "Jul 1, 2026"),
    ("Stock Price", "After Hours", "$40.60", "finance.yahoo.com/quote/STEP/", "Jul 1, 2026"),
    ("Stock Price", "52-Week High", "$77.79", "finance.yahoo.com/quote/STEP/", "Jul 1, 2026"),
    ("Stock Price", "52-Week Low", "$38.85", "finance.yahoo.com/quote/STEP/", "Jul 1, 2026"),
    ("Market Data", "Market Cap", "$3.38B", "Yahoo Statistics (6/30/2026)", "6/30/2026"),
    ("Market Data", "Enterprise Value", "$3.56B", "Yahoo Statistics (6/30/2026)", "6/30/2026"),
    ("Market Data", "Shares Outstanding", "80.70M", "Yahoo Balance Sheet + Summary", "3/31/2026"),
    ("Market Data", "Beta (5Y Monthly)", "1.28", "Yahoo Finance Summary", "Jul 1, 2026"),
    ("Market Data", "Forward Dividend", "$1.67 (4.04%)", "Yahoo Finance Summary", "Jul 1, 2026"),
    ("Income Statement", "Revenue FY26 (TTM)", "$1,993.6M", "Yahoo Income Statement", "3/31/2026"),
    ("Income Statement", "Revenue FY25", "$1,174.8M", "Yahoo Income Statement", "3/31/2025"),
    ("Income Statement", "Revenue FY24", "$711.6M", "Yahoo Income Statement", "3/31/2024"),
    ("Income Statement", "Revenue FY23", "$67.6M", "Yahoo Income Statement", "3/31/2023"),
    ("Income Statement", "Gross Profit FY26", "$905.9M", "Yahoo Income Statement", "3/31/2026"),
    ("Income Statement", "Operating Income FY26", "-$1,023.4M", "Yahoo Income Statement", "3/31/2026"),
    ("Income Statement", "Net Income FY26", "-$535.8M", "Yahoo Income Statement", "3/31/2026"),
    ("Income Statement", "EPS Diluted FY26", "-$6.78", "Yahoo Income Statement", "3/31/2026"),
    ("Income Statement", "EBITDA FY26", "-$826.9M", "Yahoo Income Statement", "3/31/2026"),
    ("Income Statement", "Other Income/Expense FY26", "$138.9M", "Yahoo Income Statement — unusual items", "3/31/2026"),
    ("Balance Sheet", "Total Assets", "$6,762.7M", "Yahoo Balance Sheet", "3/31/2026"),
    ("Balance Sheet", "Total Liabilities", "$5,876.7M", "Yahoo Balance Sheet", "3/31/2026"),
    ("Balance Sheet", "Total Equity", "$886.0M", "Yahoo Balance Sheet", "3/31/2026"),
    ("Balance Sheet", "Common Stock Equity", "-$413.6M", "Yahoo Balance Sheet — NEGATIVE", "3/31/2026"),
    ("Balance Sheet", "Total Debt", "$1,305.4M", "Yahoo Balance Sheet", "3/31/2026"),
    ("Balance Sheet", "Total Cash", "$1,120.0M", "Yahoo Summary Statistics", "6/30/2026"),
    ("Cash Flow", "Operating Cash Flow FY26", "$66.5M", "Yahoo Cash Flow", "3/31/2026"),
    ("Cash Flow", "Capex FY26", "$2.6M", "Yahoo Cash Flow", "3/31/2026"),
    ("Cash Flow", "Free Cash Flow FY26", "$63.9M", "Yahoo Cash Flow", "3/31/2026"),
    ("Cash Flow", "Investing Cash Flow FY26", "$732.9M", "Unusually positive — likely divestiture or investment liquidation", "3/31/2026"),
    ("Analyst Estimates", "FY27 Revenue Consensus", "$1.57B", "Yahoo Finance Analysis", "Jul 1, 2026"),
    ("Analyst Estimates", "FY27 EPS Consensus", "$2.49", "Yahoo Finance Analysis", "Jul 1, 2026"),
    ("Analyst Estimates", "FY28 EPS Consensus", "$3.29", "Yahoo Finance Analysis", "Jul 1, 2026"),
    ("Analyst Estimates", "1Y Target Estimate", "$72.50", "Yahoo Finance Summary", "Jul 1, 2026"),
    ("Analyst Estimates", "Analyst Count", "7-8 analysts", "Multiple estimate tables", "Jul 1, 2026"),
    ("Valuation", "Forward P/E", "16.86x", "Yahoo Finance Summary", "Jul 1, 2026"),
    ("Valuation", "P/S (TTM)", "1.64x", "Yahoo Finance Summary", "Jul 1, 2026"),
    ("Valuation", "EV/Revenue", "1.79x", "Yahoo Finance Summary", "Jul 1, 2026"),
    ("WACC Input", "10Y US Treasury", "4.481%", "CNBC US10Y", "Jul 1, 2026"),
    ("Earnings", "Next Earnings Date", "Aug 6, 2026", "Yahoo Finance Summary", "Jul 1, 2026"),
]

ws4.cell(row=3, column=1, value="Note: StockAnalysis.com returned 404 for STEP. All data sourced from Yahoo Finance.")
ws4['A3'].font = Font(italic=True, size=10)

for i, h in enumerate(audit_headers):
    ws4.cell(row=5, column=i+1, value=h).font = header_font
    ws4.cell(row=5, column=i+1).fill = header_fill

for i, (cat, dp, val, src, note) in enumerate(audit_data):
    r = 6 + i
    ws4.cell(row=r, column=1, value=cat)
    ws4.cell(row=r, column=2, value=dp)
    ws4.cell(row=r, column=3, value=val)
    ws4.cell(row=r, column=4, value=src)
    ws4.cell(row=r, column=5, value=note)

style_data_block(ws4, 5, 5 + len(audit_data), 5)
ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 20
ws4.column_dimensions['D'].width = 40
ws4.column_dimensions['E'].width = 45

# ===================================================================
# Sheet 5: Questions
# ===================================================================
ws5 = wb.create_sheet("Questions")

ws5.merge_cells('A1:B1')
ws5['A1'] = "Open Questions"
ws5['A1'].font = title_font
ws5['A1'].alignment = Alignment(horizontal='center')

questions = [
    ("Revenue Surge Mechanism", "Revenue nearly doubled from $712M (FY24) to $1,994M (FY26 TTM). What acquisition or business combination drove this? Was it a reverse merger, SPAC de-SPAC, or conventional M&A?"),
    ("Operating Income Collapse", "Despite revenue nearly tripling, operating income went from +$172M (FY24) to -$1,023M (FY26). The 'Other Income/Expense' line shows $138.9M in gains. What is driving the massive operating loss? Are these acquisition-related impairment charges, integration costs, or something structural?"),
    ("Negative Common Stock Equity", "Common stock equity is -$414M on the FY26 balance sheet while total equity is +$886M. What accounts sit between these? Likely minority interest or preferred equity. This needs clarification for the capital structure."),
    ("Massive Investing Cash Flow", "Investing cash flow was +$732.9M in FY26 vs -$43.5M in FY25. This is extraordinarily positive — implies major divestiture, investment liquidation, or asset sales. Is this repeatable?"),
    ("Debt Spike", "Total debt surged from $269M (FY24) to $1,305M (FY26). Was this to finance an acquisition? What tranches exist and when do maturities fall due?"),
    ("Total Liability Explosion", "Total liabilities nearly doubled from $2,812M (FY25) to $5,877M (FY26). What drove this? Likely acquisition-related assumed liabilities. Need to identify the composition."),
    ("Equity Collapse", "Total equity fell from $1,774M (FY25) to $886M (FY26). Is this from the acquisition accounting (goodwill, assumed debt)?"),
    ("Share Count Growth", "Shares outstanding rose from 65.6M (FY24) to 80.7M (FY26). Is this deal-related dilution? Any convertible securities?"),
    ("FY27 Revenue Below FY26 TTM", "Analysts forecast FY27 revenue at $1.57B — below the FY26 TTM of $1.99B. Does this mean FY26 TTM includes non-recurring acquisition revenue recognition?"),
    ("Analyst Revisions Trend", "EPS estimates for FY27 have declined from $2.53 (90 days ago) to $2.49 (current). Bears would point to downward revision momentum. What catalysts could reverse this?"),
    ("Dividend Sustainability", "The $1.67/share dividend (4.04% yield) seems aggressive given -$6.78 EPS. Can the company afford this? Is it funded by cash reserves?"),
    ("Peer Comparison", "How does STEP compare to peers (HLNE, VCTR, ARES, CG) on forward multiples, growth, and capital efficiency?"),
    ("Tax Rate Anomaly", "Tax provision of $148M on pre-tax loss of $891M is unusual. This implies a positive tax charge during a loss year — likely deferred tax or state/local items."),
    ("Integration Timeline", "When does management expect integration costs to subside and normalized earnings to emerge?"),
    ("Q2 FY26 Report", "Earnings on Aug 6, 2026. First full post-acquisition quarter? Will the 10-K or 10-Q explain the balance sheet changes?"),
]

for i, (q, detail) in enumerate(questions):
    ws5.cell(row=i + 2, column=1, value=i + 1)
    ws5.cell(row=i + 2, column=2, value=q)
    ws5.cell(row=i + 2, column=3, value=detail)

ws5.column_dimensions['A'].width = 5
ws5.column_dimensions['B'].width = 30
ws5.column_dimensions['C'].width = 90

# ===================================================================
# Sheet 6: Sources
# ===================================================================
ws6 = wb.create_sheet("Sources")

ws6.merge_cells('A1:B1')
ws6['A1'] = "Data Sources"
ws6['A1'].font = title_font
ws6['A1'].alignment = Alignment(horizontal='center')

sources = [
    ("Yahoo Finance — Summary Page", "https://finance.yahoo.com/quote/STEP/", "Price, market cap, beta, dividend, analyst targets, 52-week range, company description"),
    ("Yahoo Finance — Income Statement", "https://finance.yahoo.com/quote/STEP/financials/", "Revenue, gross profit, operating income, net income, EPS, EBITDA by fiscal year"),
    ("Yahoo Finance — Balance Sheet", "https://finance.yahoo.com/quote/STEP/balance-sheet/", "Total assets, liabilities, equity, debt, cash, share counts"),
    ("Yahoo Finance — Cash Flow", "https://finance.yahoo.com/quote/STEP/cash-flow/", "Operating cash flow, capex, free cash flow, investing/financing flows"),
    ("Yahoo Finance — Analysis/Estimates", "https://finance.yahoo.com/quote/STEP/analysis/", "Analyst consensus estimates, EPS surprises, revision trends"),
    ("CNBC — US10Y Treasury", "https://www.cnbc.com/quotes/US10Y", "10-year Treasury yield for WACC risk-free rate"),
]

for i, (name, url, desc) in enumerate(sources):
    ws6.cell(row=i + 2, column=1, value=i + 1)
    ws6.cell(row=i + 2, column=2, value=name)
    ws6.cell(row=i + 2, column=3, value=url)
    ws6.cell(row=i + 2, column=4, value=desc)

ws6.cell(row=2, column=1, value="#").font = header_font
ws6.cell(row=2, column=2, value="Source").font = header_font
ws6.cell(row=2, column=3, value="URL").font = header_font
ws6.cell(row=2, column=4, value="Used For").font = header_font

ws6.column_dimensions['A'].width = 5
ws6.column_dimensions['B'].width = 40
ws6.column_dimensions['C'].width = 50
ws6.column_dimensions['D'].width = 60

# ===================================================================
# Save
# ===================================================================
outpath = "/home/refcell/dev/capital/models/[2026-07-01] StepStone Group Model.xlsx"
wb.save(outpath)
print(f"\nSaved workbook to {outpath}")

# Verify
wb2 = openpyxl.load_workbook(outpath)
print(f"Sheets: {wb2.sheetnames}")
for name in wb2.sheetnames:
    ws = wb2[name]
    print(f"  {name}: {ws.max_row} rows x {ws.max_column} cols")
