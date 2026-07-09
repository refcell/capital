"""Build 6-sheet valuation model for Bloom Energy Corporation (BE).

Sources:
  - Yahoo Finance Income Statement / Balance Sheet / Cash Flow / Statistics / Analysis
  - CNBC US10Y: 4.581% (2026-07-08)
  - Quote date: 2026-07-08, close $254.29

All $ figures in millions unless noted otherwise.
StockAnalysis.com returned 404 for BE; Yahoo Finance is the primary source.
"""

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ───────────────────────────────────────────────────────────
HERE = Path(__file__).resolve().parent
XLSX = HERE / "[2026-07-08] Bloom Energy Model.xlsx"

# ── Raw inputs ───────────────────────────────────────────────────────
PRICE = 254.29
SHARES_M = 284.44          # implied shares outstanding (Yahoo Stats, 2026-07-08)
MC_B = 77.05              # $B market cap
EV_B = 77.51              # $B enterprise value
NET_DEBT_B = EV_B - MC_B  # ~$0.46B — EV minus MC as net debt proxy

# Income statement ($M, from Yahoo Finance /financials/)
# TTM       FY2025   FY2024   FY2023   FY2022
REV     = {"TTM": 2449.0, "FY2025": 2024.0, "FY2024": 1473.9, "FY2023": 1333.5, "FY2022": 1199.1}
GP      = {"TTM": 724.2,  "FY2025": 587.4,  "FY2024": 404.6,  "FY2023": 197.8,  "FY2022": 148.3}
GM_PCT  = {"TTM": 0.296,  "FY2025": 0.290,  "FY2024": 0.274,  "FY2023": 0.149,  "FY2022": 0.124}
OP_IN   = {"TTM": 164.1,  "FY2025": 72.8,   "FY2024": 22.9,   "FY2023": -208.9,"FY2022": -261.0}
NI      = {"TTM": 6.0,    "FY2025": -88.4,  "FY2024": -29.2,  "FY2023": -302.1,"FY2022": -301.4}
EPS_D   = {"TTM": -0.04,  "FY2025": -0.37,  "FY2024": -0.13,  "FY2023": -1.42,  "FY2022": -1.62}
EBITDA  = {"TTM": 112.7,  "FY2025": 20.1,   "FY2024": 89.3,   "FY2023": -135.1,"FY2022": -198.9}
FCF     = {"TTM": 229.6,  "FY2025": 57.2,   "FY2024": 33.1,   "FY2023": -456.3,"FY2022": -308.5}
OCF     = {"TTM": 298.2,  "FY2025": 113.9,  "FY2024": 92.0,   "FY2023": -372.5,"FY2022": -191.7}
CAPEX   = {"TTM": 68.7,   "FY2025": 56.8,   "FY2024": 58.9,   "FY2023": 83.7,   "FY2022": 116.8}

# Balance sheet ($M, from Yahoo Finance /balance-sheet/)
TOTAL_DEBT_FY25 = 2992.0    # FY2025 total debt
CASH_MQ = 2490.0            # Q1 FY2026 total cash per Yahoo Stats
# Note: Cash flow shows $2.5B issu + $1B repay TTM. Use EV-MC as net debt proxy.

# Cash flow ($M, from Yahoo Finance /cash-flow/)
# FCF, OCF, Capex above

# Tax, beta, rates
# Tax rate: TTM effective is very low because income is tiny ($2.75M / $12.7M = 21.6%)
# Use 21% corporate rate plus state = ~25% forward
TAX_RATE = 0.21            # statutory federal rate; FY2025 effective was 3.2% but income negative
BETA = 3.74                # 5Y monthly beta (Yahoo Stats)
RISK_FREE = 0.04581        # 10Y US Treasury, CNBC 2026-07-08
ERP = 0.05                 # equity risk premium

# Analyst estimates (Yahoo Finance /analysis/, 2026-07-08)
EPS_F2026 = 2.17           # 25 analysts
EPS_F2027 = 4.43           # 24 analysts
REV_F2026 = 3750.0         # 26 analysts
REV_F2027 = 6430.0         # 27 analysts

# Q1 FY2026 actual: Revenue $751.05M, Earnings $138.06M
# EPS Q1 actual: $0.44 vs estimate $0.13 (242% surprise)

# ── Computed ──────────────────────────────────────────────────────
COE = RISK_FREE + BETA * ERP  # CAPM cost of equity
COST_OF_DEBT = RISK_FREE + 0.03  # high-beta company; 300 bps spread

# Because net debt is tiny relative to MC (~$460M vs ~$77B), weights are ~all equity
EQUITY_WEIGHT = MC_B / (MC_B + max(NET_DEBT_B, 0.01))
DEBT_WEIGHT = max(NET_DEBT_B, 0.01) / (MC_B + max(NET_DEBT_B, 0.01))
WACC = EQUITY_WEIGHT * COE + DEBT_WEIGHT * COST_OF_DEBT * (1 - TAX_RATE)

PE_TRAIL = "N/A (-)"       # trailing EPS negative
PE_FWD_2026 = PRICE / EPS_F2026
PE_FWD_2027 = PRICE / EPS_F2027
PS = MC_B / (REV["TTM"] / 1000)
PFCF = MC_B / (FCF["TTM"] / 1000) if FCF["TTM"] else None
EV_FCF = EV_B / (FCF["TTM"] / 1000) if FCF["TTM"] else None
EV_SALES = EV_B / (REV["TTM"] / 1000)
EV_EBITDA_TTM = EV_B / (EBITDA["TTM"] / 1000) if EBITDA["TTM"] > 0 else None

# Earnings growth estimates (from Yahoo: BE vs S&P500)
# FY2026: BE +184.95%, S&P500 +23.72%
# FY2027: BE +104.70%, S&P500 +18.14%

# FCF sufficiency check: FCF * 10 vs net debt
# $229.6M * 10 = $2.3B >> $0.46B net debt → FCF framework is VALID
print(f"WACC = {WACC:.4f} ({WACC*100:.2f}%)")
print(f"COE = {COE:.4f}, CDE = {COST_OF_DEBT:.4f}")
print(f"Equity weight = {EQUITY_WEIGHT:.4f}, Debt weight = {DEBT_WEIGHT:.4f}")

# ── Styles ────────────────────────────────────────────────────────
HDR_FONT = Font(bold=True, size=14)
SUB_FONT = Font(bold=True, size=11)
BOLD = Font(bold=True)
PCT_FMT = '0.0%'
PCT2_FMT = '0.00%'
DOLB_FMT = '$#,##0.0" B"'
DOLM_FMT = '$#,##0.0" M"'
DOL1_FMT = '$#,##0.0'
DOL2_FMT = '$#,##0.00'
COMMA_FMT = '#,##0.0'
INT_FMT = '#,##0'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LIGHT_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
WHITE_FONT = Font(bold=True, color="FFFFFF")
BEAR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BASE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BULL_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")


def write_table(ws, start_row, headers, data, col_widths=None):
    """Write headers + data rows starting at start_row. Returns last row."""
    ncols = len(headers)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ri = start_row
    for row_data in data:
        ri += 1
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center' if ci > 1 else 'left',
                                      vertical='center')
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    return ri


# ── Build workbook ───────────────────────────────────────────────
wb = openpyxl.Workbook()

# ==================================================================
# Sheet 1: Valuation
# ==================================================================
ws1 = wb.active
ws1.title = "Valuation"
ws1.merge_cells('A1:E1')
cell = ws1['A1']
cell.value = "Bloom Energy Corporation (BE) — Valuation Model"
cell.font = HDR_FONT
cell.alignment = CENTER

# Title block
title_block = [
    ("Company:", "Bloom Energy Corporation"),
    ("Ticker:", "NYSE: BE"),
    ("Sector:", "Industrials — Electrical Equipment / Energy Technology"),
    ("Date:", "2026-07-08"),
    ("Price:", PRICE),
    ("Shares Outstanding:", f"{SHARES_M:.2f}M"),
    ("Market Cap:", f"${MC_B:.2f}B"),
    ("Enterprise Value:", f"${EV_B:.2f}B"),
    ("Primary Lens:", "Forward P/E + FCF Multiple + EV/EBITDA; Growth-transition play"),
    ("Stance:", "Cautiously Bullish — revenue inflection confirmed, but multiple is extreme"),
]
for i, (label, val) in enumerate(title_block, 2):
    ws1.cell(row=i, column=1, value=label).font = BOLD
    ws1.cell(row=i, column=2, value=val)

# Valuation metrics
ws1.cell(row=13, column=1, value="Key Valuation Metrics").font = SUB_FONT
metrics = [
    ("Metric", "Value", "Comment"),
    ("Trailing P/E", PE_TRAIL, "TTM diluted EPS = -$0.04; earnings just turned positive"),
    ("Forward P/E (FY2026)", round(PE_FWD_2026, 1), f"Consensus EPS ${EPS_F2026:.2f} (25 analysts)"),
    ("Forward P/E (FY2027)", round(PE_FWD_2027, 1), f"Consensus EPS ${EPS_F2027:.2f} (24 analysts)"),
    ("P/Sales", round(PS, 2), f"MC ${MC_B:.2f}B / TTM rev ${REV['TTM']/1000:.2f}B"),
    ("P/FCF (TTM)", round(PFCF, 1), f"TTM FCF ${FCF['TTM']:.0f}M; strong FCF generation"),
    ("EV/FCF (TTM)", round(EV_FCF, 1), f"EV ${EV_B:.2f}B / FCF ${FCF['TTM']:.0f}M"),
    ("EV/Sales", round(EV_SALES, 2), f"EV ${EV_B:.2f}B / Sales ${REV['TTM']/1000:.2f}B"),
    ("EV/EBITDA (TTM)", round(EV_EBITDA_TTM, 1) if EV_EBITDA_TTM else "N/A",
     f"EBITDA TTM ${EBITDA['TTM']:.0f}M; turnaround in progress"),
]
write_table(ws1, 14, metrics[0], metrics[1:], col_widths=[22, 18, 60])

# ==================================================================
# Sheet 2: WACC
# ==================================================================
ws2 = wb.create_sheet("WACC")
ws2.merge_cells('A1:D1')
cell = ws2['A1']
cell.value = "WACC Calculation — CAPM Method"
cell.font = HDR_FONT
cell.alignment = CENTER

wacc_data = [
    ("Component", "Value", "Source / Notes"),
    ("Risk-Free Rate (10Y US)", RISK_FREE, "CNBC US10Y, 2026-07-08"),
    ("Equity Risk Premium", ERP, "Standard assumption: 5%"),
    ("Beta (5Y Monthly)", BETA, "Yahoo Finance Statistics — extremely high (volatile, post-SPAC growth)" ),
    ("Cost of Equity (CAPM)", round(COE, 4),
     f"={RISK_FREE:.4f} + {BETA} x {ERP:.0%} = {COE:.4f}"),
    ("Cost of Debt (pre-tax)", round(COST_OF_DEBT, 4),
     "Risk-free + 300bps spread (high-beta industrials)"),
    ("Tax Rate", TAX_RATE, "Federal statutory 21%; FY effective unreliable (negative income)"),
    ("Market Cap ($B)", MC_B, "Yahoo Finance, 2026-07-08"),
    ("Net Debt ($B)", round(NET_DEBT_B, 2), "EV - MC proxy; ~0.46B — nearly net cash"),
    ("Equity Weight", round(EQUITY_WEIGHT, 4), "MC / (MC + Net Debt) — almost all equity"),
    ("Debt Weight", round(DEBT_WEIGHT, 4), "Net Debt / (MC + Net Debt) — negligible"),
    ("WACC", round(WACC, 4),
     f"={EQUITY_WEIGHT:.4f} x {COE:.4f} + {DEBT_WEIGHT:.4f} x {COST_OF_DEBT:.4f} x (1-{TAX_RATE:.0%})"),
]
write_table(ws2, 3, wacc_data[0], wacc_data[1:], col_widths=[30, 20, 60])

# ==================================================================
# Sheet 3: Scenarios
# ==================================================================
ws3 = wb.create_sheet("Scenarios")
ws3.merge_cells('A1:N1')
cell = ws3['A1']
cell.value = "Bear / Base / Bull Scenario Analysis — 5-Year Projection"
cell.font = HDR_FONT
cell.alignment = CENTER

# Scenario assumptions anchored in analyst consensus
# Consensus FY2026 rev: $3.75B (85% growth), FY2027 rev: $6.43B (72% growth)
# 5Y CAGR from $2.45B TTM base: if 70% growth, ~5Y ~$16B+ unrealistic
# We project 5 years from current TTM base

# Consensus revenue bridge:
# FY2026: $3.75B, FY2027: $6.43B
# After FY2027, growth slows to 20-35% as base gets large
# Bear: consensus low-end, then deceleration
# Base: consensus mid, then gradual normalization
# Bull: consensus high-end, sustained growth

# All in $M for consistency
BASE_REV = REV["TTM"]  # 2449.0 M

# Revenue projections (simple growth path based on consensus)
# Bear: FY2026 $3.52B (low est), FY2027 $4.39B (low est), then 5%/yr
# Base: FY2026 $3.75B, FY2027 $6.43B, then 15-20%
# Bull: FY2026 $4.28B (high), FY2027 $8.37B (high), then 25-30%

scenarios_raw = {
    "Bear": {
        "path": [3520.0, 4390.0, 5200.0, 5960.0, 6750.0],  # Y1-Y5 revenue
        "fcf_margin": 0.05,
        "exit_multiple": 25,  # Exit on FY5 FCF @ P/E; or EV/FCF
        "weight": 0.20,
        "fill": BEAR_FILL,
    },
    "Base": {
        "path": [3750.0, 6430.0, 8000.0, 9600.0, 11500.0],
        "fcf_margin": 0.08,
        "exit_multiple": 35,
        "weight": 0.50,
        "fill": BASE_FILL,
    },
    "Bull": {
        "path": [4280.0, 8370.0, 11000.0, 14000.0, 18000.0],
        "fcf_margin": 0.10,
        "exit_multiple": 45,
        "weight": 0.30,
        "fill": BULL_FILL,
    },
}

# Compute derived values — ALL in $M for consistency
# Terminal FCF = Terminal Revenue * FCF margin
# Implied EV = Terminal FCF * Exit Multiple
# Equity Value = Implied EV - Net Debt (all in $M)
# Target Price = Equity Value / Shares (in M)

NET_DEBT_M = NET_DEBT_B * 1000  # $460M

headers3 = ("Scenario", "Revenue CAGR (5Y)", "Terminal Revenue ($M)",
            "FCF Margin", "Terminal FCF ($M)", "Exit EV/FCF Multiple",
            "Implied EV ($M)", "Less Net Debt ($M)", "Shares (M)",
            "Target Price", "Upside %", "Weight", "Weighted $/Share")

rows3 = []
for name, s in [("Bear", scenarios_raw["Bear"]),
               ("Base", scenarios_raw["Base"]),
               ("Bull", scenarios_raw["Bull"])]:
    path = s["path"]
    term_rev = path[-1]  # Year 5 terminal revenue ($M)

    # CAGR from TTM base ($M) to terminal
    cagr = (term_rev / BASE_REV) ** (1/5) - 1

    term_fcf = round(term_rev * s["fcf_margin"], 1)  # $M
    implied_ev = round(term_fcf * s["exit_multiple"], 1)  # $M
    eq_value = implied_ev - NET_DEBT_M  # $M: EV - net debt
    target = round(eq_value / SHARES_M, 2) if eq_value > 0 else 0  # $/share
    upside = (target / PRICE) - 1
    weighted = round(target * s["weight"], 2)

    rows3.append((
        name, cagr, round(term_rev), s["fcf_margin"],
        round(term_fcf, 1), s["exit_multiple"],
        round(implied_ev), round(NET_DEBT_M, 0),
        SHARES_M, round(target, 2), upside, s["weight"], weighted
    ))

    print(f"{name}: CAGR={cagr:.1%}, TermRev=${term_rev:.0f}M, "
          f"FCF=${term_fcf:.0f}M, EV=${implied_ev:.0f}M, "
          f"Target=${target:.2f}, Upside={upside:.1%}")

# Write at row 3
vrow = 3
for ci, h in enumerate(headers3, 1):
    cell = ws3.cell(row=vrow, column=ci, value=h)
    cell.font = WHITE_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

for ri, row_data in enumerate(rows3, vrow + 1):
    for ci, val in enumerate(row_data, 1):
        cell = ws3.cell(row=ri, column=ci, value=val)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if ci == 1:
            cell.fill = scenarios_raw[row_data[0]]["fill"]
            cell.font = BOLD

# Summary
summary_row = vrow + len(rows3) + 1
ws3.cell(row=summary_row, column=1, value="Probability-Weighted FV").font = BOLD
weighted_fv = sum(r[len(headers3) - 1] for r in rows3)
ws3.cell(row=summary_row, column=10, value=round(weighted_fv, 2)).font = BOLD
ws3.cell(row=summary_row, column=11, value=round(weighted_fv / PRICE - 1, 4)).font = BOLD
for ci in range(1, len(headers3) + 1):
    ws3.cell(row=summary_row, column=ci).border = THIN_BORDER

# Note
note_row = summary_row + 1
ws3.cell(row=note_row, column=1,
         value="Note: Scenarios use revenue paths anchored to Yahoo Finance analyst consensus "
               "(FY2026 avg: $3.75B, FY2027 avg: $6.43B). Years 3-5 are extrapolated. "
               "Exit multiples are EV/FCF — forward P/E is the primary valuation lens.").font = Font(italic=True, size=9)

# Total probability check
prob_row = summary_row + 1
ws3.cell(row=prob_row, column=1, value="Total Probability").font = BOLD
ws3.cell(row=prob_row, column=12, value=1.00).font = BOLD
for ci in range(1, len(headers3) + 1):
    ws3.cell(row=prob_row, column=ci).border = THIN_BORDER

widths3 = [12, 16, 20, 12, 18, 18, 16, 16, 12, 14, 12, 10, 16]
for ci, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

# ==================================================================
# Sheet 4: Actuals Source Audit
# ==================================================================
ws4 = wb.create_sheet("Actuals Source Audit")
ws4.merge_cells('A1:E1')
cell = ws4['A1']
cell.value = "Data Source Audit — Every Data Point Traced"
cell.font = HDR_FONT
cell.alignment = CENTER

audit_headers = ("Data Point", "Value", "Source URL", "Date Access", "Notes")
audit_data = [
    ("Stock Price", "$254.29", "Yahoo Finance /quote/BE/", "2026-07-08",
     "Close at 4:00:03 PM EDT; -5.67% intraday"),
    ("Market Cap", "$77.05B", "Yahoo Finance Statistics", "2026-07-08", "Per Yahoo Stats valuation measures"),
    ("Enterprise Value", "$77.51B", "Yahoo Finance Statistics", "2026-07-08", "MC + net debt"),
    ("Shares Outstanding", "284.44M", "Yahoo Finance Statistics", "2026-07-08", "Implied; up from 229.1M FY2025"),
    ("TTM Revenue", "$2,449.0M", "Yahoo Finance /financials/", "2026-07-08", "Trailing 12 months; all in thousands"),
    ("FY2025 Revenue", "$2,024.0M", "Yahoo Finance /financials/", "2026-07-08", "FY 12/31/2025"),
    ("FY2024 Revenue", "$1,473.9M", "Yahoo Finance /financials/", "2026-07-08", "FY 12/31/2024"),
    ("FY2023 Revenue", "$1,333.5M", "Yahoo Finance /financials/", "2026-07-08", "FY 12/31/2023"),
    ("FY2022 Revenue", "$1,199.1M", "Yahoo Finance /financials/", "2026-07-08", "FY 12/31/2022"),
    ("FY2025 Gross Profit", "$587.4M", "Yahoo Finance /financials/", "2026-07-08", "Gross margin 29.0%"),
    ("FY2025 Operating Income", "$72.8M", "Yahoo Finance /financials/", "2026-07-08", "Turned positive from -$209M"),
    ("FY2025 Net Income", "-$88.4M", "Yahoo Finance /financials/", "2026-07-08",
     "Still negative; other expense -$137M drove losses"),
    ("TTM Net Income", "$6.0M", "Yahoo Finance /financials/", "2026-07-08",
     "Turned barely positive; EPS -$0.04 diluted"),
    ("TTM Diluted EPS", "-$0.04", "Yahoo Finance /financials/", "2026-07-08", "Near breakeven"),
    ("FY2025 Total Debt", "$2,992M", "Yahoo Finance /balance-sheet/", "2026-07-08",
     "Total debt, FY2025; up from $1,530M FY2024"),
    ("FY2025 Total Assets", "$4,397M", "Yahoo Finance /balance-sheet/", "2026-07-08",
     "Up from $2,657M; large step-up"),
    ("FY2025 Total Equity", "$793M", "Yahoo Finance /balance-sheet/", "2026-07-08", "Common equity $769M"),
    ("Total Cash Q1'26", "$2,490M", "Yahoo Finance Statistics", "2026-07-08",
     "Most recent quarter; per-shae $8.76"),
    ("TTM Operating Cash Flow", "$298.2M", "Yahoo Finance /cash-flow/", "2026-07-08", "OCF TTM"),
    ("TTM Capex", "$68.7M", "Yahoo Finance /cash-flow/", "2026-07-08", "Capex TTM"),
    ("TTM Free Cash Flow", "$229.6M", "Yahoo Finance /cash-flow/", "2026-07-08", "OCF - Capex"),
    ("Beta (5Y Monthly)", "3.74", "Yahoo Finance Statistics", "2026-07-08",
     "Extremely high — volatile growth stock"),
    ("FY2026 EPS Consensus", "$2.17", "Yahoo Finance /analysis/", "2026-07-08",
     "25 analysts; non-GAAP/normalized"),
    ("FY2027 EPS Consensus", "$4.43", "Yahoo Finance /analysis/", "2026-07-08",
     "24 analysts; up from $2.17 FY2026"),
    ("FY2026 Rev Consensus", "$3.75B", "Yahoo Finance /analysis/", "2026-07-08",
     "26 analysts; vs TTM $2.45B"),
    ("FY2027 Rev Consensus", "$6.43B", "Yahoo Finance /analysis/", "2026-07-08",
     "27 analysts; massive growth implied"),
    ("Q1 FY2026 Actual Rev", "$751.05M", "Yahoo Finance /analysis/", "2026-07-08",
     "Vs $401.24M Q1 2025"),
    ("Q1 FY2026 Actual EPS", "$0.44", "Yahoo Finance /analysis/", "2026-07-08",
     "Vs est $0.13; 242% surprise"),
    ("10Y US Treasury", "4.581%", "CNBC US10Y", "2026-07-08",
     "Yield at 10:08 PM EDT"),
    ("Next Earnings Date", "July 28, 2026", "Yahoo Finance Scout summary", "2026-07-08",
     "Q2 FY2026 earnings"),
    ("52-Week High/Low", "$351.28 / $24.04", "Yahoo Finance Statistics", "2026-07-08",
     "+785.72% 52-week change"),
]
write_table(ws4, 3, audit_headers, audit_data, col_widths=[28, 18, 38, 14, 55])

# ==================================================================
# Sheet 5: Questions
# ==================================================================
ws5 = wb.create_sheet("Questions")
ws5.merge_cells('A1:C1')
cell = ws5['A1']
cell.value = "Open Questions And Due Diligence Items"
cell.font = HDR_FONT
cell.alignment = CENTER

questions = [
    ("1.", "Massive Asset & Debt Step-Up FY2025",
     "Total assets jumped from $2.66B to $4.40B (+65%) and total debt from $1.53B to $2.99B (+95%) in FY2025. "
     "Cash flow shows $2.5B in debt issuance. What drove this? Was there an acquisition, facility purchase, or "
     "convertible bond issuance? The share count expanded 22% (229.1M -> 280.0M) suggesting equity financing or "
     "stock consideration for deals."),
    ("2.", "Earnings Quality — Other Expense Drag",
     "FY2025 operating income was $72.8M (3.6%) but net income was -$88.4M. 'Other income/expense' was -$137.4M, "
     "and net interest expense was -$19.8M. The 'other expense' is 36x larger than the prior year's -$12M. "
     "What is driving this? Restructuring charges? Impairments? Stock-based compensation? "
     "This is the biggest question for earnings quality."),
    ("3.", "FCF Turnaround Story",
     "FCF swung from -$456M (FY2023) to +$57M (FY2025) to +$230M TTM. But FY2025 FCF was only $57M on $2.02B "
     "revenue = 2.8% FCF margin. TTM FCF margin is 9.4% on $2.45B = $230M. Is the improvement from operating "
     "leverage, margin expansion, or reduced capex intensity?"),
    ("4.", "Share Count Dynamics",
     "Shares went from 205.7M (FY2022) to 229.1M (FY2025) to 280.0M issued (FY2025 end) and now 284.4M. "
     "The 22% jump in FY2025 is deal-related. The 1.6% jump from FY2025 to current is likely options/SBC. "
     "At $254/share and $77B MC, further dilution from employee equity programs is per-share material."),
    ("5.", "Competitive Position vs. Fuel Cell Rivals",
     "Bloom Energy makes solid-oxide fuel cells for stationary power (data centers, industrials). "
     "Rivals include FuelCell Energy (FCEL), Plug Power (PLUG), Ballard (BLDP), and Vertiv (VRT — "
     "power infrastructure). Bloom's advantage is commercialization — it's the only one shipping at scale. "
     "But is the moat defensible vs. battery + grid alternatives?"),
    ("6.", "Customer Concentration and Channel Risk",
     "Does Bloom Energy rely on a few large hyperscale or industrial customers? Data center demand could be "
     "concentrated. If 2-3 customers drop deployments, does revenue collapse?"),
    ("7.", "Valuation Multiple Sustainability",
     "At 120x forward P/E (FY2026) and 29x P/S, the stock is pricing in near-flawless execution. "
     "52-week change is +786%. Is this a bubble or justified by growth? The consensus EPS path from $2.17 -> "
     "$4.43 in one year is extraordinary. Any stumble could cause massive multiple compression."),
    ("8.", "Government Subsidy & Incentive Exposure",
     "Bloom Energy benefits from IRA (Inflation Reduction Act) tax credits for domestic manufacturing and "
     "clean hydrogen. How much of the current revenue and margin benefit is subsidy-driven vs. organic? "
     "If political winds change or credits expire, does the unit economics break?"),
    ("9.", "Debt Maturity Wall",
     "With $3.0B in total debt, what is the maturity profile? Are there near-term maturities requiring "
     "refinancing in a rising rate environment? The $2.5B issuance in FY2025/T TM must have terms — "
     "interest rates, covenants, maturity dates."),
    ("10.", "Gross Margin Expansion Driver",
     "Gross margin expanded from 12.4% (FY2022) to 29.0% (FY2025) to 29.6% TTM. What drove this? "
     "Scale economies? Product mix (more direct sales vs. third-party)? Component cost reduction? "
     "Is 30% a ceiling or is there more upside?"),
]

vrow = 2
for num, title, detail in questions:
    ws5.cell(row=vrow, column=1, value=num).font = BOLD
    ws5.cell(row=vrow, column=2, value=title).font = BOLD
    ws5.cell(row=vrow, column=3, value=detail)
    ws5.cell(row=vrow, column=3).alignment = Alignment(wrap_text=True, vertical='top')
    ws5.row_dimensions[vrow].height = 65
    vrow += 1

ws5.column_dimensions['A'].width = 6
ws5.column_dimensions['B'].width = 35
ws5.column_dimensions['C'].width = 105

# ==================================================================
# Sheet 6: Sources
# ==================================================================
ws6 = wb.create_sheet("Sources")
ws6.merge_cells('A1:C1')
cell = ws6['A1']
cell.value = "Data Sources"
cell.font = HDR_FONT
cell.alignment = CENTER

sources = [
    ("1.", "Yahoo Finance — BE Quote", "https://finance.yahoo.com/quote/BE/"),
    ("2.", "Yahoo Finance — BE Income Statement", "https://finance.yahoo.com/quote/BE/financials/"),
    ("3.", "Yahoo Finance — BE Balance Sheet", "https://finance.yahoo.com/quote/BE/balance-sheet/"),
    ("4.", "Yahoo Finance — BE Cash Flow", "https://finance.yahoo.com/quote/BE/cash-flow/"),
    ("5.", "Yahoo Finance — BE Key Statistics", "https://finance.yahoo.com/quote/BE/key-statistics/"),
    ("6.", "Yahoo Finance — BE Analyst Estimates", "https://finance.yahoo.com/quote/BE/analysis/"),
    ("7.", "Yahoo Finance — BE Profile", "https://finance.yahoo.com/quote/BE/profile/"),
    ("8.", "CNBC — US10Y 10-Year Treasury Yield", "https://www.cnbc.com/quotes/US10Y"),
    ("9.", "StockAnalysis.com — BE (returned 404; Yahoo Finance used as primary source)",
     "https://stockanalysis.com/quote/BE/"),
]

vrow = 2
for ci, h in enumerate(["#", "Description", "URL"], 1):
    cell = ws6.cell(row=vrow, column=ci, value=h)
    cell.font = WHITE_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

for num, desc, url in sources:
    vrow += 1
    ws6.cell(row=vrow, column=1, value=num)
    ws6.cell(row=vrow, column=2, value=desc)
    ws6.cell(row=vrow, column=3, value=url)
    for ci in range(1, 4):
        ws6.cell(row=vrow, column=ci).border = THIN_BORDER

ws6.column_dimensions['A'].width = 6
ws6.column_dimensions['B'].width = 55
ws6.column_dimensions['C'].width = 60

# ── Save and verify ──────────────────────────────────────
wb.save(str(XLSX))
print(f"\nSaved: {XLSX}")

# Verify
wv = openpyxl.load_workbook(str(XLSX))
print(f"Sheets: {wv.sheetnames}")
for sn in wv.sheetnames:
    ws = wv[sn]
    print(f"  {sn}: {ws.max_row} rows x {ws.max_column} cols")
wv.close()
print("\nModel build complete.")
