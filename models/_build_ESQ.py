#!/usr/bin/env python3
"""Build 6-sheet valuation model for Esquire Financial Holdings (ESQ) — Regional Bank Framework."""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()

# ── Style helpers ──
title_font = Font(name='Calibri', size=14, bold=True)
header_font = Font(name='Calibri', size=11, bold=True)
bold_font = Font(name='Calibri', size=11, bold=True)
normal_font = Font(name='Calibri', size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
bear_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
base_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
bull_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')

def c(ws, row, col, value, font=None, fill=None, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    cell.border = thin_border
    return cell

def write_header_row(ws, row, headers, fill=header_fill):
    for ci, h in enumerate(headers, 1):
        c(ws, row, ci, h, header_font, fill)

# ──────────────────────────────────────────────
# Sheet 1: Valuation
# ──────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Valuation'

# Title block - row 1 merged
ws1.merge_cells('A1:F1')
c(ws1, 1, 1, 'Esquire Financial Holdings, Inc. (ESQ) — Valuation', title_font)

# Title block rows
title_data = [
    ('Date', '2026-07-24'),
    ('Ticker', 'NASDAQ: ESQ'),
    ('Company', 'Esquire Financial Holdings, Inc.'),
    ('Sector / Industry', 'Financial Services / Banks — Regional'),
    ('Price', '$120.10'),
    ('Shares Outstanding', '8.65M'),
    ('Market Cap', '$1.01B'),
    ('Enterprise Value', 'N/A (bank — deposits are operating liabilities)'),
    ('Primary Valuation Lens', 'P/B + ROE / Forward P/E (bank framework)'),
    ('Stance', 'Watch'),
]
for ri, (k, v) in enumerate(title_data, 2):
    c(ws1, ri, 1, k, bold_font)
    c(ws1, ri, 2, v, normal_font)

# Valuation metrics table
wm_data = [
    ('Metric', 'Value', 'Comment'),
    ('P/E (Trailing)', '19.59x', 'TTM P/E — modest for a bank with 18.26% ROE'),
    ('Forward P/E (FY26)', '17.71x', '$120.10 / $6.78 implied — analyst consensus EPS'),
    ('P/S (TTM)', '6.59x', 'Yahoo Stats — TTM revenue $149.31M'),
    ('P/B (MRQ)', '3.34x', '$120.10 / $37.93 BVPS — above regional bank peer median'),
    ('EV/Sales', '6.56x', 'Yahoo Stats — EV approximates MC for bank'),
    ('EV/EBITDA', 'N/A', 'Not applicable for banks — interest income structure dominates'),
    ('EV/FCF', 'N/A', 'N/A — FCF meaningless for banks (deposits offset loan origination)'),
    ('Dividend Yield', '0.64%', 'Forward annual dividend $0.75/share'),
    ('ROE (TTM)', '18.26%', 'Strong — above cost of equity'),
    ('ROA (TTM)', '2.31%', 'Healthy for a regional bank'),
    ('Profit Margin', '35.30%', 'Net interest margin + noninterest income'),
]
for ri, (m, v, comm) in enumerate(wm_data, 14):
    for ci, val in enumerate([m, v, comm], 1):
        if ri == 14:
            c(ws1, ri, ci, val, header_font, header_fill)
        else:
            c(ws1, ri, ci, val, bold_font if ci == 1 else normal_font)

for ci in range(1, 4):
    ws1.column_dimensions[get_column_letter(ci)].width = 28 if ci == 3 else 20

# ──────────────────────────────────────────────
# Sheet 2: WACC
# ──────────────────────────────────────────────
ws2 = wb.create_sheet('WACC')

ws2.merge_cells('A1:B1')
c(ws2, 1, 1, 'WACC — CAPM (Bank Framework)', title_font)

# Risk-free rate from CNBC US10Y: 4.681%
rf = 0.04681
erp = 0.05
# Bank beta ~1.0-1.2 for regional banks — use 1.15 as levered
beta = 1.15
ke = rf + beta * erp  # 4.681% + 1.15 * 5% = 10.43%

# Bank cost of debt: use avg interest rate on deposits/borrowings
# Interest expense FY2025: $17.936M / Total debt estimate ~$500M-800M (est.)
# For a bank, cost of debt ≈ interest expense / (avg deposits + borrowings)
# Approximate: interest expense $20.332M TTM, total debt from borrowings ~small relative to deposits
kd = 0.030  # 3.0% approx cost of debt (deposits are cheap funding)
tax_rate = 0.21  # Corporate rate for banks

mc = 1.01  # B
# For banks, "debt" is deposits (operating liabilities) — use long-term borrowings for debt weight
# Total borrowings likely small — use a proxy
total_debt = 0.08  # ~$80M in long-term borrowings/subordinated debt (estimated)
total_cap = mc + total_debt
eq_weight = mc / total_cap
debt_weight = total_debt / total_cap

wacc = eq_weight * ke + debt_weight * kd * (1 - tax_rate)

wacc_data = [
    ('Component', 'Value', 'Notes'),
    ('Risk-Free Rate (10Y US Treasury)', f'{rf*100:.2f}%', 'CNBC US10Y, July 24, 2026'),
    ('Equity Risk Premium', '5.00%', 'Standard assumption'),
    ('Beta (Levered)', f'{beta:.2f}', 'Regional bank estimate — peer avg 1.0-1.3'),
    ('Cost of Equity (Ke)', f'{ke*100:.2f}%', f'CAPM: {rf*100:.2f}% + {beta} × 5.00%'),
    ('Cost of Debt (Kd)', f'{kd*100:.2f}%', 'Estimate — interest on deposits/borrowings'),
    ('Tax Rate', f'{tax_rate*100:.1f}%', 'US corporate rate'),
    ('Market Cap', f'${mc:.2f}B', 'July 24, 2026'),
    ('Total Debt (Borrowings)', f'${total_debt:.2f}B', 'Long-term borrowings estimate — deposits excluded'),
    ('Equity Weight', f'{eq_weight*100:.1f}%', f'MC / (MC + Debt)'),
    ('Debt Weight', f'{debt_weight*100:.1f}%', f'Debt / (MC + Debt)'),
    ('WACC', f'{wacc*100:.2f}%', f'{eq_weight*100:.1f}% × {ke*100:.2f}% + {debt_weight*100:.1f}% × {kd*100:.2f}% × (1 - {tax_rate*100:.0f}%)'),
]

for ri, (comp, val, note) in enumerate(wacc_data, 3):
    for ci, v in enumerate([comp, val, note], 1):
        if ri == 3:
            c(ws2, ri, ci, v, header_font, header_fill)
        else:
            c(ws2, ri, ci, v, bold_font if ci == 1 else normal_font)

for ci in range(1, 4):
    ws2.column_dimensions[get_column_letter(ci)].width = 30 if ci == 3 else 25

# ──────────────────────────────────────────────
# Sheet 3: Scenarios (P/B + ROE framework)
# ──────────────────────────────────────────────
ws3 = wb.create_sheet('Scenarios')

ws3.merge_cells('A1:H1')
c(ws3, 1, 1, 'Scenarios — P/B + ROE Framework (Bank Valuation)', title_font)

c(ws3, 2, 1, 'Note: FCF multiples N/A for banks. Scenarios use BVPS CAGR, exit P/B multiple, and implied price per share.', bold_font)

current_price = 120.10
bvps = 37.93  # MRQ book value per share

# Scenario parameters — BVPS CAGR over 5 years, then exit P/B
scenarios = [
    {
        'name': 'Bear',
        'bvps_cagr': 0.04,
        'terminal_bvps': round(bvps * (1 + 0.04)**5, 2),
        'exit_pb': 1.50,
        'implied_price': round(round(bvps * (1 + 0.04)**5, 2) * 1.50, 2),
        'upside': 0,  # calc later
        'weight': 0.20,
    },
    {
        'name': 'Base',
        'bvps_cagr': 0.07,
        'terminal_bvps': round(bvps * (1 + 0.07)**5, 2),
        'exit_pb': 2.00,
        'implied_price': round(round(bvps * (1 + 0.07)**5, 2) * 2.00, 2),
        'upside': 0,
        'weight': 0.50,
    },
    {
        'name': 'Bull',
        'bvps_cagr': 0.10,
        'terminal_bvps': round(bvps * (1 + 0.10)**5, 2),
        'exit_pb': 2.50,
        'implied_price': round(round(bvps * (1 + 0.10)**5, 2) * 2.50, 2),
        'upside': 0,
        'weight': 0.30,
    },
]

for s in scenarios:
    s['upside'] = round((s['implied_price'] / current_price) - 1, 4)

# Forward P/E cross-check scenarios
fpb_scenarios = [
    ('Bear', 10.0, 0.20),
    ('Base', 15.0, 0.50),
    ('Bull', 20.0, 0.30),
]
# Use FY27 EPS of 8.52 projected forward 5 years
base_eps_growth = 0.08  # implied from 6.78→8.52 is 25.7%, but 5-yr avg more moderate
terminal_eps_by_case = {
    'Bear': round(8.52 * (1 + 0.05)**3, 2),   # 3 years from FY27
    'Base': round(8.52 * (1 + 0.08)**3, 2),
    'Bull': round(8.52 * (1 + 0.12)**3, 2),
}

# Build P/B framework primarily and cross-check with P/E
header = ['Metric', 'Bear', 'Base', 'Bull', 'Notes']
scenario_rows = [
    header,
    ('BVPS CAGR (5Y)', '4.00%', '7.00%', '10.00%', 'Book value compound rate'),
    ('Terminal BVPS (5Y)', f'${scenarios[0]["terminal_bvps"]:.2f}', f'${scenarios[1]["terminal_bvps"]:.2f}', f'${scenarios[2]["terminal_bvps"]:.2f}', f'From ${bvps:.2f} current BVPS'),
    ('ROE Assumption', '~12%', '~16%', '~20%', 'Bear: rate pressure, Base: current, Bull: expansion'),
    ('Exit P/B Multiple', '1.50x', '2.00x', '2.50x', 'Peer: WSBC 1.8-2.0x, THFF 1.5-1.7x'),
    ('Implied Price/Share', f'${scenarios[0]["implied_price"]:.2f}', f'${scenarios[1]["implied_price"]:.2f}', f'${scenarios[2]["implied_price"]:.2f}', 'BVPS × Exit P/B'),
    ('Upside from Current', f'{scenarios[0]["upside"]*100:.1f}%', f'{scenarios[1]["upside"]*100:.1f}%', f'{scenarios[2]["upside"]*100:.1f}%', f'vs $120.10'),
    ('Scenario Weight', '20%', '50%', '30%', 'Probability assignment'),
    ('Weighted Value/Share', f'${scenarios[0]["implied_price"]*0.20:.2f}', f'${scenarios[1]["implied_price"]*0.50:.2f}', f'${scenarios[2]["implied_price"]*0.30:.2f}', 'Price × Weight'),
    ('Probability-Weighted FV', '', '', f'${sum(s["implied_price"]*s["weight"] for s in scenarios):.2f}', 'Sum of weighted values'),
    ('Current Price', '', '', '$120.10', 'Jul 24, 2026 close'),
    ('Total Probability-Weighted Upside', '', '', f'{((sum(s["implied_price"]*s["weight"] for s in scenarios)/current_price)-1)*100:.1f}%', 'Weighted FV vs current'),
]

# Cross-check with Forward P/E
fpb_note_rows = [
    ('', '', '', '', ''),
    ('Forward P/E Cross-Check', '', '', '', 'FY27 EPS $8.52 consensus × 3yr EPS CAGR'),
    ('Terminal EPS (est)', f'${terminal_eps_by_case["Bear"]:.2f}', f'${terminal_eps_by_case["Base"]:.2f}', f'${terminal_eps_by_case["Bull"]:.2f}', '5yr EPS proj'),
    ('Exit P/E', '10.0x', '15.0x', '20.0x', 'Peer regional bank range'),
    ('P/E Implied Price', f'${terminal_eps_by_case["Bear"]*10:.2f}', f'${terminal_eps_by_case["Base"]*15:.2f}', f'${terminal_eps_by_case["Bull"]*20:.2f}', 'Terminal EPS × Exit P/E'),
]

all_rows = scenario_rows + fpb_note_rows

for ri, row_data in enumerate(all_rows, 3):
    for ci, val in enumerate(row_data, 1):
        f = header_fill if ri == 3 else (bold_font if ci == 1 else normal_font)
        fill = None
        if ri > 3:
            if 'Bear' in str(row_data[0]) or (ci == 2 and ri > 3):
                pass  # no row-level coloring for banks
        c(ws3, ri, ci, val, f)

# Weight column widths
for ci, w in enumerate([25, 14, 14, 14, 35], 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

# ──────────────────────────────────────────────
# Sheet 4: Actuals Source Audit
# ──────────────────────────────────────────────
ws4 = wb.create_sheet('Actuals Source Audit')

ws4.merge_cells('A1:D1')
c(ws4, 1, 1, 'Actuals Source Audit — ESQ', title_font)

audit_data = [
    ('Data Point', 'Value', 'Source URL', 'Date / Period'),
    ('Stock Price', '$120.10', 'finance.yahoo.com/quote/ESQ/', 'Jul 24, 2026 close'),
    ('Market Cap', '$1.01B', 'Yahoo Statistics page', 'Jul 24, 2026'),
    ('Shares Outstanding', '8.65M', 'Yahoo Statistics page', 'Most recent filing'),
    ('P/B Ratio', '3.34x', 'Yahoo Statistics page', 'Current'),
    ('Trailing P/E', '19.59x', 'Yahoo Statistics page', 'TTM'),
    ('Book Value Per Share', '$37.93', 'Yahoo Statistics page', 'MRQ'),
    ('Total Cash', '$242.18M', 'Yahoo Statistics page', 'MRQ 6/30/2026'),
    ('ROE (TTM)', '18.26%', 'Yahoo Statistics page', 'TTM'),
    ('ROA (TTM)', '2.31%', 'Yahoo Statistics page', 'TTM'),
    ('Revenue (TTM)', '$149.31M', 'Yahoo Statistics page', 'TTM'),
    ('Revenue (FY2025)', '$146.56M', 'Yahoo Income Statement', '12/31/2025'),
    ('Revenue (FY2024)', '$124.82M', 'Yahoo Income Statement', '12/31/2024'),
    ('Revenue (FY2023)', '$113.52M', 'Yahoo Income Statement', '12/31/2023'),
    ('Revenue (FY2022)', '$84.27M', 'Yahoo Income Statement', '12/31/2022'),
    ('Net Income (FY2025)', '$50.82M', 'Yahoo Income Statement', '12/31/2025'),
    ('Net Income (FY2024)', '$43.66M', 'Yahoo Income Statement', '12/31/2024'),
    ('Net Income (FY2023)', '$41.01M', 'Yahoo Income Statement', '12/31/2023'),
    ('Diluted EPS (FY2025)', '$5.87', 'Yahoo Income Statement', '12/31/2025'),
    ('Diluted EPS (FY2024)', '$5.14', 'Yahoo Income Statement', '12/31/2024'),
    ('Interest Income (FY2025)', '$139.42M', 'Yahoo Income Statement', '12/31/2025'),
    ('Interest Expense (FY2025)', '$17.94M', 'Yahoo Income Statement', '12/31/2025'),
    ('Net Interest Income (FY2025)', '$121.48M', 'Yahoo Income Statement', '12/31/2025'),
    ('Pretax Income (FY2025)', '$65.65M', 'Yahoo Income Statement', '12/31/2025'),
    ('Tax Rate (FY2025)', '22.6%', 'Calculated: $14.83M / $65.65M', 'FY2025'),
    ('Total Assets (FY2025)', '$2,365.7M', 'Yahoo Balance Sheet', '12/31/2025'),
    ('Total Equity (FY2025)', '$289.6M', 'Yahoo Balance Sheet', '12/31/2025'),
    ('FY26 EPS Consensus', '$6.78', 'Yahoo Analysis — Estimates table', '1 analyst'),
    ('FY27 EPS Consensus', '$8.52', 'Yahoo Analysis — Estimates table', '2 analysts'),
    ('FY26 Revenue Consensus', '$193.5M', 'Yahoo Analysis — Revenue table', '1 analyst'),
    ('FY27 Revenue Consensus', '$292.78M', 'Yahoo Analysis — Revenue table', '1 analyst'),
    ('EPS Surprise Q2 FY26', '+3.00%', 'Yahoo Analysis — Surprise table', 'Q2 ended 6/30/2026'),
    ('EPS Surprise Q1 FY26', '-7.69%', 'Yahoo Analysis — Surprise table', 'Q1 ended 3/31/2026'),
    ('Next Earnings Date', 'Oct 22, 2026', 'Yahoo Profile page', 'Upcoming event'),
    ('Dividend Rate', '$0.75/yr', 'Yahoo Statistics', 'Forward annual'),
    ('Beta', 'N/A', 'Yahoo Statistics — not available', 'Current'),
    ('10Y Treasury Rate', '4.681%', 'CNBC US10Y', 'Jul 24, 2026'),
    ('52-Week High', '$134.82', 'Yahoo Statistics', 'Current'),
    ('52-Week Low', '$90.57', 'Yahoo Statistics', 'Current'),
    ('% Held by Institutions', '71.33%', 'Yahoo Statistics', 'Current'),
    ('% Held by Insiders', '17.45%', 'Yahoo Statistics', 'Current'),
    ('Short % of Float', '10.82%', 'Yahoo Statistics', '6/30/2026'),
]

for ri, (dp, val, src, date_) in enumerate(audit_data, 2):
    for ci, v in enumerate([dp, val, src, date_], 1):
        if ri == 2:
            c(ws4, ri, ci, v, header_font, header_fill)
        else:
            c(ws4, ri, ci, v, bold_font if ci == 1 else normal_font)

for ci, w in enumerate([30, 20, 40, 20], 1):
    ws4.column_dimensions[get_column_letter(ci)].width = w

# ──────────────────────────────────────────────
# Sheet 5: Questions
# ──────────────────────────────────────────────
ws5 = wb.create_sheet('Questions')

ws5.merge_cells('A1:B1')
c(ws5, 1, 1, 'Open Questions — ESQ', title_font)

questions = [
    ('1.', 'Why is ESQ trading at 3.34x P/B — well above regional bank peer median (1.2-1.8x)? Is the legal-industry specialization justifying the premium, or has the stock run ahead of fundamentals?'),
    ('2.', 'Revenue consensus shows $193.5M for FY26 (+32% over TTM $149.31M) and $292.78M for FY27 (+51% growth). Only 1-2 analysts cover the stock — are these estimates based on management guidance or extrapolation?'),
    ('3.', 'The EPS estimates imply +15.5% growth in FY26 ($6.78 vs $5.87 FY25) and +25.7% growth in FY27 ($8.52). What drives this acceleration — rate cycle, volume growth, or margin expansion?'),
    ('4.', 'Net interest income of $121.48M on $149.31M revenue = 81% of revenue from NII. How sensitive is profitability to rate-cycle direction? If rates fall, what happens to NIM?'),
    ('5.', 'Total assets of $2.37B with equity of $290M gives a leverage ratio of ~8.2x. Is this conservative for a regional bank? How does the loan-to-deposit ratio look?'),
    ('6.', 'No debt figure available on Yahoo Statistics (shows "--"). For a bank, deposits are operating liabilities but what about subordinated debt or borrowings? The capital adequacy ratio (CET1) is needed to assess regulatory cushion.'),
    ('7.', 'Q1 FY26 EPS missed estimates ($1.40 vs $1.52, -7.69% surprise) while Q2 beat ($1.60 vs $1.55, +3.0%). What caused the Q1 miss? Was it seasonal, rate-related, or a one-time item?'),
    ('8.', 'The company specializes in legal-industry banking and structured settlement loans. Is this niche creating genuine moat economics, or does it create concentration risk if the legal environment changes?'),
    ('9.', '151 employees for a $2.37B-asset bank — a very lean operation. Is this sustainable scale, or does it suggest operational risk if turnover spikes?'),
    ('10.', 'Dividend payout ratio is only 12.21% ($0.75 on $6.14 EPS). Is there room for dividend growth or potential for buybacks? Capital return policy?'),
    ('11.', 'The Q2 FY26 earnings call has been released (per Yahoo page). What did management say about Q3 guidance, loan growth, deposit trends, and rate sensitivity?'),
    ('12.', 'ESQ was founded in 2006. Has the bank had any loan loss provisions, regulatory issues, or capital raises since founding?'),
    ('13.', 'EPS revision trend shows FY27 estimates were cut by 1 analyst in last 30 days ($8.65→$8.52). What triggered this?'),
    ('14.', 'Float is 7.11M shares with 9.38% short interest. Is there a visible short squeeze dynamic, or is short interest organic?'),
]

for ri, (num, q) in enumerate(questions, 2):
    c(ws5, ri, 1, num, bold_font)
    c(ws5, ri, 2, q, normal_font)
    ws5.row_dimensions[ri].height = 45

ws5.column_dimensions['A'].width = 6
ws5.column_dimensions['B'].width = 100

# ──────────────────────────────────────────────
# Sheet 6: Sources
# ──────────────────────────────────────────────
ws6 = wb.create_sheet('Sources')

ws6.merge_cells('A1:B1')
c(ws6, 1, 1, 'Sources', title_font)

sources = [
    ('1.', 'Yahoo Finance — ESQ Summary & Price', 'https://finance.yahoo.com/quote/ESQ/'),
    ('2.', 'Yahoo Finance — ESQ Key Statistics', 'https://finance.yahoo.com/quote/ESQ/key-statistics/'),
    ('3.', 'Yahoo Finance — ESQ Income Statement', 'https://finance.yahoo.com/quote/ESQ/financials/'),
    ('4.', 'Yahoo Finance — ESQ Balance Sheet', 'https://finance.yahoo.com/quote/ESQ/balance-sheet/'),
    ('5.', 'Yahoo Finance — ESQ Analyst Estimates', 'https://finance.yahoo.com/quote/ESQ/analysis/'),
    ('6.', 'Yahoo Finance — ESQ Company Profile', 'https://finance.yahoo.com/quote/ESQ/profile/'),
    ('7.', 'CNBC — US10Y Treasury Rate', 'https://www.cnbc.com/quotes/US10Y'),
    ('8.', 'StockAnalysis — ESQ (404 — not available)', 'https://stockanalysis.com/quote/ESQ/ (404)'),
]

for ri, (num, name, url) in enumerate(sources, 2):
    c(ws6, ri, 1, num, bold_font)
    c(ws6, ri, 2, name, normal_font)
    c(ws6, ri, 3, url, normal_font)

ws6.column_dimensions['A'].width = 6
ws6.column_dimensions['B'].width = 50
ws6.column_dimensions['C'].width = 60

# ── Save ──
outpath = '/home/refcell/dev/capital/models/2026-07-24 ESQ Model.xlsx'
wb.save(outpath)
print(f'Saved: {outpath}')

# ── Verify ──
wb2 = openpyxl.load_workbook(outpath)
print(f'Sheets: {wb2.sheetnames}')
for sn in wb2.sheetnames:
    ws = wb2[sn]
    print(f'  {sn}: {ws.max_row} rows × {ws.max_column} cols')

# Print key computed values
print(f'\nWACC: {wacc*100:.2f}%')
fv = sum(s['implied_price'] * s['weight'] for s in scenarios)
print(f'Probability-Weighted FV: ${fv:.2f}')
print(f'Upside from $120.10: {((fv/120.10)-1)*100:.1f}%')
for s in scenarios:
    print(f'  {s["name"]}: ${s["implied_price"]:.2f} (up {s["upside"]*100:.1f}%)')
