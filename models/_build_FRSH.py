#!/usr/bin/env python3
"""Build the Freshworks six-sheet valuation workbook."""
from math import isclose
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT = Path(__file__).with_name("2026-09-02 Freshworks Model.xlsx")
PRICE = 13.46
SHARES = 260.92  # millions
MARKET_CAP = 3512.0  # millions
ENTERPRISE_VALUE = 2882.0  # millions
DEBT = 34.59  # millions, primarily leases
CASH = 664.11  # cash and marketable securities, millions
NET_CASH = CASH - DEBT
TTM_REVENUE = 903.87
TTM_FCF = 233.75
FY26_REVENUE = 965.0
FY26_FCF = 263.67

# Five-year discounted terminal FCF framework. Cash is conservatively reduced
# in scenarios because repurchases and acquisitions are consuming the balance.
SCENARIOS = {
    "Bear": {"rev_cagr": 0.07, "fcf_margin": 0.20, "multiple": 9.0, "net_cash": 100.0, "shares": 250.0, "discount": 0.105, "weight": 0.20},
    "Base": {"rev_cagr": 0.12, "fcf_margin": 0.245, "multiple": 12.0, "net_cash": 250.0, "shares": 240.0, "discount": 0.0925, "weight": 0.55},
    "Bull": {"rev_cagr": 0.15, "fcf_margin": 0.28, "multiple": 15.0, "net_cash": 350.0, "shares": 230.0, "discount": 0.085, "weight": 0.25},
}
for case in SCENARIOS.values():
    case["terminal_revenue"] = FY26_REVENUE * (1 + case["rev_cagr"]) ** 5
    case["terminal_fcf"] = case["terminal_revenue"] * case["fcf_margin"]
    case["terminal_ev"] = case["terminal_fcf"] * case["multiple"]
    case["terminal_price"] = (case["terminal_ev"] + case["net_cash"]) / case["shares"]
    case["target"] = case["terminal_price"] / (1 + case["discount"]) ** 5
weighted_value = sum(case["target"] * case["weight"] for case in SCENARIOS.values())

assert SCENARIOS["Bear"]["target"] < PRICE
assert abs(SCENARIOS["Base"]["target"] - 14.38) / 14.38 < 0.20

rf = 4.792
erp = 5.00
beta = 0.86
cost_equity = rf + beta * erp
pretax_debt_cost = 5.00
tax_rate = 24.0  # management's FY2026 projected non-GAAP tax rate
equity_weight = MARKET_CAP / (MARKET_CAP + DEBT)
debt_weight = 1 - equity_weight
wacc = equity_weight * cost_equity + debt_weight * pretax_debt_cost * (1 - tax_rate / 100)

wb = Workbook()
navy, blue, gold, gray, white = "17365D", "D9EAF7", "D8B34B", "E7E6E6", "FFFFFF"
thin = Side(style="thin", color="A6A6A6")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def put(ws, row, col, value, *, bold=False, fill=None, color="000000", size=10, wrap=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Aptos", size=size, bold=bold, color=color)
    cell.fill = PatternFill("solid", fgColor=fill) if fill else PatternFill(fill_type=None)
    cell.border = border
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    return cell


def title(ws, text, end_col=5, subtitle=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    put(ws, 1, 1, text, bold=True, fill=navy, color=white, size=15)
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        put(ws, 2, 1, subtitle, bold=True, fill=blue)


def header(ws, row, values):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=True, fill=navy, color=white)


def widths(ws, values):
    for col, width in enumerate(values, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A3"


# 1. Valuation
ws = wb.active
ws.title = "Valuation"
title(ws, "Freshworks (FRSH) — Valuation", 4, "Quote date: September 1, 2026 | Model date: September 2, 2026")
facts = [
    ("Company", "Freshworks Inc.", "AI-powered employee and customer service SaaS"),
    ("Ticker", "NASDAQ: FRSH", "Technology / application software"),
    ("Price", PRICE, "StockAnalysis September 1 close"),
    ("Shares outstanding (M)", SHARES, "Filing-date shares; current Class A share class is lower because Class B is included in total"),
    ("Market capitalization ($M)", MARKET_CAP, "StockAnalysis"),
    ("Enterprise value ($M)", ENTERPRISE_VALUE, "StockAnalysis"),
    ("Net cash ($M)", NET_CASH, "Cash and investments less total debt"),
    ("Primary valuation lens", "Discounted terminal FCF", "GAAP net income is distorted by a deferred-tax benefit; FCF is more representative"),
    ("Stance", "Watch", "Execution is improving, but the current price already discounts much of the near-term operating progress"),
]
header(ws, 3, ["Field", "Value", "Comment"])
for row, values in enumerate(facts, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))

header(ws, 15, ["Valuation metric", "Value", "Interpretation"])
metrics = [
    ("Trailing P/E", 20.61, "Misleading: TTM net income includes a large deferred-tax benefit"),
    ("Forward P/E", 17.49, "Based on non-GAAP adjusted EPS"),
    ("P/S", 3.89, "Moderate for mid-teens SaaS growth and mid-20s adjusted margin"),
    ("P/FCF", 15.02, "6.7% equity FCF yield before treating SBC as an economic cost"),
    ("EV/FCF", 12.33, "Primary current cash-flow cross-check"),
    ("EV/Sales", 3.19, "Net cash lowers enterprise valuation"),
    ("EV/EBITDA", 50.14, "GAAP EBITDA remains a weak denominator during the profitability transition"),
    ("FCF yield", TTM_FCF / MARKET_CAP, "$233.75M TTM FCF / $3.512B market cap"),
]
for row, values in enumerate(metrics, 16):
    for col, value in enumerate(values, 1):
        cell = put(ws, row, col, value, bold=(col == 1))
        if values[0] == "FCF yield" and col == 2:
            cell.number_format = "0.0%"
widths(ws, [30, 24, 82, 14])

# 2. WACC
ws = wb.create_sheet("WACC")
title(ws, "Freshworks — WACC", 4, "CAPM and debt-weighted cost of capital")
header(ws, 3, ["Component", "Value", "Source / formula"])
wacc_rows = [
    ("Risk-free rate", rf / 100, "CNBC U.S. 10-year Treasury on September 1, 2026"),
    ("Equity risk premium", erp / 100, "Model assumption"),
    ("Levered beta", beta, "StockAnalysis five-year beta"),
    ("Cost of equity", cost_equity / 100, "Risk-free rate + beta × ERP"),
    ("Pre-tax cost of debt", pretax_debt_cost / 100, "Normalized lease/debt assumption; debt is immaterial"),
    ("Tax rate", tax_rate / 100, "Company FY2026 projected non-GAAP tax rate"),
    ("Market cap ($M)", MARKET_CAP, "StockAnalysis"),
    ("Total debt ($M)", DEBT, "StockAnalysis standardized figure"),
    ("Equity weight", equity_weight, "Market cap / (market cap + debt)"),
    ("Debt weight", debt_weight, "Debt / (market cap + debt)"),
    ("After-tax debt cost", pretax_debt_cost / 100 * (1 - tax_rate / 100), "Kd × (1 − tax rate)"),
    ("WACC", wacc / 100, "E/V × Ke + D/V × Kd × (1 − t)"),
]
for row, values in enumerate(wacc_rows, 4):
    for col, value in enumerate(values, 1):
        cell = put(ws, row, col, value, bold=(col == 1 or values[0] == "WACC"), fill=(gold if values[0] == "WACC" else None))
        if col == 2 and values[0] not in {"Levered beta", "Market cap ($M)", "Total debt ($M)"}:
            cell.number_format = "0.00%"
widths(ws, [30, 20, 76, 14])

# 3. Scenarios
ws = wb.create_sheet("Scenarios")
title(ws, "Freshworks — Discounted Terminal FCF Scenarios", 6, "Five-year terminal value discounted to September 2026; all company figures in USD millions")
header(ws, 3, ["Metric", "Bear", "Base", "Bull", "Notes"])
scenario_rows = [
    ("Revenue CAGR (5Y)", *(SCENARIOS[k]["rev_cagr"] for k in ("Bear", "Base", "Bull")), "FY2026 consensus / guide midpoint is approximately $965M"),
    ("Terminal revenue ($M)", *(SCENARIOS[k]["terminal_revenue"] for k in ("Bear", "Base", "Bull")), "FY2026 base compounded five years"),
    ("Adjusted FCF margin", *(SCENARIOS[k]["fcf_margin"] for k in ("Bear", "Base", "Bull")), "Bear assumes SBC and competition pressure; bull assumes sustained Rule of 40 economics"),
    ("Terminal FCF ($M)", *(SCENARIOS[k]["terminal_fcf"] for k in ("Bear", "Base", "Bull")), "Terminal revenue × FCF margin"),
    ("Exit FCF multiple", *(SCENARIOS[k]["multiple"] for k in ("Bear", "Base", "Bull")), "9x / 12x / 15x"),
    ("Implied EV ($M)", *(SCENARIOS[k]["terminal_ev"] for k in ("Bear", "Base", "Bull")), "Terminal FCF × exit multiple"),
    ("Net cash adjustment ($M)", *(SCENARIOS[k]["net_cash"] for k in ("Bear", "Base", "Bull")), "Lower than current net cash to reflect acquisitions and continued repurchases"),
    ("Shares outstanding (M)", *(SCENARIOS[k]["shares"] for k in ("Bear", "Base", "Bull")), "Repurchases offset SBC at different rates"),
    ("Terminal price", *(SCENARIOS[k]["terminal_price"] for k in ("Bear", "Base", "Bull")), "Undiscounted year-five value per share"),
    ("Discount rate", *(SCENARIOS[k]["discount"] for k in ("Bear", "Base", "Bull")), "Scenario risk adjustment around CAPM WACC"),
    ("Present target price", *(SCENARIOS[k]["target"] for k in ("Bear", "Base", "Bull")), "Terminal price discounted five years"),
    ("Upside / (downside)", *((SCENARIOS[k]["target"] / PRICE - 1) for k in ("Bear", "Base", "Bull")), "Versus $13.46 close"),
    ("Probability", *(SCENARIOS[k]["weight"] for k in ("Bear", "Base", "Bull")), "20% / 55% / 25%"),
    ("Weighted value/share", *(SCENARIOS[k]["target"] * SCENARIOS[k]["weight"] for k in ("Bear", "Base", "Bull")), "Contribution to fair value"),
]
for row, values in enumerate(scenario_rows, 4):
    for col, value in enumerate(values, 1):
        cell = put(ws, row, col, value, bold=(col == 1))
        if col in (2, 3, 4) and values[0] in {"Revenue CAGR (5Y)", "Adjusted FCF margin", "Discount rate", "Upside / (downside)", "Probability"}:
            cell.number_format = "0.0%"
        elif col in (2, 3, 4) and values[0] in {"Terminal price", "Present target price", "Weighted value/share"}:
            cell.number_format = "$0.00"
put(ws, 19, 1, "Probability-weighted fair value", bold=True, fill=gold)
put(ws, 19, 2, weighted_value, bold=True, fill=gold).number_format = "$0.00"
put(ws, 20, 1, "Upside from current price", bold=True, fill=gold)
put(ws, 20, 2, weighted_value / PRICE - 1, bold=True, fill=gold).number_format = "0.0%"
put(ws, 22, 1, "Framework note", bold=True, fill=gray)
put(ws, 22, 2, "The model uses FCF rather than GAAP EPS because Freshworks' TTM net income includes a large deferred-tax benefit. FCF is not fully clean: SBC was $127.7M TTM, acquisitions consumed $75.4M, and repurchases consumed $409.7M. Scenario cash balances therefore assume current net cash is partly deployed.", fill=gray)
ws.merge_cells("B22:E22")
widths(ws, [31, 18, 18, 18, 84, 14])

# 4. Actuals Source Audit
ws = wb.create_sheet("Actuals Source Audit")
title(ws, "Freshworks — Actuals Source Audit", 5, "Financial statement figures in USD millions unless noted")
header(ws, 3, ["Data point", "Value", "Source URL", "Source date", "Notes"])
sa = "https://stockanalysis.com/stocks/frsh"
ir = "https://ir.freshworks.com/news/news-details/2026/Freshworks-Reports-Record-Second-Quarter-2026-Results/default.aspx"
audit = [
    ("Stock price", "$13.46", f"{sa}/", "2026-09-01", "Official close; after-hours excluded"),
    ("Market cap", "$3.512B", f"{sa}/financials/ratios/", "2026-09-01", "Daily market statistic"),
    ("Enterprise value", "$2.882B", f"{sa}/statistics/", "2026-09-02", "EV less MC is consistent with $629.5M net cash"),
    ("Shares outstanding", "260.92M", f"{sa}/statistics/", "2026-09-02", "Down 5.56% YoY"),
    ("Revenue TTM", "$903.87M", f"{sa}/financials/", "2026-06-30", "+15.57%"),
    ("Gross profit TTM", "$767.96M", f"{sa}/financials/", "2026-06-30", "84.96% margin"),
    ("Operating income TTM", "$37.50M", f"{sa}/financials/", "2026-06-30", "4.15% GAAP margin"),
    ("Net income TTM", "$185.20M", f"{sa}/financials/", "2026-06-30", "Distorted by a large deferred-tax benefit"),
    ("Cash and investments", "$664.11M", f"{sa}/financials/balance-sheet/", "2026-06-30", "$494.67M cash plus $169.44M short-term investments"),
    ("Total debt", "$34.59M", f"{sa}/financials/balance-sheet/", "2026-06-30", "Primarily leases"),
    ("Deferred revenue", "$403.37M", f"{sa}/financials/balance-sheet/", "2026-06-30", "Current plus long-term"),
    ("Operating cash flow TTM", "$246.72M", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "27.3% of revenue"),
    ("Capital expenditures TTM", "$12.97M", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "Low physical capital intensity"),
    ("Free cash flow TTM", "$233.75M", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "25.86% margin"),
    ("Stock-based compensation TTM", "$127.65M", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "54.6% of reported FCF"),
    ("Repurchases TTM", "$409.70M", f"{sa}/financials/cash-flow-statement/", "2026-06-30", "Exceeds FCF and reduces net cash"),
    ("FY2026 revenue consensus", "$965.0M", f"{sa}/forecast/", "2026-08-09", "16 analysts; +15.0%"),
    ("FY2026 adjusted EPS", "$0.67", f"{sa}/forecast/", "2026-08-09", "Non-GAAP diluted; range $0.66-$0.68"),
    ("FY2027 headline revenue", "$1.10B", f"{sa}/forecast/", "2026-08-09", "+14.2%; detailed table gated"),
    ("FY2027 headline EPS", "$0.83", f"{sa}/forecast/", "2026-08-09", "+24.4%; non-GAAP"),
    ("Average analyst target", "$14.38", f"{sa}/forecast/", "2026-08-09", "16 analysts; $10-$25 range"),
    ("Beta", "0.86", f"{sa}/statistics/", "2026-09-02", "Five-year beta"),
    ("U.S. 10Y Treasury", "4.792%", "https://www.cnbc.com/2026/09/01/bonds-treasury-yields-middle-east-tensions.html", "2026-09-01", "Risk-free-rate input"),
    ("Q2 revenue", "$237.377M", ir, "2026-06-30", "+16% YoY"),
    ("Q2 non-GAAP operating income", "$55.928M", ir, "2026-06-30", "23.6% margin"),
    ("Q2 adjusted FCF", "$57.659M", ir, "2026-06-30", "24.3% margin"),
    ("Q2 NDR", "104%", ir, "2026-06-30", "Down from 106% in Q1 2026 and Q2 2025"),
    (">$100K ARR customers", "1,746", ir, "2026-06-30", "+25% YoY"),
    ("Freddy AI enterprise attach", ">71%", ir, "2026-06-30", "Share of new enterprise deals using Copilot"),
]
for row, values in enumerate(audit, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
widths(ws, [31, 21, 95, 16, 68])

# 5. Questions
ws = wb.create_sheet("Questions")
title(ws, "Freshworks — Open Questions", 4, "Items that can change valuation or conviction")
header(ws, 3, ["#", "Question", "Why it matters", "Best evidence / next check"])
questions = [
    (1, "How much of the 104% NDR reflects smaller-customer churn versus contraction among larger enterprises?", "Expansion is only modestly above 100%; durable mid-teens growth requires new-logo velocity or better expansion.", "Cohort ARR and product-level retention disclosure."),
    (2, "Can EX ARR sustain mid-20s growth after the current enterprise adoption wave?", "Freshservice appears to be the fastest-growing and most strategically valuable franchise.", "Quarterly EX ARR and large-deal growth."),
    (3, "How much incremental ARR and revenue does Freddy AI produce rather than protect?", "A 71% attach rate is encouraging, but attach does not prove monetization or net retention uplift.", "AI ARR, usage pricing and cohort uplift."),
    (4, "Why did Q2 non-GAAP EPS decline to $0.17 from $0.18 despite revenue and operating-income growth?", "Tax normalization and share-count mechanics can obscure core operating leverage.", "Detailed non-GAAP net-income bridge."),
    (5, "What portion of the TTM $185.2M GAAP net income is repeatable after the deferred-tax benefit?", "Trailing P/E materially overstates normalized GAAP earnings power.", "Tax footnote and valuation allowance release schedule."),
    (6, "Is $127.7M of TTM SBC declining quickly enough relative to revenue and FCF?", "SBC equals 14.1% of revenue and 54.6% of reported FCF; it is a real economic cost.", "SBC by function and grant-date dilution."),
    (7, "Can repurchases remain accretive after spending $409.7M TTM, well above $233.8M FCF?", "Buybacks reduced shares but also drove net cash down 29% YoY.", "Average repurchase price, authorization and cash floor."),
    (8, "What drove the $75.4M TTM acquisition outflow and $51.3M increase in goodwill since FY2025?", "Acquisitions may accelerate product breadth but can dilute organic returns and complicate FCF quality.", "FireHydrant purchase accounting and revenue contribution."),
    (9, "How should Device42 and FireHydrant cross-sell into Freshservice be measured?", "The platform thesis depends on attaching discovery and incident management, not just owning more products.", "Attach rates, win rates and bundle pricing."),
    (10, "Does FedRAMP In Process convert into meaningful public-sector bookings?", "Government can lengthen duration and increase switching costs, but sales cycles are long.", "Authorization timeline and federal pipeline."),
    (11, "How durable is gross margin above 85% as AI inference and cloud usage scale?", "AI monetization can expand ARR while compressing gross margin if pricing lags compute cost.", "AI gross margin and hosting-cost trend."),
    (12, "Will the May 2026 restructuring produce durable G&A leverage without slowing product execution?", "Q2 GAAP profitability included restructuring charges and a product leadership transition.", "Headcount, hiring plan and expense guidance."),
    (13, "What does the CPTO transition from Srini Raghavan to Ryan Manning change in product priorities?", "Leadership change during an AI platform transition creates execution and retention risk.", "Product roadmap and senior engineering attrition."),
    (14, "How does Freshworks win against ServiceNow upmarket and Atlassian/Zendesk in the mid-market?", "The value proposition is simplicity; enterprise expansion can erode that advantage.", "Competitive win/loss rates and implementation time."),
    (15, "When is the next earnings release after the already-reported August 4 quarter?", "Q3 will test 14%-15% guide growth, AI attach and whether NDR stabilizes.", "Investor-relations calendar and Q3 release."),
]
for row, values in enumerate(questions, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
widths(ws, [7, 80, 68, 62])

# 6. Sources
ws = wb.create_sheet("Sources")
title(ws, "Freshworks — Sources", 4, "Public sources accessed September 2, 2026")
header(ws, 3, ["#", "Source", "URL", "Use"])
sources = [
    (1, "StockAnalysis overview", f"{sa}/", "Price, profile, headline operating and market data"),
    (2, "StockAnalysis financials", f"{sa}/financials/", "Historical income statement, segments, cash/debt and margins"),
    (3, "StockAnalysis balance sheet", f"{sa}/financials/balance-sheet/", "Cash, investments, debt, deferred revenue and equity"),
    (4, "StockAnalysis cash flow", f"{sa}/financials/cash-flow-statement/", "OCF, capex, FCF, SBC, acquisitions and repurchases"),
    (5, "StockAnalysis statistics", f"{sa}/statistics/", "Valuation, EV, beta, shares and analyst summary"),
    (6, "StockAnalysis forecast", f"{sa}/forecast/", "Revenue, adjusted EPS and price-target consensus"),
    (7, "StockAnalysis ratios", f"{sa}/financials/ratios/", "Historical valuation and yield ratios"),
    (8, "StockAnalysis profile", f"{sa}/company/", "Company identity, products, management and recent filings"),
    (9, "Freshworks Q2 2026 results", ir, "Quarterly actuals, operating metrics, guidance and non-GAAP reconciliations"),
    (10, "CNBC U.S. 10-year Treasury", "https://www.cnbc.com/2026/09/01/bonds-treasury-yields-middle-east-tensions.html", "Risk-free rate"),
    (11, "Freshworks investor relations", "https://ir.freshworks.com/", "Primary company materials and future updates"),
]
for row, values in enumerate(sources, 4):
    for col, value in enumerate(values, 1):
        put(ws, row, col, value, bold=(col == 1))
widths(ws, [7, 36, 105, 68])

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = sheet.dimensions

wb.save(OUT)
check = load_workbook(OUT, data_only=False)
expected = ["Valuation", "WACC", "Scenarios", "Actuals Source Audit", "Questions", "Sources"]
assert check.sheetnames == expected, check.sheetnames
assert isclose(check["Scenarios"]["B19"].value, weighted_value, rel_tol=1e-12)
print(f"WACC: {wacc:.2f}%")
for name, case in SCENARIOS.items():
    print(f"{name}: ${case['target']:.2f} ({case['target'] / PRICE - 1:.1%})")
print(f"Probability-weighted fair value: ${weighted_value:.2f} ({weighted_value / PRICE - 1:.1%})")
print(f"Wrote and verified {OUT}")
