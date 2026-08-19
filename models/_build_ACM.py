#!/usr/bin/env python3
"""Build 6-sheet ACM (AECOM) valuation model."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()

# ---- Style helpers ----
hdr = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)
subtitle_font = Font(bold=True, size=11, italic=True)
money_fmt = '#,##0.00'
pct_fmt = '0.00%'
num_fmt = '#,##0'
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

def c(ws, row, col, val, font=None, fmt=None, border=None, fill=None):
    cell = ws.cell(row=row, column=col, value=val)
    if font: cell.font = font
    if fmt: cell.number_format = fmt
    if border: cell.border = border
    if fill: cell.fill = fill
    return cell

def write_table(ws, start_row, headers, data, col_widths=None):
    for ci, h in enumerate(headers, 1):
        cell = c(ws, start_row, ci, h, font=hdr, fill=header_fill, border=thin_border)
    for ri, row_data in enumerate(data, start_row + 1):
        for ci, val in enumerate(row_data, 1):
            f = light_fill if (ri - start_row) % 2 == 0 else None
            cell = c(ws, ri, ci, val, border=thin_border, fill=f)
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

# =====================================================================
# Sheet 1: Valuation
# =====================================================================
ws1 = wb.active
ws1.title = 'Valuation'
ws1.merge_cells('A1:F1')
c(ws1, 1, 1, 'AECOM (ACM) - Valuation Summary', font=title_font)

# Title block
title_block = [
    ('Company', 'AECOM'),
    ('Ticker', 'NYSE: ACM'),
    ('Date', str(date.today())),
    ('Price', '$61.72 (Aug 18, 2026 close)'),
    ('Shares Outstanding', '128.7M'),
    ('Market Cap', '$8.01B'),
    ('Enterprise Value', '$10.20B (Yahoo Key Stats)'),
    ('Total Debt', '$3.33B (Key Stats, MRQ 6/30/2026)'),
    ('Total Cash', '$1.02B (MRQ)'),
    ('Net Debt (EV-MC)', '$2.19B'),
    ('Primary Lens', 'Forward P/E — fallen angel post-earnings-whiplash'),
    ('Stance', 'Watch — earnings crater, down 49% from high, Q2 miss was -134%'),
    ('Beta (5Y)', '0.92'),
    ('Sector', 'Industrials / Engineering & Construction'),
]

for i, (field, val) in enumerate(title_block, 2):
    c(ws1, i, 1, field, font=hdr if i == 2 else None)
    c(ws1, i, 2, val)

# Valuation multiples table
c(ws1, 18, 1, 'Valuation Multiples', font=hdr)
headers = ['Metric', 'Value', 'Comment']
data = [
    ['Trailing P/E', '21.91x', 'Elevated because TTm EPS depressed by Q2 miss (-$0.50 actual vs +$1.46 est)'],
    ['Forward P/E', '9.61x (FY26 consensus)', 'Cheap on FY26 EPS $3.98 — but FY26 includes the cratered Q2. FY27 fwd P/E ~9.6x on $6.42'],
    ['P/S', '0.53x', 'Deep discount for engineering services — near all-time revenue levels'],
    ['EV/Revenue', '0.66x', 'Enterprise-levelcheapness; reflects debt burden vs revenue stability'],
    ['EV/EBITDA', '10.87x', 'Reasonable for engineering; peers (FLR ~8-12x, ACWX ~7-9x)'],
    ['P/Book', '3.65x', 'Below prior peak (6.90x); compressed from 52-week high'],
    ['PEG (5Y Exp)', '0.65', 'Implies <1x on expected growth; aggressive growth assumptions baked in'],
    ['Dividend Yield', '1.91%', '$1.19/share annualized; payout ratio 41.9% — sustainable'],
    ['52-Week Range', '$60.35 - $135.52', 'Stock trading at the absolute bottom of 52-week range. Down -48.6% YTD'],
]
write_table(ws1, 19, headers, data, [25, 35, 65])

for ci in range(1, 4):
    ws1.column_dimensions[get_column_letter(ci)].width = [25, 35, 65][ci-1]

# =====================================================================
# Sheet 2: WACC
# =====================================================================
ws2 = wb.create_sheet('WACC')
ws2.merge_cells('A1:D1')
c(ws2, 1, 1, 'WACC Calculation — AECOM (ACM)', font=title_font)

wacc_data = [
    ('Risk-Free Rate (10Y US Treasury)', '4.70%', 'CNBC US10Y, Aug 18 2026'),
    ('Equity Risk Premium', '5.00%', 'Standard assumption'),
    ('', '', ''),
    ('Beta (5Y Monthly)', '0.92', 'Yahoo Key Statistics'),
    ('Cost of Equity (CAPM)', '9.26%', '4.70% + 0.92 * 5.00%'),
    ('', '', ''),
    ('Interest Expense (FY25)', '$184.3M', 'Yahoo Income Statement'),
    ('Cost of Debt (pre-tax)', '5.71%', '184.3M / 3,229M total debt ~5.71%'),
    ('Tax Rate', '18.68%', '97,755 / 521,009 FY25 tax provision'),
    ('Cost of Debt (after-tax)', '4.65%', '5.71% * (1 - 0.1868)'),
    ('', '', ''),
    ('Market Cap', '$8,010M', 'Yahoo Key Stats, Current'),
    ('Total Debt', '$3,330M', 'Yahoo Key Stats, MRQ'),
    ('Total Capitalization', '$11,340M', 'MC + Debt'),
    ('Equity Weight', '70.63%', '8,010 / 11,340'),
    ('Debt Weight', '29.37%', '3,330 / 11,340'),
    ('', '', ''),
    ('WACC', '8.04%', '=0.7063*9.26% + 0.2937*4.65%'),
]

for i, (field, val, note) in enumerate(wacc_data, 3):
    c(ws2, i, 1, field, font=hdr if field == 'WACC' else None)
    c(ws2, i, 2, val)
    c(ws2, i, 3, note)

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 45

# Print WACC
rf = 0.0470
erp = 0.05
beta = 0.92
ke = rf + beta * erp
tax_rate = 0.1868
kd_pct = 0.0571
kd_after_tax = kd_pct * (1 - tax_rate)
eq_wt = 8010 / 11340
debt_wt = 3330 / 11340
wacc = eq_wt * ke + debt_wt * kd_after_tax
print(f"WACC: rf={rf}, erp={erp}, beta={beta}, ke={ke:.4f}, kd_at={kd_after_tax:.4f}")
print(f"  Equity wt: {eq_wt:.4f}, Debt wt: {debt_wt:.4f}, WACC={wacc:.4f}")

# =====================================================================
# Sheet 3: Scenarios
# =====================================================================
ws3 = wb.create_sheet('Scenarios')
ws3.merge_cells('A1:I1')
c(ws3, 1, 1, 'Forward P/E Scenario Analysis — AECOM (ACM)', font=title_font)

c(ws3, 3, 1, 'Framework: Forward P/E on analyst-consensus EPS. Primary lens for engineering/construction services. FCF multiple as cross-check.', font=subtitle_font)

# Scenario inputs
# FY25 EPS was $4.21. FY26 consensus $3.98. FY27 consensus $6.42.
# Revenue: $15.39B TTM, $15.3B FY26 est, $16.87B FY27 est (+10.3% YoY)
# Net income FY25: $561.8M, TTM: $288M (cratered by Q2 miss)
# Revenue CAGR for next 5 years: modest 2-4% for engineering/construction
# Adjusted FCF margin: ~4.5% (FCF ~$685M on $15.39B revenue = 4.45%)

shares_mm = 128.7
current_price = 61.72
mc_mm = 8010
ev_mm = 10200
net_debt_mm = ev_mm - mc_mm  # 2,190

print(f"\nShares (M): {shares_mm}")
print(f"Current price: ${current_price}")
print(f"MC: ${mc_mm}M, EV: ${ev_mm}M, Net Debt: ${net_debt_mm}M")

# Scenario parameters — EPS-based forward P/E framework
scenarios_data = [
    # Revenue CAGR, Terminal Revenue ($M, Yr5), Adj FCF Margin, Terminal FCF ($M), Exit FCF Multiple, Implied EV ($M), Less Net Debt ($M), Shares (M), Target Price, Upside%, Weight, Weighted Value/Share
    # BEAR: Revenue flat/slight decline, FCF margin compresses, multiple compresses to 8x
    # FY27 consensus revenue $16.87B, but bear says revenue stalls
    # FY27 EPS consensus $6.42 — bear uses compressed EPS of $5.00
    # Exit P/E: 10x (below current 9.61x fwd) — earnings whiplash premium
    ('Revenue CAGR (next 5Y)', '0%', '3%', '5%'),
    ('Terminal Revenue (Yr 5)', '$15.3B', '$17.7B', '$19.5B'),
    ('Adjusted FCF Margin', '3.5%', '4.5%', '5.5%'),
    ('Terminal FCF ($M)', '536M', '797M', '1,073M'),
    ('Exit FCF Multiple (cross-check)', '8x', '10x', '12x'),
    ('Implied EV ($M)', '4,288', '7,970', '12,876'),
    ('Less Net Debt ($M)', '(2,190)', '(2,190)', '(2,190)'),
    ('Equity Value ($M)', '2,098', '5,780', '10,686')]

# Forward P/E based scenario — primary framework
# BEAR: EPS $5.00/share, P/E 10x → $50.00 target
# BASE: EPS $6.42/share (FY27 consensus), P/E 12x → $77.04 target
# BULL: EPS $8.50/share, P/E 15x → $127.50 target

# Let's compute more carefully
# FY27 EPS consensus: $6.42. Forward P/E at current price: 61.72/6.42 = 9.61x
# Bear: multiple compresses to 9x, EPS at $5.00 → $45.00 target
# Base: multiple expands to 12x on FY27 EPS $6.42 → $77.04 target
# Bull: multiple at 15x, EPS at $8.50 → $127.50 target

bear_eps = 5.00
base_eps = 6.42
bull_eps = 8.50

bear_pe = 9
base_pe = 12
bull_pe = 15

bear_target = bear_eps * bear_pe
base_target = base_eps * base_pe
bull_target = bull_eps * bull_pe

bear_upside = (bear_target / current_price - 1)
base_upside = (base_target / current_price - 1)
bull_upside = (bull_target / current_price - 1)

print(f"\nScenario targets:")
print(f"  Bear: ${bear_eps} EPS * {bear_pe}x = ${bear_target:.2f} ({bear_upside*100:.1f}%)")
print(f"  Base: ${base_eps} EPS * {base_pe}x = ${base_target:.2f} ({base_upside*100:.1f}%)")
print(f"  Bull: ${bull_eps} EPS * {bull_pe}x = ${bull_target:.2f} ({bull_upside*100:.1f}%)")

# FCF cross-check targets
bear_fcf = 536  # 15.3B * 3.5%
base_fcf = 797  # 17.7B * 4.5%
bull_fcf = 1073  # 19.5B * 5.5%

bear_fcf_ev = bear_fcf * 8
base_fcf_ev = base_fcf * 10
bull_fcf_ev = bull_fcf * 12

bear_fcf_eq = bear_fcf_ev - net_debt_mm
base_fcf_eq = base_fcf_ev - net_debt_mm
bull_fcf_eq = bull_fcf_ev - net_debt_mm

bear_fcf_price = bear_fcf_eq / shares_mm
base_fcf_price = base_fcf_eq / shares_mm
bull_fcf_price = bull_fcf_eq / shares_mm

print(f"  FCF cross-check Bear: EV=${bear_fcf_ev}M - ${net_debt_mm}M debt = ${bear_fcf_eq}M / {shares_mm}M = ${bear_fcf_price:.2f}")
print(f"  FCF cross-check Base: EV=${base_fcf_ev}M - ${net_debt_mm}M debt = ${base_fcf_eq}M / {shares_mm}M = ${base_fcf_price:.2f}")
print(f"  FCF cross-check Bull: EV=${bull_fcf_ev}M - ${net_debt_mm}M debt = ${bull_fcf_eq}M / {shares_mm}M = ${bull_fcf_price:.2f}")

# Scenario table
headers_s = ['Metric', 'Bear', 'Base', 'Bull', 'Notes']
scenarios = [
    ('EPS (Year 1 anchor)', f'${bear_eps:.2f}', f'${base_eps:.2f}', f'${bull_eps:.2f}', 'Bear below FY27 consensus of $6.42'),
    ('Exit P/E', f'{bear_pe}x', f'{base_pe}x', f'{bull_pe}x', 'Eng/construction peer norms: 10-16x'),
    ('', '', '', '', ''),
    ('Implied Target Price', f'${bear_target:.2f}', f'${base_target:.2f}', f'${bull_target:.2f}', 'Primary: EPS * Exit P/E'),
    ('Upside from Current', f'{bear_upside*100:.1f}%', f'{base_upside*100:.1f}%', f'{bull_upside*100:.1f}%', f'From ${current_price}'),
    ('', '', '', '', ''),
    ('FCF Cross-Check Target', f'${bear_fcf_price:.2f}', f'${base_fcf_price:.2f}', f'${bull_fcf_price:.2f}', 'Terminal FCF * exit mult - net debt / shares'),
    ('', '', '', '', ''),
    ('Probability Weight', '20%', '50%', '30%', ''),
    ('Weighted Value/Share', f'${bear_target*0.20:.2f}', f'${base_target*0.50:.2f}', f'${bull_target*0.30:.2f}', ''),
    ('', '', '', '', ''),
    ('Probability-Weighted FV', '', '', f'${bear_target*0.20 + base_target*0.50 + bull_target*0.30:.2f}', ''),
    ('Current Price', '', '', f'${current_price}', ''),
    ('Probability-Weighted Upside', '', '', f'{(bear_target*0.20 + base_target*0.50 + bull_target*0.30)/current_price - 1:.1%}', ''),
]

fv = bear_target*0.20 + base_target*0.50 + bull_target*0.30
print(f"\nFV: ${fv:.2f}, Upside: {(fv/current_price - 1)*100:.1f}%")

write_table(ws3, 5, headers_s, scenarios, [30, 15, 15, 15, 40])

for ci in range(1, 6):
    ws3.column_dimensions[get_column_letter(ci)].width = [30, 15, 15, 15, 50][ci-1]

# Actuals Source Audit
ws4 = wb.create_sheet('Actuals Source Audit')
ws4.merge_cells('A1:E1')
c(ws4, 1, 1, 'Actuals Source Audit — AECOM (ACM)', font=title_font)

audit_headers = ['Data Point', 'Value', 'Source URL', 'Date', 'Notes']
audit_data = [
    ['Stock Price', '$61.72', 'Yahoo Finance /quote/ACM/', 'Aug 18, 2026', 'Market close'],
    ['After Hours', '$62.20', 'Yahoo Finance', 'Aug 18, 2026', '+0.48 (+0.78%)'],
    ['Market Cap', '$8.01B', 'Yahoo Key Statistics', 'Current (Aug 2026)', 'Quarterly tab'],
    ['Enterprise Value', '$10.20B', 'Yahoo Key Statistics', 'Current', 'Same quarter'],
    ['Total Debt', '$3.33B', 'Yahoo Key Statistics', 'MRQ 6/30/2026', 'Key Stats value — uses higher/more conservative'],
    ['Total Debt (BS)', '$3.23B', 'Yahoo Balance Sheet', 'FY25 (9/30/2025)', 'Within $100M of Key Stats — no material discrepancy'],
    ['Net Debt (BS)', '$1.13B', 'Yahoo Balance Sheet', 'FY25', 'Total debt $3.23B - cash $2.10B. Net debt from EV-MC = $2.19B.'],
    ['Total Cash', '$1.02B', 'Yahoo Key Statistics', 'MRQ', 'Cash per share = $7.95'],
    ['Shares Outstanding', '128.7M', 'Yahoo Key Statistics', 'Current', 'Implied shares match basic shares on IS'],
    ['Beta (5Y)', '0.92', 'Yahoo Key Statistics', 'Current', 'Monthly beta'],
    ['Revenue TTM', '$15.39B', 'Yahoo Income Statement', 'TTM', 'Annual: $16.14B (FY25), $16.11B (FY24)'],
    ['Gross Profit TTM', '$874M', 'Yahoo Income Statement', 'TTM', 'Margin: 5.68%'],
    ['Operating Income TTM', '$715M', 'Yahoo Income Statement', 'TTM', 'Margin: 4.64%. FY25: $1.06B (6.56%)'],
    ['Net Income TTM', '$288M', 'Yahoo Income Statement', 'TTM', 'Heavily impacted by Q2 FY26 miss. FY25: $562M'],
    ['Diluted EPS TTM', '$2.19', 'Yahoo Income Statement', 'TTM', 'FY25: $4.21 (peak)'],
    ['EBITDA TTM', '$723M', 'Yahoo Income Statement', 'TTM', 'S&P Global calculated on Key Stats: $932M'],
    ['Operating Cash Flow', '$822M', 'Yahoo Cash Flow', 'FY25 (9/30/2025)', 'FCF: $685M after $137M capex'],
    ['Total Assets', '$12.20B', 'Yahoo Balance Sheet', 'FY25', 'Annual data'],
    ['Total Liabilities', '$9.50B', 'Yahoo Balance Sheet', 'FY25', ''],
    ['Common Stock Equity', '$2.49B', 'Yahoo Balance Sheet', 'FY25', 'Tangible book value: -$1.39B (negative from goodwill)'],
    ['Analyst Est FY26 EPS', '$3.98', 'Yahoo Analysis/Estimates', 'Aug 2026', '10 analysts. Low $3.95, High $4.05. GAAP vs non-GAAP: GAAP shown.'],
    ['Analyst Est FY27 EPS', '$6.42', 'Yahoo Analysis/Estimates', 'Aug 2026', '12 analysts. Low $5.91, High $7.01'],
    ['Analyst Est FY26 Rev', '$15.3B', 'Yahoo Analysis', 'Aug 2026', '3 analysts. Low $15.18B, High $15.54B'],
    ['Analyst Est FY27 Rev', '$16.87B', 'Yahoo Analysis', 'Aug 2026', '2 analysts. Low $16.3B, High $17.84B'],
    ['Q2 FY26 EPS Actual', '-$0.50', 'Yahoo Analysis/EPS History', 'Quarter ended 3/31/2026', 'Estimated $1.46. Miss of -134.17%. CRITICAL EARNINGS WHIPLASH.'],
    ['10Y Treasury', '4.704%', 'CNBC US10Y', 'Aug 18, 2026', 'Used for WACC risk-free rate'],
    ['FY26 Revenue (Q3)', '$3.8B', 'Yahoo Analysis/Revenue vs Earnings', 'Aug 2026', 'Q2 FY26 actual. Estimate: $4.08B for Q3 (Sep 2026)'],
    ['Next Earnings Date', 'Q3 FY26', 'Yahoo Profile - Recent Events', 'Aug 10, 2026 earnings date', 'Q3 results already released (Aug 10); transcript published Aug 17'],
]

write_table(ws4, 3, audit_headers, audit_data, [25, 18, 35, 20, 50])
for ci in range(1, 6):
    ws4.column_dimensions[get_column_letter(ci)].width = [28, 18, 38, 22, 55][ci-1]

# =====================================================================
# Sheet 5: Questions
# =====================================================================
ws5 = wb.create_sheet('Questions')
ws5.merge_cells('A1:D1')
c(ws5, 1, 1, 'Open Questions — AECOM (ACM)', font=title_font)

questions = [
    ('1', 'Q2 FY26 EPS miss of -134%: Management estimated $1.46 but reported -$0.50. The miss comes from a combination of one-time project losses, cost overruns, or provisioning events. What specifically drove the -$0.50 EPS? Was it a single project write-down or systemic margin erosion?', 'Critical — defines earnings quality'),
    ('2', 'EPS revision cascade: FY26 consensus dropped from $5.97 to $3.98 in 30+ days — a $2.00/share (33%) downward revision. Was the prior estimate based on pre-demotion guidance? What was management saying in Q1/Q2 calls that drove prior high expectations?', 'Defines earnings trajectory'),
    ('3', 'Revenue headwinds: QoQ revenue collapsed (Q1 FY26: $4.08B est → Q2 actual: $3.8B actual). Is the $15.39B TTM revenue reflecting a structural revenue decline or cyclical timing of construction project completions?', 'Revenue growth vs cyclicality'),
    ('4', 'Goodwill composition: Tangible book value is -$1.39B on $2.49B common equity — meaning goodwill and intangibles of ~$4.4B (177% of equity). Any impairment risk if earnings stay depressed?', 'Balance sheet quality'),
    ('5', 'Debt trajectory: Total debt increased from $3.03B (FY25) to $3.33B (MRQ). Issuance of debt was $3.4B in FY25 vs repayment of $3.27B. Why the debt increase during a period when OCF was $822M?', 'Capital allocation concern'),
    ('6', 'Buyback activity: Repurchased $388M of stock in FY25 at prices well above current $61.72 (peak was $135+). Were buybacks deployed at inflated multiples, destroying long-term value?', 'Governance / capital allocation'),
    ('7', 'Segment profitability: AECOM reports Americas, International, and AECOM Capital. Which segment drove the revenue decline? International exposure creates FX risk — is currency headwind masking or amplifying the operating story?', 'Segment-level analysis needed'),
    ('8', 'Backlog visibility: Engineering/construction is backlog-driven. What is the current backlog-to-revenue ratio? A declining backlog ratio would signal future revenue contraction before it appears in the P&L.', 'Leading indicator'),
    ('9', 'AECOM Capital division: The capital/investment arm ($1.8B+ of total assets) exposes the company to real estate market risk. In a declining commercial real estate market, this could impair valuation and tie up balance sheet capacity.', 'Real estate risk exposure'),
    ('10', 'Q3 FY26 earnings: Results were released Aug 10, 2026, with the earnings call transcript available Aug 17. What did the Q3 results show? Was the Q2 miss a one-time event or the continuation of a trend?', 'Catalyst follow-up'),
]

q_headers = ['#', 'Question', 'Why It Matters']
write_table(ws5, 3, q_headers, questions, [5, 80, 35])

# =====================================================================
# Sheet 6: Sources
# =====================================================================
ws6 = wb.create_sheet('Sources')
ws6.merge_cells('A1:C1')
c(ws6, 1, 1, 'Sources — AECOM (ACM)', font=title_font)

sources = [
    ('1', 'Yahoo Finance — ACM Stock Summary', 'https://finance.yahoo.com/quote/ACM/'),
    ('2', 'Yahoo Finance — ACM Key Statistics / Valuation Measures', 'https://finance.yahoo.com/quote/ACM/key-statistics/'),
    ('3', 'Yahoo Finance — ACM Income Statement', 'https://finance.yahoo.com/quote/ACM/financials/'),
    ('4', 'Yahoo Finance — ACM Balance Sheet', 'https://finance.yahoo.com/quote/ACM/balance-sheet/'),
    ('5', 'Yahoo Finance — ACM Cash Flow Statement', 'https://finance.yahoo.com/quote/ACM/cash-flow/'),
    ('6', 'Yahoo Finance — ACM Analyst Estimates / Earnings Trends', 'https://finance.yahoo.com/quote/ACM/analysis/'),
    ('7', 'Yahoo Finance — ACM Company Profile', 'https://finance.yahoo.com/quote/ACM/profile/'),
    ('8', 'Yahoo Finance — ACM News', 'https://finance.yahoo.com/quote/ACM/news/'),
    ('9', 'CNBC — US 10-Year Treasury Yield', 'https://www.cnbc.com/quotes/US10Y'),
    ('10', 'StockAnalysis.com — ACM (404 — unavailable)', 'https://stockanalysis.com/quote/ACM/'),
    ('11', 'Motley Fool — Q3 FY26 Earnings Call Transcript', 'Via Yahoo Finance news tab'),
    ('12', 'StockStory — Q2 Earnings Review / Analyst Questions', 'Via Yahoo Finance news tab'),
]

s_headers = ['#', 'Description', 'URL']
write_table(ws6, 3, s_headers, sources, [5, 55, 50])

# =====================================================================
# Save
# =====================================================================
outfile = '/home/refcell/dev/capital/models/[2026-08-18] AECOM Model.xlsx'
wb.save(outfile)
print(f"\nModel saved to: {outfile}")
print("DONE")
