#!/usr/bin/env python3
"""Build 6-sheet valuation model for Liberty Global Ltd. (LBTYA).
Data sourced from Yahoo Finance as of July 29, 2026.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
import datetime

# === KEY DATA (Yahoo Finance, Jul 29, 2026) ===
PRICE = 10.53
MKT_CAP_MM = 3480        # $3.48B
EV_MM = 10130            # $10.13B
SHARES_MM = 176.72       # Common shares outstanding (Yahoo Stat)
IMPLIED_SHARES_MM = 341.09  # Implied with convertible sub equity
BETA = 0.74
P_S = 0.71
P_B = 0.38
EV_REV = 2.07
EV_EBITDA = 15.11
REVENUE_TTM_MM = 4884.8    # $4.88B
GROSS_PROFIT_TTM_MM = 3281.0
OP_INCOME_TTM_MM = 55.7
OCF_TTM_MM = 1271.2        # $1.27B
FCF_TTM_MM = -254.8        # Negative due to heavy fiber capex
TOTAL_DEBT_MM = 9334.6     # $9.33B
TOTAL_CASH_MM = 2821.1     # $2.82B (MRQ from BS)
NET_DEBT_MM = 6513.5       # EV - MC proxy
BVPS = 28.11               # $28.11 per share (MRQ)
RISK_FREE = 4.681          # 10Y US Treasury (CNBC)

# === WACC ===
ERP = 5.0
COST_EQUITY = RISK_FREE + BETA * ERP  # ~8.96%
COST_DEBT = 4.5   # Estimated pre-tax cost of debt (BBB-rated telecom)
TAX_RATE = 20.0   # Approximate effective rate (European operations)
TOTAL_CAP_MM = MKT_CAP_MM + TOTAL_DEBT_MM
EQ_WT = MKT_CAP_MM / TOTAL_CAP_MM
DEBT_WT = TOTAL_DEBT_MM / TOTAL_CAP_MM
WACC = EQ_WT * COST_EQUITY + DEBT_WT * COST_DEBT * (1 - TAX_RATE / 100)

# === SCENARIOS (P/B framework primary, EV/EBITDA cross-check) ===
# Liberty Global is a capital-intensive telecom — P/B and EV/EBITDA are the
# correct valuation lenses. FCF multiples are inapplicable during heavy
# fiber-buildout cycles.

# Current BVPS ~$28.11, current share count 176.72M, price $10.53
# The key variable: can they reduce debt, maintain OCF, and grow BVPS?

# Scenario 1: Bear — multiple stays compressed, debt overhang persists
# P/B 0.35x at year 5, shares decline 10% from buybacks, BVPS flat
bear_pb = 0.35
bear_bvps_5y = 28.11  # Flat BVPS (debt overhang)
bear_shares = 176.72 * 0.90  # 10% buyback
bear_target = bear_pb * bear_bvps_5y

# Scenario 2: Base — P/B re-rates modestly, BVPS grows via OCF/debt reduction
base_pb = 0.50
base_bvps_5y = 28.11 * (1.06)**5  # 6% BVPS CAGR from OCF + buybacks
base_shares = 176.72 * 0.80  # 20% buyback over 5 years
base_target = base_pb * base_bvps_5y

# Scenario 3: Bull — P/B normalizes, Ziggo IPO unlocks value, debt reduced
bull_pb = 0.65
bull_bvps_5y = 28.11 * (1.10)**5  # 10% BVPS CAGR
bull_shares = 176.72 * 0.70  # 30% buyback
bull_target = bull_pb * bull_bvps_5y

# Probabilities
bear_wt = 0.20
base_wt = 0.50
bull_wt = 0.30
WTD_FV = bear_target * bear_wt + base_target * base_wt + bull_target * bull_wt

print(f"WACC: {WACC:.2f}%")
print(f"Bear target: ${bear_target:.2f} (0.35x P/B)")
print(f"Base target: ${base_target:.2f} (0.50x P/B)")
print(f"Bull target: ${bull_target:.2f} (0.65x P/B)")
print(f"Weighted FV: ${WTD_FV:.2f}  ({(WTD_FV-PRICE)/PRICE*100:+.1f}% from ${PRICE})")

# Sanity check: targets should be plausible ($5-20 range)
assert 5 < bear_target < 30, f"Bear target {bear_target} seems wrong"
assert 5 < base_target < 40, f"Base target {base_target} seems wrong"
assert 5 < bull_target < 60, f"Bull target {bull_target} seems wrong"
print("All targets in plausible range. Building workbook...")

# === BUILD WORKBOOK ===
wb = Workbook()

# Style objects
title_font = Font(bold=True, size=14, name="Calibri")
bold_font = Font(bold=True, size=11, name="Calibri")
normal_font = Font(size=11, name="Calibri")
header_fill = PatternFill("solid", fgColor="D9E2F3")
header_border = Border(bottom=Side(style="thin"))
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

def write_title(ws, text):
    """Write title in merged row 1"""
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = text
    c.font = title_font

def write_header(ws, row, col_start, headers):
    """Write formatted header row"""
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start+i, value=h)
        c.font = bold_font
        c.fill = header_fill
        c.border = header_border

def write_row(ws, row, col_start, values):
    """Write data row"""
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=col_start+i, value=v)
        c.font = normal_font
        c.border = thin_border

def set_col_widths(ws, widths):
    """Set column widths"""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ========== SHEET 1: Valuation ==========
ws = wb.active
ws.title = "Valuation"
write_title(ws, "Liberty Global Ltd. (LBTYA) - Valuation Summary")
set_col_widths(ws, [25, 20, 20, 40, 20, 15])

# Title block
row = 3
write_row(ws, row, 1, ["Ticker:", "LBTYA", "Exchange: NASDAQ (NasdaqGS)"])
row += 1
write_row(ws, row, 1, ["Company:", "Liberty Global Ltd.", "Bermuda-domiciled telecom holding co."])
row += 1
write_row(ws, row, 1, ["Date:", "2026-07-29", "As of close Jul 29, 2026"])
row += 1
write_row(ws, row, 1, ["Price:", f"${PRICE:.2f}", "NASDAQ close"])
row += 1
write_row(ws, row, 1, ["Shares Outstanding:", f"{SHARES_MM:.1f}M", f"Common (Yahoo Stat); Implied: {IMPLIED_SHARES_MM:.1f}M w/convertibles"])
row += 1
write_row(ws, row, 1, ["Market Cap:", f"${MKT_CAP_MM/1000:.2f}B", ""])
row += 1
write_row(ws, row, 1, ["Enterprise Value:", f"${EV_MM/1000:.2f}B", "MC + $9.33B debt - $2.82B cash"])
row += 1
write_row(ws, row, 1, ["Primary Lens:", "P/B and EV/EBITDA", "Telecom infrastructure — FCF multiples inapplicable during fiber buildout"])
row += 1
write_row(ws, row, 1, ["Stance:", "Watch", "P/B 0.38x below book but earnings quality is transitional"])

# Valuation metrics table
row += 2
ws.cell(row=row, column=1, value="KEY VALUATION METRICS").font = bold_font
row += 1
write_header(ws, row, 1, ["Metric", "Value", "Comment"])
row += 1
val_data = [
    ["P/E (Trailing)", "N/A", "Negative TTM earnings (-$3.04B net income). GAAP EPS -$9.30 distorted by 'Other Income/Expense' items."],
    ["Forward P/E", "N/A", "Negative GAAP forward estimates; analysts use non-GAAP/adjusted metrics."],
    ["P/S (TTM)", f"{P_S:.2f}x", "$3.48B / $4.88B TTM revenue. Below 1x on consumer staples-level multiple."],
    ["P/B (MRQ)", f"{P_B:.2f}x", "$10.53 / $28.11 BVPS. Trading at 38% of book value."],
    ["EV/EBITDA", f"{EV_EBITDA:.1f}x", "$10.13B / ~$670M S&P EBITDA. Note: Yahoo GAAP EBITDA is -$1.25B; S&P calculated EBITDA ~$1.14B."],
    ["EV/Sales", f"{EV_REV:.2f}x", "$10.13B / $4.88B. Reasonable for European cable operators."],
    ["Beta (5Y)", f"{BETA:.2f}", "Low volatility — defensive telecom characteristics."],
    ["BVPS", f"${BVPS:.2f}", "Tangible book value per common share."],
    ["Revenue TTM", f"${REVENUE_TTM_MM/1000:.2f}B", "$4.88B across UK, Netherlands, Belgium, Ireland, Slovakia."],
    ["OCF TTM", f"${OCF_TTM_MM/1000:.2f}B", "$1.27B operating cash flow — strong cash generation for debt reduction."],
    ["Net Debt", f"${NET_DEBT_MM/1000:.2f}B", "Ev - MC proxy. High leverage but manageable with $1.27B OCF."],
    ["Debt/Equity", "98.6%", "Total debt / total equity (MRQ). Elevated but telecom-normative."],
]
for vd in val_data:
    write_row(ws, row, 1, vd)
    row += 1

# ========== SHEET 2: WACC ==========
ws = wb.create_sheet("WACC")
write_title(ws, "WACC Calculation - Liberty Global Ltd. (LBTYA)")
set_col_widths(ws, [30, 18, 25, 40])

row = 3
write_header(ws, row, 1, ["Parameter", "Value", "Source", "Note"])
row += 1
wacc_data = [
    ["Risk-Free Rate (10Y US)", f"{RISK_FREE:.3f}%", "CNBC (cnbc.com/quotes/US10Y)", "As of Jul 29, 2026, 10:34PM EDT"],
    ["Equity Risk Premium", f"{ERP:.1f}%", "Assumed", "Standard 5% ERP"],
    ["Beta (Levered, 5Y Monthly)", f"{BETA:.2f}", "Yahoo Finance Key Statistics", "Low — defensive telecom characteristics"],
    ["Cost of Equity (CAPM)", f"{COST_EQUITY:.2f}%", f"{RISK_FREE:.1f}% + {BETA:.2f} x {ERP:.1f}%"],
    ["Cost of Debt (Pre-Tax)", f"{COST_DEBT:.1f}%", "Estimated", f"BBB-rated European telecom; weighted avg on debt book"],
    ["Tax Rate", f"{TAX_RATE:.1f}%", "Effective (approx.)", "Multi-jurisdictional (Bermuda HQ + European ops)"],
    ["Market Cap (million $)", f"{MKT_CAP_MM:.0f}", "Yahoo Finance Statistics"],
    ["Total Debt (million $)", f"{TOTAL_DEBT_MM:.0f}", "Yahoo Finance Balance Sheet"],
    ["Total Capitalization", f"{TOTAL_CAP_MM:.0f}", "MC + Debt"],
    ["Equity Weight", f"{EQ_WT:.2%}", f"{MKT_CAP_MM:.0f}/{TOTAL_CAP_MM:.0f}"],
    ["Debt Weight", f"{DEBT_WT:.2%}", f"{TOTAL_DEBT_MM:.0f}/{TOTAL_CAP_MM:.0f}"],
    ["WACC", f"{WACC:.2f}%", f"{EQ_WT:.2%} x {COST_EQUITY:.2f}% + {DEBT_WT:.2%} x {COST_DEBT:.1f}% x (1-{TAX_RATE/100:.3f})"],
]
for wd in wacc_data:
    write_row(ws, row, 1, wd)
    row += 1

# ========== SHEET 3: Scenarios ==========
ws = wb.create_sheet("Scenarios")
write_title(ws, "Liberty Global (LBTYA) - Scenario Analysis (P/B Framework, 5-Year Horizon)")
set_col_widths(ws, [30, 18, 18, 18, 45])

row = 3
write_header(ws, row, 1, ["Metric", "Bear", "Base", "Bull", "Note"])
row += 1

# Revenue growth assumptions
bear_rev_cagr = 0.5       # Low growth, flat in local currencies
base_rev_cagr = 3.0       # Modest organic growth
bull_rev_cagr = 5.0       # Acceleration from mobile convergence

# Terminal revenue (5Y)
bear_term_rev = REVENUE_TTM_MM * (1 + bear_rev_cagr/100)**5
base_term_rev = REVENUE_TTM_MM * (1 + base_rev_cagr/100)**5
bull_term_rev = REVENUE_TTM_MM * (1 + bull_rev_cagr/100)**5

# EV/EBITDA exit multiples (cross-check)
bear_ev_ebitda = 12.0
base_ev_ebitda = 15.0
bull_ev_ebitda = 18.0

scenarios = [
    ["Revenue CAGR (5Y)", f"{bear_rev_cagr:.1f}%", f"{base_rev_cagr:.1f}%", f"{bull_rev_cagr:.1f}%", "Nominal, FX-neutral"],
    ["Terminal Revenue (5Y)", f"${bear_term_rev:.0f}M", f"${base_term_rev:.0f}M", f"${bull_term_rev:.0f}M", "Revenue projection"],
    ["Exit P/B Multiple", f"{bear_pb:.2f}x", f"{base_pb:.2f}x", f"{bull_pb:.2f}x", "Primary valuation lens"],
    ["Exit EV/EBITDA (cross-check)", f"{bear_ev_ebitda:.0f}x", f"{base_ev_ebitda:.0f}x", f"{bull_ev_ebitda:.0f}x", "Secondary validation"],
    ["BVPS CAGR (5Y)", "0%", "6%", "10%", "From OCF deployment + buybacks"],
    ["Terminal BVPS", f"${bear_bvps_5y:.2f}", f"${base_bvps_5y:.2f}", f"${bull_bvps_5y:.2f}", "BVPS at year 5"],
    ["Terminal Shares (M)", f"{bear_shares:.1f}", f"{base_shares:.1f}", f"{bull_shares:.1f}", "Post-buyback"],
    ["Implied Target Price", f"${bear_target:.2f}", f"${base_target:.2f}", f"${bull_target:.2f}", "P/B x Terminal BVPS"],
    ["Upside from Current", f"{(bear_target-PRICE)/PRICE*100:.1f}%", f"{(base_target-PRICE)/PRICE*100:.1f}%", f"{(bull_target-PRICE)/PRICE*100:.1f}%", f"vs ${PRICE}"],
    ["Scenario Weight", f"{bear_wt:.0%}", f"{base_wt:.0%}", f"{bull_wt:.0%}", "Probability assignment"],
    ["Weighted Value/Share", f"${bear_target*bear_wt:.2f}", f"${base_target*base_wt:.2f}", f"${bull_target*bull_wt:.2f}", "Target x Weight"],
]
for s in scenarios:
    write_row(ws, row, 1, s)
    row += 1

row += 1
ws.cell(row=row, column=1, value="Probability-Weighted Fair Value").font = bold_font
c = ws.cell(row=row, column=2, value=f"${WTD_FV:.2f}")
c.font = bold_font
ws.cell(row=row, column=3, value=f"{(WTD_FV-PRICE)/PRICE*100:+.1f}% from ${PRICE:.2f}").font = bold_font
row += 2

# Framework notes
ws.cell(row=row, column=1, value="FRAMEWORK NOTES").font = bold_font
row += 1
notes = [
    "1. P/B is primary lens because Liberty Global is a capital-intensive telecom with heavy debt, transitional GAAP earnings, and fiber buildout cycle distorting FCF.",
    "2. GAAP net income is meaningless for valuation: -$3.04B TTM from $2.41B TTM of 'Other Income/Expense' charges (amortization of intangibles from Virgin Media O2 acquisition, non-cash items).",
    "3. S&P Global calculated EBITDA of ~$1.14B TTM is the better earnings proxy vs. Yahoo GAAP EBITDA of -$1.25B.",
    f"4. Operating cash flow of ${OCF_TTM_MM/1000:.2f}B TTM is the real earnings power metric — sufficient for debt reduction and modest buybacks.",
    "5. Share count already declining: 459M (FY22) -> 177M (Jul 2026), a 61% reduction from M&A reverse recapitalizations.",
    "6. Net debt of $6.51B is the key variable — OCF must reduce this over time for P/B to re-rate.",
    f"7. Ziggo IPO (planned Amsterdam listing 2027) could unlock significant value not yet reflected in current ${PRICE} price.",
    f"8. Current P/B of {P_B:.2f}x is deep discount for telecom sector (peer median ~0.6-0.9x). Even the bear case at 0.35x P/B implies modest downside only.",
]
for n in notes:
    ws.cell(row=row, column=1, value=n).font = Font(size=10, name="Calibri")
    row += 1

# ========== SHEET 4: Actuals Source Audit ==========
ws = wb.create_sheet("Actuals Source Audit")
write_title(ws, "Liberty Global (LBTYA) - Data Source Audit")
set_col_widths(ws, [30, 20, 40, 18, 35])

row = 3
write_header(ws, row, 1, ["Data Point", "Value", "Source", "Date", "Notes"])
row += 1

audit = [
    ["Stock Price", f"${PRICE:.2f}", "Yahoo Finance Summary", "Jul 29, 2026 close", "NASDAQ real-time"],
    ["Market Cap", f"${MKT_CAP_MM/1000:.2f}B", "Yahoo Finance Statistics", "Jul 29, 2026", "Intraday MC"],
    ["Enterprise Value", f"${EV_MM/1000:.2f}B", "Yahoo Finance Statistics", "Jul 29, 2026", "MC + Total Debt - Cash"],
    ["Shares Outstanding", f"{SHARES_MM:.1f}M", "Yahoo Finance Key Statistics", "Jul 2026", "Common shares; Implied: {IMPLIED_SHARES_MM:.1f}M w/subsidiary converts"],
    ["52W Range", "$9.43 - $13.52", "Yahoo Finance", "Jul 2026", "Currently near midpoint"],
    ["Beta (5Y Monthly)", f"{BETA:.2f}", "Yahoo Finance Key Statistics", "Jul 29, 2026", "Low volatility"],
    ["Revenue TTM", f"${REVENUE_TTM_MM/1000:.2f}B", "Yahoo Finance Income Statement", "TTM Jun 2026", "4,884,800 (thousands)"],
    ["Gross Profit TTM", f"${GROSS_PROFIT_TTM_MM/1000:.2f}B", "Yahoo Finance IS", "TTM", "67.2% gross margin"],
    ["Operating Income TTM", f"${OP_INCOME_TTM_MM/1000:.1f}M", "Yahoo Finance IS", "TTM", "1.1% operating margin"],
    ["Net Income TTM", "-$3.04B", "Yahoo Finance IS", "TTM", "Distorted by -$2.41B Other Income/Expense"],
    ["Diluted EPS TTM", "-$9.30", "Yahoo Finance IS", "TTM", "Not usable for valuation"],
    ["EBITDA (Yahoo GAAP)", "-$1.25B", "Yahoo Finance IS", "TTM", "Massively distorted; don't use"],
    ["EBITDA (S&P Calc)", "~$1.14B", "Yahoo Finance Statistics page", "TTM", "S&P Global methodology; better proxy"],
    ["Operating CF TTM", f"${OCF_TTM_MM/1000:.2f}B", "Yahoo Finance Cash Flow", "TTM", "$1,271,200 (thousands)"],
    ["Capex TTM", "-$1.53B", "Yahoo Finance Cash Flow", "TTM", "Heavy fiber buildout — capital intensive"],
    ["Free Cash Flow TTM", "-$254.8M", "Yahoo Finance Cash Flow", "TTM", "Negative due to capex cycle; not normalized"],
    ["Total Debt", f"${TOTAL_DEBT_MM/1000:.2f}B", "Yahoo Finance Balance Sheet", "FY2025", "9,334,600 (thousands)"],
    ["Total Cash (MRQ)", f"${TOTAL_CASH_MM/1000:.2f}B", "Yahoo Finance Balance Sheet", "MRQ Jun 2026", "2,821,100 (thousands)"],
    ["Net Debt", f"${NET_DEBT_MM/1000:.2f}B", "Calculated: Debt - Cash", "Jul 2026", "Debt-heavy but OCF supports service"],
    ["Total Assets", "$22.6B", "Yahoo Finance Balance Sheet", "FY2025", "22,595,900 (thousands)"],
    ["Total Equity", "$9.95B", "Yahoo Finance Balance Sheet", "FY2025", "9,945,800 (thousands)"],
    ["BVPS", f"${BVPS:.2f}", "Yahoo Finance Key Statistics", "MRQ", "Book value per common share"],
    ["Tangible Book Value", "$4.88B", "Yahoo Finance Balance Sheet", "FY2025", "Goodwill-adjusted equity"],
    ["P/S Ratio", f"{P_S:.2f}x", "Yahoo Finance Statistics", "Jul 29, 2026", "Current"],
    ["P/B Ratio", f"{P_B:.2f}x", "Yahoo Finance Statistics", "Jul 29, 2026", "Current"],
    ["EV/Revenue", f"{EV_REV:.2f}x", "Yahoo Finance Statistics", "Jul 29, 2026", "Current"],
    ["EV/EBITDA", f"{EV_EBITDA:.1f}x", "Yahoo Finance Statistics", "Jul 29, 2026", "S&P-calculated EBITDA"],
    ["10Y Treasury Yield", f"{RISK_FREE:.3f}%", "CNBC (cnbc.com/quotes/US10Y)", "Jul 29, 2026", "Risk-free rate for WACC"],
    ["Fiscal Year End", "December 31", "Yahoo Finance Profile", "Annual", "Calendar year"],
    ["Q2 FY26 Revenue", "$1.17B", "Yahoo Finance Analysis", "Jul 2026", "In line with estimates"],
    ["Q2 FY26 EPS (Non-GAAP)", "-$1.07 actual, -$0.86 est", "Yahoo Finance Analysis", "Jul 2026", "Missed by $0.21"],
    ["Analyst Avg PT", "$15.05", "Yahoo Finance Analysis", "Jul 2026", "High: $25, Low: $9.90"],
    ["Debt/Equity Ratio", "98.6%", "Yahoo Finance Statistics", "MRQ", "Elevated leverage"],
    ["Quarterly Rev Growth YoY", "-7.7%", "Yahoo Finance Statistics", "TTM", "Decline in Q2 base"],
    ["Buybacks TTM", "-$90.1M", "Yahoo Finance Cash Flow", "TTM", "Modest vs previous years"],
    ["Short % of Float", "6.14%", "Yahoo Finance Key Statistics", "Jul 15, 2026", "Moderate short interest"],
    ["Employees", "6,636", "Yahoo Finance Profile", "Current", "European telecom workforce"],
    ["Capital Leases", "$740M", "Yahoo Finance Balance Sheet", "FY2025", "739,700 (thousands)"],
]
for a in audit:
    write_row(ws, row, 1, a)
    row += 1

# ========== SHEET 5: Questions ==========
ws = wb.create_sheet("Questions")
write_title(ws, "Liberty Global (LBTYA) - Open Questions")
set_col_widths(ws, [5, 70, 25])

row = 3
write_header(ws, row, 1, ["#", "Question", "Category"])
row += 1

questions = [
    ["1", "What comprises the -$2.41B TTM 'Other Income/Expense'? This line item is the single largest earnings distortion — likely amortization of Virgin Media O2 acquisition intangibles, impairment charges, or accounting adjustments. What is the breakdown?", "Earnings Quality"],
    ["2", "What is the detailed debt maturity schedule? $9.33B total debt with net debt of $6.51B — what comes due in 2027-2030 and at what rates? Rate reset risk on European debt?", "Capital Structure"],
    ["3", "How does the debt reduction plan align with OCF of $1.27B TTM? At current pace, how many years to meaningfully reduce net debt?", "Capital Structure"],
    ["4", "Is the fiber capex cycle ($1.53B TTM, crushing FCF) cyclical or structural? If cyclical, when does the buildout end and FCF recover?", "Cash Flow"],
    ["5", "Q2 FY26 EPS missed by $0.21 (-$1.07 vs -0.86 est). Was this a one-off or structural margin pressure?", "Earnings Quality"],
    ["6", "What is the exact timeline and structure of the planned Ziggo IPO (Amsterdam listing, targeted 2027)? How much value could unlock?", "Strategy"],
    ["7", "The share count fell 61% from FY22 (459M) to current (177M). Was this from reverse recapitalization or aggressive buybacks?", "Capital Structure"],
    ["8", "Does the implied shares outstanding of 341M (vs 177M common) represent LBRDA/LILAK convertible subsidiary equity? What is the conversion trigger?", "Capital Structure"],
    ["9", "How does the Virgin Media O2 mobile convergence (UK full mobile operation since 2021) impact competitive positioning vs. BT/EE, Vodafone?", "Competitive"],
    ["10", "Quarterly revenue growth of -7.7% YoY in Q2 FY26 — is this organic decline or timing/sequencing from the prior year base?", "Growth"],
    ["11", "What is the effective tax rate treatment for Bermuda-domiciled company with European operations? TTM tax provision was $191.5M on -$2.81B pretax — not informative for forward projections.", "Taxation"],
    ["12", "Multiple 8-K filings (Apr-Jul 2026) for 'Corporate Changes & Voting Matters' — what corporate actions are these? Governance changes? Board composition?", "Governance"],
    ["13", "How does the ONE Connect and Horizon 5 technology platform create competitive durability against UK Openreach fiber rollout and Dutch market consolidation?", "Competitive"],
    ["14", "What is the customer concentration risk across UK/Netherlands/Belgium/Ireland/Slovakia? Regulatory exposure in each jurisdiction?", "Risk"],
    ["15", "Are there preferred stock obligations or minority interest claims that should be subtracted from reported equity for true common shareholder value?", "Capital Structure"],
]
for q in questions:
    write_row(ws, row, 1, q)
    row += 1

# ========== SHEET 6: Sources ==========
ws = wb.create_sheet("Sources")
write_title(ws, "Liberty Global (LBTYA) - Data Sources")
set_col_widths(ws, [5, 80])

row = 3
write_header(ws, row, 1, ["#", "Source"])
row += 1

sources = [
    ["1", "Yahoo Finance Summary — finance.yahoo.com/quote/LBTYA/ (price, market cap, profile, recent news)"],
    ["2", "Yahoo Finance Key Statistics — finance.yahoo.com/quote/LBTYA/key-statistics/ (P/S, P/B, EV, beta, shares)"],
    ["3", "Yahoo Finance Income Statement — finance.yahoo.com/quote/LBTYA/financials/ (revenue, gross profit, operating income, net income, EPS, EBITDA)"],
    ["4", "Yahoo Finance Balance Sheet — finance.yahoo.com/quote/LBTYA/balance-sheet/ (total assets, liabilities, equity, debt, cash, BVPS)"],
    ["5", "Yahoo Finance Cash Flow — finance.yahoo.com/quote/LBTYA/cash-flow/ (OCF, capex, FCF, buybacks, end cash)"],
    ["6", "Yahoo Finance Analysis — finance.yahoo.com/quote/LBTYA/analysis/ (analyst estimates, price targets, earnings trends)"],
    ["7", "Yahoo Finance Profile — finance.yahoo.com/quote/LBTYA/profile/ (executives, sector, employees, fiscal year, events)"],
    ["8", "CNBC 10Y Treasury — cnbc.com/quotes/US10Y (risk-free rate for WACC: 4.681% on Jul 29, 2026)"],
    ["9", "Morningstar Research Reports (via Yahoo Finance) — multiple coverage notes from 2024-2026"],
    ["10", "GuruFocus / MarketBeat — Q2 2026 earnings call highlights (via Yahoo Finance news tab)"],
    ["11", "GlobeNewswire — Liberty Global Q2 2026 results press release, investor call scheduling, Ziggo IPO announcements, executive appointments"],
    ["12", "Insider Monkey — fiber consolidation strategy and UK broadband coverage articles"],
]
for s in sources:
    write_row(ws, row, 1, s)
    row += 1

# === SAVE ===
outpath = "/home/refcell/dev/capital/models/[2026-07-29] Liberty Global Model.xlsx"
wb.save(outpath)
print(f"Workbook saved to {outpath}")
print("BUILD COMPLETE")
