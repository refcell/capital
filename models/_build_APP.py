"""Build 6-sheet valuation model for AppLovin Corporation (APP).

Sources:
  - Yahoo Finance Income Statement / Balance Sheet / Cash Flow / Statistics / Analysis
  - StockAnalysis.com returned 404; Yahoo Finance is the primary source.
  - CNBC US10Y: 4.676% (2026-08-06)
  - Quote date: 2026-08-06, close $335.67

All $ figures in millions unless noted otherwise.

Key characteristics:
  - Fallen angel: down 55% from 52W high of $745.61; crashed -19.66% on Aug 6 alone
  - AI-powered mobile advertising platform (Axon AI, MAX bidding, Adjust, Wurl)
  - Revenue $6.16B TTM, +87.6% vs FY2024; $4.4B FCF TTM = 71.4% FCF margin
  - Heavy buybacks: $2.17B TTM; share count declining (338M -> 306M ~10% in recent periods)
  - Interest expense $205M vs operating income $4.75B = only 4.3% coverage ratio
  - Analyst consensus: $16.02 EPS FY26, $21.17 EPS FY27
  - Primary lens: Forward P/E scenarios (Fallen Angel pattern per skill notes)
  - Beta 2.53, currently at absolute bottom of 52W range
"""

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──
HERE = Path(__file__).resolve().parent
XLSX = HERE / "[2026-08-06] AppLovin Model.xlsx"

# ── Raw inputs ──
PRICE = 335.67
SHARES_M = 305.73           # shares outstanding (Yahoo Stats, Aug 6 2026)
MC_B = 133.00               # $B market cap
EV_B = 133.75               # $B enterprise value
TOTAL_DEBT_M = 3544.7       # from Yahoo Balance Sheet Q1 FY26
CASH_M = 2754.6             # from Yahoo Key Stats Q1 FY26 total cash
NET_DEBT_M = EV_B * 1000 - MC_B * 1000  # $750M from Yahoo stats EV-MC = $0.75B
# Actually: EV $133.75B - MC $133.00B = $0.75B = $750M
NET_DEBT_M = 750.0

# Income statement ($M)
REV_TTM = 6164.2
REV_F25 = 5480.7
REV_F24 = 3224.1
REV_F23 = 1841.8
REV_F22 = 2817.1

GP_TTM = 5447.1
GP_F25 = 4815.6

OP_IN_TTM = 4751.9
OP_IN_F25 = 4151.9
OP_IN_F24 = 1910.96

NI_TTM = 3962.6
NI_F25 = 3333.3
NI_F24 = 1577.1

EPS_D_TTM = 11.64
EPS_D_F25 = 9.75
EPS_D_F24 = 4.53

EBITDA_TTM = 4943.6
EBITDA_F25 = 4354.7

FCF_TTM = 4402.5
FCF_F25 = 3942.8
OCF_TTM = 4430.8

# Balance sheet
TOTAL_DEBT_BS_M = 3544.7  # FY2025 total debt from BS
CASH_BS_M = 2487.1        # FY2025 end cash
# Note: Net debt from BS: $3,544.7M debt - $2,754.6M cash (Q1 FY26) = ~$790M
# EV-MC proxy: $750M — close enough

# Analyst estimates (Yahoo Finance /analysis/, Aug 6 2026)
REV_F26_C = 8180.0        # consensus FY2026 revenue $8.18B, 29 analysts
REV_F27_C = 10620.0       # consensus FY2027 revenue $10.62B, 31 analysts
EPS_F26_C = 16.02         # consensus FY2026 EPS, 23 analysts
EPS_F27_C = 21.17         # consensus FY2027 EPS, 25 analysts
EPS_Q2_F26_C = 3.75       # Q2 FY26 estimate
EPS_Q3_F26_C = 4.07

# Recent earnings surprise history:
# Q1 FY26: EPS actual $3.56 vs est $3.44 (+3.5% surprise, beat)
# FY25 Q4: EPS actual $3.24 vs est $2.94 (+10.0% surprise)
# FY25 Q3: EPS actual $2.45 vs est $2.39 (+2.6% surprise)
# FY25 Q2: EPS actual $2.39 vs est $1.98 (+20.4% surprise)
# Pattern: 4 consecutive beats, accelerating surprise

# Tax rate, beta, rates
TAX_RATE = 0.147           # TTM effective ~14.7%
BETA = 2.53                # 5Y monthly beta (Yahoo Stats)
RISK_FREE = 0.04676        # 10Y US Treasury, CNBC Aug 6 2026
ERP = 0.05                 # equity risk premium

# ── Computed ──
COE = RISK_FREE + BETA * ERP  # CAPM cost of equity ~0.1733 = 17.33%
COST_OF_DEBT = RISK_FREE + 0.025  # ~250 bps spread over risk-free
EQUITY_WEIGHT = MC_B / (MC_B + max(NET_DEBT_M / 1000, 0.01))
DEBT_WEIGHT = max(NET_DEBT_M / 1000, 0.01) / (MC_B + max(NET_DEBT_M / 1000, 0.01))
WACC = EQUITY_WEIGHT * COE + DEBT_WEIGHT * COST_OF_DEBT * (1 - TAX_RATE)

# Valuation multiples
PE_TRAIL_TTM = PRICE / EPS_D_TTM          # 34.43x
PE_FWD_F26 = PRICE / EPS_F26_C            # 20.95x
PE_FWD_F27 = PRICE / EPS_F27_C            # 15.86x
PS = MC_B / (REV_TTM / 1000)              # ~21.58x
PB = MC_B / (2134.7 / 1000)               # ~62.3x (book equity $2.13B FY25)
EV_REVENUE = EV_B / (REV_TTM / 1000)      # ~21.70x
EV_EBITDA = EV_B / (EBITDA_TTM / 1000)    # ~27.06x
PE_FCF = MC_B / (FCF_TTM / 1000)          # ~30.21x
EV_FCF = EV_B / (FCF_TTM / 1000)          # ~30.38x

# ── Scenario assumptions (Fallen Angel framework: Forward P/E primary) ──
# 5-year revenue CAGR to bridge TTM $6.16B through FY27 consensus
# FY27 rev $10.62B = 2-year CAGR of ~28.3% from current TTM
# 5-year CAGR then decelerates

# Bear: Multiple compresses further, growth decelerates sharply, revenue hits FY27 consensus but multiple drops to 10-12x
# Base: Hitting consensus, multiple normalizes to 18-22x (still premium for high growth)
# Bull: Beats consensus, multiple holds near current 21x, growth sustains

scenarios = {
    'bear': {
        'rev_cagr_5y': 0.12,     # decelerates to 12% - growth slows substantially
        'terminal_rev_5y': REV_F27_C * 1.12**3,  # ~$14.9B (decelerated from $10.6B)
        'op_margin': 0.60,       # compresses from 78% to 60% - competitive pressure
        'terminal_FCF_margin': 0.55,
        'exit_PE': 14.0,         # compressed to value-growth peer levels
    },
    'base': {
        'rev_cagr_5y': 0.18,     # 18% sustainable long-run
        'terminal_rev_5y': REV_F27_C * 1.18**3,  # ~$16.2B
        'op_margin': 0.70,       # stays strong at 70%
        'terminal_FCF_margin': 0.65,
        'exit_PE': 20.0,         # growth premium maintained
    },
    'bull': {
        'rev_cagr_5y': 0.25,     # 25% - AI advertising tailwind sustains
        'terminal_rev_5y': REV_F27_C * 1.25**3,  # ~$20.8B
        'op_margin': 0.76,       # near current levels
        'terminal_FCF_margin': 0.72,
        'exit_PE': 28.0,         # premium growth multiple
    }
}

# Weights
weights = {'bear': 0.20, 'base': 0.50, 'bull': 0.30}

# ── Style helpers ──
title_font = Font(name='Calibri', bold=True, size=16)
subtitle_font = Font(name='Calibri', bold=True, size=11, color='666666')
header_font = Font(name='Calibri', bold=True, size=10)
data_font = Font(name='Calibri', size=10)
bold_font = Font(name='Calibri', bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

def ws_title(ws, text):
    """Write merged title at row 1."""
    ws.merge_cells('A1:F1')
    ws['A1'] = text
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    for col in range(1, 7):
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

def ws_subtitle(ws, row, text):
    """Write subtitle at given row."""
    c = ws.cell(row=row, column=1, value=text)
    c.font = subtitle_font

def write_table(ws, headers, data, start_row=3):
    """Write header row + data rows with borders and formatting."""
    # Header
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    # Data
    for ri, row_data in enumerate(data, start_row + 1):
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = data_font
            c.border = thin_border
            if ci == 1:
                c.font = bold_font
    # Column widths
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18


# ── Build ──
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════
# Sheet 1: Valuation
# ══════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Valuation'
ws_title(ws1, 'AppLovin Corporation (NasdaqGS: APP)')
ws_subtitle(ws1, 2, f'Valuation Model — Data as of {PRICE} on 2026-08-06 (close)')

val_headers = ['Field', 'Value', 'Comment']
val_data = [
    ['Company', 'AppLovin Corporation', 'AI-powered mobile advertising'],
    ['Ticker', 'NasdaqGS: APP', ''],
    ['Date', '2026-08-06', ''],
    ['Close Price', f'${PRICE}', 'Down -$82.13 (-19.66%) — massive selloff day'],
    ['Shares Outstanding', f'{SHARES_M}M', 'Yahoo Stats Aug 6; declining from ~352M FY23'],
    ['Market Cap', f'${MC_B:.2f}B', ''],
    ['Enterprise Value', f'${EV_B:.2f}B', 'Net debt ~$750M via EV-MC proxy'],
    ['Primary Valuation Lens', 'Forward P/E', 'Fallen Angel pattern — FCF multiple distorted by recent run-up multiple'],
    ['Stance', 'Watch', 'Near absolute 52W low; 20.9x fwd P/E — multiple compressing but earnings acceleration may not sustain'],
    ['', '', ''],
    ['PE Trailing (TTM)', f'{PE_TRAIL_TTM:.2f}x', f'On dil EPS {EPS_D_TTM:.2f}'],
    ['Fwd PE FY26', f'{PE_FWD_F26:.2f}x', f'On consensus EPS ${EPS_F26_C:.2f} (23 analysts)'],
    ['Fwd PE FY27', f'{PE_FWD_F27:.2f}x', f'On consensus EPS ${EPS_F27_C:.2f} (25 analysts)'],
    ['P/S', f'{PS:.2f}x', f'On TTM revenue ${REV_TTM/1000:.2f}B'],
    ['P/FCF', f'{PE_FCF:.1f}x', f'On TTM FCF ${FCF_TTM/1000:.1f}B'],
    ['EV/FCF', f'{EV_FCF:.1f}x', f''],
    ['EV/Sales', f'{EV_REVENUE:.1f}x', f''],
    ['EV/EBITDA', f'{EV_EBITDA:.1f}x', f''],
    ['P/B', f'{PB:.0f}x', f'Book equity $2.13B FY25 — extremely high for asset-light'],
    ['Beta (5Y Monthly)', f'{BETA}', 'Yahoo Stats'],
    ['52W High', '$745.61', f'Stock down {((745.61-PRICE)/745.61)*100:.0f}% from peak'],
    ['52W Low', '$332.19', f'Stock at {((PRICE-332.19)/332.19)*100:.0f}% above absolute low'],
]
write_table(ws1, val_headers, val_data, 3)

# ══════════════════════════════════════════════
# Sheet 2: WACC
# ══════════════════════════════════════════════
ws2 = wb.create_sheet('WACC')
ws_title(ws2, 'WACC Calculation — CAPM')
ws_subtitle(ws2, 2, 'Components as of 2026-08-06')

wacc_data = [
    ['Risk-Free Rate (10Y US)', f'{RISK_FREE*100:.3f}%', 'CNBC US10Y, 4.676%'],
    ['Equity Risk Premium', '5.00%', 'Assumed'],
    ['Beta (Levered, 5Y Monthly)', f'{BETA}', 'Yahoo Stats'],
    ['Cost of Equity (CAPM)', f'{COE*100:.2f}%', f'{RISK_FREE:.4f} + {BETA} x 0.05'],
    ['Cost of Debt (pre-tax)', f'{COST_OF_DEBT*100:.2f}%', f'RF + 250bps spread'],
    ['Tax Rate', f'{TAX_RATE*100:.1f}%', f'TTM effective ~{TAX_RATE*100:.1f}%'],
    ['Market Cap', f'${MC_B:.2f}B', ''],
    ['Total Debt', f'${TOTAL_DEBT_BS_M/1000:.2f}B', 'Yahoo BS FY2025'],
    ['Net Debt (EV-MC proxy)', f'${NET_DEBT_M}M', ''],
    ['Equity Weight', f'{EQUITY_WEIGHT*100:.2f}%', ''],
    ['Debt Weight', f'{DEBT_WEIGHT*100:.2f}%', ''],
    ['WACC', f'{WACC*100:.2f}%', ''],
]
write_table(ws2, ['Component', 'Value', 'Note'], wacc_data, 3)

# Print WACC for verification
print(f"WACC: {WACC*100:.2f}%")

# ══════════════════════════════════════════════
# Sheet 3: Scenarios (Forward P/E framework)
# ══════════════════════════════════════════════
ws3 = wb.create_sheet('Scenarios')
ws_title(ws3, 'Scenario Analysis — Forward P/E Framework')
ws_subtitle(ws3, 2, 'Fallen Angel pattern: primary lens = Forward P/E on consensus')

# Compute scenario targets
# Using forward P/E framework: target = terminal EPS × exit P/E
# Terminal EPS derived from: terminal revenue × terminal op margin × (1-tax) / shares
# But revenue 5y from FY27 = FY27 rev × (1+cagr)^3 = terminal revenue

bear_term_rev = scenarios['bear']['terminal_rev_5y']  # ~$14.9B
base_term_rev = scenarios['base']['terminal_rev_5y']  # ~$16.2B
bull_term_rev = scenarios['bull']['terminal_rev_5y']  # ~$20.8B

def calc_target(term_rev_M, op_margin, tax_rate, exit_pe, shares_M):
    term_op_in = term_rev_M * op_margin
    term_ni = term_op_in * (1 - tax_rate)
    term_eps = term_ni / shares_M
    target = term_eps * exit_pe
    return term_eps, target

bear_eps, bear_target = calc_target(bear_term_rev, scenarios['bear']['op_margin'], TAX_RATE, scenarios['bear']['exit_PE'], SHARES_M)
base_eps, base_target = calc_target(base_term_rev, scenarios['base']['op_margin'], TAX_RATE, scenarios['base']['exit_PE'], SHARES_M)
bull_eps, bull_target = calc_target(bull_term_rev, scenarios['bull']['op_margin'], TAX_RATE, scenarios['bull']['exit_PE'], SHARES_M)

# FCF-margin cross-check
# FCF margin 5Y: terminal_rev * FCF_margin
bear_term_fcf = bear_term_rev * scenarios['bear']['terminal_FCF_margin']
base_term_fcf = base_term_rev * scenarios['base']['terminal_FCF_margin']
bull_term_fcf = bull_term_rev * scenarios['bull']['terminal_FCF_margin']

# Weights and FV
bear_w, base_w, bull_w = weights['bear'], weights['base'], weights['bull']
weighted_fv = bear_w * bear_target + base_w * base_target + bull_w * bull_target

# Upside calculations
bear_upside = (bear_target - PRICE) / PRICE * 100
base_upside = (base_target - PRICE) / PRICE * 100
bull_upside = (bull_target - PRICE) / PRICE * 100
weighted_upside = (weighted_fv - PRICE) / PRICE * 100

scen_headers = ['Scenario Driver', 'Bear', 'Base', 'Bull', 'Note']
scen_data = [
    ['Revenue CAGR (5Y)', f'{scenarios["bear"]["rev_cagr_5y"]*100:.0f}%', f'{scenarios["base"]["rev_cagr_5y"]*100:.0f}%', f'{scenarios["bull"]["rev_cagr_5y"]*100:.0f}%', 'From FY27 baseline'],
    ['Terminal Revenue (5Y, $MM)', f'{bear_term_rev:.0f}', f'{base_term_rev:.0f}', f'{bull_term_rev:.0f}', ''],
    ['Op Margin', f'{scenarios["bear"]["op_margin"]*100:.0f}%', f'{scenarios["base"]["op_margin"]*100:.0f}%', f'{scenarios["bull"]["op_margin"]*100:.0f}%', ''],
    ['Terminal EPS (5Y)', f'${bear_eps:.2f}', f'${base_eps:.2f}', f'${bull_eps:.2f}', ''],
    ['Exit P/E', f'{scenarios["bear"]["exit_PE"]:.0f}x', f'{scenarios["base"]["exit_PE"]:.0f}x', f'{scenarios["bull"]["exit_PE"]:.0f}x', ''],
    ['Target Price', f'${bear_target:.2f}', f'${base_target:.2f}', f'${bull_target:.2f}', ''],
    ['Upside %', f'{bear_upside:.1f}%', f'{base_upside:.1f}%', f'{bull_upside:.1f}%', f'From ${PRICE}'],
    ['Weight', f'{bear_w*100:.0f}%', f'{base_w*100:.0f}%', f'{bull_w*100:.0f}%', ''],
    ['Weighted Value/Share', f'${bear_w*bear_target:.2f}', f'${base_w*base_target:.2f}', f'${bull_w*bull_target:.2f}', ''],
    ['', '', '', '', ''],
    ['Probability-Weighted FV', '', '', f'${weighted_fv:.2f}', 'Sum of weighted'],
    ['Current Price', '', '', f'${PRICE}', ''],
    ['Weighted Upside', '', '', f'{weighted_upside:.1f}%', ''],
]

# Add FCF cross-check section
scen_data.append(['', '', '', '', ''])
scen_data.append(['FCF Cross-Check (Margin)', '', '', '', ''])
scen_data.append(['Terminal FCF Margin', f'{scenarios["bear"]["terminal_FCF_margin"]*100:.0f}%', f'{scenarios["base"]["terminal_FCF_margin"]*100:.0f}%', f'{scenarios["bull"]["terminal_FCF_margin"]*100:.0f}%', ''])
scen_data.append(['Terminal FCF ($MM)', f'{bear_term_fcf:.0f}', f'{base_term_fcf:.0f}', f'{bull_term_fcf:.0f}', ''])

# Add note frame
scen_data.append(['', '', '', '', ''])
scen_data.append(['Framework Note', '', '', '', ''])
scen_data.append(['Primary Lens', '', '', '', 'Forward P/E (Fallen Angel pattern)'])
scen_data.append(['FCF Framework', '', '', '', 'N/A — 71.4% FCF margin from asset-light adtech; multiples not meaningful vs peers'])
scen_data.append(['EV/EBITDA', '', '', '', f'$133.75B / 71.4% margin'])
scen_data.append(['EV/EBITDA', '', '', f'{EV_EBITDA:.1f}x', f'$ {EV_B:.2f}B / EBITDA $ {EBITDA_TTM/1000:.1f}B'])

write_table(ws3, scen_headers, scen_data, 3)

# ══════════════════════════════════════════════════
# Sheet 4: Actuals Source Audit
# ══════════════════════════════════════════════
ws4 = wb.create_sheet('Actuals Source Audit')
ws_title(ws4, 'Actuals Source Audit')
ws_subtitle(ws4, 2, 'Every data point with source URL, date, and verification notes')

audit_headers = ['Data Point', 'Value', 'Source URL', 'Date', 'Notes']
audit_data = [
    ['Stock Price', f'${PRICE}', 'Yahoo Finance /quote/APP/', '2026-08-06', 'Close price'],
    ['Price Change', '-$82.13 (-19.66%)', 'Yahoo Finance /quote/APP/', '2026-08-06', 'Massive selloff'],
    ['Overnight/After hours', '~$337.51', 'Yahoo Finance /quote/APP/', '2026-08-06', 'Blue Ocean ATS activity'],
    ['Market Cap', f'${MC_B:.2f}B', 'Yahoo Finance Statistics', '2026-08-06', ''],
    ['Enterprise Value', f'${EV_B:.2f}B', 'Yahoo Finance Statistics', '2026-08-06', ''],
    ['Shares Outstanding', f'{SHARES_M}M', 'Yahoo Finance Statistics', '2026-08-06', 'Implied shares per Key Stats'],
    ['Implied Shares Outstanding', '335.94M', 'Yahoo Finance Statistics', '2026-08-06', 'With subsidiary equity conversions'],
    ['52W High', '$745.61', 'Yahoo Finance Statistics', '2026-08-06', ''],
    ['52W Low', '$332.19', 'Yahoo Finance Statistics', '2026-08-06', ''],
    ['Beta (5Y Mo)', f'{BETA}', 'Yahoo Finance Statistics', '2026-08-06', ''],
    ['TTM Revenue', f'${REV_TTM/1000:.1f}B', 'Yahoo Finance /financials/', 'As of FY2025 Q4/TTM', 'All $000s converted to $M'],
    ['FY25 Revenue', f'${REV_F25/1000:.1f}B', 'Yahoo Finance /financials/', 'FY2025', ''],
    ['FY24 Revenue', f'${REV_F24/1000:.1f}B', 'Yahoo Finance /financials/', 'FY2024', ''],
    ['TTM Gross Profit', f'${GP_TTM/1000:.1f}B', 'Yahoo Finance /financials/', 'TTM', ''],
    ['TTM Operating Income', f'${OP_IN_TTM/1000:.1f}B', 'Yahoo Finance /financials/', 'TTM', ''],
    ['TTM Net Income', f'${NI_TTM/1000:.1f}B', 'Yahoo Finance /financials/', 'TTM', ''],
    ['TTM EPS (Dil)', f'${EPS_D_TTM}', 'Yahoo Finance /financials/', 'TTM', ''],
    ['TTM EBITDA', f'${EBITDA_TTM/1000:.1f}B', 'Yahoo Finance /financials/', 'TTM', ''],
    ['TTM FCF', f'${FCF_TTM/1000:.1f}B', 'Yahoo Finance /cash-flow/', 'TTM', 'Levered free cash flow = OCF + Investing CF'],
    ['TTM OCF', f'${OCF_TTM/1000:.1f}B', 'Yahoo Finance /cash-flow/', 'TTM', ''],
    ['Total Debt', f'${TOTAL_DEBT_BS_M}M', 'Yahoo Finance /balance-sheet/', 'FY25/Key Stats', 'Total debt per BS'],
    ['Total Cash', f'${CASH_M}M', 'Yahoo Finance Key Stats Statistics', 'Q1 FY26', ''],
    ['Net Debt', f'${NET_DEBT_M}M', 'EV-MC Proxy', '2026-08-06', 'EV - MC'],
    ['Next Earnings', 'Nov 4, 2026 3:00 PM EST', 'Yahoo Finance /profile/', '2026-08-06', 'Next earnings date'],
    ['EPS F26 Consensus', f'${EPS_F26_C}', 'Yahoo Finance /analysis/', '2026-08-06', '23 analysts, GAAP'],
    ['EPS F27 Consensus', f'${EPS_F27_C}', 'Yahoo Finance /analysis/', '2026-08-06', '25 analysts, GAAP'],
    ['Rev F26 Consensus', f'${REV_F26_C/1000:.2f}B', 'Yahoo Finance /analysis/', '2026-08-06', '29 analysts'],
    ['Rev F27 Consensus', f'${REV_F27_C/1000:.2f}B', 'Yahoo Finance /analysis/', '2026-08-06', '31 analysts'],
    ['Q1 FY26 EPS Actual', '$3.56', 'Yahoo Finance /analysis/', 'Apr 2026', 'Beat est $3.44 by +3.4%'],
    ['Q2 FY26 EPS Est', f'${EPS_Q2_F26_C}', 'Yahoo Finance /analysis/', '2026-08-06', '17 analysts'],
    ['Q3 FY26 EPS Est', f'${EPS_Q3_F26_C}', 'Yahoo Finance /analysis/', '2026-08-06', ''],
    ['10Y US Treasury', f'{RISK_FREE*100:.3f}%', 'CNBC /quotes/US10Y', '2026-08-06', ''],
]
write_table(ws4, audit_headers, audit_data, 3)

# ════════════════════════════════════════════
# Sheet 5: Questions
# ════════════════════════════════════════════
ws5 = wb.create_sheet('Questions')
ws_title(ws5, 'Open Questions')
ws_subtitle(ws5, 2, 'Issues requiring follow-up or monitoring')

quest_headers = ['#', 'Question', 'Category', 'Priority']
quest_data = [
    [1, 'The -19.66% selloff on Aug 6, 2026: What triggered this crash? Was it sector-wide (DDOG also -19%), company-specific, or market-wide?', 'Catalyst', 'High'],
    [2, 'Revenue growth sustainability: TTM revenue $6.16B vs FY27 consensus $10.62B requires ~28% CAGR for 2 years. Can Axon AI and MAX maintain double-digit growth as TAM saturates?', 'Growth', 'High'],
    [3, 'Operating margin of 78.15% TTM is extraordinary — what is the floor? If competitive pressure (Trade Desk, Magnite, Meta) compresses even to 60-65%, how much does EPS decelerate?', 'Margins', 'High'],
    [4, 'Adjust acquisition integration: Adjust was acquired for ~$1.33B in May 2024. How much revenue/margin contribution comes from Adjust vs. organic Axon expansion?', 'M&A', 'Medium'],
    [5, 'Wurl CTV platform: Wurl is a connected TV platform — what is the growth trajectory and margin profile vs. the core mobile business?', 'Diversification', 'Medium'],
    [6, 'Buyback trajectory: $2.17B TTM, shares declining from ~352M to 306M. Is the buyback rate sustainable given $3.54B debt? Interest coverage at 4.3% is manageable but watch the trajectory.', 'Capital Allocation', 'High'],
    [7, 'Apple privacy/IDFA impact: AppLovin\'s Axon AI model is designed to work post-ATT. What is the actual performance attribution accuracy vs. pre-ATT? Has Apple\'s ATT framework already been fully priced in?', 'Regulatory', 'High'],
    [8, 'Customer concentration: Does the advertising business have significant exposure to top advertisers or app categories that could drop rapidly?', 'Concentration', 'Medium'],
    [9, 'Debt trajectory: Total debt ~$3.54B. What is the maturity profile? Any significant maturities in 2026-2028?', 'Debt', 'Medium'],
    [10, 'Tax rate anomaly: TTM effective tax rate of ~14.7% vs ~13% in FY25 vs 1.4% in FY24. Is the lower rate from R&D credits/foreign ops? Is upward normalization likely?', 'Tax', 'Medium'],
    [11, 'Q2 FY26 earnings beat ($3.56 vs $3.44 est) — first evidence of the pattern reversal. Can the Q3 estimate of $4.07 be hit? Any guidance on Q3 provided?', 'Earnings', 'High'],
    [12, 'SBC policy: What is the stock-based compensation trajectory? Does it dilute the buyback benefit?', 'Dilution', 'Low'],
    [13, 'Implied shares outstanding (335.94M) vs actual (305.73M): 30M share gap — what subsidiary equity is convertible and at what terms?', 'Capital Structure', 'Low'],
    [14, 'Short interest at 3.74% of shares outstanding. Has the short interest changed materially post-Aug 6 crash?', 'Sentiment', 'Low'],
    [15, 'Management commentary on AI-driven ad spending: Has management specifically addressed whether the AI boom is structural or cyclical for mobile advertising spend?', 'Guidance', 'Medium'],
]
write_table(ws5, quest_headers, quest_data, 3)
ws5.column_dimensions[get_column_letter(2)].width = 50

# ══════════════════════════════════════════════
# Sheet 6: Sources
# ════════════════════════════════════════════
ws6 = wb.create_sheet('Sources')
ws_title(ws6, 'Data Sources')
ws_subtitle(ws6, 2, 'All data accessed on 2026-08-06')

src_headers = ['#', 'Source', 'URL', 'Content']
src_data = [
    [1, 'Yahoo Finance — Quote/Summary', 'https://finance.yahoo.com/quote/APP/', 'Price, market cap, EV'],
    [2, 'Yahoo Finance — Income Statement', 'https://finance.yahoo.com/quote/APP/financials/', 'Revenue, GP, operating income, NI, EPS, EBITDA'],
    [3, 'Yahoo Finance — Balance Sheet', 'https://finance.yahoo.com/quote/APP/balance-sheet/', 'Assets, Liabilities, Debt, Cash, Equity'],
    [4, 'Yahoo Finance — Cash Flow', 'https://finance.yahoo.com/quote/APP/cash-flow/', 'OCF, Investing CF, Financing CF, FCF, Buybacks, Capex'],
    [5, 'Yahoo Finance — Key Statistics', 'https://finance.yahoo.com/quote/APP/key-statistics/', 'Shares, MC, EV, P/E, P/S, Beta, 52W range'],
    [6, 'Yahoo Finance — Analysis', 'https://finance.yahoo.com/quote/APP/analysis/', 'Analyst estimates, EPS consensus, revenue estimates, earnings history'],
    [7, 'Yahoo Finance — Profile', 'https://finance.yahoo.com/quote/APP/profile/', 'Company description, sector, industry, employees, next earnings'],
    [8, 'CNBC — US10Y', 'https://www.cnbc.com/quotes/US10Y', '10Y Treasury yield: 4.676%'],
    [9, 'StockAnalysis.com', 'https://stockanalysis.com/quote/APP/', 'Returned 404 — not available for this ticker'],
]
write_table(ws6, src_headers, src_data, 3)
ws6.column_dimensions[get_column_letter(3)].width = 45
ws6.column_dimensions[get_column_letter(4)].width = 35

# ── Save ──
wb.save(str(XLSX))
print(f"Saved: {XLSX}")

# ── Verification ──
print(f"\n=== VERIFICATION ===")
print(f"WACC: {WACC*100:.4f}%")
print(f"COE: {COE*100:.2f}%")
print(f"\nScenario Targets:")
print(f"  Bear:  ${bear_target:.2f} (EPS ${bear_eps:.2f}, {scenarios['bear']['exit_PE']}x PE) — {bear_upside:.1f}% upside")
print(f"  Base :  ${base_target:.2f} (EPS ${base_eps:.2f}, {scenarios['base']['exit_PE']}x PE) — {base_upside:.1f}% upside")
print(f"  Bull:  ${bull_target:.2f} (EPS ${bull_eps:.2f}, {scenarios['bull']['exit_PE']}x PE) — {bull_upside:.1f}% upside")
print(f"\nProbability-Weighted FV: ${weighted_fv:.2f} — {weighted_upside:.1f}% upside from ${PRICE}")
print(f"\nSanity check: FCF margin = {FCF_TTM/REV_TTM*100:.1f}%, OCF/Revenue = {OCF_TTM/REV_TTM*100:.1f}%")
print(f"Buyback/OP_IN_TTM: $2,173M / $4,752M = {2173/4752*100:.0f}% of OCF")
print(f"\nAnalyst consensus check: FY26 rev growth from TTM = {(8180/6164.2-1)*100:.1f}%, FY27 rev growth = {(10620/8180-1)*100:.1f}%")
