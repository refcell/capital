#!/usr/bin/env python3
"""
Build script for PSNL (Personalis, Inc.) valuation model
Date: 2026-08-25
Company: Personalis, Inc. — Cancer Genomics / Liquid Biopsy / Sequencing Platform
Sector: Healthcare / Diagnostics & Research
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Helpers ──────────────────────────────────────────────────────────
def bold_font(size=11, color="000000"):
    return Font(bold=True, size=size, color=color)

def header_font():
    return Font(bold=True, size=10, underline="single", color="FFFFFF")

def alt_fill():
    return PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

def blue_fill():
    return PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

def green_fill():
    return PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

def light_blue_fill():
    return PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

def thin_border():
    return Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

def c(ws, row, col, value, font_style=None, fill=None, alignment=None, border_style=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font_style:
        cell.font = font_style
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border_style:
        cell.border = border_style
    return cell

def write_table(ws, start_row, headers, data, alt_row_color=False):
    """Write header + data table with formatting."""
    for ci, h in enumerate(headers, 1):
        cell = c(ws, start_row, ci, h, header_font(), blue_fill(),
                 Alignment(horizontal="center", vertical="center", wrap_text=True), thin_border())
    for ri, row_data in enumerate(data, start_row + 1):
        for ci, val in enumerate(row_data, 1):
            f = alt_fill() if alt_row_color and (ri - start_row) % 2 == 0 else None
            c(ws, ri, ci, val, border_style=thin_border(), fill=f)
    return start_row + len(data) + 1

# ── Company data ─────────────────────────────────────────────────────
price = 17.25
market_cap = 1.82  # $B
ev = 1.65  # $B
shares_mm = market_cap / price * 1000  # = ~105.5M
revenue_ttm = 69.67  # $M
gross_profit_ttm = 8.838  # $M
operating_income_ttm = -114.309  # $M
net_income_ttm = -107.179  # $M
eps_ttm = -1.09  # diluted
total_cash_mm = 212.66  # $M from MRQ
total_debt_mm = 42.2  # ~implied from 16.16% x equity
operating_cf_ttm = -92.298  # $M
levered_fcf_ttm = -56.03  # $M
beta = 2.29
risk_free_rate = 4.641  # 10Y Treasury, Aug 25 2026
erp = 5.0
cost_of_equity = risk_free_rate + beta * erp  # 4.641 + 2.29*5 = 16.09%
tax_rate = 0.0  # No meaningful tax when unprofitable
total_debt_for_wacc_mm = 42.2
cost_of_debt = 3.5  # estimated
td_weight = total_debt_for_wacc_mm / (market_cap * 1000 + total_debt_for_wacc_mm)
eq_weight = 1 - td_weight
wacc = eq_weight * cost_of_equity + td_weight * cost_of_debt * (1 - tax_rate)

print(f"WACC = {wacc:.2f}%")
print(f"Cost of equity = {cost_of_equity:.2f}%")
print(f"Shares = {shares_mm:.1f}M")

# ── Sheet 1: Valuation ───────────────────────────────────────────────
ws1 = wb.active  # type: ignore
ws1.title = "Valuation"

# Title block
ws1.merge_cells("A1:F1")
c(ws1, 1, 1, f"PSNL — Personalis, Inc.", bold_font(16), alignment=Alignment(horizontal="center"))
ws1.merge_cells("A2:F2")
c(ws1, 2, 1, f"Valuation Summary | Data as of August 25, 2026", bold_font(11, "666666"),
    alignment=Alignment(horizontal="center"))
ws1.merge_cells("A3:F3")
c(ws1, 3, 1, "This is investment research, not personalized financial advice.",
    Font(italic=True, size=9, color="999999"), alignment=Alignment(horizontal="center"))

# Title data
title_data = [
    ("Ticker", "PSNL", "NasdaqGS"),
    ("Company", "Personalis, Inc.", "Cancer Genomics / Liquid Biopsy / Sequencing Platform"),
    ("Sector / Industry", "Healthcare / Diagnostics & Research", ""),
    ("As Of", "Aug 25, 2026", "Market close"),
    ("Stock Price", f"${price:.2f}", "Yahoo Finance"),
    ("Shares Outstanding", f"{shares_mm:.1f}M", "Implied from MC / Price"),
    ("Market Cap", f"${market_cap:.2f}B", "Yahoo Finance Statistics"),
    ("Enterprise Value", f"${ev:.2f}B", "Yahoo Finance Statistics"),
    ("Total Cash", f"${total_cash_mm:.0f}M", "Yahoo Finance Key Stats MRQ"),
    ("Total Debt", f"${total_debt_mm:.0f}M", "Implied 16.16% D/E x $261.2M equity"),
    ("Primary Valuation Lens", "P/S + Cash NAV floor", "Pre-profitability: DCF/FCF invalid"),
    ("Stance", "Watch", "Fully monetized on catalysts; 6-12Q to profitability"),
]

for ri, (field, val, note) in enumerate(title_data, 4):
    c(ws1, ri, 1, field, border_style=thin_border(), fill=light_blue_fill())
    c(ws1, ri, 2, val, bold_font(), border_style=thin_border())
    c(ws1, ri, 3, note, border_style=thin_border())

# Valuation metrics table
metrics_headers = ["Metric", "Value", "Comment"]
metrics_data = [
    ("Trailing P/E", "N/A", "Net income negative — trailing P/E mechanically meaningless"),
    ("Forward P/E", "N/A", "EPS remains negative through FY27 consensus (−$0.94)"),
    ("Price/Sales (TTM)", f"{market_cap / revenue_ttm:.1f}x", "vs. peer range 15-35x for diagnostics/genomics"),
    ("Price/Book (MRQ)", "7.75x", "Per Yahoo Finance Key Stats"),
    ("EV/Revenue", f"{ev / revenue_ttm * 1000:.1f}x", "=" + f"${ev*1000:.0f}M / ${revenue_ttm:.0f}M ≈ {ev*1000/revenue_ttm:.0f}x"),
    ("EV/EBITDA", "N/A", "EBITDA negative ($−97.2M TTM)"),
    ("P/FCF", "N/A", "FCF negative ($−56.0M TTM)"),
    ("Revenue per Share", f"${revenue_ttm / (shares_mm/1000):.2f}", f"=${revenue_ttm:.2f}M / {shares_mm/1000:.1f}M shares"),
    ("Cash per Share", f"${total_cash_mm / (shares_mm / 1000):.2f}", f"=${total_cash_mm:.0f}M / {(shares_mm/1000):.1f}M = NAV floor"),
    ("Beta (5Y)", f"{beta:.2f}", "High-beta growth/genomics volatility"),
    ("Gross Margin TTM", f"{revenue_ttm / 1000 * 100:.0f}% → {gross_profit_ttm / revenue_ttm * 100:.1f}%", "Severely compressed from 31.7% FY24 peak"),
    ("Net Margin TTM", f"{net_income_ttm / revenue_ttm * 100:.1f}%", "Pre-profitability — structural operating losses"),
    ("ROE TTM", "-50.32%", "Negative — burning through equity base"),
    ("ROA TTM", "-25.03%", "Per Yahoo Finance Key Stats"),
    ("Analyst Avg PT", "$15.44", "Below current $17.25; 7 analysts (1 Strong Buy, 4 Buy, 2 Hold)"),
    ("Analyst Low PT", "$13.00", "Lake Street Hold target"),
    ("52-Week Range", "$4.65 − $18.89", "Currently within $1.64 of 52W high"),
]

next_row = write_table(ws1, len(title_data) + 5, metrics_headers, metrics_data, alt_row_color=True)

# Column widths
for col in range(1, 4):
    ws1.column_dimensions[get_column_letter(col)].width = 22 if col == 3 else 18

# ── Sheet 2: WACC ────────────────────────────────────────────────────
ws2 = wb.create_sheet("WACC")

ws2.merge_cells("A1:C1")
c(ws2, 1, 1, "PSNL — Weighted Average Cost of Capital", bold_font(14))
ws2.merge_cells("A2:C2")
c(ws2, 2, 1, "CAPM Framework | Beta elevated for pre-profitability genomics",
    Font(italic=True, size=10, color="666666"))

wacc_data = [
    ("", "", ""),
    ("CAPM Components", "", ""),
    ("Risk-Free Rate (10Y US Treasury)", f"{risk_free_rate:.3f}%", "CNBC Aug 25, 2026"),
    ("Equity Risk Premium", f"{erp:.1f}%", "Market assumption"),
    ("Beta (5Y Monthly)", f"{beta:.2f}", "Yahoo Finance Key Stats"),
    ("", "", ""),
    ("Cost of Equity (Ke)", f"{cost_of_equity:.2f}%", f"Rf + Beta × ERP = {risk_free_rate:.2f} + {beta:.2f} × {erp:.1f}"),
    ("", "", ""),
    ("Cost of Debt (Kd)", f"{cost_of_debt:.1f}%", "Estimated — minimal debt"),
    ("Tax Rate", f"{tax_rate:.0f}%", "No meaningful tax when unprofitable"),
    ("", "", ""),
    ("Capital Structure Inputs", "", ""),
    ("Total Debt", f"${total_debt_for_wacc_mm:.1f}M", "~16% of equity = ~$42M"),
    ("Market Cap", f"${market_cap*1000:.0f}M", f"${price:.2f} × {(shares_mm/1000):.1f}M"),
    ("Equity Weight", f"{eq_weight:.2%}", f"MC / (MC + D)"),
    ("Debt Weight", f"{td_weight:.2%}", f"D / (MC + D)"),
    ("", "", ""),
    ("WACC", f"{wacc:.2f}%", f"Weighted: {eq_weight:.0%} × Ke({cost_of_equity:.2f}%) + {td_weight:.0%} × Kd({cost_of_debt:.1f}%) × (1-0%)"),
]

for ri, (field, val, note) in enumerate(wacc_data, 3):
    cell = c(ws2, ri, 1, field, border_style=thin_border())
    if field and "WACC" in field and "Weighted" in field:
        cell.font = bold_font(12)
        cell.fill = green_fill()
    elif field and not note:
        cell.font = bold_font(10)
    c(ws2, ri, 2, val, border_style=thin_border(), font_style=bold_font() if (field and not note) else None)
    c(ws2, ri, 3, note, border_style=thin_border())
    if "WACC" in field and "Weighted" in field:
        ws2.cell(row=ri, column=2).fill = green_fill()

ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 42

# ── Sheet 3: Scenarios ───────────────────────────────────────────────
ws3 = wb.create_sheet("Scenarios")

ws3.merge_cells("A1:K1")
c(ws3, 1, 1, "PSNL — Bear / Base / Bull Scenario Analysis (P/S Framework)", bold_font(14))
ws3.merge_cells("A2:K2")
c(ws3, 2, 1, "Pre-profitability genomics: DCF/FCF invalid. Bear anchored near current price. "
              "Revenue growth + gross margin trajectory are primary drivers.",
    Font(italic=True, size=10, color="666666"))
ws3.merge_cells("A3:K3")
c(ws3, 3, 1, "NOTE: FCF and DCF framework are NOT used — operating losses of $88-114M/yr make them meaningless.",
    Font(color="CC0000", size=10, bold=True))

# Scenario headers at row 5
scenario_headers = ["Metric", "Bear", "Base", "Bull", "Note"]

# Revenue growth and gross margin assumptions
# Current: $69.67M revenue, ~12.7% GM, Tempus deal catalyst
# Analyst consensus: $82.47M (FY26), $110.2M (FY27)
# 5-year base: assume reaches ~$250M+ by year 5

scenarios = [
    ("", "", "", "", ""),
    ("Assumptions", "", "", "", ""),
    ("Revenue CAGR (5Y)", "12%", "20%", "32%", "Bear: slow; Base: consensus-like; Bull: Tempus-scale commercialization"),
    ("Terminal Revenue (Year 5)", "$126M", "$173M", "$290M", "Compound from $69.7M TTM"),
    ("Adjusted Gross Margin", "14%", "22%", "30%", "Bear: stays compressed; Bull: approaches FY24 peak"),
    ("Terminal Gross Profit", "$17.6M", "$38.1M", "$87.0M", "Rev × GM%"),
    ("Operating Expense Ratio", "180% of Rev", "140% of Rev", "100% of Rev", "Stags behind rev growth in Bull = operating leverage"),
    ("OpEx $M", "$226.8M", "$242.2M", "$290.0M", "CAGR lower than revenue"),
    ("Terminal Operating Income", "-$209.2M", "-$204.1M", "-$203.0M", "Still negative year 5 in all cases"),
    ("", "", "", "", ""),
    ("Exit Valuation via P/S", "", "", "", ""),
    ("Exit P/S Multiple", "8x", "15x", "25x", "Bear: distressed pre-profit; Base: sector norm; Bull: growth premium"),
    ("Implied Market Cap", "$1.01B", "$2.60B", "$7.26B", "Terminal Rev × P/S"),
    ("Less: Net Debt Adj", "$(170.5M)", "$(170.5M)", "$(170.5M)", "Subtract cash surplus (Net Cash = $170.5M, NEG debt)"),
    ("Implied Enterprise Value", "$1.18B", "$2.77B", "$7.43B", "MC + Net Debt (neg = plus cash)"),
    ("", "", "", "", ""),
    ("Dilution Assumptions (5Y)", "", "", "", ""),
    ("Future Dilution Factor (1+X)", "1.50x", "1.30x", "1.15x", "Bear: needs more capital; Bull: self-funding sooner"),
    ("Diluted Share Count (M)", "158.3", "137.1", "121.4", f"Current {shares_mm/1000:.1f}M × factor"),
    ("", "", "", ""),
    ("Target Price", "", "", "", ""),
    ("Target Price / Share", f"${1010/158.3:.2f}", f"${2600/137.1:.2f}", f"${7260/121.4:.2f}", "Implied MC / diluted shares"),
    ("Upside from Current", f"{(1010/158.3 / price - 1)*100:.0f}%", f"{(2600/137.1 / price - 1)*100:.0f}%", f"{(7260/121.4 / price - 1)*100:.0f}%", ""),
    ("vs 52W High ($18.89)", f"{(1010/158.3 / 18.89 - 1)*100:.0f}%", f"{(2600/137.1 / 18.89 - 1)*100:.0f}%", f"{(7260/121.4 / 18.89 - 1)*100:.0f}%", ""),
    ("", "", "", "", ""),
    ("Probability Weights", "", "", "", ""),
    ("Bear Weight", "25%", "", "", "Catalysts fail, GM stays compressed"),
    ("Base Weight", "50%", "", "", "Consensus-ish, slow path to profitability"),
    ("Bull Weight", "25%", "", "", "Tempus deal scales, GM recovers"),
    ("", "", "", "", ""),
    ("Probability-Weighted FV", "", "", "", ""),
    ("Weighted Value / Share", "", "", f"${0.25*1010/158.3 + 0.50*2600/137.1 + 0.25*7260/121.4:.2f}", "Sum of weights × targets"),
    ("Total Probability-Weighted FV", "", "", f"${0.25*1010/158.3 + 0.50*2600/137.1 + 0.25*7260/121.4:.2f}", ""),
    ("Current Price", "", "", f"${price:.2f}", ""),
    ("Implied Upside", "", "", f"{(0.25*1010/158.3 + 0.50*2600/137.1 + 0.25*7260/121.4) / price - 1:.1%}", ""),
    ("", "", "", "", ""),
    ("Cash NAV Floor", "", "", "", ""),
    ("Cash / Share (Current)", "", "", f"${total_cash_mm / (shares_mm / 1000):.2f}", f"${total_cash_mm:.0f}M / {(shares_mm/1000):.1f}M"),
    ("Cash / Share (5Y Diluted Base)", "", "", f"${150 / 137.1:.2f}", "Burn-adjusted cash / diluted shares"),
    ("", "", "", "", ""),
    ("Framework Note", "", "", "", ""),
    ("Primary Framework", "", "", "P/S + Cash NAV", "DCF/FCF invalid — operating losses $88-114M/yr"),
    ("Cross-Check", "", "", "P/B at 7.75x current", "Declining BVPS — equity being consumed"),
    ("Key Variable", "", "", "Gross margin trajectory", "GM was 31.7% in FY24; TTM only 12.7%"),
]

write_table(ws3, 5, scenario_headers, scenarios, alt_row_color=True)

# Widths
for col, w in enumerate([22, 14, 14, 14, 44], 1):
    ws3.column_dimensions[get_column_letter(col)].width = w

print(f"Scenario Bear target: ${1010/158.3:.2f} ({(1010/158.3 / price - 1)*100:.0f}% upside)")
print(f"Scenario Base target: ${2600/137.1:.2f} ({(2600/137.1 / price - 1)*100:.0f}% upside)")
print(f"Scenario Bull target: ${7260/121.4:.2f} ({(7260/121.4 / price - 1)*100:.0f}% upside)")
fv = 0.25*1010/158.3 + 0.50*2600/137.1 + 0.25*7260/121.4
print(f"Probability-Weighted FV: ${fv:.2f} ({(fv/price-1)*100:.0f}% upside)")

# ── Sheet 4: Actuals Source Audit ────────────────────────────────────
ws4 = wb.create_sheet("Actuals Source Audit")

ws4.merge_cells("A1:E1")
c(ws4, 1, 1, "PSNL — Data Source Audit", bold_font(14))

audit_headers = ["Data Point", "Value", "Source", "Date", "Notes"]
audit_data = [
    ("Stock Price", "$17.25", "Yahoo Finance", "Aug 25, 2026", "Market close"),
    ("After Hours Price", "$17.07", "Yahoo Finance", "Aug 25, 2026", "Post-market ATS"),
    ("Market Cap", "$1.82B", "Yahoo Finance Key Stats", "Aug 25, 2026", "MRQ"),
    ("Enterprise Value", "$1.65B", "Yahoo Finance Key Stats", "Aug 25, 2026", "MRQ"),
    ("Shares Outstanding", "~105.5M", "Implied: MC/Price", "Aug 25, 2026", "$1.82B/$17.25"),
    ("Beta (5Y Monthly)", "2.29", "Yahoo Finance Key Stats", "Aug 25, 2026", ""),
    ("Revenue TTM", "$69.67M", "Yahoo Finance IS", "TTM as of Jun 2026", "All numbers in thousands"),
    ("Gross Profit TTM", "$8.84M", "Yahoo Finance IS", "TTM", "GM severely compressed to 12.7%"),
    ("OpEx TTM", "$123.15M", "Yahoo Finance IS", "TTM", "Operating expense far exceeds revenue"),
    ("Operating Income TTM", "-$114.31M", "Yahoo Finance IS", "TTM", ""),
    ("Net Income TTM", "-$107.18M", "Yahoo Finance IS", "TTM", "Consistent operating losses"),
    ("EPS Diluted TTM", "-$1.09", "Yahoo Finance IS", "TTM", "Basic = diluted (no options dilution counted)"),
    ("EBITDA TTM", "-$97.22M", "Yahoo Finance IS", "TTM", "Negative — DCF invalid"),
    ("Operating CF TTM", "-$92.30M", "Yahoo Finance CF", "TTM", "Cash burn continues"),
    ("Levered FCF TTM", "-$56.03M", "Yahoo Finance Key Stats", "TTM", "Structurally negative"),
    ("Total Cash MRQ", "$212.66M", "Yahoo Finance Key Stats", "MRQ Jun 2026", "Primary liquidity buffer"),
    ("Total Debt/Equity MRQ", "16.16%", "Yahoo Finance Key Stats", "MRQ Jun 2026", "Low leverage — mostly cash-funded"),
    ("P/S Ratio TTM", "23.93x", "Yahoo Finance Key Stats", "TTM", "High even for growth diagnostics"),
    ("P/B Ratio MRQ", "7.75x", "Yahoo Finance Key Stats", "MRQ", "Elevated for unprofitable company"),
    ("Profit Margin TTM", "-153.84%", "Yahoo Finance Key Stats", "TTM", ""),
    ("ROE TTM", "-50.32%", "Yahoo Finance Key Stats", "TTM", ""),
    ("ROA TTM", "-25.03%", "Yahoo Finance Key Stats", "TTM", ""),
    ("52W Range", "$4.65 − $18.89", "Yahoo Finance Summary", "Aug 25, 2026", "Near 52W high"),
    ("Next Earnings Date", "Nov 3, 2026", "Yahoo Finance Summary", "Estimated", "Q3 FY26"),
    ("Analyst Avg PT", "$15.44", "Yahoo Finance Analysis", "Aug 25, 2026", "7 analysts"),
    ("Analyst Low PT", "$13.00", "Yahoo Finance Analysis", "Aug 25, 2026", "Lake Street Hold"),
    ("FY26 Revenue Est.", "$82.47M", "Yahoo Finance Analysis", "Consensus", "82 analysts avg"),
    ("FY27 Revenue Est.", "$110.20M", "Yahoo Finance Analysis", "Consensus", "82 analysts avg"),
    ("FY26 EPS Est.", "-$1.09", "Yahoo Finance Analysis", "GAAP", "Worse than FY25 -0.91"),
    ("FY27 EPS Est.", "-$0.94", "Yahoo Finance Analysis", "GAAP", "Still negative"),
    ("10Y Treasury Rate", "4.641%", "CNBC", "Aug 25, 2026", "Used for WACC"),
    ("Employees", "259", "Yahoo Finance Profile", "Current", ""),
    ("Fiscal Year End", "Dec 31", "Yahoo Finance Profile", "Current", ""),
]

write_table(ws4, 3, audit_headers, audit_data, alt_row_color=True)

for col, w in enumerate([24, 15, 22, 18, 38], 1):
    ws4.column_dimensions[get_column_letter(col)].width = w

# ── Sheet 5: Questions ───────────────────────────────────────────────
ws5 = wb.create_sheet("Questions")

ws5.merge_cells("A1:C1")
c(ws5, 1, 1, "PSNL — Open Questions", bold_font(14))

questions = [
    ("Q1", "What drove the gross margin collapse from 31.7% (FY24) to 12.7% (TTM)?",
     "FY24 GM was 31.7% ($26.8M/$84.6M). FY25 collapsed to 22.7% ($15.8M/$69.6M). "
     "TTM at 12.7% ($8.8M/$69.7M). Is this a structural reset (lower-priced test volume) or transitional (cost restructuring)?"),
    ("Q2", "How meaningful is the Tempus AI collaboration for revenue?",
     "Yahoo Finance news shows Tempus AI-PSNL deal as a major catalyst. What % of forward revenue does this collaboration represent? "
     "Is it revenue-bearing or just a distribution partnership?"),
    ("Q3", "What is the path to operating profitability?",
     "Operating losses of $68-114M/year for 4 consecutive years. Revenue ~$70-85M. Operating expenses ~$95-123M/year. "
     "Does the company need 3-4x revenue just to breakeven at current cost structure?"),
    ("Q4", "What is the dilution trajectory?",
     "Shares grew from 45.7M (FY22) to 89.2M (FY25) to 97.6M (TTM) — ~113% dilution in 3 years. "
     "At current cash burn of ~$75-92M/year and $213M cash, runway is ~2-3 years without more raises."),
    ("Q5", "Is the stock overextended after 267% 1-year return?",
     "Up 267% YoY driven by Tempus deal + Moderna mRNA vaccine catalyst. Currently at $17.25, within $1.64 of 52W high. "
     "Analyst avg PT of $15.44 is BELOW current price — suggests limited upside consensus."),
    ("Q6", "What explains the revenue decline from FY24 ($84.6M) to FY25 ($69.6M)?",
     "-17.7% revenue drop. Was this loss of research contracts (NIH grants ending?), customer churn, or a one-time FY24 spike?"),
    ("Q7", "How does PSNL differentiate from Guardant Health (GH) and GRAIL?",
     "Competitive landscape: GH ($22B MC), GRAIL ($3.7B MC) do liquid biopsy/MRD. PSNL at $1.82B. "
     "Does PSNL's sequencing platform give it a cost advantage?"),
    ("Q8", "What is the $136.4M TTM financing cash flow funding?",
     "Massive financing inflow TTM ($136.4M) vs prior year $130.3M. Equity raises? Convertible debt? "
     "What was the raise price vs current $17.25 — indicative of insider valuation expectations?"),
    ("Q9", "Can 259 employees justify the current cost structure?",
     "OpEx of $123M TTM on 259 employees = ~$475K/employee/year. For a genomics company, is this R&D-heavy or is there SG&A bloat?"),
    ("Q10", "What is the burn rate and cash runway?",
     "Operating CF -$92.3M TTM = ~$23M/quarter. Cash $212.7M → ~9 quarters at current burn. "
     "Will need another raise by Q4 FY27/Q1 FY28 at current pace."),
]

q_headers = ["#", "Question", "Context"]
write_table(ws5, 3, q_headers, questions)

ws5.column_dimensions["A"].width = 7
ws5.column_dimensions["B"].width = 42
ws5.column_dimensions["C"].width = 65

# ── Sheet 6: Sources ─────────────────────────────────────────────────
ws6 = wb.create_sheet("Sources")

ws6.merge_cells("A1:B1")
c(ws6, 1, 1, "Primary Data Sources", bold_font(14))

sources = [
    ("1", "Yahoo Finance — PSNL Quote Page", "finance.yahoo.com/quote/PSNL/"),
    ("2", "Yahoo Finance — Key Statistics", "finance.yahoo.com/quote/PSNL/key-statistics/"),
    ("3", "Yahoo Finance — Income Statement (Annual)", "finance.yahoo.com/quote/PSNL/financials/"),
    ("4", "Yahoo Finance — Balance Sheet (Annual)", "finance.yahoo.com/quote/PSNL/balance-sheet/"),
    ("5", "Yahoo Finance — Cash Flow (Annual)", "finance.yahoo.com/quote/PSNL/cash-flow/"),
    ("6", "Yahoo Finance — Analyst Estimates & Trends", "finance.yahoo.com/quote/PSNL/analysis/"),
    ("7", "Yahoo Finance — Company Profile", "finance.yahoo.com/quote/PSNL/profile/"),
    ("8", "CNBC — US 10Y Treasury Yield", "www.cnbc.com/quotes/US10Y"),
    ("9", "StockAnalysis.com — 404 (not available for PSNL)", "stockanalysis.com/quote/PSNL/"),
    ("10", "Yahoo Finance News Tab — Recent headlines", "finance.yahoo.com/quote/PSNL/news/"),
    ("11", "Yahoo Finance — Analyst Ratings summary", "From Analysis page extraction"),
    ("", "", ""),
]

s_headers = ["#", "Source", "URL"]
write_table(ws6, 3, s_headers, sources)

ws6.column_dimensions["A"].width = 6
ws6.column_dimensions["B"].width = 48
ws6.column_dimensions["C"].width = 45

# ── Save ─────────────────────────────────────────────────────────────
outpath = "/home/refcell/dev/capital/models/[2026-08-25] PSNL Model.xlsx"
wb.save(outpath)
print(f"\nWorkbook saved to {outpath}")

# Verification
wb2 = openpyxl.load_workbook(outpath)
print(f"Sheets: {wb2.sheetnames}")
for s in wb2.worksheets:
    print(f"  {s.title}: {s.max_row} rows x {s.max_column} cols")
