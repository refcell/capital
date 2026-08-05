"""Build 6-sheet valuation model for Genius Sports Limited (GENI).

Sources:
  - Yahoo Finance Income Statement / Balance Sheet / Cash Flow / Statistics / Analysis
  - CNBC US10Y: 4.611% (2026-08-04)
  - Quote date: 2026-08-04, close $8.08
  - StockAnalysis.com returned 404; Yahoo Finance is primary source.

All $ figures in millions unless noted otherwise.

Key characteristics:
  - Growth SaaS/sports-data company approaching profitability
  - Revenue growing 23%+ YoY, gross margin expanding (0.8% → 22.7% in 3 years)
  - Still operating at a loss ($-157M TTM op income)
  - FCF negative TTM (-$30.3M) from heavy capex ($81.1M) — data center/infrastructure investment
  - $144M capital raise TTM
  - Primary valuation lens: Forward P/E and P/S — FCF framework invalid while negative
  - 18-19 analyst coverage, Q1 FY26 beat (+65% EPS surprise)
"""

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ───────────────────────────────────────────────────────────
HERE = Path(__file__).resolve().parent
XLSX = HERE / "[2026-08-04] Genius Sports Model.xlsx"

# ── Raw inputs ─────────────────────────────────────────────────────
PRICE = 8.08
SHARES_M = 267.34      # implied shares outstanding (Yahoo Stats, Aug 4 2026)
MC_B = 1.85            # $B market cap
EV_B = 1.69            # $B enterprise value
CASH_M = 197.44        # Q1 FY2026 total cash per Yahoo Stats (or End Cash $194.3M on CF stmt)
TOTAL_DEBT_M = 30.55   # capital lease obligations per Yahoo Stats Q1 FY2026
NET_DEBT_M = CASH_M - TOTAL_DEBT_M  # ~$166.9M net cash position
# Verify EV-MC: 1.69 - 1.85 = -0.16 billion = -$160M, consistent with net cash

# Income statement ($M, from Yahoo Finance /financials/)
# TTM       FY2025   FY2024   FY2023   FY2022
REV     = {"TTM": 713.45, "FY2025": 669.49, "FY2024": 510.89, "FY2023": 412.98, "FY2022": 341.03}
GP      = {"TTM": 161.96, "FY2025": 153.84, "FY2024": 128.71, "FY2023":  69.01, "FY2022":   2.86}
GM_PCT  = {"TTM": 0.227,  "FY2025": 0.229,  "FY2024": 0.252,  "FY2023": 0.167,  "FY2022": 0.008}
OP_IN   = {"TTM": -157.35,"FY2025": -141.38,"FY2024":  -56.29,"FY2023":  -71.66,"FY2022": -181.20}
NI      = {"TTM": -158.85,"FY2025": -111.58,"FY2024":  -63.04,"FY2023":  -85.53,"FY2022": -181.64}
EPS_D   = {"TTM": -0.62,  "FY2025": -0.44,  "FY2024": -0.27,  "FY2023": -0.38,  "FY2022": -0.91}
EBITDA  = {"TTM": -101.25,"FY2025":  -70.86,"FY2024":   16.28,"FY2023":    5.64,"FY2022": -112.68}
FCF     = {"TTM":  -30.34,"FY2025":    5.50,"FY2024":   17.61,"FY2023":  -34.27,"FY2022": -51.01}
OCF     = {"TTM":   50.78,"FY2025":   86.40,"FY2024":   81.86,"FY2023":   14.88,"FY2022":  -3.46}
CAPEX   = {"TTM":   81.12,"FY2025":   80.89,"FY2024":   64.26,"FY2023":   49.14,"FY2022":  47.55}
DA      = {"TTM":   75.10,"FY2025":   70.52,"FY2024":   72.57,"FY2023":   77.31,"FY2022":  68.53}

# Balance sheet ($M, from Yahoo Finance /balance-sheet/)
CASH_FY25_M = 280.56     # FY2025 end cash
DEBT_TOTAL_FY25_M = 30.50
TANGIBLE_BOOK_FY25_M = 242.22   # net tangible assets
SHARES_ISSUED_FY25_M = 250.41   # share issued count per BS

# Tax, beta, rates
# Beta: 1.88 (5Y Monthly, Yahoo Stats)
# Effective tax rate is unusual — use 21% corporate rate
TAX_RATE = 0.21
BETA = 1.88
RISK_FREE = 0.04611    # 10Y US Treasury, CNBC Aug 4 2026
ERP = 0.05             # equity risk premium

# Analyst estimates (Yahoo Finance /analysis/, Aug 4 2026)
# 18 rev / 18 EPS for Q2 FY26, 19 rev / 18 EPS for FY26, 18 EPS for FY27
EPS_F2026 = 0.52       # FY2026 EPS consensus, non-GAAP normalized
EPS_F2027 = 0.97       # FY2027 EPS consensus, non-GAAP normalized
REV_F2026 = 1000.0     # FY2026 revenue consensus ($1B)
REV_F2027 = 1290.0     # FY2027 revenue consensus ($1.29B)

# Q1 FY26 EPS actual: $0.016 vs est $0.01 (+65% surprise)
# Q1 FY26 revenue: $187.95M, earnings: -$55.47M
# Historical EPS: Q1 FY26 actual $0.02 vs est $0.01 (beat)

# Capital raise: $144M issuance of capital stock TTM

# ── FCF Framework Check ──────────────────────────────────────────
# FCF = -$30.3M TTM (negative). FY2025 FCF = $5.5M.
# FCF framework is INVALID — negative FCF and capex investment cycle.
# Also: net debt is actually net CASH ($166.9M). No debt amplification issue.
# Primary framework: Forward P/E on analyst consensus.
# Cross-check: P/S ratio.

# ── Computed ─────────────────────────────────────────────────────
COE = RISK_FREE + BETA * ERP  # CAPM cost of equity
COST_OF_DEBT = RISK_FREE + 0.02  # 200 bps spread — very low debt
EQUITY_WEIGHT = MC_B / (MC_B + max(NET_DEBT_M / 1000, 0.01))
DEBT_WEIGHT = max(NET_DEBT_M / 1000, 0.01) / (MC_B + max(NET_DEBT_M / 1000, 0.01))
WACC = EQUITY_WEIGHT * COE + DEBT_WEIGHT * COST_OF_DEBT * (1 - TAX_RATE)

# Valuation metrics
PE_TRAIL = "N/A (-)"          # trailing EPS negative
PE_FWD_2026 = PRICE / EPS_F2026 if EPS_F2026 > 0 else None
PE_FWD_2027 = PRICE / EPS_F2027 if EPS_F2027 > 0 else None
PS = MC_B / (REV["TTM"] / 1000)
EV_REV = EV_B / (REV["TTM"] / 1000)
# P/FCF and EV/FCF: N/A — FCF negative TTM
PB = MC_B / (REV["TTM"] / 1000) * (REV["TTM"] / 1000) / (TANGIBLE_BOOK_FY25_M / 1000)
# Actually P/B = MC / Book Value
BOOK_VALUE_B = TANGIBLE_BOOK_FY25_M / 1000  # $242.22M
P_BOOK = MC_B / BOOK_VALUE_B  # ~7.64x on tangible book
# But common stock equity is $724.5M → P/B = 1.85 / 0.725 = 2.55x (matches Yahoo's 2.67x)
P_BOOK_COMMON = MC_B / (724.475 / 1000)  # ~2.55x

# Revenue growth rates
REV_GROW_FY23_FY22 = (412.98 - 341.03) / 341.03
REV_GROW_FY24_FY23 = (510.89 - 412.98) / 412.98
REV_GROW_FY25_FY24 = (669.49 - 510.89) / 510.89
REV_GROW_TTM_FY25 = (713.45 - 669.49) / 669.49

# Gross margin expansion is the key: 0.8% → 22.7% in 3 years
# This is the single most important metric — if GM continues to expand,
# operating losses narrow toward breakeven.

# Scenario Parameters ──────────────────────────────────────────
# Framework: Forward P/E — uses analyst EPS consensus as anchor
# Bear: profitability delayed, margins compress, multiple contracts to single digits
# Base: analysts right, operating leverage delivers, premium multiple for niche monopoly
# Bull: GM continues to expand toward 30%+, market share deepens, multiple sustains

# Revenue CAGR assumptions (from FY2027 $1.29B base):
BEAR_REV_CAGR_5Y = 0.08   # organic growth decelerates; sports betting regulations slow down
BASE_REV_CAGR_5Y = 0.16   # continues analyst trajectory
BULL_REV_CAGR_5Y = 0.22   # accelerates as betting markets expand globally

# EPS projections (5-year, reflects operating leverage trajectory):
# FY2026: $0.52, FY2027: $0.97 — base is already doubling EPS in one year
BEAR_EPS_5Y = 0.35        # barely profitable; OpEx runs hot, GM stalls at ~20%
BASE_EPS_5Y = 1.60        # GM expands to 28%, OpEx normalizes to 35% of revenue
BULL_EPS_5Y = 2.80        # GM hits 32%, OpEx falls to 30% — platform economics kick in

# 10-year:
BEAR_EPS_10Y = 0.50       # still marginal profitability after a decade
BASE_EPS_10Y = 3.00       # robust profitable platform
BULL_EPS_10Y = 5.00       # best-in-class margins on $4B+ revenue

# Exit multiples (P/E):
BEAR_PE = 12.0            # compressed to below-growth-peer norms
BASE_PE = 20.0            # growth tech at stable profitability (peer median)
BULL_PE = 28.0            # monopoly premium in sports data + streaming

# Scenario computations
BEAR_5Y_TARGET = BEAR_EPS_5Y * BEAR_PE
BASE_5Y_TARGET = BASE_EPS_5Y * BASE_PE
BULL_5Y_TARGET = BULL_EPS_5Y * BULL_PE

# 10-year targets
BEAR_10Y_TARGET = BEAR_EPS_10Y * BEAR_PE
BASE_10Y_TARGET = BASE_EPS_10Y * BASE_PE
BULL_10Y_TARGET = BULL_EPS_10Y * BULL_PE

# Weights (probability): BEAR 25%, BASE 50%, BULL 25%
BEAR_W = 0.25
BASE_W = 0.50
BULL_W = 0.25

# Weighted FV
WTD_FV = BEAR_W * BEAR_5Y_TARGET + BASE_W * BASE_5Y_TARGET + BULL_W * BULL_5Y_TARGET

import sys, warnings
warnings.filterwarnings('ignore', category=UserWarning)
print(f"WACC = {WACC:.4f} ({WACC*100:.2f}%)")
print(f"COE = {COE:.4f} ({COE*100:.2f}%), CDE = {COST_OF_DEBT:.4f}")
print(f"Equity weight = {EQUITY_WEIGHT:.4f}, Debt weight = {DEBT_WEIGHT:.4f}")
print(f"Fwd P/E FY26 = {PE_FWD_2026:.1f}x, FY27 = {PE_FWD_2027:.1f}x")
print(f"P/S = {PS:.2f}x, P/B (common) = {P_BOOK_COMMON:.2f}x")
print(f"Scenario targets — Bear 5Y: ${BEAR_5Y_TARGET:.2f}, Base: ${BASE_5Y_TARGET:.2f}, Bull: ${BULL_5Y_TARGET:.2f}")
print(f"Probability-Weighted FV: ${WTD_FV:.2f}")
print(f"Upside from ${PRICE}: {((WTD_FV - PRICE) / PRICE * 100):.1f}%")

# ── Styles ────────────────────────────────────────────────────────
HDR_FONT = Font(bold=True, size=14)
SUB_FONT = Font(bold=True, size=11)
BOLD = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
HDR_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
HDR_FONT2 = Font(bold=True, color="000000")
WRAP = Alignment(wrap_text=True, vertical="top")

def c(ws, row, col, value, font=None, border=False, fill=False, align=None):
    """Helper: write cell value with optional formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if border:
        cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    else:
        cell.alignment = WRAP
    return cell

def write_table(ws, data, start_row, col_widths=None):
    """Write a table starting at start_row. data is list of lists."""
    for ri, row_data in enumerate(data, start_row):
        for ci, val in enumerate(row_data, 1):
            is_header = ri == start_row
            c(ws, ri, ci, val,
              font=BOLD if is_header else None,
              border=True, fill=HDR_FILL if is_header else False)
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

# ── Create workbook ──────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════
# Sheet 1: Valuation
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Valuation"

# Merge title
ws1.merge_cells("A1:F1")
c(ws1, 1, 1, "Genius Sports Limited (GENI) — Valuation Model", font=HDR_FONT)

# Title block
ws1.merge_cells("A2:F2")
c(ws1, 2, 1, f"As of August 4, 2026 — Price $8.08, NYSE:GENI", font=SUB_FONT)

title_data = [
    ["Company", "Genius Sports Limited", ""],
    ["Ticker", "NYSE: GENI", ""],
    ["Date", "2026-08-04", ""],
    ["Close Price", f"${PRICE:.2f}", "NYSE close Aug 4, 2026 (+14.29%)"],
    ["Shares Outstanding", f"{SHARES_M:.1f}M", "Yahoo Finance Key Statistics, Q1 FY2026"],
    ["Market Cap", f"${MC_B * 1000:.0f}M / ${MC_B:.2f}B", "Yahoo Finance"],
    ["Enterprise Value", f"${EV_B * 1000:.0f}M / ${EV_B:.2f}B", "Yahoo Finance — actually below MC due to net cash"],
    ["Cash", f"${CASH_M:.0f}M", "Q1 FY2026 Yahoo Stats — net cash of ~$167M"],
    ["Total Debt", f"${TOTAL_DEBT_M:.0f}M", "Capital lease obligations, Q1 FY2026"],
    ["Net Cash Position", f"${NET_DEBT_M:.0f}M", f"Cash ${CASH_M:.0f}M less debt ${TOTAL_DEBT_M:.0f}M"],
    ["Primary Lens", "Forward P/E", "TCF invalid: FCF negative TTM from capex cycle"],
    ["Stance", "Watch", "Pre-profitability growth; earnings call in 2 days"],
    ["Sector", "Communication Services", "Internet Content & Information"],
]

for i, (field, value, note) in enumerate(title_data, 3):
    c(ws1, i, 1, field, font=BOLD, border=True, fill=False)
    c(ws1, i, 2, value, border=True)
    c(ws1, i, 3, note, border=True)

ws1.column_dimensions["A"].width = 18
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 55

# Valuation metrics table
c(ws1, len(title_data) + 4, 1, "Valuation Metrics", font=SUB_FONT)
c(ws1, len(title_data) + 4, 1, "Valuation Metrics", font=SUB_FONT)

metrics_data = [
    ["Metric", "Value", "Comment"],
    ["Trailing P/E", "N/A (negative EPS)", f"TTM EPS ${EPS_D['TTM']:.2f} — company not yet profitable"],
    ["Forward P/E (FY2026)", f"{PE_FWD_2026:.1f}x", f"Using consensus ${EPS_F2026:.2f} EPS, 18 analysts"],
    ["Forward P/E (FY2027)", f"{PE_FWD_2027:.1f}x", f"Using consensus ${EPS_F2027:.2f} EPS"],
    ["P/Sales (TTM)", f"{PS:.2f}x", "Revenue $713.5M TTM. Reasonable for growth SaaS."],
    ["EV/Sales", f"{EV_REV:.2f}x", "EV below MC due to net cash position."],
    ["P/Book (Common)", f"{P_BOOK_COMMON:.2f}x", f"Common equity $724.5M"],
    ["P/Tangible Book", f"{P_BOOK:.2f}x", f"Tangible book $242.2M; intangibles significant from M&A"],
    ["P/FCF", "N/A", "FCF negative TTM (-$30.3M) from capex investment cycle"],
    ["EV/FCF", "N/A", "Same reason — FCF negative"],
    ["EV/EBITDA", "N/A", "EBITDA negative TTM (-$101.3M)"],
    ["Beta (5Y)", f"{BETA:.2f}", "Yahoo Finance Key Statistics — high-beta growth stock"],
]

for i, (metric, val, comment) in enumerate(metrics_data, len(title_data) + 5):
    c(ws1, i, 1, metric, font=BOLD if metric == metrics_data[0][0] else None, border=True, fill=HDR_FILL if i == len(title_data)+5 else False)
    c(ws1, i, 2, val, border=True)
    c(ws1, i, 3, comment, border=True)

# ═══════════════════════════════════════════════════════════════
# Sheet 2: WACC
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("WACC")

ws2.merge_cells("A1:E1")
c(ws2, 1, 1, "WACC Calculation — Genius Sports Limited", font=HDR_FONT)

wacc_data = [
    ["Component", "Value", "Source / Notes"],
    ["Risk-Free Rate (10Y US Treasury)", f"{RISK_FREE * 100:.3f}%", "CNBC US10Y, Aug 4 2026"],
    ["Equity Risk Premium", f"{ERP * 100:.1f}%", "Standard assumption"],
    ["Beta (Levered, 5Y Monthly)", f"{BETA:.2f}", "Yahoo Finance Key Statistics"],
    ["Cost of Equity (CAPM)", f"{COE * 100:.2f}%", f"= {RISK_FREE*100:.2f}% + {BETA:.2f} × {ERP*100:.0f}%"],
    ["Cost of Debt", f"{COST_OF_DEBT * 100:.2f}%", f"Risk-free + 200 bps spread"],
    ["Tax Rate", f"{TAX_RATE * 100:.0f}%", "US/UK corporate rate proxy"],
    ["Market Cap", f"${MC_B:.2f}B", "Price $8.08 × 267.3M shares"],
    ["Total Debt", f"${TOTAL_DEBT_M:.0f}M", "Capital lease obligations, Q1 FY2026"],
    ["Net Cash/Cash", f"${CASH_M:.0f}M", "Q1 FY2026 total cash"],
    ["Equity Weight", f"{EQUITY_WEIGHT:.4f}", "Near 100% — company is effectively all-equity financed"],
    ["Debt Weight", f"{DEBT_WEIGHT:.4f}", "Debt is trivial relative to MC"],
    ["WACC", f"{WACC * 100:.2f}%", f"= {EQUITY_WEIGHT:.2%} × {COE*100:.2f}% + {DEBT_WEIGHT:.2%} × {COST_OF_DEBT*100:.2f}% × (1 - {TAX_RATE})"],
]

for i, (comp, val, note) in enumerate(wacc_data, 2):
    c(ws2, i, 1, comp, font=BOLD if comp == wacc_data[0][0] else None, border=True, fill=HDR_FILL if i==2 else False)
    c(ws2, i, 2, val, border=True)
    c(ws2, i, 3, note, border=True)

ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 60

# ═══════════════════════════════════════════════════════════════
# Sheet 3: Scenarios
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Scenarios")

ws3.merge_cells("A1:H1")
c(ws3, 1, 1, "Scenario Analysis — Genius Sports Limited", font=HDR_FONT)
ws3.merge_cells("A2:H2")
c(ws3, 2, 1, "Framework: Forward P/E (not FCF — FCF negative TTM from capex investment cycle)", font=SUB_FONT)

# Scenario headers and data
# Key: this uses P/E framework because FCF is negative TTM
scenarios_data = [
    ["", "Bear", "Base", "Bull", "Weight", "Weighted", "", ""],
    ["Revenue CAGR (5Y)", "12.0%", "18.0%", "25.0%", "", "", "", ""],
    ["Terminal Revenue (5Y) from FY27 $1.29B", f"${(1290*(1+BEAR_REV_CAGR_5Y)**5):.0f}M", f"${(1290*(1+BASE_REV_CAGR_5Y)**5):.0f}M", f"${(1290*(1+BULL_REV_CAGR_5Y)**5):.0f}M", "", "", "", ""],
    ["Terminal EPS (5Y)", f"${BEAR_EPS_5Y:.2f}", f"${BASE_EPS_5Y:.2f}", f"${BULL_EPS_5Y:.2f}", "", "", "", ""],
    ["Operating Margin (5Y)", "~5%", "~12%", "~18%", "", "", "", ""],
    ["Terminal FCF Margin (5Y)", "~2%", "~8%", "~14%", "", "", "", ""],
    ["Exit P/E Multiple", f"{BEAR_PE:.0f}x", f"{BASE_PE:.0f}x", f"{BULL_PE:.0f}x", "", "", "", ""],
    ["Target Price (5Y)", f"${BEAR_5Y_TARGET:.2f}", f"${BASE_5Y_TARGET:.2f}", f"${BULL_5Y_TARGET:.2f}", "", "", "", ""],
    ["Upside from $8.08", f"{(BEAR_5Y_TARGET/PRICE - 1)*100:.0f}%", f"{(BASE_5Y_TARGET/PRICE - 1)*100:.0f}%", f"{(BULL_5Y_TARGET/PRICE - 1)*100:.0f}%", "", "", "", ""],
    ["Case Weight", f"{BEAR_W:.0%}", f"{BASE_W:.0%}", f"{BULL_W:.0%}", "", "", "", ""],
    ["Weighted Value/Share", f"${BEAR_W * BEAR_5Y_TARGET:.2f}", f"${BASE_W * BASE_5Y_TARGET:.2f}", f"${BULL_W * BULL_5Y_TARGET:.2f}", "", "", "", ""],
    ["", "", "", "", "", "", "", ""],
    ["Probability-Weighted FV", "", "", f"${WTD_FV:.2f}", "Sum of weighted", "", "", ""],
    ["Current Price", "", "", f"${PRICE:.2f}", "", "", "", ""],
    ["Upside from Current", "", "", f"{(WTD_FV/PRICE - 1)*100:.1f}%", "", "", "", ""],
]

for ri, row_data in enumerate(scenarios_data, 3):
    for ci, val in enumerate(row_data, 1):
        is_header = ri == 3
        is_summary = ri >= 12
        c(ws3, ri, ci, val,
          font=BOLD if is_header or is_summary else None,
          border=True, fill=HDR_FILL if is_header else False)

ws3.column_dimensions["A"].width = 35
for ci in range(2, 9):
    ws3.column_dimensions[get_column_letter(ci)].width = 16

# 10-year addendum below
c(ws3, 18, 1, "10-Year Extended Scenarios (P/E Framework)", font=SUB_FONT)
ext_data = [
    ["", "Bear", "Base", "Bull"],
    ["Terminal EPS (10Y)", f"${BEAR_EPS_10Y:.2f}", f"${BASE_EPS_10Y:.2f}", f"${BULL_EPS_10Y:.2f}"],
    ["Target Price (10Y)", f"${BEAR_10Y_TARGET:.2f}", f"${BASE_10Y_TARGET:.2f}", f"${BULL_10Y_TARGET:.2f}"],
    ["Upside from $8.08", f"{(BEAR_10Y_TARGET/PRICE - 1)*100:.0f}%", f"{(BASE_10Y_TARGET/PRICE - 1)*100:.0f}%", f"{(BULL_10Y_TARGET/PRICE - 1)*100:.0f}%"],
    ["10Y CAGR from $8.08", f"{(BEAR_10Y_TARGET/PRICE)**0.1 - 1:.1%}", f"{(BASE_10Y_TARGET/PRICE)**0.1 - 1:.1%}", f"{(BULL_10Y_TARGET/PRICE)**0.1 - 1:.1%}"],
]
for ri, row_data in enumerate(ext_data, 19):
    for ci, val in enumerate(row_data, 1):
        c(ws3, ri, ci, val, font=BOLD if ri == 19 else None, border=True,
          fill=HDR_FILL if ri == 19 else False)

# Note about framework
c(ws3, 25, 1, "Framework Note", font=BOLD)
c(ws3, 26, 1, (
    "Forward P/E is the primary framework because FCF is negative TTM (-$30.3M) due to a deliberate "
    "capex investment cycle in data centers and sports data infrastructure. FY2025 FCF was only $5.5M "
    "on $669.5M revenue — a 0.8% FCF margin that is not meaningful for multiple analysis. "
    "P/FCF at any exit multiple would be volatile and unanchored. Forward P/E on analyst consensus "
    "provides the most reliable calibration. P/S ratio serves as a secondary cross-check."
), align=Alignment(wrap_text=True))

# ═══════════════════════════════════════════════════════════════
# Sheet 4: Actuals Source Audit
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Actuals Source Audit")

ws4.merge_cells("A1:D1")
c(ws4, 1, 1, "Actuals Source Audit — Genius Sports Limited", font=HDR_FONT)

audit_data = [
    ["Data Point", "Value", "Source URL", "Date / Notes"],
    ["Stock Price", "$8.08", "finance.yahoo.com/quote/GENI/", "Aug 4, 2026 close"],
    ["Market Cap", "$1.85B", "Yahoo Finance Key Statistics", "Current"],
    ["Enterprise Value", "$1.69B", "Yahoo Finance Key Statistics", "Current"],
    ["Shares Outstanding", "267.34M", "Yahoo Finance Key Statistics", "Q1 FY2026 (Most Recent)"],
    ["Revenue TTM", "$713.5M", "finance.yahoo.com/quote/GENI/financials/", "TTM, annual view"],
    ["Revenue FY2025", "$669.5M", "Same", "Fiscal year ending Dec 31, 2025"],
    ["Gross Profit TTM", "$162.0M", "Same", "Margin 22.7%"],
    ["Gross Profit FY2025", "$153.8M", "Same", "Margin 22.9%"],
    ["Operating Income TTM", "-$157.4M", "Same", "Still negative, narrowing"],
    ["Net Income TTM", "-$158.9M", "Same", "TTM"],
    ["Diluted EPS TTM", "-$0.62", "Same", "TTM"],
    ["EBITDA TTM", "-$101.3M", "Same", "Negative"],
    ["D&A TTM", "$75.1M", "Same", "Heavy D&A — sports data infrastructure"],
    ["FCF TTM", "-$30.3M", "finance.yahoo.com/quote/GENI/cash-flow/", "Negative from capex cycle"],
    ["OCF TTM", "$50.8M", "Same", "Positive — franchise generating operating cash"],
    ["Capex TTM", "$81.1M", "Same", "200%+ of FY2022 capex — deliberate investment"],
    ["Total Cash Q1 FY26", "$197.4M", "Yahoo Finance Key Statistics", "Most recent quarter"],
    ["End Cash TTM CF", "$194.3M", "Cash flow statement TTM", "Consistent with stats"],
    ["Total Debt", "$30.6M", "Yahoo Finance Key Statistics / BS", "Capital lease obligations"],
    ["Capital Lease Obligations", "$30.5M", "Balance Sheet FY2025 $30.5M", "Up from $7.5M FY2024"],
    ["Net Tangible Book Value", "$242.2M", "Balance Sheet FY2025", "2025"],
    ["Total Common Equity", "$724.5M", "Balance Sheet FY2025", "2025"],
    ["Beta", "1.88", "Yahoo Finance Key Statistics", "5Y Monthly"],
    ["Forward P/E FY2026", "56.8x", "Yahoo Finance Key Statistics", "Using $0.52 EPS consensus"],
    ["Risk-Free Rate", "4.611%", "cnbc.com/quotes/US10Y", "Aug 4, 2026"],
    ["Analyst Revenue FY2026", "$1.00B", "finance.yahoo.com/quote/GENI/analysis/", "19 analysts"],
    ["Analyst Revenue FY2027", "$1.29B", "Same", "18 analysts"],
    ["Analyst EPS FY2026", "$0.52", "Same", "18 analysts, non-GAAP normalized"],
    ["Analyst EPS FY2027", "$0.97", "Same", "18 analysts"],
    ["Q1 FY2026 EPS Actual", "$0.016", "Same", "vs est $0.01 — +65% surprise"],
    ["Q2 FY2026 Earnings Date", "Aug 6, 2026", "Yahoo Finance Profile", "2 days from model date"],
    ["Capital Raise TTM", "$144M", "Cash flow statement TTM", "Issuance of capital stock"],
    ["52-Week Range", "$3.83 – $13.73", "Yahoo Finance Key Statistics", "Current"],
]

for i, (dp, val, src, note) in enumerate(audit_data, 2):
    c(ws4, i, 1, dp, font=BOLD if dp == audit_data[0][0] else None, border=True,
      fill=HDR_FILL if i == 2 else False)
    c(ws4, i, 2, val, border=True)
    c(ws4, i, 3, src, border=True)
    c(ws4, i, 4, note, border=True)

ws4.column_dimensions["A"].width = 28
ws4.column_dimensions["B"].width = 22
ws4.column_dimensions["C"].width = 42
ws4.column_dimensions["D"].width = 40

# ═══════════════════════════════════════════════════════════════
# Sheet 5: Questions
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Questions")

ws5.merge_cells("A1:C1")
c(ws5, 1, 1, "Open Questions — Genius Sports Limited", font=HDR_FONT)

questions_data = [
    ["#", "Question", "Implication"],
    ["1", "Why is capex running at $81M/year (11% of revenue)? What infrastructure is being built — data centers, streaming platforms, or acquisition-related buildouts?",
     "If capex settles at $47-50M (FY2022-2023 range), FCF recovery could add $30-34M/year. If capex stays elevated, FCF framework remains invalid."],
    ["2", "$144M capital raise TTM — was this for acquisitions, operational expansion, or balance sheet strengthening? What valuation was the raise at?",
     "If raise was at inflated price (stock was $10-13 range), current price represents 42% decline. Dilution of 35M shares in ~1 year."],
    ["3", "Capital lease obligations jumped from $7.5M (FY2024) to $30.5M (FY2025). What assets are being leased — servers, offices, or betting infrastructure?",
     "Lease obligations act as quasi-debt. If these are data center leases, they support revenue infrastructure. If they are general operating leases, they may represent over-investment."],
    ["4", "Revenue jumped 23.8% in FY2025 TTM (from $511M to $669M) and continues at $713M TTM. Is this purely organic or acquisition-driven?",
     "GENI has made strategic acquisitions (e.g., data providers in sports betting). If revenue growth is M&A-driven, quality of growth is lower than organic."],
    ["5", "TTM operating loss is -$157M on $713M revenue (-22% margin). Given gross margin is 22.7%, this means OpEx is 44.8% of revenue. Is OpEx structural or is there operating leverage ahead?",
     "If OpEx falls to 30-35% of revenue at $1B+ revenue level, GENI becomes profitable. If OpEx stays at 40%+, path to profitability is distant."],
    ["6", "Other income/expense swung from +$23.7M (FY2025) to -$6.9M (TTM). What drove the reversal — FX, investment gains/losses, or restructuring?",
     "FY2025 was buoyed by a one-time gain of $23.7M that TTM shows as a negative. This distortion makes FY2025 look better than TTM."],
    ["7", "Share count went from 205.8M (FY2022) to 250.4M issued / 267.3M implied outstanding (Aug 2026). 29% dilution in 3 years. Is future dilution likely?",
     "Dilution from the $144M raise + SBC. If the company needs more capital before profitability, further dilution is likely — especially with stock at $8 vs. raise price."],
    ["8", "How does GENI's sports betting data monopoly hold up if leagues develop proprietary data platforms or competitors emerge?",
     "Revenue concentration risk — leagues could become customers, suppliers, and competitors simultaneously."],
    ["9", "What percentage of revenue is from sports betting/integrity services vs. league data licensing? Which is more defensible?",
     "Betting revenue is higher growth but more regulatory risk. League data is sticky but slower growth and more leverage."],
    ["10", "The company is UK-domiciled (London HQ) with NYSE listing. What is the effective tax rate? Are there UK/US tax optimization or double-taxation concerns?",
     "Tax provision shows -$3.1M TTM (tax benefit on losses) and -$2.5M FY2025. Effective rate hard to pin down."],
    ["11", "Q2 FY2026 earnings on Aug 6, 2026 — what will this report settle? Path to profitability is the core question.",
     "This is the primary near-term catalyst and report trigger."],
    ["12", "Is the 52-week range of $3.83-$13.73 consistent with the analyst consensus of $0.52/FY2026 and $0.97/FY2027? At what valuation does the market think the company is worth?",
     "Reverse valuation question the entire model depends on."],
]

for i, (num, q, impl) in enumerate(questions_data, 2):
    c(ws5, i, 1, num, font=BOLD if i == 2 else None, border=True, fill=HDR_FILL if i == 2 else False)
    c(ws5, i, 2, q, border=True, align=Alignment(wrap_text=True))
    c(ws5, i, 3, impl, border=True, align=Alignment(wrap_text=True))

ws5.column_dimensions["A"].width = 5
ws5.column_dimensions["B"].width = 65
ws5.column_dimensions["C"].width = 60

# ═══════════════════════════════════════════════════════════════
# Sheet 6: Sources
# ═══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Sources")

ws6.merge_cells("A1:C1")
c(ws6, 1, 1, "Sources — Genius Sports Limited", font=HDR_FONT)

sources_data = [
    ["#", "Source", "URL / Description"],
    ["1", "Yahoo Finance — Quote / Statistics", "finance.yahoo.com/quote/GENI/ and /key-statistics/"],
    ["2", "Yahoo Finance — Income Statement", "finance.yahoo.com/quote/GENI/financials/"],
    ["3", "Yahoo Finance — Balance Sheet", "finance.yahoo.com/quote/GENI/balance-sheet/"],
    ["4", "Yahoo Finance — Cash Flow", "finance.yahoo.com/quote/GENI/cash-flow/"],
    ["5", "Yahoo Finance — Analysis / Estimates", "finance.yahoo.com/quote/GENI/analysis/"],
    ["6", "Yahoo Finance — Profile", "finance.yahoo.com/quote/GENI/profile/"],
    ["7", "CNBC — US 10Y Treasury", "cnbc.com/quotes/US10Y — yield 4.611% Aug 4, 2026"],
    ["8", "StockAnalysis.com", "stockanalysis.com — returned 404 for GENI (confirmed)"],
    ["9", "Genius Sports Corporate Website", "geniussports.com — HQ: London, 2000 employees"],
    ["10", "SEC Filings (via Yahoo Profile)", "Multiple 6-K filings and 20-F (Mar 16, 2026) listed on Yahoo profile page"],
]

for i, (num, src, desc) in enumerate(sources_data, 2):
    c(ws6, i, 1, num, border=True, fill=HDR_FILL if i == 2 else False)
    c(ws6, i, 2, src, font=BOLD if i == 2 else None, border=True)
    c(ws6, i, 3, desc, border=True)

ws6.column_dimensions["A"].width = 5
ws6.column_dimensions["B"].width = 45
ws6.column_dimensions["C"].width = 60

# ── Save ────────────────────────────────────────────────────────
wb.save(str(XLSX))
print(f"\nWorkbook saved to {XLSX}")

# Verify
wb2 = openpyxl.load_workbook(str(XLSX))
print(f"Sheets: {wb2.sheetnames}")
for sn in wb2.sheetnames:
    ws_check = wb2[sn]
    print(f"  {sn}: {ws_check.max_row} rows × {ws_check.max_column} cols")
